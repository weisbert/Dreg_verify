"""复现候选 iddq-1：dft 门控的 MUX 根，analyze_signal 漏 iddq 门。
对比 analyze_signal(mux 根).vectors / dft_gate  vs  build_for_topout(.sv)  vs  topout_view_models。
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook as _load
from openpyxl.utils import column_index_from_string as _ci

import make_mirror_btlp as M
from dreg_verify.excel_model import load_workbook
from dreg_verify import topout as T
from dreg_verify import resolver as R

PATH = "scratchpad/_repro_mux_iddq.xlsx"

# 1) 造基础夹具(含 d_en_vco_fc_ls 那套，含 mux d_bt_lp_lna_itrim)
M.build_dft_gated_ls(PATH)

# 2) 在 dft 页给 mux 输出 d_bt_lp_lna_itrim 加一行 iddq 门控观测: E=B?0:A
wb_x = _load(PATH)
dft = wb_x["dft"]
r = dft.max_row + 1
def setc(col, val):
    dft.cell(row=r, column=_ci(col), value=val)
setc("A", "d_bt_lp_lna_itrim_to_dft[3:0]")          # A: 本级输入(被门观测的网)
setc("B", "d_bt_lp_pll_dig_dft_iddq_mode_to_dft")   # B: iddq 门
setc("D", "d_bt_lp_lna_itrim")                       # D: 被门控的输出基名(=mux out_base)
setc("E", "B?0:A")                                    # E: 门控表达式(iddq=1 → 0)
setc("F", "16")
setc("H", "Jiao Yexiang")
wb_x.save(PATH)

# 3) 解析 + 三路对比
wb = load_workbook(PATH)
print("dft keys:", sorted((wb.dft or {}).keys()))
gate = (wb.dft or {}).get("d_bt_lp_lna_itrim")
print("dft gate for d_bt_lp_lna_itrim:", gate)

topo = next(t for t in wb.topout if t.name.lower().startswith("d_bt_lp_lna_itrim"))
print("\nTopout name:", topo.name)

# ---- 路径 A: analyze_signal (GUI 真值表编辑器走它) ----
res = T.analyze_signal(wb, R.Resolver(wb), topo)
print("\n[analyze_signal] root.kind =", res.root.kind, " status =", res.status)
print("[analyze_signal] n_vectors =", len(res.vectors))
print("[analyze_signal] res.dft_gate =", res.dft_gate)
n_dftpitch = sum(1 for v in res.vectors if getattr(v, "dft_pitch", False))
print("[analyze_signal] DFT 拍数 =", n_dftpitch)
n_extraf = sum(1 for v in res.vectors if getattr(v, "extra_forces", None))
print("[analyze_signal] 带 extra_forces 向量数 =", n_extraf)

# ---- 路径 B: build_for_topout (.sv 导出) ----
b = T.build_for_topout(wb, only=[topo.name])
# 找到 mux 块
sv_block = None
for ln, st in b["blocks"]:
    if st.get("topout_name", "").lower() == topo.name.lower():
        sv_block = (ln, st)
print("\n[build_for_topout/.sv] n_vectors =", sv_block[1].get("n_vectors") if sv_block else "NONE")
if sv_block:
    txt = "\n".join(sv_block[0])
    print("[build_for_topout/.sv] 出现 iddq_mode 行数 =", txt.lower().count("iddq_mode"))
    print("[build_for_topout/.sv] 出现 IDDQ 漏电态 =", "IDDQ" in txt or "iddq" in txt.lower())

# ---- 路径 C: topout_view_models (报告/GUI 富表) ----
vms = T.topout_view_models(wb)
vm = next((m for m in vms if m["name"].lower() == topo.name.lower()), None)
print("\n[topout_view_models] kind =", vm["kind"] if vm else "NONE")
print("[topout_view_models] n tests =", len(vm["tests"]) if vm else "NONE")
if vm:
    in_labels = [i.get("label") or i.get("name") or str(i) for i in vm["inputs"]]
    print("[topout_view_models] input labels =", in_labels)
    has_iddq = any("iddq" in str(i).lower() for i in in_labels)
    print("[topout_view_models] 有 iddq 输入行 =", has_iddq)

print("\n=== 结论 ===")
print("analyze(GUI真值表)向量 = %d, dft_gate = %r, DFT拍 = %d"
      % (len(res.vectors), res.dft_gate, n_dftpitch))
print(".sv 向量 = %s" % (sv_block[1].get("n_vectors") if sv_block else "NONE"))
print("view_models tests = %s" % (len(vm["tests"]) if vm else "NONE"))
