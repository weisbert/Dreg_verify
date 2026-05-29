# -*- coding: utf-8 -*-
"""
inspect_excel.py — 读取 Dreg 核心 Excel，导出结构化文本，方便粘贴给 Claude 分析。

用法:
    python inspect_excel.py <excel路径>
        [--rows N]            每个 sheet 导出的样本行数 (默认 20)
        [--out 输出.txt]      输出文件 (默认在 Excel 同目录下 <名字>_inspect.txt)
        [--sheets a,b,c]      只导出这些 sheet (默认全部, 关键页自动排前面)
        [--compact]           精简模式：去冗余段落, logic 表达式列(L)完整不截断 (推荐, 输出最短)
        [--split]             每个 sheet 单独存一个 txt, 方便分批发送
        [--anon-signals]      对“信号名类”内容做“保结构”脱敏 (见下)
        [--mask-owners]       把 owner 列(P)的人名替换成 OwnerN
        [--maxlen N]          单元格内容最大显示长度 (默认 80; logic 的 L/M 列始终不截断)

想让输出最短就用:
    python inspect_excel.py "核心文件.xlsx" --compact --mask-owners --rows 10 \
        --sheets logic,regmap,NamingRule,for_test
还嫌长就再加 --split, 然后一页一页发我。

脱敏说明:
  - 默认【不脱敏】，导出原文；请你自己过目后再决定要不要手动删改。
  - --anon-signals 会把字母词元替换成 w1/w2/...，但【保留】下划线、方括号、位宽数字、
    分隔符与表达式运算符；logic 的真值表达式列(L) 与类型列(M) 始终保持原文（核心信息且不敏感）。
    注意：开了它之后，命名规则里像 "logic" 这种字面插入串也会被替换，我可能要回头跟你确认几个字面前缀/中缀。
  - 重要：请尽量保留【信号名结构】和【表达式列】，否则生成器的核心逻辑没法还原。

依赖: openpyxl  ->  pip install openpyxl
限制: 仅支持 .xlsx。若是老的 .xls，请用 Excel 另存为 .xlsx 再跑。
"""

import argparse
import os
import re
import sys

# Windows 控制台常是 GBK，打印中文/emoji 会崩；统一改 UTF-8（文件输出本就 UTF-8，不受影响）
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    sys.exit("缺少 openpyxl，请先运行:  pip install openpyxl")

# 让 --expr-forms 能用生成器自带的表达式求值器做"解析覆盖"自检
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from dreg_verify import expr as _expr
except Exception:   # noqa: BLE001  （缺包时 --expr-forms 退化为仅做形态归并，不做解析校验）
    _expr = None

# 关键 sheet 排在最前，其余按原顺序跟在后面
PRIORITY_SHEETS = ["logic", "regmap", "NamingRule", "for_test", "main",
                   "total_memory_map", "mux", "level_shift", "SignalPath"]

PROFILE_SCAN_CAP = 5000   # 列取值统计最多扫描多少行，防止 max_row 被撑爆

# ---------- 脱敏 ----------
_word_map = {}
_owner_map = {}


def anon_signal(text):
    """保结构脱敏：仅替换字母词元，保留 _ [] : 数字 等结构。"""
    def repl(m):
        key = m.group(0).lower()
        if key not in _word_map:
            _word_map[key] = "w%d" % (len(_word_map) + 1)
        return _word_map[key]
    return re.sub(r"[A-Za-z]+", repl, text)


def mask_owner(text):
    key = text.strip().lower()
    if not key:
        return text
    if key not in _owner_map:
        _owner_map[key] = "Owner%d" % (len(_owner_map) + 1)
    return _owner_map[key]


# ---------- 单元格取值 ----------
def cell_str(value, maxlen=80, no_trunc=False):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        s = str(int(value))
    else:
        s = str(value)
    s = s.replace("\n", "\\n").replace("\r", "")
    if not no_trunc and maxlen and len(s) > maxlen:
        s = s[:maxlen] + "…(+%d)" % (len(s) - maxlen)
    return s


def read_rows(ws, max_row, cap):
    """返回 [[v,v,...], ...]，按行；超出 cap 截断。"""
    upto = min(max_row or 0, cap)
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=upto, values_only=True):
        rows.append(list(r))
    return rows


def is_blank_row(row):
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in row)


# ---------- 各列脱敏策略 ----------
def transform_cell(sheet_name, col_letter, value, args):
    """根据开关与列位置返回（可能脱敏后的）值。"""
    if value is None or not isinstance(value, str):
        return value
    sn = sheet_name.lower()
    # logic 的表达式列 L、类型列 M 永远不脱敏（符号逻辑/枚举，不敏感且是核心信息）
    if sn == "logic" and col_letter in ("L", "M"):
        return value
    # owner 列 P：仅在 --mask-owners 时替换
    if sn == "logic" and col_letter == "P":
        return mask_owner(value) if args.mask_owners else value
    # 其余字符串：仅在 --anon-signals 时保结构脱敏
    if args.anon_signals:
        return anon_signal(value)
    return value


def render_row_line(rows, ri, max_col, sheet_name, args):
    """把一行渲染成 '[row N] A=.. | C=..' （跳过空单元格）。logic 的 L 列不截断。"""
    is_logic = sheet_name.lower() == "logic"
    cells = []
    for c in range(max_col):
        letter = get_column_letter(c + 1)
        raw = rows[ri][c] if c < len(rows[ri]) else None
        v = transform_cell(sheet_name, letter, raw, args)
        no_trunc = is_logic and letter in ("L", "M")
        sv = cell_str(v, args.maxlen, no_trunc)
        if sv != "":
            cells.append("%s=%s" % (letter, sv))
    return "[row %d] %s" % (ri + 1, " | ".join(cells))


# ---------- 输出 ----------
def dump_sheet(out, ws, sheet_name, args):
    max_row = ws.max_row
    max_col = ws.max_column
    cap = 20000 if args.find else PROFILE_SCAN_CAP
    rows = read_rows(ws, max_row, cap)
    is_logic = sheet_name.lower() == "logic"

    out.append("=" * 70)
    out.append("==== SHEET: %s ====" % sheet_name)
    out.append("dimensions(原始报告): %s 行 x %s 列   (实际扫描 %d 行)"
               % (max_row, max_col, len(rows)))

    if not rows:
        out.append("(空表)")
        return

    # 列字母 ↔ 表头(第1行)
    header = rows[0]
    out.append("--- 列映射 (字母 -> 第1行内容) ---")
    line = []
    for c in range(max_col):
        letter = get_column_letter(c + 1)
        h = transform_cell(sheet_name, letter,
                           header[c] if c < len(header) else None, args)
        line.append("%s=%s" % (letter, cell_str(h, args.maxlen)))
    out.append(" | ".join(line))

    # 关键词过滤模式：只打印"任意单元格含关键词"的行（跨整页）
    if args.find:
        keywords = [k.strip().lower() for k in args.find.split(",") if k.strip()]
        out.append("--- 匹配行 (关键词: %s) ---" % keywords)
        matched = 0
        for ri in range(len(rows)):
            rowtext = " ".join("" if v is None else str(v)
                               for v in rows[ri]).lower()
            if any(k in rowtext for k in keywords):
                out.append(render_row_line(rows, ri, max_col, sheet_name, args))
                matched += 1
                if matched >= args.find_max:
                    out.append("... (达到每页上限 %d, 截断)" % args.find_max)
                    break
        out.append("(本页匹配 %d 行)" % matched)
        return

    # 前几行原文（非精简模式才输出，帮我判断表头是否跨多行）
    if not args.compact:
        out.append("--- 前 5 行原文 ---")
        for ri in range(min(5, len(rows))):
            out.append(render_row_line(rows, ri, max_col, sheet_name, args))

    # 样本数据行（跳过空行）
    out.append("--- 样本数据行 (最多 %d 行, 跳过空行) ---" % args.rows)
    shown = 0
    for ri in range(1, len(rows)):
        if shown >= args.rows:
            break
        if is_blank_row(rows[ri]):
            continue
        out.append(render_row_line(rows, ri, max_col, sheet_name, args))
        shown += 1

    # logic 页：列画像（精简模式也保留，很有用）；竖排仅非精简模式输出
    if is_logic:
        dump_logic_profile(out, rows, max_col, sheet_name, args)
        if not args.compact:
            dump_logic_vertical(out, rows, max_col, sheet_name, args)


def dump_logic_profile(out, rows, max_col, sheet_name, args):
    out.append("--- [logic] 列画像 (非空数, 去重数, 取值/样本) ---")
    header = rows[0]
    for c in range(max_col):
        letter = get_column_letter(c + 1)
        vals = []
        for ri in range(1, len(rows)):
            v = rows[ri][c] if c < len(rows[ri]) else None
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            vals.append(transform_cell(sheet_name, letter, v, args))
        nonempty = len(vals)
        if nonempty == 0:
            continue
        distinct = list(dict.fromkeys(cell_str(v, args.maxlen) for v in vals))
        hname = cell_str(header[c] if c < len(header) else None, 40)
        if len(distinct) <= 20:
            detail = "取值=%s" % distinct
        else:
            detail = "样本=%s" % distinct[:6]
        out.append("  col %s [%s]: 非空=%d, 去重=%d, %s"
                   % (letter, hname, nonempty, len(distinct), detail))


def dump_logic_vertical(out, rows, max_col, sheet_name, args):
    out.append("--- [logic] 前 %d 个数据行竖排 (L/M 列不截断) ---" % args.rows)
    header = rows[0]
    shown = 0
    for ri in range(1, len(rows)):
        if shown >= args.rows:
            break
        if is_blank_row(rows[ri]):
            continue
        out.append("[row %d]" % (ri + 1))
        for c in range(max_col):
            letter = get_column_letter(c + 1)
            raw = rows[ri][c] if c < len(rows[ri]) else None
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                continue
            v = transform_cell(sheet_name, letter, raw, args)
            no_trunc = letter in ("L", "M")
            hname = cell_str(header[c] if c < len(header) else None, 30)
            out.append("    %-3s [%s]: %s"
                       % (letter, hname, cell_str(v, args.maxlen, no_trunc)))
        shown += 1


def safe_name(s):
    return re.sub(r"[^A-Za-z0-9_.-]", "_", s)


# ───────────────────────────── 表达式形态覆盖报告 ─────────────────────────────
def _struct_form(expr_text):
    """把表达式抽象成"结构形态"：变量→V, 常量→K, 其余算子/括号保留。
    用于把 A?C:B 与 X?Y:Z 归并到同一形态 V?V:V。无求值器时用简易正则降级。"""
    if _expr is not None:
        try:
            toks = _expr.tokenize(expr_text)
            out = []
            for t in toks:
                if t.kind == "EOF":
                    break
                if t.kind == "ID":
                    out.append("V")
                elif t.kind in ("BASED", "DEC"):
                    out.append("K")
                else:
                    out.append(t.text)
            return "".join(out)
        except Exception:  # noqa: BLE001
            pass
    s = re.sub(r"\d+'[sSbBoOdDhH][0-9a-fA-FxXzZ_\?]+", "K", expr_text)
    s = re.sub(r"[A-Za-z_]\w*", "V", s)
    return re.sub(r"\s+", "", s)


def expr_forms_report(ws, header_row=2):
    """扫描 logic 页 L 列，按"原始表达式"和"结构形态"两级归并，并用求值器逐条试解析。"""
    L = column_index_from_string("L")
    K = column_index_from_string("K")
    R = column_index_from_string("R")
    by_expr = {}          # 原始表达式 -> {count, examples:[(R,K)]}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        l = cell_str(row[L - 1] if L - 1 < len(row) else None, 0, no_trunc=True).strip() \
            if (L - 1 < len(row) and row[L - 1] is not None) else ""
        k = cell_str(row[K - 1] if K - 1 < len(row) else None, 0, no_trunc=True).strip() \
            if (K - 1 < len(row) and row[K - 1] is not None) else ""
        if l == "" or k == "":
            continue
        rno = cell_str(row[R - 1] if R - 1 < len(row) else None, 0)
        e = by_expr.setdefault(l, {"count": 0, "examples": []})
        e["count"] += 1
        if len(e["examples"]) < 4:
            e["examples"].append((rno, k))

    # 按结构形态归并
    by_form = {}
    parse_fail = []
    for expr_text, info in by_expr.items():
        form = _struct_form(expr_text)
        f = by_form.setdefault(form, {"count": 0, "exprs": []})
        f["count"] += info["count"]
        f["exprs"].append(expr_text)
        # 解析校验（用占位宽度，纯语法/结构层面）
        if _expr is not None:
            try:
                node = _expr.parse(expr_text)
                widths = {v: 1 for v in _expr.collect_vars(node)}
                _expr.classify_vars(node, _expr.Env(widths))
            except Exception as ex:  # noqa: BLE001
                parse_fail.append((expr_text, str(ex), info["examples"]))

    out = []
    out.append("=" * 70)
    out.append("==== 表达式形态覆盖报告 (logic.L 列) ====")
    out.append("不同表达式: %d 种；不同结构形态: %d 种；求值器可用: %s"
               % (len(by_expr), len(by_form), "是" if _expr else "否(仅形态归并)"))
    if _expr is not None:
        out.append("解析失败: %d 种 %s"
                   % (len(parse_fail), "✅ 全部可解析" if not parse_fail else "⚠ 见下"))
    out.append("")
    out.append("--- 按结构形态归并 (形态 | 出现次数 | 实例表达式) ---")
    for form, f in sorted(by_form.items(), key=lambda kv: -kv[1]["count"]):
        uniq = list(dict.fromkeys(f["exprs"]))
        out.append("  [%3d次] %s" % (f["count"], form))
        for e in uniq[:6]:
            out.append("           例: %s" % e)
        if len(uniq) > 6:
            out.append("           ...(+%d 个同形态表达式)" % (len(uniq) - 6))

    out.append("")
    out.append("--- 全部不同表达式 (按出现次数) ---")
    for expr_text, info in sorted(by_expr.items(), key=lambda kv: -kv[1]["count"]):
        status = ""
        if _expr is not None:
            status = " [解析失败]" if any(expr_text == pf[0] for pf in parse_fail) else " [OK]"
        ex = ", ".join("R%s:%s" % (r, k) for r, k in info["examples"])
        out.append("  [%3d次]%s %s   ← %s" % (info["count"], status, expr_text, ex))

    if parse_fail:
        out.append("")
        out.append("--- ⚠ 求值器无法解析的表达式 (需扩展 dreg_verify/expr.py) ---")
        for expr_text, err, examples in parse_fail:
            ex = ", ".join("R%s:%s" % (r, k) for r, k in examples)
            out.append("  表达式: %s" % expr_text)
            out.append("    错误: %s" % err)
            out.append("    出现于: %s" % ex)
    return out, len(by_expr), len(parse_fail)


def run_expr_forms(args):
    wb = openpyxl.load_workbook(args.excel, data_only=True, read_only=True)
    ws = None
    for s in wb.sheetnames:
        if s.lower() == "logic":
            ws = wb[s]
            break
    if ws is None:
        wb.close()
        sys.exit("找不到 'logic' 页；现有页：%s" % wb.sheetnames)
    lines, n_expr, n_fail = expr_forms_report(ws, header_row=2)
    wb.close()
    base = os.path.splitext(args.excel)[0]
    out_path = args.out if args.out else base + "_exprforms.txt"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print("已导出表达式形态报告: %s" % out_path)
    print("不同表达式 %d 种；求值器解析失败 %d 种。" % (n_expr, n_fail))
    if n_fail:
        print("⚠ 有未覆盖形态——把报告里 [解析失败] 那几条发我，我扩展求值器。")
    else:
        print("✅ 全部表达式均可被求值器解析（覆盖完整）。" if _expr
              else "（未装 dreg_verify，仅做了形态归并，未做解析校验。）")


# ───────────────────────────── 信号聚焦调试模式 (--signal) ─────────────────────────────
# 目的：调一个具体输出信号(如 d_en_refbuf)时，一条命令把它在各页的相关信息全抠出来：
#   ① logic 页该信号的定义行(竖排, L/M 不截断)；
#   ② 自动展开它表达式里的输入名(如 testmode)，在其它页一并定位(抓"testmode value error")；
#   ③ for_test / *_tc 等"块状页"按【整块】抠(含横向 T0..T63 向量, 不再漏行)；
#   ④ regmap / total_memory_map 抠出"目标 + 各输入"的定义行。
_SIG_INPUT_COLS = list("ABCDEFGHIJ")


def _cell_by_letter(row, letter):
    """按列字母从 values_only 的 row(tuple) 取值；越界返回 None。"""
    idx = column_index_from_string(letter) - 1
    return row[idx] if 0 <= idx < len(row) else None


def _strip_sig(name):
    """去位宽 [..] 与 _to_logic 后缀，返回小写基名，用于跨页匹配。空→''。"""
    s = "" if name is None else str(name).strip()
    s = re.sub(r"\[[^\]]*\]", "", s)              # 去 [8:0] / [2]
    for suf in ("_to_logic", "_TO_LOGIC"):
        if s.endswith(suf):
            s = s[: -len(suf)]
    return s.strip().lower()


def _row_contains(row, keywords):
    """整行任意单元格(字符串化小写)是否含任一关键词(子串匹配)。"""
    text = " ".join("" if v is None else str(v) for v in row).lower()
    return any(k in text for k in keywords)


def _has_row2_header(sheet_name):
    """tmm 是块结构无统一表头；其余页(logic/regmap/for_test)真表头在第 2 行。"""
    return sheet_name.lower() not in ("total_memory_map", "total memory map", "memory_map")


def _header_map_line(ws, args, header_row=2):
    """轻量读第 header_row 行，渲染成 '字母=表头' 列映射，帮助解读列。"""
    hdr = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=header_row, values_only=True), start=1):
        if ri == header_row:
            hdr = row
    if not hdr:
        return None
    parts = []
    for c in range(len(hdr)):
        h = hdr[c]
        if h is None or (isinstance(h, str) and h.strip() == ""):
            continue
        parts.append("%s=%s" % (get_column_letter(c + 1), cell_str(h, args.maxlen)))
    return "--- 列映射(第%d行表头) --- %s" % (header_row, " | ".join(parts))


def _dump_matching_rows(out, ws, sn, keywords, args):
    """普通页(regmap/tmm/mux/…)：抠出任意单元格含关键词的行(含目标信号及其各输入)。"""
    out.append("")
    out.append("=" * 70)
    out.append("==== %s : 含关键词的行 ====" % sn)
    if _has_row2_header(sn):
        hl = _header_map_line(ws, args)
        if hl:
            out.append(hl)
    cap = 20000
    matched = 0
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, cap),
                                          values_only=True), start=1):
        if _has_row2_header(sn) and ri <= 2:
            continue   # 跳过表头行(仅对有统一表头的页；tmm 无表头, 行 1-2 是真数据)
        if _row_contains(row, keywords):
            cells = []
            for c in range(len(row)):
                raw = row[c]
                if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                    continue
                cells.append("%s=%s" % (get_column_letter(c + 1), cell_str(raw, args.maxlen)))
            out.append("[row %d] %s" % (ri, " | ".join(cells)))
            matched += 1
            if matched >= args.find_max:
                out.append("...(达每页上限 %d, 截断)" % args.find_max)
                break
    out.append("(本页命中 %d 行)" % matched)


def _dump_fortest_blocks(out, ws, sn, keywords, args):
    """块状页(for_test / *_tc)：A 列=输出名，每个输出一个行块；T0..T63 横向铺在右侧。
    用 A 列(去位宽基名)变化界定块边界；A 命中目标关键词的块整块 dump(含横向所有列)。"""
    out.append("")
    out.append("=" * 70)
    out.append("==== %s : 信号整块(含横向 T 向量) ====" % sn)
    if _has_row2_header(sn):
        hl = _header_map_line(ws, args)
        if hl:
            out.append(hl)
    cap = 20000
    rows = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, cap),
                                          values_only=True), start=1):
        rows.append((ri, row))
    n = len(rows)
    # 块起点：A 列基名非空且与上一个非空 A 不同
    starts, last = [], None
    for idx, (_ri, row) in enumerate(rows):
        a = _strip_sig(_cell_by_letter(row, "A"))
        if a and a != last:
            starts.append(idx)
        if a:
            last = a
    dumped = 0
    for si, start_idx in enumerate(starts):
        _ri, row = rows[start_idx]
        if _strip_sig(_cell_by_letter(row, "A")) not in keywords:
            continue
        end_idx = starts[si + 1] if si + 1 < len(starts) else n
        end_idx = min(end_idx, start_idx + args.block_tail)
        head_a = cell_str(_cell_by_letter(rows[start_idx][1], "A"), args.maxlen)
        out.append("")
        out.append("---- 块 A=%s  (行 %d..%d) ----"
                   % (head_a, rows[start_idx][0], rows[min(end_idx, n) - 1][0]))
        for k in range(start_idx, min(end_idx, n)):
            ri2, row2 = rows[k]
            cells = []
            for c in range(len(row2)):
                raw = row2[c]
                if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                    continue
                cells.append("%s=%s" % (get_column_letter(c + 1), cell_str(raw, args.maxlen)))
            if cells:
                out.append("[row %d] %s" % (ri2, " | ".join(cells)))
        dumped += 1
    if dumped == 0:
        out.append("⚠ 未找到 A 列等于目标信号的块；退化为'含关键词的行':")
        _dump_matching_rows(out, ws, sn, keywords, args)
    else:
        out.append("(共 dump %d 个块)" % dumped)


def _is_block_sheet(sheet_name):
    low = sheet_name.lower()
    return ("for_test" in low) or ("fortest" in low) or low.endswith("_tc") or low.endswith("tc")


def run_signal(args):
    wb = openpyxl.load_workbook(args.excel, data_only=True, read_only=True)
    all_sheets = wb.sheetnames
    targets = [t.strip() for t in args.signal.split(",") if t.strip()]
    base_kw = {_strip_sig(t) for t in targets} | {t.lower() for t in targets}
    base_kw = {k for k in base_kw if k}

    out = []
    out.append("Dreg 信号聚焦调试导出 (--signal)")
    out.append("文件: %s" % os.path.basename(args.excel))
    out.append("目标信号/关键词: %s" % targets)
    out.append("全部 sheet (%d): %s" % (len(all_sheets), all_sheets))

    # ── Pass 1：logic 页定位目标行，收集其输入名做 1 级展开(抓 testmode 等) ──
    logic_name = next((s for s in all_sheets if s.lower() == "logic"), None)
    expanded = set(base_kw)
    logic_hits = []
    if logic_name:
        ws = wb[logic_name]
        header = None
        for ri, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if ri == 2:
                header = row
                continue
            if ri < 2:
                continue
            kbase = _strip_sig(_cell_by_letter(row, "K"))
            if kbase in base_kw or _row_contains(row, base_kw):
                logic_hits.append((ri, row))
                if args.signal_expand > 0:
                    for col in _SIG_INPUT_COLS:
                        b = _strip_sig(_cell_by_letter(row, col))
                        if b:
                            expanded.add(b)
        out.append("")
        out.append("=" * 70)
        out.append("==== logic 命中行 (目标信号定义；L/M 列不截断) ====")
        if not logic_hits:
            out.append("⚠ 在 logic 页未找到目标信号的行 —— 核对名字/位宽/前缀(K 列原文)。")
        for ri, row in logic_hits:
            out.append("[logic row %d]" % ri)
            for c in range(len(row)):
                raw = row[c]
                if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                    continue
                letter = get_column_letter(c + 1)
                hname = cell_str(header[c] if header and c < len(header) else None, 30)
                no_trunc = letter in ("L", "M")
                out.append("    %-3s [%s]: %s" % (letter, hname, cell_str(raw, args.maxlen, no_trunc)))
        out.append("")
        out.append("展开后的关键词集合(目标 + 其 logic 输入, 用于在其它页定位 testmode 等):")
        out.append("  %s" % sorted(expanded))

    # ── Pass 2：其它页 ──
    if args.sheets:
        want = [s.strip() for s in args.sheets.split(",") if s.strip()]
        sheets = [s for s in all_sheets if s in want]
    else:
        sheets = list(all_sheets)
    for sn in sheets:
        if logic_name and sn == logic_name:
            continue
        ws = wb[sn]
        if _is_block_sheet(sn):
            _dump_fortest_blocks(out, ws, sn, expanded, args)
        else:
            _dump_matching_rows(out, ws, sn, expanded, args)

    wb.close()
    out_path = args.out if args.out else os.path.splitext(args.excel)[0] + "_signal.txt"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(out) + "\n")
    print("已导出信号聚焦报告: %s" % out_path)
    print("命中 logic 行: %d；展开关键词: %d 个；行数: %d。"
          % (len(logic_hits), len(expanded), len(out)))
    print("请把该 txt 全文贴给我。")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel", help="Excel 文件路径 (.xlsx)")
    ap.add_argument("--rows", type=int, default=20)
    ap.add_argument("--out", default=None)
    ap.add_argument("--sheets", default=None, help="逗号分隔, 只导出这些 sheet")
    ap.add_argument("--compact", action="store_true")
    ap.add_argument("--split", action="store_true")
    ap.add_argument("--anon-signals", action="store_true")
    ap.add_argument("--mask-owners", action="store_true")
    ap.add_argument("--maxlen", type=int, default=80)
    ap.add_argument("--find", default=None,
                    help="逗号分隔关键词; 跨整页只导出任意单元格含关键词的行")
    ap.add_argument("--find-max", type=int, default=100,
                    help="--find 时每页最多输出多少匹配行 (默认 100)")
    ap.add_argument("--expr-forms", action="store_true",
                    help="只生成表达式形态覆盖报告: 枚举 logic.L 所有不同表达式/结构形态, "
                         "并用 dreg_verify 求值器逐条试解析(OK/FAIL), 闭合覆盖风险")
    ap.add_argument("--signal", default=None,
                    help="信号聚焦调试: 逗号分隔的信号名/关键词(如 d_en_refbuf,testmode)。"
                         "跨页抠出: logic 定义行(竖排) + 自动展开其输入 + for_test/*_tc 整块"
                         "(含横向 T 向量) + regmap/tmm 定义行。专为调单个信号设计。")
    ap.add_argument("--signal-expand", type=int, default=1,
                    help="--signal 时自动展开 logic 输入的层数(默认 1: 连同 testmode 等输入一起抠)。"
                         "设 0 则只抠目标信号本身。")
    ap.add_argument("--block-tail", type=int, default=120,
                    help="--signal 抠 for_test 类整块时, 块头后最多带多少行(默认 120)。")
    args = ap.parse_args()

    if not os.path.isfile(args.excel):
        sys.exit("找不到文件: %s" % args.excel)
    if not args.excel.lower().endswith(".xlsx"):
        sys.exit("只支持 .xlsx，请先在 Excel 里另存为 .xlsx")

    if args.expr_forms:
        run_expr_forms(args)
        return

    if args.signal:
        run_signal(args)
        return

    wb = openpyxl.load_workbook(args.excel, data_only=True, read_only=True)
    all_sheets = wb.sheetnames

    if args.sheets:
        want = [s.strip() for s in args.sheets.split(",") if s.strip()]
        sheets = [s for s in all_sheets if s in want]
        missing = [s for s in want if s not in all_sheets]
    else:
        prio = [s for s in PRIORITY_SHEETS if s in all_sheets]
        rest = [s for s in all_sheets if s not in prio]
        sheets = prio + rest
        missing = []

    header_lines = []
    header_lines.append("Dreg Excel 结构导出")
    header_lines.append("文件: %s" % os.path.basename(args.excel))
    header_lines.append("全部 sheet (%d): %s" % (len(all_sheets), all_sheets))
    header_lines.append("本次导出 sheet: %s" % sheets)
    if missing:
        header_lines.append("⚠ 指定但不存在的 sheet: %s" % missing)
    header_lines.append("开关: compact=%s, split=%s, anon_signals=%s, "
                        "mask_owners=%s, rows=%s, maxlen=%s"
                        % (args.compact, args.split, args.anon_signals,
                           args.mask_owners, args.rows, args.maxlen))

    base = os.path.splitext(args.excel)[0]

    if args.split:
        written = []
        for sn in sheets:
            out = list(header_lines)
            out.append("")
            dump_sheet(out, wb[sn], sn, args)
            path = "%s_inspect_%s.txt" % (base, safe_name(sn))
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(out))
            written.append((path, len(out)))
        wb.close()
        print("已分页导出 %d 个文件 (可分批发送):" % len(written))
        for p, n in written:
            print("  %s  (%d 行)" % (p, n))
        print("建议先发 *_logic.txt 给我。")
        return

    out = list(header_lines)
    for sn in sheets:
        out.append("")
        dump_sheet(out, wb[sn], sn, args)
    wb.close()

    out_path = args.out if args.out else base + "_inspect.txt"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(out))
    print("已导出: %s" % out_path)
    print("行数: %d" % len(out))
    print("请过目内容后，把 txt 全文贴给我。")


if __name__ == "__main__":
    main()
