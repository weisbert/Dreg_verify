# -*- coding: utf-8 -*-
"""make_mirror_excel.py — 镜像真实 Hi1108 WL_RFTRX 表的开发夹具 Excel。

目标（第二十三轮重建，2026-06-08）：用【真实信号名 + 真实地址 + 真实 mux/regmap/dft 结构】
复现公司机真表 `excel/Hi1108V100_WL_RFTRX_C0C1_DREG_to_dig_95P.xlsx` 上用户报的三个问题，
不依赖公司机即可在本地跑生成器复现。

数据来源（第二十三轮 inspect_excel --signal / --account 真抽）：
  · temp_code 上游 mux56(行113/114)：mode? local(RW) : tsensor(RO)，输出 d_wl_rf_temp_code，后缀 to_mux。
  · mixer2g_trim mux157(行1085-1097)：控制 = temp_code[3:1] 切片，8 选 1，**有重复行**(t0-t4 出现两遍)，后缀 to_dft。
  · mixer5g_trim mux158(行1098-1105)：控制 = temp_code[3:1]，干净 8 选 1，后缀 to_dft。
  · slna_1st_bias_trim_gain_cal mux152(行855-872)：18 行挤一个输出名 —— 前 9 行选 1st_bias 源、
    后 9 行选 2nd_bias 源，**case 完全相同**(gtune[3:0])→ 同 case 不同源【冲突】(designer 复制漏改输出名)。
  · lp5g_rxrf_lna_lctune mux156(行893-1084)：3 控制 8-bit case {rx5g_en, band_5g_sel[3:0], c_code[2:0]}，
    rx5g_off 高 band 段(b8+)的 case 值【错写成】跟 rx5g_on 低 band 段(b0+)一样 → 同 case 不同源【冲突】。
  · dft 页(行188/192/193/194)：上面 4 个输出都被 d_wl_rf_trx_reg_dft_iddq_mode 门控，表达式 B?0:A。
  真实地址：temp_code@d60(bit0=mode,7:4=local)；mixer2g_trim@d388；mixer5g_trim@d389/d390。

第二十四轮新增（band_trim mode=0 漏生成）：
  · band_trim mux(真表行1170-1173)：控制 = d_wl_rf_band_2g_sel[1:0] 切片，4 选 1 → b0..b3_trim，后缀 to_dft。
  · band_2g_sel 上游 mux(真表行15-16,N=7)：freq_sel_mode? local(RW) : linectrl_band_sel。
    关键差异(vs mixer5g_trim 的 tsensor)：mode=0 源 d_wl_rf_linectrl_band_sel 是【logic 页输出】(SECTION 3 实证)、
    经 _to_mux 衔接网喂上游 mux —— resolver 判它 needs-prefix，被载体扫描静默排除 → mode=0 那半张表无声漏掉。
    修后：没配探针前缀 → cascade_alt_skipped 记缺口；配了 scan_rtl 前缀(→prefixed-wire) → 正常出 mode=0。

它严格按 excel_model 的列约定造 5 页(logic/regmap/total_memory_map/mux/dft)。
for_test 页【留空】——真表这份文件里 for_test 也是空的(designer 真值表在另一个 .xlsx)。

用法:  python make_mirror_excel.py [输出路径=mirror_wl_dreg.xlsx]
"""

import sys

import openpyxl
from openpyxl.utils import column_index_from_string


def _set(ws, row, mapping):
    for col, val in mapping.items():
        ws.cell(row=row, column=column_index_from_string(col), value=val)


# ───────────── 地址（十进制；regmap 用 d<n>，tmm 用 h<X> 十六进制，两边一致）─────────────
ADDR_TEMP = 60           # temp_code: bit0=mode, 7:4=local
ADDR_TSENSOR = 160       # linectrl_tsensor RO（mode=0 line 路）
ADDR_M2G = 388           # mixer2g_trim t0..t7（2bit 打包进一个寄存器）
ADDR_M5G_A = 389         # mixer5g_trim t0..t3（3bit，4位对齐）
ADDR_M5G_B = 390         # mixer5g_trim t4..t7
ADDR_IDDQ = 0x47         # dft iddq_mode RO（门控）——真表 WL_DFT_IDDQ=h47（2026-06-10 探查实证；
                         # 地址影响输入行排序：真表上 iddq 在 temp_code(0x3C) 与 tsensor(0xA0) 之间）
ADDR_FREQSEL = 70        # freq_sel_mode RW（band_2g_sel 上游 mode 选择）
ADDR_BAND2G = 71         # band_2g_sel_local[3:0] RW（mode=1 主载体）
ADDR_PFB_TRIM = 72       # pfb b0..b3_trim 4×4bit 打包（band_trim 的数据源）
# 第二十四轮 audit 揭示的两个【新】形态（前半文件实证）：
ADDR_RX5GEN = 80         # rx5g_en_mode(bit0) + rx5g_en_local(bit1) RW —— *_en 型 mode mux
ADDR_GM = 81             # lp5g_gm_itrim off/on 数据源（4bit×2 打包）RW
ADDR_BWCTRL = 90         # bwctrl_mode(bit0) + bwctrl_local[2:1] RW —— 深层(第2层)级联的 mode mux
ADDR_BWLINE = 91         # linectrl_rfabb_bwctrl[1:0] RO【寄存器】(force 名即可,不需前缀——这是 [~] 与 [XX] 的关键差别)
ADDR_CORNER = 92         # lpf_corner_sel 数据源（2bit×2 打包）RW —— 中间层 mux 载体
ADDR_CMAIN = 93          # lpf_cmain 数据源（4bit×2 打包）RW —— 最下游数据


def _d(n):
    return "d%d" % n


def _h(n):
    return "h%X" % n


# ───────────── regmap：D=Reg_Name F=Reg Type G=Signal_Name H=地址 J..Y=bit15..bit0(Y=bit0) AE=owner ─────────────
def _regmap(ws, r, reg, typ, signal, bit_msb, bit_lsb, addr_dec, owner="WL"):
    _set(ws, r, {"D": reg, "F": typ, "G": signal, "H": _d(addr_dec), "AE": owner})
    for b in range(bit_lsb, bit_msb + 1):
        col = column_index_from_string("Y") - b      # Y=bit0
        ws.cell(row=r, column=col, value=1)
    return r + 1


# ───────────── total_memory_map：A=名 B=bit F=地址 D=DIG_TOP_PIN H=类型 ─────────────
def _tmm_reg(ws, r, name, addr_dec):
    _set(ws, r, {"A": name, "B": _h(addr_dec), "C": "0", "D": "RW", "E": "register def"})
    return r + 1


def _tmm_field(ws, r, name, bit, addr_dec, dig, typ, desc=""):
    _set(ws, r, {"A": name, "B": bit, "C": "0", "D": dig, "E": desc or "field",
                 "F": _h(addr_dec), "H": typ})
    return r + 1


# ───────────── mux：A=input B~E=ctrl F=case G=out H=去向 I=top L=owner N=组号 ─────────────
def _mrow(ws, r, mux_input, ctrls, case, out, group_no, dest="to_mux", owner="WL", top=0):
    """ctrls: 单个控制名(str) 或 控制名列表(写入 B/C/D/E)。"""
    if isinstance(ctrls, str):
        ctrls = [ctrls]
    m = {"A": mux_input, "F": case, "G": out, "H": dest, "I": top, "L": owner, "N": group_no}
    for i, c in enumerate(ctrls):
        m[chr(ord("B") + i)] = c          # B,C,D,E
    _set(ws, r, m)
    return r + 1


def build(path):
    wb = openpyxl.Workbook()

    # ========================= logic 页 =========================
    #   4 个问题信号都是 mux 输出，logic 不直接产它们；但 band_2g_sel 的 mode=0 备用源
    #   d_wl_rf_linectrl_band_sel 是【logic 页输出】(SECTION 3 实证：在 logic.K，M 列含 to_mux)——
    #   不是 RO 寄存器。这正是 band_trim mode=0 漏生成的根因：它被 resolver 判 needs-prefix(logic→mux 衔接网)。
    ws = wb.active
    ws.title = "logic"
    _set(ws, 1, {"A": "<<输入信号(列字母=表达式变量)>>", "K": "<<输出>>", "L": "<<真值表达式>>"})
    _set(ws, 2, {"A": "in_A", "B": "in_B", "K": "logic输出名", "L": "表达式",
                 "M": "type", "N": "top_output", "O": "Notes", "P": "owner", "R": "no"})
    # linectrl_band_sel：线控 band 选择，由管脚线控输入透传 → 喂 mux(to_mux)。top_output=0(内部网)。
    _set(ws, 3, {"A": "d_wl_rf_band_sel_line_to_logic[3:0]", "K": "d_wl_rf_linectrl_band_sel[3:0]",
                 "L": "A", "M": "to_mux", "N": "0", "O": "线控 band 选择(管脚)", "P": "yangteng", "R": "1"})
    # rx5g_en_line：*_en 型 mode mux 的 mode=0 源，也是 logic→mux 网，但命名是 *_line(非 linectrl_*)——
    # 验证修复对【非 linectrl 命名】的线控源也成立(audit [XX] 里 rx5g_en_line/txcal5g_en_line/... 这一族)。
    _set(ws, 4, {"A": "d_wl_rf_rx5g_en_pin_to_logic", "K": "d_wl_rf_rx5g_en_line",
                 "L": "A", "M": "to_mux", "N": "0", "O": "线控 rx5g 使能(管脚)", "P": "shenzheng", "R": "2"})

    # ========================= mux 页 =========================
    mx = wb.create_sheet("mux")
    _set(mx, 2, {"A": "mux_input", "B": "ctrl1", "C": "ctrl2", "D": "ctrl3", "F": "case",
                 "G": "mux_out", "H": "dest", "I": "top", "L": "owner", "N": "组号"})
    r = 3

    # ── mux56：temp_code[3:0] = case(temp_code_mode){0:线控 tsensor(RO); 1:temp_code_local(RW)}，后缀 to_mux ──
    r = _mrow(mx, r, "d_wl_rf_linectrl_tsensor_to_mux[3:0]", "d_wl_rf_temp_code_mode_to_mux",
              "1'b0", "d_wl_rf_temp_code[3:0]", 56, dest="to_mux", owner="yangteng")
    r = _mrow(mx, r, "d_wl_rf_temp_code_local_to_mux[3:0]", "d_wl_rf_temp_code_mode_to_mux",
              "1'b1", "d_wl_rf_temp_code[3:0]", 56, dest="to_mux", owner="yangteng")

    # ── mux152：slna_1st_bias_trim_gain_cal —— 18 行挤一个输出名（同 case 不同源【冲突】）──
    #    前 9 行选 1st_bias g8..g0，后 9 行选 2nd_bias g8..g0，gtune[3:0] case 完全相同。
    def _gtune_case(k):              # g8→1xxx；g7..g0→0111..0000
        return "4'b1xxx" if k == 8 else "4'b" + format(k, "04b")
    for fam in ("1st", "2nd"):
        for k in (8, 7, 6, 5, 4, 3, 2, 1, 0):
            r = _mrow(mx, r,
                      "d_bt_rx_slna_%s_bias_trim_gain_cal_g%d_wl_to_mux[3:0]" % (fam, k),
                      "d_bt_slna_gtune_wl_to_mux[3:0]", _gtune_case(k),
                      "d_bt_rx_slna_1st_bias_trim_gain_cal_wl[3:0]", 152,
                      dest="to_dft", owner="Konglingshan")

    # ── mux156：lp5g_rxrf_lna_lctune —— 3 控制 8-bit case，rx5g_off 高 band 段 case 错写成 on 低 band 段（冲突）──
    #    case 编码 = {rx5g_en[7], band_5g_sel[6:3], c_code[2:0]}；缩小版：band {0} on + {0} off + {8} off(buggy)。
    lctune_ctrls = ["d_wl_rf_rx5g_en_to_mux",
                    "d_wl_rf_band_5g_sel_to_mux[3:0]",
                    "d_wl_rf_c_code_to_mux[2:0]"]

    def _lctune_row(band, corner, onoff, buggy=False):
        nonlocal r
        rx5g_en = 1 if onoff == "on" else 0
        case_band = (band - 8) if buggy else band         # buggy: off 高 band 复用 on 低 band 的 case 空间
        case_en = 1 if buggy else rx5g_en                 # buggy: bit7 错写成 1（=on）
        val = (case_en << 7) | ((case_band & 0xF) << 3) | (corner & 0x7)
        case = "8'b" + format(val, "08b")
        src = "d_wl_rf_lp5g_rx_lna_lctune_lut_b%d_c%d_rx5g_%s_to_mux[5:0]" % (band, corner, onoff)
        r = _mrow(mx, r, src, lctune_ctrls, case,
                  "d_wl_rf_lp5g_rxrf_lna_lctune[5:0]", 156, dest="to_dft", owner="shenzheng")
    for corner in range(5):                       # b0 rx5g_on（case bit7=1, band=0）
        _lctune_row(0, corner, "on")
    for corner in range(5):                       # b0 rx5g_off（case bit7=0, band=0）—— 正确
        _lctune_row(0, corner, "off")
    for corner in range(5):                       # b8 rx5g_off（case 错写成 bit7=1, band=0）→ 撞 b0_on
        _lctune_row(8, corner, "off", buggy=True)

    # ── mux157：mixer2g_trim[1:0] = case(temp_code[3:1]) 8 选 1，**含重复行**(t0-t4 出现两遍)，后缀 to_dft ──
    def _m2g(k):
        nonlocal r
        r = _mrow(mx, r, "d_wl_rf_lo2g5g_mixer2g_trim_t%d_to_mux[1:0]" % k,
                  "d_wl_rf_temp_code_to_mux[3:1]", "3'b" + format(k, "03b"),
                  "d_wl_rf_lo2g5g_mixer2g_trim[1:0]", 157, dest="to_dft", owner="law/chenhao")
    for k in range(5):    # t0..t4
        _m2g(k)
    for k in range(5):    # t0..t4 再来一遍（真表的重复行 → 死分支去重）
        _m2g(k)
    for k in (5, 6, 7):   # t5..t7
        _m2g(k)

    # ── mux158：mixer5g_trim[2:0] = case(temp_code[3:1]) 干净 8 选 1，后缀 to_dft ──
    for k in range(8):
        r = _mrow(mx, r, "d_wl_rf_lo2g5g_mixer5g_trim_t%d_to_mux[2:0]" % k,
                  "d_wl_rf_temp_code_to_mux[3:1]", "3'b" + format(k, "03b"),
                  "d_wl_rf_lo2g5g_mixer5g_trim[2:0]", 158, dest="to_dft", owner="law/chenhao")

    # ── mux7：band_2g_sel[3:0] = case(freq_sel_mode){0:线控 linectrl_band_sel(logic→mux 网); 1:band_2g_sel_local(RW)} ──
    #    与 mux56(temp_code) 同形，但 mode=0 源是【logic 页输出】而非 RO 寄存器 —— 这是 band_trim mode=0 漏生成的根因。
    r = _mrow(mx, r, "d_wl_rf_linectrl_band_sel_to_mux[3:0]", "d_wl_rf_freq_sel_mode_to_mux",
              "1'b0", "d_wl_rf_band_2g_sel[3:0]", 7, dest="to_mux", owner="yangteng")
    r = _mrow(mx, r, "d_wl_rf_band_2g_sel_local_to_mux[3:0]", "d_wl_rf_freq_sel_mode_to_mux",
              "1'b1", "d_wl_rf_band_2g_sel[3:0]", 7, dest="to_mux", owner="yangteng")

    # ── mux200：lcbufc0_2g_pfb_band_trim[3:0] = case(band_2g_sel[1:0] 切片) 4 选 1 → b0..b3_trim，后缀 to_dft ──
    #    控制是上游 mux7 输出的 [1:0] 切片（两级级联：band_trim ← band_2g_sel ← freq_sel_mode）。
    for k in range(4):
        r = _mrow(mx, r, "d_wl_rf_lo2g5g_lcbufc0_2g_pfb_b%d_trim_to_mux[3:0]" % k,
                  "d_wl_rf_band_2g_sel_to_mux[1:0]", "2'b" + format(k, "02b"),
                  "d_wl_rf_lo2g5g_lcbufc0_2g_pfb_band_trim[3:0]", 200, dest="to_dft", owner="law/chenhao")

    # ── 【新形态A：*_en 型 mode mux + *_line 线控源】mux32 rx5g_en = rx5g_en_mode? local(RW) : rx5g_en_line(logic→mux) ──
    #    下游 gm_itrim 由 rx5g_en(1bit) 级联控制 → [XX] mode=0 缺失(rx5g_en_line 需前缀)；命名 *_line 非 linectrl_*。
    r = _mrow(mx, r, "d_wl_rf_rx5g_en_line_to_mux", "d_wl_rf_rx5g_en_mode_to_mux",
              "1'b0", "d_wl_rf_rx5g_en", 32, dest="to_mux", owner="shenzheng")
    r = _mrow(mx, r, "d_wl_rf_rx5g_en_local_to_mux", "d_wl_rf_rx5g_en_mode_to_mux",
              "1'b1", "d_wl_rf_rx5g_en", 32, dest="to_mux", owner="shenzheng")
    for k, onoff in ((0, "off"), (1, "on")):
        r = _mrow(mx, r, "d_wl_rf_lp5g_gm_itrim_%s_to_mux[3:0]" % onoff,
                  "d_wl_rf_rx5g_en_to_mux", "1'b%d" % k,
                  "d_wl_rf_lp5g_gm_itrim[3:0]", 201, dest="to_dft", owner="shenzheng")

    # ── 【新形态B：深层(第2层)级联 + RO 寄存器 mode=0 源】lpf_cmain ← lpf_corner_sel ← bwctrl(mode mux) ──
    #    mode=0 源 linectrl_rfabb_bwctrl 是【RO 寄存器】(force 名即可,不需前缀)——audit 标 [~]：
    #    根因不是前缀，而是两分支只展开【最近一层】，深层 mode mux 的 mode=0 没被驱动。这是【代码可修】的桶。
    r = _mrow(mx, r, "d_wl_rf_linectrl_rfabb_bwctrl_to_mux[1:0]", "d_wl_rf_bwctrl_mode_to_mux",
              "1'b0", "d_wl_rf_bwctrl[1:0]", 57, dest="to_mux", owner="luqi")
    r = _mrow(mx, r, "d_wl_rf_bwctrl_local_to_mux[1:0]", "d_wl_rf_bwctrl_mode_to_mux",
              "1'b1", "d_wl_rf_bwctrl[1:0]", 57, dest="to_mux", owner="luqi")
    for k in range(2):    # 中间层 mux58：lpf_corner_sel[1:0] = case(bwctrl[1:0]) → corner RW 源
        r = _mrow(mx, r, "d_wl_rf_lpf_corner%d_to_mux[1:0]" % k,
                  "d_wl_rf_bwctrl_to_mux[1:0]", "2'b" + format(k, "02b"),
                  "d_wl_rf_lpf_corner_sel[1:0]", 58, dest="to_mux", owner="luqi")
    for k in range(2):    # 最下游 mux202：lpf_cmain[3:0] = case(lpf_corner_sel[0] 切片) → cmain RW 源
        r = _mrow(mx, r, "d_wl_rf_lpf_cmain%d_to_mux[3:0]" % k,
                  "d_wl_rf_lpf_corner_sel_to_mux[0]", "1'b%d" % k,
                  "d_wl_rf_lpf_cmain[3:0]", 202, dest="to_dft", owner="luqi")

    # ========================= dft 页（4 个输出被 iddq_mode 门控，B?0:A）=========================
    dft = wb.create_sheet("dft")
    _set(dft, 2, {"A": "A", "B": "gate(to_dft)", "C": "C", "D": "observed_out", "E": "gate_expr"})
    r = 3
    for out_sig in ("d_bt_rx_slna_1st_bias_trim_gain_cal_wl[3:0]",
                    "d_wl_rf_lp5g_rxrf_lna_lctune[5:0]",
                    "d_wl_rf_lo2g5g_mixer2g_trim[1:0]",
                    "d_wl_rf_lo2g5g_mixer5g_trim[2:0]",
                    "d_wl_rf_lo2g5g_lcbufc0_2g_pfb_band_trim[3:0]"):   # 真表 dft 行 203 也门控它
        _set(dft, r, {"A": out_sig.split("[")[0] + "_to_dft",
                      "B": "d_wl_rf_trx_reg_dft_iddq_mode_to_dft",
                      "D": out_sig, "E": "B?0:A"})
        r += 1

    # ========================= regmap 页 =========================
    rm = wb.create_sheet("regmap")
    _set(rm, 2, {"D": "Reg_Name", "F": "Reg Type", "G": "Signal_Name",
                 "H": "addr", "I": "default", "J": "b15", "Y": "b0", "AE": "owner"})
    r = 3
    r = _regmap(rm, r, "WL_TEMP_ctrl", "RW", "d_wl_rf_temp_code_mode", 0, 0, ADDR_TEMP, "yangteng")
    r = _regmap(rm, r, "WL_TEMP_ctrl", "RW", "d_wl_rf_temp_code_local[3:0]", 7, 4, ADDR_TEMP, "yangteng")
    r = _regmap(rm, r, "TSENSOR_LINE", "RO", "d_wl_rf_linectrl_tsensor", 3, 0, ADDR_TSENSOR, "yangteng")
    # mixer2g_trim 8 源打包进 d388（bit1:0,3:2,...,15:14）
    for k in range(8):
        r = _regmap(rm, r, "MIXER_CTRL1", "RW", "d_wl_rf_lo2g5g_mixer2g_trim_t%d[1:0]" % k,
                    2 * k + 1, 2 * k, ADDR_M2G, "law/chenhao")
    # mixer5g_trim 8 源 3bit、4位对齐：t0..t3@d389 / t4..t7@d390
    for k in range(4):
        r = _regmap(rm, r, "MIXER_CTRL2", "RW", "d_wl_rf_lo2g5g_mixer5g_trim_t%d[2:0]" % k,
                    4 * k + 2, 4 * k, ADDR_M5G_A, "law/chenhao")
    for k in range(4, 8):
        r = _regmap(rm, r, "MIXER_CTRL3", "RW", "d_wl_rf_lo2g5g_mixer5g_trim_t%d[2:0]" % k,
                    4 * (k - 4) + 2, 4 * (k - 4), ADDR_M5G_B, "law/chenhao")
    r = _regmap(rm, r, "DFT_IDDQ", "RO", "d_wl_rf_trx_reg_dft_iddq_mode", 0, 0, ADDR_IDDQ, "yangteng")
    # band_trim 级联链：freq_sel_mode(RW) + band_2g_sel_local(RW)；mode=0 源 linectrl_band_sel 不在此(它是 logic 输出)
    r = _regmap(rm, r, "LO_CTRL_mode", "RW", "d_wl_rf_freq_sel_mode", 0, 0, ADDR_FREQSEL, "yangteng")
    r = _regmap(rm, r, "BAND2G_LOCAL", "RW", "d_wl_rf_band_2g_sel_local[3:0]", 3, 0, ADDR_BAND2G, "yangteng")
    for k in range(4):    # b0..b3_trim 4bit 打包进 d72（band_trim 的数据源，RW）
        r = _regmap(rm, r, "PFB_TRIM", "RW", "d_wl_rf_lo2g5g_lcbufc0_2g_pfb_b%d_trim[3:0]" % k,
                    4 * k + 3, 4 * k, ADDR_PFB_TRIM, "law/chenhao")
    # 新形态A：rx5g_en mode mux + gm_itrim 数据
    r = _regmap(rm, r, "RX5GEN", "RW", "d_wl_rf_rx5g_en_mode", 0, 0, ADDR_RX5GEN, "shenzheng")
    r = _regmap(rm, r, "RX5GEN", "RW", "d_wl_rf_rx5g_en_local", 1, 1, ADDR_RX5GEN, "shenzheng")
    r = _regmap(rm, r, "GM", "RW", "d_wl_rf_lp5g_gm_itrim_off[3:0]", 3, 0, ADDR_GM, "shenzheng")
    r = _regmap(rm, r, "GM", "RW", "d_wl_rf_lp5g_gm_itrim_on[3:0]", 7, 4, ADDR_GM, "shenzheng")
    # 新形态B：bwctrl mode mux(深层) + 中间/下游数据；linectrl_rfabb_bwctrl 是 RO【寄存器】(force 名即可)
    r = _regmap(rm, r, "BWCTRL", "RW", "d_wl_rf_bwctrl_mode", 0, 0, ADDR_BWCTRL, "luqi")
    r = _regmap(rm, r, "BWCTRL", "RW", "d_wl_rf_bwctrl_local[1:0]", 2, 1, ADDR_BWCTRL, "luqi")
    r = _regmap(rm, r, "BWLINE", "RO", "d_wl_rf_linectrl_rfabb_bwctrl[1:0]", 1, 0, ADDR_BWLINE, "luqi")
    for k in range(2):
        r = _regmap(rm, r, "CORNER", "RW", "d_wl_rf_lpf_corner%d[1:0]" % k, 2 * k + 1, 2 * k, ADDR_CORNER, "luqi")
    for k in range(2):
        r = _regmap(rm, r, "CMAIN", "RW", "d_wl_rf_lpf_cmain%d[3:0]" % k, 4 * k + 3, 4 * k, ADDR_CMAIN, "luqi")

    # ========================= total_memory_map 页 =========================
    tmm = wb.create_sheet("total_memory_map")
    r = 1
    r = _tmm_reg(tmm, r, "WL_TEMP_ctrl", ADDR_TEMP)
    r = _tmm_field(tmm, r, "d_wl_rf_temp_code_mode", "0", ADDR_TEMP, "N", "RW", "Tsensor mode select(1:RF reg 0:line)")
    r = _tmm_field(tmm, r, "d_wl_rf_temp_code_local[3:0]", "7:4", ADDR_TEMP, "N", "RW", "Temperature Code local 4bit")
    r = _tmm_reg(tmm, r, "TSENSOR_LINE", ADDR_TSENSOR)
    r = _tmm_field(tmm, r, "d_wl_rf_linectrl_tsensor", "3:0", ADDR_TSENSOR, "Y", "RO", "温度码线控(管脚,只读,force)")
    r = _tmm_reg(tmm, r, "MIXER_CTRL1", ADDR_M2G)
    for k in range(8):
        r = _tmm_field(tmm, r, "d_wl_rf_lo2g5g_mixer2g_trim_t%d" % k,
                       "%d:%d" % (2 * k + 1, 2 * k), ADDR_M2G, "N", "RW", "lo2g5g mixer2g gain trim t%d" % k)
    r = _tmm_reg(tmm, r, "MIXER_CTRL2", ADDR_M5G_A)
    for k in range(4):
        r = _tmm_field(tmm, r, "d_wl_rf_lo2g5g_mixer5g_trim_t%d" % k,
                       "%d:%d" % (4 * k + 2, 4 * k), ADDR_M5G_A, "N", "RW", "lo2g5g mixer5g gain trim t%d" % k)
    r = _tmm_reg(tmm, r, "MIXER_CTRL3", ADDR_M5G_B)
    for k in range(4, 8):
        r = _tmm_field(tmm, r, "d_wl_rf_lo2g5g_mixer5g_trim_t%d" % k,
                       "%d:%d" % (4 * (k - 4) + 2, 4 * (k - 4)), ADDR_M5G_B, "N", "RW",
                       "lo2g5g mixer5g gain trim t%d" % k)
    r = _tmm_reg(tmm, r, "DFT_IDDQ", ADDR_IDDQ)
    r = _tmm_field(tmm, r, "d_wl_rf_trx_reg_dft_iddq_mode", "0", ADDR_IDDQ, "Y", "RO", "IDDQ DFT 门控(管脚,只读,force)")
    # band_trim 级联链寄存器
    r = _tmm_reg(tmm, r, "LO_CTRL_mode", ADDR_FREQSEL)
    r = _tmm_field(tmm, r, "d_wl_rf_freq_sel_mode", "0", ADDR_FREQSEL, "N", "RW", "WL freq mode(1:RF reg 0:line control)")
    r = _tmm_reg(tmm, r, "BAND2G_LOCAL", ADDR_BAND2G)
    r = _tmm_field(tmm, r, "d_wl_rf_band_2g_sel_local[3:0]", "3:0", ADDR_BAND2G, "N", "RW", "band 2g select local(mode=1)")
    r = _tmm_reg(tmm, r, "PFB_TRIM", ADDR_PFB_TRIM)
    for k in range(4):
        r = _tmm_field(tmm, r, "d_wl_rf_lo2g5g_lcbufc0_2g_pfb_b%d_trim" % k,
                       "%d:%d" % (4 * k + 3, 4 * k), ADDR_PFB_TRIM, "N", "RW",
                       "lcbufc0 2g pfb b%d trim" % k)
    # 新形态A：rx5g_en mode mux + gm_itrim
    r = _tmm_reg(tmm, r, "RX5GEN", ADDR_RX5GEN)
    r = _tmm_field(tmm, r, "d_wl_rf_rx5g_en_mode", "0", ADDR_RX5GEN, "N", "RW", "rx5g en mode(1:reg 0:line)")
    r = _tmm_field(tmm, r, "d_wl_rf_rx5g_en_local", "1", ADDR_RX5GEN, "N", "RW", "rx5g en local")
    r = _tmm_reg(tmm, r, "GM", ADDR_GM)
    r = _tmm_field(tmm, r, "d_wl_rf_lp5g_gm_itrim_off[3:0]", "3:0", ADDR_GM, "N", "RW", "lp5g gm itrim off")
    r = _tmm_field(tmm, r, "d_wl_rf_lp5g_gm_itrim_on[3:0]", "7:4", ADDR_GM, "N", "RW", "lp5g gm itrim on")
    # 新形态B：bwctrl mode mux(深层) + linectrl_rfabb_bwctrl 是 RO 寄存器(force 名即可,不需前缀) + 中间/下游
    r = _tmm_reg(tmm, r, "BWCTRL", ADDR_BWCTRL)
    r = _tmm_field(tmm, r, "d_wl_rf_bwctrl_mode", "0", ADDR_BWCTRL, "N", "RW", "rfabb bwctrl mode(1:reg 0:line)")
    r = _tmm_field(tmm, r, "d_wl_rf_bwctrl_local[1:0]", "2:1", ADDR_BWCTRL, "N", "RW", "rfabb bwctrl local")
    r = _tmm_reg(tmm, r, "BWLINE", ADDR_BWLINE)
    r = _tmm_field(tmm, r, "d_wl_rf_linectrl_rfabb_bwctrl[1:0]", "1:0", ADDR_BWLINE, "Y", "RO", "rfabb bwctrl 线控(RO寄存器,force名)")
    r = _tmm_reg(tmm, r, "CORNER", ADDR_CORNER)
    for k in range(2):
        r = _tmm_field(tmm, r, "d_wl_rf_lpf_corner%d[1:0]" % k, "%d:%d" % (2 * k + 1, 2 * k),
                       ADDR_CORNER, "N", "RW", "lpf corner sel src %d" % k)
    r = _tmm_reg(tmm, r, "CMAIN", ADDR_CMAIN)
    for k in range(2):
        r = _tmm_field(tmm, r, "d_wl_rf_lpf_cmain%d[3:0]" % k, "%d:%d" % (4 * k + 3, 4 * k),
                       ADDR_CMAIN, "N", "RW", "lpf cmain data %d" % k)

    # ========================= for_test 页（留空：真表这份文件 for_test 也是空的）=========================
    ft = wb.create_sheet("for_test")
    _set(ft, 1, {"A": "case", "B": "输入寄存器地址", "C": "value", "D": "需要验证的顶层信号",
                 "E": "预计输出", "F": "验证信号的输入信号"})
    ft.cell(row=2, column=1, value="（本文件 for_test 为空 —— designer 真值表在另一个 .xlsx；§4 回填对照目标）")

    # ========================= 其余代表性 stub（工具不读）=========================
    _stub_sheet(wb, "iddq",
                ["A", "B", "C", "out put signal", "logic expression", "abserve", "dft_input"],
                [["", "", "", "(本表问题信号的门控在 dft 页，不在 iddq 页)", "", "", ""]])
    _stub_sheet(wb, "level_shift",
                ["level_shift_input", "level", "level_shift_output_name", "suffix", "top_out", "Notes", "Owner"],
                [["d_wl_rf_temp_code_mode", "down", "d_wl_rf_temp_code_mode_ls", "ls", "0", "stub", "WL"]])
    _stub_sheet(wb, "Topout",
                ["Owner", "TOP Out Signal", "Reg Name", "Offset", "Reg Description", "RegAddress"],
                [["WL", "(本表 4 个问题信号 top_out=0，非顶层输出)", "", "", "", ""]])

    wb.save(path)
    return path


def _stub_sheet(wb, name, headers, samples):
    ws = wb.create_sheet(name)
    ws.cell(row=1, column=1, value="【STUB】%s —— 工具不读此页；代表性示意" % name)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    for ri, row in enumerate(samples, start=3):
        for c, v in enumerate(row, start=1):
            ws.cell(row=ri, column=c, value=v)
    return ws


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "mirror_wl_dreg.xlsx"
    build(out)
    print("已生成镜像 Excel:", out)
