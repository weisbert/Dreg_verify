# -*- coding: utf-8 -*-
"""
inspect_fortest.py — 摸清 for_test(测试用例)页的真实排版。

背景：我们之前的工具假设 for_test 是「A 列=输出名、每输出一个行块」。但 Hi1107C WL
真表跑全清单时，A 列出现的是 case / aa / bb / case2 / cas / aes 这种『标签』，而不是
信号名 —— 说明这页排版和假设不同：信号名八成在别的列，或者用了合并单元格(竖向合并的
单元格除左上角外，openpyxl 读出来全是 None，会被误判成『一个超长块』)。

这个脚本只做一件事：把 for_test 页的真实列结构摸清楚 ——
  · 哪列是信号名(对照 logic.K 基名 / d_ 前缀)；
  · 有没有合并单元格、合并在哪几列、跨多少行(解释『9582 行一个块』之谜)；
  · 每列的表头 + 取值样本(认出 T0..T63 向量列 / 地址列 / 类型列)；
  · 几处『锚点行』整行竖排(看一条真实测试用例长什么样)。

只读，不碰任何产品代码(dreg_verify/)。输出写 UTF-8 文件 —— Windows 控制台是 GBK，
直接 print 中文会变成 ?，所以一律写 .txt，跑完把文件全文贴回来即可。

用法(在装了 openpyxl 的那个 python 环境里跑)：
    python inspect_fortest.py 真表.xlsx
        [--sheet for_test]    目标页(默认 for_test；找不到就列出可选页名退出)
        [--at 4,76,419]       额外锚点行号：在这些行竖排 dump 整行(看真实用例)
        [--span 5]            每个锚点往下连带 dump 多少行(默认 5)
        [--cols 80]           列画像/扫描最多看前多少列(默认 80，T 向量列很多)
        [--rows 10000]        最多扫描多少行(默认 10000)
        [--out 文件]          默认 <excel名>_fortest.txt

依赖：openpyxl。仅支持 .xlsx。
"""
import argparse
import os
import re
import sys

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass


def safe_print(s):
    try:
        print(s)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, "encoding", None) or "ascii"
        print(s.encode(enc, "replace").decode(enc, "replace"))


try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少 openpyxl，请先运行:  pip install openpyxl")


# ───────────────────────────── 基础工具 ─────────────────────────────
def cell_str(value, maxlen=0):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        s = str(int(value))
    else:
        s = str(value)
    s = s.replace("\n", "\\n").replace("\r", "")
    if maxlen and len(s) > maxlen:
        s = s[:maxlen] + "…(+%d)" % (len(s) - maxlen)
    return s


def strip_sig(name):
    """去位宽 [..] 与 _to_logic/_to_mux/_to_dft 后缀，返回小写基名，用于跨页匹配。"""
    s = "" if name is None else str(name).strip()
    s = re.sub(r"\[[^\]]*\]", "", s)
    for suf in ("_to_logic", "_to_mux", "_to_dft"):
        if s.lower().endswith(suf):
            s = s[: -len(suf)]
    return s.strip().lower()


def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


# ───────────────────────────── 读页(含合并单元格还原) ─────────────────────────────
def load_grid(ws, max_rows, max_cols):
    """整页读成 grid[r][c]（0 基）；并把竖向/横向合并单元格的值『还原』到合并区每个格子，
    这样列画像与信号检测不会因合并而漏掉真实值。返回 (grid, merges)。
    merges: [(min_row,max_row,min_col,max_col,value)]（1 基行列），仅含跨多格的真合并。"""
    maxr = min(ws.max_row or 0, max_rows)
    maxc = min(ws.max_column or 0, max_cols)
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=maxr, max_col=maxc, values_only=True):
        rows.append(list(r))
    # 规整成定长
    grid = []
    for r in rows:
        if len(r) < maxc:
            r = r + [None] * (maxc - len(r))
        grid.append(r)
    merges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row > maxr or mr.min_col > maxc:
            continue
        av = grid[mr.min_row - 1][mr.min_col - 1] if mr.min_row - 1 < len(grid) else None
        if mr.max_row > mr.min_row or mr.max_col > mr.min_col:
            merges.append((mr.min_row, mr.max_row, mr.min_col, mr.max_col, av))
        for rr in range(mr.min_row, min(mr.max_row, maxr) + 1):
            for cc in range(mr.min_col, min(mr.max_col, maxc) + 1):
                grid[rr - 1][cc - 1] = av
    return grid, merges, maxr, maxc


def collect_logic_k(wb):
    """logic 页 K 列(输出名)基名集合，用于认信号名列。无 logic 页返回空集。"""
    name = next((s for s in wb.sheetnames if s.lower() == "logic"), None)
    if not name:
        return set()
    ws = wb[name]
    out = set()
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 20000),
                                          values_only=True), start=1):
        if ri <= 2:
            continue
        v = row[10] if len(row) > 10 else None   # K 列
        b = strip_sig(v)
        if b:
            out.add(b)
    return out


# ───────────────────────────── SECTION 0：sheet 总览 ─────────────────────────────
def dump_overview(out, wb, sheet):
    out.append("=" * 72)
    out.append("==== SECTION 0: 全部 sheet 总览 ====")
    for sn in wb.sheetnames:
        ws = wb[sn]
        mark = "  ← 目标页" if sn == sheet else ""
        out.append("  %-22s %s 行 x %s 列%s" % (sn, ws.max_row, ws.max_column, mark))


# ───────────────────────────── SECTION 1：原文表头 ─────────────────────────────
def dump_raw_head(out, grid, maxc, n=4):
    out.append("")
    out.append("--- 1a. 前 %d 行原文(横排，判断表头与列含义) ---" % n)
    for r in range(min(n, len(grid))):
        cells = []
        for c in range(maxc):
            v = grid[r][c]
            if is_blank(v):
                continue
            cells.append("%s=%s" % (get_column_letter(c + 1), cell_str(v, 40)))
        out.append("[row %d] %s" % (r + 1, " | ".join(cells)))


# ───────────────────────────── SECTION 1b：合并单元格 ─────────────────────────────
def dump_merges(out, merges, maxc, head_cols=10):
    out.append("")
    out.append("--- 1b. 合并单元格(只看前 %d 列；解释『一个块好几千行』之谜) ---" % head_cols)
    vert = [m for m in merges if m[2] <= head_cols and (m[1] - m[0]) >= 1]
    if not vert:
        out.append("  (前 %d 列没有跨行合并 —— 信号名不是靠合并铺开的)" % head_cols)
        return
    vert.sort(key=lambda m: (m[2], m[0]))
    shown = 0
    for (r0, r1, c0, c1, v) in vert:
        span = r1 - r0 + 1
        colrng = get_column_letter(c0) + (("-" + get_column_letter(c1)) if c1 > c0 else "")
        out.append("  col %-4s 行 %d..%d (%d 行合并): %s"
                   % (colrng, r0, r1, span, cell_str(v, 50)))
        shown += 1
        if shown >= 80:
            out.append("  ...(合并区过多，截断前 80 个)")
            break
    out.append("  合计跨行合并块: %d" % len(vert))


# ───────────────────────────── SECTION 1c：列画像 + 信号名列检测 ─────────────────────────────
def dump_column_profile(out, grid, maxr, maxc, logic_k, header_rows=3):
    out.append("")
    out.append("--- 1c. 列画像(表头取前 %d 行；数据从第 %d 行起统计) ---"
               % (header_rows, header_rows + 1))
    sig_candidates = []   # (col_letter, hit_logic, hit_dprefix, nonempty)
    for c in range(maxc):
        letter = get_column_letter(c + 1)
        hdr = []
        for hr in range(min(header_rows, len(grid))):
            hv = cell_str(grid[hr][c], 24)
            if hv:
                hdr.append(hv)
        vals = []
        for r in range(header_rows, maxr):
            v = grid[r][c] if c < len(grid[r]) else None
            if is_blank(v):
                continue
            vals.append(cell_str(v))
        if not vals and not hdr:
            continue
        distinct = list(dict.fromkeys(vals))
        # 信号名检测
        bases = {strip_sig(v) for v in distinct}
        bases.discard("")
        hit_logic = len(bases & logic_k)
        hit_dpref = sum(1 for b in bases if b.startswith("d_"))
        if vals and (hit_logic >= 3 or hit_dpref >= 3):
            sig_candidates.append((letter, hit_logic, hit_dpref, len(vals), len(distinct)))
        sample = distinct[:12]
        hdr_s = " / ".join(hdr) if hdr else "-"
        tag = ""
        if hit_logic or hit_dpref:
            tag = "  [信号名? logic.K命中=%d, d_前缀=%d]" % (hit_logic, hit_dpref)
        out.append("  col %-4s [表头:%s] 非空=%d 去重=%d%s"
                   % (letter, hdr_s, len(vals), len(distinct), tag))
        out.append("        样本=%s" % sample)
    out.append("")
    out.append("--- 1d. 信号名列候选(降序) ---")
    if sig_candidates:
        sig_candidates.sort(key=lambda x: -(x[1] * 1000 + x[2]))
        for letter, hl, hd, ne, dis in sig_candidates:
            out.append("  ★ col %s: logic.K命中=%d, d_前缀=%d, 非空=%d, 去重=%d"
                       % (letter, hl, hd, ne, dis))
        out.append("  → 最可能的信号名列: col %s" % sig_candidates[0][0])
    else:
        out.append("  ⚠ 没有任何列像信号名(logic.K命中/ d_ 前缀都不足) —— "
                   "可能信号名带前缀或格式特殊，看 1a/1e 原文再判断。")


# ───────────────────────────── SECTION 2：锚点行竖排 ─────────────────────────────
def dump_anchors(out, grid, maxr, maxc, header_rows, anchors, span):
    out.append("")
    out.append("=" * 72)
    out.append("==== SECTION 2: 锚点行整行竖排(看真实测试用例) ====")
    h = grid[header_rows - 1] if len(grid) >= header_rows else []
    for a in anchors:
        if a < 1 or a > maxr:
            out.append("")
            out.append("[锚点行 %d 越界(总 %d 行)，跳过]" % (a, maxr))
            continue
        out.append("")
        out.append("---- 锚点行 %d (连带往下 %d 行) ----" % (a, span))
        for r in range(a - 1, min(a - 1 + span, maxr)):
            cells_any = False
            line = ["  [row %d]" % (r + 1)]
            for c in range(maxc):
                v = grid[r][c] if c < len(grid[r]) else None
                if is_blank(v):
                    continue
                cells_any = True
                hname = cell_str(h[c] if c < len(h) else None, 24)
                line.append("      %-4s [%s]: %s" % (get_column_letter(c + 1), hname, cell_str(v)))
            if cells_any:
                out.append("\n".join(line))


# ───────────────────────────── SECTION 3：按测试组(每输出一组) ─────────────────────────────
# for_test 列固定语义(0 基)：A=0 case标记, B=1 regaddr, C=2 regval, D=3 输出, E=4 期望,
# F=5 输入信号, J=9 RO, K=10 bits, N=13 组号, O=14 起为 T1,T2,...
COL_A, COL_B, COL_C, COL_D, COL_E, COL_F, COL_I, COL_J, COL_K, COL_N = 0, 1, 2, 3, 4, 5, 8, 9, 10, 13
T_START = 14
TAG_KW = ("iddq", "tsensor", "dft")


def build_groups(grid, maxr):
    """按 col A 非空(『case』标记)切组 —— 每个输出一组。返回 [(start_idx, end_idx)] (0 基, 含头不含尾)。
    数据从第 3 行(idx 2)起；col A 非空即新组起点。"""
    starts = [r for r in range(2, maxr) if not is_blank(grid[r][COL_A])]
    groups = []
    for i, s in enumerate(starts):
        e = starts[i + 1] if i + 1 < len(starts) else maxr
        groups.append((s, e))
    return groups


def _group_info(grid, s, e, maxc):
    """抽一组的概要：输出(col D)/期望(col E)/组号(col N)/输入列表/标记。"""
    outs, exps, inputs = [], [], []
    nval = ""
    tags = set()
    last_t = T_START - 1
    for r in range(s, e):
        row = grid[r]
        d = cell_str(row[COL_D]) if COL_D < len(row) else ""
        if d:
            outs.append(d)
        ev = cell_str(row[COL_E]) if COL_E < len(row) else ""
        if ev:
            exps.append(ev)
        if not nval:
            nval = cell_str(row[COL_N]) if COL_N < len(row) else ""
        f = cell_str(row[COL_F]) if COL_F < len(row) else ""
        if f:
            tvals = []
            for c in range(T_START, maxc):
                tv = cell_str(row[c]) if c < len(row) else ""
                tvals.append(tv)
                if tv != "":
                    last_t = max(last_t, c)
            inputs.append({
                "row": r + 1,
                "F": f,
                "RO": cell_str(row[COL_J]) if COL_J < len(row) else "",
                "K": cell_str(row[COL_K]) if COL_K < len(row) else "",
                "B": cell_str(row[COL_B]) if COL_B < len(row) else "",
                "C": cell_str(row[COL_C]) if COL_C < len(row) else "",
                "I": cell_str(row[COL_I]) if COL_I < len(row) else "",
                "T": tvals,
            })
        txt = " ".join(cell_str(v) for v in row if not is_blank(v)).lower()
        for k in TAG_KW:
            if k in txt:
                tags.add(k)
    n_t = max(0, last_t - T_START + 1)
    return {"out": outs, "exp": exps, "N": nval, "inputs": inputs, "tags": tags, "n_t": n_t}


def dump_group_directory(out, grid, groups, maxc):
    out.append("")
    out.append("=" * 72)
    out.append("==== SECTION 3: 测试组目录(每输出一组，共 %d 组) ====" % len(groups))
    out.append("  组号N | 行范围 | 输入数 | 标记 | 输出(col D) | 期望(col E)")
    for (s, e) in groups:
        info = _group_info(grid, s, e, maxc)
        tagmark = ("<" + ",".join(sorted(info["tags"])) + ">") if info["tags"] else ""
        out.append("  N=%-3s 行%d..%d (%d行) 输入%d %-22s 输出=%s 期望=%s"
                   % (info["N"], s + 1, e, e - s, len(info["inputs"]), tagmark,
                      ";".join(info["out"])[:60] or "(无)",
                      ";".join(info["exp"])[:40] or "(无)"))


def dump_groups_full(out, grid, groups, maxc, match_subs):
    """dump 匹配 match_subs(对 col D 输出名子串，小写) 的组的完整真值表。"""
    out.append("")
    out.append("=" * 72)
    out.append("==== SECTION 4: 命中组完整真值表 (匹配: %s) ====" % match_subs)
    hit = 0
    for (s, e) in groups:
        info = _group_info(grid, s, e, maxc)
        outname = ";".join(info["out"]).lower()
        if not any(sub in outname for sub in match_subs):
            continue
        hit += 1
        out.append("")
        out.append("-" * 72)
        out.append("【组 N=%s  行 %d..%d】 输出=%s  期望=%s"
                   % (info["N"], s + 1, e, ";".join(info["out"]), ";".join(info["exp"]) or "(无)"))
        n_t = info["n_t"]
        thdr = " ".join("T%d" % (i + 1) for i in range(n_t))
        out.append("  输入信号 [RO|bits|regaddr=值] :  %s" % thdr)
        for inp in info["inputs"]:
            tv = " ".join((inp["T"][i] if i < len(inp["T"]) and inp["T"][i] != "" else ".")
                          for i in range(n_t))
            reg = ("%s=%s" % (inp["B"], inp["C"])) if (inp["B"] or inp["C"]) else ""
            meta = "RO=%s|bit%s%s" % (inp["RO"] or "?", inp["K"] or "?",
                                     ("|" + reg) if reg else "")
            out.append("    %-44s [%s]" % (inp["F"], meta))
            out.append("        T: %s" % tv)
    if hit == 0:
        out.append("  ⚠ 没有 col D 输出名匹配 %s 的组。看 SECTION 3 目录挑准确名字再 --out-sig。" % match_subs)
    else:
        out.append("")
        out.append("(共 dump %d 组)" % hit)


# ───────────────────────────── main ─────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="摸清 for_test 测试用例页的真实排版")
    ap.add_argument("excel")
    ap.add_argument("--sheet", default="for_test")
    ap.add_argument("--at", default="", help="额外锚点行号，逗号分隔(如 4,76,419)")
    ap.add_argument("--span", type=int, default=5, help="每锚点往下连带 dump 行数(默认 5)")
    ap.add_argument("--cols", type=int, default=80, help="最多看前多少列(默认 80)")
    ap.add_argument("--rows", type=int, default=10000, help="最多扫描多少行(默认 10000)")
    ap.add_argument("--groups", action="store_true",
                    help="打印测试组目录(每输出一组：组号/行范围/输入数/iddq·dft·tsensor标记/输出/期望)")
    ap.add_argument("--out-sig", default="",
                    help="逗号分隔的输出名子串(对 col D)；命中组 dump 完整真值表(含 T 向量)。"
                         "如 --out-sig mixer2g_trim,mixer5g_trim")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    if not os.path.isfile(args.excel):
        sys.exit("找不到文件: %s" % args.excel)
    if not args.excel.lower().endswith(".xlsx"):
        sys.exit("只支持 .xlsx")

    # 要拿合并单元格 → 不能 read_only；data_only 取缓存值
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    sheet = next((s for s in wb.sheetnames if s == args.sheet), None)
    if sheet is None:
        sheet = next((s for s in wb.sheetnames if s.lower() == args.sheet.lower()), None)
    if sheet is None:
        wb.close()
        sys.exit("找不到页 '%s'；现有页: %s" % (args.sheet, wb.sheetnames))

    ws = wb[sheet]
    grid, merges, maxr, maxc = load_grid(ws, args.rows, args.cols)
    logic_k = collect_logic_k(wb)

    # 锚点：默认头部 + 几个全清单已知的块起点；并入 --at
    default_anchors = [3, 4, 76, 419]
    extra = [int(x) for x in args.at.split(",") if x.strip().isdigit()]
    anchors = sorted(set(default_anchors + extra))

    out = []
    out.append("Dreg for_test 页排版探查 (inspect_fortest.py)")
    out.append("文件: %s" % os.path.basename(args.excel))
    out.append("目标页: %s (%d 行 x %d 列；实际扫描 %d 行 x %d 列)"
               % (sheet, ws.max_row, ws.max_column, maxr, maxc))
    out.append("logic.K 基名集合规模(用于认信号名列): %d" % len(logic_k))

    match_subs = [s.strip().lower() for s in args.out_sig.split(",") if s.strip()]
    group_mode = args.groups or match_subs

    dump_overview(out, wb, sheet)
    if not group_mode:
        # 默认：摸结构(列画像/合并/锚点行)
        out.append("")
        out.append("=" * 72)
        out.append("==== SECTION 1: [%s] 列结构 ====" % sheet)
        dump_raw_head(out, grid, maxc)
        dump_merges(out, merges, maxc)
        dump_column_profile(out, grid, maxr, maxc, logic_k)
        dump_anchors(out, grid, maxr, maxc, 3, anchors, args.span)
    else:
        groups = build_groups(grid, maxr)
        if args.groups:
            dump_group_directory(out, grid, groups, maxc)
        if match_subs:
            dump_groups_full(out, grid, groups, maxc, match_subs)

    wb.close()
    out_path = args.out if args.out else os.path.splitext(args.excel)[0] + "_fortest.txt"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(out) + "\n")
    safe_print("已导出 for_test 排版探查报告: %s" % out_path)
    safe_print("共 %d 行。请把该 txt 全文贴回来。" % len(out))


if __name__ == "__main__":
    main()
