# -*- coding: utf-8 -*-
"""make_sample_excel.py — 生成一个"看得见、能跑"的示例 Dreg 核心 Excel。

它严格按 excel_model 的真实列约定造三页(logic / regmap / total_memory_map)，
比 tests/fixtures.py 更丰富：覆盖直通、二选一 mux、门控(&~)、多位总线、
同地址多字段合并、RO(force)与 RW(RF_WRITE)混合等典型情况，且每个输入都能解析干净。

用法:  python make_sample_excel.py [输出路径]
然后:  python -m dreg_verify.gui   →  加载该 .xlsx  →  点信号看「测试项」标签页
"""

import sys

import openpyxl
from openpyxl.utils import column_index_from_string


def _set(ws, row, mapping):
    for col, val in mapping.items():
        ws.cell(row=row, column=column_index_from_string(col), value=val)


# ───────────── total_memory_map：寄存器块 + 字段 ─────────────
def _tmm_reg(ws, r, name, addr):
    _set(ws, r, {"A": name, "B": addr, "C": "0", "D": "RW", "E": "register def"})
    return r + 1


def _tmm_field(ws, r, name, bit, addr, dig, typ, desc=""):
    # 字段行：A=名 B=bit号/范围 C=复位 D=DIG TOP PIN(Y/N) E=功能 F=该字段地址(hex) H=类型
    _set(ws, r, {"A": name, "B": bit, "C": "0", "D": dig, "E": desc or "field",
                 "F": addr, "H": typ})
    return r + 1


def _regmap(ws, r, reg, typ, signal, bit, owner):
    # regmap：D=Reg_Name F=Reg Type G=Signal_Name J..Y=bit15..bit0(Y=bit0) AE=owner
    _set(ws, r, {"D": reg, "F": typ, "G": signal, "AE": owner})
    # bit 列：Y=bit0, X=bit1, ... 这里只为非空标记，bit 位置以 tmm 为准
    col = column_index_from_string("Y") - bit
    ws.cell(row=r, column=col, value=1)
    return r + 1


def build(path):
    wb = openpyxl.Workbook()

    # ========================= logic 页 =========================
    ws = wb.active
    ws.title = "logic"
    _set(ws, 1, {"A": "<<输入信号(列字母=表达式变量) A..J>>", "K": "<<输出>>", "L": "<<真值表达式>>"})
    _set(ws, 2, {"A": "in_A", "B": "in_B", "C": "in_C", "D": "in_D", "J": "in_J",
                 "K": "logic输出名", "L": "表达式", "M": "type", "N": "top_output",
                 "O": "Notes", "P": "owner", "R": "no"})

    rows = [
        # 1) 门控二选一 mux：A 选 C/B，再被 J(iddq)门控拉0。控制位 A、J；数据 B、C。
        {"A": "d_bt_lp_linelocal_mode_ctrl_to_logic",
         "B": "d_bt_lp_bt_mode_sel_to_logic",
         "C": "d_bt_lp_bt_mode_sel_local_to_logic",
         "J": "d_bt_lp_iddq_to_logic",
         "K": "d_logic_bt_lp_reserve", "L": "(A?C:B)&(~J)",
         "M": "to_logic", "N": 1, "O": "门控mux示例", "P": "Alice", "R": 1},

        # 2) 多位(3bit)二选一 mux：A 选 3 位总线 C/B。展示多位 force/RF_WRITE。
        {"A": "d_bt_lp_lna_line_sel_to_logic",
         "B": "d_bt_lp_lna_agc_line_to_logic[2:0]",
         "C": "d_bt_lp_lna_agc_local_to_logic[2:0]",
         "K": "d_logic_bt_lp_lna_agc[2:0]", "L": "A?C:B",
         "M": "to_mux", "N": 1, "O": "多位mux示例", "P": "Bob", "R": 2},

        # 3) 直通：输出 = 单个输入。最简单的一条。
        {"A": "d_bt_lp_en_refbuf_cfg_to_logic",
         "K": "d_en_refbuf", "L": "A",
         "M": "ls", "N": 1, "O": "直通示例", "P": "Alice", "R": 3},

        # 4) 与/或 + 取反混合：(A&B)|(~C)。A=RW, B=RO管脚, C=RO。混合驱动方式。
        {"A": "d_bt_lp_clk_byp_to_logic",
         "B": "d_bt_lp_test_en_to_logic",
         "C": "d_bt_lp_iddq_to_logic",
         "K": "d_logic_clk_byp_out", "L": "(A&B)|(~C)",
         "M": "to_logic", "N": 1, "O": "与或取反示例(RW+RO混合)", "P": "Carol", "R": 4},

        # 5) 多输入与门 + 门控：A&B&(~C)。三个控制位全组合。
        {"A": "d_bt_lp_clk_byp_to_logic",
         "B": "d_bt_lp_lna_line_sel_to_logic",
         "C": "d_bt_lp_iddq_to_logic",
         "K": "d_logic_pll_en", "L": "A&B&(~C)",
         "M": "to_logic", "N": 1, "O": "多控制位与门示例", "P": "Alice", "R": 5},

        # 6) 同地址合并示例：A(0x11 bit2) 选 B(0x11 bit1:0) / C(0x13 bit1:0)。
        #    A 与 B 同在寄存器 0x11 → 生成时合并成一条 RF_WRITE(0x11, ...)。
        {"A": "d_bt_lp_clk_byp_to_logic",
         "B": "d_bt_lp_clk_div_sel_to_logic[1:0]",
         "C": "d_bt_lp_clk_div_alt_to_logic[1:0]",
         "K": "d_logic_clk_sel_out[1:0]", "L": "A?B:C",
         "M": "to_mux", "N": 1, "O": "同地址RF_WRITE合并示例", "P": "Carol", "R": 6},
    ]
    for i, rd in enumerate(rows, start=3):
        _set(ws, i, rd)

    # ========================= regmap 页 =========================
    rm = wb.create_sheet("regmap")
    _set(rm, 2, {"D": "Reg_Name", "F": "Reg Type", "G": "Signal_Name",
                 "I": "default", "J": "b15", "Y": "b0", "Z": "suffix", "AE": "owner"})
    r = 3
    r = _regmap(rm, r, "BT_MODE", "RW", "d_bt_lp_bt_mode_sel", 0, "Alice")
    r = _regmap(rm, r, "LINELOCAL", "RW", "d_bt_lp_linelocal_mode_ctrl", 0, "Alice")
    r = _regmap(rm, r, "LINELOCAL", "RW", "d_bt_lp_bt_mode_sel_local", 1, "Alice")
    r = _regmap(rm, r, "IDDQ_CTRL", "RO", "d_bt_lp_iddq", 0, "Alice")
    r = _regmap(rm, r, "CLK_CFG", "RW", "d_bt_lp_clk_byp", 2, "Carol")

    # ========================= total_memory_map 页 =========================
    tmm = wb.create_sheet("total_memory_map")
    r = 1
    r = _tmm_reg(tmm, r, "BT_MODE", "h2C")
    r = _tmm_field(tmm, r, "d_bt_lp_bt_mode_sel", "0", "h2C", "N", "RW", "BT 模式选择数据")
    r = _tmm_reg(tmm, r, "LINELOCAL_CTRL", "h2D")
    r = _tmm_field(tmm, r, "d_bt_lp_linelocal_mode_ctrl", "0", "h2D", "N", "RW", "line/local 选择(控制位)")
    r = _tmm_field(tmm, r, "d_bt_lp_bt_mode_sel_local", "1", "h2D", "N", "RW", "本地 BT 模式数据")
    r = _tmm_reg(tmm, r, "IDDQ_CTRL", "h30")
    r = _tmm_field(tmm, r, "d_bt_lp_iddq", "0", "h30", "Y", "RO", "IDDQ 门控(管脚,只读)")
    r = _tmm_reg(tmm, r, "LNA_SEL", "h31")
    r = _tmm_field(tmm, r, "d_bt_lp_lna_line_sel", "0", "h31", "N", "RW", "LNA line 选择(控制位)")
    r = _tmm_reg(tmm, r, "LNA_AGC", "h32")
    r = _tmm_field(tmm, r, "d_bt_lp_lna_agc_local", "2:0", "h32", "N", "RW", "本地 AGC 3位数据")
    r = _tmm_reg(tmm, r, "LNA_AGC_LINE", "h33")
    r = _tmm_field(tmm, r, "d_bt_lp_lna_agc_line", "2:0", "h33", "Y", "RO", "line AGC 3位(管脚,只读)")
    r = _tmm_reg(tmm, r, "REFBUF", "h10")
    r = _tmm_field(tmm, r, "d_bt_lp_en_refbuf_cfg", "0", "h10", "N", "RW", "refbuf 使能配置")
    r = _tmm_reg(tmm, r, "CLK_CFG", "h11")
    r = _tmm_field(tmm, r, "d_bt_lp_clk_div_sel", "1:0", "h11", "N", "RW", "时钟分频选择 2位")
    r = _tmm_field(tmm, r, "d_bt_lp_clk_byp", "2", "h11", "N", "RW", "时钟旁路(控制位,同寄存器)")
    r = _tmm_reg(tmm, r, "TEST_EN", "h12")
    r = _tmm_field(tmm, r, "d_bt_lp_test_en", "0", "h12", "Y", "RO", "测试使能(管脚,只读)")
    r = _tmm_reg(tmm, r, "CLK_ALT", "h13")
    r = _tmm_field(tmm, r, "d_bt_lp_clk_div_alt", "1:0", "h13", "N", "RW", "备用分频 2位")

    wb.save(path)
    return path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "sample_dreg.xlsx"
    build(out)
    print("已生成示例 Excel:", out)
