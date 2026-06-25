# -*- coding: utf-8 -*-
"""
diag_one_topout.py — 探查【单个 Topout 信号在真表里到底长什么样】，并核对工具的判定（2026-06-24）。

针对【已解析】的信号（diag_topout_rename.py 只列未解析的，扫不到它）。把这个信号：
  · 在各结构页(level_shift/dft/regmap/tmm/logic/mux)出现的【原始行】(带列名)全 dump 出来；
  · 改名链【逐跳】展示(顶层名 ←(哪页)← 源名 …)；
  · 寄存器字段的 bit_msb/bit_lsb/地址/类型 + 算出的字段宽(核对位宽告警对不对)；
  · 工具最终的解析分类 + 真值表向量数 + issues(黄字告警)；
  · 为这一个信号生成的 .sv 断言片段(看探针贴的到底是哪个名)。

用法（在能打开真表的机器上跑，把整段输出贴回给 Claude）：
    .venv\\Scripts\\python.exe tools/diag_one_topout.py 真表.xlsx d_vco_en_faston_ls
    # 可再补额外名字一起溯源(查门控信号定义在哪页)：
    .venv\\Scripts\\python.exe tools/diag_one_topout.py 真表.xlsx d_vco_en_faston_ls fll_active d_bt_lp_pll_dig_dft_iddq_mode
"""
import sys

# Windows 中文控制台默认 GBK，emoji(⛔/⚠ 等)编不进去会崩 → 兜底永不崩(编不了的字符替成 ?)。
try:
    sys.stdout.reconfigure(errors="replace")
except Exception:   # noqa: BLE001
    pass

import os, sys  # noqa: E402,F811 —— path bootstrap：tools/ 下脚本上溯仓库根、找到 dreg_verify 包
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dreg_verify import excel_model as M
from dreg_verify import topout as T
from dreg_verify import resolver as R

SKIP_SHEETS = {"for_test", "for_test_gen_tc", "main", "namingrule"}
OUT_COL = {"logic": 11, "dft": 4, "iddq": 4, "mux": 7, "level_shift": 3}


def _base(v):
    if v is None:
        return ""
    return M._strip_width(str(v))[0].strip()


def _col_letter(i):
    s = ""
    while i > 0:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def main(path, signal, extras=None):
    import openpyxl
    extras = [e.strip() for e in (extras or []) if e.strip()]
    wb = M.load_workbook(path)
    name = signal.strip()
    low = name.lower()
    print("装载: %s" % path)
    print("sheets: %s" % wb.sheet_names)
    print("\n探查信号: %s\n" % name)

    # ── ① 它在 Topout 页(B 列)的原始行 ──
    print("=== ① 它在 Topout 页(要验清单)里的原始行 ===")
    topo = None
    for t in wb.topout:
        if t.name.lower() == low or _base(t.raw).lower() == low:
            topo = t
            break
    if topo is None:
        print("   [!] 在 wb.topout 里没找到这个名字。它可能不是 Topout 要验信号(是源名?)，")
        print("     仍继续探查它在各结构页的出现情况。")
    else:
        print("   row%-4d  name=%s  width=%s  owner=%s  raw=%r"
              % (topo.row, topo.name, topo.width, topo.owner, topo.raw))
        print("   (旁注 reg_*: reg_name=%r addr=%r reset=%r)"
              % (topo.reg_name, topo.reg_addr, topo.reg_reset))

    # ── ② 改名表里跟它相关的条目 ──
    dft_map = T._dft_rename_map(wb)
    ls_map = T._ls_rename_map(wb)
    full = T._rename_map(wb)
    print("\n=== ② 改名表(工具自动认出的)里跟它相关的条目 ===")
    print("   dft 观测改名表里 %r → %r" % (low, dft_map.get(low)))
    print("   level_shift 改名表里 %r → %r" % (low, ls_map.get(low)))
    print("   (合并表 %r → %r)" % (low, full.get(low)))

    # ── ③ 改名链逐跳 ──
    print("\n=== ③ 改名链逐跳回溯（顶层名 → 源名 → …）===")
    seen, cur, hops, chain = {low}, low, 0, [low]
    while hops < 8:
        nxt = full.get(cur)
        if nxt is None or nxt in seen:
            break
        which = "dft观测改名" if cur in dft_map else ("level_shift电平移位" if cur in ls_map else "?")
        print("   第%d跳: %s ←(%s)← %s" % (hops + 1, cur, which, nxt))
        seen.add(nxt); chain.append(nxt); cur = nxt; hops += 1
    if hops == 0:
        print("   （没有改名：顶层名直接就能/不能解析，不走桥接）")
    else:
        print("   链尾源名 = %s" % cur)

    # ── ④ resolve_root 的最终判定 ──
    logic_idx, mux_idx = T.build_index(wb)
    root = T.resolve_root(wb, name, logic_idx, mux_idx, rename=full)
    print("\n=== ④ 工具最终解析判定 (resolve_root) ===")
    print("   分类 kind     = %s" % root.kind)
    print("   命中名 matched= %s" % root.matched_name)
    print("   源名 source   = %s" % root.source_name)
    print("   探针名 probe  = %s   (None=用源对象自身名)" % root.probe_name)
    print("   reg_kind      = %s" % root.reg_kind)
    print("   dft_obs_name  = %s   (≠None=途经带 iddq 门的 dft 观测层 → analyze 应据它把 iddq 门钉成输入)"
          % getattr(root, "dft_obs_name", None))
    print("   note          = %s" % root.note)

    # ── ⑤ 若是直连寄存器：dump 字段 bit_msb/bit_lsb/地址/类型，核对位宽 ──
    src_for_reg = (root.source_name or name).lower()
    rk, ent, found = T._register_kind(wb, src_for_reg)
    print("\n=== ⑤ 寄存器字段实情 (核对『字段宽』告警对不对) ===")
    if ent is None:
        print("   源名 %r 不在 regmap/tmm —— 不是直连寄存器根" % src_for_reg)
    else:
        bm, bl = getattr(ent, "bit_msb", None), getattr(ent, "bit_lsb", None)
        fw = (int(bm) - int(bl) + 1) if (bm is not None and bl is not None) else 1
        print("   来源页       = %s" % found)
        print("   字段名       = %s" % getattr(ent, "name", getattr(ent, "signal", "?")))
        print("   类型 reg_type= %s  (RO=回读跳过 / RW=可驱可验)" % rk)
        print("   bit_msb      = %s" % bm)
        print("   bit_lsb      = %s" % bl)
        print("   => 字段宽    = %d bit   %s" % (fw, "（多 bit → 会触发『按字段全宽验』黄字告警）"
                                                 if fw > 1 else "（1 bit → 不告警）"))
        print("   地址 address = %s" % getattr(ent, "address", None))
        print("   Topout 名声明的 width = %s" % (topo.width if topo else "?"))

    # ── ⑥ analyze_signal：真值表 + issues(黄字告警来源) ──
    if topo is not None:
        resolver = R.Resolver(wb)
        res = T.analyze_signal(wb, resolver, topo, root=root, mode="min", max_tests=256)
        print("\n=== ⑥ 真值表分析结果 (analyze_signal) ===")
        print("   status   = %s" % res.status)
        print("   out_width= %s   (最终用来生成/断言的位宽)" % res.out_width)
        print("   用例数    = %d" % len(res.vectors))
        print("   issues(黄字告警):")
        if res.issues:
            for s in res.issues:
                print("      [!] %s" % s)
        else:
            print("      （无）")
        # cone 直接绑定的输入基名（注意：pin_dft_gate 钉的 iddq 门在【向量 extra_forces】里、不在 bindings）
        ibases = sorted({(getattr(b, "base", "") or "").lower()
                         for b in (res.bindings or {}).values() if b is not None})
        print("   cone bindings 输入基名(%d): %s" % (len(ibases), ", ".join(ibases) or "(无)"))
        if res.meta.get("iddq_skipped"):
            print("   [!] meta.iddq_skipped = %s" % res.meta["iddq_skipped"])

        # ── ⑥b dft 门(iddq)钉入实情：精确定位「为什么 iddq 门漏了」 ──
        print("\n=== ⑥b dft 门(iddq)钉入实情（排查漏门：dft_obs_name / wb.dft / 门 resolve / 向量 force）===")
        obs = getattr(root, "dft_obs_name", None)
        print("   root.dft_obs_name = %r   (None → 不会调 pin_dft_gate → iddq 必漏)" % obs)
        dft = getattr(wb, "dft", None) or {}
        print("   wb.dft 共 %d 个被门控观测输出；以下逐个核对【dft_obs_name + 改名链每跳】是否门控:" % len(dft))
        to_check = []
        if obs:
            to_check.append(obs.lower())
        for h in chain:
            if h not in to_check:
                to_check.append(h)
        for h in to_check:
            g = dft.get(h)
            if g is None:
                print("     · wb.dft[%r] = None（此名在 dft 页不是被 iddq 门控的观测输出）" % h)
                continue
            gb = g.get("gate_base")
            print("     · wb.dft[%r]: gate_base=%r gate_raw=%r transparent=%r"
                  % (h, gb, g.get("gate_raw"), g.get("transparent")))
            info = {"raw": gb, "base": gb, "width": 1, "msb": None, "lsb": None}
            try:
                gbnd = resolver.resolve("dft_gate_" + str(gb), info)
                print("       门 %r resolve → resolved=%s kind=%s (需 RO 才能 pin)  note=%s"
                      % (gb, gbnd.resolved, getattr(gbnd, "kind", None), getattr(gbnd, "note", "")))
            except Exception as ex:   # noqa: BLE001
                print("       门 resolve 异常: %r" % ex)
        ef = set()
        for v in res.vectors:
            for tup in (getattr(v, "extra_forces", None) or []):
                ef.add(tup[0])
        print("   向量里实际钉上的 extra_force 网（iddq 门若钉成功=在这里）: %s"
              % (", ".join(sorted(ef)) or "(无)"))

    # ── ⑦ 这个信号在各结构页的所有原始行 ──
    print("\n=== ⑦ 它(及其源名/额外名)在各结构页的【原始行】===")
    targets = {low} | set(chain) | {e.lower() for e in extras}
    if root.source_name:
        targets.add(root.source_name.lower())
    print("   匹配名集合: %s" % ", ".join(sorted(targets)))
    raw = openpyxl.load_workbook(path, data_only=True, read_only=True)
    headers = {}
    for ws in raw.worksheets:
        if ws.title.lower() in SKIP_SHEETS:
            continue
        hdr = {}
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for c, v in enumerate(row, 1):
                if v is not None and c not in hdr:
                    hdr[c] = str(v).strip()
        headers[ws.title] = hdr
    any_hit = False
    for ws in raw.worksheets:
        title = ws.title
        if title.lower() in SKIP_SHEETS:
            continue
        outc = OUT_COL.get(title.lower())
        for r, row in enumerate(ws.iter_rows(values_only=True), 1):
            hit_cols = [c for c, v in enumerate(row, 1) if _base(v).lower() in targets]
            if not hit_cols:
                continue
            any_hit = True
            star = "★输出列(此页把它当输出/改名定义行)" if outc in hit_cols else ""
            cells = []
            for c, v in enumerate(row, 1):
                if v is None or str(v).strip() == "":
                    continue
                hdr = headers.get(title, {}).get(c, "")
                cells.append("%s%s=%s" % (_col_letter(c), ("(%s)" % hdr) if hdr else "",
                                          str(v).strip()))
                if len(cells) >= 14:
                    cells.append("…"); break
            print("  [%s] row%-4d %s" % (title, r, star))
            print("      " + " | ".join(cells))
    if not any_hit:
        print("  （在所有结构页都没出现）")

    # ── ⑧ 为这一个信号生成的 .sv 断言片段 ──
    print("\n=== ⑧ 为这一个信号生成的 .sv（看探针贴的是哪个名）===")
    try:
        text, _b = T.render_topout_sv(wb, mode="min", max_tests=32, only=[name])
        for ln in text.splitlines():
            print("   " + ln)
    except Exception as ex:   # noqa: BLE001
        print("   生成 .sv 失败: %r" % ex)

    # ── ⑨ for_test 金标准：这个信号那一组的输入/写值原始行（看真实需要哪些输入）──
    print("\n=== ⑨ for_test 金标准：含 %r 的那一组原始行(看真实输入有哪些) ===" % name)
    ftn = next((s for s in raw.sheetnames if s.lower() == "for_test"), None)
    if ftn is None:
        print("   （没有 for_test 页）")
    else:
        ws = raw[ftn]
        rows = []          # (行号, [A..J 文本])
        hit = []
        ft_targets = {low} | {e.lower() for e in extras}
        if root.source_name:
            ft_targets.add(root.source_name.lower())
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_col=15, values_only=True), 1):
            cells = ["" if v is None else str(v).strip() for v in row]
            rows.append((ri, cells))
            joined = " ".join(cells).lower()
            if any(t in joined for t in ft_targets):
                hit.append(ri)
        if not hit:
            print("   [X] for_test 里没有含该信号(或源名/额外名)的行——可能 for_test 此信号为空，"
                  "或名字带本脚本没归一掉的后缀")
        else:
            lo = max(0, min(hit) - 20)
            hi = min(len(rows), max(hit) + 4)
            print("   命中行: %s   (dump 窗口 row%d..%d)"
                  % (", ".join(str(h) for h in hit[:12]) + (" …" if len(hit) > 12 else ""),
                     lo + 1, hi))
            for (ri, cells) in rows[lo:hi]:
                show = [("%s=%s" % (_col_letter(c + 1), v)) for c, v in enumerate(cells) if v]
                if show:
                    mark = " ←命中" if ri in hit else ""
                    print("   row%-5d %s%s" % (ri, " | ".join(show[:12]), mark))

    # ── ⑩ 额外名(门控信号)在结构页哪里被【定义为输出】──
    if extras:
        print("\n=== ⑩ 额外名(门控信号)溯源：它们在哪页被当输出定义 ===")
        for ex in extras:
            exl = ex.lower()
            print("  -- %s --" % ex)
            found_any = False
            for ws in raw.worksheets:
                title = ws.title
                if title.lower() in SKIP_SHEETS:
                    continue
                outc = OUT_COL.get(title.lower())
                for r, row in enumerate(ws.iter_rows(values_only=True), 1):
                    cols = [c for c, v in enumerate(row, 1) if _base(v).lower() == exl]
                    if not cols:
                        continue
                    found_any = True
                    star = "★输出列(此页定义它)" if outc in cols else "(出现在输入列)"
                    cells = []
                    for c, v in enumerate(row, 1):
                        if v is None or str(v).strip() == "":
                            continue
                        hdr = headers.get(title, {}).get(c, "")
                        cells.append("%s%s=%s" % (_col_letter(c), ("(%s)" % hdr) if hdr else "",
                                                  str(v).strip()))
                        if len(cells) >= 12:
                            cells.append("…"); break
                    print("    [%s] row%-4d %s  %s" % (title, r, star, " | ".join(cells)))
            if not found_any:
                print("    （在结构页没出现——可能是顶层 pin / RTL 内部网，不在真表结构里）")

    print("\n=== 怎么核对 ===")
    print(" · ①⑦ 看真表原文：level_shift 页有没有 %s 的输出行(C 列=顶层口真名)。" % low)
    print(" · ③ 看改名链对不对：是不是 1 跳到源寄存器。")
    print(" · ⑤ 看字段宽：bit_msb/bit_lsb 算出的宽是不是真的 2 → 黄字告警是否成立。")
    print(" · ⑧ 看 .sv 的 assert 探的是顶层名(_ls)、RF_WRITE 写的是源寄存器、位宽对不对。")
    print(" · ⑨ for_test 金标准：真实需要哪些输入(若比 cone 多→工具漏了门控层)。")
    print(" · ⑩ 多出来的门控信号定义在哪页——决定能不能 cone 进来(在结构页=能补；只在 RTL=真表缺口)。")
    print(" · 把整段贴回给 Claude 一起核对判断。")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python tools/diag_one_topout.py 真表.xlsx 信号名 [额外名1 额外名2 …]")
        print("例:   python tools/diag_one_topout.py Hi1108_Pilot_BT_LP_DREG_95P_20260623.xlsx d_vco_en_faston_ls")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3:])
