# -*- coding: utf-8 -*-
"""
diag_topout_rename.py — 探查【Topout 未解析信号到底在哪一页被改名】（2026-06-24）。

背景：有些 Topout 顶层名（如 A）在 logic 页叫别的名（如 A_sm），中间某一页观测后改了名。
工具按 logic/mux/regmap/tmm 候选名解析 Topout 名时对不上 → 标『未解析』、建不出真值表。
要补桥接，先得知道【改名发生在哪一页、源名是谁】——本脚本就是来回答这个的，不改任何文件。

用法（在能打开真表的机器上跑，把整段输出贴回给 Claude）：
    python tools/diag_topout_rename.py 真表.xlsx
    # 或限定只看前 N 个未解析名： python tools/diag_topout_rename.py 真表.xlsx 30

输出三段：
  ① Topout 分类汇总（logic/mux/直连寄存器/RO/未解析 各多少）
  ② 未解析名清单
  ③ 逐个未解析名：它在各【结构页】(logic/mux/dft/iddq/level_shift/ISO/regmap/tmm/SignalPath)
     的哪些行出现 + 整行内容 + 是否落在该页的【输出列】(★=改名定义行，输入列就是源名)。
     for_test/for_test_gen_tc(测试数据, 动辄上万行)不扫。
"""
import sys

import os, sys  # noqa: E402,F811 —— path bootstrap：tools/ 下脚本上溯仓库根、找到 dreg_verify 包
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dreg_verify import excel_model as M
from dreg_verify import topout as T

# 各结构页的【输出列】列号(1-based)：名字落这列 = 该页把它当输出定义 → 那一行的输入列就是『源』。
OUT_COL = {"logic": 11, "dft": 4, "iddq": 4, "mux": 7, "level_shift": 3}
# 不扫的页(测试数据, 太大)
SKIP_SHEETS = {"for_test", "for_test_gen_tc", "main", "namingrule"}
STRUCT_HINT = ["logic", "mux", "dft", "iddq", "level_shift", "iso", "regmap",
               "total_memory_map", "signalpath", "topout"]


def _base(v):
    if v is None:
        return ""
    return M._strip_width(str(v))[0].strip()


def _col_letter(i):
    s = ""
    while i > 0:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def main(path, limit=None):
    import openpyxl
    wb = M.load_workbook(path)
    print("装载: %s" % path)
    print("sheets: %s\n" % wb.sheet_names)

    # ── ① 分类汇总 ──
    logic_idx, mux_idx = T.build_index(wb)
    rename = T._rename_map(wb)              # 当前工具能自动认出的改名(dft + level_shift，链式)
    by_kind, unresolved = {}, []
    for topo in wb.topout:
        root = T.resolve_root(wb, topo.name, logic_idx, mux_idx, rename=rename)
        by_kind[root.kind] = by_kind.get(root.kind, 0) + 1
        if root.kind == T.UNRESOLVED:
            unresolved.append(topo.name)
    print("=== ① Topout 分类汇总（%d 个）===" % len(wb.topout))
    for k in ("logic", "mux", "register", "ro-readback", "unresolved"):
        print("   %-14s %d" % (k, by_kind.get(k, 0)))
    print("   （当前工具自动认出 %d 条改名[dft+level_shift,链式]：%s）"
          % (len(rename), ", ".join("%s←%s" % (k, v) for k, v in list(rename.items())[:8])
             + (" …" if len(rename) > 8 else "")))

    if not unresolved:
        print("\n✅ 没有未解析的 Topout 信号——全部能解析，不需要改名桥接。")
        return
    print("\n=== ② 未解析(疑似改名/埋件) %d 个 ===" % len(unresolved))
    print("   " + ", ".join(unresolved))

    # ── ③ 逐个未解析名：全结构页定位 ──
    raw = openpyxl.load_workbook(path, data_only=True, read_only=True)
    # 预读每页表头(第2行)做列名提示
    headers = {}
    for ws in raw.worksheets:
        if ws.title.lower() in SKIP_SHEETS:
            continue
        hdr = {}
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
            for c, v in enumerate(row, 1):
                if v is not None and c not in hdr:
                    hdr[c] = str(v).strip()
        headers[ws.title] = hdr

    todo = unresolved[:limit] if limit else unresolved
    print("\n=== ③ 逐个未解析名定位（★=落在该页输出列=改名定义行，看同行输入列即源名）===")
    for name in todo:
        nlow = name.lower()
        print("\n--- %s ---" % name)
        any_hit = False
        for ws in raw.worksheets:
            title = ws.title
            if title.lower() in SKIP_SHEETS:
                continue
            outc = OUT_COL.get(title.lower())
            for r, row in enumerate(ws.iter_rows(values_only=True), 1):
                hit_cols = [c for c, v in enumerate(row, 1) if _base(v).lower() == nlow]
                if not hit_cols:
                    continue
                any_hit = True
                is_out = (outc in hit_cols)
                star = "★ 输出列(改名定义行)" if is_out else ""
                # 整行非空单元（带列名提示），最多 12 个
                cells = []
                for c, v in enumerate(row, 1):
                    if v is None or str(v).strip() == "":
                        continue
                    hdr = headers.get(title, {}).get(c, "")
                    cells.append("%s=%s" % (_col_letter(c) + (("(%s)" % hdr) if hdr else ""),
                                            str(v).strip()))
                    if len(cells) >= 12:
                        cells.append("…")
                        break
                print("  [%s] row%-4d %s" % (title, r, star))
                print("      " + " | ".join(cells))
        if not any_hit:
            print("  （在所有结构页都没出现——可能是真·埋件/真表⊋DUT，或名字带本脚本没归一掉的后缀）")

    print("\n=== 怎么看 ===")
    print(" · 找每个未解析名带 ★ 的行 = 哪一页把它当输出『改了名』；该行【输入列】的名字(去 _to_xxx)=源名。")
    print(" · 若源名在 logic 页有同名输出行 → 桥接到 logic 源、断言贴顶层名即可（工具可自动补）。")
    print(" · 把本段贴回给 Claude：会据『改名在哪页 + 源名规律』把桥接接到对的页。")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python tools/diag_topout_rename.py 真表.xlsx [只看前N个未解析]")
        sys.exit(1)
    main(sys.argv[1], int(sys.argv[2]) if len(sys.argv) > 2 else None)
