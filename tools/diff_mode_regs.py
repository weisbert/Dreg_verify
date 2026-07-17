# -*- coding: utf-8 -*-
"""
diff_mode_regs.py — 模式寄存器差异审计：对一个目录里的多个寄存器表 excel，按寄存器名
对齐，找出【同一寄存器在不同文件里默认值/地址/类型不一致】的地方，输出审计 xlsx。

⚠ 本脚本【自适应】识别寄存器表结构（列名靠关键词猜），第一职责是把【它认到的结构】
如实打印 + 写进输出的「结构报告」sheet。跑一次先核对结构报告：
  · 认对了 → 「默认值差异」sheet 就是你要的审计结果；
  · 认错了（表头行/列映射不对、寄存器数明显不对）→ 把控制台输出或结构报告贴回来，
    我按你真实的列名改精确版（不用打字描述，数据会告诉我）。

当前【假设】（不符就是要调的地方，已全部显式列出）：
  1. 目录顶层每个 .xlsx/.xlsm = 一个待对比的寄存器表（≥2 个才谈得上 diff）。
  2. 每个文件里寄存器表在某个 sheet；表头行 = 前 8 行中能同时匹配到「寄存器名列」
     +「默认值或地址列」且关键词命中最多的那行。
  3. 列识别关键词（小写子串，按优先级）：
       名称  = signal_name / signal name / reg_name / register / signal / 信号 / 名称
       默认值= default value / default / reset / 初始 / 默认
       地址  = address / addr / offset / 地址
       类型  = reg type / type / 读写 / 类型
  4. 对齐键 = 寄存器名（剥位宽 [msb:lsb] + 小写）。
  5. 差异 = 某寄存器的【默认值】(主) 在各文件间去重后 >1 个取值（空值算一种、参与比较），
     或【某文件缺这个寄存器】。默认值按原文 strip 比较（不做 0x0==0 归一，如实呈现）。
  6. 名字含 "mode" 的寄存器额外高亮（模式寄存器审计重点）。

用法：
    python tools\\diff_mode_regs.py <目录> [--out 模式差异审计.xlsx]
        [--recursive]   连子目录一起扫（默认只顶层）
        [--only-mode]   只审计名字含 mode 的寄存器
        [--sheet 名]    强制指定寄存器表 sheet 名（跳过自动识别）
依赖：openpyxl。
"""
import argparse
import glob
import os
import re
import sys

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少 openpyxl：pip install openpyxl")

KW = {
    "name":    ["signal_name", "signal name", "reg_name", "register", "signal", "信号", "名称"],
    "default": ["default value", "default", "reset", "初始", "默认"],
    "address": ["address", "addr", "offset", "地址"],
    "type":    ["reg type", "type", "读写", "类型"],
}
_WIDTH = re.compile(r"\s*\[[^\]]*\]\s*$")
_BLANK = "(空)"       # 默认值为空时的占位——空 vs 有值也算差异（审计要看得见）


def _s(v):
    return "" if v is None else str(v).strip()


def _norm_key(name):
    return _WIDTH.sub("", _s(name)).lower()


def _match_col(header_cells):
    """一行表头 → {role: col_index}（0基）。每个 role 按【关键词优先级】取列——
    先试高优先关键词(如 signal_name)扫全行，命中即定；没有再降级(reg_name)。
    否则 Reg_Name(块名) 排在 Signal_Name(信号名) 前会被列顺序抢走 → 对齐键错。"""
    lows = [_s(c).lower() for c in header_cells]
    found = {}
    for role, kws in KW.items():
        for kw in kws:
            hit = next((ci for ci, low in enumerate(lows) if low and kw in low), None)
            if hit is not None:
                found[role] = hit
                break
    return found


def _detect_table(ws):
    """在一个 sheet 里找寄存器表：返回 (表头行1基, {role:col}) 或 None。
    表头 = 前 8 行里能同时认到 name + (default 或 address) 且命中列数最多的那行。"""
    best = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        cols = _match_col(row or [])
        if "name" in cols and ("default" in cols or "address" in cols):
            score = len(cols)
            if best is None or score > best[0]:
                best = (score, ri, cols)
    return (best[1], best[2]) if best else None


def _read_regs(path, forced_sheet=None):
    """读一个 excel 的寄存器表 → (sheet, 表头行, {role:col}, {key:{role:值,'_raw_name':原名}})。
    找不到寄存器表 → (None, None, None, {})。"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheets = [forced_sheet] if forced_sheet else wb.sheetnames
        for sh in sheets:
            if sh not in wb.sheetnames:
                continue
            ws = wb[sh]
            det = _detect_table(ws)
            if det is None:
                continue
            hrow, cols = det
            name_ci = cols["name"]
            regs = {}
            for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                nm = _s(row[name_ci]) if name_ci < len(row) else ""
                if not nm:
                    continue
                key = _norm_key(nm)
                if not key:
                    continue
                rec = {"_raw_name": nm}
                for role, ci in cols.items():
                    if role == "name":
                        continue
                    rec[role] = _s(row[ci]) if ci < len(row) else ""
                regs.setdefault(key, rec)      # 同名首个采纳
            return sh, hrow, cols, regs
        return None, None, None, {}
    finally:
        wb.close()


def main():
    ap = argparse.ArgumentParser(description="模式寄存器差异审计（自适应识别，先核对结构报告）")
    ap.add_argument("directory")
    ap.add_argument("--out", default="模式差异审计.xlsx")
    ap.add_argument("--recursive", action="store_true")
    ap.add_argument("--only-mode", action="store_true")
    ap.add_argument("--sheet", default=None)
    args = ap.parse_args()

    if not os.path.isdir(args.directory):
        sys.exit("目录不存在: %s" % args.directory)
    pat = "**/*" if args.recursive else "*"
    files = sorted(f for ext in ("xlsx", "xlsm")
                   for f in glob.glob(os.path.join(args.directory, pat + "." + ext),
                                      recursive=args.recursive)
                   if not os.path.basename(f).startswith("~$"))
    if not files:
        sys.exit("目录里没找到 .xlsx/.xlsm: %s" % args.directory)

    print("=" * 70)
    print("扫描目录: %s（%d 个文件）" % (args.directory, len(files)))
    per_file, struct_rows = [], []
    for f in files:
        base = os.path.basename(f)
        try:
            sh, hrow, cols, regs = _read_regs(f, args.sheet)
        except Exception as ex:  # noqa: BLE001
            print("  ⛔ %s 打开失败: %r" % (base, ex))
            struct_rows.append([base, "(打开失败)", "", "", 0, repr(ex)])
            continue
        if sh is None:
            print("  ⚠ %s 没认出寄存器表（无 sheet 同时含 名称+默认值/地址 列）" % base)
            struct_rows.append([base, "(未识别)", "", "", 0, "无匹配表头"])
            continue
        colmap = ", ".join("%s=%s列" % (r, get_column_letter(c + 1))
                           for r, c in sorted(cols.items(), key=lambda x: x[1]))
        print("  · %-34s sheet=%-16s 表头行%s  %s  → %d 个寄存器"
              % (base, sh, hrow, colmap, len(regs)))
        struct_rows.append([base, sh, hrow, colmap, len(regs), ""])
        per_file.append((base, regs))

    if len(per_file) < 2:
        print("\n⚠ 能识别的寄存器表 < 2 个，无法做跨文件 diff（只出结构报告）。")

    all_keys = {}
    for base, regs in per_file:
        for k, rec in regs.items():
            all_keys.setdefault(k, rec["_raw_name"])
    file_names = [p[0] for p in per_file]

    diff_rows = []
    for k in sorted(all_keys):
        raw = all_keys[k]
        is_mode = "mode" in k
        if args.only_mode and not is_mode:
            continue
        defaults, missing = [], []       # defaults[i]: None=缺该寄存器 / 字符串(可空)=默认值
        for base, regs in per_file:
            rec = regs.get(k)
            if rec is None:
                defaults.append(None)
                missing.append(base)
            else:
                defaults.append(rec.get("default", ""))
        # 差异判定：present 里（含空值占位）去重 >1，或有文件缺失
        present_tokens = {(_BLANK if d == "" else d) for d in defaults if d is not None}
        has_diff = len(present_tokens) > 1 or (missing and len(present_tokens) >= 1)
        if has_diff:
            diff_rows.append((raw, is_mode, defaults, missing, present_tokens))

    print("\n对齐 %d 个寄存器，发现 %d 个有差异%s。"
          % (len(all_keys), len(diff_rows), "（仅 mode）" if args.only_mode else ""))

    out = openpyxl.Workbook()
    RED = PatternFill("solid", fgColor="FFC7CE")
    YEL = PatternFill("solid", fgColor="FFF2CC")
    HDR = Font(bold=True)

    ws1 = out.active
    ws1.title = "结构报告"
    ws1.append(["文件", "sheet", "表头行", "列映射（核对是否认对）", "寄存器数", "备注"])
    for c in ws1[1]:
        c.font = HDR
    for r in struct_rows:
        ws1.append(r)
    ws1.freeze_panes = "A2"

    ws2 = out.create_sheet("默认值差异")
    ws2.append(["寄存器", "是否mode"] + ["%s 默认值" % n for n in file_names] + ["缺失于"])
    for c in ws2[1]:
        c.font = HDR
    for raw, is_mode, defaults, missing, tokens in diff_rows:
        row = [raw, "是" if is_mode else ""] \
            + [("(缺失)" if d is None else (d if d != "" else _BLANK)) for d in defaults] \
            + [", ".join(missing)]
        ws2.append(row)
        r = ws2.max_row
        if is_mode:
            ws2.cell(r, 1).fill = YEL
        for ci in range(len(file_names)):     # 差异行整排值标红，便于横向比对
            ws2.cell(r, 3 + ci).fill = RED
    ws2.freeze_panes = "C2"
    if ws2.max_row == 1:
        ws2.append(["（无差异——所有文件的同名寄存器默认值一致）"])

    out.save(args.out)
    print("\n已写出: %s" % os.path.abspath(args.out))
    print("→ 先看「结构报告」sheet / 上面控制台的列映射：认对了差异表就是结果；"
          "认错了把这段输出贴回来，我按真实列改精确版。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
