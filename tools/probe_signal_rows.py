# -*- coding: utf-8 -*-
"""
probe_signal_rows.py — 把某一族信号（默认子串 en_cnt）在【原始真表每一页】出现的
所有行原样 dump 出来，格式让 Claude 能直接照抄进本地镜像 100% 复现。外加工具当前
对这些信号的解析判定（root 类型 / testcase 数 / 是否带 iddq / issues / .sv 断言前几行）。

真表 .xlsx 传不给 Claude（保密），所以流程是：你在公司机跑这个脚本 → 贴回 .txt →
Claude 拿原始行重建镜像复现。只读，不改真表、不碰任何产品代码。

用法（仓库根目录，装了 openpyxl 的环境 = 仓库 .venv）：
    .venv\\Scripts\\python.exe tools\\probe_signal_rows.py 真表.xlsx en_cnt
        [en_cnt d_en_cnt_to_crg ...]   多个子串/精确名一起抓（默认只 en_cnt）
        [--rows-per-sheet 80]          每页最多 dump 多少匹配行（默认 80）
        [--out 文件]                   默认 <真表名>_rows.txt

依赖：openpyxl。仅支持 .xlsx。跑完把 .txt 全文贴回来。
"""
import argparse
import os
import sys

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:   # noqa: BLE001
        pass

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少 openpyxl，请先 pip install openpyxl（或用仓库 .venv 的 python）")


def _cell(v):
    if v is None:
        return None
    s = str(v).replace("\n", "\\n").replace("\r", "")
    return s if len(s) <= 120 else s[:117] + "..."


def _row_str(row_vals, max_cols=40):
    parts = []
    for ci, v in enumerate(row_vals[:max_cols]):
        cv = _cell(v)
        if cv is not None and cv != "":
            parts.append("%s=%r" % (get_column_letter(ci + 1), str(v)))
    return "   ".join(parts)


def main():
    ap = argparse.ArgumentParser(description="dump 信号族在真表每页的原始行（只读）")
    ap.add_argument("excel")
    ap.add_argument("needles", nargs="*", default=None,
                    help="要抓的子串/信号名（默认 en_cnt）")
    ap.add_argument("--rows-per-sheet", type=int, default=80)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    needles = [n.lower() for n in (args.needles or ["en_cnt"])]
    out_path = args.out or (os.path.splitext(os.path.basename(args.excel))[0] + "_rows.txt")
    L = []

    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    try:
        import subprocess
        head = subprocess.run(["git", "-C", here, "rev-parse", "--short", "HEAD"],
                              capture_output=True, text=True, timeout=10).stdout.strip()
    except Exception:   # noqa: BLE001
        head = "(git 不可用)"
    L.append("工具 HEAD: %s" % head)
    L.append("Excel: %s" % args.excel)
    L.append("抓取子串: %s" % ", ".join(needles))

    # ── ① 每页：表头(前2行) + 所有匹配行，原样 dump ──
    wb = openpyxl.load_workbook(args.excel, data_only=True, read_only=True)
    L.append("\n全部页名: %s" % list(wb.sheetnames))
    for s in wb.sheetnames:
        ws = wb[s]
        header_lines, hit_lines, n_hit = [], [], 0
        for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rs = _row_str(row)
            if ri <= 2 and rs:                        # 表头两行（帮 Claude 对齐列含义）
                header_lines.append("  [表头]行%d:  %s" % (ri, rs))
            hit = any(v is not None and any(nd in str(v).lower() for nd in needles)
                      for v in row)
            if hit:
                n_hit += 1
                if n_hit <= args.rows_per_sheet:
                    hit_lines.append("  行%-5d %s" % (ri, rs))
        if not hit_lines:
            continue                                  # 该页无匹配 → 跳过（不啰嗦）
        L.append("\n══ 页 %r —— %d 个匹配行 ══" % (s, n_hit))
        L.extend(header_lines)
        L.append("  ----")
        L.extend(hit_lines)
        if n_hit > args.rows_per_sheet:
            L.append("  …（还有 %d 行未显示，--rows-per-sheet 调大）" % (n_hit - args.rows_per_sheet))
    wb.close()

    # ── ② 工具判定：load_workbook 后，对匹配到的完整信号名 resolve_root + analyze ──
    L.append("\n\n══ 工具当前判定（load_workbook + resolve_root/analyze）══")
    sys.path.insert(0, here)
    try:
        import types
        from dreg_verify import excel_model as M
        from dreg_verify import topout as T
        from dreg_verify import resolver as R
        m = M.load_workbook(args.excel)

        # dft loud 通道（本轮新增）：有没有认不出的门形态 / 同名重复带门行
        if getattr(m, "dft_unrecognized", None):
            L.append("\n[dft 认不出门形态的行]（可能漏门，本轮起 loud）：")
            for u in m.dft_unrecognized:
                if any(nd in u.get("out", "") for nd in needles):
                    L.append("  第%s行 out=%s expr=%r —— %s"
                             % (u.get("row"), u.get("out"), u.get("expr"), u.get("why")))
        if getattr(m, "dft_dups", None):
            for ob, ds in m.dft_dups.items():
                if any(nd in ob for nd in needles):
                    L.append("\n[dft 同名重复带门行] out=%s：额外 %d 条（已按首行采纳）" % (ob, len(ds)))

        # 收集所有匹配的候选信号名：Topout 页 + dft 观测输出 + logic/mux 输出 + regmap 字段
        names = set()
        for t in (m.topout or []):
            if any(nd in t.name.lower() for nd in needles):
                names.add(t.name)
        for ob in (m.dft or {}):
            if any(nd in ob for nd in needles):
                names.add(ob)
        for sig in m.logic:
            if any(nd in (sig.out_base or "").lower() for nd in needles):
                names.add(sig.out_base)
        for g in (m.mux or []):
            if any(nd in (g.out_base or "").lower() for nd in needles):
                names.add(g.out_base)

        topo_by_name = {t.name.lower(): t for t in (m.topout or [])}
        for nm in sorted(names):
            L.append("\n── 信号 %r ──" % nm)
            root = T.resolve_root(m, nm)
            L.append("  resolve_root: kind=%s  matched=%s  source=%s  dft_obs=%s  dft_bridged=%s"
                     % (root.kind, getattr(root, "matched_name", None),
                        getattr(root, "source_name", None),
                        getattr(root, "dft_obs_name", None),
                        getattr(root, "dft_bridged", None)))
            if getattr(root, "note", ""):
                L.append("  note: %s" % root.note)
            in_topout = nm.lower() in topo_by_name
            topo = topo_by_name.get(nm.lower()) or types.SimpleNamespace(
                name=nm, width=1, owner="")
            try:
                res = T.analyze_signal(m, R.Resolver(m), topo, mode="max", max_tests=64)
                forced = sorted({wl.lower() for v in res.vectors
                                 for (wl, _x, _w) in (getattr(v, "extra_forces", None) or [])})
                L.append("  analyze: status=%s  testcase数=%d  含iddq=%s  在Topout清单=%s"
                         % (res.status, len(res.vectors),
                            any("iddq" in f for f in forced), in_topout))
                if forced:
                    L.append("    force的网: %s" % forced)
                for iss in (res.issues or []):
                    L.append("    ⚠ %s" % iss)
                try:
                    text, _ = T.render_topout_sv(m, mode="max", max_tests=64, only=[nm])
                    asserts = [l.strip() for l in text.splitlines() if "assert (" in l]
                    for a in asserts[:6]:
                        L.append("    .sv: %s" % a)
                    if len(asserts) > 6:
                        L.append("    .sv: …（共 %d 条断言）" % len(asserts))
                except Exception as ex:   # noqa: BLE001
                    L.append("    (render_sv 失败: %r)" % ex)
            except Exception as ex:       # noqa: BLE001
                L.append("  analyze 失败: %r" % ex)
    except Exception as ex:               # noqa: BLE001
        L.append("  ⛔ load_workbook/分析失败: %r" % ex)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print("已写出: %s（%d 行）—— 把这个文件全文贴回来" % (out_path, len(L)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
