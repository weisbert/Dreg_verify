# -*- coding: utf-8 -*-
"""dft 门控解析/归属加固回归（2026-07-17，主线逻辑设计 review 的 🔴 修复）：

  A2  read_dft 按 E 条件字母取门列（此前盲取物理 B 列——E="A?0:B" 门/功能对调=假绿）
  A3  门形态容差（B[0]/括号/取反/定宽零）+ 认不出的 '?' 行 loud 记账（此前静默漏门）
  A4  同名多条带门行 first-wins + dft_dups 告警（此前 last-wins 静默、与 R35 regmap 口径相反）
  A1  G 列（去向）交叉核对：带门行去向与本支不一致 → 不套门 + ⚠ 门归属存疑（此前
      out_base 兜底不看去向——同源扇出时兄弟支的门被串上，3bde25d 只堵了桥接支）
  A6  iddq 已是 mux 显式输入 → 不再 pin/补拍（此前同拍同网 force 两次，pin 赢=选路拍假红）

评审报告：docs/主线逻辑设计review_20260717.md。无这些形态的旧表零影响（mirror e2e 另测）。
"""

import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import make_mirror_btlp                            # noqa: E402
from dreg_verify import excel_model as M           # noqa: E402
from dreg_verify import generator                  # noqa: E402
from dreg_verify import topout as T                # noqa: E402
from dreg_verify import resolver as R              # noqa: E402


# ───────────────────────── 表达式解析（A2/A3 单元） ─────────────────────────

def test_parse_dft_gate_expr_forms():
    P = M._parse_dft_gate_expr
    # 基本两形态 + 条件字母就是谁写在前面（不是列位置）
    assert P("B?0:A") == ("B", 0)
    assert P("B ? 0 : A") == ("B", 0)
    assert P("B?A:0") == ("B", 1)
    assert P("A?0:B") == ("A", 0)          # 门写在 A 列的合法拼法（E1）
    assert P("C?0:A") == ("C", 0)
    # 容差：位切片 / 括号 / 定宽零 / 分支切片
    assert P("B[0]?0:A") == ("B", 0)
    assert P("(B)?0:A") == ("B", 0)
    assert P("B?0:(A)") == ("B", 0)
    assert P("B?0:A[3:0]") == ("B", 0)
    assert P("B?16'h0:A") == ("B", 0)
    assert P("B?A:1'b0") == ("B", 1)
    # 取反 = 两支互换
    assert P("!B?A:0") == ("B", 0)         # ≡ B?0:A
    assert P("~B?0:A") == ("B", 1)         # ≡ B?A:0
    # 不认（不猜）：压 1 门 / 非三目 / 条件是常量 / 两支都非常量
    assert P("B?1:A") is None
    assert P("A&B") is None
    assert P("0?0:A") is None
    assert P("B?C:A") is None
    assert P("A") is None
    assert P("") is None and P(None) is None


# ───────────────────────── 工作簿级回归（mirror + 注入行） ─────────────────────────

_IDDQ = "d_bt_lp_pll_dig_dft_iddq_mode"


def _logic_row(lg, out, expr="A?B:C"):
    r = lg.max_row + 1
    for c, v in [(1, "d_bt_lp_linelocal_mode_ctrl_to_logic"),
                 (2, "d_bt_lp_linectrl_rx_en_to_logic"),
                 (3, "d_bt_lp_rx_en_local_to_logic"),
                 (11, out), (12, expr), (16, "Yao Wang")]:
        lg.cell(r, c, v)


def _dft_row(dft, a, b, d, e, g=""):
    r = dft.max_row + 1
    for c, v in [(1, a), (2, b), (4, d), (5, e), (6, "16"), (8, "Yao Wang")]:
        if v:
            dft.cell(r, c, v)
    if g:
        dft.cell(r, 7, g)
    return r


@pytest.fixture(scope="module")
def hardened_path(tmp_path_factory):
    p = str(tmp_path_factory.mktemp("dg") / "hardened.xlsx")
    make_mirror_btlp.build(p)
    wb = openpyxl.load_workbook(p)
    lg, dft, ls, top = wb["logic"], wb["dft"], wb["level_shift"], wb["Topout"]
    # A2：门写在 A 列（E="A?0:B"）——功能源在 B 列
    _logic_row(lg, "src_sm")
    _dft_row(dft, _IDDQ + "_to_dft", "src_sm_to_dft", "x_top", "A?0:B")
    # A4：同名两条带门行（首行 B?0:A / 次行 B?A:0）
    _logic_row(lg, "s1")
    _dft_row(dft, "s1_to_dft", _IDDQ + "_to_dft", "dup_out", "B?0:A")
    _dft_row(dft, "s1_to_dft", _IDDQ + "_to_dft", "dup_out", "B?A:0")
    # A3a：容差拼法 B[0]?0:A 要认出门
    _logic_row(lg, "v_src")
    _dft_row(dft, "v_src_to_dft", _IDDQ + "_to_dft", "v_top", "B[0]?0:A")
    # A3b：B?1:A（压 1 门，语义未确认）→ 不套门但 loud
    _logic_row(lg, "w_top")
    _dft_row(dft, "w_top_to_dft", _IDDQ + "_to_dft", "w_top", "B?1:A")
    # A1：G 列交叉核对——g_sig 的带门行去向 to_crg（≠本支 ls）；h_sig 的去向 to_ls（=本支）；
    #     k_sig 带门行去向 to_ls 但 Topout 验的是裸名（≠ls 支）
    for base, g in (("g_sig", "to_crg"), ("h_sig", "to_ls"), ("k_sig", "to_ls")):
        _logic_row(lg, base)
        _dft_row(dft, base + "_to_dft", _IDDQ + "_to_dft", base, "B?0:A", g=g)
    for base in ("g_sig", "h_sig"):
        r = ls.max_row + 1
        for c, v in [(1, base + "_to_ls"), (2, "STD_SR_L2H"), (3, base + "_ls"),
                     (5, 1), (6, "Yao Wang")]:
            ls.cell(r, c, v)
    for nm in ("x_top", "dup_out", "v_top", "w_top", "g_sig_ls", "h_sig_ls", "k_sig"):
        r = top.max_row + 1
        top.cell(r, 1, "Yao Wang"); top.cell(r, 2, nm)
    wb.save(p)
    return p


def _analyze(wb, name):
    topo = next(t for t in wb.topout if t.name == name)
    return T.analyze_signal(wb, R.Resolver(wb), topo, mode="max", max_tests=64)


def _iddq_evidence(res):
    """(dft_gate 非空?, extra_forces 里有 iddq?, DFT 拍数)"""
    forced = {wl.lower() for v in res.vectors
              for (wl, _x, _w) in (getattr(v, "extra_forces", None) or [])}
    return (res.dft_gate is not None, any("iddq" in f for f in forced),
            sum(1 for v in res.vectors if getattr(v, "dft_pitch", False)))


def test_a2_gate_letter_maps_to_column(hardened_path):
    """⭐A2：E="A?0:B" → 门取 A 列(iddq)、功能源取 B 列(src_sm)——此前门/功能对调=真源不验。"""
    wb = M.load_workbook(hardened_path)
    g = wb.dft["x_top"]
    assert g["gate_letter"] == "A" and g["gate_base"] == _IDDQ
    assert T._dft_rename_map(wb).get("x_top") == "src_sm"    # 改名边=功能源，不再指向门寄存器
    root = T.resolve_root(wb, "x_top")
    assert root.kind == T.LOGIC and root.source_name == "src_sm"
    assert root.dft_obs_name == "x_top"
    res = _analyze(wb, "x_top")
    assert res.status == "ok" and _iddq_evidence(res) == (True, True, 1)


def test_a4_duplicate_gated_rows_first_wins_and_warn(hardened_path):
    """⭐A4：同名两条带门行 → 首行采纳(transparent=0)+dft_dups 记帐+分析时 ⚠（此前 last-wins 静默）。"""
    wb = M.load_workbook(hardened_path)
    assert wb.dft["dup_out"]["transparent"] == 0             # 首行 B?0:A（last-wins 时=1）
    assert len(wb.dft_dups.get("dup_out", [])) == 1
    res = _analyze(wb, "dup_out")
    assert res.status == "ok" and _iddq_evidence(res)[0]
    assert any("已按首行采纳" in s for s in res.issues), res.issues


def test_a3_variant_spelling_recognized(hardened_path):
    """A3：B[0]?0:A 拼法认出门（此前静默漏门）。"""
    wb = M.load_workbook(hardened_path)
    assert "v_top" in wb.dft
    res = _analyze(wb, "v_top")
    assert res.status == "ok" and _iddq_evidence(res) == (True, True, 1)


def test_a3_unrecognized_gate_is_loud(hardened_path):
    """⭐A3：B?1:A（压 1 门，不猜语义）→ 不套门，但 dft_unrecognized 记账 + 分析 ⚠（此前静默）。"""
    wb = M.load_workbook(hardened_path)
    assert "w_top" not in wb.dft
    assert any(u["out"] == "w_top" for u in wb.dft_unrecognized)
    res = _analyze(wb, "w_top")
    assert res.status == "ok" and _iddq_evidence(res) == (False, False, 0)
    assert any("认不出门形态" in s for s in res.issues), res.issues


def test_a1_gate_dest_mismatch_suppresses_fallback(hardened_path):
    """⭐A1：带门行去向 G=to_crg、Topout 验 _ls 支 → 不套门 + ⚠ 门归属存疑
    （E2 的镜像：此前 out_base 兜底不看 G 列，兄弟支的门被串上）。"""
    wb = M.load_workbook(hardened_path)
    res = _analyze(wb, "g_sig_ls")
    assert res.status == "ok" and _iddq_evidence(res) == (False, False, 0)
    assert any("门归属存疑" in s for s in res.issues), res.issues


def test_a1_gate_dest_match_keeps_gate(hardened_path):
    """对照：带门行去向 G=to_ls、本支就是 _ls 支 → 门照套、无存疑 ⚠（d_en_vco_fc 类不回归）。"""
    wb = M.load_workbook(hardened_path)
    res = _analyze(wb, "h_sig_ls")
    assert res.status == "ok" and _iddq_evidence(res) == (True, True, 1)
    assert not any("门归属存疑" in s for s in res.issues), res.issues


def test_a1_bare_hit_with_ls_dest_suppressed(hardened_path):
    """A1：带门行去向 G=to_ls 但 Topout 验的是裸名（非 ls 支）→ 带门网=<名>_to_ls≠裸名，不套门+⚠。"""
    wb = M.load_workbook(hardened_path)
    res = _analyze(wb, "k_sig")
    assert res.status == "ok" and _iddq_evidence(res) == (False, False, 0)
    assert any("门归属存疑" in s for s in res.issues), res.issues


def test_display_gate_same_policy_as_analyze(hardened_path):
    """A13 收口：形态分类(_result_gate_info)与 analyze 同一 _gate_obs_for——g_sig_ls 不标 GATED、
    h_sig_ls 标 GATED（此前两处各写一份，改一处忘一处=口径漂移温床）。"""
    wb = M.load_workbook(hardened_path)
    assert T._result_gate_info(wb, _analyze(wb, "g_sig_ls")) is None
    assert T._result_gate_info(wb, _analyze(wb, "h_sig_ls")) is not None


# ───────────────────────── A6：iddq 已是 mux 显式输入 → 不再 pin/补拍 ─────────────────────────

def test_a6_mux_gate_as_explicit_input_no_double_force(tmp_path):
    """⭐A6：iddq 同时是 mux 的显式控制输入且 mux 输出被 dft 门控 → 同拍不再对同一网 force 两次
    （此前 pin 排在控制 force 之后=pin 赢，"iddq=1 选 case1"实际走 case0=断言假红）。"""
    from test_mux_wl import _build_mux_wb, _mrow
    wb = _build_mux_wb(
        str(tmp_path / "a6.xlsx"),
        tmm_fields=[
            ("IDDQ", "h11", "d_iddq_mode", "0", "Y", "RO"),    # 门=RO 线，同时当 mux 控制
            ("S0", "h20", "d_s0[1:0]", "1:0", "N", "RW"),
            ("S1", "h21", "d_s1[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux[1:0]", "d_iddq_mode_to_mux", "1'b0", "d_g[1:0]", 1),
            _mrow("d_s1_to_mux[1:0]", "d_iddq_mode_to_mux", "1'b1", "d_g[1:0]", 1),
        ],
        dft_rows=[("d_g_to_dft[1:0]", "d_iddq_mode_to_dft", "d_g[1:0]", "B?0:A")],
    )
    txt = generator.render(generator.build(wb, generator.GenOptions(mux_mode="min")))
    # 门已是显式输入：不补 DFT 拍（无 release），每拍恰好一条 iddq force（控制扫 0/1 各一次）
    assert "release `ENV_RF.d_iddq_mode;" not in txt
    forces = [l for l in txt.splitlines() if "force `ENV_RF.d_iddq_mode=" in l]
    assert len(forces) == 2, forces                       # 修前=4(每拍双 force)+DFT 拍
    assert any("=16'h0;" in l for l in forces) and any("=16'h1;" in l for l in forces)
    # report 双轨同口径
    rep = generator.report(wb, generator.GenOptions(mux_mode="min"))
    tbl = next(t for t in rep["tables"] if t["signal"].startswith("d_g"))
    labels = [x["label"] for x in tbl["inputs"]]
    assert labels.count("d_iddq_mode") == 1, labels       # 真值表里 iddq 只一行（修前两行）


# ═════════════════════ 评审剩 5 条 🟡（A7/A8/A9/A11/A12，2026-07-17 续修）═════════════════════
"""
  A7  build_index 候选名撞车 setdefault 静默首个（验错行=假绿）→ _index_collisions 记账 + ⚠
  A8  直接命中静默压掉 dft 改名边（同名双源）→ resolve_root 命中 dft 改名边即 ⚠
  A9  pin_dft_gate 门非可 force → 【已 loud】：同拍 _append_dft_vectors 先跑并置 iddq_skipped
      （証伪：非静默，故不改生产码，仅锁定 loud 行为，防将来有人拆掉 append 兄弟）
  A11 Topout 页同名重复行 → 复用 row_aid = 重复 assert 标号 = 非法 SV → 按名 seen 去重 + dup-name 记账
  A12 make_negative 不拷 extra_forces → 带门 passthrough 根负向拍丢门 pin → 与 clone_vector 对齐拷贝
"""


def _ls_row(ls, in_net, out_name):
    r = ls.max_row + 1
    for c, v in [(1, in_net), (2, "STD_SR_L2H"), (3, out_name), (5, 1), (6, "Yao Wang")]:
        ls.cell(r, c, v)


def _top_row(top, nm):
    r = top.max_row + 1
    top.cell(r, 1, "Yao Wang"); top.cell(r, 2, nm)


_RW_GATE = "d_bt_lp_bt_mode_sel_local"       # 镜像里的 RW 字段（非 RO、且非 rw9 的 logic 输入）——当 A9 门


@pytest.fixture(scope="module")
def yellow_path(tmp_path_factory):
    p = str(tmp_path_factory.mktemp("yl") / "yellow.xlsx")
    make_mirror_btlp.build(p)
    wb = openpyxl.load_workbook(p)
    lg, dft, ls, top = wb["logic"], wb["dft"], wb["level_shift"], wb["Topout"]
    # A7：字面 K=coll_sig_ls 行(A&B) vs coll_sig 经 level_shift 的 _ls_name=coll_sig_ls
    _logic_row(lg, "coll_sig", "A?B:C")
    _ls_row(ls, "coll_sig_to_ls", "coll_sig_ls")
    _logic_row(lg, "coll_sig_ls", "A&B")
    _top_row(top, "coll_sig_ls")
    # A8：logic 行 shadow_top + dft 改名 shadow_src→shadow_top 并存（直接命中 logic，改名边被压）
    _logic_row(lg, "shadow_top", "A?B:C")
    _logic_row(lg, "shadow_src", "A|B")
    _dft_row(dft, "shadow_src_to_dft", "", "shadow_top", "A")
    _top_row(top, "shadow_top")
    _top_row(top, "shadow_src")                     # 对照：普通 logic（非双源）——A8 不应误报
    # A9：logic 输出 rw9 被【RW 门】(非 RO)门控（identity 观测 = 非改名，避开 A8 双源）
    _logic_row(lg, "rw9", "A?B:C")
    _dft_row(dft, "rw9_to_dft", _RW_GATE + "_to_dft", "rw9", "B?0:A")
    _top_row(top, "rw9")
    # A11：Topout 页 clk_force_on(register 根，镜像已有 1 行)再加 2 行同名
    _top_row(top, "clk_force_on")
    _top_row(top, "clk_force_on")
    wb.save(p)
    return p


# ───────────────────────── A7：撞名可见化 ─────────────────────────

def test_a7_index_collision_is_loud(yellow_path):
    """⭐A7：coll_sig_ls 被两个 logic 对象认领(字面行 + coll_sig 的 _ls_name)→ 解析仍按行序取首个
    (不改胜负)，但 res.issues 出 ⚠ 撞名(此前静默、验错行=假绿无告警)。"""
    wb = M.load_workbook(yellow_path)
    coll = wb._index_collisions if hasattr(wb, "_index_collisions") else None
    # build_index 尚未跑时属性可能未建——resolve 一次触发
    root = T.resolve_root(wb, "coll_sig_ls")
    assert "coll_sig_ls" in (getattr(wb, "_index_collisions", None) or {})
    assert root.kind == T.LOGIC                       # 仍解析(取首个)，不因撞名变 UNRESOLVED
    res = _analyze(wb, "coll_sig_ls")
    assert any("候选名" in s and "多个源对象" in s for s in res.issues), res.issues


# ───────────────────────── A8：双源可见化 ─────────────────────────

def test_a8_direct_hit_shadows_dft_edge_is_loud(yellow_path):
    """⭐A8：shadow_top 既是 logic 直接命中、又是 dft 改名(shadow_src→shadow_top)输出 → 直接命中
    被采纳、dft 改名边零消费，但 res.issues 出 ⚠ 双源(此前 logic/mux 侧无 R-vco-faston 的续走/告警)。"""
    wb = M.load_workbook(yellow_path)
    root = T.resolve_root(wb, "shadow_top")
    assert root.kind == T.LOGIC and not root.renamed
    assert "shadow_top" in T._dft_rename_edges(wb)     # 确是双源前提
    res = _analyze(wb, "shadow_top")
    assert any("双源" in s and "dft 改名输出" in s for s in res.issues), res.issues


def test_a8_no_false_warn_on_plain_signal(yellow_path):
    """对照：普通信号(非双源)不误报 A8 ⚠——干净表零影响的护栏。"""
    wb = M.load_workbook(yellow_path)
    res = _analyze(wb, "shadow_src")                   # 只是普通 logic，非 dft 改名输出
    assert not any("双源" in s for s in res.issues), res.issues


# ───────────────────────── A9：非可 force 门已 loud（証伪，锁定） ─────────────────────────

def test_a9_nonro_gate_is_already_loud(yellow_path):
    """A9【証伪】：门=RW 字段(非 RO 不可 force)时 pin_dft_gate 返回 None，但同拍 _append_dft_vectors
    先跑并置 meta['iddq_skipped']（→块顶 ⚠）——非静默。锁定此 loud 行为，防将来拆掉 append 兄弟。"""
    wb = M.load_workbook(yellow_path)
    res = _analyze(wb, "rw9")
    assert res.status == "ok"
    assert res.dft_gate is None                         # RW 门非 RO → 不 pin
    assert res.meta.get("iddq_skipped"), res.meta       # 但 loud：iddq_skipped 已置（块顶 ⚠ 走它）
    assert _RW_GATE in res.meta["iddq_skipped"]
    assert not any("双源" in s for s in res.issues)     # identity 观测=非改名边 → 不误触 A8


# ───────────────────────── A11：重复 Topout 行去重 + 撞键告警 ─────────────────────────

def test_a11_duplicate_topout_rows_dedup_no_illegal_sv(yellow_path):
    """⭐A11：Topout 页 clk_force_on 3 行 → 按 Topout 名 seen 去重，只产出 1 块、其余记 dup-name；
    绝不产出重复 assert 标号(非法 SV，elaboration 才炸)。summary 出 n_dup_topout_names 撞键告警。"""
    import re
    wb = M.load_workbook(yellow_path)
    b = T.build_for_topout(wb, mode="min")
    # 撞键告警可见
    assert "clk_force_on" in b["dup_topout_names"]
    assert b["summary"]["n_dup_topout_names"] >= 1
    # 只产出 1 个 clk_force_on 块；另 2 行记 dup-name（不静默丢、不产出）
    emitted = [(l, s) for l, s in b["blocks"]
               if s.get("n_vectors", 0) > 0 and "clk_force_on" in "\n".join(l).lower()]
    assert len(emitted) == 1, len(emitted)
    dupname = [a for a in b["accounted"]
               if a["name"].lower() == "clk_force_on" and a["status"] == "dup-name"]
    assert len(dupname) == 2, dupname
    # 全局不变量：任一 assert 标号只出现在一个产出块里（否则=非法 SV）
    per_id_blocks = {}
    for l, s in b["blocks"]:
        if s.get("n_vectors", 0) <= 0:
            continue
        for i in set(re.findall(r"assert_(\d+)_", "\n".join(l))):
            per_id_blocks[i] = per_id_blocks.get(i, 0) + 1
    assert all(c == 1 for c in per_id_blocks.values()), \
        [i for i, c in per_id_blocks.items() if c > 1]


# ───────────────────────── A12：负向拍继承门 pin ─────────────────────────

def test_a12_make_negative_inherits_gate_pin():
    """⭐A12：make_negative 拷源向量 extra_forces/release_nets(与 clone_vector 对齐)——带门 passthrough
    根的负向拍此前 extra_forces=[] 丢门 pin(GUI 重排向量即爆)。"""
    from dreg_verify import vectors as V
    tv = V.TestVector(0, {"A": 1}, 1, 1)
    tv.extra_forces = [("`ENV_RF.iddq", 0, 1)]
    tv.release_nets = ["`ENV_RF.iddq"]
    neg = V.make_negative(tv)
    assert neg.is_negative
    assert neg.extra_forces == [("`ENV_RF.iddq", 0, 1)]     # 门 pin 随负向拍带上
    assert neg.release_nets == ["`ENV_RF.iddq"]
    # 且是新列表(不与源共享引用，改一个不动另一个)
    neg.extra_forces.append(("x", 1, 1))
    assert tv.extra_forces == [("`ENV_RF.iddq", 0, 1)]


def test_a12_negative_no_forces_when_source_ungated():
    """对照：源向量无门 pin(常规 logic 负向，pin 在负向之后跑)→ 负向拍 extra_forces 仍空 → 逐字节不变。"""
    from dreg_verify import vectors as V
    tv = V.TestVector(0, {"A": 1}, 1, 1)
    neg = V.make_negative(tv)
    assert neg.extra_forces == [] and neg.release_nets == []
