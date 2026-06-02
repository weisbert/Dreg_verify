# -*- coding: utf-8 -*-
"""对抗式审查确认的 8 个问题的回归测试（每条对应一个 review finding）。"""

import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dreg_verify import expr as E             # noqa: E402
from dreg_verify import vectors as V          # noqa: E402
from dreg_verify import sv_writer as W        # noqa: E402
from dreg_verify import excel_model as M      # noqa: E402
from dreg_verify import generator as G       # noqa: E402
from dreg_verify.excel_model import _normalize_type, read_tmm, DregWorkbook, TmmField, LogicSignal  # noqa: E402
from dreg_verify.resolver import Resolver, InputBinding  # noqa: E402


# ── #1/#8 除法：解析与求值一致，且结果正确 ──
def test_division_consistent():
    assert E.eval_expr("A/B", {"A": 4, "B": 4}, {"A": 8, "B": 2}, 4)[0] == 4
    assert E.eval_expr("A/B", {"A": 4, "B": 4}, {"A": 8, "B": 0}, 4)[0] == 0   # 除0→0


# ── #2 有符号字面量：显式拒绝（不再静默按无符号）──
def test_signed_literal_rejected():
    with pytest.raises(E.ExprError):
        E.parse("4'sb1010")
    with pytest.raises(E.ExprError):
        E.eval_expr("4'sb1010", {}, {}, 4)


# ── #3 rule4：字段值按字段位宽裁剪，溢出不污染相邻字段 ──
def test_rule4_field_width_clamp():
    foo = InputBinding("A", "foo", "foo", width=3, kind="RW", address=0x10,
                       reg_lsb=0, reg_msb=0, wire="foo", found_in="tmm")   # 实际只占 bit0
    bar = InputBinding("B", "bar", "bar", width=1, kind="RW", address=0x10,
                       reg_lsb=1, reg_msb=1, wire="bar", found_in="tmm")
    vec = V.TestVector(0, {"A": 5, "B": 1}, exp_value=0, exp_width=1)       # foo=5 溢出 bit0
    lines, unresolved = W._build_drive_lines(vec, {"A": foo, "B": bar}, ["A", "B"])
    text = "\n".join(lines)
    assert not unresolved
    assert "16'h3" in text             # foo 截为 bit0=1, bar=bit1 → 0x3
    assert "16'h7" not in text         # 不应是 0x7（溢出污染）


# ── #4 截断/去重统计自洽：向量数 + 去重数 + 丢弃数 == 计划组合数 ──
def test_truncate_dedup_accounting():
    class _B:
        def __init__(self, w):
            self.width = w
    bindings = {x: _B(1) for x in "ABCD"}
    node = E.parse("(A?C:B)|D")        # 控制 A,D(各1位→4组合)，数据 B,C
    vecs, meta = V.generate_vectors(node, bindings, out_width=1, mode="max", max_tests=7)
    control, _ = E.classify_vars(node, E.Env({x: 1 for x in "ABCD"}))
    planned = (2 ** len(control)) * 5  # max=5 主题
    assert "deduped" in meta
    assert len(vecs) + meta["deduped"] + meta.get("dropped", 0) == planned
    assert meta["truncated"] is True


# ── #5 resolver 歧义后缀匹配：不静默选第一个，判 UNKNOWN + note ──
def test_resolver_ambiguous_suffix():
    tmm = {
        "d_x_bt_mode_sel": TmmField("d_x_bt_mode_sel", 0, 0, 0x2C, "RW", "N", "R1"),
        "d_y_lna_line_sel": TmmField("d_y_lna_line_sel", 0, 0, 0x31, "RW", "N", "R2"),
    }
    wb = DregWorkbook(logic=[], regmap={}, tmm=tmm, sheet_names=[])
    b = Resolver(wb).resolve("A", {"base": "sel", "width": 1, "raw": "sel"})
    assert b.kind == "UNKNOWN"
    assert not b.resolved
    assert "模糊匹配" in b.note


def test_resolver_single_suffix_still_works():
    tmm = {"d_x_bt_mode_sel": TmmField("d_x_bt_mode_sel", 0, 0, 0x2C, "RW", "N", "R1")}
    wb = DregWorkbook(logic=[], regmap={}, tmm=tmm, sheet_names=[])
    b = Resolver(wb).resolve("A", {"base": "bt_mode_sel", "width": 1, "raw": "x"})
    assert b.kind == "RW" and b.address == 0x2C and b.resolved   # 唯一后缀匹配仍可用


def test_resolver_ambiguous_honors_override():
    tmm = {
        "d_x_bt_mode_sel": TmmField("d_x_bt_mode_sel", 0, 0, 0x2C, "RW", "N", "R1"),
        "d_y_lna_line_sel": TmmField("d_y_lna_line_sel", 0, 0, 0x31, "RW", "N", "R2"),
    }
    wb = DregWorkbook(logic=[], regmap={}, tmm=tmm, sheet_names=[])
    b = Resolver(wb, force_overrides=["sel"]).resolve("A", {"base": "sel", "width": 1, "raw": "x"})
    assert b.kind == "RO"   # 用户显式覆盖时即便歧义也尊重


# ── #6 _normalize_type：各种 RO/RW 写法 ──
@pytest.mark.parametrize("text,expect", [
    ("RW", "RW"), ("RO", "RO"), ("R", "RO"), ("W", "RW"),
    ("R/W", "RW"), ("READ/WRITE", "RW"), ("READ WRITE", "RW"), ("Read Write", "RW"),
    ("WO", "RW"), ("W/O", "RW"), ("READ ONLY", "RO"), ("Read Only", "RO"),
    ("READ", "RO"), ("", None),
])
def test_normalize_type(text, expect):
    assert _normalize_type(text) == expect


# ── #7 dig_top_pin：'NA'/'N/A' 不被误判为 'N' ──
def test_dig_top_pin_exact():
    wb = openpyxl.Workbook()
    ws = wb.active
    from openpyxl.utils import column_index_from_string as ci

    def setrow(r, m):
        for col, v in m.items():
            ws.cell(row=r, column=ci(col), value=v)

    setrow(1, {"A": "REG1", "B": "h40", "D": "RW"})            # 寄存器行
    setrow(2, {"A": "f_na", "B": "0", "F": "h40", "D": "NA", "H": "RO"})
    setrow(3, {"A": "f_y", "B": "1", "F": "h40", "D": "Y", "H": "RO"})
    setrow(4, {"A": "f_n", "B": "2", "F": "h40", "D": "N", "H": "RW"})
    fields = read_tmm(ws)
    assert fields["f_na"].dig_top_pin is None      # 'NA' → None（不再误判 'N'）
    assert fields["f_y"].dig_top_pin == "Y"
    assert fields["f_n"].dig_top_pin == "N"


# ── 追加: owner 名字含空格的筛选（大小写无关 + 折叠多余空格）──
def test_owner_filter_with_spaces():
    sig = LogicSignal(row=3, out_name="d_x", out_width=1, expr="A", suffix="ls",
                      top_output=1, notes="", owner="Wei Yu", assert_id="1",
                      inputs={"A": {"raw": "a_to_logic", "base": "a", "width": 1,
                                    "msb": None, "lsb": None}})
    wb = DregWorkbook(logic=[sig], regmap={}, tmm={}, sheet_names=[])
    for q in ["Wei Yu", "wei yu", "Wei  Yu", "  WEI YU  "]:
        assert len(G.select_signals(wb, G.GenOptions(owners=[q]))) == 1, q
    assert len(G.select_signals(wb, G.GenOptions(owners=["Alice"]))) == 0


# ── 追加: 级联中间信号 + wire 兜底（对应真表诊断里的 pll_n*/*_mux_out 等）──
def _logic(out_name, out_width, expr, inputs, aid, owner="C", suffix="ls", top_output=1):
    return LogicSignal(row=1, out_name=out_name, out_width=out_width, expr=expr,
                       suffix=suffix, top_output=top_output, notes="", owner=owner,
                       assert_id=aid, inputs=inputs)


def test_chained_and_wire_fallback():
    # plln1 是一个 top_output 的 32bit logic 输出；plln2 把 plln1 当输入(级联，可见 wire)
    s1 = _logic("d_plln1[31:0]", 32, "A",
                {"A": {"raw": "int_n_to_logic[31:0]", "base": "int_n", "width": 32,
                       "msb": 31, "lsb": 0}}, "60", top_output=1)
    s2 = _logic("d_plln2[31:0]", 32, "A",
                {"A": {"raw": "d_plln1", "base": "d_plln1", "width": 1,
                       "msb": None, "lsb": None}}, "61", suffix="to_mux", top_output=1)
    wb = DregWorkbook(logic=[s1, s2], regmap={}, tmm={}, sheet_names=[])
    res = Resolver(wb)
    # s1.A=int_n 表里查无 → wire 兜底 force
    b1 = res.resolve_signal_inputs(s1)["A"]
    assert b1.kind == "RO" and b1.found_in == "wire" and b1.resolved
    # s2.A=d_plln1 是 top_output 输出 → 级联可 force，宽度取真实 32
    b2 = res.resolve_signal_inputs(s2)["A"]
    assert b2.kind == "RO" and b2.found_in == "logic" and b2.width == 32


def test_chained_ls_output_forces_rtl_name():
    # ls 输出的 RTL 网名 = K 列名 + _ls（lpbt_dig_top.v 实证）。
    # 本例输入原文"直接写上游输出名"(d_pllen, 无 _to_logic 后缀) → 读的就是上游输出网
    # → force RTL 网名(_ls)。若原文带 _to_logic 则是 regfile 前级信号，force 基名
    # （见 test_cascade_via_to_logic_forces_base_not_ls / test_self_ref_*）。
    s1 = _logic("d_pllen[1:0]", 2, "A",
                {"A": {"raw": "cfg_to_logic", "base": "cfg", "width": 2, "msb": 1, "lsb": 0}},
                "70", suffix="ls", top_output=1)
    s2 = _logic("d_chained", 1, "A",
                {"A": {"raw": "d_pllen", "base": "d_pllen", "width": 2, "msb": 1, "lsb": 0}},
                "71", suffix="to_mux", top_output=1)
    wb = DregWorkbook(logic=[s1, s2], regmap={}, tmm={}, sheet_names=[])
    # rtl_name: 后缀插在基名之后、位宽切片之前
    assert s1.rtl_name == "d_pllen_ls[1:0]"
    assert s2.rtl_name == "d_chained"                  # 非 ls 原样
    b = Resolver(wb).resolve_signal_inputs(s2)["A"]
    assert b.kind == "RO" and b.found_in == "logic"
    assert b.wire == "d_pllen_ls"                      # force 用 RTL 网名
    assert b.wire_lhs == "d_pllen_ls[1:0]"


def test_chained_to_internal_flagged():
    # 输入是内部信号(top_output=0) → 探不到，应标 UNKNOWN/logic-internal 而非 force
    s1 = _logic("pll_n1[31:0]", 32, "A",
                {"A": {"raw": "int_n", "base": "int_n", "width": 32, "msb": 31, "lsb": 0}},
                "1", top_output=0)                       # 内部信号
    s2 = _logic("pll_n2[31:0]", 32, "A",
                {"A": {"raw": "pll_n1", "base": "pll_n1", "width": 1, "msb": None, "lsb": None}},
                "2", top_output=1)
    wb = DregWorkbook(logic=[s1, s2], regmap={}, tmm={}, sheet_names=[])
    b = Resolver(wb).resolve_signal_inputs(s2)["A"]
    assert b.kind == "UNKNOWN" and b.found_in == "logic-internal" and not b.resolved
    assert "内部信号" in b.note


# ── 追加(2026-06-02 真实 bug d2a_cnt_sclk): RO 自引用输入不能 force 本行输出网 ──
# RTL 实证(BT_LP_DREG.v / BT_LP_DREG_sig_logic.v):
#   input  d2a_cnt_sclk;                  ← 顶层输入端口(RO, to_logicro_reg 可读回)
#   wire   d2a_cnt_sclk_to_logic;         ← regfile 导出的前级信号
#   output d2a_cnt_sclk_ls;               ← sig_logic 输出
#   assign d2a_cnt_sclk_ls = d_bt_lp_pll_dig_dft_iddq_mode_to_logic ? 1'b0 : d2a_cnt_sclk_to_logic;
# 旧 bug: 输入 A=d2a_cnt_sclk_to_logic 去后缀得基名 d2a_cnt_sclk → 命中"它自己"的 logic 输出行
# → 被当级联 force 输出网 d2a_cnt_sclk_ls → 把被验证的输出钉死，断言全部失真。
def _d2a_logic_row():
    return _logic("d2a_cnt_sclk", 1, "B?1'b0:A",
                  {"A": {"raw": "d2a_cnt_sclk_to_logic", "base": "d2a_cnt_sclk",
                         "width": 1, "msb": None, "lsb": None},
                   "B": {"raw": "d_bt_lp_pll_dig_dft_iddq_mode_to_logic",
                         "base": "d_bt_lp_pll_dig_dft_iddq_mode",
                         "width": 1, "msb": None, "lsb": None}},
                  "100", owner="Yao Wang", suffix="ls", top_output=1)


def _d2a_tmm():
    # 镜像真表: d2a_cnt_sclk 在 to_logicro_reg_59(h3B) bit7 (RO 读回)；iddq_mode 在 readro_reg_41(h29) bit2
    return {"d2a_cnt_sclk": TmmField("d2a_cnt_sclk", 7, 7, 0x3B, "RO", None, "to_logicro_reg_59"),
            "d_bt_lp_pll_dig_dft_iddq_mode":
                TmmField("d_bt_lp_pll_dig_dft_iddq_mode", 2, 2, 0x29, "RO", "Y", "readro_reg_41")}


def test_self_ref_ro_input_forces_base_not_own_output():
    sig = _d2a_logic_row()
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=_d2a_tmm(), sheet_names=[])
    bindings = Resolver(wb).resolve_signal_inputs(sig)
    a, b = bindings["A"], bindings["B"]
    # 输出探针仍是 RTL 网名(带 _ls)——这部分本来就对
    assert sig.rtl_name == "d2a_cnt_sclk_ls"
    # ⭐ 输入 A 必须 force 前级原始信号(顶层端口名)，绝不能是本行输出网 _ls
    assert a.kind == "RO" and a.resolved
    assert a.wire == "d2a_cnt_sclk"
    assert a.wire_lhs == "d2a_cnt_sclk"
    # B(iddq, RO 管脚) 照常 force 基名
    assert b.kind == "RO" and b.wire == "d_bt_lp_pll_dig_dft_iddq_mode"


def test_self_ref_ro_input_e2e_sv():
    """全链路: build → render，.sv 里 force 前级信号、assert 输出网，且绝无 force 输出网。"""
    sig = _d2a_logic_row()
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=_d2a_tmm(), sheet_names=[])
    res = G.build(wb, G.GenOptions())
    assert res["summary"]["n_generated"] == 1 and res["summary"]["n_skipped"] == 0
    text = G.render(res)
    # 驱动: force 前级原始信号
    assert "force `ENV_RF.d2a_cnt_sclk=" in text
    assert "force `ENV_RF.d_bt_lp_pll_dig_dft_iddq_mode=" in text
    # 断言: 探输出网(_ls)
    assert "assert (`ENV_RF.d2a_cnt_sclk_ls==" in text
    # ⭐ 绝不能 force 输出网(把被验证输出钉死)
    assert "force `ENV_RF.d2a_cnt_sclk_ls" not in text


def test_self_ref_input_not_in_tmm_still_generates():
    """自引用输入即使 tmm/regmap 查无(found_in=self-input)，也不算 wire 兜底风险——
    RTL 结构(端口→regfile→_to_logic→logic)由 logic 行本身证明存在，默认照常生成。"""
    sig = _d2a_logic_row()
    tmm = {"d_bt_lp_pll_dig_dft_iddq_mode":
           TmmField("d_bt_lp_pll_dig_dft_iddq_mode", 2, 2, 0x29, "RO", "Y", "readro_reg_41")}
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=tmm, sheet_names=[])
    b = Resolver(wb).resolve_signal_inputs(sig)["A"]
    assert b.kind == "RO" and b.found_in == "self-input" and b.resolved
    assert b.wire == "d2a_cnt_sclk"
    assert "自引用" in b.note
    # 默认(非 include-risky)也能生成，不被当 wire 兜底跳过
    res = G.build(wb, G.GenOptions())
    assert res["summary"]["n_generated"] == 1 and res["summary"]["n_skipped"] == 0


def test_cascade_via_to_logic_forces_base_not_ls():
    """级联输入原文带 _to_logic(如 d_en_refbuf_to_logic)：读的是 regfile 导出的前级信号，
    force 基名 d_en_refbuf；不能 force 上游输出网 d_en_refbuf_ls(两根不同的 wire)。
    与 test_chained_ls_output_forces_rtl_name 对照——那个原文直接写输出名(无 _to_logic)，
    才 force RTL 网名(_ls)。"""
    s1 = _logic("d_en_refbuf", 1, "A",
                {"A": {"raw": "d_en_refbuf_cfg_to_logic", "base": "d_en_refbuf_cfg",
                       "width": 1, "msb": None, "lsb": None}},
                "50", suffix="ls", top_output=1)
    s2 = _logic("d_downstream", 1, "A",
                {"A": {"raw": "d_en_refbuf_to_logic", "base": "d_en_refbuf",
                       "width": 1, "msb": None, "lsb": None}},
                "51", suffix="to_mux", top_output=1)
    wb = DregWorkbook(logic=[s1, s2], regmap={}, tmm={}, sheet_names=[])
    b = Resolver(wb).resolve_signal_inputs(s2)["A"]
    assert b.kind == "RO" and b.found_in == "logic"
    assert b.wire == "d_en_refbuf"                      # 前级信号基名
    assert b.wire != "d_en_refbuf_ls"                   # 不是上游输出网
    assert "前级" in b.note


def test_multibit_force_wire_has_slice():
    # 多位 force wire 的 LHS 要带位宽切片 [msb:lsb]；标量不带（对齐真实 VBA）
    bus = InputBinding("A", "x", "d_lna", width=3, kind="RO", address=None, reg_lsb=None,
                       reg_msb=None, wire="d_lna", found_in="wire", slice_msb=2, slice_lsb=0)
    scalar = InputBinding("B", "y", "d_en", width=1, kind="RO", address=None, reg_lsb=None,
                          reg_msb=None, wire="d_en", found_in="wire")
    assert bus.wire_lhs == "d_lna[2:0]"
    assert scalar.wire_lhs == "d_en"
    vec = V.TestVector(0, {"A": 5, "B": 1}, exp_value=0, exp_width=1)
    lines, _ = W._build_drive_lines(vec, {"A": bus, "B": scalar}, ["A", "B"])
    text = "\n".join(lines)
    assert "force `ENV_RF.d_lna[2:0]=16'h5;" in text          # 多位带切片
    assert "force `ENV_RF.d_en=16'h1;" in text                # 标量不带
    assert '"d_lna[2:0]"' in text                             # 消息里 wire 名也带切片


def test_force_literal_width_adaptive():
    # 32bit wire 的 force 字面量应是 32'h，不被截成 16'h
    b = InputBinding("A", "d_plln1", "d_plln1", width=32, kind="RO", address=None,
                     reg_lsb=None, reg_msb=None, wire="d_plln1", found_in="logic")
    vec = V.TestVector(0, {"A": 0x12345678}, exp_value=0, exp_width=32)
    lines, _ = W._build_drive_lines(vec, {"A": b}, ["A"])
    assert "32'h12345678" in "\n".join(lines)


def test_exclude_signals_and_regex():
    s1 = _logic("pll_n1[31:0]", 32, "A", {"A": {"raw": "x", "base": "x", "width": 1,
                                                  "msb": None, "lsb": None}}, "1")
    s2 = _logic("d_logic_bt_lp_reserve", 1, "A", {"A": {"raw": "y", "base": "y", "width": 1,
                                                         "msb": None, "lsb": None}}, "2")
    wb = DregWorkbook(logic=[s1, s2], regmap={}, tmm={}, sheet_names=[])
    # 按名排除
    sel = G.select_signals(wb, G.GenOptions(exclude=["pll_n1"]))
    assert [s.out_name for s in sel] == ["d_logic_bt_lp_reserve"]
    # 按正则排除 datapath
    sel2 = G.select_signals(wb, G.GenOptions(exclude_regex="pll_n|_to_dsm"))
    assert [s.out_name for s in sel2] == ["d_logic_bt_lp_reserve"]


def test_skip_risky_signal_by_default():
    # 输入 close_ready_flag 不在 tmm/regmap/logic → wire兜底(非可驱动 net)→ 默认跳过(对齐 VBA)
    s = _logic("d_logic_x", 1, "A",
               {"A": {"raw": "close_ready_flag", "base": "close_ready_flag", "width": 1,
                      "msb": None, "lsb": None}}, "130")
    wb = DregWorkbook(logic=[s], regmap={}, tmm={}, sheet_names=[])
    res = G.build(wb, G.GenOptions(top_output_only=False))
    assert res["summary"]["n_generated"] == 0
    assert res["summary"]["n_skipped"] == 1
    assert res["skipped"][0][0] == "d_logic_x"
    # --include-risky 强制生成(force by name，可能 elaboration 失败，但用户显式要求)
    res2 = G.build(wb, G.GenOptions(top_output_only=False, include_risky=True))
    assert res2["summary"]["n_generated"] == 1


def test_clean_signal_not_skipped():
    from dreg_verify.excel_model import TmmField
    tmm = {"d_en": TmmField("d_en", 0, 0, 0x10, "RW", "N", "R")}
    s = _logic("d_logic_y", 1, "A",
               {"A": {"raw": "d_en", "base": "d_en", "width": 1, "msb": None, "lsb": None}}, "5")
    wb = DregWorkbook(logic=[s], regmap={}, tmm=tmm, sheet_names=[])
    res = G.build(wb, G.GenOptions(top_output_only=False))
    assert res["summary"]["n_generated"] == 1 and res["summary"]["n_skipped"] == 0


def test_no_wire_fallback_keeps_unknown():
    s1 = _logic("d_x", 1, "A",
                {"A": {"raw": "foo", "base": "foo", "width": 1, "msb": None, "lsb": None}}, "1")
    wb = DregWorkbook(logic=[s1], regmap={}, tmm={}, sheet_names=[])
    b = Resolver(wb, wire_fallback=False).resolve_signal_inputs(s1)["A"]
    assert b.kind == "UNKNOWN" and not b.resolved


# ── 追加: tmm/regmap 字段名带位宽([8:0])时去标注后匹配（真表里 int_n[8:0] 这类）──
def test_tmm_regmap_strip_width_in_field_names():
    from openpyxl.utils import column_index_from_string as ci
    from dreg_verify.excel_model import read_tmm, read_regmap

    twb = openpyxl.Workbook()
    tws = twb.active

    def sr(r, m):
        for c, v in m.items():
            tws.cell(row=r, column=ci(c), value=v)
    sr(1, {"A": "INT_N", "B": "h2", "D": "RW"})                       # 寄存器定义行
    sr(2, {"A": "int_n[8:0]", "B": "15:7", "C": "d45", "D": "N", "F": "h2", "H": "RW"})
    fields = read_tmm(tws)
    assert "int_n" in fields and "int_n[8:0]" not in fields
    assert fields["int_n"].address == 0x2
    assert fields["int_n"].bit_msb == 15 and fields["int_n"].bit_lsb == 7

    rwb = openpyxl.Workbook()
    rws = rwb.active

    def sr2(r, m):
        for c, v in m.items():
            rws.cell(row=r, column=ci(c), value=v)
    sr2(2, {"G": "Signal_Name", "F": "Reg Type", "H": "Address"})     # 表头 row2
    sr2(3, {"G": "int_n[8:0]", "F": "RW", "H": "d2"})                 # d2 = 十进制 2
    reg = read_regmap(rws)
    assert "int_n" in reg and reg["int_n"].address == 2 and reg["int_n"].reg_type == "RW"


def test_real_pll_field_resolves_to_rfwrite():
    # 复现真表：logic 输入基名 int_n，tmm 字段名 int_n[8:0]（h2, 15:7）→ 应解析为 RW/RF_WRITE
    from openpyxl.utils import column_index_from_string as ci
    from dreg_verify.excel_model import read_tmm
    twb = openpyxl.Workbook()
    tws = twb.active

    def sr(r, m):
        for c, v in m.items():
            tws.cell(row=r, column=ci(c), value=v)
    sr(1, {"A": "PFD", "B": "hD", "D": "RW"})
    sr(2, {"A": "d_pfd_en_lnmode[1:0]", "B": "15:14", "D": "N", "F": "hD", "H": "RW"})
    wb = DregWorkbook(logic=[], regmap={}, tmm=read_tmm(tws), sheet_names=[])
    b = Resolver(wb).resolve("A", {"base": "d_pfd_en_lnmode", "width": 2, "raw": "x"})
    assert b.kind == "RW" and b.address == 0xD and b.reg_lsb == 14 and b.resolved


# ── 追加: 同名输入(同一物理信号占多个变量)共享取值，RF_WRITE 不重复 ──
def test_same_base_inputs_share_value():
    from dreg_verify.excel_model import TmmField
    tmm = {"d_pfd": TmmField("d_pfd", 15, 14, 0xD, "RW", "N", "PFD")}
    s = _logic("d_pfd", 2, "A?B:B",
               {"A": {"raw": "d_pfd", "base": "d_pfd", "width": 2, "msb": None, "lsb": None},
                "B": {"raw": "d_pfd", "base": "d_pfd", "width": 2, "msb": None, "lsb": None}},
               "1", suffix="to_mux")
    wb = DregWorkbook(logic=[s], regmap={}, tmm=tmm, sheet_names=[])
    res = Resolver(wb)
    bindings = res.resolve_signal_inputs(s)
    vecs, _ = V.generate_vectors(E.parse(s.expr), bindings, 2, mode="max")
    for v in vecs:                                  # 同一物理信号 → A、B 必同值
        assert v.assignments["A"] == v.assignments["B"]
    lines, _ = W._build_drive_lines(vecs[0], bindings, ["A", "B"])
    assert "\n".join(lines).count("`RF_WRITE(10'hD,") == 1   # 同字段只写一次，不重复


# ── #8 重复次数取自变量值时不被声明位宽截断 ──
def test_repeat_count_not_truncated():
    # A 声明 1 位但取值 2 → {2{B}} = 2'b11 = 3
    assert E.eval_expr("{A{B}}", {"A": 1, "B": 1}, {"A": 2, "B": 1}, 4)[0] == 0b11
    # 字面量次数不受影响
    assert E.eval_expr("{3{C}}", {"C": 1}, {"C": 1}, 3)[0] == 0b111


# ── 追加(2026-06-02 对抗式审查): 自引用的边界路径 —— 内部自引用 / 切片位宽 / 前缀 / 报告 / 负向 ──
def test_internal_self_ref_flagged_logic_internal():
    """内部信号(top_output=0)自引用：没有顶层端口可 force → 必须标 logic-internal，
    不能按『自引用前级』force 基名(那个网在顶层不存在 → CUVUNF)。
    build 时 cone 会报循环引用错误(组合环，fail-loud)，绝不产出错误 .sv。"""
    sig = _logic("d_int", 1, "A&B",
                 {"A": {"raw": "d_int_to_logic", "base": "d_int", "width": 1,
                        "msb": None, "lsb": None},
                  "B": {"raw": "cfg_to_logic", "base": "cfg", "width": 1,
                        "msb": None, "lsb": None}},
                 "300", suffix="to_logic", top_output=0)
    tmm = {"cfg": TmmField("cfg", 0, 0, 0x10, "RW", "N", "R")}
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=tmm, sheet_names=[])
    b = Resolver(wb).resolve_signal_inputs(sig)["A"]
    assert b.kind == "UNKNOWN" and b.found_in == "logic-internal" and not b.resolved
    assert "组合环" in b.note
    # build: cone 循环引用 → errors，不产出(也不静默 force 不存在的网)
    res = G.build(wb, G.GenOptions(top_output_only=False))
    assert res["summary"]["n_generated"] == 0
    assert len(res["errors"]) == 1 and "循环引用" in res["errors"][0][2]


def test_self_ref_slice_width_not_overwidened():
    """自引用输入带显式切片(如 [1:0])：求值位宽=切片宽度，不能被输出全宽覆盖
    (否则枚举值越界 → force 进切片被截断 → 断言期望永不成立)。"""
    sig = _logic("d_xbus[3:0]", 4, "B?4'b0:A",
                 {"A": {"raw": "d_xbus_to_logic[1:0]", "base": "d_xbus", "width": 2,
                        "msb": 1, "lsb": 0},
                  "B": {"raw": "d_iddq_to_logic", "base": "d_iddq", "width": 1,
                        "msb": None, "lsb": None}},
                 "301", suffix="ls", top_output=1)
    tmm = {"d_xbus": TmmField("d_xbus", 3, 0, 0x3B, "RO", None, "R"),
           "d_iddq": TmmField("d_iddq", 2, 2, 0x29, "RO", "Y", "R")}
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=tmm, sheet_names=[])
    a = Resolver(wb).resolve_signal_inputs(sig)["A"]
    assert a.width == 2                      # 切片自身宽度，不是输出全宽 4
    assert a.wire_lhs == "d_xbus[1:0]"       # force LHS 带切片
    # 无显式切片的自引用：位宽推断仍取输出全宽(原行为保留)
    sig2 = _logic("d_ybus[3:0]", 4, "A",
                  {"A": {"raw": "d_ybus_to_logic", "base": "d_ybus", "width": 1,
                         "msb": None, "lsb": None}},
                  "302", suffix="ls", top_output=1)
    wb2 = DregWorkbook(logic=[sig2], regmap={}, tmm={}, sheet_names=[])
    a2 = Resolver(wb2).resolve_signal_inputs(sig2)["A"]
    assert a2.width == 4


def test_self_ref_multibit_force_lhs_has_slice():
    """多 bit 自引用：force LHS 带 [msb:lsb] 切片；绝不出现标量形式或输出网。"""
    sig = _logic("d2a_bus[2:0]", 3, "B?3'b0:A",
                 {"A": {"raw": "d2a_bus_to_logic[2:0]", "base": "d2a_bus", "width": 3,
                        "msb": 2, "lsb": 0},
                  "B": {"raw": "d_iddq_to_logic", "base": "d_iddq", "width": 1,
                        "msb": None, "lsb": None}},
                 "303", suffix="ls", top_output=1)
    tmm = {"d2a_bus": TmmField("d2a_bus", 2, 0, 0x3B, "RO", None, "R"),
           "d_iddq": TmmField("d_iddq", 2, 2, 0x29, "RO", "Y", "R")}
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=tmm, sheet_names=[])
    a = Resolver(wb).resolve_signal_inputs(sig)["A"]
    assert a.kind == "RO" and a.wire == "d2a_bus" and a.width == 3
    assert a.wire_lhs == "d2a_bus[2:0]"
    text = G.render(G.build(wb, G.GenOptions()))
    assert "force `ENV_RF.d2a_bus[2:0]=" in text
    assert "force `ENV_RF.d2a_bus=" not in text       # 不能丢切片变标量
    assert "force `ENV_RF.d2a_bus_ls" not in text     # 不能是输出网
    assert "assert (`ENV_RF.d2a_bus_ls[2:0]==" in text


def test_self_ref_input_honors_wire_prefix():
    """自引用输入 + 探针前缀映射：force 路径带前缀、found_in 保持 tmm(不被改写)。"""
    sig = _d2a_logic_row()
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=_d2a_tmm(), sheet_names=[])
    a = Resolver(wb, wire_prefixes={"d2a_cnt_sclk": "U_BT_LP_DREG"}).resolve_signal_inputs(sig)["A"]
    assert a.kind == "RO"
    assert a.wire == "U_BT_LP_DREG.d2a_cnt_sclk"
    assert a.found_in == "tmm"
    assert "自引用" in a.note and "探针前缀" in a.note


def test_self_ref_prefix_e2e_sv():
    """e2e: 同一信号名同时是输入 wire 与输出探针，前缀对两者都生效，且绝不 force 任何 _ls 网。"""
    sig = _d2a_logic_row()
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=_d2a_tmm(), sheet_names=[])
    text = G.render(G.build(wb, G.GenOptions(
        probe_prefixes={"d2a_cnt_sclk": "U_BT_LP_DREG"})))
    assert "force `ENV_RF.U_BT_LP_DREG.d2a_cnt_sclk=" in text
    assert "assert (`ENV_RF.U_BT_LP_DREG.d2a_cnt_sclk_ls==" in text
    assert "force `ENV_RF.U_BT_LP_DREG.d2a_cnt_sclk_ls" not in text
    assert "force `ENV_RF.d2a_cnt_sclk_ls" not in text


def test_analyze_signal_self_input_clean():
    """analyze_signal: 自引用信号状态=clean，输入 net 显示 force 基名(无 _ls)。"""
    sig = _d2a_logic_row()
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=_d2a_tmm(), sheet_names=[])
    a = G.analyze_signal(Resolver(wb), sig, wb=wb)
    assert a["status"] == "clean"
    inp_a = next(i for i in a["inputs"] if i["letter"] == "A")
    assert inp_a["found_in"] == "tmm"            # 真表里 d2a 在 tmm(to_logicro RO 读回)
    assert inp_a["net"] == "force `ENV_RF.d2a_cnt_sclk"
    assert "_ls" not in inp_a["net"]
    assert inp_a["resolved"]


def test_diagnose_self_input_categories():
    """diagnose 分类: 自引用字段在 tmm → force_ro；不在 tmm(self-input) → force_chained。
    两条 found_in 路径都不能落进 force_wire/unknown。"""
    sig = _d2a_logic_row()
    wb1 = DregWorkbook(logic=[sig], regmap={}, tmm=_d2a_tmm(), sheet_names=[])
    d1 = G.diagnose(wb1, G.GenOptions(top_output_only=False))
    assert d1["cats"]["force_ro"] == 2 and d1["cats"]["force_chained"] == 0
    assert d1["cats"]["force_wire"] == 0 and d1["cats"]["unknown"] == 0
    # 不在 tmm → self-input → 计入级联/自引用
    tmm2 = {"d_bt_lp_pll_dig_dft_iddq_mode":
            TmmField("d_bt_lp_pll_dig_dft_iddq_mode", 2, 2, 0x29, "RO", "Y", "R")}
    wb2 = DregWorkbook(logic=[_d2a_logic_row()], regmap={}, tmm=tmm2, sheet_names=[])
    d2 = G.diagnose(wb2, G.GenOptions(top_output_only=False))
    assert d2["cats"]["force_chained"] == 1 and d2["cats"]["force_ro"] == 1
    assert d2["cats"]["force_wire"] == 0 and d2["cats"]["unknown"] == 0


def test_negative_self_ref_e2e():
    """负向用例 + 自引用：断言侧翻转(_NEG)，驱动侧不变(force 基名，绝不 force 输出网)。"""
    sig = _d2a_logic_row()
    wb = DregWorkbook(logic=[sig], regmap={}, tmm=_d2a_tmm(), sheet_names=[])
    res = G.build(wb, G.GenOptions(neg_all=True))
    assert res["summary"]["n_negative"] >= 1
    text = G.render(res)
    assert "_NEG" in text
    assert "force `ENV_RF.d2a_cnt_sclk=" in text
    assert "force `ENV_RF.d2a_cnt_sclk_ls" not in text
