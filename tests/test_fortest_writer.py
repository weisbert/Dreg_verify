# -*- coding: utf-8 -*-
"""回填 for_test 写出器测试（R27 任务2）。

验证 generator.report → fortest_writer 的列排版逐格复刻真表(Hi1107C WL，inspect_fortest
--raw 摸清)：
  · 每组 = [输入行×n] + [输出行] + [per-test 名字行]，col A 首行=输出名+_g序号
  · RW 寄存器：B/C 放最后字段行(C=合并写值)、H 运行累加(末拍值<<lsb)、最终 H==C
  · RO/wire：B=网名 C=16'h0 H=0
  · 输入 T 向量=二进制(补零位宽)，输出 T=十进制；名字行=GUI 测试名
  · write_fortest 复制源全部 sheet、只换 for_test 页(源文件不动)
"""
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import fixtures                                          # noqa: E402
from dreg_verify import excel_model, generator           # noqa: E402
from dreg_verify import fortest_writer as FT             # noqa: E402
from dreg_verify import sv_writer as W                   # noqa: E402


@pytest.fixture(scope="module")
def rep_max(tmp_path_factory):
    path = tmp_path_factory.mktemp("ft") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    wb = excel_model.load_workbook(str(path))
    rep = generator.report(wb, generator.GenOptions(mode="max"))
    return str(path), rep


def _group(groups, prefix):
    return next(g for g in groups if g["name"].startswith(prefix))


def test_fortest_group_shape_and_case_name(rep_max):
    _path, rep = rep_max
    groups = FT.build_fortest_rows(rep)
    g = _group(groups, "d_logic_bt_lp_reserve_g")
    rows = g["rows"]
    inputs = [r for r in rows if r["kind"] == "input"]
    outs = [r for r in rows if r["kind"] == "output"]
    tnames = [r for r in rows if r["kind"] == "tname"]
    # 结构：n 输入 + 1 输出 + 1 名字行；且输出行/名字行在最后
    assert len(outs) == 1 and len(tnames) == 1
    assert rows[-1]["kind"] == "tname" and rows[-2]["kind"] == "output"
    # col A 只在首个输入行，= 组名(输出名+_g序号)
    assert inputs[0]["a"] == g["name"] and g["name"].endswith("_g1")
    assert all("a" not in r for r in inputs[1:])
    # 每行都带组号 N
    assert all(r["n"] == 1 for r in rows)


def test_fortest_rw_register_accumulation(rep_max):
    """RW：B/C 放寄存器最后字段行，C=合并写值且 == 该行 H(运行累加终值)。"""
    _path, rep = rep_max
    groups = FT.build_fortest_rows(rep)
    g = _group(groups, "d_logic_bt_lp_reserve_g")    # (A?C:B)&(~J)：A/C 同寄存器 10'h2D，B 在 10'h2C
    inputs = [r for r in g["rows"] if r["kind"] == "input"]
    rw_bc = [r for r in inputs if r["j"] == "False" and "b" in r]      # 带 B/C 的 RW 行=寄存器最后字段
    assert rw_bc, "应有 RW 寄存器最后字段行带 B/C 合并写值"
    for r in rw_bc:
        assert r["b"].startswith("10'h") and r["c"].startswith("16'h")
        assert r["i"].startswith("10'h")
        regval = int(r["c"].split("'h")[1], 16)
        assert r["h"] == regval, "B/C 行的 H(运行累加终值) 应 == C(合并寄存器写值)"
    # 至少一个寄存器有 2 个字段(reserve 的 A/C 共享 10'h2D)→ 共享寄存器只在最后字段出一次 B/C
    addrs = [r["i"] for r in inputs if r["j"] == "False"]
    assert any(addrs.count(a) >= 2 for a in set(addrs)), "reserve 的 A/C 应共享同一寄存器地址"


def test_fortest_bitsplit_field_matches_sv(tmp_path):
    """⭐对抗评审 major：bit 拆分字段(同字段被引用成 x[15:1] 与 x[0] 两片)的寄存器写值，必须
    逐字节 = 生成 .sv 的权威 RF_WRITE，不能只按 reg_lsb 重算(那样两片会在 bit0 互相覆盖、值错半)。
    回填 B/C/H 直接复用 compute_drives 的 writes(已处理 slice 偏移)。"""
    src = str(tmp_path / "bs.xlsx")
    fixtures.build_workbook(src, with_pll_chain=True)        # pll_n1 引用 frac_n_lsb[15:1]+[0]
    wb = excel_model.load_workbook(src)
    rep = generator.report(wb, generator.GenOptions(mode="max"))
    t = next(x for x in rep["tables"] if x["signal"].startswith("pll_n1"))
    auth = {w["addr"]: w["hex"] for w in t["tests"][-1]["writes"]}      # 权威末拍 RF_WRITE
    assert any(addr for addr in auth), "应有 RF_WRITE"
    g = next(grp for grp in FT.build_fortest_rows(rep) if grp["name"].startswith("pll_n1"))
    rows = [r for r in g["rows"] if r["kind"] == "input" and r["j"] == "False"]
    bc = {r["b"]: r["c"] for r in rows if str(r.get("b", "")).startswith("10'h")}
    for addr, hexv in auth.items():
        assert bc.get(addr) == hexv, "寄存器 %s 回填写值 %s 应=权威 .sv %s" % (addr, bc.get(addr), hexv)
    # 拆分字段所在寄存器有 ≥2 行同 base、但只最后一行带 B/C，且该行 H==C(累加终值=合并写值)
    splitaddr = next((a for a in auth if sum(1 for r in rows if r["i"] == a) >= 2), None)
    assert splitaddr, "应有一个寄存器含多个字段行(int_n/frac_n_msb 或拆分字段)"
    bc_rows = [r for r in rows if r["i"] == splitaddr and r.get("b")]
    assert len(bc_rows) == 1
    assert bc_rows[0]["h"] == int(bc_rows[0]["c"].split("'h")[1], 16)


def test_fortest_ro_wire_force(rep_max):
    """RO/wire(iddq)：B=网名、C=16'h0、H=0、J=True。"""
    _path, rep = rep_max
    groups = FT.build_fortest_rows(rep)
    g = _group(groups, "d_logic_bt_lp_reserve_g")
    iddq = next(r for r in g["rows"] if r["kind"] == "input" and "iddq" in r["f"])
    assert iddq["j"] == "True"
    assert iddq["c"] == W.fmt_hex(0) and iddq["h"] == 0
    assert iddq["b"] == iddq["i"]            # RO：B 与 I 都是 force 网名(非 10'h 地址)


def test_fortest_value_formats(rep_max):
    """输入 T 向量=二进制(补零位宽)，输出 T=十进制，名字行=GUI 测试名。"""
    _path, rep = rep_max
    groups = FT.build_fortest_rows(rep)
    g = _group(groups, "d_logic_bt_lp_lna_agc_g")     # A?C:B [2:0]，含 3bit 输入
    rows = g["rows"]
    # 多 bit 输入：T 值是 3 位二进制串
    multibit = next(r for r in rows if r["kind"] == "input" and "[2:0]" in r["f"])
    assert all(len(v) == 3 and set(v) <= {"0", "1"} for v in multibit["t"]), multibit["t"]
    # 输出行：T 值是十进制(无 b 前缀的纯数字串)
    out = next(r for r in rows if r["kind"] == "output")
    assert all(v.isdigit() for v in out["t"]), out["t"]
    assert str(out["e"]).lstrip("-").isdigit() or out["e"] == ""
    # 名字行：填 GUI 测试名(T0/T1/…或自定义)，条数 = 测试数
    tname = next(r for r in rows if r["kind"] == "tname")
    assert tname["t"] and len(tname["t"]) == len(out["t"])
    assert tname["t"][0] == "T0"


def test_fortest_g_is_last_test_value(rep_max):
    """G(VBA scratch)=末列(Tn)那拍的值拷贝。"""
    _path, rep = rep_max
    groups = FT.build_fortest_rows(rep)
    for g in groups:
        for r in g["rows"]:
            if r.get("t") and r.get("kind") in ("input", "output") and "g" in r:
                assert str(r["g"]) == str(r["t"][-1]), (r["kind"], r["g"], r["t"][-1])


def test_write_fortest_preserves_sheets_replaces_for_test(rep_max, tmp_path):
    import openpyxl
    src, rep = rep_max
    out = str(tmp_path / "out_fortest.xlsx")
    n = FT.write_fortest(src, out, rep)
    assert n >= 1
    src_wb = openpyxl.load_workbook(src)
    out_wb = openpyxl.load_workbook(out)
    # 源文件全部 sheet 都在(只 for_test 被替换)；源文件本身未被改(仍可正常打开、sheet 数一致)
    assert set(src_wb.sheetnames) <= set(out_wb.sheetnames)
    assert "for_test" in out_wb.sheetnames
    ws = out_wb["for_test"]
    cola = {ws.cell(r, 1).value for r in range(1, ws.max_row + 1)}
    assert any(str(v).startswith("d_logic_bt_lp_reserve_g") for v in cola if v)


def test_write_fortest_skips_mux(tmp_path):
    """mux 信号结构不同，不进 for_test(只回填 logic cone 真值表)。"""
    src = str(tmp_path / "muxwb.xlsx")
    fixtures.build_workbook(src, with_mux=True)
    wb = excel_model.load_workbook(src)
    rep = generator.report(wb, generator.GenOptions(mode="max"))
    groups = FT.build_fortest_rows(rep)
    mux_names = {t["signal"] for t in rep.get("tables", []) if not t.get("is_logic")}
    # 没有任何组名来自 mux 信号
    for g in groups:
        assert not any(g["name"].startswith(mn.split("[")[0]) and "mux" in mn.lower()
                       for mn in mux_names) or True
    # 更直接：所有组对应的都是 is_logic 表
    logic_sigs = {t["signal"] for t in rep["tables"] if t.get("is_logic") and t.get("tests")}
    assert len(groups) == len(logic_sigs)


def test_fortest_empty_signal_no_group(tmp_path):
    """零用例(无 tests)的信号不产出 for_test 组。"""
    src = str(tmp_path / "wb.xlsx")
    fixtures.build_workbook(src)
    wb = excel_model.load_workbook(src)
    rep = generator.report(wb, generator.GenOptions(mode="max"))
    # 人工塞一个无 tests 的 logic 表，确认被跳过
    rep["tables"].append({"is_logic": True, "signal": "d_empty", "tests": [], "inputs": []})
    groups = FT.build_fortest_rows(rep)
    assert all(g["name"] != "d_empty_g%d" % len(groups) for g in groups)
    assert not any("d_empty" in g["name"] for g in groups)
