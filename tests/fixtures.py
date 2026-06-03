# -*- coding: utf-8 -*-
"""构造一个镜像真实 Dreg Excel 结构的合成工作簿，用于端到端测试（拿不到真表时的 oracle）。

含已对照过 .sv 的 reserve(R=106) 与 lna_agc(R=108)，以及一个 ls 直通信号。
列布局严格按 excel_model 的约定：logic A..J=输入(列字母=变量字母), K=输出, L=表达式,
M=type, N=top_output, O=notes, P=owner, R=序号。表头在第 2 行。
"""

import openpyxl


def _set_row(ws, rownum, mapping):
    """mapping: {列字母: 值}。"""
    from openpyxl.utils import column_index_from_string
    for col, val in mapping.items():
        ws.cell(row=rownum, column=column_index_from_string(col), value=val)


def build_workbook(path, with_pll_chain=False, with_mux=False):
    """with_pll_chain=True 额外加入 pll_n←pll_n2←pll_n1 三层 cone 链
    （镜像 2026-06-02 真表 Hi1108 的 pll_n 结构，用于 cone 展开测试）。
    with_mux=True 额外加入 mux 页（镜像 2026-06-03 真表 mux 页排版；默认 False
    保证现有测试的 sheet/信号计数完全不变）。"""
    wb = openpyxl.Workbook()

    # ───────── logic 页 ─────────
    ws = wb.active
    ws.title = "logic"
    # row1 = 合并标题(随意)，row2 = 真表头
    _set_row(ws, 1, {"A": "inputs", "K": "logic"})
    _set_row(ws, 2, {
        "A": "in_A", "B": "in_B", "C": "in_C", "J": "in_J",
        "K": "logic_out_signal_name", "L": "logic expression", "M": "suffix",
        "N": "top_output", "O": "Notes", "P": "owner", "R": "no",
    })
    # reserve: (A?C:B)&(~J)  A=linelocal_mode_ctrl B=bt_mode_sel C=bt_mode_sel_local J=iddq
    _set_row(ws, 3, {
        "A": "d_bt_lp_linelocal_mode_ctrl_to_logic",
        "B": "d_bt_lp_bt_mode_sel_to_logic",
        "C": "d_bt_lp_bt_mode_sel_local_to_logic",
        "J": "d_bt_lp_iddq_to_logic",
        "K": "d_logic_bt_lp_reserve",
        "L": "(A?C:B)&(~J)",
        "M": "to_logic", "N": 0, "O": "to_pll_ctrl", "P": "Alice", "R": 106,
    })
    # lna_agc: A?C:B  [2:0]  A=lna_line_sel B=lna_agc_line(RO,3) C=lna_agc_local(RW,3)
    _set_row(ws, 4, {
        "A": "d_bt_lp_lna_line_sel_to_logic",
        "B": "d_bt_lp_lna_agc_line_to_logic[2:0]",
        "C": "d_bt_lp_lna_agc_local_to_logic[2:0]",
        "K": "d_logic_bt_lp_lna_agc[2:0]",
        "L": "A?C:B",
        "M": "to_mux", "N": 1, "O": "to_datapath", "P": "Bob", "R": 108,
    })
    # ls 直通: A  输出名无 d_logic_ 前缀 (en_refbuf)
    _set_row(ws, 5, {
        "A": "d_bt_lp_en_refbuf_cfg_to_logic",
        "K": "d_en_refbuf",
        "L": "A",
        "M": "ls", "N": 1, "O": "to_top", "P": "Alice", "R": 50,
    })
    # pll_n 三层 cone 链（镜像真表 logic row3/4/5：内部信号 pll_n1/pll_n2 逐层代入）
    if with_pll_chain:
        _set_row(ws, 6, {
            "A": "int_n_to_logic[8:0]",
            "B": "frac_n_msb_to_logic[6:0]",
            "C": "frac_n_lsb_to_logic[15:1]",
            "D": "frac_n_lsb_to_logic[0]",
            "E": "d_xo_freq_sel_to_logic",
            "K": "pll_n1[31:0]",
            "L": "E?{1'b0,A,B,C}:{A,B,C,D}",
            "M": "to_logic", "N": 0, "O": "internal", "P": "Yao Wang", "R": 1,
        })
        _set_row(ws, 7, {
            "A": "pll_n1_to_logic[31]",
            "B": "pll_n1_to_logic[30:23]",
            "C": "pll_n1_to_logic[22:16]",
            "D": "pll_n1_to_logic[15:0]",
            "E": "en_dig_clk_div2_to_logic",
            "K": "pll_n2[31:0]",
            "L": "E?{B,C,D,1'b0}:{A,B,C,D}",
            "M": "to_logic", "N": 0, "O": "internal", "P": "Yao Wang", "R": 2,
        })
        _set_row(ws, 8, {
            "A": "pll_n2_to_logic[31]",
            "B": "pll_n2_to_logic[30:23]",
            "C": "pll_n2_to_logic[22:16]",
            "D": "pll_n2_to_logic[15:0]",
            "E": "en_dig_clk_div4_to_logic",
            "K": "pll_n[31:0]",
            "L": "E?{B,C,D,1'b0}:{A,B,C,D}",
            "N": 1, "O": "to_pll_ctrl", "P": "Yao Wang", "R": 3,
        })
        # d_pfd_en_lnmode（镜像真表 row21）：输入 C=mon_active 是 PLL 控制器 FSM 输出，
        # 不在 RF 表里(wire 兜底)，且实际位于 U_BT_LP_PLL_DIG 子模块内部 → 需要探针前缀
        _set_row(ws, 9, {
            "A": "d_pfd_en_lnmode_to_logic[1]",
            "B": "d_pfd_en_lnmode_to_logic[0]",
            "C": "mon_active_to_logic",
            "F": "d_bt_lp_pll_dig_dft_iddq_mode_to_logic",
            "K": "d_pfd_en_lnmode",
            "L": "(A?B:C)&(~F)",
            "M": "ls", "N": 1, "O": "lnmode", "P": "Yao Wang", "R": 19,
        })

    # ───────── regmap 页 ─────────
    rm = wb.create_sheet("regmap")
    _set_row(rm, 2, {
        "D": "Reg_Name", "F": "Reg Type", "G": "Signal_Name", "I": "default",
        "J": "b15", "Y": "b0", "Z": "suffix", "AE": "owner",
    })
    # 仅作类型/兜底来源，bit 位置以 tmm 为准；这里给几条
    _regmap_row(rm, 3, "BT_MODE", "RW", "d_bt_lp_bt_mode_sel", bit0=True)
    _regmap_row(rm, 4, "LINELOCAL", "RW", "d_bt_lp_linelocal_mode_ctrl", bit0=True)
    _regmap_row(rm, 5, "LINELOCAL", "RW", "d_bt_lp_bt_mode_sel_local", bit1=True)
    _regmap_row(rm, 6, "IDDQ_CTRL", "RO", "d_bt_lp_iddq", bit0=True)

    # ───────── total_memory_map 页 ─────────
    tmm = wb.create_sheet("total_memory_map")
    r = 1
    # 列：A=name B=bit/addr C=reset D=DIGPIN/type E=func F=field_addr G=- H=type
    r = _tmm_reg(tmm, r, "BT_MODE", "h2C")
    r = _tmm_field(tmm, r, "d_bt_lp_bt_mode_sel", "0", addr="h2C", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "LINELOCAL_CTRL", "h2D")
    r = _tmm_field(tmm, r, "d_bt_lp_linelocal_mode_ctrl", "0", addr="h2D", dig="N", typ="RW")
    r = _tmm_field(tmm, r, "d_bt_lp_bt_mode_sel_local", "1", addr="h2D", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "IDDQ_CTRL", "h30")
    r = _tmm_field(tmm, r, "d_bt_lp_iddq", "0", addr="h30", dig="Y", typ="RO")
    r = _tmm_reg(tmm, r, "LNA_SEL", "h31")
    r = _tmm_field(tmm, r, "d_bt_lp_lna_line_sel", "0", addr="h31", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "LNA_AGC", "h32")
    r = _tmm_field(tmm, r, "d_bt_lp_lna_agc_local", "2:0", addr="h32", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "LNA_AGC_LINE", "h33")
    r = _tmm_field(tmm, r, "d_bt_lp_lna_agc_line", "2:0", addr="h33", dig="Y", typ="RO")
    r = _tmm_reg(tmm, r, "REFBUF", "h10")
    r = _tmm_field(tmm, r, "d_bt_lp_en_refbuf_cfg", "0", addr="h10", dig="N", typ="RW")

    # pll_n cone 链的叶子寄存器（地址/位段 1:1 取自真表 total_memory_map）
    if with_pll_chain:
        r = _tmm_reg(tmm, r, "TOP_EN", "h1")
        r = _tmm_field(tmm, r, "en_dig_clk_div2", "7", addr="h1", dig="N", typ="RW")
        r = _tmm_field(tmm, r, "en_dig_clk_div4", "6", addr="h1", dig="N", typ="RW")
        r = _tmm_reg(tmm, r, "INT_N", "h2")
        r = _tmm_field(tmm, r, "int_n[8:0]", "15:7", addr="h2", dig="N", typ="RW")
        r = _tmm_field(tmm, r, "frac_n_msb[6:0]", "6:0", addr="h2", dig="N", typ="RW")
        r = _tmm_reg(tmm, r, "INT_N", "h3")
        r = _tmm_field(tmm, r, "frac_n_lsb[15:0]", "15:0", addr="h3", dig="N", typ="RW")
        r = _tmm_reg(tmm, r, "REFBUF2", "hC")
        r = _tmm_field(tmm, r, "d_xo_freq_sel", "0", addr="hC", dig="N", typ="RW")
        # d_pfd_en_lnmode 用：寄存器位自身 RW + iddq RO；mon_active 故意不入表(wire 兜底)
        r = _tmm_reg(tmm, r, "ANA_CTRL", "hD")
        r = _tmm_field(tmm, r, "d_pfd_en_lnmode[1:0]", "15:14", addr="hD", dig="N", typ="RW")
        r = _tmm_reg(tmm, r, "IDDQ_REG", "h29")
        r = _tmm_field(tmm, r, "d_bt_lp_pll_dig_dft_iddq_mode", "2", addr="h29", dig="Y", typ="RO")

    # ───────── mux 页（with_mux=True 才生成）─────────
    # 镜像真表排版：A=mux_input B=ctrl_sig1 C/D/E=预留 F=case G=mux_out I=top_out L=Owner N=组号
    # 控制信号 = logic 页已有的 to_mux 行 d_logic_bt_lp_lna_agc（行4）→ logic↔mux 衔接真实成立：
    #   line 路径: force d_bt_lp_lna_agc_line[2:0] + RF_WRITE h31 bit0=0
    #   local 路径: RF_WRITE h32(lna_agc_local) + RF_WRITE h31 bit0=1
    if with_mux:
        # mux 数据寄存器入 tmm（g1/g2 同地址不同字段 = 镜像真表 t1..t4 同住一个 16bit 寄存器）
        r = _tmm_reg(tmm, r, "RCCAL_I_A", "h40")
        r = _tmm_field(tmm, r, "d_bt_lp_rccal_i_g1[3:0]", "3:0", addr="h40", dig="N", typ="RW")
        r = _tmm_field(tmm, r, "d_bt_lp_rccal_i_g2[3:0]", "7:4", addr="h40", dig="N", typ="RW")
        r = _tmm_reg(tmm, r, "RCCAL_I_B", "h41")
        r = _tmm_field(tmm, r, "d_bt_lp_rccal_i_g3[3:0]", "3:0", addr="h41", dig="N", typ="RW")
        r = _tmm_reg(tmm, r, "BIAS_Q", "h42")
        r = _tmm_field(tmm, r, "d_bt_lp_bias_q_t1[1:0]", "1:0", addr="h42", dig="N", typ="RW")
        r = _tmm_field(tmm, r, "d_bt_lp_bias_q_t2[1:0]", "3:2", addr="h42", dig="N", typ="RW")

        mx = wb.create_sheet("mux")
        _set_row(mx, 1, {"O": "There is No error in the selected rows."})
        _set_row(mx, 2, {
            "A": "mux_input", "B": "mux_ctrl_sig1", "C": "mux_ctrl_sig2",
            "D": "mux_ctrl_sig3", "E": "mux_ctrl_sig4", "F": "case", "G": "mux_out",
            "H": "suffix", "I": "top_out", "J": "Notes", "L": "Owner", "N": 0,
        })
        # 组1: rccal_i = case(lna_agc) 三选一, 含一个 don't-care case (3'b10x)
        _mux_row(mx, 3, "d_bt_lp_rccal_i_g1_to_mux[3:0]", "3'b010",
                 "d_bt_lp_rccal_i[3:0]", owner="Alice", n=1)
        _mux_row(mx, 4, "d_bt_lp_rccal_i_g2_to_mux[3:0]", "3'b011",
                 "d_bt_lp_rccal_i[3:0]", owner="Alice", n=1)
        _mux_row(mx, 5, "d_bt_lp_rccal_i_g3_to_mux[3:0]", "3'b10x",
                 "d_bt_lp_rccal_i[3:0]", owner="Alice", n=1)
        # (reserved) 行: F 列空 → read_mux 应过滤
        _set_row(mx, 6, {"A": "(reserved)", "G": "(reserved)", "I": 0})
        # 组2: bias_q = case(lna_agc) 二选一
        _mux_row(mx, 7, "d_bt_lp_bias_q_t1_to_mux[1:0]", "3'b000",
                 "d_bt_lp_bias_q[1:0]", owner="Bob", n=2)
        _mux_row(mx, 8, "d_bt_lp_bias_q_t2_to_mux[1:0]", "3'b001",
                 "d_bt_lp_bias_q[1:0]", owner="Bob", n=2)

    wb.save(path)
    return path


def build_wl_workbook(path):
    """构造镜像 WL_RFTRX 真表结构的合成工作簿（2026-06-03 两轮 inspect_mux 实证）。

    与 LPBT（build_workbook）的关键差异，全部按真表 1:1 还原：
      ① 控制信号大多是寄存器直出（RW，RF_WRITE 直接驱动），不是 logic 页行
      ② 数据输入分 RO 线控（linectrl_*，force）和 RW 本地/lut（RF_WRITE）
      ③ mux 套 mux 级联（组2 bwctrl 的输出是组3/组4 的控制信号）
      ④ 多控制拼接（组4：case = {lut_en(1bit), bwctrl(4bit)}，B 高位，含 don't-care）
      ⑤ top_out 全 0（输出不在 DUT 顶层，探针需 scan_rtl 前缀）
      ⑥ mux 输出反过来喂 logic 输入（mux→logic 级联）

    五个 mux 组：
      组1 lna_gain[2:0]   = case(lna_gain_ctrl_mode)     {0:线控RO; 1:local RW}    ← WL 最常见形态
      组2 bwctrl[3:0]     = case(bwctrl_mode)            {0:线控RO; 1:local RW}    ← 被组3/4 级联
      组3 tx_bwctrl[4:0]  = case(bwctrl[3:0])            {0..3: lut0..3}           ← mux 级联控制
      组4 tx_rc_code[5:0] = case({lut_en,bwctrl[3:0]})   {0xxxx:local; 1NNNN:lutN} ← 多控制+级联+x位
      组5 dpd_path[1:0]   = case(fb_en)                  {0:dpd_a; 1:dpd_b}        ← logic 行控制(LPBT形态)
    """
    wb = openpyxl.Workbook()

    # ───────── logic 页 ─────────
    ws = wb.active
    ws.title = "logic"
    _set_row(ws, 1, {"A": "inputs", "K": "logic"})
    _set_row(ws, 2, {
        "A": "in_A", "B": "in_B", "C": "in_C", "J": "in_J",
        "K": "logic_out_signal_name", "L": "logic expression", "M": "suffix",
        "N": "top_output", "O": "Notes", "P": "owner", "R": "no",
    })
    # 组5 的控制信号：logic 行（LPBT 形态，A?C:B = 模式选 line/local）
    _set_row(ws, 3, {
        "A": "d_wl_rf_fb_mode_to_logic",
        "B": "d_wl_rf_fb_line_to_logic",
        "C": "d_wl_rf_fb_local_to_logic",
        "K": "d_wl_rf_fb_en",
        "L": "A?C:B",
        "M": "to_mux", "N": 1, "O": "ctrl for dpd_path mux", "P": "Owner1", "R": 1,
    })
    # mux→logic 级联：logic 输入 = 组1 mux 的输出（真表 to_logic 去向的还原）
    _set_row(ws, 4, {
        "A": "d_wl_rf_lna_gain_to_logic[2:0]",
        "K": "d_wl_rf_lna_gain_dly[2:0]",
        "L": "A",
        "M": "ls", "N": 1, "O": "downstream of mux", "P": "Owner1", "R": 2,
    })

    # ───────── regmap 页（最小化，类型以 tmm 为准）─────────
    rm = wb.create_sheet("regmap")
    _set_row(rm, 2, {
        "D": "Reg_Name", "F": "Reg Type", "G": "Signal_Name", "I": "default",
        "J": "b15", "Y": "b0", "Z": "suffix", "AE": "owner",
    })

    # ───────── total_memory_map 页 ─────────
    tmm = wb.create_sheet("total_memory_map")
    r = 1
    # 模式寄存器（控制信号寄存器直出，RW）
    r = _tmm_reg(tmm, r, "WL_MODE_CTRL", "h50")
    r = _tmm_field(tmm, r, "d_wl_rf_lna_gain_ctrl_mode", "0", addr="h50", dig="N", typ="RW")
    r = _tmm_field(tmm, r, "d_wl_rf_bwctrl_mode", "1", addr="h50", dig="N", typ="RW")
    r = _tmm_field(tmm, r, "d_wl_rf_rc_code_lut_en", "2", addr="h50", dig="N", typ="RW")
    # 组1 数据：local（RW）+ 线控（RO）
    r = _tmm_reg(tmm, r, "WL_LNA_GAIN", "h51")
    r = _tmm_field(tmm, r, "d_wl_rf_lna_gain_local[2:0]", "2:0", addr="h51", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "WL_LNA_LINE", "h52")
    r = _tmm_field(tmm, r, "d_wl_rf_linectrl_lna_gain[2:0]", "2:0", addr="h52", dig="Y", typ="RO")
    # 组2 数据
    r = _tmm_reg(tmm, r, "WL_BWCTRL", "h53")
    r = _tmm_field(tmm, r, "d_wl_rf_bwctrl_local[3:0]", "3:0", addr="h53", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "WL_BW_LINE", "h54")
    r = _tmm_field(tmm, r, "d_wl_rf_linectrl_bwctrl[3:0]", "3:0", addr="h54", dig="Y", typ="RO")
    # 组3 数据：lut0..3（RW，两两同住一个寄存器 = 真表排布）
    r = _tmm_reg(tmm, r, "WL_TX_BW_LUT01", "h55")
    r = _tmm_field(tmm, r, "d_wl_rf_tx_bw_lut0[4:0]", "4:0", addr="h55", dig="N", typ="RW")
    r = _tmm_field(tmm, r, "d_wl_rf_tx_bw_lut1[4:0]", "9:5", addr="h55", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "WL_TX_BW_LUT23", "h56")
    r = _tmm_field(tmm, r, "d_wl_rf_tx_bw_lut2[4:0]", "4:0", addr="h56", dig="N", typ="RW")
    r = _tmm_field(tmm, r, "d_wl_rf_tx_bw_lut3[4:0]", "9:5", addr="h56", dig="N", typ="RW")
    # 组4 数据：local + lut0..3
    r = _tmm_reg(tmm, r, "WL_RC_LOCAL", "h57")
    r = _tmm_field(tmm, r, "d_wl_rf_tx_rc_local[5:0]", "5:0", addr="h57", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "WL_RC_LUT01", "h58")
    r = _tmm_field(tmm, r, "d_wl_rf_tx_rc_lut0[5:0]", "5:0", addr="h58", dig="N", typ="RW")
    r = _tmm_field(tmm, r, "d_wl_rf_tx_rc_lut1[5:0]", "11:6", addr="h58", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "WL_RC_LUT23", "h59")
    r = _tmm_field(tmm, r, "d_wl_rf_tx_rc_lut2[5:0]", "5:0", addr="h59", dig="N", typ="RW")
    r = _tmm_field(tmm, r, "d_wl_rf_tx_rc_lut3[5:0]", "11:6", addr="h59", dig="N", typ="RW")
    # 组5（logic 控制）数据 + fb_en 的 logic 输入寄存器
    r = _tmm_reg(tmm, r, "WL_DPD_PATH", "h5A")
    r = _tmm_field(tmm, r, "d_wl_rf_dpd_a[1:0]", "1:0", addr="h5A", dig="N", typ="RW")
    r = _tmm_field(tmm, r, "d_wl_rf_dpd_b[1:0]", "3:2", addr="h5A", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "WL_FB_CTRL", "h5B")
    r = _tmm_field(tmm, r, "d_wl_rf_fb_mode", "0", addr="h5B", dig="N", typ="RW")
    r = _tmm_field(tmm, r, "d_wl_rf_fb_local", "1", addr="h5B", dig="N", typ="RW")
    r = _tmm_reg(tmm, r, "WL_FB_LINE", "h5C")
    r = _tmm_field(tmm, r, "d_wl_rf_fb_line", "0", addr="h5C", dig="Y", typ="RO")

    # ───────── mux 页（top_out 全 0 = 真表实况）─────────
    mx = wb.create_sheet("mux")
    _set_row(mx, 1, {"O": "There is No error in the selected rows."})
    _set_row(mx, 2, {
        "A": "mux_input", "B": "mux_ctrl_sig1", "C": "mux_ctrl_sig2",
        "D": "mux_ctrl_sig3", "E": "mux_ctrl_sig4", "F": "case", "G": "mux_out",
        "H": "to_logic", "I": "top_out", "J": "Notes", "L": "Owner", "N": 0,
    })
    # 组1: lna_gain = case(ctrl_mode) {0:线控RO; 1:local RW}   ← H=to_logic(喂 logic 页行4)
    _wl_mux_row(mx, 3, "d_wl_rf_linectrl_lna_gain_to_mux[2:0]",
                {"B": "d_wl_rf_lna_gain_ctrl_mode_to_mux"}, "1'b0",
                "d_wl_rf_lna_gain[2:0]", h="to_logic", n=1)
    _wl_mux_row(mx, 4, "d_wl_rf_lna_gain_local_to_mux[2:0]",
                {"B": "d_wl_rf_lna_gain_ctrl_mode_to_mux"}, "1'b1",
                "d_wl_rf_lna_gain[2:0]", h="to_logic", n=1)
    # 组2: bwctrl = case(bwctrl_mode) {0:线控; 1:local}   ← H=to_mux(输出喂组3/4 的控制)
    _wl_mux_row(mx, 5, "d_wl_rf_linectrl_bwctrl_to_mux[3:0]",
                {"B": "d_wl_rf_bwctrl_mode_to_mux"}, "1'b0",
                "d_wl_rf_bwctrl[3:0]", h="to_mux", n=2)
    _wl_mux_row(mx, 6, "d_wl_rf_bwctrl_local_to_mux[3:0]",
                {"B": "d_wl_rf_bwctrl_mode_to_mux"}, "1'b1",
                "d_wl_rf_bwctrl[3:0]", h="to_mux", n=2)
    # 组3: tx_bwctrl = case(bwctrl[3:0]) {0..3: lut0..3}   ← 控制信号是组2 的输出（mux 级联）
    for i in range(4):
        _wl_mux_row(mx, 7 + i, "d_wl_rf_tx_bw_lut%d_to_mux[4:0]" % i,
                    {"B": "d_wl_rf_bwctrl_to_mux[3:0]"}, "4'b%s" % format(i, "04b"),
                    "d_wl_rf_tx_bwctrl[4:0]", h="to_dft", n=3)
    # 组4: tx_rc_code = case({lut_en, bwctrl}) {0xxxx:local; 1NNNN:lutN}   ← 多控制拼接(B 高位)
    _wl_mux_row(mx, 11, "d_wl_rf_tx_rc_local_to_mux[5:0]",
                {"B": "d_wl_rf_rc_code_lut_en_to_mux", "C": "d_wl_rf_bwctrl_to_mux[3:0]"},
                "5'b0xxxx", "d_wl_rf_tx_rc_code[5:0]", h="to_dft", n=4)
    for i in range(4):
        _wl_mux_row(mx, 12 + i, "d_wl_rf_tx_rc_lut%d_to_mux[5:0]" % i,
                    {"B": "d_wl_rf_rc_code_lut_en_to_mux", "C": "d_wl_rf_bwctrl_to_mux[3:0]"},
                    "5'b1%s" % format(i, "04b"), "d_wl_rf_tx_rc_code[5:0]", h="to_dft", n=4)
    # 组5: dpd_path = case(fb_en) {0:dpd_a; 1:dpd_b}   ← 控制信号是 logic 行（LPBT 形态）
    _wl_mux_row(mx, 16, "d_wl_rf_dpd_a_to_mux[1:0]",
                {"B": "d_wl_rf_fb_en_to_mux"}, "1'b0",
                "d_wl_rf_dpd_path[1:0]", h="to_dft", n=5)
    _wl_mux_row(mx, 17, "d_wl_rf_dpd_b_to_mux[1:0]",
                {"B": "d_wl_rf_fb_en_to_mux"}, "1'b1",
                "d_wl_rf_dpd_path[1:0]", h="to_dft", n=5)

    wb.save(path)
    return path


def _wl_mux_row(ws, row, mux_input, ctrls, case, mux_out, h="to_logic", owner="Owner1", n=None):
    """WL mux 页数据行：ctrls={"B": 控制1, "C": 控制2, ...}，top_out 固定 0（真表实况）。"""
    cells = {"A": mux_input, "F": case, "G": mux_out, "H": h, "I": 0, "L": owner, "N": n}
    cells.update(ctrls)
    _set_row(ws, row, cells)


def _mux_row(ws, row, mux_input, case, mux_out, owner="", n=None,
             ctrl="d_logic_bt_lp_lna_agc_to_mux[2:0]"):
    """mux 页数据行（列语义与真表一致）。"""
    _set_row(ws, row, {"A": mux_input, "B": ctrl, "F": case, "G": mux_out,
                       "I": 1, "J": mux_out, "L": owner, "N": n})


def _regmap_row(ws, row, reg, typ, signal, bit0=False, bit1=False):
    from openpyxl.utils import column_index_from_string, get_column_letter
    _set_row(ws, row, {"D": reg, "F": typ, "G": signal})
    # J..Y = bit15..bit0；bit0 在 Y 列, bit1 在 X 列
    if bit0:
        ws.cell(row=row, column=column_index_from_string("Y"), value=1)
    if bit1:
        ws.cell(row=row, column=column_index_from_string("X"), value=1)


def _tmm_reg(ws, r, name, addr):
    _set_row(ws, r, {"A": name, "B": addr, "C": "0", "D": "RW", "E": "reg def"})
    return r + 1


def _tmm_field(ws, r, name, bit, addr, dig, typ):
    _set_row(ws, r, {"A": name, "B": bit, "C": "0", "D": dig, "E": "field",
                     "F": addr, "H": typ})
    return r + 1


if __name__ == "__main__":
    import sys
    build_workbook(sys.argv[1] if len(sys.argv) > 1 else "synthetic_dreg.xlsx")
    print("done")
