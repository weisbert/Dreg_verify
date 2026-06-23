# -*- coding: utf-8 -*-
"""Cone 展开测试 — 用真表 pll_n←pll_n2←pll_n1 三层链做 oracle（2026-06-02 实证数据）。

for_test 对 pll_n 的处理（块 行154..160）：
  驱动 6 个叶子寄存器: en_dig_clk_div2(0x1 bit7) en_dig_clk_div4(0x1 bit6)
                       int_n[8:0](0x2 15:7)  frac_n_msb[6:0](0x2 6:0)
                       frac_n_lsb[15:0](0x3) d_xo_freq_sel(0xC bit0)
  T0: int_n=256 frac_n_msb=64 frac_n_lsb=0xFFFF 其余=1
  → RF_WRITE 0x1=0xC0, 0x2=0x8040, 0x3=0xFFFF, 0xC=0x1
  期望(按 logic 表达式) pll_n = pll_n1 << 2 = 0x0081FFFC（for_test 写的 32'b1 是错的，不抄）。
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import fixtures                                     # noqa: E402
from dreg_verify import cone, excel_model, generator  # noqa: E402
from dreg_verify import expr as E                   # noqa: E402
from dreg_verify import resolver as R               # noqa: E402
from dreg_verify import vectors as V                # noqa: E402


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    path = tmp_path_factory.mktemp("cone") / "synthetic_pll.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    return excel_model.load_workbook(str(path))


@pytest.fixture(scope="module")
def resolver(wb):
    return R.Resolver(wb)


def test_level_shift_top_port_probe(tmp_path):
    """新表 level_shift 页 → logic 输出探针定到顶层电平移位口 _ls（2026-06-12 Hi1108 Pilot）。

    d_en_refbuf: M=to_dft、top_output=0、自引用输入 d_en_refbuf_to_logic（撞名）——裸名在 RTL
    不存在，真顶层口是 d_en_refbuf_ls（level_shift 页 + lpbt_dig_top.v output 双证）。验证：
      ① 探针网名 = d_en_refbuf_ls（不是裸名、不是 _to_logic）
      ② 『logic 加尾缀』开关 ON/OFF 都压不动它（_ls_name 最优先，治『点了加尾缀没生效』）
      ③ 无 level_shift 条目的 d_plain 不受影响（自引用排除后仍裸名）
    """
    path = tmp_path / "levelshift.xlsx"
    fixtures.build_levelshift_workbook(str(path))
    wb = excel_model.load_workbook(str(path))
    refbuf = next(s for s in wb.logic if s.out_base == "d_en_refbuf")
    plain = next(s for s in wb.logic if s.out_base == "d_plain")
    assert refbuf._ls_name == "d_en_refbuf_ls"
    assert refbuf.rtl_name == "d_en_refbuf_ls"
    assert refbuf.rtl_base_full == "d_en_refbuf_ls"
    # 加尾缀开关两态都压不动 _ls 顶层口（这正是用户『点了 logic 加尾缀没生效』的根因解法）
    refbuf._append_to_logic = False
    assert refbuf.rtl_name == "d_en_refbuf_ls"
    refbuf._append_to_logic = True
    assert refbuf.rtl_name == "d_en_refbuf_ls"
    # 对照：不过 level_shift 的输出按原逻辑（M=to_dft 非 ls + 自引用排除 → 裸名，无 _to_logic）
    assert plain._ls_name is None
    assert plain.rtl_name == "d_plain"
    # ④ 走 level_shift 的输出：前缀【只认 _ls 真网名，不认 K 列裸名】
    #    (2026-06-12：续2 曾对 level_shift 输出一律清前缀，把 pll_n 配的真前缀也误清 → CUVUNF；改为按真名匹配)
    # 4a. d_en_refbuf：真名=d_en_refbuf_ls。按裸名 d_en_refbuf 配的前缀(指向移位前/消费侧)→忽略
    assert generator.probe_prefix_for(
        refbuf, generator.GenOptions(probe_prefixes={"d_en_refbuf": "U_CONSUMER"})) == ""
    # 4b. 按【_ls 真网名】配 → 生效
    assert generator.probe_prefix_for(
        refbuf, generator.GenOptions(probe_prefixes={"d_en_refbuf_ls": "U_RIGHT"})) == "U_RIGHT"
    # 4c. pll_n：经 level_shift 但真名仍是裸名 pll_n、RTL 在子模块——用户按 pll_n 配的前缀必须生效(复刻真实 CUVUNF)
    pll = next(s for s in wb.logic if s.out_base == "pll_n")
    assert pll._ls_name == "pll_n"
    assert pll.rtl_name == "pll_n[31:0]"
    assert generator.probe_prefix_for(
        pll, generator.GenOptions(probe_prefixes={"pll_n": "U_BT_LP_PLL_DIG"})) == "U_BT_LP_PLL_DIG"


# ───────────── expr.Part 节点 ─────────────
def test_part_eval_and_width():
    node = E.Part(E.Var("A"), 30, 23)
    env = E.Env({"A": 32}, {"A": 0x40207FFF})
    assert E.self_width(node, env) == 8
    # 0x40207FFF bits[30:23] = 0x80
    assert E.evaluate(node, env, 8)[0] == 0x80
    # Part 嵌套: ((A)[30:23])[7]  → bit7 of 0x80 = 1
    nested = E.Part(E.Part(E.Var("A"), 30, 23), 7, 7)
    assert E.evaluate(nested, env, 1)[0] == 1
    assert E.collect_vars(nested) == ["A"]


# ───────────── 展开结构 ─────────────
def test_expand_pll_chain_leaves(wb, resolver):
    sig = next(s for s in wb.logic if s.out_base == "pll_n")
    node, bindings = cone.expand(sig, wb, resolver)
    # 叶子 = 6 个真寄存器，全部 RW 可驱动
    assert set(bindings) == {"INT_N", "FRAC_N_MSB", "FRAC_N_LSB",
                             "D_XO_FREQ_SEL", "EN_DIG_CLK_DIV2", "EN_DIG_CLK_DIV4"}
    assert all(b.kind == "RW" and b.address is not None for b in bindings.values())
    # frac_n_lsb 两个切片([15:1]与[0])合并成一个 16bit 叶子
    assert bindings["FRAC_N_LSB"].width == 16
    assert bindings["INT_N"].width == 9
    # 地址来自 tmm
    assert bindings["EN_DIG_CLK_DIV2"].address == 0x1
    assert bindings["INT_N"].address == 0x2
    assert bindings["FRAC_N_LSB"].address == 0x3
    assert bindings["D_XO_FREQ_SEL"].address == 0xC


def test_cone_leaves_carry_excel_sources(wb, resolver):
    """⭐cone 展开后的字母溯源(2026-06-03 用户报告的显示 bug)：

    修复前 GUI『输入信号』表/报告的字母列显示大写基名(INT_N)——与信号列重复且对不回 Excel。
    修复后叶子带 xl_letters = Excel 来源坐标：
      根行(pll_n)的直接输入 → 列字母本身("E")
      上游行展开来的叶子   → "上游行名.字母"("pll_n1.A")"""
    sig = next(s for s in wb.logic if s.out_base == "pll_n")
    node, bindings = cone.expand(sig, wb, resolver)
    groups = V.input_groups(node, bindings)
    by_base = {g["base"].lower(): g for g in groups}
    # 根行直接输入：en_dig_clk_div4 = pll_n 行的 E 列
    assert by_base["en_dig_clk_div4"]["xl_letters"] == ["E"]
    # 上游行的叶子："行名.字母"
    assert by_base["en_dig_clk_div2"]["xl_letters"] == ["pll_n2.E"]
    assert by_base["int_n"]["xl_letters"] == ["pll_n1.A"]
    assert by_base["frac_n_msb"]["xl_letters"] == ["pll_n1.B"]
    assert by_base["d_xo_freq_sel"]["xl_letters"] == ["pll_n1.E"]
    # 同一寄存器的两个切片(frac_n_lsb[15:1]+[0]) → 来源累积成 C、D 两条
    assert by_base["frac_n_lsb"]["xl_letters"] == ["pll_n1.C", "pll_n1.D"]
    # 多路径不爆炸：pll_n 的 A/B/C/D ×  pll_n2 的 A/B/C/D 共 16 条路径到达 int_n，
    # 但来源坐标只有 1 条(去重)
    assert len(by_base["int_n"]["xl_letters"]) == 1
    # letters(表达式变量名/求值用)保持叶子基名不变——求值/驱动路径不受显示字段影响
    assert by_base["int_n"]["letters"] == ["INT_N"]


def test_cone_expand_chain(wb, resolver):
    """⭐展开链(2026-06-03 用户功能)：cone 展开时记录展开过程——本行+逐层代入的上游行，
    每行带 Excel 原式 + "字母代入真实信号名"的等价形式（GUI/HTML 报告显示）。"""
    sig = next(s for s in wb.logic if s.out_base == "pll_n")
    chain = []
    node, bindings = cone.expand(sig, wb, resolver, chain_out=chain)
    # 顺序: 本行(根) → pll_n2 → pll_n1 (DFS)；同一上游行被 A/B/C/D 多次到达只记一次
    assert [c["out"] for c in chain] == ["pll_n", "pll_n2", "pll_n1"]
    # 原式 = Excel L 列原文
    assert chain[0]["expr"] == "E?{B,C,D,1'b0}:{A,B,C,D}"
    assert chain[2]["expr"] == "E?{1'b0,A,B,C}:{A,B,C,D}"
    # 代入形式: 字母 → 真实信号名(去 _to_logic 的基名 + Excel 单元格切片)
    assert chain[0]["subst"] == ("en_dig_clk_div4?{pll_n2[30:23],pll_n2[22:16],pll_n2[15:0],1'b0}"
                                 ":{pll_n2[31],pll_n2[30:23],pll_n2[22:16],pll_n2[15:0]}")
    assert chain[2]["subst"] == ("d_xo_freq_sel?{1'b0,int_n[8:0],frac_n_msb[6:0],frac_n_lsb[15:1]}"
                                 ":{int_n[8:0],frac_n_msb[6:0],frac_n_lsb[15:1],frac_n_lsb[0]}")
    # 不传 chain_out 的现有调用不受影响(返回值/叶子完全一致)
    _node2, bindings2 = cone.expand(sig, wb, resolver)
    assert set(bindings2) == set(bindings)


def test_expand_signal_chain_out(wb, resolver):
    """generator.expand_signal 透传 chain_out；非 cone 信号不动它(保持空)。"""
    sig = next(s for s in wb.logic if s.out_base == "pll_n")
    chain = []
    _n, _b, expanded = generator.expand_signal(wb, resolver, sig, chain_out=chain)
    assert expanded and [c["out"] for c in chain] == ["pll_n", "pll_n2", "pll_n1"]
    sig2 = next(s for s in wb.logic if s.out_base == "d_en_refbuf")
    chain2 = []
    _n, _b, expanded2 = generator.expand_signal(wb, resolver, sig2, chain_out=chain2)
    assert not expanded2 and chain2 == []


def test_report_and_html_carry_chain(wb, tmp_path):
    """报告数据带 chain；HTML 报告 cone 信号真值表上方渲染展开链。"""
    from dreg_verify import cli
    rep = generator.report(wb, generator.GenOptions(top_output_only=False))
    by_sig = {t["signal"]: t for t in rep["tables"]}
    assert [c["out"] for c in by_sig["pll_n[31:0]"]["chain"]] == ["pll_n", "pll_n2", "pll_n1"]
    assert by_sig["d_en_refbuf"]["chain"] == []          # 非 cone 信号链为空
    # HTML 渲染（esc 会把 1'b0 的引号转义，断言避开常量）
    path = tmp_path / "r.html"
    cli.write_report(str(path), rep, "synthetic.xlsx")
    html = path.read_text(encoding="utf-8")
    assert "展开链" in html
    assert "① pll_n = E?" in html
    assert "③ pll_n1 = E?" in html
    assert "pll_n2[30:23],pll_n2[22:16],pll_n2[15:0]" in html          # 代入形式
    assert "d_xo_freq_sel?{1" in html


def test_non_cone_groups_xl_letters_equal_letters(wb, resolver):
    """非 cone 信号：xl_letters = 表达式字母本身(A/B/C…)，显示与修复前完全一致。"""
    for base in ("d_pfd_en_lnmode", "d_en_refbuf"):
        sig = next(s for s in wb.logic if s.out_base == base)
        node = E.parse(sig.expr)
        bindings = resolver.resolve_signal_inputs(sig)
        for g in V.input_groups(node, bindings):
            assert g["xl_letters"] == g["letters"]


def test_cone_report_letters_show_excel_sources(wb):
    """HTML/CSV 报告真值表：cone 信号的字母列 = Excel 来源坐标，不再是大写基名。"""
    rep = generator.report(wb, generator.GenOptions(signals=["pll_n"], top_output_only=False))
    t = rep["tables"][0]
    letters_by_label = {i["label"]: i["letters"] for i in t["inputs"]}
    assert letters_by_label["int_n[8:0]"] == "pll_n1.A"
    assert letters_by_label["en_dig_clk_div4"] == "E"
    assert letters_by_label["frac_n_lsb[15:0]"] == "pll_n1.C,pll_n1.D"
    # 大写基名(修复前的显示)不再出现
    assert "INT_N" not in str(letters_by_label.values())


def test_expand_expected_value_matches_logic(wb, resolver):
    """for_test T0 输入 → 期望值按 logic 表达式 = 0x0081FFFC（pll_n1 左移 2 位）。"""
    sig = next(s for s in wb.logic if s.out_base == "pll_n")
    node, bindings = cone.expand(sig, wb, resolver)
    widths = {k: b.width for k, b in bindings.items()}
    values = {"INT_N": 256, "FRAC_N_MSB": 64, "FRAC_N_LSB": 0xFFFF,
              "D_XO_FREQ_SEL": 1, "EN_DIG_CLK_DIV2": 1, "EN_DIG_CLK_DIV4": 1}
    val, w = E.evaluate(node, E.Env(widths, values), sig.out_width)
    # pll_n1 = {1'b0, 256(9b), 64(7b), 0x7FFF(15b)} = 0x40207FFF；左移两次(高位丢弃) = 0x0081FFFC
    assert w == 32
    assert val == 0x0081FFFC
    # 控制位 = 三个时钟/频率选择，数据 = int_n/frac_n
    control, data = E.classify_vars(node, E.Env(widths))
    assert set(control) == {"D_XO_FREQ_SEL", "EN_DIG_CLK_DIV2", "EN_DIG_CLK_DIV4"}
    assert set(data) == {"INT_N", "FRAC_N_MSB", "FRAC_N_LSB"}


def test_expand_only_when_needed(wb, resolver):
    """没有内部输入的信号不做 cone 展开（reserve 原样）。"""
    sig = next(s for s in wb.logic if s.out_base == "d_logic_bt_lp_reserve")
    node, bindings, expanded = generator.expand_signal(wb, resolver, sig)
    assert not expanded
    assert set(bindings) == {"A", "B", "C", "J"}


def test_expand_cycle_detection(resolver):
    """内部信号循环引用(top→a→b→a) → ConeError 而非死循环。"""
    from dreg_verify.excel_model import DregWorkbook, LogicSignal
    mk = lambda name, inp, r, top: LogicSignal(       # noqa: E731
        row=1, out_name=name + "[3:0]", out_width=4, expr="A",
        suffix="to_logic", top_output=top, notes="", owner="", assert_id=r,
        inputs={"A": {"raw": inp, "base": inp, "width": 4, "msb": 3, "lsb": 0}})
    wb2 = DregWorkbook(logic=[mk("sig_top", "sig_a", "1", 1),
                              mk("sig_a", "sig_b", "2", 0),
                              mk("sig_b", "sig_a", "3", 0)],
                       regmap={}, tmm={}, sheet_names=[])
    with pytest.raises(cone.ConeError):
        cone.expand(wb2.logic[0], wb2, R.Resolver(wb2))


# ───────────── 端到端 .sv ─────────────
def test_pll_n_e2e_sv(wb):
    opts = generator.GenOptions(signals=["pll_n"], mode="min")
    res = generator.build(wb, opts)
    assert res["summary"]["n_generated"] == 1      # 不再被跳过
    assert res["summary"]["n_skipped"] == 0
    text = generator.render(res)
    # 驱动 = 4 个地址的 RF_WRITE（同址合并），无 force
    assert "`RF_WRITE(10'h1," in text
    assert "`RF_WRITE(10'h2," in text
    assert "`RF_WRITE(10'h3," in text
    assert "`RF_WRITE(10'hC," in text
    assert "force" not in text
    # 断言探 pll_n（带位宽），期望 32 位
    assert "assert (`ENV_RF.pll_n[31:0]==32'b" in text
    # 标号 = R 列序号
    assert "assert_3_T0:" in text


def test_pll_n_vectors_cover_t0(wb, resolver):
    """min 模式控制位全组合：3 个控制位 → 至少 8 个用例；期望值全部按表达式计算。"""
    sig = next(s for s in wb.logic if s.out_base == "pll_n")
    node, bindings = cone.expand(sig, wb, resolver)
    vecs, meta = V.generate_vectors(node, bindings, sig.out_width, mode="min")
    assert len(vecs) >= 8
    widths = {k: b.width for k, b in bindings.items()}
    for v in vecs:
        val, _ = E.evaluate(node, E.Env(widths, v.assignments), sig.out_width)
        assert val == v.exp_value


def test_pll_n_report_and_verifiability(wb):
    """报告：pll_n 真值表输入为 6 个叶子寄存器；可验证性=clean(经 cone 展开)。"""
    rep = generator.report(wb, generator.GenOptions(signals=["pll_n"], top_output_only=False))
    assert len(rep["tables"]) == 1
    labels = [i["label"] for i in rep["tables"][0]["inputs"]]
    assert any("int_n" in lb for lb in labels)
    assert any("frac_n_lsb" in lb for lb in labels)
    verif = {v["signal"]: v["status"] for v in rep["verifiability"]["signals"]}
    assert verif["pll_n[31:0]"] == "clean"


def test_internal_signals_still_skipped_by_default(wb):
    """pll_n1/pll_n2 自身是内部信号(top_output=0)：默认仍不出现在产物里。"""
    opts = generator.GenOptions(top_output_only=True)
    res = generator.build(wb, opts)
    names = {st["out_name"] for _l, st in res["blocks"]}
    assert "pll_n[31:0]" in names
    assert "pll_n1[31:0]" not in names and "pll_n2[31:0]" not in names


# ───────────── 同名信号多 bit 切片（A=[1] + B=[0]） ─────────────
def test_multi_slice_inputs_expanded_as_rows(wb, resolver):
    """回归(2026-06-02 严重 bug + 用户拍板展示方式)：d_pfd_en_lnmode 的 A=[1]、B=[0] 是同一
    寄存器的两个 bit —— 各自独立成行（A→[1]、B→[0]），取值互相独立，不许坍塌共享。"""
    sig = next(s for s in wb.logic if s.out_base == "d_pfd_en_lnmode")
    node = E.parse(sig.expr)
    bindings = resolver.resolve_signal_inputs(sig)
    groups = V.input_groups(node, bindings)
    by_label = {g["label"]: g for g in groups}
    # A、B 分开两行，各 1bit，标注各自切片
    assert "d_pfd_en_lnmode[1]" in by_label and "d_pfd_en_lnmode[0]" in by_label
    assert by_label["d_pfd_en_lnmode[1]"]["letters"] == ["A"]
    assert by_label["d_pfd_en_lnmode[0]"]["letters"] == ["B"]
    assert by_label["d_pfd_en_lnmode[1]"]["width"] == 1

    # A=1(bit1), B=0(bit0)：互相独立 —— 修复前 A、B 被坍塌成共享一个值
    bv = {by_label["d_pfd_en_lnmode[1]"]["key"]: 1,
          by_label["d_pfd_en_lnmode[0]"]["key"]: 0,
          "mon_active": 0, "d_bt_lp_pll_dig_dft_iddq_mode": 0}
    vec = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width)
    assert vec.assignments["A"] == 1 and vec.assignments["B"] == 0
    assert vec.exp_value == 0                               # (1?0:0)&~0 = 0
    # A=0, C=1 → 期望 1
    bv2 = dict(bv, **{by_label["d_pfd_en_lnmode[1]"]["key"]: 0, "mon_active": 1})
    assert V.make_vector_from_base_values(node, bindings, groups, bv2, sig.out_width).exp_value == 1
    # 自动向量：A、B 各自独立取值 → 存在 A != B 的用例（坍塌时不可能）
    vecs, _meta = V.generate_vectors(node, bindings, sig.out_width, mode="min")
    assert any(v.assignments["A"] != v.assignments["B"] for v in vecs)


def test_multi_slice_rf_write_combines_field(wb):
    """切片合并写寄存器：A=1(bit1) B=0(bit0) → 字段值 0b10，落在 0xD 的 [15:14] → 0x8000。"""
    sig = next(s for s in wb.logic if s.out_base == "d_pfd_en_lnmode")
    from dreg_verify import resolver as R
    res = R.Resolver(wb, wire_prefixes={"mon_active": "U_BT_LP_PLL_DIG"})
    node = E.parse(sig.expr)
    bindings = res.resolve_signal_inputs(sig)
    groups = V.input_groups(node, bindings)
    by_label = {g["label"]: g for g in groups}
    bv = {by_label["d_pfd_en_lnmode[1]"]["key"]: 1,
          by_label["d_pfd_en_lnmode[0]"]["key"]: 0,
          "mon_active": 1, "d_bt_lp_pll_dig_dft_iddq_mode": 0}
    vec = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width)
    from dreg_verify import sv_writer as W
    forces, writes, unres = W.compute_drives(vec, bindings, E.collect_vars(node))
    assert not unres
    by_addr = {w["addr"]: w["hex"] for w in writes}
    assert by_addr["10'hD"] == "16'h8000"          # 0b10 << 14
    # mon_active force 带前缀、iddq force 无前缀
    wires = {f["wire"] for f in forces}
    assert "U_BT_LP_PLL_DIG.mon_active" in wires
    assert "d_bt_lp_pll_dig_dft_iddq_mode" in wires


# ───────────── 探针前缀（输出网在 ENV_RF 子模块里） ─────────────
def test_probe_prefix_in_sv(wb):
    """pll_n 实际在 U_BT_LP_PLL_DIG 内部 → 探针 = `ENV_RF.U_BT_LP_PLL_DIG.pll_n[31:0]。"""
    opts = generator.GenOptions(signals=["pll_n"], mode="min",
                                probe_prefixes={"pll_n": "U_BT_LP_PLL_DIG"})
    text = generator.render(generator.build(wb, opts))
    assert "assert (`ENV_RF.U_BT_LP_PLL_DIG.pll_n[31:0]==32'b" in text
    assert "`ENV_RF.pll_n[31:0]==" not in text
    # 驱动(RF_WRITE)不受前缀影响
    assert "`RF_WRITE(10'h1," in text


def test_probe_prefix_only_for_named_signal(wb):
    """前缀只作用于映射里的信号，其它信号探针不变。"""
    opts = generator.GenOptions(signals=["pll_n", "d_en_refbuf"],
                                probe_prefixes={"pll_n": "U_BT_LP_PLL_DIG"})
    text = generator.render(generator.build(wb, opts))
    assert "`ENV_RF.U_BT_LP_PLL_DIG.pll_n[31:0]==" in text
    assert "`ENV_RF.d_en_refbuf_ls==" in text          # 不带前缀(且带 _ls 后缀)


def test_force_prefix_for_input_wire(wb):
    """mon_active 风格：force 的输入 wire 在 ENV_RF 子模块里 → force 路径带前缀，
    且配置前缀后不再被当 wire 兜底风险跳过。"""
    # 无前缀：mon_active 是 wire 兜底 → 默认 risky → 整个信号被跳过
    res0 = generator.build(wb, generator.GenOptions(signals=["d_pfd_en_lnmode"]))
    assert res0["summary"]["n_generated"] == 0 and res0["summary"]["n_skipped"] == 1

    # 配置前缀：mon_active 在 U_BT_LP_PLL_DIG 下 → 信号可生成，force 带前缀
    opts = generator.GenOptions(signals=["d_pfd_en_lnmode"],
                                probe_prefixes={"mon_active": "U_BT_LP_PLL_DIG"})
    res = generator.build(wb, opts)
    assert res["summary"]["n_generated"] == 1 and res["summary"]["n_skipped"] == 0
    text = generator.render(res)
    assert "force `ENV_RF.U_BT_LP_PLL_DIG.mon_active=" in text
    # 输出探针不带前缀(映射里没有 d_pfd_en_lnmode)，且 ls 信号自动加 _ls 后缀
    assert "assert (`ENV_RF.d_pfd_en_lnmode_ls==" in text
    # 其它输入不受影响：寄存器位 RF_WRITE、iddq force 无前缀
    assert "`RF_WRITE(10'hD," in text
    assert "force `ENV_RF.d_bt_lp_pll_dig_dft_iddq_mode=" in text


def test_probe_prefix_by_rtl_name_with_ls_suffix(wb):
    """回归(2026-06-02 真实 bug)：scan_rtl 导出的映射 key 是 RTL 网名(带 _ls 后缀)，
    必须能命中 Excel K 列名为 d_pfd_en_lnmode(无后缀) 的 ls 信号。"""
    # 映射 key 用带 _ls 后缀的 RTL 网名（scan_rtl 输出格式）
    opts = generator.GenOptions(signals=["d_pfd_en_lnmode"],
                                probe_prefixes={"d_pfd_en_lnmode_ls": "U_BT_LP_PLL_DIG",
                                                "mon_active": "U_BT_LP_PLL_DIG"})
    res = generator.build(wb, opts)
    assert res["summary"]["n_generated"] == 1
    text = generator.render(res)
    # 探针带前缀（且名字带 _ls）
    assert "assert (`ENV_RF.U_BT_LP_PLL_DIG.d_pfd_en_lnmode_ls==" in text
    assert "`ENV_RF.d_pfd_en_lnmode_ls==" not in text
    # K 列名做 key 同样生效（两套名字都认）
    opts2 = generator.GenOptions(signals=["d_pfd_en_lnmode"],
                                 probe_prefixes={"d_pfd_en_lnmode": "U_BT_LP_PLL_DIG",
                                                 "mon_active": "U_BT_LP_PLL_DIG"})
    text2 = generator.render(generator.build(wb, opts2))
    assert "assert (`ENV_RF.U_BT_LP_PLL_DIG.d_pfd_en_lnmode_ls==" in text2


def test_wire_prefix_by_rtl_name_for_cascade_input(wb, resolver):
    """级联输入指向 ls 输出时，force 网名带 _ls 后缀——映射 key 用 RTL 网名(_ls)也要能命中。"""
    from dreg_verify import resolver as R
    # d_en_refbuf 是 ls 输出(非寄存器字段)：级联 force 网名 = d_en_refbuf_ls
    res2 = R.Resolver(wb, wire_prefixes={"d_en_refbuf_ls": "U_BT_LP_PLL_DIG"})
    b = res2.resolve("A", {"raw": "d_en_refbuf", "base": "d_en_refbuf",
                           "width": 1, "msb": None, "lsb": None})
    assert b.kind == "RO" and b.found_in == "logic"
    assert b.wire == "U_BT_LP_PLL_DIG.d_en_refbuf_ls"


def test_internal_signal_rtl_name_to_logic_suffix(wb):
    """内部信号 RTL 网名规则（2026-06-02 BT_LP_DREG_sig_logic.v 实证）：

    被下游 logic 行以 <名>_to_logic 引用的内部信号 → RTL 名 = K + "_to_logic"（pll_n1/pll_n2）；
    未被引用的内部信号（reserve）→ K 原文；顶层输出（pll_n）→ K 原文。"""
    by_base = {s.out_base: s for s in wb.logic}
    assert by_base["pll_n1"].rtl_name == "pll_n1_to_logic[31:0]"     # 被 pll_n2 引用
    assert by_base["pll_n2"].rtl_name == "pll_n2_to_logic[31:0]"     # 被 pll_n 引用
    assert by_base["pll_n"].rtl_name == "pll_n[31:0]"                # 顶层输出, 原样
    assert by_base["d_logic_bt_lp_reserve"].rtl_name == "d_logic_bt_lp_reserve"   # 未被引用, 原样


def test_internal_signal_verifiable_with_prefix(wb):
    """top_output=0 的 pll_n1 也可以验：探针 = `ENV_RF.<层级>.pll_n1_to_logic[31:0]，
    输入是 RW 寄存器照常 RF_WRITE。"""
    prefix = "U_BT_LP_PLL_DIG.U_BT_LP_DREG.U_SIG_LOGIC"
    opts = generator.GenOptions(signals=["pll_n1"], mode="min",
                                top_output_only=False,
                                probe_prefixes={"pll_n1_to_logic": prefix})
    res = generator.build(wb, opts)
    assert res["summary"]["n_generated"] == 1 and res["summary"]["n_skipped"] == 0
    text = generator.render(res)
    assert "assert (`ENV_RF.%s.pll_n1_to_logic[31:0]==32'b" % prefix in text
    # 驱动不受影响: int_n/frac_n 等 RW 寄存器
    assert "`RF_WRITE(10'h2," in text and "`RF_WRITE(10'h3," in text


def test_to_logic_suffix_toggle(tmp_path):
    """append_to_logic 开关（2026-06-11 Hi1108）：默认补 _to_logic 探针网名；关时直接探基名网；
    _ls 尾缀不受此开关影响（只压 _to_logic 那一支）。

    用【独立 wb】（不复用模块级 fixture）：本测试会把 append_to_logic=False 经 Resolver 写到
    wb.logic 信号上（持久态，GUI 懒读需要），复用模块 wb 会污染后续直接读 rtl_name 的测试。"""
    path = tmp_path / "to_logic_toggle.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    wb = excel_model.load_workbook(str(path))
    # 默认（开）：pll_n1(top_out=0, 被 pll_n2 引用) 探针 = pll_n1_to_logic
    on = generator.render(generator.build(wb, generator.GenOptions(
        signals=["pll_n1"], top_output_only=False,
        probe_prefixes={"pll_n1_to_logic": "U_X"})))
    assert "`ENV_RF.U_X.pll_n1_to_logic[31:0]==" in on
    # 关：pll_n1 探针 = 基名 pll_n1（不补 _to_logic），前缀按基名 key 命中
    off = generator.render(generator.build(wb, generator.GenOptions(
        signals=["pll_n1"], top_output_only=False, append_to_logic=False,
        probe_prefixes={"pll_n1": "U_X"})))
    assert "`ENV_RF.U_X.pll_n1[31:0]==" in off
    assert "pll_n1_to_logic" not in off                    # 尾缀确实没补
    # _ls 不受开关影响：d_pfd_en_lnmode(M=ls) 关开关后仍探 d_pfd_en_lnmode_ls
    off_ls = generator.render(generator.build(wb, generator.GenOptions(
        signals=["d_pfd_en_lnmode"], append_to_logic=False,
        probe_prefixes={"d_pfd_en_lnmode_ls": "U_BT_LP_PLL_DIG",
                        "mon_active": "U_BT_LP_PLL_DIG"})))
    assert "d_pfd_en_lnmode_ls==" in off_ls


def test_probe_prefix_no_silent_miss_when_suffix_off(tmp_path):
    """探针前缀【不静默失配】(2026-06-11 Hi1108 rxiq_phase_ctrl CUVUNF 实证)：

    尾缀开关关时断言探针退成裸名，但 scan_rtl 导出 / 用户手配的前缀 key 往往是带 _to_logic 的
    【全名】。修复前：前缀查找只认裸名 rtl_base → 全名 key 对不上 → 前缀被静默丢掉 → 裸名直贴
    `ENV_RF → elaboration CUVUNF。修复后：全名也参与匹配，用户配的前缀照常生效(贴在裸名上)。"""
    path = tmp_path / "prefix_no_miss.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    wb = excel_model.load_workbook(str(path))
    # 关尾缀 + 前缀按【带 _to_logic 的全名】配（scan_rtl 风格 / 复现用户实际配法）
    off = generator.render(generator.build(wb, generator.GenOptions(
        signals=["pll_n1"], top_output_only=False, append_to_logic=False,
        probe_prefixes={"pll_n1_to_logic": "U_PLL_DIG"})))
    assert "`ENV_RF.U_PLL_DIG.pll_n1[31:0]==" in off    # 前缀贴上了（不再裸名直贴 ENV_RF）
    assert "pll_n1_to_logic" not in off                  # 断言探针仍是裸名（尾缀确实关着）
    assert "`ENV_RF.pll_n1[31:0]==" not in off           # 没有「无前缀裸名」那条 → 不会 CUVUNF


def test_cli_no_to_logic_suffix(tmp_path):
    """CLI --no-to-logic-suffix 端到端：pll_n1 探针落到基名网（不补 _to_logic）；默认仍带 _to_logic。"""
    from dreg_verify import cli
    path = tmp_path / "atl_cli.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    out_on = tmp_path / "on.sv"
    cli.main(["--excel", str(path), "--signals", "pll_n1", "--include-internal",
              "--probe-prefix", "pll_n1_to_logic=U_X", "--out", str(out_on)])
    on = out_on.read_text(encoding="utf-8")
    assert "`ENV_RF.U_X.pll_n1_to_logic[31:0]==" in on
    out_off = tmp_path / "off.sv"
    cli.main(["--excel", str(path), "--signals", "pll_n1", "--include-internal",
              "--no-to-logic-suffix", "--probe-prefix", "pll_n1=U_X", "--out", str(out_off)])
    off = out_off.read_text(encoding="utf-8")
    assert "`ENV_RF.U_X.pll_n1[31:0]==" in off
    assert "pll_n1_to_logic" not in off


def test_ref_suffix_excludes_self_reference(tmp_path):
    """自引用不算下游引用(2026-06-11 Hi1108 d_wl_rf_dpd_s2d_bias_en 实证)：输出 X 读【自己的】
    <X>_to_logic 当输入时，那是 X 的前级原始信号(regfile 透传)，不是下游消费 X 输出网的引用——
    据此补尾缀会让 X 的输出探针网名 = 它自己的输入网名(撞名探错)。故自引用不补；别的信号引用才补。"""
    import openpyxl
    from openpyxl.utils import column_index_from_string
    path = tmp_path / "selfref.xlsx"
    wb_x = openpyxl.Workbook(); ws = wb_x.active; ws.title = "logic"

    def setrow(r, d):
        for col, v in d.items():
            ws.cell(row=r, column=column_index_from_string(col), value=v)
    setrow(2, {"A": "in_A", "K": "out", "L": "expr", "M": "suffix", "N": "top_output", "R": "no"})
    # 自引用：d_self 读自己的 d_self_to_logic（前级信号）→ 不该补尾缀
    setrow(3, {"A": "d_self_to_logic", "K": "d_self", "L": "A", "M": "to_logic", "N": 0, "R": 1})
    # 真·下游引用：d_prod 被【别的】信号 d_cons 以 d_prod_to_logic 引用 → 该补尾缀
    setrow(4, {"A": "regx_to_logic", "K": "d_prod", "L": "A", "M": "to_logic", "N": 0, "R": 2})
    setrow(5, {"A": "d_prod_to_logic", "K": "d_cons", "L": "A", "M": "to_logic", "N": 0, "R": 3})
    wb_x.save(str(path))

    w = excel_model.load_workbook(str(path))
    by = {s.out_base: s for s in w.logic}
    assert by["d_self"].ref_suffix == ""              # 自引用 → 不补尾缀
    assert by["d_self"].rtl_name == "d_self"          # 探裸名（≠ 自己的输入 d_self_to_logic，不撞名）
    assert by["d_prod"].ref_suffix == "_to_logic"     # 被别的信号引用 → 真·下游引用 → 补尾缀
    assert by["d_prod"].rtl_name == "d_prod_to_logic"


def test_ref_suffix_to_mux_from_mux_page():
    """尾缀随 Excel：被 mux 页以 <名>_to_mux 引用的 logic 输出 → RTL 探针网名带 _to_mux（不是 _to_logic）。
    _apply_mux_ref_suffix 跨页回填；开关关时回到基名。"""
    sig = excel_model.LogicSignal(row=1, out_name="d_foo[2:0]", out_width=3, expr="A",
                                  suffix="", top_output="0", notes="", owner="", assert_id="1",
                                  inputs={})
    assert sig.ref_suffix == "" and sig.rtl_name == "d_foo[2:0]"        # 未引用 → 裸名
    case = excel_model.MuxCase(row=3, case_raw="1'b0", input_raw="d_foo_to_mux[2:0]",
                               input_base="d_foo", input_width=3, input_msb=2, input_lsb=0)
    grp = excel_model.MuxGroup(group_no=1, out_name="d_bar[2:0]", out_width=3,
                               ctrl_raw="d_sel_to_mux", ctrl_base="d_sel", ctrl_width=1,
                               owner="", top_output="0", cases=[case])
    excel_model._apply_mux_ref_suffix([sig], [grp])
    assert sig.ref_suffix == "_to_mux"                                  # 尾缀来自 Excel 的 mux 引用
    assert sig.rtl_name == "d_foo_to_mux[2:0]"                          # 探针网名带 _to_mux
    assert sig.to_logic_ref is False                                   # 兼容旧名：不是 _to_logic
    sig._append_to_logic = False                                       # 开关关 → 回基名
    assert sig.rtl_name == "d_foo[2:0]"


def test_probe_prefix_cli_parse():
    from dreg_verify.cli import _parse_probe_prefixes
    assert _parse_probe_prefixes(["pll_n=U_BT_LP_PLL_DIG", "x=A.B"]) == {
        "pll_n": "U_BT_LP_PLL_DIG", "x": "A.B"}
    assert _parse_probe_prefixes([]) == {}
    with pytest.raises(SystemExit):
        _parse_probe_prefixes(["no_equal_sign"])


def test_parse_probe_prefix_lines():
    """映射文本解析：支持注释/空行/多级层级路径；空路径与坏行跳过；首尾点去除。"""
    text = """
# 这是注释
pll_n=U_BT_LP_PLL_DIG

mon_active=U_BT_LP_PLL_DIG.DIG_1
empty_prefix=
no_equal_sign
spaced  =  U_X.Y.
"""
    assert generator.parse_probe_prefix_lines(text) == {
        "pll_n": "U_BT_LP_PLL_DIG",
        "mon_active": "U_BT_LP_PLL_DIG.DIG_1",
        "spaced": "U_X.Y",
    }
    assert generator.parse_probe_prefix_lines("") == {}
    assert generator.parse_probe_prefix_lines(None) == {}


def test_parse_probe_prefix_merged_format():
    """合并格式：『路径:』组头 + 其下信号名（每行一个）。"""
    text = """
# 合并格式
U_TOP1:
    SIG1
    SIG2
    SIG3
U_TOP1.TOP2:
    SIG4
    SIG5
"""
    assert generator.parse_probe_prefix_lines(text) == {
        "sig1": "U_TOP1", "sig2": "U_TOP1", "sig3": "U_TOP1",
        "sig4": "U_TOP1.TOP2", "sig5": "U_TOP1.TOP2"}


def test_parse_probe_prefix_inline_format():
    """合并格式·一行多个：组头之下信号名逗号/空格分隔（scan_rtl 默认导出）。"""
    text = """
U_TOP1:
    SIG1, SIG2, SIG3
U_TOP1.TOP2:
    SIG3, SIG4
"""
    # SIG3 在两个路径下出现 → 后者覆盖前者（dict 语义，与每行一个一致）
    assert generator.parse_probe_prefix_lines(text) == {
        "sig1": "U_TOP1", "sig2": "U_TOP1",
        "sig3": "U_TOP1.TOP2", "sig4": "U_TOP1.TOP2"}
    # 边界：尾逗号、多空格、逗号+空格混排、位宽切片整体作一个名（不被内部冒号当组头/拆分）、
    # 行尾 # 注释被忽略（不变成假信号名）
    edge = "U_X:\n    A, B,\n    C   D\n    E,F  G\n    dout[7:0]\n    H, I  # 这是注释\n"
    assert generator.parse_probe_prefix_lines(edge) == {
        "a": "U_X", "b": "U_X", "c": "U_X", "d": "U_X",
        "e": "U_X", "f": "U_X", "g": "U_X", "dout[7:0]": "U_X",
        "h": "U_X", "i": "U_X"}


def test_parse_probe_prefix_mixed_and_edges():
    """三种格式可混用；组头之前的裸行仍跳过；空路径组头=顶层无前缀（跳过其下信号）。"""
    text = """
loose_no_group
flat_sig=U_FLAT.SUB.
U_GRP:
    grouped_sig, second_sig
:
    at_top_sig
"""
    assert generator.parse_probe_prefix_lines(text) == {
        "flat_sig": "U_FLAT.SUB",      # 扁平条目去尾点，不影响后续组头
        "grouped_sig": "U_GRP",
        "second_sig": "U_GRP"}         # 同一行逗号分隔的多个名
    # loose_no_group（组头前裸行）、at_top_sig（空路径组下）均不产生映射


def test_render_probe_prefix_grouped_round_trip():
    """render_probe_prefix_grouped 输出（内联）可被 parse 无损还原；空映射 → 空串。"""
    mapping = {"pll_n": "U_BT_LP_PLL_DIG", "mon_active": "U_BT_LP_PLL_DIG",
               "xxx": "U_BT_LP_PLL_DIG.DIG_1"}
    text = generator.render_probe_prefix_grouped(mapping)
    assert "U_BT_LP_PLL_DIG:" in text and "    mon_active, pll_n" in text   # 内联、排序后逗号分隔
    assert "U_BT_LP_PLL_DIG.DIG_1:" in text and "    xxx" in text
    assert generator.parse_probe_prefix_lines(text) == mapping     # 往返无损
    assert generator.render_probe_prefix_grouped({}) == ""


def test_probe_prefix_file_cli(tmp_path):
    """CLI --probe-prefix-file 读 GUI 导出的映射文件；命令行 --probe-prefix 同名覆盖文件。"""
    from dreg_verify.cli import _parse_probe_prefixes
    f = tmp_path / "probe_prefixes.txt"
    f.write_text("# 项目通用映射\npll_n=U_BT_LP_PLL_DIG\nmon_active=U_OLD\n", encoding="utf-8")
    assert _parse_probe_prefixes([], str(f)) == {
        "pll_n": "U_BT_LP_PLL_DIG", "mon_active": "U_OLD"}
    # 命令行优先于文件
    assert _parse_probe_prefixes(["mon_active=U_NEW"], str(f))["mon_active"] == "U_NEW"
    # 文件不存在 → 明确报错
    with pytest.raises(SystemExit):
        _parse_probe_prefixes([], str(tmp_path / "missing.txt"))


def test_probe_prefix_analyze_signal(wb, resolver):
    sig = next(s for s in wb.logic if s.out_base == "pll_n")
    a = generator.analyze_signal(resolver, sig, wb=wb, probe_prefix="U_BT_LP_PLL_DIG")
    assert a["out_net"] == "`ENV_RF.U_BT_LP_PLL_DIG.pll_n[31:0]"
    assert a["status"] == "clean"


# ───────────── GUI 冒烟（离屏） ─────────────
def test_gui_cone_signal_editor(tmp_path_factory):
    """GUI 点 pll_n → 测试项编辑器显示 6 个叶子寄存器列；状态=clean；生成不抛异常。"""
    pytest.importorskip("PySide6")
    from PySide6 import QtWidgets
    QtWidgets.QApplication.instance() or QtWidgets.QApplication([])
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("gui_cone") / "synthetic_pll.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    w = gui.MainWindow()
    w.path_edit.setText(str(path)); w.on_load()
    sig = next(s for s in w.signals if s.out_base == "pll_n")
    w._load_test_items(sig)
    # 输入分组 = 6 个叶子寄存器（不是 pll_n2 切片）
    bases = {g["base"].lower() for g in w._ti_groups}
    assert bases == {"int_n", "frac_n_msb", "frac_n_lsb",
                     "d_xo_freq_sel", "en_dig_clk_div2", "en_dig_clk_div4"}
    # 状态列：cone 展开后输入全部可驱动 → clean
    row = next(r for r in range(w.table.rowCount()) if w._sig_of_row(r).out_base == "pll_n")
    assert w._analysis[row]["status"] == "clean"
    assert w._analysis[row]["cone"] is True
    # 全选导出（全部范围）能产出 pll_n 的块
    res = generator.build(w.wb, w._opts(["pll_n"]))
    assert res["summary"]["n_generated"] == 1 and res["summary"]["n_skipped"] == 0


def test_gui_cone_inputs_table_shows_excel_sources(tmp_path_factory):
    """⭐GUI 显示回归(2026-06-03 用户报告)：cone 信号的『输入信号』表字母列 = Excel 来源坐标。

    修复前：字母列显示大写基名(INT_N/EN_DIG_CLK_DIV4…)，与信号列重复、对不回 Excel；
    修复后：根行直接输入=列字母(E)；上游行叶子="行名.字母"(pll_n1.A)；头部带展开标记。
    非 cone 信号(d_pfd_en_lnmode)仍显示 A/B/C… 不变。"""
    gui, w = _pll_window(tmp_path_factory, "gui_cone_sources")
    # ── cone 信号 pll_n ──
    sig = next(s for s in w.signals if s.out_base == "pll_n")
    w._load_test_items(sig)
    ti = w.ti_inputs
    letter_by_signal = {ti.item(r, 1).text(): ti.item(r, 0).text() for r in range(ti.rowCount())}
    assert letter_by_signal["int_n[8:0]"] == "pll_n1.A"
    assert letter_by_signal["en_dig_clk_div4"] == "E"
    assert letter_by_signal["en_dig_clk_div2"] == "pll_n2.E"
    assert letter_by_signal["frac_n_lsb[15:0]"] == "pll_n1.C,pll_n1.D"
    # 大写基名(修复前的显示)不再出现在字母列
    assert "INT_N" not in letter_by_signal.values()
    assert "EN_DIG_CLK_DIV4" not in letter_by_signal.values()
    # 真值表行表头 = 信号名(2026-06-03 用户拍板：不用字母/坐标)——cone 信号即叶子寄存器名；
    # Excel 来源坐标(pll_n1.A)仍在『输入信号』表字母列与行表头 tooltip 里
    headers = [w.ti_table.verticalHeaderItem(i).text() for i in range(len(w._ti_groups))]
    assert any(h.startswith("int_n") for h in headers)        # 叶子寄存器名(小写、带位宽)
    assert not any(h.startswith("INT_N") for h in headers)    # 不是大写变量基名(修复前的显示)
    tips = [w.ti_table.verticalHeaderItem(i).toolTip() for i in range(len(w._ti_groups))]
    assert any("pll_n1.A" in t for t in tips)                 # 来源坐标在 tooltip
    # 头部带"已展开上游"标记，提示字母列为什么不是本行 A/B/C
    assert "已展开上游" in w.ti_header.text()
    # ── 非 cone 信号对照：显示不变、无展开标记 ──
    sig2 = next(s for s in w.signals if s.out_base == "d_pfd_en_lnmode")
    w._load_test_items(sig2)
    letters = {w.ti_inputs.item(r, 0).text() for r in range(w.ti_inputs.rowCount())}
    assert letters >= {"A", "B", "C"}
    assert "已展开上游" not in w.ti_header.text()


def test_gui_cone_chain_panel(tmp_path_factory):
    """⭐『展开链』面板(2026-06-03 用户功能)：cone 信号显示展开过程，非 cone/mux 隐藏。

    每个链节两行：『① 行名 = Excel 原式』+『(对齐) = 字母代入真实信号名』。"""
    gui, w = _pll_window(tmp_path_factory, "gui_chain")
    # ── cone 信号 pll_n：面板可见，链=本行→pll_n2→pll_n1 ──
    sig = next(s for s in w.signals if s.out_base == "pll_n")
    w._load_test_items(sig)
    # 用 isHidden()（widget 自身 show/hide 状态，由 _populate_chain 显式设）而非 isVisibleTo(w)：
    # 2026-06-23 重构后旧视图住『排查(旧)』分页（非默认 tab），isVisibleTo 会被外层 tab 连累；
    # isHidden() 直接验 _populate_chain 的 show/hide 逻辑，与外层 tab 无关（更精准、不弱化）。
    assert not w.ti_chain.isHidden() and not w.ti_chain_cap.isHidden()
    text = w.ti_chain.toPlainText()
    lines = text.splitlines()
    assert len(lines) == 6                                   # 3 链节 × (原式+代入) 2 行
    assert lines[0] == "① pll_n = E?{B,C,D,1'b0}:{A,B,C,D}"  # Excel 原式
    assert lines[1].lstrip().startswith("= en_dig_clk_div4?{pll_n2[30:23]")   # 代入信号名
    assert lines[2].startswith("② pll_n2 = ")
    assert lines[4].startswith("③ pll_n1 = ")
    assert "d_xo_freq_sel?{1'b0,int_n[8:0]" in lines[5]
    # 链数据与 _ti_chain 一致
    assert [c["out"] for c in w._ti_chain] == ["pll_n", "pll_n2", "pll_n1"]
    # ── 非 cone 信号：面板隐藏 ──
    sig2 = next(s for s in w.signals if s.out_base == "d_pfd_en_lnmode")
    w._load_test_items(sig2)
    assert w.ti_chain.isHidden() and w.ti_chain_cap.isHidden()
    # ── 再切回 cone 信号：面板恢复显示(状态切换无残留) ──
    w._load_test_items(sig)
    assert not w.ti_chain.isHidden()


def test_gui_probe_prefix_flows_to_sv(tmp_path_factory):
    """GUI 设置探针前缀 → 状态列 out_net 带前缀 → 生成的 .sv 探针带前缀 → 持久化结构正确。"""
    pytest.importorskip("PySide6")
    from PySide6 import QtWidgets
    QtWidgets.QApplication.instance() or QtWidgets.QApplication([])
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("gui_pfx") / "synthetic_pll.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    w = gui.MainWindow()
    w.path_edit.setText(str(path)); w.on_load()
    sig = next(s for s in w.signals if s.out_base == "pll_n")
    idx = w.signals.index(sig)
    # 模拟『设置探针前缀』(绕过输入对话框直接走数据路径)
    w._probe_prefixes[sig.out_name.lower()] = "U_BT_LP_PLL_DIG"
    w._analysis[idx] = gui.generator.analyze_signal(w._resolver, sig, wb=w.wb,
                                                    probe_prefix=w._prefix_of(sig))
    w._populate_table()
    assert w._analysis[idx]["out_net"].startswith("`ENV_RF.U_BT_LP_PLL_DIG.")
    # 前缀列显示（输出前缀 → 带"输出→"标识）
    row = next(r for r in range(w.table.rowCount()) if w._sig_of_row(r).out_base == "pll_n")
    assert w.table.item(row, gui.COL_PREFIX).text() == "输出→U_BT_LP_PLL_DIG"
    # 生成走 _opts → .sv 探针带前缀
    res = generator.build(w.wb, w._opts(["pll_n"]))
    text = generator.render(res)
    assert "`ENV_RF.U_BT_LP_PLL_DIG.pll_n[31:0]==" in text


def _pll_window(tmp_path_factory, sub):
    pytest.importorskip("PySide6")
    from PySide6 import QtWidgets
    QtWidgets.QApplication.instance() or QtWidgets.QApplication([])
    from dreg_verify import gui
    path = tmp_path_factory.mktemp(sub) / "synthetic_pll.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    w = gui.MainWindow()
    w.path_edit.setText(str(path)); w.on_load()
    return gui, w


def test_gui_search_by_input_signal(tmp_path_factory):
    """按输入信号搜索：搜 mon_active → 只列出用它做输入的输出信号(d_pfd_en_lnmode)。"""
    gui, w = _pll_window(tmp_path_factory, "gui_search")
    w.name_edit.setText("mon_active")
    visible = [w._sig_of_row(r).out_name for r in range(w.table.rowCount())
               if not w.table.isRowHidden(r)]
    assert visible == ["d_pfd_en_lnmode"]
    # 输入名正则同样支持
    w.name_edit.setText("^mon_act.*e$")
    visible = [w._sig_of_row(r).out_name for r in range(w.table.rowCount())
               if not w.table.isRowHidden(r)]
    assert visible == ["d_pfd_en_lnmode"]
    # 清空搜索 → 全部显示
    w.name_edit.setText("")
    assert all(not w.table.isRowHidden(r) for r in range(w.table.rowCount()))


def test_gui_prefix_column_shows_input_effect(tmp_path_factory):
    """输入信号配置前缀后，『探针前缀』列显示 mon_active→U_BT_LP_PLL_DIG（蓝色）。"""
    gui, w = _pll_window(tmp_path_factory, "gui_pfxcol")
    # 配置输入前缀 + 输出前缀
    w._probe_prefixes = {"mon_active": "U_BT_LP_PLL_DIG", "pll_n": "U_BT_LP_PLL_DIG"}
    w._reanalyze_all()
    # d_pfd_en_lnmode：输入 mon_active 受影响
    row = next(r for r in range(w.table.rowCount())
               if w._sig_of_row(r).out_base == "d_pfd_en_lnmode")
    cell = w.table.item(row, gui.COL_PREFIX)
    assert "mon_active→U_BT_LP_PLL_DIG" in cell.text()
    assert "U_BT_LP_PLL_DIG.mon_active" in cell.toolTip()   # 完整 force 路径
    # pll_n：输出受影响
    row2 = next(r for r in range(w.table.rowCount())
                if w._sig_of_row(r).out_base == "pll_n")
    assert "输出→U_BT_LP_PLL_DIG" in w.table.item(row2, gui.COL_PREFIX).text()
    # 不相关的信号前缀列为空
    row3 = next(r for r in range(w.table.rowCount())
                if w._sig_of_row(r).out_base == "d_logic_bt_lp_reserve")
    assert w.table.item(row3, gui.COL_PREFIX).text() == ""


def test_gui_prefix_update_keeps_checked_signals(tmp_path_factory):
    """配置探针前缀(OK)后整表重建：用户已勾选的『选』/『负向』必须原样保留。"""
    from PySide6 import QtCore
    gui, w = _pll_window(tmp_path_factory, "gui_keep_checks")
    row_pll = next(r for r in range(w.table.rowCount())
                   if w._sig_of_row(r).out_base == "pll_n")
    row_res = next(r for r in range(w.table.rowCount())
                   if w._sig_of_row(r).out_base == "d_logic_bt_lp_reserve")
    w.table.item(row_pll, gui.COL_SEL).setCheckState(QtCore.Qt.Checked)
    w.table.item(row_res, gui.COL_NEG).setCheckState(QtCore.Qt.Checked)   # 同时触发负向定制
    # 模拟『设置探针前缀』点 OK 后的数据路径
    w._probe_prefixes = {"mon_active": "U_BT_LP_PLL_DIG"}
    w._save_probe_prefixes()
    w._reanalyze_all()
    # 勾选保持（行可能因排序变化，按信号重新定位）
    row_pll2 = next(r for r in range(w.table.rowCount())
                    if w._sig_of_row(r).out_base == "pll_n")
    row_res2 = next(r for r in range(w.table.rowCount())
                    if w._sig_of_row(r).out_base == "d_logic_bt_lp_reserve")
    assert w.table.item(row_pll2, gui.COL_SEL).checkState() == QtCore.Qt.Checked
    assert w.table.item(row_res2, gui.COL_NEG).checkState() == QtCore.Qt.Checked
    # 负向定制状态也未丢
    assert "d_logic_bt_lp_reserve" in w._neg_only


def test_gui_preview_refreshes_after_prefix_change(tmp_path_factory):
    """回归(2026-06-02)：改探针前缀后，已打开的 .sv 预览必须自动按新前缀重算——
    之前预览是旧快照，导致『导出有前缀、预览没有』的困惑。"""
    from PySide6 import QtCore
    gui, w = _pll_window(tmp_path_factory, "gui_preview_sync")
    # 勾选 pll_n + 打开预览（此时无前缀）
    row = next(r for r in range(w.table.rowCount()) if w._sig_of_row(r).out_base == "pll_n")
    w.table.item(row, gui.COL_SEL).setCheckState(QtCore.Qt.Checked)
    w.on_preview()
    assert "`ENV_RF.pll_n[31:0]==" in w.preview.toPlainText()
    # 配置前缀（模拟映射编辑器点 OK 的数据路径）→ 预览自动刷新为带前缀
    w._probe_prefixes = {"pll_n": "U_BT_LP_PLL_DIG"}
    w._save_probe_prefixes()
    w._reanalyze_all()
    assert "`ENV_RF.U_BT_LP_PLL_DIG.pll_n[31:0]==" in w.preview.toPlainText()
    assert "`ENV_RF.pll_n[31:0]==" not in w.preview.toPlainText()


def test_gui_single_signal_preview_has_prefix(tmp_path_factory):
    """回归(2026-06-02)：编辑器『预览本信号.sv』漏传 probe_prefix/node → 单信号预览无前缀。"""
    gui, w = _pll_window(tmp_path_factory, "gui_ti_preview")
    w._probe_prefixes = {"pll_n": "U_BT_LP_PLL_DIG"}
    w._reanalyze_all()
    sig = next(s for s in w.signals if s.out_base == "pll_n")
    w._load_test_items(sig)
    w.on_ti_preview_signal()
    text = w.preview.toPlainText()
    assert "`ENV_RF.U_BT_LP_PLL_DIG.pll_n[31:0]==" in text
    # cone 信号的驱动也要正常（RF_WRITE 不能因漏传 node 而缺失）
    assert "`RF_WRITE(10'h1," in text


def test_gui_prefix_mapping_covers_input_wire(tmp_path_factory):
    """GUI 映射编辑器数据路径：配置 mon_active 前缀 → _reanalyze_all → 信号变 clean，force 带前缀。"""
    pytest.importorskip("PySide6")
    from PySide6 import QtWidgets
    QtWidgets.QApplication.instance() or QtWidgets.QApplication([])
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("gui_wire_pfx") / "synthetic_pll.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    w = gui.MainWindow()
    w.path_edit.setText(str(path)); w.on_load()
    sig = next(s for s in w.signals if s.out_base == "d_pfd_en_lnmode")
    idx = w.signals.index(sig)
    # 无前缀: mon_active wire 兜底 → 非 clean
    assert w._analysis[idx]["status"] == "wire-fallback"
    # 配置前缀并全表重析（与映射编辑器保存后路径一致）
    w._probe_prefixes = {"mon_active": "U_BT_LP_PLL_DIG"}
    w._reanalyze_all()
    assert w._analysis[idx]["status"] == "clean"
    nets = {i["base"]: i["net"] for i in w._analysis[idx]["inputs"]}
    assert "U_BT_LP_PLL_DIG.mon_active" in nets["mon_active"]
    # 导出路径同样生效
    text = generator.render(generator.build(w.wb, w._opts([sig.out_name])))
    assert "force `ENV_RF.U_BT_LP_PLL_DIG.mon_active=" in text


# ───────────── 自引用 × cone 组合（2026-06-02 对抗式审查确证的 high bug）─────────────
def _mk_logic(out_name, out_width, expr, inputs, aid, suffix="ls", top_output=1):
    return excel_model.LogicSignal(row=1, out_name=out_name, out_width=out_width, expr=expr,
                                   suffix=suffix, top_output=top_output, notes="", owner="T",
                                   assert_id=aid, inputs=inputs)


def test_cone_self_ref_leaf_forces_base_not_ls():
    """自引用(原文不带 _to_logic 简写) + 同信号引用内部信号(触发 cone)：

    叶子重解析(_full_binding)必须保持自引用语义(self_base 透传) → force 基名，
    绝不能 force 被断言的输出网 _ls。审查实证：不透传 self_base 时 cone 路径会让原 d2a bug 复活。"""
    s_x = _mk_logic("d_x", 1, "A&B",
                    {"A": {"raw": "d_x", "base": "d_x", "width": 1, "msb": None, "lsb": None},
                     "B": {"raw": "d_y_to_logic", "base": "d_y", "width": 1,
                           "msb": None, "lsb": None}},
                    "200", suffix="ls", top_output=1)
    s_y = _mk_logic("d_y", 1, "A",
                    {"A": {"raw": "cfg_to_logic", "base": "cfg", "width": 1,
                           "msb": None, "lsb": None}},
                    "201", suffix="to_logic", top_output=0)
    tmm = {"cfg": excel_model.TmmField("cfg", 0, 0, 0x10, "RW", "N", "R"),
           "d_x": excel_model.TmmField("d_x", 1, 1, 0x3B, "RO", None, "R")}
    wb2 = excel_model.DregWorkbook(logic=[s_x, s_y], regmap={}, tmm=tmm, sheet_names=[])
    res = R.Resolver(wb2)
    _node, leaves = cone.expand(s_x, wb2, res)
    leaf = leaves["D_X"]
    assert leaf.kind == "RO"
    assert leaf.wire == "d_x"               # force 基名(前级原始信号)
    assert leaf.wire != "d_x_ls"            # 绝不能 force 被断言的输出网
    # e2e: .sv 里 force 基名 + assert 输出网，绝无 force 输出网
    text = generator.render(generator.build(wb2, generator.GenOptions()))
    assert "force `ENV_RF.d_x=" in text
    assert "force `ENV_RF.d_x_ls" not in text
    assert "assert (`ENV_RF.d_x_ls==" in text


def test_cone_leaf_cascade_self_ref_upstream_forces_base():
    """cone 叶子原文带 _to_logic、上游行自引用(d2a 形态)：regfile 前级 → force 基名，不是 _ls。

    锁定 _full_binding 的 raw=b.raw 传递——mutation 审查确认改回 b.base 会让此叶子塌成 d_up_ls。"""
    s_up = _mk_logic("d_up", 1, "B?1'b0:A",
                     {"A": {"raw": "d_up_to_logic", "base": "d_up", "width": 1,    # 自引用
                            "msb": None, "lsb": None},
                      "B": {"raw": "cfg_to_logic", "base": "cfg", "width": 1,
                            "msb": None, "lsb": None}},
                     "210", suffix="ls", top_output=1)
    s_mid = _mk_logic("d_mid", 1, "A",
                      {"A": {"raw": "d_up_to_logic", "base": "d_up", "width": 1,
                             "msb": None, "lsb": None}},
                      "211", suffix="to_logic", top_output=0)
    s_top = _mk_logic("d_top", 1, "A",
                      {"A": {"raw": "d_mid_to_logic", "base": "d_mid", "width": 1,
                             "msb": None, "lsb": None}},
                      "212", suffix="ls", top_output=1)
    wb2 = excel_model.DregWorkbook(logic=[s_up, s_mid, s_top], regmap={}, tmm={}, sheet_names=[])
    _node, leaves = cone.expand(s_top, wb2, R.Resolver(wb2))
    leaf = leaves["D_UP"]
    assert leaf.kind == "RO" and leaf.found_in == "logic"
    assert leaf.wire == "d_up"              # 前级信号基名(regfile 导出的透传)
    assert leaf.wire != "d_up_ls"
    assert "前级" in leaf.note


def _computed_upstream_chain():
    """d_top → d_mid(内部) → d_up(top 输出、不自引用=计算网) 三层链，
    且 d_up 的源头 cfg 是 RW 寄存器（两种级联模式都能驱动到底）。"""
    s_up = _mk_logic("d_up", 1, "A",
                     {"A": {"raw": "cfg_to_logic", "base": "cfg", "width": 1,     # 不自引用
                            "msb": None, "lsb": None}},
                     "230", suffix="ls", top_output=1)
    s_mid = _mk_logic("d_mid", 1, "A",
                      {"A": {"raw": "d_up_to_logic", "base": "d_up", "width": 1,
                             "msb": None, "lsb": None}},
                      "231", suffix="to_logic", top_output=0)
    s_top = _mk_logic("d_top", 1, "A",
                      {"A": {"raw": "d_mid_to_logic", "base": "d_mid", "width": 1,
                             "msb": None, "lsb": None}},
                      "232", suffix="ls", top_output=1)
    tmm = {"cfg": excel_model.TmmField("cfg", 0, 0, 0x10, "RW", "N", "R")}
    return excel_model.DregWorkbook(logic=[s_up, s_mid, s_top], regmap={}, tmm=tmm,
                                    sheet_names=[])


def test_cone_leaf_computed_upstream_cone_mode_default():
    """⭐cone 模式(默认)：上游计算网继续展开到底 → 叶子是 cfg 寄存器 → 纯 RF_WRITE，无需前缀。

    d_top → d_mid(内部,展开) → d_up(计算网,继续展开) → cfg(RW 叶子)"""
    wb2 = _computed_upstream_chain()
    _node, leaves = cone.expand(next(s for s in wb2.logic if s.out_base == "d_top"),
                                wb2, R.Resolver(wb2))
    # 一路展开到底：唯一叶子 = cfg 寄存器
    assert set(leaves) == {"CFG"}
    assert leaves["CFG"].kind == "RW" and leaves["CFG"].address == 0x10
    # e2e: 默认选项即可生成，驱动全是 RF_WRITE，级联网/上游名完全不出现
    res = generator.build(wb2, generator.GenOptions(signals=["d_top"]))
    assert res["summary"]["n_generated"] == 1 and res["summary"]["n_skipped"] == 0
    text = generator.render(res)
    assert "`RF_WRITE(10'h10," in text
    assert "d_up" not in text          # 上游信号(基名/_ls/_to_logic)被代入消掉，完全不出现
    assert "d_mid" not in text         # 内部信号同样消掉
    assert "force" not in text         # 叶子只有 RW 寄存器 → 全 RF_WRITE 无 force
    assert "assert (`ENV_RF.d_top_ls==" in text


def test_cone_leaf_computed_upstream_force_mode():
    """force 级联网模式：上游计算网不展开 → force 字面 _to_logic 网(需前缀)。

    没配前缀 → needs-prefix(整个信号默认跳过保 elaborate)；配前缀 → 正常生成。"""
    wb2 = _computed_upstream_chain()
    s_top = next(s for s in wb2.logic if s.out_base == "d_top")
    # 没配前缀: 叶子标 needs-prefix, force 目标 = 字面 _to_logic 网
    _node, leaves = cone.expand(s_top, wb2, R.Resolver(wb2, cascade_mode="force"))
    leaf = leaves["D_UP"]
    assert leaf.kind == "RO" and leaf.found_in == "needs-prefix"
    assert leaf.wire == "d_up_to_logic"
    assert leaf.wire not in ("d_up", "d_up_ls")
    # build 默认跳过(给原因)，不会产出 CUVUNF 的 .sv
    res = generator.build(wb2, generator.GenOptions(signals=["d_top"], cascade_mode="force"))
    assert res["summary"]["n_generated"] == 0 and res["summary"]["n_skipped"] == 1
    assert "需要探针前缀" in str(res["skipped"][0][2])
    # 配前缀(scan_rtl 产物)后正常生成
    res2 = generator.build(wb2, generator.GenOptions(
        signals=["d_top"], cascade_mode="force",
        probe_prefixes={"d_up_to_logic": "U_DREG.U_SIG_LOGIC"}))
    assert res2["summary"]["n_generated"] == 1
    text = generator.render(res2)
    assert "force `ENV_RF.U_DREG.U_SIG_LOGIC.d_up_to_logic=" in text
    assert "force `ENV_RF.d_up=" not in text
    assert "d_up_ls" not in text


def test_cone_leaf_shorthand_cascade_still_forces_ls():
    """对照: cone 叶子原文『直接写上游输出名』(无 _to_logic) → 仍 force 上游 RTL 网名(_ls)。"""
    s_up = _mk_logic("d_up", 1, "A",
                     {"A": {"raw": "cfg_to_logic", "base": "cfg", "width": 1,
                            "msb": None, "lsb": None}},
                     "220", suffix="ls", top_output=1)
    s_mid = _mk_logic("d_mid", 1, "A",
                      {"A": {"raw": "d_up", "base": "d_up", "width": 1,
                             "msb": None, "lsb": None}},
                      "221", suffix="to_logic", top_output=0)
    s_top = _mk_logic("d_top", 1, "A",
                      {"A": {"raw": "d_mid_to_logic", "base": "d_mid", "width": 1,
                             "msb": None, "lsb": None}},
                      "222", suffix="ls", top_output=1)
    wb2 = excel_model.DregWorkbook(logic=[s_up, s_mid, s_top], regmap={}, tmm={}, sheet_names=[])
    _node, leaves = cone.expand(s_top, wb2, R.Resolver(wb2))
    assert leaves["D_UP"].wire == "d_up_ls"


def test_cone_cycle_falls_back_to_force_when_base_is_register():
    """cone 成环但内部信号基名是顶层真寄存器（linectrl 形态：logic 输出撞名 RO 寄存器 d61）
    → expand_signal 回退为 force 顶层基名（for_test 那招），不再抛 cone 失败。"""
    s = _mk_logic("d_band_sel", 4, "A",
                  {"A": {"raw": "d_band_sel_to_logic", "base": "d_band_sel", "width": 4,
                         "msb": 3, "lsb": 0}},
                  "2", suffix="to_mux", top_output=0)
    tmm = {"d_band_sel": excel_model.TmmField("d_band_sel", 0, 5, 0x3D, "RO", "N", "R")}
    wb2 = excel_model.DregWorkbook(logic=[s], regmap={}, tmm=tmm, sheet_names=[])
    res = R.Resolver(wb2)
    with pytest.raises(cone.ConeError):           # 裸 cone.expand 仍会成环
        cone.expand(s, wb2, res)
    notes = []
    node, bindings, expanded = generator.expand_signal(wb2, res, s, fallback_notes=notes)
    assert expanded is False                      # 没展开（走了回退）
    b = bindings["A"]
    assert b.kind == "RO" and b.resolved is True
    assert b.wire == "d_band_sel"                 # force 顶层基名 = for_test
    assert b.found_in in ("tmm", "regmap")
    assert "回退force" in (b.note or "")
    assert notes and "cone 成环" in notes[0]
    assert "d_band_sel" not in res.force_overrides  # 回退不污染 resolver 全局 override


def test_cone_cycle_no_register_still_errors():
    """cone 成环且内部信号顶层无寄存器 → 回退兜不住 → 仍报 cone 失败（不乱 force 不存在的网→CUVUNF）。"""
    s = _mk_logic("d_pure", 1, "A",
                  {"A": {"raw": "d_pure_to_logic", "base": "d_pure", "width": 1,
                         "msb": None, "lsb": None}},
                  "3", suffix="to_logic", top_output=0)
    wb2 = excel_model.DregWorkbook(logic=[s], regmap={}, tmm={}, sheet_names=[])
    with pytest.raises(cone.ConeError):
        generator.expand_signal(wb2, R.Resolver(wb2), s)


def test_gui_force_signals_flows_to_resolver_and_opts(tmp_path_factory):
    """GUI『强制 force 信号』：force_signals 流进 resolver.force_overrides 与 _opts(GenOptions)
    （= 生成/.sv 路径都按 force 顶层基名处理；状态列与产物一致）。"""
    pytest.importorskip("PySide6")
    from PySide6 import QtWidgets
    QtWidgets.QApplication.instance() or QtWidgets.QApplication([])
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("gui_force") / "synthetic_pll.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)
    w = gui.MainWindow()
    w.path_edit.setText(str(path)); w.on_load()
    assert "pll_n2" not in w._resolver.force_overrides       # 初始没有
    w._force_signals = {"pll_n2"}
    w._reanalyze_all()
    assert "pll_n2" in w._resolver.force_overrides            # 流进 resolver（GUI 状态/分析）
    assert "pll_n2" in w._opts(["pll_n"]).force_overrides     # 流进 GenOptions（build/预览/.sv）
