# -*- coding: utf-8 -*-
"""WL_RFTRX 结构 mux 支持的测试（2026-06-03 第十四轮）。

WL 表与 LPBT 表的结构差异（两轮 inspect_mux 实证）：
  ① 多控制信号：case = {B,C,D,E} 拼接（B 高位），可含 don't-care x 位
  ② 控制三来源：logic 行 / 寄存器直出(RW→RF_WRITE) / 上游 mux 输出(级联)
  ③ 数据三来源：RW(RF_WRITE) / RO 线控(force) / 级联
  ④ top_out 全 0：输出不在顶层，探针需 scan_rtl 前缀

LPBT 兼容性测试在 test_mux.py（一个都不能动）；本文件只测 WL 新行为。
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dreg_verify import excel_model, expr as E, generator, mux_gen, resolver  # noqa: E402
from fixtures import build_wl_workbook, build_workbook  # noqa: E402


@pytest.fixture(scope="module")
def wl_wb(tmp_path_factory):
    path = tmp_path_factory.mktemp("wl") / "wl_dreg.xlsx"
    build_wl_workbook(str(path))
    return excel_model.load_workbook(str(path))


@pytest.fixture(scope="module")
def lpbt_wb(tmp_path_factory):
    path = tmp_path_factory.mktemp("lpbt") / "lpbt_dreg.xlsx"
    build_workbook(str(path), with_mux=True)
    return excel_model.load_workbook(str(path))


def _grp(wb, out_base):
    g = next((g for g in wb.mux if g.out_base == out_base), None)
    assert g is not None, "找不到 mux 组 %s（现有: %s）" % (out_base, [x.out_base for x in wb.mux])
    return g


# ───────────── ① 数据层：多控制读取 ─────────────
def test_wl_read_mux_group_count(wl_wb):
    """5 个组全部读出（按 G 基名归并）。"""
    assert len(wl_wb.mux) == 5
    assert [g.group_no for g in wl_wb.mux] == [1, 2, 3, 4, 5]


def test_wl_multi_ctrl_collected(wl_wb):
    """组4（多控制）：ctrls 按 B/C 列序收集，B 在首=case 高位。"""
    g4 = _grp(wl_wb, "d_wl_rf_tx_rc_code")
    assert len(g4.ctrls) == 2
    assert g4.is_multi_ctrl
    assert g4.ctrls[0].letter == "B"
    assert g4.ctrls[0].base == "d_wl_rf_rc_code_lut_en"
    assert g4.ctrls[0].width == 1
    assert g4.ctrls[1].letter == "C"
    assert g4.ctrls[1].base == "d_wl_rf_bwctrl"
    assert g4.ctrls[1].width == 4
    # case 总宽 = 1+4 = 5（与 F 列 5'b... 对上）
    assert g4.ctrl_total_width == 5
    # 兼容别名指向第一个控制（B 列）
    assert g4.ctrl_base == "d_wl_rf_rc_code_lut_en"
    assert g4.ctrl_width == 1


def test_wl_single_ctrl_unchanged_shape(wl_wb):
    """单控制组：ctrls 长度 1，兼容别名与 ctrls[0] 一致（LPBT 形状）。"""
    g1 = _grp(wl_wb, "d_wl_rf_lna_gain")
    assert len(g1.ctrls) == 1
    assert not g1.is_multi_ctrl
    assert g1.ctrls[0].letter == "B"
    assert g1.ctrl_base == "d_wl_rf_lna_gain_ctrl_mode"
    assert g1.ctrl_width == 1
    assert g1.ctrl_total_width == 1


def test_wl_top_out_zero(wl_wb):
    """WL 输出全部 top_out=0（不在顶层）。"""
    for g in wl_wb.mux:
        assert not g.is_top, "%s 应为非顶层输出" % g.out_base


def test_wl_multi_ctrl_expr_text(wl_wb):
    """多控制组的 expr 文本显示拼接 {c1,c2}。"""
    g4 = _grp(wl_wb, "d_wl_rf_tx_rc_code")
    assert g4.expr.startswith("case({d_wl_rf_rc_code_lut_en,d_wl_rf_bwctrl})")
    # 单控制组不变
    g1 = _grp(wl_wb, "d_wl_rf_lna_gain")
    assert g1.expr.startswith("case(d_wl_rf_lna_gain_ctrl_mode)")


def test_wl_ctrl_mismatch_rows_empty(wl_wb):
    """fixture 同组各行控制列一致 → ctrl_mismatch_rows 全空。"""
    for g in wl_wb.mux:
        assert g.ctrl_mismatch_rows == []


def test_lpbt_mux_ctrls_compat(lpbt_wb):
    """LPBT 表：单控制 → ctrls 长度 1 且兼容别名逐字节不变（test_mux.py 的断言镜像）。"""
    g1 = lpbt_wb.mux[0]
    assert len(g1.ctrls) == 1
    assert g1.ctrls[0].letter == "B"
    assert g1.ctrl_base == "d_logic_bt_lp_lna_agc"
    assert g1.ctrl_width == 3
    assert g1.ctrl_total_width == 3
    assert not g1.is_multi_ctrl
    assert g1.ctrl_mismatch_rows == []


def test_muxgroup_old_constructor_compat():
    """旧构造方式（单控制三标量，不传 ctrls）仍可用——test_mux.py 的 _mini_mux_group 依赖。"""
    grp = excel_model.MuxGroup(group_no=1, out_name="d_x[3:0]", out_width=4,
                               ctrl_raw="d_ctrl_to_mux[2:0]", ctrl_base="d_ctrl",
                               ctrl_width=3, owner="", top_output=1, cases=[])
    assert len(grp.ctrls) == 1
    assert grp.ctrls[0].base == "d_ctrl"
    assert grp.ctrl_base == "d_ctrl"
    assert grp.ctrl_width == 3
    assert grp.ctrl_total_width == 3
    # 无控制信号（空 ctrl_base）→ ctrls 空，别名回退到传入值
    grp2 = excel_model.MuxGroup(group_no=2, out_name="d_y", out_width=1,
                                ctrl_raw="", ctrl_base="", ctrl_width=1,
                                owner="", top_output=1, cases=[])
    assert grp2.ctrls == []
    assert grp2.ctrl_base == ""
    assert grp2.ctrl_total_width == 1


# ───────────── ② split_case_value / concat_case_parts ─────────────
def test_split_case_value_golden():
    """黄金样本（真表 row149-154）：5'b10010 按 [1,4] 拆 → lut_en=1, bwctrl=2。"""
    v, w, dc = E.parse_case_literal("5'b10010")
    assert (v, w, dc) == (0b10010, 5, 0)
    assert E.split_case_value(v, dc, [1, 4]) == [(1, 0), (2, 0)]


def test_split_case_value_dontcare():
    """5'b0xxxx 按 [1,4] 拆 → lut_en=0(确定), bwctrl=全 don't-care。"""
    v, w, dc = E.parse_case_literal("5'b0xxxx")
    assert (v, w, dc) == (0, 5, 0b01111)
    assert E.split_case_value(v, dc, [1, 4]) == [(0, 0), (0, 0b1111)]


def test_split_case_value_single():
    """单控制退化：拆分结果 = 原值。"""
    assert E.split_case_value(0b010, 0, [3]) == [(0b010, 0)]


def test_split_concat_roundtrip():
    """split ↔ concat 互逆。"""
    widths = [1, 4, 2]
    for value, dc in [(0b1001011, 0), (0b0000000, 0b0011110), (0b1111111, 0b1000001)]:
        parts = E.split_case_value(value, dc, widths)
        triples = [(p[0], w, p[1]) for p, w in zip(parts, widths)]
        assert E.concat_case_parts(triples) == (value, sum(widths), dc)


# ───────────── ③ 数据层：组结构细节 ─────────────
def test_wl_cascade_group_cases(wl_wb):
    """组3（级联控制）：4 个 case 与 lut0..3 一一对应。"""
    g3 = _grp(wl_wb, "d_wl_rf_tx_bwctrl")
    assert len(g3.cases) == 4
    assert [c.case_raw for c in g3.cases] == ["4'b0000", "4'b0001", "4'b0010", "4'b0011"]
    assert [c.input_base for c in g3.cases] == [
        "d_wl_rf_tx_bw_lut0", "d_wl_rf_tx_bw_lut1", "d_wl_rf_tx_bw_lut2", "d_wl_rf_tx_bw_lut3"]
    # 控制信号 = 组2 的输出（级联）
    assert g3.ctrl_base == "d_wl_rf_bwctrl"
    g2 = _grp(wl_wb, "d_wl_rf_bwctrl")
    assert g3.ctrl_base == g2.out_base


def test_wl_multi_ctrl_group_cases(wl_wb):
    """组4（多控制）：5 个 case（1 个 don't-care local + 4 个 lut）。"""
    g4 = _grp(wl_wb, "d_wl_rf_tx_rc_code")
    assert len(g4.cases) == 5
    assert g4.cases[0].case_raw == "5'b0xxxx"
    assert g4.cases[0].input_base == "d_wl_rf_tx_rc_local"
    assert g4.cases[1].case_raw == "5'b10000"
    assert g4.cases[4].input_base == "d_wl_rf_tx_rc_lut3"


def test_wl_logic_ctrl_group(wl_wb):
    """组5（logic 行控制 = LPBT 形态）：控制信号在 logic 页存在。"""
    g5 = _grp(wl_wb, "d_wl_rf_dpd_path")
    assert g5.ctrl_base == "d_wl_rf_fb_en"
    logic_bases = {s.out_base.lower() for s in wl_wb.logic}
    assert g5.ctrl_base.lower() in logic_bases


# ───────────── ④ 解析层：控制/数据三来源 ─────────────
@pytest.fixture(scope="module")
def wl_resolver(wl_wb):
    return resolver.Resolver(wl_wb)


def test_wl_ctrl_reg_direct(wl_wb, wl_resolver):
    """组1：控制信号是寄存器直出（RW）→ source='reg'，RF_WRITE 驱动，无 issue。"""
    g1 = _grp(wl_wb, "d_wl_rf_lna_gain")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g1)
    assert exp["issues"] == []
    assert len(exp["ctrl_drivers"]) == 1
    d = exp["ctrl_drivers"][0]
    assert d["source"] == "reg"
    assert d["key"] == "c:reg"                       # 单控制 idx=0 保持 'c:' 前缀
    assert d["binding"].kind == "RW" and d["binding"].address == 0x50
    # used_vars: 控制在前数据在后
    assert exp["used_vars"][0] == "c:reg"
    assert exp["data_keys"] == ["d:0", "d:1"]


def test_wl_data_ro_force_allowed(wl_wb, wl_resolver):
    """组1：线控数据（RO→force）放行，local 数据（RW→RF_WRITE）；旧版"必须RW"硬门已开。"""
    g1 = _grp(wl_wb, "d_wl_rf_lna_gain")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g1)
    assert exp["issues"] == []
    b_line = exp["bindings"]["d:0"]                  # linectrl（RO 线控）
    b_local = exp["bindings"]["d:1"]                 # local（RW）
    assert b_line.kind == "RO"
    assert b_local.kind == "RW" and b_local.address == 0x51


def test_wl_ctrl_mux_cascade_recipe(wl_wb, wl_resolver):
    """组3：控制信号 = 组2 的输出 → source='mux'，上游配方（优先 RW 载体 + 上游控制驱动）。"""
    g3 = _grp(wl_wb, "d_wl_rf_tx_bwctrl")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g3)
    assert exp["issues"] == []
    d = exp["ctrl_drivers"][0]
    assert d["source"] == "mux"
    assert d["upstream"].out_base == "d_wl_rf_bwctrl"
    recipe = d["recipe"]
    # 载体优先 RW = 组2 的 local 数据（case 1'b1, h53），不是 RO 线控
    assert recipe["carrier_ci"] == 1
    assert recipe["carrier_key"] == "m2.d:1"
    assert recipe["bindings"]["m2.d:1"].kind == "RW"
    # 上游控制 = bwctrl_mode 寄存器直出；载体 case 1'b1 → 上游控制驱到 1
    assert len(recipe["ctrl_drivers"]) == 1
    assert recipe["ctrl_drivers"][0]["source"] == "reg"
    assert recipe["ctrl_drivers"][0]["key"] == "m2.c:reg"
    assert recipe["ctrl_values"] == [1]


def test_wl_multi_ctrl_drivers(wl_wb, wl_resolver):
    """组4：两个控制 → 2 个 driver（B=寄存器直出，C=mux 级联）。"""
    g4 = _grp(wl_wb, "d_wl_rf_tx_rc_code")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g4)
    assert exp["issues"] == []
    assert len(exp["ctrl_drivers"]) == 2
    assert exp["ctrl_drivers"][0]["source"] == "reg"           # B: lut_en
    assert exp["ctrl_drivers"][0]["key"] == "c:reg"
    assert exp["ctrl_drivers"][1]["source"] == "mux"           # C: bwctrl（级联到组2）
    assert exp["ctrl_drivers"][1]["upstream"].group_no == 2


def test_wl_logic_ctrl_lpbt_path(wl_wb, wl_resolver):
    """组5：logic 行控制（LPBT 形态）→ source='logic' + 兼容别名 line/local/ctrl_sig 都在。"""
    g5 = _grp(wl_wb, "d_wl_rf_dpd_path")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g5)
    assert exp["issues"] == []
    d = exp["ctrl_drivers"][0]
    assert d["source"] == "logic"
    # LPBT 兼容别名（make_mux_vectors 现有路径靠它们）
    assert exp["ctrl_sig"] is not None and exp["ctrl_sig"].out_base == "d_wl_rf_fb_en"
    assert exp["line"] is not None and exp["line"]["kind"] == "RO"     # fb_line
    assert exp["local"] is not None and exp["local"]["kind"] == "RW"   # fb_local
    assert exp["line"]["key"] == "c:B" and exp["local"]["key"] == "c:C"


def test_wl_resolver_recognizes_mux_output(wl_wb):
    """resolver：logic 输入 = mux 组输出（mux→logic 级联）→ found_in='mux-output'（不再 wire 兜底）。"""
    r = resolver.Resolver(wl_wb)
    sig = next(s for s in wl_wb.logic if s.out_base == "d_wl_rf_lna_gain_dly")
    b = r.resolve_signal_inputs(sig)["A"]
    assert b.found_in == "mux-output"
    assert b.kind == "RO"
    assert b.wire == "d_wl_rf_lna_gain_to_logic"     # force 字面衔接网（不是基名）
    assert "mux" in b.note


def test_wl_resolver_mux_output_with_prefix(wl_wb):
    """配置探针前缀后：mux 输出衔接网 → prefixed-wire（确认存在，可生成）。"""
    r = resolver.Resolver(wl_wb,
                          wire_prefixes={"d_wl_rf_lna_gain_to_logic": "U_DREG.U_MUX"})
    sig = next(s for s in wl_wb.logic if s.out_base == "d_wl_rf_lna_gain_dly")
    b = r.resolve_signal_inputs(sig)["A"]
    assert b.found_in == "prefixed-wire"
    assert b.wire == "U_DREG.U_MUX.d_wl_rf_lna_gain_to_logic"


def test_wl_cascade_cycle_detection(wl_wb, wl_resolver):
    """mux 级联成环 → issue（不递归爆栈）。"""
    import copy
    # 人造环：组2 的控制改成组3 的输出（2→3→2）。
    # 控制位宽要与组2 自己的 case 位宽(1bit)一致——否则位宽不一致的 issue 先挡住，到不了环检测。
    wb2 = copy.copy(wl_wb)
    g2 = copy.deepcopy(_grp(wl_wb, "d_wl_rf_bwctrl"))
    g3 = copy.deepcopy(_grp(wl_wb, "d_wl_rf_tx_bwctrl"))
    g2.ctrls = [excel_model.MuxCtrl("B", "d_wl_rf_tx_bwctrl_to_mux",
                                    "d_wl_rf_tx_bwctrl", 1)]
    wb2.mux = [g2, g3]
    r = resolver.Resolver(wb2)
    exp = mux_gen.expand_mux_group(wb2, r, g3)
    assert any("成环" in i for i in exp["issues"]), exp["issues"]


def test_wl_cascade_width_mismatch_blocked(wl_wb):
    """上游载体 case 位宽与上游控制拼接总宽不一致 → issue（不静默生成错位 case）。"""
    import copy
    wb2 = copy.copy(wl_wb)
    g2 = copy.deepcopy(_grp(wl_wb, "d_wl_rf_bwctrl"))
    g3 = copy.deepcopy(_grp(wl_wb, "d_wl_rf_tx_bwctrl"))
    # 组2 的控制声明成 5bit，但它的 case 是 1'b0/1'b1 → 位宽不一致
    g2.ctrls = [excel_model.MuxCtrl("B", "d_wl_rf_tx_bwctrl_to_mux[4:0]",
                                    "d_wl_rf_tx_bwctrl", 5)]
    wb2.mux = [g2, g3]
    r = resolver.Resolver(wb2)
    exp = mux_gen.expand_mux_group(wb2, r, g3)
    assert any("不一致" in i for i in exp["issues"]), exp["issues"]


# ───────────── ⑤ 生成层：通用向量生成 ─────────────
def test_wl_vectors_reg_ctrl(wl_wb, wl_resolver):
    """组1（寄存器直出控制 + RO/RW 混合数据）：每 case 一条，c:reg 写 case 值。"""
    g1 = _grp(wl_wb, "d_wl_rf_lna_gain")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g1)
    vecs, meta = mux_gen.make_mux_vectors(g1, exp, mode="min")
    assert len(vecs) == 2                                # 2 个 case
    assert meta["scan_path"] == "direct"
    assert not meta["value_collision"]
    # T0: case 1'b0 → ctrl_mode=0 选线控；数据互异值 [1,2]（3bit 窄寄存器从 1 递增）
    t0 = vecs[0]
    assert t0.assignments["c:reg"] == 0
    assert t0.assignments["d:0"] == 1 and t0.assignments["d:1"] == 2
    assert t0.exp_value == 1                             # 选中线控 → 期望=线控 force 的值
    # T1: case 1'b1 → ctrl_mode=1 选 local
    t1 = vecs[1]
    assert t1.assignments["c:reg"] == 1
    assert t1.exp_value == 2                             # 选中 local → 期望=local 写的值


def test_wl_vectors_cascade_ctrl(wl_wb, wl_resolver):
    """组3（mux 级联控制）：每条向量含上游配方（m2.c:reg=1 选 local + m2.d:1=case值）。"""
    g3 = _grp(wl_wb, "d_wl_rf_tx_bwctrl")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g3)
    vecs, meta = mux_gen.make_mux_vectors(g3, exp, mode="min")
    assert len(vecs) == 4                                # 4 个 case（lut0..3）
    # 每条向量：上游 bwctrl_mode=1（选 local）、上游 bwctrl_local=本 case 的值
    for ci, v in enumerate(vecs):
        assert v.assignments["m2.c:reg"] == 1, "上游模式必须选 local 路径"
        assert v.assignments["m2.d:1"] == ci, "上游载体寄存器写 case 值 %d" % ci
    # 数据互异值 0xA 递减（5bit lut 寄存器），期望=被选中 lut 的值
    assert vecs[0].exp_value == 0xA and vecs[2].exp_value == 0x8
    # 上游配方键也在 used_vars 里（compute_drives 要写它们）
    assert "m2.c:reg" in exp["used_vars"] and "m2.d:1" in exp["used_vars"]


def test_wl_vectors_multi_ctrl_golden(wl_wb, wl_resolver):
    """组4（多控制拼接，黄金样本）：5'b10010 拆成 lut_en=1 / bwctrl=2 分别驱动。"""
    g4 = _grp(wl_wb, "d_wl_rf_tx_rc_code")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g4)
    vecs, meta = mux_gen.make_mux_vectors(g4, exp, mode="min")
    assert len(vecs) == 5                                # 5 个 case（x 位取 0）
    assert meta["multi_ctrl"]
    assert meta["ctrl_sources"] == ["reg", "mux"]
    # T0: case 5'b0xxxx（x 取 0）→ lut_en=0，bwctrl=0 → 选 local
    t0 = vecs[0]
    assert t0.assignments["c:reg"] == 0                  # B 列控制 lut_en（高位段）
    assert t0.assignments["m2.d:1"] == 0                 # C 列控制 bwctrl（经上游 mux）
    assert t0.exp_value == 0xA                           # local 的互异值
    # T3: case 5'b10010（黄金样本）→ lut_en=1，bwctrl=2 → 选 lut2
    t3 = vecs[3]
    assert t3.assignments["c:reg"] == 1
    assert t3.assignments["m2.d:1"] == 2
    assert t3.exp_value == 0x7                           # lut2 的互异值（0xA,0x9,0x8,0x7,0x6 的第4个）


def test_wl_vectors_logic_ctrl_lpbt_path(wl_wb, wl_resolver):
    """组5（logic 行控制）：走 LPBT line/local 双路径生成器（向量结构与 LPBT 相同）。"""
    g5 = _grp(wl_wb, "d_wl_rf_dpd_path")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g5)
    vecs, meta = mux_gen.make_mux_vectors(g5, exp, mode="min")
    # LPBT 形态：2 case line 扫 + 1 条 local 抽测 = 3
    assert len(vecs) == 3
    assert meta["scan_path"] == "line"
    assert meta["other_path_scan"] == "probe"


def test_wl_coverage_levels(wl_wb, wl_resolver):
    """通用形态覆盖度：精简 < 全面 < 穷举。组4 含级联控制(C=bwctrl→组2，mode=0 RO/mode=1 RW
    两分支) → 全面/穷举补 item④ 备用载体(mode=0)轮：全面 DC=0、穷举 DC 全展（第二十二轮）。"""
    g4 = _grp(wl_wb, "d_wl_rf_tx_rc_code")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g4)
    res = {m: mux_gen.make_mux_vectors(g4, exp, mode=m) for m in ("min", "max", "exhaustive")}
    n = {m: len(v) for m, (v, _meta) in res.items()}
    # min=5（每 case 1 条）
    # max=20（case0 的 4 x 位展开成 16 +4）+5（反码轮）+5（级联 alt 轮，每 case 1 条 mode=0）=30
    # exhaustive=20+5+20（alt 轮 DC 全展：case0 16 +4）=45
    assert n["min"] == 5
    assert n["max"] == 20 + 5 + 5
    assert n["exhaustive"] == 20 + 5 + 20
    assert n["min"] < n["max"] < n["exhaustive"]            # 三档严格递增（级联组）
    # 级联 alt 轮（mode=0 备用载体）只在全面/穷举出现
    assert res["min"][1].get("cascade_alt") is None
    assert res["max"][1].get("cascade_alt") is True
    assert len([v for v in res["max"][0] if "alt载体" in (v.note or "")]) == 5
    assert len([v for v in res["exhaustive"][0] if "alt载体" in (v.note or "")]) == 20
    # 反码轮的期望值 = 反码互异值（与 alt 轮 note 不同，计数不相串）
    vecs_max = res["max"][0]
    inv = [v for v in vecs_max if "inverted data" in (v.note or "")]
    assert len(inv) == 5
    base_vals, _ = mux_gen.alloc_distinct_values(g4, exp["bindings"], exp["data_keys"])
    inv_vals, _ = mux_gen.alloc_inverted_values(g4, base_vals, exp["bindings"], exp["data_keys"])
    assert [v.exp_value for v in inv] == inv_vals


def test_wl_cascade_two_branch(wl_wb, wl_resolver):
    """⭐ item④ 级联两分支（第二十二轮）：组3 控制=组2 输出(case0=RO 线控/case1=RW local)。
    精简只走主载体(mode=1 写 local)；全面/穷举补 alt 轮(mode=0 force RO linectrl)，对齐 designer
    16 拍签核表（mode=1 半张 + mode=0 半张）。"""
    g3 = _grp(wl_wb, "d_wl_rf_tx_bwctrl")
    exp = mux_gen.expand_mux_group(wl_wb, wl_resolver, g3)
    assert exp["issues"] == []
    rec = exp["ctrl_drivers"][0]["recipe"]
    # 双载体：主=RW local(ci1)、alt=RO 线控(ci0)
    assert rec["carrier_ci"] == 1 and rec["carrier_alt_ci"] == 0
    assert rec["alt_skipped_reason"] == ""
    assert exp["bindings"][rec["carrier_key"]].kind == "RW"
    assert exp["bindings"][rec["carrier_alt_key"]].kind == "RO"
    # 精简：4 条全主载体(mode=1 选 local)，无 alt
    vmin, mmin = mux_gen.make_mux_vectors(g3, exp, mode="min")
    assert len(vmin) == 4
    assert all(v.assignments["m2.c:reg"] == 1 for v in vmin)
    assert mmin.get("cascade_alt") is None
    # 全面：8 主(mode=1) + 4 alt(mode=0)
    vmax, mmax = mux_gen.make_mux_vectors(g3, exp, mode="max")
    main = [v for v in vmax if v.assignments.get("m2.c:reg") == 1]
    alt = [v for v in vmax if v.assignments.get("m2.c:reg") == 0]
    assert len(main) == 8 and len(alt) == 4
    assert mmax.get("cascade_alt") is True
    # alt 向量：上游模式驱到 0（选 RO 线控分支）、写【备用载体】(linectrl) 而非主载体(local)
    for v in alt:
        assert rec["carrier_alt_key"] in v.assignments      # 写了 RO 线控载体（渲染层 force）
        assert rec["carrier_key"] not in v.assignments      # 没写主 RW 载体
        assert "alt载体" in (v.note or "")
    # alt 选值覆盖 4 个 case（每 case 各一条，slice_lsb=0 → 载体=case 值原样）
    assert sorted(v.assignments[rec["carrier_alt_key"]] for v in alt) == [0, 1, 2, 3]
    # ⭐ 渲染级回归（评审 wf_657e87b6 major）：min 主轮【不得】驱动备用载体(RO linectrl)——
    # 旧 compute_drives 对未驱动的 used_var 凭空补 force=0，污染产物且破坏 min 字节一致。
    only3 = dict(probe_prefixes=WL_PREFIXES, signals={"d_wl_rf_tx_bwctrl"})
    txt_min = generator.render(generator.build(wl_wb, generator.GenOptions(mux_mode="min", **only3)))
    assert "linectrl_bwctrl" not in txt_min, "min 主轮不应出现备用载体 linectrl_bwctrl 的驱动"
    txt_max = generator.render(generator.build(wl_wb, generator.GenOptions(mux_mode="max", **only3)))
    assert "force `ENV_RF.d_wl_rf_linectrl_bwctrl" in txt_max     # 全面 alt 轮才 force 它


def test_coverage_decouple_genoptions():
    """⭐ 第二十二轮：GenOptions 覆盖档位 logic/mux 解耦——各自覆盖对应侧；未传则回退旧 (mode,exhaustive)。"""
    G = generator
    # 未解耦：两侧都回退到 (mode, exhaustive)，与旧行为逐字节一致
    o = G.GenOptions(mode="max", exhaustive=False)
    assert o.logic_vec_params() == ("max", False)
    assert o.mux_cov_mode() == "max"
    o2 = G.GenOptions(mode="min", exhaustive=True)
    assert o2.logic_vec_params() == ("min", True)
    assert o2.mux_cov_mode() == "exhaustive"
    # 解耦：各侧只跟自己的档，互不影响
    d = G.GenOptions(mode="min", exhaustive=False, logic_mode="min", mux_mode="exhaustive")
    assert d.logic_vec_params() == ("min", False)
    assert d.mux_cov_mode() == "exhaustive"
    d2 = G.GenOptions(mode="max", exhaustive=True, logic_mode="max", mux_mode="min")
    assert d2.logic_vec_params() == ("max", False)   # logic_mode 覆盖，忽略 exhaustive
    assert d2.mux_cov_mode() == "min"
    # 非法值忽略 → 回退到 (mode, exhaustive)，不抛错
    d3 = G.GenOptions(mode="min", logic_mode="bogus", mux_mode="")
    assert d3.logic_vec_params() == ("min", False)
    assert d3.mux_cov_mode() == "min"


def test_coverage_decouple_build_independent(lpbt_wb):
    """⭐ build() 端到端：logic 侧与 mux 侧覆盖档互不串档（改一侧另一侧用例数不动）。"""
    G = generator

    def counts(**kw):
        res = G.build(lpbt_wb, G.GenOptions(**kw))
        lc = sum(st["n_vectors"] for _l, st in res["blocks"] if not st.get("is_mux"))
        mc = sum(st["n_vectors"] for _l, st in res["blocks"] if st.get("is_mux"))
        return lc, mc

    l_min, m_min = counts(logic_mode="min", mux_mode="min")
    l_max, m_max = counts(logic_mode="max", mux_mode="max")
    assert l_min < l_max, "logic 全面应比精简多用例: %d/%d" % (l_min, l_max)   # logic 旋钮生效
    assert m_min < m_max, "mux 全面应比精简多用例: %d/%d" % (m_min, m_max)     # mux 旋钮生效
    # logic 精简 + mux 全面 → logic 跟 logic 档(=l_min)、mux 跟 mux 档(=m_max)，互不串
    assert counts(logic_mode="min", mux_mode="max") == (l_min, m_max)
    # logic 全面 + mux 精简 → 反向同理
    assert counts(logic_mode="max", mux_mode="min") == (l_max, m_min)
    # 未传新参 = 回退旧 (mode,exhaustive)：默认(min)==两侧精简；--mode max==两侧全面
    assert counts() == (l_min, m_min)
    assert counts(mode="max") == (l_max, m_max)


def test_sig_cov_genoptions():
    """⭐ 单点覆盖度：GenOptions.sig_cov 按信号名压过全局；name 缺省/未命中=纯走全局(逐字节不变)。"""
    G = generator
    o = G.GenOptions(mode="min", exhaustive=False,
                     sig_cov={"d_foo": "exhaustive", "D_BAR": "max"})
    # 命中单点：logic 与 mux 都以单点档为准；键大小写不敏感
    assert o.logic_vec_params("d_foo") == ("max", True)      # exhaustive→(max,True)
    assert o.mux_cov_mode("d_foo") == "exhaustive"
    assert o.logic_vec_params("D_BAR") == ("max", False)
    assert o.mux_cov_mode("d_bar") == "max"
    # 未命中 / name 缺省：回退全局(=旧行为，与不传 sig_cov 逐字节一致)
    assert o.logic_vec_params("d_other") == ("min", False)
    assert o.mux_cov_mode("d_other") == "min"
    assert o.logic_vec_params() == ("min", False)
    assert o.mux_cov_mode() == "min"
    # 单点压过【解耦的】全局档；未命中信号仍跟随解耦全局
    d = G.GenOptions(logic_mode="min", mux_mode="min", sig_cov={"d_x": "max"})
    assert d.logic_vec_params("d_x") == ("max", False)
    assert d.mux_cov_mode("d_x") == "max"
    assert d.logic_vec_params("d_y") == ("min", False)
    assert d.mux_cov_mode("d_y") == "min"
    # 非法档位忽略 → 回退全局，不抛错
    bad = G.GenOptions(mode="min", sig_cov={"d_z": "bogus", "d_w": ""})
    assert bad.logic_vec_params("d_z") == ("min", False)
    assert bad.mux_cov_mode("d_w") == "min"


def test_sig_cov_build_per_signal(lpbt_wb):
    """⭐ build() 端到端：单点覆盖度只改【指定信号】用例数，其余信号跟随全局不动；logic+mux 各举一例。"""
    G = generator

    def counts(**kw):
        res = G.build(lpbt_wb, G.GenOptions(**kw))
        return {st["out_name"]: (st["n_vectors"], bool(st.get("is_mux")))
                for _l, st in res["blocks"]}

    cmin = {n: v for n, (v, _m) in counts(mode="min").items()}
    full = counts(mode="max")
    cmax = {n: v for n, (v, _m) in full.items()}

    def pick(is_mux):
        for n, (v, m) in full.items():
            if m == is_mux and cmax[n] > cmin.get(n, 0):
                return n
        return None

    lname, mname = pick(False), pick(True)
    assert lname and mname, "需要至少一个 logic + 一个 mux 信号在全面档用例更多"

    # 全局精简 + 只给这两个信号设单点全面
    res = G.build(lpbt_wb, G.GenOptions(mode="min", sig_cov={lname: "max", mname: "max"}))
    c = {st["out_name"]: st["n_vectors"] for _l, st in res["blocks"]}
    assert c[lname] == cmax[lname], "单点 logic 信号应取全面用例数"
    assert c[mname] == cmax[mname], "单点 mux 信号应取全面用例数"
    others = [n for n in cmin if n not in (lname, mname)]
    assert others, "需要旁证信号"
    assert all(c[n] == cmin[n] for n in others), "未设单点的信号应仍跟随全局精简"


def test_sig_cov_report_per_signal(lpbt_wb):
    """⭐ report() 与 build() 双轨一致：报告里单点信号的用例数也按单点档（HTML 真值表不漏档）。"""
    G = generator

    def rep_counts(**kw):
        rep = G.report(lpbt_wb, G.GenOptions(**kw))
        return {r["signal"]: r["n_tests"] for r in rep["summary"]}

    rmin = rep_counts(mode="min")
    rmax = rep_counts(mode="max")
    lname = next((n for n, v in rmax.items() if v > rmin.get(n, 0)), None)
    assert lname, "需要一个全面档用例更多的信号"
    r = rep_counts(mode="min", sig_cov={lname: "max"})
    assert r[lname] == rmax[lname]
    other = next(n for n in rmin if n != lname)
    assert r[other] == rmin[other]


def test_wl_key_role():
    """键角色判断（generator/gui/report 共用）。"""
    assert mux_gen.key_role("c:A") == "ctrl"
    assert mux_gen.key_role("c:reg") == "ctrl"
    assert mux_gen.key_role("c1:reg") == "ctrl"
    assert mux_gen.key_role("c2:A") == "ctrl"
    assert mux_gen.key_role("d:0") == "data"
    assert mux_gen.key_role("m2.c:reg") == "upstream"
    assert mux_gen.key_role("m57.d:1") == "upstream"


# ───────────── ⑥ build/report 端到端 ─────────────
WL_PREFIX = "U_WL_DREG.U_RF_MUX"
WL_PREFIXES = {
    "d_wl_rf_lna_gain": WL_PREFIX, "d_wl_rf_bwctrl": WL_PREFIX,
    "d_wl_rf_tx_bwctrl": WL_PREFIX, "d_wl_rf_tx_rc_code": WL_PREFIX,
    "d_wl_rf_dpd_path": WL_PREFIX,
    # mux→logic 级联衔接网（logic 行 lna_gain_dly 的输入）
    "d_wl_rf_lna_gain_to_logic": WL_PREFIX,
}


def test_wl_build_generates_without_prefix_with_warning(wl_wb):
    """没配探针前缀：top_out=0 的 mux 组【照常生成】裸名探针 + 警告（2026-06-03 用户拍板：
    工具不替用户假设 top_out=0=埋深，探得到就过、真 CUVUNF 再配前缀）。"""
    res = generator.build(wl_wb, generator.GenOptions())
    # 5 个组全部生成（不再因 top_out=0 跳过）
    assert res["summary"]["n_mux_generated"] == 5
    assert res["summary"]["n_mux_warnings"] == 5
    # 警告文案提示 scan_rtl（裸名探针，CUVUNF 时配前缀）
    assert all("scan_rtl" in why for _n, _a, why in res["mux_warnings"])
    # .sv 里是裸名探针（无前缀），块顶有 ⚠ 警告注释
    text = "\n".join("\n".join(lines) for lines, _ in res["blocks"])
    assert "assert_mux1_T0:" in text
    assert "`ENV_RF.d_wl_rf_lna_gain[2:0]==" in text       # 裸名，无层级前缀
    assert "// ⚠" in text                                   # 块顶警告注释


def test_wl_build_prefix_silences_warning(wl_wb):
    """配了探针前缀的组：无警告（探针带前缀）。"""
    res = generator.build(wl_wb, generator.GenOptions(probe_prefixes=WL_PREFIXES))
    assert res["summary"]["n_mux_generated"] == 5
    assert res["summary"]["n_mux_warnings"] == 0


def test_wl_build_with_prefixes(wl_wb):
    """配置探针前缀后全部 5 个组生成；.sv 内容覆盖三来源驱动 + 前缀探针。"""
    opts = generator.GenOptions(probe_prefixes=WL_PREFIXES)
    res = generator.build(wl_wb, opts)
    assert res["summary"]["n_mux_groups"] == 5
    assert res["summary"]["n_mux_generated"] == 5, res["skipped"]
    text = "\n".join("\n".join(lines) for lines, _ in res["blocks"])
    # ① 寄存器直出控制：RF_WRITE 模式寄存器（h50）
    assert "`RF_WRITE(10'h50," in text
    # ② RO 线控数据：force（无前缀——线控是顶层 RO 寄存器）
    assert "force `ENV_RF.d_wl_rf_linectrl_lna_gain[2:0]=" in text
    # ③ 输出探针带层级前缀（top_out=0）
    assert "`ENV_RF.%s.d_wl_rf_lna_gain[2:0]==" % WL_PREFIX in text
    assert "`ENV_RF.%s.d_wl_rf_tx_bwctrl[4:0]==" % WL_PREFIX in text
    # ④ mux 级联：上游载体寄存器（bwctrl_local h53）被 RF_WRITE
    assert "`RF_WRITE(10'h53," in text
    # ⑤ assert 标签
    assert "assert_mux1_T0:" in text and "assert_mux4_T0:" in text


def test_wl_build_include_risky_generates_without_prefix(wl_wb):
    """--include-risky：没前缀也强制生成（探针不带前缀，用户自担 CUVUNF 风险）。"""
    res = generator.build(wl_wb, generator.GenOptions(include_risky=True))
    assert res["summary"]["n_mux_generated"] == 5


def test_wl_report_sync_with_build(wl_wb):
    """report 与 build 双轨同步：同口径生成；多控制键正确分类。"""
    # 无前缀：top_out=0 组照常生成（error 列空），警告进 warning 列
    rep0 = generator.report(wl_wb, generator.GenOptions())
    mux_rows0 = [r for r in rep0["summary"] if r["type"] == "mux"]
    assert len(mux_rows0) == 5
    assert all(r["error"] == "" for r in mux_rows0)                 # 不再跳过
    assert all("scan_rtl" in r.get("warning", "") for r in mux_rows0)  # 警告提示 scan_rtl
    assert all(r["n_tests"] > 0 for r in mux_rows0)                 # 真生成了向量
    # 有前缀：全部生成，tables 有 5 个 mux 表
    opts = generator.GenOptions(probe_prefixes=WL_PREFIXES)
    rep = generator.report(wl_wb, opts)
    mux_rows = [r for r in rep["summary"] if r["type"] == "mux"]
    assert all(r["error"] == "" for r in mux_rows), [r["error"] for r in mux_rows]
    mux_tables = [t for t in rep["tables"] if t.get("kind") == "mux"]
    assert len(mux_tables) == 5
    # 组3 的输入行包含上游配方键（标记为"上游mux"角色）
    g3_table = next(t for t in mux_tables if "tx_bwctrl" in t["signal"])
    letters = [r["letters"] for r in g3_table["inputs"]]
    assert any("上游mux" in s for s in letters), letters
    # build 与 report 的生成组数一致
    res = generator.build(wl_wb, opts)
    assert res["summary"]["n_mux_generated"] == len([r for r in mux_rows if not r["error"]])


def test_wl_analyze_mux_group_status(wl_wb, wl_resolver):
    """analyze_mux_group（GUI 状态源）：top_out=0 无前缀 → bare-probe【非阻断】（照常生成裸名）；
    有前缀 → clean。"""
    g1 = _grp(wl_wb, "d_wl_rf_lna_gain")
    a0 = generator.analyze_mux_group(wl_resolver, wl_wb, g1)
    assert a0["status"] == "bare-probe"
    assert "scan_rtl" in a0["error"]
    assert a0["blocking"] is False                  # 软提示，不挡生成（真值表照常）
    a1 = generator.analyze_mux_group(wl_resolver, wl_wb, g1, probe_prefix=WL_PREFIX)
    assert a1["status"] == "clean"


def test_wl_analyze_force_net_blocker_is_blocking(tmp_path):
    """对照：force 子模块衔接网缺前缀（force 级联模式）→ needs-prefix 且 blocking=True（真阻断）。"""
    wb = _build_mux_wb(
        str(tmp_path / "blk.xlsx"),
        tmm_fields=[("D0", "h11", "d_d0[2:0]", "2:0", "Y", "RO"),
                    ("D1", "h12", "d_d1[2:0]", "2:0", "N", "RW")],
        mux_rows=[
            # 控制是别的 mux 的输出，force 级联模式 → force 衔接网需前缀
            _mrow("d_d0_to_mux[2:0]", "d_up_mux_to_mux", "1'b0", "d_gg[2:0]", 1),
            _mrow("d_d1_to_mux[2:0]", "d_up_mux_to_mux", "1'b1", "d_gg[2:0]", 1),
            _mrow("d_x0_to_mux[2:0]", "d_sel_to_mux", "1'b0", "d_up_mux[2:0]", 2),
            _mrow("d_x1_to_mux[2:0]", "d_sel_to_mux", "1'b1", "d_up_mux[2:0]", 2),
        ],
    )
    r = resolver.Resolver(wb, cascade_mode="force")
    a = generator.analyze_mux_group(r, wb, _grp(wb, "d_gg"))
    # 控制衔接网 d_up_mux_to_mux 在子模块内，force 模式没前缀 → 真阻断
    assert a["status"] in ("needs-prefix", "unresolved")
    if a["status"] == "needs-prefix":
        assert a["blocking"] is True


def test_narrow_field_mux_uses_marker_method(tmp_path):
    """字段太窄(3 个 1-bit 数据，互异值装不下)：不再判假绿跳过，改『点名法』生成——
    每 case 一条：被测数据寄存器=标记(非0)、其余=0，期望=标记 → routing 全验，任何位宽都成立。"""
    wb = _build_mux_wb(
        str(tmp_path / "fg.xlsx"),
        tmm_fields=[
            ("CTRL", "h10", "d_ctrl[1:0]", "1:0", "N", "RW"),
            ("N0", "h20", "d_n0", "0", "N", "RW"),     # 三个 1-bit 数据字段：3 个互异值装不下
            ("N1", "h21", "d_n1", "0", "N", "RW"),
            ("N2", "h22", "d_n2", "0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_n0_to_mux", "d_ctrl_to_mux[1:0]", "2'b00", "d_fgout", 1),
            _mrow("d_n1_to_mux", "d_ctrl_to_mux[1:0]", "2'b01", "d_fgout", 1),
            _mrow("d_n2_to_mux", "d_ctrl_to_mux[1:0]", "2'b10", "d_fgout", 1),
        ],
    )
    r = resolver.Resolver(wb)
    grp = _grp(wb, "d_fgout")
    exp = mux_gen.expand_mux_group(wb, r, grp)
    vecs, meta = mux_gen.make_mux_vectors(grp, exp, mode="min")
    assert meta["marker_method"] is True                 # 切到点名法
    assert meta["value_collision"] is False              # 不再算假绿
    assert len(vecs) == 3                                 # 每 case 一条
    dks = exp["data_keys"]
    for ci, v in enumerate(vecs):                         # 被测=标记非0，其余=0，期望=标记
        assert v.exp_value != 0
        for j, dk in enumerate(dks):
            assert (v.assignments[dk] != 0) == (j == ci)
    # 不再 false-green：analyze 能生成（输出 top_out=0 → 裸名探针/或需前缀）
    a = generator.analyze_mux_group(r, wb, grp)
    assert a["status"] in ("clean", "needs-prefix", "bare-probe"), (a["status"], a.get("error"))


def test_wl_negative_vectors(wl_wb):
    """负向用例（反例）对通用形态向量同样可用。"""
    opts = generator.GenOptions(probe_prefixes=WL_PREFIXES, neg_all=True)
    res = generator.build(wl_wb, opts)
    assert res["summary"]["n_mux_generated"] == 5
    text = "\n".join("\n".join(lines) for lines, _ in res["blocks"])
    assert "_NEG:" in text                               # 负向 assert 标签


# ───────────── ⑦ 对抗式审查确认的 5 个缺陷的回归（2026-06-03 第十四轮 review）─────────────
from fixtures import _set_row, _tmm_reg, _tmm_field  # noqa: E402


def _build_mux_wb(path, tmm_fields, mux_rows, regmap_fields=None, logic_rows=None, dft_rows=None):
    """造一个只有 logic(默认空)/tmm/regmap/mux 四页的最小工作簿，给级联/边界回归用。

    tmm_fields: [(reg_name, addr, field_name(可带[msb:lsb]), bit, dig, typ), ...]
    mux_rows:   [{列字母: 值}, ...]（直接 _set_row 到 mux 页数据行，row 从 3 起）
    regmap_fields: [(reg, typ, signal), ...] 可选
    logic_rows: [{列字母: 值}, ...] 可选——往 logic 页加行（A..J=输入, K=输出, L=表达式,
                M=suffix, N=top_output, R=序号），用于复现"mux 数据/控制输入是某 logic 输出"的形态。
    dft_rows: [(A=<out>_to_dft, B=<gate>_to_dft, D=<out>, E=门控表达式), ...] 可选——造 dft 页。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "logic"
    _set_row(ws, 2, {"A": "in_A", "K": "out", "L": "expr", "M": "suffix",
                     "N": "top", "P": "owner", "R": "no"})
    for ri, cells in enumerate(logic_rows or [], start=3):
        _set_row(ws, ri, cells)
    rm = wb.create_sheet("regmap")
    _set_row(rm, 2, {"D": "Reg", "F": "Type", "G": "Signal", "J": "b15", "Y": "b0", "AE": "owner"})
    from openpyxl.utils import column_index_from_string
    for ri, (reg, typ, signal) in enumerate(regmap_fields or [], start=3):
        _set_row(rm, ri, {"D": reg, "F": typ, "G": signal})
        rm.cell(row=ri, column=column_index_from_string("Y"), value=1)
    tmm = wb.create_sheet("total_memory_map")
    r = 1
    for reg_name, addr, fname, bit, dig, typ in tmm_fields:
        r = _tmm_reg(tmm, r, reg_name, addr)
        r = _tmm_field(tmm, r, fname, bit, addr=addr, dig=dig, typ=typ)
    mx = wb.create_sheet("mux")
    _set_row(mx, 2, {"A": "mux_input", "B": "c1", "C": "c2", "D": "c3", "E": "c4",
                     "F": "case", "G": "mux_out", "H": "to_logic", "I": "top", "L": "Owner", "N": 0})
    for ri, cells in enumerate(mux_rows, start=3):
        _set_row(mx, ri, cells)
    if dft_rows:
        ds = wb.create_sheet("dft")
        _set_row(ds, 2, {"A": "A", "B": "B", "C": "C", "D": "out put signal", "E": "logic expression"})
        for ri, (a, b, d, e) in enumerate(dft_rows, start=3):
            _set_row(ds, ri, {"A": a, "B": b, "D": d, "E": e})
    wb.save(path)
    return excel_model.load_workbook(path)


def _mrow(inp, ctrl_b, case, out, n, h="to_dft"):
    return {"A": inp, "B": ctrl_b, "F": case, "G": out, "H": h, "I": 0, "L": "O1", "N": n}


def test_review_fix1_narrow_rw_carrier_at_ci1_rejected(tmp_path):
    """[审查#1 critical] 上游载体在 ci>=1 的窄 RW 字段：_effective_width 必须按真实字段位宽算
    （此前传 ci 而非 0 → clamp 被跳过 → 窄字段误判够宽）。修后窄 RW 被拒，改选够宽的 RO 线。"""
    wb = _build_mux_wb(
        str(tmp_path / "c1.xlsx"),
        tmm_fields=[
            ("UCTRL", "h10", "d_uctrl", "0", "N", "RW"),
            ("ULINE", "h11", "d_uline[3:0]", "3:0", "Y", "RO"),   # 4bit RO 线（够宽载体）
            ("ULOCAL", "h12", "d_ulocal[1:0]", "1:0", "N", "RW"),  # 字段仅 2bit（窄 RW）
            ("DD0", "h20", "d_dd0[3:0]", "3:0", "N", "RW"),
            ("DD1", "h21", "d_dd1[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            # 上游组1：out d_umux[3:0]，控制 d_uctrl；ci0=line(RO)，ci1=local(窄 RW，声明 4bit)
            _mrow("d_uline_to_mux[3:0]", "d_uctrl_to_mux", "1'b0", "d_umux[3:0]", 1),
            _mrow("d_ulocal_to_mux[3:0]", "d_uctrl_to_mux", "1'b1", "d_umux[3:0]", 1),
            # 下游组2：控制 = d_umux（级联），case 需要 4bit 值
            _mrow("d_dd0_to_mux[3:0]", "d_umux_to_mux[3:0]", "4'b1010", "d_dmux[3:0]", 2),
            _mrow("d_dd1_to_mux[3:0]", "d_umux_to_mux[3:0]", "4'b0101", "d_dmux[3:0]", 2),
        ],
    )
    r = resolver.Resolver(wb)
    g_dn = _grp(wb, "d_dmux")
    exp = mux_gen.expand_mux_group(wb, r, g_dn)
    recipe = exp["ctrl_drivers"][0]["recipe"]
    # 窄 RW(ci=1, 字段2bit) 被正确拒绝（eff=2<out_width=4），改选够宽的 RO 线(ci=0)
    assert recipe["carrier_ci"] == 0
    assert recipe["carrier_key"] == "m1.d:0"
    assert recipe["carrier_eff_width"] == 4


def test_review_fix1_narrow_only_carrier_skips_group(tmp_path):
    """[审查#1] 唯一载体是窄 RW（修前会被误判够宽并产假向量）→ 修后整组无载体 → 跳过给原因。"""
    wb = _build_mux_wb(
        str(tmp_path / "c1b.xlsx"),
        tmm_fields=[
            ("UCTRL", "h10", "d_uctrl", "0", "N", "RW"),
            ("UL0", "h11", "d_ul0[1:0]", "1:0", "N", "RW"),       # 窄
            ("ULOCAL", "h12", "d_ulocal[1:0]", "1:0", "N", "RW"),  # 窄
            ("DD0", "h20", "d_dd0[3:0]", "3:0", "N", "RW"),
            ("DD1", "h21", "d_dd1[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_ul0_to_mux[3:0]", "d_uctrl_to_mux", "1'b0", "d_umux[3:0]", 1),
            _mrow("d_ulocal_to_mux[3:0]", "d_uctrl_to_mux", "1'b1", "d_umux[3:0]", 1),
            _mrow("d_dd0_to_mux[3:0]", "d_umux_to_mux[3:0]", "4'b1010", "d_dmux[3:0]", 2),
            _mrow("d_dd1_to_mux[3:0]", "d_umux_to_mux[3:0]", "4'b0101", "d_dmux[3:0]", 2),
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_dmux"))
    # 两个窄载体都 eff=2<4 → 没有可用载体 → issue
    assert any("载体" in i for i in exp["issues"]), exp["issues"]


def test_review_fix2_wire_fallback_data_skipped(tmp_path):
    """[审查#2 critical] mux 数据输入 wire 兜底（表里查无）→ 必须进 prefix_risks → 默认跳过，
    不能静默生成 force 裸名（RTL 顶层不存在的网 → elaboration CUVUNF）。"""
    wb = _build_mux_wb(
        str(tmp_path / "c2.xlsx"),
        tmm_fields=[
            ("CMODE", "h10", "d_cmode", "0", "N", "RW"),
            ("DLOCAL", "h12", "d_dlocal[2:0]", "2:0", "N", "RW"),
        ],
        mux_rows=[
            # 数据 ci0 = 表里查无的衔接网（既不在 tmm/regmap、也不是 logic/mux 输出）→ wire 兜底
            _mrow("d_mystery_internal_to_mux[2:0]", "d_cmode_to_mux", "1'b0", "d_gout[2:0]", 1),
            _mrow("d_dlocal_to_mux[2:0]", "d_cmode_to_mux", "1'b1", "d_gout[2:0]", 1),
        ],
    )
    r = resolver.Resolver(wb)
    grp = _grp(wb, "d_gout")
    exp = mux_gen.expand_mux_group(wb, r, grp)
    b = exp["bindings"]["d:0"]
    assert b.found_in == "wire"           # 确认走了 wire 兜底
    # 即便配了输出探针前缀（隔离输出风险），wire 兜底输入仍要进 risks
    opts = generator.GenOptions(probe_prefixes={"d_gout": "U_X"})
    risks = generator.mux_prefix_risks(grp, exp, opts)
    assert any("wire" in why or "CUVUNF" in why for _t, _n, why in risks), risks
    # build 默认跳过（不生成 force 裸名）
    res = generator.build(wb, opts)
    assert res["summary"]["n_mux_generated"] == 0
    text = "\n".join("\n".join(lines) for lines, _ in res["blocks"])
    assert "d_mystery_internal" not in text


def test_review_fix3_carrier_wider_than_output_clamped(tmp_path):
    """[审查#3 major] 载体寄存器比上游输出宽：控制值按上游【输出】位宽截断，
    越界 case 的向量被丢弃（此前用载体位宽 → 高位漏过 → 假绿）。"""
    wb = _build_mux_wb(
        str(tmp_path / "c3.xlsx"),
        tmm_fields=[
            ("UCTRL", "h10", "d_uctrl", "0", "N", "RW"),
            ("ULINE", "h11", "d_uline[3:0]", "3:0", "Y", "RO"),
            ("ULOCAL", "h12", "d_ulocal[5:0]", "5:0", "N", "RW"),   # 载体 6bit（宽于输出 4bit）
            ("DD0", "h20", "d_dd0[3:0]", "3:0", "N", "RW"),
            ("DD1", "h21", "d_dd1[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_uline_to_mux[3:0]", "d_uctrl_to_mux", "1'b0", "d_umux[3:0]", 1),
            _mrow("d_ulocal_to_mux[5:0]", "d_uctrl_to_mux", "1'b1", "d_umux[3:0]", 1),
            # 下游控制声明 6bit、case 需要超过上游输出(4bit)的高位 → 物理上达不到 → 丢弃
            _mrow("d_dd0_to_mux[3:0]", "d_umux_to_mux[5:0]", "6'b110000", "d_dmux[3:0]", 2),
            _mrow("d_dd1_to_mux[3:0]", "d_umux_to_mux[5:0]", "6'b000001", "d_dmux[3:0]", 2),
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_dmux"))
    vecs, meta = mux_gen.make_mux_vectors(_grp(wb, "d_dmux"), exp, mode="min")
    # case 6'b110000(=48) 需要上游输出 48，但上游只 4bit → 48&0xF=0 ≠ 48 → 该向量丢弃
    assert meta["dropped"] >= 1
    assert all("6'b110000" not in (v.note or "") for v in vecs)


def test_review_fix4_upstream_ctrl_truncated_dropped(tmp_path):
    """[审查#4 major] 载体 case 需要的上游控制值超出上游控制寄存器实际位宽 →
    上游被截断选错 case → 该向量必须丢弃（此前不回验上游是否真命中载体 case）。"""
    wb = _build_mux_wb(
        str(tmp_path / "c4.xlsx"),
        tmm_fields=[
            ("UCTRL", "h10", "d_uctrl", "0", "N", "RW"),          # 控制字段仅 1bit
            ("UD0", "h11", "d_ud0[3:0]", "3:0", "Y", "RO"),
            ("UD1", "h12", "d_ud1[3:0]", "3:0", "Y", "RO"),
            ("UWIDE", "h13", "d_uwide[3:0]", "3:0", "N", "RW"),    # 够宽 RW 载体，但在 case 2'b10
            ("DD0", "h20", "d_dd0[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            # 上游控制 d_uctrl 声明 2bit（cases 2bit），但 tmm 字段仅 1bit
            _mrow("d_ud0_to_mux[3:0]", "d_uctrl_to_mux[1:0]", "2'b00", "d_umux[3:0]", 1),
            _mrow("d_ud1_to_mux[3:0]", "d_uctrl_to_mux[1:0]", "2'b01", "d_umux[3:0]", 1),
            _mrow("d_uwide_to_mux[3:0]", "d_uctrl_to_mux[1:0]", "2'b10", "d_umux[3:0]", 1),
            # 下游：控制 = d_umux，挑一个 case
            _mrow("d_dd0_to_mux[3:0]", "d_umux_to_mux[3:0]", "4'b1010", "d_dmux[3:0]", 2),
        ],
    )
    r = resolver.Resolver(wb)
    g_dn = _grp(wb, "d_dmux")
    exp = mux_gen.expand_mux_group(wb, r, g_dn)
    recipe = exp["ctrl_drivers"][0]["recipe"]
    # 载体 = case 2'b10 的 d_uwide（RW 优先），需要上游控制驱到 2，但 1bit 字段截断 2→0
    assert recipe["carrier_ci"] == 2 and recipe["ctrl_values"] == [2]
    vecs, meta = mux_gen.make_mux_vectors(g_dn, exp, mode="min")
    # 上游控制截断后选不中载体 case → 向量被丢弃（不静默生成假断言）
    assert meta["dropped"] >= 1
    assert any("截断" in rr for rr in meta.get("dropped_reasons", []))


def test_review_fix5_blank_continuation_rows_not_mismatch(tmp_path):
    """[审查#5 major] 续行控制列留空（合并单元格 / 控制只写组首行）→ 视为继承首行，
    不误判 mismatch（此前整组被跳过 + 误导『请核对 Excel』，是 LPBT 排版的回归）。"""
    wb = _build_mux_wb(
        str(tmp_path / "c5.xlsx"),
        tmm_fields=[
            ("CMODE", "h10", "d_cmode", "0", "N", "RW"),
            ("D0", "h11", "d_d0[2:0]", "2:0", "Y", "RO"),
            ("D1", "h12", "d_d1[2:0]", "2:0", "N", "RW"),
        ],
        mux_rows=[
            # 首行写控制；续行 B 列留空（模拟合并单元格/只写首行）
            _mrow("d_d0_to_mux[2:0]", "d_cmode_to_mux", "1'b0", "d_gg[2:0]", 1),
            {"A": "d_d1_to_mux[2:0]", "F": "1'b1", "G": "d_gg[2:0]", "H": "to_dft",
             "I": 0, "L": "O1", "N": 1},        # B 留空
        ],
    )
    grp = _grp(wb, "d_gg")
    assert grp.ctrl_mismatch_rows == []        # 续行空控制不算 mismatch
    assert len(grp.ctrls) == 1 and grp.ctrl_base == "d_cmode"
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, grp)
    assert not any("不一致" in i for i in exp["issues"]), exp["issues"]
    # 配前缀后照常生成（不被误跳过）
    res = generator.build(wb, generator.GenOptions(probe_prefixes={"d_gg": "U_X"}))
    assert res["summary"]["n_mux_generated"] == 1


def test_review_fix5_genuine_mismatch_still_caught(tmp_path):
    """[审查#5] 续行写了【不同的非空】控制信号 → 仍要抓（不能把真错误也放过）。"""
    wb = _build_mux_wb(
        str(tmp_path / "c5b.xlsx"),
        tmm_fields=[
            ("CMODE", "h10", "d_cmode", "0", "N", "RW"),
            ("CMODE2", "h13", "d_other", "0", "N", "RW"),
            ("D0", "h11", "d_d0[2:0]", "2:0", "Y", "RO"),
            ("D1", "h12", "d_d1[2:0]", "2:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_d0_to_mux[2:0]", "d_cmode_to_mux", "1'b0", "d_gg[2:0]", 1),
            _mrow("d_d1_to_mux[2:0]", "d_other_to_mux", "1'b1", "d_gg[2:0]", 1),  # 控制变了
        ],
    )
    grp = _grp(wb, "d_gg")
    assert grp.ctrl_mismatch_rows == [4]       # 续行控制不同 → 仍记 mismatch
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, grp)
    assert any("不一致" in i for i in exp["issues"])


# ───────────── ⑧ mux 数据输入是 top_output=0 logic 输出的 _to_mux 衔接网（2026-06-04 WL linectrl/_line）─────────────
# 真表 270 组里 ~41 个"数据输入未解析"的根因：mux 页 A 列写的是显式 X_to_mux 真网，但基名 X 撞上一个
# top_output=0 的 logic 输出 → resolve 在"内部信号"兜底短路（cone 还会假成环）→ 未解析跳过。
# 修复：原文带 _to_mux 时不走内部信号兜底，落到 force 字面网（needs-prefix）。


def _mux_skip_reason(wb, res, out_base):
    """从 compose_account 取某 mux 组的去向+原因（= 用户跑 --account 看到的那一行）。"""
    items = generator.compose_account(wb, generator.GenOptions(), res)
    it = next(i for i in items if i["kind"] == "mux" and i["name"].startswith(out_base))
    return it["disposition"], it["reason"]


def test_wl_linectrl_collision_data_input_forces_net(tmp_path):
    """linectrl 家族（名字撞车 + 纯透传）：mux 数据输入 X_to_mux，X 既是 top_output=0 logic 输出、
    又与一个 RO 寄存器同名。修前→resolve 在"内部信号"兜底短路→未解析跳过；修后→force 字面网。"""
    wb = _build_mux_wb(
        str(tmp_path / "linectrl.xlsx"),
        tmm_fields=[
            ("MODE", "h10", "d_wl_rf_mode", "0", "N", "RW"),
            ("BAND", "h11", "d_wl_rf_band_sel[5:0]", "5:0", "Y", "RO"),   # 底层 RO（与 logic 输出同名）
            ("LOCAL", "h12", "d_wl_rf_local[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_wl_rf_band_sel_to_mux[3:0]", "d_wl_rf_mode_to_mux", "1'b0", "d_wl_rf_out[3:0]", 1),
            _mrow("d_wl_rf_local_to_mux[3:0]", "d_wl_rf_mode_to_mux", "1'b1", "d_wl_rf_out[3:0]", 1),
        ],
        logic_rows=[
            # 纯透传 + 名字撞车：logic 输出 d_wl_rf_band_sel[3:0] = A（同名 RO 寄存器的低 4 位）
            {"A": "d_wl_rf_band_sel_to_logic[3:0]", "K": "d_wl_rf_band_sel[3:0]",
             "L": "A", "M": "to_mux", "N": 0, "R": 1},
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_wl_rf_out"))
    b = exp["bindings"]["d:0"]                              # 线控数据输入
    assert b.found_in == "needs-prefix", (b.found_in, b.note)
    assert b.kind == "RO"
    assert b.resolved is True                              # ⭐ 不再"未解析"
    assert b.wire == "d_wl_rf_band_sel_to_mux"            # force 字面衔接网
    assert not any("未解析" in i for i in exp["issues"]), exp["issues"]
    # --account 视角：没前缀 → 跳过原因是"缺前缀"(丙类)，不是"未解析"(丁类)
    res0 = generator.build(wb, generator.GenOptions())
    disp, reason = _mux_skip_reason(wb, res0, "d_wl_rf_out")
    assert disp == "跳过"
    assert "前缀" in reason and "未解析" not in reason, reason
    # 配前缀 → 整组生成
    res1 = generator.build(wb, generator.GenOptions(probe_prefixes={
        "d_wl_rf_out": "U_X.U_MUX", "d_wl_rf_band_sel_to_mux": "U_X.U_SIGLOGIC"}))
    assert res1["summary"]["n_mux_generated"] == 1


def test_wl_line_family_data_input_forces_net(tmp_path):
    """_line 家族（无名字撞车 + 门控三元 B?A:0）：mux 数据输入 X_to_mux，X 是 top_output=0 logic
    输出（与寄存器不同名）。同样应 force 字面 _to_mux 网（needs-prefix），不判未解析。"""
    wb = _build_mux_wb(
        str(tmp_path / "line.xlsx"),
        tmm_fields=[
            ("MODE", "h10", "d_wl_rf_tx_en_mode", "0", "N", "RW"),
            ("TXLINE", "h11", "d_wl_rf_linectrl_tx_en", "0", "Y", "RO"),    # _line 底层 RO 线控寄存器
            ("FREQ", "h12", "d_wl_rf_freq_5g_sel_reg", "0", "N", "RW"),     # 门控 B（简化为寄存器）
            ("LOCAL", "h13", "d_wl_rf_tx_en_local", "0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_wl_rf_tx_en_line_to_mux", "d_wl_rf_tx_en_mode_to_mux", "1'b0", "d_wl_rf_tx_en", 1),
            _mrow("d_wl_rf_tx_en_local_to_mux", "d_wl_rf_tx_en_mode_to_mux", "1'b1", "d_wl_rf_tx_en", 1),
        ],
        logic_rows=[
            # _line 形态：d_wl_rf_tx_en_line = B?A:1'b0（A=RO 线控, B=频选），top_output=0
            {"A": "d_wl_rf_linectrl_tx_en_to_logic", "B": "d_wl_rf_freq_5g_sel_reg_to_logic",
             "K": "d_wl_rf_tx_en_line", "L": "B?A:1'b0", "M": "to_mux", "N": 0, "R": 1},
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_wl_rf_tx_en"))
    b = exp["bindings"]["d:0"]
    assert b.found_in == "needs-prefix", (b.found_in, b.note)
    assert b.kind == "RO" and b.resolved is True
    assert b.wire == "d_wl_rf_tx_en_line_to_mux"
    assert not any("未解析" in i for i in exp["issues"]), exp["issues"]


def test_internal_to_logic_input_still_cone_expandable(tmp_path):
    """守护：修复只对 _to_mux 生效。logic 页输入引用 top_output=0 内部信号（_to_logic 后缀）时
    仍标 logic-internal（cone 展开路径不受影响）——logic 页输入从不带 _to_mux 后缀。"""
    wb = _build_mux_wb(
        str(tmp_path / "cone.xlsx"),
        tmm_fields=[("R0", "h10", "d_leaf_reg", "0", "N", "RW")],
        mux_rows=[  # 给个最小 mux 组占位（_build_mux_wb 需要 mux 页）
            _mrow("d_leaf_reg_to_mux", "d_leaf_reg_to_mux", "1'b0", "d_dummy", 9)],
        logic_rows=[
            # 内部信号 d_inner（top_output=0）；下游 d_top 以 d_inner_to_logic 引用它
            {"A": "d_leaf_reg_to_logic", "K": "d_inner", "L": "A", "M": "to_logic", "N": 0, "R": 1},
            {"A": "d_inner_to_logic", "K": "d_top", "L": "A", "M": "ls", "N": 1, "R": 2},
        ],
    )
    r = resolver.Resolver(wb)
    sig = next(s for s in wb.logic if s.out_base == "d_top")
    b = r.resolve_signal_inputs(sig)["A"]
    # _to_logic 内部信号 → 仍走 cone 可展开标记（未被 _to_mux 例外改变）
    assert b.found_in in ("logic-internal", "logic-computed"), (b.found_in, b.note)


def test_wl_control_internal_logic_output_forces_net(tmp_path):
    """控制侧·纯 OR 族（fb_en=A|B，输入全是 top_output=0 内部信号）：控制本身是内部 logic 输出，
    经显式 _to_mux 网喂 mux。无透传路径 → 改 force 字面 _to_mux 网（needs-prefix），不判未解析。"""
    wb = _build_mux_wb(
        str(tmp_path / "ctrl_or.xlsx"),
        tmm_fields=[
            ("LEAF", "h10", "d_leaf", "0", "N", "RW"),
            ("D0", "h20", "d_data0[3:0]", "3:0", "N", "RW"),
            ("D1", "h21", "d_data1[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_data0_to_mux[3:0]", "d_fb_en_to_mux", "1'b0", "d_cout[3:0]", 1),
            _mrow("d_data1_to_mux[3:0]", "d_fb_en_to_mux", "1'b1", "d_cout[3:0]", 1),
        ],
        logic_rows=[
            # 两个内部 enable（top_output=0），作为控制表达式的输入
            {"A": "d_leaf_to_logic", "K": "d_en_a", "L": "A", "M": "to_logic", "N": 0, "R": 1},
            {"A": "d_leaf_to_logic", "K": "d_en_b", "L": "A", "M": "to_logic", "N": 0, "R": 2},
            # 控制 fb_en = A|B（两输入都是内部信号）→ discover_ctrl_paths 找不到 line/local
            {"A": "d_en_a_to_logic", "B": "d_en_b_to_logic", "K": "d_fb_en",
             "L": "A|B", "M": "to_mux", "N": 0, "R": 3},
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_cout"))
    d = exp["ctrl_drivers"][0]
    assert d["source"] == "mux-force", (d["source"], exp["issues"])
    b = d["binding"]
    assert b.found_in == "needs-prefix", (b.found_in, b.note)
    assert b.kind == "RO" and b.resolved is True
    assert b.wire == "d_fb_en_to_mux"                      # force 字面控制衔接网
    assert not any("未解析" in i for i in exp["issues"]), exp["issues"]
    # --account：没前缀 → "缺前缀"(丙类)，不是"未解析"(丁类)
    res0 = generator.build(wb, generator.GenOptions())
    disp, reason = _mux_skip_reason(wb, res0, "d_cout")
    assert disp == "跳过"
    assert "前缀" in reason and "未解析" not in reason, reason
    # 配控制网前缀 → 整组生成
    res1 = generator.build(wb, generator.GenOptions(probe_prefixes={
        "d_cout": "U_X.U_MUX", "d_fb_en_to_mux": "U_X.U_SIGLOGIC"}))
    assert res1["summary"]["n_mux_generated"] == 1


def test_wl_control_gated_logic_output_forces_net(tmp_path):
    """控制侧·门控族（tx_filter_en=A&(B|C)，三输入全内部）：与纯 OR 同理，门控表达式同样无透传
    路径 → force 控制 _to_mux 网。守护"四种表达式形态统一触发"中的 A&(B|C)。"""
    wb = _build_mux_wb(
        str(tmp_path / "ctrl_gate.xlsx"),
        tmm_fields=[
            ("LEAF", "h10", "d_leaf", "0", "N", "RW"),
            ("D0", "h20", "d_data0[3:0]", "3:0", "N", "RW"),
            ("D1", "h21", "d_data1[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_data0_to_mux[3:0]", "d_txf_en_to_mux", "1'b0", "d_cout2[3:0]", 1),
            _mrow("d_data1_to_mux[3:0]", "d_txf_en_to_mux", "1'b1", "d_cout2[3:0]", 1),
        ],
        logic_rows=[
            {"A": "d_leaf_to_logic", "K": "d_en_a", "L": "A", "M": "to_logic", "N": 0, "R": 1},
            {"A": "d_leaf_to_logic", "K": "d_en_b", "L": "A", "M": "to_logic", "N": 0, "R": 2},
            {"A": "d_leaf_to_logic", "K": "d_en_c", "L": "A", "M": "to_logic", "N": 0, "R": 3},
            {"A": "d_en_a_to_logic", "B": "d_en_b_to_logic", "C": "d_en_c_to_logic",
             "K": "d_txf_en", "L": "A&(B|C)", "M": "to_mux", "N": 0, "R": 4},
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_cout2"))
    d = exp["ctrl_drivers"][0]
    assert d["source"] == "mux-force", (d["source"], exp["issues"])
    assert d["binding"].wire == "d_txf_en_to_mux"
    assert d["binding"].found_in == "needs-prefix"
    assert not any("未解析" in i for i in exp["issues"]), exp["issues"]


def test_lpbt_logic_control_passthrough_not_hijacked(tmp_path):
    """守护：控制是 logic 输出但有干净 line(RO)/local(RW) 透传路径（LPBT 形态）时，仍走 logic
    级联（source='logic'），不被 force 兜底劫持——force 只在『无透传路径』时接管。"""
    wb = _build_mux_wb(
        str(tmp_path / "lpbt_ctrl.xlsx"),
        tmm_fields=[
            ("MODE", "h10", "d_mode", "0", "N", "RW"),
            ("LINE", "h11", "d_line", "0", "Y", "RO"),
            ("LOCAL", "h12", "d_local", "0", "N", "RW"),
            ("D0", "h20", "d_data0[3:0]", "3:0", "N", "RW"),
            ("D1", "h21", "d_data1[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_data0_to_mux[3:0]", "d_ctrl_to_mux", "1'b0", "d_cout3[3:0]", 1),
            _mrow("d_data1_to_mux[3:0]", "d_ctrl_to_mux", "1'b1", "d_cout3[3:0]", 1),
        ],
        logic_rows=[
            # 控制 d_ctrl = mode?local:line（line=RO 透传, local=RW 透传）→ 有 line/local 路径
            {"A": "d_line_to_logic", "B": "d_mode_to_logic", "C": "d_local_to_logic",
             "K": "d_ctrl", "L": "B?C:A", "M": "to_mux", "N": 0, "R": 1},
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_cout3"))
    d = exp["ctrl_drivers"][0]
    assert d["source"] == "logic", (d["source"], exp["issues"])   # 没被 force 劫持
    assert d["line"] is not None or d["local"] is not None


# ───────────── 同源喂多 case：互异值按物理寄存器分配（2026-06-04 回归） ─────────────
def test_wl_shared_data_source_across_cases_generates(tmp_path):
    """⭐回归（2026-06-04，真表 vga_cfb_dpd_ss_lut5 症状）：同一个数据寄存器喂多个 case
    （d:0/d:1/d:2 全是 *_t0）。旧版互异值【按 case 序号】分配 → 同一寄存器被写 3 个不同值 →
    emit 的 by_base 冲突检查把每条向量都丢弃 → 空向量 → 误报"控制信号没有可用的驱动路径"。
    修复：互异值【按物理寄存器】分配，同源 case 复用同值 → 正常生成、不再空。"""
    wb = _build_mux_wb(
        str(tmp_path / "shared.xlsx"),
        tmm_fields=[
            ("CTRL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),      # 2bit 寄存器直出控制
            ("T0", "h20", "d_src_t0[2:0]", "2:0", "N", "RW"),
            ("T1", "h21", "d_src_t1[2:0]", "2:0", "N", "RW"),
        ],
        mux_rows=[
            # 3 个 case：00→t0, 01→t0(同源!), 10→t1
            _mrow("d_src_t0_to_mux[2:0]", "d_sel_to_mux[1:0]", "2'b00", "d_shared[2:0]", 1),
            _mrow("d_src_t0_to_mux[2:0]", "d_sel_to_mux[1:0]", "2'b01", "d_shared[2:0]", 1),
            _mrow("d_src_t1_to_mux[2:0]", "d_sel_to_mux[1:0]", "2'b10", "d_shared[2:0]", 1),
        ],
    )
    r = resolver.Resolver(wb)
    grp = _grp(wb, "d_shared")
    exp = mux_gen.expand_mux_group(wb, r, grp)
    assert not exp["issues"], exp["issues"]
    # 互异值：同源(t0)的两个 case 必须同值；t0 != t1（不同寄存器才需互异）
    values, collision = mux_gen.alloc_distinct_values(grp, exp["bindings"], exp["data_keys"])
    assert values[0] == values[1], "同一寄存器喂的两个 case 必须分配同值（否则 by_base 冲突全丢）"
    assert values[0] != values[2], "不同寄存器必须互异（mux 选路才能验）"
    assert not collision                                   # 2 个不同寄存器，3bit 装得下
    # 端到端：向量真的生成了（不再空、不再被 by_base 冲突全丢）
    vecs, meta = mux_gen.make_mux_vectors(grp, exp, mode="min")
    assert vecs, "同源喂多 case 不该再产生空向量"
    assert meta["dropped"] == 0, meta.get("dropped_reasons")
    assert not meta.get("value_collision")
    # build 流程同样生成（不再落进 skipped）
    import copy
    wb2 = copy.copy(wb)
    wb2.mux = [grp]
    res = generator.build(wb2, generator.GenOptions(types=["mux"], include_risky=True))
    assert res["summary"]["n_mux_generated"] == 1


def test_wl_shared_source_narrow_marker_marks_whole_register(tmp_path):
    """同源喂多 case + 字段太窄（点名法路径）：被测寄存器喂的【所有】case 一起给标记，
    只把别的寄存器置 0——否则同源 case 一个标记一个 0 仍会撞 by_base 冲突。"""
    wb = _build_mux_wb(
        str(tmp_path / "shared_narrow.xlsx"),
        tmm_fields=[
            ("CTRL", "h10", "d_sel2[1:0]", "1:0", "N", "RW"),
            ("T0", "h20", "d_n_t0", "0", "N", "RW"),          # 1bit
            ("T1", "h21", "d_n_t1", "0", "N", "RW"),          # 1bit
            ("T2", "h22", "d_n_t2", "0", "N", "RW"),          # 1bit
        ],
        mux_rows=[
            _mrow("d_n_t0_to_mux", "d_sel2_to_mux[1:0]", "2'b00", "d_narrow", 1),
            _mrow("d_n_t0_to_mux", "d_sel2_to_mux[1:0]", "2'b01", "d_narrow", 1),  # 同源 t0
            _mrow("d_n_t1_to_mux", "d_sel2_to_mux[1:0]", "2'b10", "d_narrow", 1),
            _mrow("d_n_t2_to_mux", "d_sel2_to_mux[1:0]", "2'b11", "d_narrow", 1),
        ],
    )
    r = resolver.Resolver(wb)
    grp = _grp(wb, "d_narrow")
    exp = mux_gen.expand_mux_group(wb, r, grp)
    assert not exp["issues"], exp["issues"]
    # 3 个不同寄存器、各 1bit（只放得下值 1，避 0）→ 互异值装不下 → 点名法
    _values, collision = mux_gen.alloc_distinct_values(grp, exp["bindings"], exp["data_keys"])
    assert collision
    # 点名 ci=0（t0）：t0 喂的两个 case(0,1)都给标记，t1/t2 置 0
    mv = mux_gen._marker_values(grp, 0, exp["bindings"], exp["data_keys"], "ones")
    assert mv[0] == 1 and mv[1] == 1 and mv[2] == 0 and mv[3] == 0, mv
    # 端到端：点名法生成、不空、不丢
    vecs, meta = mux_gen.make_mux_vectors(grp, exp, mode="min")
    assert vecs and meta.get("marker_method") and meta["dropped"] == 0, meta.get("dropped_reasons")


def test_wl_cascaded_control_with_shared_data_source_generates(tmp_path):
    """⭐回归（完全贴合真表 vga_cfb_dpd_ss_lut5 形态）：控制是级联 mux(temp_code) + 下游同一个
    数据寄存器喂多个 case。控制侧完全正常解析（全 RW/regmap 命中），却因数据侧 by_base 冲突空向量
    → 误报"控制信号没有可用的驱动路径"（指错了地方）。修复后正常生成。"""
    wb = _build_mux_wb(
        str(tmp_path / "casc_shared.xlsx"),
        tmm_fields=[
            ("TCMODE", "h10", "d_tcmode", "0", "N", "RW"),
            ("TCLINE", "h11", "d_tcline[3:0]", "3:0", "Y", "RO"),     # 够宽载体（RO 线）
            ("TCLOCAL", "h12", "d_tclocal[3:0]", "3:0", "N", "RW"),   # 够宽载体（RW，优先）
            ("DT0", "h20", "d_dt0[2:0]", "2:0", "N", "RW"),
            ("DT1", "h21", "d_dt1[2:0]", "2:0", "N", "RW"),
            ("DT2", "h22", "d_dt2[2:0]", "2:0", "N", "RW"),
        ],
        mux_rows=[
            # 上游组1：控制 mux d_tc = case(tcmode){0:line; 1:local}
            _mrow("d_tcline_to_mux[3:0]", "d_tcmode_to_mux", "1'b0", "d_tc[3:0]", 1),
            _mrow("d_tclocal_to_mux[3:0]", "d_tcmode_to_mux", "1'b1", "d_tc[3:0]", 1),
            # 下游组2：控制=d_tc(级联)，5 个 case 映射 t0/t0/t1/t1/t2（同一寄存器喂多 case）
            _mrow("d_dt0_to_mux[2:0]", "d_tc_to_mux[3:0]", "4'b0000", "d_dout[2:0]", 2),
            _mrow("d_dt0_to_mux[2:0]", "d_tc_to_mux[3:0]", "4'b0001", "d_dout[2:0]", 2),  # 同源 t0
            _mrow("d_dt1_to_mux[2:0]", "d_tc_to_mux[3:0]", "4'b0010", "d_dout[2:0]", 2),
            _mrow("d_dt1_to_mux[2:0]", "d_tc_to_mux[3:0]", "4'b0011", "d_dout[2:0]", 2),  # 同源 t1
            _mrow("d_dt2_to_mux[2:0]", "d_tc_to_mux[3:0]", "4'b0100", "d_dout[2:0]", 2),
        ],
    )
    r = resolver.Resolver(wb)
    grp = _grp(wb, "d_dout")
    exp = mux_gen.expand_mux_group(wb, r, grp)
    assert not exp["issues"], exp["issues"]
    assert exp["ctrl_drivers"][0]["source"] == "mux"      # 控制确是级联 mux，正常解析
    # 端到端：向量生成、不空、不丢（修复前这里空 → 误报"控制无驱动路径"）
    vecs, meta = mux_gen.make_mux_vectors(grp, exp, mode="min")
    assert vecs, "级联控制+同源数据不该再产生空向量"
    assert meta["dropped"] == 0, meta.get("dropped_reasons")
    # 同源 t0 的两个 case 在同一条向量里写同值
    dv, _coll = mux_gen.alloc_distinct_values(grp, exp["bindings"], exp["data_keys"])
    assert dv[0] == dv[1] and dv[2] == dv[3] and dv[0] != dv[2] != dv[4]


def test_empty_vector_reason_surfaces_dropped_reasons():
    """向量为空时不再一律甩"控制无驱动路径"：meta 有逐 case 丢弃原因就抛真因（去重）。"""
    # 有 dropped_reasons → 抛真因
    meta = {"dropped_reasons": [
        "case 2'b00: 寄存器 d_x 同时被下游数据与上游级联配方写不同值（冲突），该向量丢弃",
        "case 2'b01: 寄存器 d_x 同时被下游数据与上游级联配方写不同值（冲突），该向量丢弃",
    ]}
    msg = generator._empty_vector_reason(meta)
    assert "都被丢弃" in msg and "冲突" in msg
    assert "控制信号没有可用的驱动路径" not in msg
    # 无 dropped_reasons（控制来源 unknown / 无扫描路径）→ 通用兜底
    assert "控制信号没有可用的驱动路径" in generator._empty_vector_reason({})


# ───────────── 嵌套 mux 自动归一化（2026-06-04，envdet5g 实证形态） ─────────────
def _nested_envdet_wb(path):
    """复刻真表 envdet5g cornercode（1607~1612 行）的两级嵌套块：
       里层 case(r_code[2:0]){00x:lut0;01x:lut1;10x:lut2;11x:lut3} + 外层 case(sel){0:自己;1:local}。"""
    return _build_mux_wb(
        path,
        tmm_fields=[
            ("RCODE", "h3B", "d_r_code[2:0]", "2:0", "N", "RW"),
            ("SEL", "h10", "d_cc_sel", "0", "N", "RW"),
            ("LUT", "h20", "d_cc_lut0[1:0]", "1:0", "N", "RW"),
            ("LUT", "h20", "d_cc_lut1[1:0]", "3:2", "N", "RW"),
            ("LUT", "h20", "d_cc_lut2[1:0]", "5:4", "N", "RW"),
            ("LUT", "h20", "d_cc_lut3[1:0]", "7:6", "N", "RW"),
            ("LOCAL", "h21", "d_cc_local[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_cc_lut0_to_mux[1:0]", "d_r_code_to_mux[2:0]", "3'b00x", "d_cc[1:0]", 1),
            _mrow("d_cc_lut1_to_mux[1:0]", "d_r_code_to_mux[2:0]", "3'b01x", "d_cc[1:0]", 1),
            _mrow("d_cc_lut2_to_mux[1:0]", "d_r_code_to_mux[2:0]", "3'b10x", "d_cc[1:0]", 1),
            _mrow("d_cc_lut3_to_mux[1:0]", "d_r_code_to_mux[2:0]", "3'b11x", "d_cc[1:0]", 1),
            _mrow("d_cc_to_mux[1:0]", "d_cc_sel_to_mux", "1'b0", "d_cc[1:0]", 1),       # 自引用=里层结果
            _mrow("d_cc_local_to_mux[1:0]", "d_cc_sel_to_mux", "1'b1", "d_cc[1:0]", 1),
        ],
    )


def test_nested_mux_normalized_to_multi_ctrl(tmp_path):
    """⭐两级嵌套块（不同行不同控制+自引用）自动折叠成多控制拼接 {sel, r_code}，6 行→5 行。"""
    wb = _nested_envdet_wb(str(tmp_path / "nested.xlsx"))
    grp = _grp(wb, "d_cc")
    # 控制变成 [sel(高), r_code(低)]，4-bit 拼接
    assert grp.is_multi_ctrl and [c.base for c in grp.ctrls] == ["d_cc_sel", "d_r_code"]
    assert grp.ctrl_total_width == 4
    assert not grp.ctrl_mismatch_rows                      # 折叠后清空，不再报"控制列不一致"
    assert grp.normalized_note and "嵌套" in grp.normalized_note
    # 自引用那行消失；5 行拼接 case 正确
    pairs = [(c.case_raw, c.input_base) for c in grp.cases]
    assert pairs == [
        ("4'b1xxx", "d_cc_local"),       # sel=1 → local（r_code don't-care）
        ("4'b000x", "d_cc_lut0"),        # sel=0 → r_code 选 lut
        ("4'b001x", "d_cc_lut1"),
        ("4'b010x", "d_cc_lut2"),
        ("4'b011x", "d_cc_lut3"),
    ], pairs
    assert all(c.input_base != "d_cc" for c in grp.cases)  # 自引用已剔除


def test_nested_mux_end_to_end_generates(tmp_path):
    """折叠后走老的多控制拼接路径：解析通、向量生成、不空（救活原本的 ✗未解析）。"""
    wb = _nested_envdet_wb(str(tmp_path / "nested2.xlsx"))
    grp = _grp(wb, "d_cc")
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, grp)
    assert not exp["issues"], exp["issues"]
    vecs, meta = mux_gen.make_mux_vectors(grp, exp, mode="min")
    assert vecs and meta["dropped"] == 0, meta.get("dropped_reasons")
    assert meta.get("multi_ctrl")
    res = generator.build(wb, generator.GenOptions(types=["mux"], include_risky=True))
    assert res["summary"]["n_mux_generated"] == 1
    # 可见性：.sv 块顶有 ⚙ 注释，report 的可验证性 detail 也带 note（designer 复核用）
    sv_lines = [ln for lines, st in res["blocks"] if st.get("is_mux") for ln in lines]
    assert any("⚙" in ln and "嵌套" in ln for ln in sv_lines), "缺 .sv 折叠注释"
    rep = generator.report(wb, generator.GenOptions(types=["mux"]))
    vdet = next(v["detail"] for v in rep["verifiability"]["signals"] if v["type"] == "mux")
    assert "⚙" in vdet and "嵌套" in vdet, vdet


def test_nested_mux_no_self_ref_stays_errored(tmp_path):
    """护栏：控制列不一致但【没有自引用】（真 typo 形态）→ 不折叠，维持 ctrl_mismatch 报错。"""
    wb = _build_mux_wb(
        str(tmp_path / "noself.xlsx"),
        tmm_fields=[
            ("RCODE", "h3B", "d_r_code[2:0]", "2:0", "N", "RW"),
            ("SEL", "h10", "d_xx_sel", "0", "N", "RW"),
            ("D0", "h20", "d_xx_d0[1:0]", "1:0", "N", "RW"),
            ("D1", "h21", "d_xx_d1[1:0]", "1:0", "N", "RW"),
            ("D2", "h22", "d_xx_d2[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_xx_d0_to_mux[1:0]", "d_r_code_to_mux[2:0]", "3'b00x", "d_xx[1:0]", 1),
            _mrow("d_xx_d1_to_mux[1:0]", "d_r_code_to_mux[2:0]", "3'b01x", "d_xx[1:0]", 1),
            _mrow("d_xx_d2_to_mux[1:0]", "d_xx_sel_to_mux", "1'b1", "d_xx[1:0]", 1),   # 控制变了但无自引用
        ],
    )
    grp = _grp(wb, "d_xx")
    assert grp.ctrl_mismatch_rows                          # 没折叠，原报错保留
    assert not grp.normalized_note
    exp = mux_gen.expand_mux_group(wb, resolver.Resolver(wb), grp)
    assert any("不一致" in i for i in exp["issues"]), exp["issues"]


def test_nested_mux_three_regimes_not_normalized(tmp_path):
    """护栏：三个控制 regime（>2 级）太复杂、不猜 → 不折叠，维持原报错。"""
    wb = _build_mux_wb(
        str(tmp_path / "three.xlsx"),
        tmm_fields=[
            ("A", "h10", "d_a3", "0", "N", "RW"),
            ("B", "h11", "d_b3", "0", "N", "RW"),
            ("C", "h12", "d_c3", "0", "N", "RW"),
            ("D0", "h20", "d_y3_d0[1:0]", "1:0", "N", "RW"),
            ("D1", "h21", "d_y3_d1[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_y3_d0_to_mux[1:0]", "d_a3_to_mux", "1'b0", "d_y3[1:0]", 1),
            _mrow("d_y3_to_mux[1:0]", "d_b3_to_mux", "1'b1", "d_y3[1:0]", 1),       # 自引用
            _mrow("d_y3_d1_to_mux[1:0]", "d_c3_to_mux", "1'b1", "d_y3[1:0]", 1),    # 第三个控制
        ],
    )
    grp = _grp(wb, "d_y3")
    assert not grp.normalized_note and grp.ctrl_mismatch_rows


def test_nested_mux_bad_width_not_normalized(tmp_path):
    """护栏：子块 case 位宽与其控制位宽对不上 → _case_bitstr 抛错 → 整组 bail（不部分折叠），
    维持原 ctrl_mismatch 报错（绝不猜一个错结构出来）。"""
    wb = _build_mux_wb(
        str(tmp_path / "badw.xlsx"),
        tmm_fields=[
            ("RCODE", "h3B", "d_r_code[2:0]", "2:0", "N", "RW"),
            ("SEL", "h10", "d_bw_sel", "0", "N", "RW"),
            ("D0", "h20", "d_bw_d0[1:0]", "1:0", "N", "RW"),
            ("LOCAL", "h21", "d_bw_local[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_bw_d0_to_mux[1:0]", "d_r_code_to_mux[2:0]", "4'b0000", "d_bw[1:0]", 1),  # 4bit case vs 3bit 控制
            _mrow("d_bw_to_mux[1:0]", "d_bw_sel_to_mux", "1'b0", "d_bw[1:0]", 1),             # 自引用
            _mrow("d_bw_local_to_mux[1:0]", "d_bw_sel_to_mux", "1'b1", "d_bw[1:0]", 1),
        ],
    )
    grp = _grp(wb, "d_bw")
    assert not grp.normalized_note and grp.ctrl_mismatch_rows


def test_normal_multi_ctrl_group_untouched(wl_wb):
    """护栏：今天能跑的正常多控制拼接组（控制每行一致）完全不进归一化——note 空、结构原样。"""
    g4 = _grp(wl_wb, "d_wl_rf_tx_rc_code")                 # 真·多控制拼接 {lut_en, bwctrl}
    assert not g4.normalized_note
    assert not g4.ctrl_mismatch_rows
    assert [c.base for c in g4.ctrls] == ["d_wl_rf_rc_code_lut_en", "d_wl_rf_bwctrl"]


# ───────── 第二十轮 A1：cone 级联控制是上游输出【切片】时的 lsb 复合（mux56→mux157 实证）─────────
def test_a1_cascade_control_slice_lsb_composed(tmp_path):
    """[A1·必修·隐藏假红] 下游控制 = 上游 mux 输出的切片 d_umux_to_mux[3:1]（slice_lsb=1）：
    case 选值 N 落在上游输出的 [3:1] 段 → 写载体前必须左移 slice_lsb（载体=N<<1，使上游输出[3:1]=N）。
    修前 mux_gen 把 N 原样写载体 → 上游输出=N、被下游切成 N>>1 → 8 个 case 错 7 个（只有真仿真暴露）。
    （真表 mux56: temp_code=case(temp_code_mode){0:线控;1:local} → mux157: mixer2g_trim=case(temp_code[3:1]){8选1}）"""
    wb = _build_mux_wb(
        str(tmp_path / "a1.xlsx"),
        tmm_fields=[
            ("UCTRL", "h10", "d_uctrl", "0", "N", "RW"),
            ("ULINE", "h11", "d_uline[3:0]", "3:0", "Y", "RO"),     # 上游 line 路（RO）
            ("ULOCAL", "h12", "d_ulocal[3:0]", "3:0", "N", "RW"),   # 上游 local 路（RW，做载体）
            ("DD0", "h20", "d_dd0[1:0]", "1:0", "N", "RW"),
            ("DD1", "h21", "d_dd1[1:0]", "1:0", "N", "RW"),
            ("DD2", "h22", "d_dd2[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            # 上游组1：d_umux[3:0] = case(d_uctrl){0:line(RO); 1:local(RW)}；载体=local=m1.d:1
            _mrow("d_uline_to_mux[3:0]", "d_uctrl_to_mux", "1'b0", "d_umux[3:0]", 1),
            _mrow("d_ulocal_to_mux[3:0]", "d_uctrl_to_mux", "1'b1", "d_umux[3:0]", 1),
            # 下游组2：控制 = d_umux 的切片 [3:1]（slice_lsb=1，3 位）；取 N=0/2/5 子集
            _mrow("d_dd0_to_mux[1:0]", "d_umux_to_mux[3:1]", "3'b000", "d_dmux[1:0]", 2),
            _mrow("d_dd1_to_mux[1:0]", "d_umux_to_mux[3:1]", "3'b010", "d_dmux[1:0]", 2),
            _mrow("d_dd2_to_mux[1:0]", "d_umux_to_mux[3:1]", "3'b101", "d_dmux[1:0]", 2),
        ],
    )
    r = resolver.Resolver(wb)                          # 默认 cone（展开上游）
    g = _grp(wb, "d_dmux")
    exp = mux_gen.expand_mux_group(wb, r, g)
    assert not exp["issues"], exp["issues"]
    assert exp["ctrl_drivers"][0]["source"] == "mux"
    assert exp["ctrl_drivers"][0]["ctrl_slice_lsb"] == 1               # 记下切片偏移
    # 显示项：lsb>0 的切片在 expr/控制描述里显出来（一眼看见非零偏移）
    assert g.expr.startswith("case(d_umux[3:1])"), g.expr
    assert generator._mux_ctrl_desc(g) == "d_umux[3:1]"
    vecs, meta = mux_gen.make_mux_vectors(g, exp, mode="min")
    # 载体（上游 local RW，键 m1.d:1）= N<<1：case 000→0, 010→2<<1=4, 101→5<<1=10
    carriers = [v.assignments["m1.d:1"] for v in vecs]
    assert carriers == [0, 4, 10], carriers
    # 上游控制 d_uctrl 被驱到 1（选中 local 载体路径，使上游输出 == 载体值）
    assert all(v.assignments["m1.c:reg"] == 1 for v in vecs)


def test_a1_cascade_full_width_control_unchanged(tmp_path):
    """[A1 向后兼容] 下游控制是上游输出全宽 d_umux_to_mux[3:0]（slice_lsb=0）→ 载体 = N 原样，
    修复对 lsb=0 是 no-op（既有级联测试不受影响）。"""
    wb = _build_mux_wb(
        str(tmp_path / "a1fw.xlsx"),
        tmm_fields=[
            ("UCTRL", "h10", "d_uctrl", "0", "N", "RW"),
            ("ULINE", "h11", "d_uline[3:0]", "3:0", "Y", "RO"),
            ("ULOCAL", "h12", "d_ulocal[3:0]", "3:0", "N", "RW"),
            ("DD0", "h20", "d_dd0[1:0]", "1:0", "N", "RW"),
            ("DD1", "h21", "d_dd1[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_uline_to_mux[3:0]", "d_uctrl_to_mux", "1'b0", "d_umux[3:0]", 1),
            _mrow("d_ulocal_to_mux[3:0]", "d_uctrl_to_mux", "1'b1", "d_umux[3:0]", 1),
            _mrow("d_dd0_to_mux[1:0]", "d_umux_to_mux[3:0]", "4'b0011", "d_dmux[1:0]", 2),
            _mrow("d_dd1_to_mux[1:0]", "d_umux_to_mux[3:0]", "4'b0101", "d_dmux[1:0]", 2),
        ],
    )
    r = resolver.Resolver(wb)
    g = _grp(wb, "d_dmux")
    exp = mux_gen.expand_mux_group(wb, r, g)
    assert exp["ctrl_drivers"][0]["ctrl_slice_lsb"] == 0
    vecs, _meta = mux_gen.make_mux_vectors(g, exp, mode="min")
    carriers = [v.assignments["m1.d:1"] for v in vecs]
    assert carriers == [0b0011, 0b0101], carriers     # 全宽：载体 = case 值原样


# ───────── 第二十轮 A2：死分支去重（先到先得）+ 同值不同源=规格矛盾硬报（mixer2g_trim 实证）─────────
def test_a2_duplicate_same_source_shadowed(tmp_path):
    """[A2] 靠后 case 与更靠前 case 控制值相同、且选同一数据源 → 死分支（Verilog 先到先得）：
    标记 shadowed、跳过其测试 + 软提示，不阻断生成。（mixer2g_trim 行1090-1094≡1085-1089 实证）"""
    wb = _build_mux_wb(
        str(tmp_path / "a2dup.xlsx"),
        tmm_fields=[
            ("SEL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
            ("S0", "h20", "d_s0[1:0]", "1:0", "N", "RW"),
            ("S1", "h21", "d_s1[1:0]", "1:0", "N", "RW"),
            ("S2", "h22", "d_s2[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[1:0]", 1),
            _mrow("d_s1_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b01", "d_g[1:0]", 1),
            _mrow("d_s0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[1:0]", 1),  # 死分支≡行3、同源
            _mrow("d_s2_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b10", "d_g[1:0]", 1),
        ],
    )
    r = resolver.Resolver(wb)
    g = _grp(wb, "d_g")
    exp = mux_gen.expand_mux_group(wb, r, g)
    assert not exp["issues"], exp["issues"]            # 同源死分支不是 issue（不阻断生成）
    assert exp["shadowed"] == {2}                      # 第三行（ci=2）= 死分支
    assert "死分支" in exp["shadowed_note"]
    vecs, meta = mux_gen.make_mux_vectors(g, exp, mode="min")
    assert len(vecs) == 3                              # 00/01/10 三条；重复的 00 不再生成
    assert meta["shadowed"] == [2]
    assert meta["shadowed_note"]
    # .sv 块顶有 ⚙ 死分支注释（看 .sv 的人能复核少了哪些行）
    res = generator.build(wb, generator.GenOptions())
    block_text = "\n".join("\n".join(lines) for lines, st in res["blocks"] if st.get("is_mux"))
    assert "死分支" in block_text


def test_a2_duplicate_diff_source_contradiction(tmp_path):
    """[A2] 靠后 case 与更靠前 case 控制值相同、却选【不同】数据源 → 规格矛盾（RTL 同选择值只能输出一个）：
    硬报 issue、整组跳过让 designer 改 Excel（不静默挑一个，否则验的是错的期望）。"""
    wb = _build_mux_wb(
        str(tmp_path / "a2bad.xlsx"),
        tmm_fields=[
            ("SEL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
            ("S0", "h20", "d_s0[1:0]", "1:0", "N", "RW"),
            ("S1", "h21", "d_s1[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[1:0]", 1),
            _mrow("d_s1_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[1:0]", 1),  # 同控制值→不同源=矛盾
        ],
    )
    r = resolver.Resolver(wb)
    g = _grp(wb, "d_g")
    exp = mux_gen.expand_mux_group(wb, r, g)
    assert any("规格矛盾" in i for i in exp["issues"]), exp["issues"]
    # build 默认跳过该组并给原因（不静默生成可能错的期望）
    res = generator.build(wb, generator.GenOptions())
    assert any("d_g" in s[0] for s in res["skipped"]), [s[0] for s in res["skipped"]]


# ───────── 第二十轮 B2：用户手填 mux 数据值（override-by-base）+ 撞值非阻断警告 ─────────
def _b2_wb(path):
    return _build_mux_wb(
        path,
        tmm_fields=[
            ("SEL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
            ("S0", "h20", "d_s0[3:0]", "3:0", "N", "RW"),
            ("S1", "h21", "d_s1[3:0]", "3:0", "N", "RW"),
            ("S2", "h22", "d_s2[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux[3:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[3:0]", 1),
            _mrow("d_s1_to_mux[3:0]", "d_sel_to_mux[1:0]", "2'b01", "d_g[3:0]", 1),
            _mrow("d_s2_to_mux[3:0]", "d_sel_to_mux[1:0]", "2'b10", "d_g[3:0]", 1),
        ],
    )


def test_b2_data_override_replaces_value(tmp_path):
    """[B2] 手填数据值按物理基名替换自动互异值；选中该源的向量 auto_out 随之改；不撞值=无警告。"""
    wb = _b2_wb(str(tmp_path / "b2.xlsx"))
    r = resolver.Resolver(wb)
    g = _grp(wb, "d_g")
    exp = mux_gen.expand_mux_group(wb, r, g)
    v, m = mux_gen.make_mux_vectors(g, exp, mode="min", data_overrides={"d_s1": 5})
    # case 2'b01 → d:1（d_s1）；手填 5 → 该数据键=5 且 auto_out(exp_value)=5
    assert v[1].assignments["d:1"] == 5
    assert v[1].exp_value == 5
    # 其余源仍自动分配、互不相同，不撞值
    assert not m.get("value_collision") and not m.get("override_collision")
    vals = {v[0].assignments["d:0"], v[1].assignments["d:1"], v[2].assignments["d:2"]}
    assert len(vals) == 3 and 5 in vals


def test_b2_data_override_collision_warns_not_skips(tmp_path):
    """[B2] 手填两路同值 → override_collision 非阻断警告(用户负责)，仍生成、不 skip 整组；
    build 把警告挂进 mux_warnings（不进 skipped）。"""
    wb = _b2_wb(str(tmp_path / "b2c.xlsx"))
    r = resolver.Resolver(wb)
    g = _grp(wb, "d_g")
    exp = mux_gen.expand_mux_group(wb, r, g)
    v, m = mux_gen.make_mux_vectors(g, exp, mode="min", data_overrides={"d_s0": 3, "d_s1": 3})
    assert m.get("override_collision") is True
    assert not m.get("value_collision")        # 不阻断
    assert v                                    # 仍生成
    res = generator.build(wb, generator.GenOptions(mux_data={"d_g": {"d_s0": 3, "d_s1": 3}}))
    assert not any("d_g" in s[0] for s in res["skipped"])        # 没被跳过
    assert any("撞值" in w[2] for w in res["mux_warnings"]), res["mux_warnings"]


# ───────── 第二十轮 对抗式审查确认问题的回归 ─────────
def test_a2_subsumption_diff_source_contradiction(tmp_path):
    """[审查#2/#3·major] x 位 case 与更窄 case 在【具体控制值】上重叠却选不同源 = 隐藏假红：
    3'b01x(→s0) 覆盖 {010,011}；后面 3'b010(→s1) 的 010 已被 s0 拥有 → 规格矛盾硬报（精确同键去重抓不到）。"""
    wb = _build_mux_wb(
        str(tmp_path / "a2sub.xlsx"),
        tmm_fields=[
            ("SEL", "h10", "d_sel[2:0]", "2:0", "N", "RW"),
            ("S0", "h20", "d_s0[3:0]", "3:0", "N", "RW"),
            ("S1", "h21", "d_s1[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux[3:0]", "d_sel_to_mux[2:0]", "3'b01x", "d_g[3:0]", 1),
            _mrow("d_s1_to_mux[3:0]", "d_sel_to_mux[2:0]", "3'b010", "d_g[3:0]", 1),
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_g"))
    assert any("规格矛盾" in i for i in exp["issues"]), exp["issues"]
    res = generator.build(wb, generator.GenOptions())
    assert any("d_g" in s[0] for s in res["skipped"])           # 默认跳过（不产假红）


def test_a2_subsumption_same_source_redundant(tmp_path):
    """[审查#2] 更窄 case 在更靠前的 x 位同源 case 内 → 死分支(被完全覆盖)，shadowed+软提示，不阻断。"""
    wb = _build_mux_wb(
        str(tmp_path / "a2subok.xlsx"),
        tmm_fields=[
            ("SEL", "h10", "d_sel[2:0]", "2:0", "N", "RW"),
            ("S0", "h20", "d_s0[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux[3:0]", "d_sel_to_mux[2:0]", "3'b01x", "d_g[3:0]", 1),
            _mrow("d_s0_to_mux[3:0]", "d_sel_to_mux[2:0]", "3'b010", "d_g[3:0]", 1),  # 010∈{010,011}，同源
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_g"))
    assert not exp["issues"], exp["issues"]
    assert exp["shadowed"] == {1}                                # 后行被完全覆盖=死分支
    assert "死分支" in exp["shadowed_note"]


def test_a1_slice_exceeds_upstream_width_flagged(tmp_path):
    """[审查#1] 下游控制切片 msb 超出上游 mux 输出位宽 → 硬报 issue（否则高位 case 选值被静默截断丢弃）。"""
    wb = _build_mux_wb(
        str(tmp_path / "a1ovf.xlsx"),
        tmm_fields=[
            ("UCTRL", "h10", "d_uctrl", "0", "N", "RW"),
            ("ULINE", "h11", "d_uline[3:0]", "3:0", "Y", "RO"),
            ("ULOCAL", "h12", "d_ulocal[3:0]", "3:0", "N", "RW"),
            ("DD0", "h20", "d_dd0[1:0]", "1:0", "N", "RW"),
            ("DD1", "h21", "d_dd1[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_uline_to_mux[3:0]", "d_uctrl_to_mux", "1'b0", "d_umux[3:0]", 1),
            _mrow("d_ulocal_to_mux[3:0]", "d_uctrl_to_mux", "1'b1", "d_umux[3:0]", 1),
            # 控制切片 [5:1]（msb=5 ≥ 上游 d_umux 位宽 4）
            _mrow("d_dd0_to_mux[1:0]", "d_umux_to_mux[5:1]", "5'b00000", "d_dmux[1:0]", 2),
            _mrow("d_dd1_to_mux[1:0]", "d_umux_to_mux[5:1]", "5'b00001", "d_dmux[1:0]", 2),
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_dmux"))
    assert any("超出上游" in i and "位宽" in i for i in exp["issues"]), exp["issues"]


def test_b2_narrow_field_override_keeps_marker_protection(tmp_path):
    """[审查#4/#8/#10·critical] 窄字段(装不下互异值)下手填一个值【不该】关掉点名法保护→假绿。
    3 个 1-bit 数据源：无 override 走点名法；手填一个后仍走点名法、不误报手填撞值，并提示手填未生效。"""
    wb = _build_mux_wb(
        str(tmp_path / "b2narrow.xlsx"),
        tmm_fields=[
            ("SEL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
            ("S0", "h20", "d_s0", "0", "N", "RW"),       # 1-bit
            ("S1", "h21", "d_s1", "0", "N", "RW"),
            ("S2", "h22", "d_s2", "0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux", "d_sel_to_mux[1:0]", "2'b00", "d_g", 1),
            _mrow("d_s1_to_mux", "d_sel_to_mux[1:0]", "2'b01", "d_g", 1),
            _mrow("d_s2_to_mux", "d_sel_to_mux[1:0]", "2'b10", "d_g", 1),
        ],
    )
    r = resolver.Resolver(wb)
    g = _grp(wb, "d_g")
    exp = mux_gen.expand_mux_group(wb, r, g)
    _v0, m0 = mux_gen.make_mux_vectors(g, exp, mode="min")
    assert m0.get("marker_method")                       # 无 override：点名法保护
    _v1, m1 = mux_gen.make_mux_vectors(g, exp, mode="min", data_overrides={"d_s0": 1})
    assert m1.get("marker_method")                       # 手填后仍点名法（修前=False=假绿）
    assert not m1.get("override_collision")              # 不归咎手填（容量问题非手填撞值）
    assert m1.get("override_ignored_marker")             # 照实提示：手填值在点名法下未生效


def test_b2_allones_override_no_spurious_collision(tmp_path):
    """[审查#6·major] 全1 手填值(0xF)在 max/exhaustive 反码轮 ~0xF&0xF=0，不该误报 override_collision。"""
    wb = _b2_wb(str(tmp_path / "b2allones.xlsx"))         # 3 个 4-bit 源（互异装得下）
    r = resolver.Resolver(wb)
    g = _grp(wb, "d_g")
    exp = mux_gen.expand_mux_group(wb, r, g)
    for mode in ("min", "max", "exhaustive"):
        _v, m = mux_gen.make_mux_vectors(g, exp, mode=mode, data_overrides={"d_s1": 0xF})
        assert not m.get("override_collision"), mode      # 全1 取反成 0 不算撞值
        assert not m.get("value_collision"), mode


def test_b2_overwidth_override_truncation_noted(tmp_path):
    """[审查#5] 超字段宽的手填值被静默截断 → meta['override_truncated'] 记下来给提示（导入/接口路径）。"""
    wb = _b2_wb(str(tmp_path / "b2ovf.xlsx"))             # d_s1 字段 4-bit
    r = resolver.Resolver(wb)
    g = _grp(wb, "d_g")
    exp = mux_gen.expand_mux_group(wb, r, g)
    _v, m = mux_gen.make_mux_vectors(g, exp, mode="min", data_overrides={"d_s1": 0x1F})
    assert m.get("override_truncated", {}).get("d_s1") == (0x1F, 0xF)


def test_used_vars_collapses_data_by_register(tmp_path):
    """[显示收拢] used_vars 按物理寄存器去重——死分支/LUT 型同源多 case 不重复列(像 for_test t0~t7 各一行)；
    data_keys 仍逐 case 保留(向量生成对齐)；compute_drives 本就按寄存器合并，故 .sv 不变。"""
    wb = _build_mux_wb(
        str(tmp_path / "collapse.xlsx"),
        tmm_fields=[
            ("SEL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
            ("S0", "h20", "d_s0[1:0]", "1:0", "N", "RW"),
            ("S1", "h21", "d_s1[1:0]", "1:0", "N", "RW"),
            ("S2", "h22", "d_s2[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[1:0]", 1),
            _mrow("d_s1_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b01", "d_g[1:0]", 1),
            _mrow("d_s0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[1:0]", 1),  # 死分支同源
            _mrow("d_s2_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b10", "d_g[1:0]", 1),
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_g"))
    data_in_used = [k for k in exp["used_vars"] if mux_gen.key_role(k) == "data"]
    assert len(data_in_used) == 3, data_in_used                       # s0/s1/s2 各一行(d:2 死分支收拢)
    assert {exp["bindings"][k].base.lower() for k in data_in_used} == {"d_s0", "d_s1", "d_s2"}
    assert len(exp["data_keys"]) == 4                                  # data_keys 仍 4 个(逐 case)


# ───────── 第二十轮 续②：DFT 路观测模式（force iddq 门到透传值）─────────
def test_dft_observe_forces_iddq_gate(tmp_path):
    """[续②] 输出在 dft 页被 iddq 门控(B?0:A)时：dft_observe=True 在产物前导 force 门到透传值(0)，
    使我们断言的基名网反映功能值（= for_test 的 iddq=0）；关时无前导。"""
    wb = _build_mux_wb(
        str(tmp_path / "dft.xlsx"),
        tmm_fields=[
            ("SEL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
            ("IDDQ", "h11", "d_iddq_mode", "0", "Y", "RO"),    # DFT 门=RO 线
            ("S0", "h20", "d_s0[1:0]", "1:0", "N", "RW"),
            ("S1", "h21", "d_s1[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[1:0]", 1),
            _mrow("d_s1_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b01", "d_g[1:0]", 1),
        ],
        dft_rows=[("d_g_to_dft[1:0]", "d_iddq_mode_to_dft", "d_g[1:0]", "B?0:A")],
    )
    # dft 页解析：d_g 被 d_iddq_mode 门控、透传值 0
    assert wb.dft.get("d_g") == {"gate_base": "d_iddq_mode",
                                 "gate_raw": "d_iddq_mode_to_dft", "transparent": 0}
    # 关：无前导 force
    res0 = generator.build(wb, generator.GenOptions(dft_observe=False))
    assert not res0.get("dft_preamble")
    # 开：前导 force iddq 门到 1'b0
    res1 = generator.build(wb, generator.GenOptions(dft_observe=True))
    pre = "\n".join(res1.get("dft_preamble") or [])
    assert "force" in pre and "d_iddq_mode" in pre and "1'b0" in pre, pre
    # render 把前导写进产物最前
    txt = generator.render(res1)
    assert "d_iddq_mode" in txt and "1'b0" in txt


def test_wl_iddq_dft(tmp_path):
    """⭐ item③ iddq DFT 态拍（第二十二轮）：被 dft 页门控(B?0:A)的输出，全面/穷举补一条 DFT 拍
    (force 门=1 → 断言输出=常量支0 → 该拍后 release)；精简不补。期望取门常量支(0)，非"透传取反"；
    release 保证不钉死后续拍(S4)。"""
    def _wb():
        return _build_mux_wb(
            str(tmp_path / "iddq.xlsx"),
            tmm_fields=[
                ("SEL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
                ("IDDQ", "h11", "d_iddq_mode", "0", "Y", "RO"),    # DFT 门=RO 线（可 force）
                ("S0", "h20", "d_s0[1:0]", "1:0", "N", "RW"),
                ("S1", "h21", "d_s1[1:0]", "1:0", "N", "RW"),
            ],
            mux_rows=[
                _mrow("d_s0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[1:0]", 1),
                _mrow("d_s1_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b01", "d_g[1:0]", 1),
            ],
            dft_rows=[("d_g_to_dft[1:0]", "d_iddq_mode_to_dft", "d_g[1:0]", "B?0:A")],
        )
    # 精简：不补 DFT 拍（保三档区别）—— iddq 门完全不出现
    txt_min = generator.render(generator.build(_wb(), generator.GenOptions(mux_mode="min")))
    assert "d_iddq_mode" not in txt_min
    # 全面：补一条 DFT 拍
    res = generator.build(_wb(), generator.GenOptions(mux_mode="max"))
    txt = generator.render(res)
    assert "force `ENV_RF.d_iddq_mode=1'b1;" in txt          # 门=1：B?0:A 的非透传支(选中常量0)
    assert "release `ENV_RF.d_iddq_mode;" in txt             # S4：拍后释放，不钉死后续拍
    # force..release 区段内断言输出==常量支 0（2-bit 输出 → 2'b00；功能值非 0，故非"取反"）
    seg = txt[txt.index("force `ENV_RF.d_iddq_mode=1'b1;"):txt.index("release `ENV_RF.d_iddq_mode;")]
    assert "==2'b00)begin" in seg, seg
    # DFT 拍是正例、期望来自 dft 门 → 不计入 designer 手填统计（designer_filled 排除 dft_pitch）
    assert res["summary"]["n_designer"] == 0
    # 穷举档同样补；exhaustive 不因 DFT 拍报错
    txt_ex = generator.render(generator.build(_wb(), generator.GenOptions(mux_mode="exhaustive")))
    assert "force `ENV_RF.d_iddq_mode=1'b1;" in txt_ex and "release `ENV_RF.d_iddq_mode;" in txt_ex
    # M3 + 报告一致：report() 也含 DFT 拍，且明细期望来源标"DFT门"（非 designer 手填）
    rep = generator.report(_wb(), generator.GenOptions(mux_mode="max"))
    dg = [d for d in rep["detail"] if d["signal"].startswith("d_g")]
    assert dg, "报告里应有 d_g 的明细行"
    assert any("DFT门" in d["exp_src"] for d in dg), dg
    # Fix C（评审 major）：报告 force 列须含 iddq 门 force（extra_forces），否则报告与 .sv 不符
    assert any("d_iddq_mode=1'b1" in (d.get("force") or "") for d in dg), [d.get("force") for d in dg]
    # Fix A（评审 blocker）：dft_observe 开时 DFT 拍后必须 force 回透传(恢复全局前导)，不能 bare release
    txt_obs = generator.render(generator.build(
        _wb(), generator.GenOptions(mux_mode="max", dft_observe=True)))
    assert "release `ENV_RF.d_iddq_mode" not in txt_obs       # 开 dft_observe：不 release（会抹前导）
    assert "force `ENV_RF.d_iddq_mode=1'b1;" in txt_obs        # DFT 拍 force 门=1
    assert "force `ENV_RF.d_iddq_mode=1'b0;" in txt_obs        # 拍后 force 回透传 0（restore 前导）


def test_wl_iddq_dft_skip_surfaced(tmp_path):
    """评审 M2/Fix D+E：iddq 门不是可 force 的 RO 网 → DFT 拍补不上，原因须在 .sv(// ⚠) 与报告里可见
    （缺口必须可见，别让"少验一支 iddq"无声无息）。"""
    wb = _build_mux_wb(
        str(tmp_path / "iddqskip.xlsx"),
        tmm_fields=[
            ("SEL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
            ("IDDQ", "h11", "d_iddq_mode", "0", "N", "RW"),    # 门=RW（非可 force RO）→ 跳过 DFT 拍
            ("S0", "h20", "d_s0[1:0]", "1:0", "N", "RW"),
            ("S1", "h21", "d_s1[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            _mrow("d_s0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_g[1:0]", 1),
            _mrow("d_s1_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b01", "d_g[1:0]", 1),
        ],
        dft_rows=[("d_g_to_dft[1:0]", "d_iddq_mode_to_dft", "d_g[1:0]", "B?0:A")],
    )
    res = generator.build(wb, generator.GenOptions(mux_mode="max"))
    assert "未补 DFT 拍" in generator.render(res)                # .sv 块顶 // ⚠ 透出（M2）
    assert any("未补 DFT 拍" in w for _n, _a, w in res["mux_warnings"])
    rep = generator.report(wb, generator.GenOptions(mux_mode="max"))
    assert any("未补 DFT 拍" in (r.get("detail") or "")          # 报告可验证性 detail 也透出（Fix E）
               for r in rep["verifiability"]["signals"]), rep["verifiability"]["signals"]


# ───────────── 第二十三轮：规格冲突(同 case 不同源)单列 + owner 透出（P1 = 1+3）─────────────
def test_spec_collision_distinct_surfacing(tmp_path):
    """designer 表里同一控制选择值却选不同数据源(如 lctune off 段 case 错写成 on 段) → 整组跳过，
    但与缺前缀/假绿等其它跳过【区分开】：expand 出结构化 spec_conflicts(带行号/两源/owner)、
    issue 文案带 owner、analyze 状态=spec-collision、账目去向=跳过·规格冲突。保持跳过(不硬生成假红)。"""
    wb = _build_mux_wb(
        str(tmp_path / "coll.xlsx"),
        tmm_fields=[
            ("CTRL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
            ("A0", "h20", "d_a0[1:0]", "1:0", "N", "RW"),
            ("B0", "h21", "d_b0[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            # 第 3 行 case 2'b00 选 d_a0；第 5 行 case 2'b00 又选【不同源】d_b0 → 规格冲突
            _mrow("d_a0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_coll[1:0]", 1),
            _mrow("d_a0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b01", "d_coll[1:0]", 1),
            _mrow("d_b0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_coll[1:0]", 1),
        ],
    )
    r = resolver.Resolver(wb)
    grp = _grp(wb, "d_coll")
    exp = mux_gen.expand_mux_group(wb, r, grp)
    # ① 结构化导出（不靠串匹配）：行号 + 两源 + owner（_mrow 默认 owner=O1）
    assert exp["spec_conflicts"], "应导出 spec_conflicts"
    sc = exp["spec_conflicts"][0]
    assert sc["later_row"] == 5 and sc["first_row"] == 3      # 第5行(后)撞第3行(先)
    assert sc["later_src"] == "d_b0" and sc["first_src"] == "d_a0"
    assert sc["owner"] == "O1"
    # ② issue 文案带 owner（报告/account reason 全文可见）
    assert any("owner: O1" in m for m in exp["issues"])
    assert any("选不同数据源" in m for m in exp["issues"])
    # ③ analyze（GUI 状态源）= spec-collision（区别于普通 unresolved）
    a = generator.analyze_mux_group(r, wb, grp)
    assert a["status"] == "spec-collision"
    # ④ 账目去向单列「跳过·规格冲突」+ build 结果带 spec_conflicts 通道
    res = generator.build(wb, generator.GenOptions())
    assert any(s["name"] == grp.out_name and s["owner"] == "O1"
               for s in res["spec_conflicts"])
    items = generator.compose_account(wb, generator.GenOptions(), res)
    it = next(x for x in items if x["name"] == grp.out_name)
    assert it["disposition"] == "跳过·规格冲突"
    assert "owner: O1" in it["reason"]


def test_spec_collision_does_not_taint_normal_skip(tmp_path):
    """对照：普通缺前缀/正常组不应被误标 spec_conflicts；shadowed 同源死分支(同 case 同源)不算冲突。"""
    wb = _build_mux_wb(
        str(tmp_path / "ok.xlsx"),
        tmm_fields=[
            ("CTRL", "h10", "d_sel[1:0]", "1:0", "N", "RW"),
            ("A0", "h20", "d_a0[1:0]", "1:0", "N", "RW"),
        ],
        mux_rows=[
            # 同 case 同源(纯死分支/冗余) → 不是规格冲突
            _mrow("d_a0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_ok[1:0]", 1),
            _mrow("d_a0_to_mux[1:0]", "d_sel_to_mux[1:0]", "2'b00", "d_ok[1:0]", 1),
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_ok"))
    assert exp["spec_conflicts"] == []                       # 同源重复不算规格冲突
    res = generator.build(wb, generator.GenOptions())
    assert res["spec_conflicts"] == []


def test_spec_collision_twin_sources_hint_dup_name(tmp_path):
    """孪生源名(1st↔2nd 等仅一小段不同) → likely_dup_name=True + 提示『疑似两个 mux 撞同一输出名/
    漏改名』(那是合法的"一个控制管多个 mux")，而非一口咬定真矛盾。仍保持跳过·待核对。"""
    wb = _build_mux_wb(
        str(tmp_path / "twin.xlsx"),
        tmm_fields=[
            ("CTRL", "h10", "d_sel[3:0]", "3:0", "N", "RW"),
            ("G1", "h20", "d_slna_1st_bias_g0[3:0]", "3:0", "N", "RW"),
            ("G2", "h21", "d_slna_2nd_bias_g0[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            # 同 case 4'b0000、同输出名 d_slna_1st_bias，却分别选 1st / 2nd 孪生源 → 疑似撞名漏改
            _mrow("d_slna_1st_bias_g0_to_mux[3:0]", "d_sel_to_mux[3:0]", "4'b0000",
                  "d_slna_1st_bias[3:0]", 1),
            _mrow("d_slna_2nd_bias_g0_to_mux[3:0]", "d_sel_to_mux[3:0]", "4'b0000",
                  "d_slna_1st_bias[3:0]", 1),
        ],
    )
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, _grp(wb, "d_slna_1st_bias"))
    sc = exp["spec_conflicts"][0]
    assert sc["likely_dup_name"] is True
    assert "另一个 mux" in sc["hint"]
    assert any("孪生" in m for m in exp["issues"])
    # 对照：d_a0 vs d_b0 不孪生 → likely_dup_name=False（沿用上面的通用提示）
    assert mux_gen._looks_like_twins("d_a0", "d_b0") is False
    assert mux_gen._looks_like_twins("d_wl_rf_2g_pfb", "d_wl_rf_5g_pfb") is True


# ───────── 第二十四轮：级联 mode=0 备用载体是 logic→mux 网(needs-prefix) → 缺口可见 + 配前缀可生成 ─────────
def test_cascade_alt_mode0_logic_carrier_visible_and_generatable(tmp_path):
    """band_trim 形态回归：级联 mux 的 mode=0 备用载体是【logic 页输出经 _to_mux 衔接网】(needs-prefix)、
    不是 RO 寄存器(对比 mixer5g_trim 的 tsensor)。修前：载体扫描把它静默排除 → mode=0 那半张表无声漏掉
    (carrier_alt=None 且无 skip 原因)；这正是『有些 mux freq_sel_mode 能出 0、band_trim 只剩 1』的根因。
    修后(第二十四轮终版,用户拍板"没前缀也要看见正确 testcase")：①没配前缀 → mode=0 照样【生成】(裸名 force
    线控网,与 bare-probe 同口径) + 标 cascade_alt_bare 警告;②配了前缀(→prefixed-wire) → mode=0 走带前缀层级、无警告。"""
    wb = _build_mux_wb(
        str(tmp_path / "band_trim.xlsx"),
        tmm_fields=[
            ("FREQSEL", "h70", "d_freq_sel_mode", "0", "N", "RW"),
            ("BANDLOC", "h71", "d_band_sel_local[3:0]", "3:0", "N", "RW"),
            ("PFB0", "h72", "d_pfb_b0_trim[3:0]", "3:0", "N", "RW"),
            ("PFB1", "h73", "d_pfb_b1_trim[3:0]", "3:0", "N", "RW"),
            ("PFB2", "h74", "d_pfb_b2_trim[3:0]", "3:0", "N", "RW"),
            ("PFB3", "h75", "d_pfb_b3_trim[3:0]", "3:0", "N", "RW"),
        ],
        logic_rows=[
            # mode=0 备用源 d_linectrl_band_sel 是 logic 输出(M=to_mux)→ 被 mux 引用时判 needs-prefix
            {"A": "d_band_sel_line_to_logic[3:0]", "K": "d_linectrl_band_sel[3:0]",
             "L": "A", "M": "to_mux", "N": 0, "R": 1},
        ],
        mux_rows=[
            # 上游 mux7 band_sel[3:0] = freq_sel_mode? local(RW) : linectrl(logic→mux 网)
            _mrow("d_linectrl_band_sel_to_mux[3:0]", "d_freq_sel_mode_to_mux", "1'b0",
                  "d_band_sel[3:0]", 7, h="to_mux"),
            _mrow("d_band_sel_local_to_mux[3:0]", "d_freq_sel_mode_to_mux", "1'b1",
                  "d_band_sel[3:0]", 7, h="to_mux"),
            # 下游 mux8 band_trim[3:0] = case(band_sel[1:0] 切片) 4 选 1 → b0..b3 trim
            _mrow("d_pfb_b0_trim_to_mux[3:0]", "d_band_sel_to_mux[1:0]", "2'b00", "d_pfb_band_trim[3:0]", 8),
            _mrow("d_pfb_b1_trim_to_mux[3:0]", "d_band_sel_to_mux[1:0]", "2'b01", "d_pfb_band_trim[3:0]", 8),
            _mrow("d_pfb_b2_trim_to_mux[3:0]", "d_band_sel_to_mux[1:0]", "2'b10", "d_pfb_band_trim[3:0]", 8),
            _mrow("d_pfb_b3_trim_to_mux[3:0]", "d_band_sel_to_mux[1:0]", "2'b11", "d_pfb_band_trim[3:0]", 8),
        ],
    )
    g = _grp(wb, "d_pfb_band_trim")
    # ① 没配前缀：mode=0 照样【生成】(裸名 force 线控网) + cascade_alt_bare 警告(不跳过、不静默)
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, g)
    vecs, meta = mux_gen.make_mux_vectors(g, exp, mode="max", max_tests=256)
    assert exp["ctrl_drivers"][0]["recipe"]["carrier_alt_ci"] is not None   # 裸名网也当 alt 载体
    assert meta.get("cascade_alt") is True
    assert meta.get("cascade_alt_bare") == "d_linectrl_band_sel"            # 警告:裸名 force、待配前缀
    assert not meta.get("cascade_alt_skipped")                              # 不再"跳过"
    altv = [v for v in vecs if "alt" in (v.note or "")]
    assert len(altv) == 4                                                   # 没前缀也出 4 条 mode=0
    # 裸名 force 的就是线控网本身(wire 不带层级前缀)
    b = exp["bindings"]
    wires = [b[k].wire for k in altv[0].assignments if k in b and "linectrl_band_sel" in (b[k].base or "").lower()]
    assert wires and all("d_linectrl_band_sel_to_mux" in w and "ENV_RF" not in w for w in wires)
    # ② 配了 scan_rtl 探针前缀：linectrl 网解析成 prefixed-wire → 走带前缀层级、无裸名警告
    r2 = resolver.Resolver(wb, wire_prefixes={"d_linectrl_band_sel": "ENV_RF.U_WL"})
    exp2 = mux_gen.expand_mux_group(wb, r2, g)
    vecs2, meta2 = mux_gen.make_mux_vectors(g, exp2, mode="max", max_tests=256)
    assert exp2["ctrl_drivers"][0]["recipe"]["carrier_alt_ci"] is not None
    assert meta2.get("cascade_alt") is True
    assert not meta2.get("cascade_alt_bare")                      # 配了前缀就不再是裸名
    altv2 = [v for v in vecs2 if "alt" in (v.note or "")]
    assert len(altv2) == 4                                        # b0..b3 各一条 mode=0
    b2 = exp2["bindings"]
    wires2 = [b2[k].wire for k in altv2[0].assignments if k in b2 and "linectrl_band_sel" in (b2[k].base or "").lower()]
    assert wires2 and all(w.startswith("ENV_RF.U_WL.") for w in wires2)   # 带层级前缀


def test_cascade_alt_mode0_depth2_ro_register(tmp_path):
    """深层(第2层)级联 mode=0 生成（第二十四轮 audit 桶二）：D ← M(中间) ← MM(mode mux)，MM 的 mode=0
    源是 RO【寄存器】(force 名即可、不需前缀)。修前两分支只展开最近一层 → 深层 mode mux 的 mode=0 静默
    漏掉(cascade_alt=None)；修后递归传播 use_alt → 生成 mode=0。断言到【赋值级】：确实把 mode 位驱到 0
    + force RO 线控载体、且不写 RW 主载体（不是挂名假 mode=0）。"""
    wb = _build_mux_wb(
        str(tmp_path / "depth2.xlsx"),
        tmm_fields=[
            ("MODE", "h70", "d_mode", "0", "N", "RW"),
            ("MMLOC", "h71", "d_mm_local[1:0]", "1:0", "N", "RW"),
            ("MMLINE", "h72", "d_mm_line[1:0]", "1:0", "Y", "RO"),   # mode=0 源 = RO 寄存器(非 logic→mux 网)
            ("MIDA", "h73", "d_mid_a[1:0]", "1:0", "N", "RW"),
            ("MIDB", "h74", "d_mid_b[1:0]", "1:0", "N", "RW"),
            ("D0", "h75", "d_down0[3:0]", "3:0", "N", "RW"),
            ("D1", "h76", "d_down1[3:0]", "3:0", "N", "RW"),
        ],
        mux_rows=[
            # MM(mux70) d_mm[1:0] = d_mode? local(RW) : line(RO 寄存器)
            _mrow("d_mm_line_to_mux[1:0]", "d_mode_to_mux", "1'b0", "d_mm[1:0]", 70, h="to_mux"),
            _mrow("d_mm_local_to_mux[1:0]", "d_mode_to_mux", "1'b1", "d_mm[1:0]", 70, h="to_mux"),
            # M(mux71, 中间层) d_mid[1:0] = case(d_mm[1:0]) → mid RW 源
            _mrow("d_mid_a_to_mux[1:0]", "d_mm_to_mux[1:0]", "2'b00", "d_mid[1:0]", 71, h="to_mux"),
            _mrow("d_mid_b_to_mux[1:0]", "d_mm_to_mux[1:0]", "2'b01", "d_mid[1:0]", 71, h="to_mux"),
            # D(mux72, 最下游) d_down[3:0] = case(d_mid[0] 切片) → down RW 源
            _mrow("d_down0_to_mux[3:0]", "d_mid_to_mux[0]", "1'b0", "d_down[3:0]", 72),
            _mrow("d_down1_to_mux[3:0]", "d_mid_to_mux[0]", "1'b1", "d_down[3:0]", 72),
        ],
    )
    g = _grp(wb, "d_down")
    r = resolver.Resolver(wb)
    exp = mux_gen.expand_mux_group(wb, r, g)
    # mode mux 在第2层：D 的直接 recipe 无 alt，但递归树里 MM 有
    assert exp["ctrl_drivers"][0]["recipe"].get("carrier_alt_ci") is None
    assert mux_gen._recipe_has_alt(exp["ctrl_drivers"][0]["recipe"]) is True
    vecs, meta = mux_gen.make_mux_vectors(g, exp, mode="max", max_tests=256)
    assert meta.get("cascade_alt") is True
    alt = [v for v in vecs if "alt" in (v.note or "")]
    assert len(alt) == 2                              # down0/down1 各一条 mode=0
    # 赋值级断言：确实驱动 mode=0 线控路径
    b = exp["bindings"]
    nv = {b[k].base.lower(): val for k, val in alt[0].assignments.items() if k in b}
    assert nv.get("d_mode") == 0                      # 第2层 mode mux 驱到 mode=0 case
    assert "d_mm_line" in nv                          # RO 线控载体被 force
    assert "d_mm_local" not in nv                     # mode=0 不写 RW 主载体


# ───────────── 第二十六轮：mux 删除测试列 / 一键清空（generator 层） ─────────────
def _first_buildable_mux(wb):
    """挑一个能建 min 档 >=2 条向量的 mux 组（generator 层删除/清空测试用）。"""
    r = resolver.Resolver(wb)
    for grp in wb.mux:
        try:
            exp = mux_gen.expand_mux_group(wb, r, grp)
            if exp["issues"]:
                continue
            vecs, meta = mux_gen.make_mux_vectors(grp, exp, mode="min")
            if len(vecs) >= 2 and not meta.get("value_collision"):
                return grp, vecs
        except Exception:  # noqa: BLE001
            continue
    raise AssertionError("无可建 mux 组")


def test_mux_cleared_skips_in_build(lpbt_wb):
    """mux_cleared → 整组零用例：不进 blocks、进 skipped 且原因带「清空」。"""
    grp, _vecs = _first_buildable_mux(lpbt_wb)
    name = grp.out_name
    res = generator.build(lpbt_wb, generator.GenOptions(signals=[name], mux_cleared=[name]))
    assert name not in {st.get("out_name") for _l, st in res["blocks"]}
    reasons = [rs for n, _a, rs in res["skipped"] if n == name]
    assert reasons and "清空" in "; ".join(str(w) for *_h, w in reasons[0])


def test_mux_dropped_filters_in_build(lpbt_wb):
    """mux_dropped(按 mux_assign_key) → 删一条测试列就少一条向量；删光 → 跳过给原因。"""
    grp, vecs = _first_buildable_mux(lpbt_wb)
    name = grp.out_name
    base = generator.build(lpbt_wb, generator.GenOptions(signals=[name]))
    base_n = sum(st["n_vectors"] for _l, st in base["blocks"] if st.get("out_name") == name)
    assert base_n == len(vecs)                      # min 档无负向
    sig0 = generator.mux_assign_key(vecs[0].assignments)
    dr = generator.build(lpbt_wb, generator.GenOptions(signals=[name], mux_dropped={name: [sig0]}))
    dr_n = sum(st["n_vectors"] for _l, st in dr["blocks"] if st.get("out_name") == name)
    assert dr_n == base_n - 1
    # 删光所有签名 → 整组跳过(零用例)，原因带「删除全部」
    allsigs = [generator.mux_assign_key(v.assignments) for v in vecs]
    dz = generator.build(lpbt_wb, generator.GenOptions(signals=[name], mux_dropped={name: allsigs}))
    assert name not in {st.get("out_name") for _l, st in dz["blocks"]}
    reasons = [rs for n, _a, rs in dz["skipped"] if n == name]
    assert reasons and "删除全部" in "; ".join(str(w) for *_h, w in reasons[0])


def test_mux_cleared_dropped_in_report(lpbt_wb):
    """报告与 build 同口径：mux_cleared → summary 行 error 带「清空」；mux_dropped → n_tests 减少。"""
    grp, vecs = _first_buildable_mux(lpbt_wb)
    name = grp.out_name
    rep0 = generator.report(lpbt_wb, generator.GenOptions(signals=[name]))
    row0 = next(r for r in rep0["summary"] if r["signal"] == name)
    base_n = row0["n_tests"]
    repc = generator.report(lpbt_wb, generator.GenOptions(signals=[name], mux_cleared=[name]))
    rowc = next(r for r in repc["summary"] if r["signal"] == name)
    assert "清空" in rowc["error"]
    sig0 = generator.mux_assign_key(vecs[0].assignments)
    repd = generator.report(lpbt_wb, generator.GenOptions(signals=[name], mux_dropped={name: [sig0]}))
    rowd = next(r for r in repd["summary"] if r["signal"] == name)
    assert rowd["n_tests"] == base_n - 1
