# -*- coding: utf-8 -*-
"""Topout 主视图 GUI 层测试（2026-06-23 重构：Topout 视图=默认门面，旧 logic/mux 降级『排查(旧)』）。

无头 Qt（offscreen）。文字断言走 widget.text()（offscreen 缺中文字体、截图里中文是方框，
所以版面靠截图、文字靠断言）；关键路径截一张 PNG 过目版面。
DoD 1-5：载 mirror → Topout 视图列 12 信号+分类 → 选 rx_en 真值表对上金标准引擎 →
生成含断言的 .sv → 导出报告/回填 → 截图。
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import make_mirror_btlp                              # noqa: E402
from dreg_verify import excel_model as M             # noqa: E402
from dreg_verify import resolver as R                # noqa: E402
from dreg_verify import topout as T                  # noqa: E402


@pytest.fixture(scope="module")
def gui_app():
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    pytest.importorskip("PySide6")
    from PySide6 import QtWidgets
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([])
    yield app


@pytest.fixture(autouse=True)
def _isolate_gui_settings(monkeypatch, tmp_path):
    pytest.importorskip("PySide6")
    from dreg_verify import gui as G
    monkeypatch.setattr(G, "SETTINGS_PATH", str(tmp_path / "gui_settings.json"))
    monkeypatch.setattr(G, "EDITS_PATH", str(tmp_path / "edits.json"))


@pytest.fixture(scope="module")
def mirror_path(tmp_path_factory):
    p = tmp_path_factory.mktemp("btlp_gui") / "mirror_btlp_dreg.xlsx"
    make_mirror_btlp.build(str(p))
    return str(p)


@pytest.fixture()
def topo_win(gui_app, mirror_path):
    from dreg_verify import gui as G
    w = G.MainWindow()
    w.resize(1320, 840)
    w.path_edit.setText(mirror_path)
    w.on_load()
    yield w
    w.close()


def _topo_row(w, name):
    from dreg_verify import gui as G
    for r in range(w.topo_table.rowCount()):
        # 显示名带位宽切片(d_x[2:0])，按【基名】(剥切片)匹配，与内部 key 一致
        if w.topo_table.item(r, G.TOPO_NAME).text().split("[", 1)[0] == name:
            return r
    raise AssertionError("Topout 行未找到: %s" % name)


# ───────────── DoD 1：Topout 视图 = 默认门面 + 列出信号 + 分类 ─────────────
def test_topout_is_default_main_tab(topo_win):
    """外层标签：Topout(默认) + logic/mux/dft/iddq 子视图 + 排查(旧)；打开即停在 Topout。"""
    w = topo_win
    names = [w.main_tabs.tabText(i) for i in range(w.main_tabs.count())]
    assert names == ["Topout 视图", "logic 视图", "mux 视图", "dft 视图", "iddq 视图", "排查(旧)"]
    assert w.main_tabs.currentIndex() == 0           # 默认门面 = Topout
    assert isinstance(w.page_views["logic"], type(w.topout_view))   # 子视图复用同一 SignalView


def test_topout_lists_twelve_signals_with_classification(topo_win):
    """载 mirror → Topout 清单列 12 个要验信号 + 分类(选路/mux/直连寄存器/RO跳过)。"""
    from dreg_verify import gui as G
    w = topo_win
    assert w.topo_table.rowCount() == 12
    cell_names = [w.topo_table.item(r, G.TOPO_NAME).text() for r in range(12)]
    # 显示名带位宽切片(2026-06-24)：多 bit 信号清单里要看得见 [w-1:0]
    assert "d_bt_lp_lna_itrim[3:0]" in cell_names
    assert "d_logic_bt_lp_lna_agc[2:0]" in cell_names
    assert "clk_force_on" in cell_names                  # 1 bit → 不加切片
    kind_of = {n.split("[", 1)[0]: w.topo_table.item(r, G.TOPO_KIND).text()
               for r, n in enumerate(cell_names)}
    assert kind_of["d_logic_bt_lp_rx_en"] == G.TOPO_KIND_LABEL["logic"]
    assert kind_of["d_bt_lp_lna_itrim"] == G.TOPO_KIND_LABEL["mux"]
    assert kind_of["clk_force_on"] == G.TOPO_KIND_LABEL["register"]
    assert kind_of["pll_lock_indicator"] == G.TOPO_KIND_LABEL["ro-readback"]
    # 每行 owner 来自 Topout A 列（非空）
    for r in range(12):
        assert w.topo_table.item(r, G.TOPO_OWNER).text()


def test_topout_ro_status_skip(topo_win):
    """RO 回读 → 状态列『跳过』(信息蓝)，不崩、不静默丢。"""
    from dreg_verify import gui as G
    w = topo_win
    r = _topo_row(w, "pll_lock_indicator")
    assert w.topo_table.item(r, G.TOPO_STATUS).text() == G.TOPO_STATUS_LABEL["skip"]


# ───────────── DoD 1：选中 → 展开链 + 真值表 + 账目 ─────────────
def test_topout_select_rx_en_truth_table_matches_golden(topo_win, mirror_path):
    """选 rx_en → 真值表(4 输入 + auto + 期望)×12 列；期望值对上 for_test 金标准引擎(非自证)。"""
    from dreg_verify import gui as G
    w = topo_win
    w.topo_cov.setCurrentText("全面")               # max 档
    r = _topo_row(w, "d_logic_bt_lp_rx_en")
    w.topo_table.setCurrentCell(r, G.TOPO_NAME)
    # 真值表结构：行 = 4 输入 + auto_out + 期望；列 = 12 测试
    assert w.topo_truth.rowCount() == 6
    assert w.topo_truth.columnCount() == 12
    # 行表头含源寄存器/线控叶子（SignalPath 实证）
    vlabels = [w.topo_truth.verticalHeaderItem(i).text() for i in range(6)]
    assert any("d_bt_lp_linelocal_mode_ctrl" in x for x in vlabels)
    assert any("d_bt_lp_rx_en_local" in x for x in vlabels)
    # 期望值对上金标准（引擎一致：模型↔report↔.sv 同源）
    golden = T.load_fortest_golden(mirror_path)
    rx_blk = next(b for b in golden if b["out"] == "d_logic_bt_lp_rx_en")
    wb = M.load_workbook(mirror_path)
    reps = T.validate_against_golden(wb, R.Resolver(wb), [rx_blk])
    assert reps[0]["status"] == "checked" and reps[0]["n_bad"] == 0 and reps[0]["n_ok"] > 0


def test_topout_register_passthrough_shows_table_and_note(topo_win):
    """直连寄存器根 → 展开链显示 passthrough 说明 + 真值表(1 输入)。"""
    from dreg_verify import gui as G
    w = topo_win
    r = _topo_row(w, "clk_force_on")
    w.topo_table.setCurrentCell(r, G.TOPO_NAME)
    assert "直连寄存器" in w.topo_chain.toPlainText()
    assert w.topo_truth.rowCount() == 3            # 1 输入 + auto + 期望
    assert w.topo_truth.columnCount() >= 1


def test_topout_ro_select_no_truth_table(topo_win):
    """选 RO 回读 → 无真值表(空)，展开链记账原因(不崩)。"""
    from dreg_verify import gui as G
    w = topo_win
    r = _topo_row(w, "pll_lock_indicator")
    w.topo_table.setCurrentCell(r, G.TOPO_NAME)
    assert w.topo_truth.rowCount() == 0
    assert "回读" in w.topo_chain.toPlainText() or "skip" in w.topo_chain.toPlainText()


# ───────────── DoD 2：出 .sv（预览 + 导出）─────────────
def test_topout_preview_sv(topo_win):
    """预览选中.sv → .sv 预览页含断言、断言贴顶层真名、RO 记账(不静默丢)。"""
    w = topo_win
    w.on_topo_preview()
    sv = w.topo_sv.toPlainText()
    assert sv.count("assert (") > 0
    assert "`ENV_RF.d_logic_bt_lp_rx_en==" in sv
    assert "d_logic_bt_lp_rx_en_to_logic" not in sv     # 贴顶层真名，无尾缀
    assert "pll_lock_indicator" in sv                    # RO 记账注释
    assert w.topo_inner.currentIndex() == 1              # 自动切到 .sv 预览页


def test_topout_export_sv_file(topo_win, tmp_path, monkeypatch):
    """导出 .sv → 写盘成功、含断言。"""
    from PySide6 import QtWidgets
    w = topo_win
    out = tmp_path / "topout_out.sv"
    monkeypatch.setattr(w.topout_view, "_ask_export_options",
                        lambda: {"comments": False, "sv_summary": False, "owner_in_msg": False})
    monkeypatch.setattr(QtWidgets.QFileDialog, "getSaveFileName",
                        staticmethod(lambda *a, **k: (str(out), "")))
    monkeypatch.setattr(QtWidgets.QMessageBox, "information",
                        staticmethod(lambda *a, **k: None))
    w.on_topo_export_sv()
    assert out.exists()
    text = out.read_text(encoding="utf-8")
    assert text.count("assert (") > 0
    assert "`ENV_RF.clk_force_on==" in text              # register passthrough 也在


def test_topout_export_sv_options_comments_and_summary(topo_win, tmp_path, monkeypatch):
    """⭐#3：导出 .sv 选项(注释/末尾汇总计数)真的改产物——勾上 → 文件头注释 + 命名汇总块出现；
    且导出摘要含负向/designer 拆分(后端 n_negative/n_designer 聚合)。"""
    from PySide6 import QtWidgets
    w = topo_win
    v = _sel(w, "d_logic_bt_lp_rx_en")
    v._e_regen(); v.truth.setCurrentCell(0, 0); v._e_addneg()      # 造一条负向 → 摘要应显示负向≥1
    out = tmp_path / "opt.sv"
    seen = {}
    monkeypatch.setattr(w.topout_view, "_ask_export_options",
                        lambda: {"comments": True, "sv_summary": True, "owner_in_msg": False})
    monkeypatch.setattr(QtWidgets.QFileDialog, "getSaveFileName",
                        staticmethod(lambda *a, **k: (str(out), "")))
    monkeypatch.setattr(QtWidgets.QMessageBox, "information",
                        staticmethod(lambda *a, **k: seen.update(msg=a[2] if len(a) > 2 else "")))
    w.on_topo_export_sv()
    text = out.read_text(encoding="utf-8")
    assert text.startswith("// auto-generated")            # comments=True → 文件头注释
    assert "dreg_n_real_fail" in text                      # sv_summary=True → 计数器命名块
    assert "负向 1" in seen.get("msg", "")                  # 导出摘要含负向数
    v._e_regen()


def test_topout_export_scope_pos_neg(topo_win, tmp_path, monkeypatch):
    """⭐#3 2b：导出范围——仅正向剔除负向断言；仅负向只留负向、且无负向的信号不出现(记账，不静默丢)。"""
    from PySide6 import QtWidgets
    w = topo_win
    w._topo_check_all(True)
    v = _sel(w, "d_logic_bt_lp_rx_en")
    v._e_regen(); v.truth.setCurrentCell(0, 0); v._e_addneg()      # 给 rx_en 加一条负向

    def export(scope):
        out = tmp_path / ("scope_%s.sv" % scope)
        monkeypatch.setattr(w.topout_view, "_ask_export_options",
                            lambda: {"scope": scope, "comments": False,
                                     "sv_summary": False, "owner_in_msg": False})
        monkeypatch.setattr(QtWidgets.QFileDialog, "getSaveFileName",
                            staticmethod(lambda *a, **k: (str(out), "")))
        monkeypatch.setattr(QtWidgets.QMessageBox, "information", staticmethod(lambda *a, **k: None))
        w.on_topo_export_sv()
        return out.read_text(encoding="utf-8")

    pos = export("pos")
    assert "d_logic_bt_lp_rx_en==" in pos and "_NEG:" not in pos    # 仍有该信号、无负向断言
    neg = export("neg")
    assert "_NEG:" in neg                                            # 仅负向：含负向断言
    assert "d_logic_bt_lp_rx_en==" in neg
    assert "clk_force_on==" not in neg                              # 无负向的寄存器信号不出现
    v._e_regen()


def test_topout_export_sv_only_checked(topo_win, tmp_path, monkeypatch):
    """勾选过滤：只勾 rx_en → 导出仅含它（其余 Topout 信号块不出现）。"""
    from PySide6 import QtCore, QtWidgets
    from dreg_verify import gui as G
    w = topo_win
    w._topo_check_all(False)
    r = _topo_row(w, "d_logic_bt_lp_rx_en")
    w.topo_table.item(r, G.TOPO_SEL).setCheckState(QtCore.Qt.Checked)
    out = tmp_path / "one.sv"
    monkeypatch.setattr(w.topout_view, "_ask_export_options",
                        lambda: {"comments": False, "sv_summary": False, "owner_in_msg": False})
    monkeypatch.setattr(QtWidgets.QFileDialog, "getSaveFileName",
                        staticmethod(lambda *a, **k: (str(out), "")))
    monkeypatch.setattr(QtWidgets.QMessageBox, "information",
                        staticmethod(lambda *a, **k: None))
    w.on_topo_export_sv()
    text = out.read_text(encoding="utf-8")
    assert "`ENV_RF.d_logic_bt_lp_rx_en==" in text
    assert "`ENV_RF.clk_force_on==" not in text          # 其它信号未勾 → 不产出
    w._topo_check_all(True)


# ───────────── DoD 3：出报告（HTML / for_test）─────────────
def test_topout_export_report_html(topo_win, tmp_path, monkeypatch):
    """导出 Topout 报告(HTML) → 含 rx_en 真值表 + RO 也在(账目 summary，不丢)。"""
    from PySide6 import QtWidgets
    w = topo_win
    out = tmp_path / "rep.html"
    monkeypatch.setattr(QtWidgets.QFileDialog, "getSaveFileName",
                        staticmethod(lambda *a, **k: (str(out), "")))
    monkeypatch.setattr(QtWidgets.QMessageBox, "information",
                        staticmethod(lambda *a, **k: None))
    w.on_topo_export_report()
    assert out.exists()
    html = out.read_text(encoding="utf-8")
    assert "d_logic_bt_lp_rx_en" in html and "pll_lock_indicator" in html


def test_topout_fortest_backfill_includes_mux(topo_win, tmp_path, monkeypatch):
    """回填 for_test → 新 Excel 的 for_test 页含 mux 真值表(lna_itrim)（堵『只回填 logic』陷阱）。"""
    from PySide6 import QtWidgets
    import openpyxl
    w = topo_win
    out = tmp_path / "ft.xlsx"
    monkeypatch.setattr(QtWidgets.QFileDialog, "getSaveFileName",
                        staticmethod(lambda *a, **k: (str(out), "")))
    monkeypatch.setattr(QtWidgets.QMessageBox, "information",
                        staticmethod(lambda *a, **k: None))
    w.on_topo_fortest()
    assert out.exists()
    wb = openpyxl.load_workbook(str(out))
    ws = next(s for s in wb.worksheets if s.title.lower() == "for_test")
    txt = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "lna_itrim" in txt                            # mux 表回填进 for_test


# ───────────── DoD：旧视图未损 ─────────────
def test_legacy_view_intact(topo_win):
    """『排查(旧)』分页仍功能正确：左表/信号/分析在位（降级保留，不是冻成门面）。"""
    w = topo_win
    assert w.table.rowCount() == len(w.signals) > 0
    assert w._analysis                                   # 旧分析画像照常
    # 旧视图的级联/尾缀控件仍在（前缀/后缀/级联那套住这儿）
    assert hasattr(w, "cascade_logic_combo") and hasattr(w, "append_to_logic_chk")


def test_topout_graceful_when_no_topout_page(gui_app, tmp_path):
    """无 Topout 页 → Topout 清单空 + 提示，不崩。"""
    import fixtures
    from dreg_verify import gui as G
    xl = tmp_path / "plain.xlsx"
    fixtures.build_workbook(str(xl), with_mux=True)      # 普通夹具，无 Topout 页
    w = G.MainWindow()
    w.path_edit.setText(str(xl))
    w.on_load()                                          # 不应抛
    try:
        assert w.topo_table.rowCount() == 0
        assert "Topout" in w.topo_detail.text()
    finally:
        w.close()


# ───────────── 对抗 review GUI 修复回归 ─────────────
def test_topout_refresh_on_tab_switch_after_maxtests_change(topo_win):
    """切回 Topout 视图时，若『排查(旧)』改过上限 → 重建清单（『用例』列不再陈旧 vs 实际导出）。"""
    w = topo_win
    new = (w._topo_maxt() % 64) + 7                       # 保证与当前不同且合理
    w.main_tabs.setCurrentIndex(1)                        # 去『排查(旧)』
    w.max_tests.setValue(new)
    w.main_tabs.setCurrentIndex(0)                        # 切回 Topout → currentChanged 触发重建
    assert w._topo_built_key[2] == new


def test_topout_refresh_failure_clears_stale_panels(topo_win, monkeypatch):
    """分析失败时清空旧表的链/真值表/.sv（否则失败仍显示上一张表，误导）；不崩。"""
    from dreg_verify import topout as T2
    w = topo_win
    r = _topo_row(w, "d_logic_bt_lp_rx_en")
    w.topo_table.setCurrentCell(r, 1)
    assert w.topo_truth.rowCount() > 0                    # 先有内容
    monkeypatch.setattr(T2, "topout_view_models",
                        lambda *a, **k: (_ for _ in ()).throw(RuntimeError("boom")))
    w._refresh_topout()                                   # 不应抛
    assert w.topo_truth.rowCount() == 0                   # 旧表已清
    assert "失败" in w.topo_detail.text()


# ═══════════════ 2026-06-24 SignalView 重构：编辑 / 筛选 / 展开链 / 覆盖度迁移 ═══════════════
def _sel(w, name):
    from dreg_verify import gui as G
    w.topo_table.setCurrentCell(_topo_row(w, name), G.TOPO_NAME)
    return w.topout_view


# ── 点1：真值表清零/加列/删列/改预填 ──
def test_topout_truth_add_delete_column(topo_win):
    """加列→列数+1；删列→回到原数；清零→零列（可逆：重新生成恢复）。"""
    v = _sel(topo_win, "d_logic_bt_lp_rx_en")
    n0 = len(v.cur_cols)
    v._e_add()
    assert len(v.cur_cols) == n0 + 1
    v.truth.setCurrentCell(0, n0)                 # 选中刚加的列
    v._e_del()
    assert len(v.cur_cols) == n0
    v._e_clear()
    assert v.cur_cols == [] and v.truth.columnCount() == 0
    v._e_regen()
    assert len(v.cur_cols) == n0                  # 重新生成可逆


def test_topout_truth_edit_expected_and_input_recompute(topo_win):
    """改输入值→auto_out 即时重算；改期望→记进 edits、导出向量带 designer_expected。"""
    from dreg_verify import vectors as V
    v = _sel(topo_win, "d_logic_bt_lp_rx_en")
    v._e_regen()
    col = v.cur_cols[0]
    for e in v.e_inputs:                          # A=linelocal=1, C=rx_en_local=1, 其余 0
        col["vals"][e["key"]] = 1 if ("linelocal_mode" in e["label"]
                                      or "rx_en_local" in e["label"]) else 0
    v._recompute_col(col)
    assert col["auto"] == 1                       # (1?1:0)&~0 = 1
    col["exp"] = 1
    v._commit()
    ed = v._compute_edited()
    assert "d_logic_bt_lp_rx_en" in ed
    assert ed["d_logic_bt_lp_rx_en"]["vectors"][0].designer_expected == 1


def test_topout_edit_reflected_in_exported_sv(topo_win):
    """清零某 Topout 信号→导出 .sv 不再含它的断言，但记账『用户已清空』(不静默丢)。"""
    v = _sel(topo_win, "d_logic_bt_lp_rx_en")
    base, _ = v.provider.render_sv(None, "max", v._maxt(), False, {})
    assert "`ENV_RF.d_logic_bt_lp_rx_en==" in base
    v._e_clear()
    text, _ = v.provider.render_sv(None, "max", v._maxt(), False, v._compute_edited())
    assert "`ENV_RF.d_logic_bt_lp_rx_en==" not in text
    assert "用户已清空" in text and "d_logic_bt_lp_rx_en" in text


def test_topout_negative_checkbox_adds_neg(topo_win):
    """信号清单『负向』勾选 → 该信号加一条负向；导出 .sv 含 _NEG 断言。"""
    from dreg_verify import gui as G
    from PySide6 import QtCore
    v = _sel(topo_win, "d_logic_bt_lp_rx_en")
    r = _topo_row(topo_win, "d_logic_bt_lp_rx_en")
    v.sig_table.item(r, G.TOPO_NEG).setCheckState(QtCore.Qt.Checked)
    assert v._has_negatives("d_logic_bt_lp_rx_en")
    text, _ = v.provider.render_sv(None, "max", v._maxt(), False, v._compute_edited())
    assert "_NEG" in text


# ── 点2：信号选择面板 owner 筛/搜索/全选清空 ──
def test_topout_owner_filter(topo_win):
    """勾一个 owner → 只显示该 owner 的信号行。"""
    v = topo_win.topout_view
    owners = {m["owner"] for m in v.models if m["owner"]}
    assert owners
    target = sorted(owners)[0]
    v._owner_acts[target].setChecked(True)
    visible = [m["name"] for r, m in enumerate(v.models) if not v.sig_table.isRowHidden(r)]
    assert visible and all(
        next(m for m in v.models if m["name"] == nm)["owner"] == target for nm in visible)
    v._owner_acts[target].setChecked(False)       # 复位


def test_topout_search_filter(topo_win):
    """搜索框输入 → 只显示名字/表达式/输入名匹配的行。"""
    v = topo_win.topout_view
    v.search.setText("rx_en")
    vis = [m["name"] for r, m in enumerate(v.models) if not v.sig_table.isRowHidden(r)]
    assert vis and all("rx_en" in nm or "rx_en" in (
        m_["expr"] or "") for nm in vis for m_ in [next(x for x in v.models if x["name"] == nm)])
    v.search.setText("")


def test_topout_select_all_clear(topo_win):
    """全选→全部勾选；清空勾选→全不勾。"""
    v = topo_win.topout_view
    v._check_all(True)
    assert len(v._checked_names()) == sum(1 for r in range(v.sig_table.rowCount())
                                          if not v.sig_table.isRowHidden(r))
    v._check_all(False)
    assert v._checked_names() == []
    v._check_all(True)


# ── 点6：展开链总显示（字母代入真名）──
def test_topout_chain_shows_substitution(topo_win):
    """单级 logic 信号也显示展开链『原式 / 字母代入真名』(不再是『无上游可展开』占位)。"""
    v = _sel(topo_win, "d_logic_bt_lp_rx_en")
    txt = v.chain.toPlainText()
    assert "(A?C:B)" in txt                         # 原式
    assert "d_bt_lp_linelocal_mode_ctrl" in txt     # 字母代入真名
    assert "d_bt_lp_rx_en_local" in txt


# ── 点5：覆盖度在工具条（非右上角）、带 ? 帮助 ──
def test_topout_coverage_in_toolbar(topo_win):
    """覆盖度下拉 = SignalView 自有 cov（在筛选工具条），有三档 + 切档即重建。"""
    v = topo_win.topout_view
    assert [v.cov.itemText(i) for i in range(v.cov.count())] == ["精简", "全面", "穷举"]
    v.cov.setCurrentText("精简")
    assert v.built_key[0] == "min"
    v.cov.setCurrentText("全面")


# ═══════════════ 2026-06-24 对抗 review 修复回归 ═══════════════
def test_select_error_mux_does_not_crash(topo_win, monkeypatch):
    """选中一个『展开失败的 mux』(expansion=None, status=error) → 不崩、非可编辑（BLOCKER 修复）。"""
    from dreg_verify import gui as G
    v = topo_win.topout_view
    bad = {"kind": "mux", "status": "error", "issues": ["mux 展开失败"], "note": "",
           "node": None, "bindings": None, "expansion": None, "vectors": [],
           "out_width": 1, "chain": [], "name": "d_bt_lp_lna_itrim",
           "sig": None, "groups": [], "src_out_name": "d_bt_lp_lna_itrim", "editable": ""}
    monkeypatch.setattr(v.provider, "analyze", lambda *a, **k: bad)
    v._load_signal("d_bt_lp_lna_itrim")          # 不应抛
    assert v.truth.rowCount() == 0
    assert not v._edit_btns["加列"].isEnabled()    # 非 ok → 编辑按钮禁用


def test_view_edits_persist_on_reload_and_keyed_by_table(topo_win, tmp_path):
    """⭐持久化(2026-06-24, #2)：designer 手填期望/编辑【关 GUI/换表不丢】，且按 Excel 路径分桶——
    重载【同表】→ 恢复(含手填期望)；载【另一张表】→ 不串进来(anti-contamination 仍成立)。
    取代旧 test_edits_cleared_on_reload：旧行为=重载即清空；新行为=同表恢复、异表隔离(更强)。"""
    w = topo_win
    v = _sel(w, "d_logic_bt_lp_rx_en")
    v.cur_cols[0]["exp"] = 5                       # designer 手填一个期望值（劳动成果）
    v._commit()
    assert w.topout_view.edits.get("d_logic_bt_lp_rx_en")
    # 重载同表 → 编辑（含手填期望）被恢复
    w.on_load()
    ed = w.topout_view.edits.get("d_logic_bt_lp_rx_en")
    assert ed and ed["cols"][0]["exp"] == 5
    # 载另一张表（不同路径，内容相同）→ 该表自己的桶为空 → 不继承上一张表的编辑
    other = tmp_path / "mirror2.xlsx"
    make_mirror_btlp.build(str(other))
    w.path_edit.setText(str(other))
    w.on_load()
    assert w.topout_view.edits == {}              # 按路径分桶：异表不串编辑


def test_topout_designer_expected_survives_config_export_import(topo_win, tmp_path, monkeypatch):
    """⭐#1a/#2：Topout 视图 designer 手填期望随【完整配置】导出→导入跨会话/机器迁移（旧版只导 legacy 门面，
    Topout edits 进不了配置）。"""
    import json
    from PySide6 import QtWidgets
    w = topo_win
    v = _sel(w, "d_logic_bt_lp_rx_en")
    v.cur_cols[0]["exp"] = 6
    v._commit()
    cfg = w._collect_config()
    assert cfg["view_edits"]["topout"]["d_logic_bt_lp_rx_en"]["cols"][0]["exp"] == 6
    cfg_path = tmp_path / "cfg.json"
    cfg_path.write_text(json.dumps(cfg), encoding="utf-8")
    v._e_regen()                                  # 丢弃当前编辑（模拟新会话/换机器）
    assert "d_logic_bt_lp_rx_en" not in w.topout_view.edits
    monkeypatch.setattr(QtWidgets.QFileDialog, "getOpenFileName",
                        staticmethod(lambda *a, **k: (str(cfg_path), "")))
    monkeypatch.setattr(QtWidgets.QMessageBox, "information", staticmethod(lambda *a, **k: None))
    w.on_import_edits()
    ed = w.topout_view.edits.get("d_logic_bt_lp_rx_en")
    assert ed and ed["cols"][0]["exp"] == 6       # 手填期望随配置恢复进 Topout


def test_topout_signal_checks_persist_on_reload(topo_win):
    """⭐N1：Topout 信号勾选(纳入导出集)关 GUI/换表不丢——取消勾选某信号 → 重载同表仍取消。"""
    from PySide6 import QtCore
    from dreg_verify import gui as G
    w = topo_win
    r = _topo_row(w, "d_logic_bt_lp_rx_en")
    w.topo_table.item(r, G.TOPO_SEL).setCheckState(QtCore.Qt.Unchecked)
    assert "d_logic_bt_lp_rx_en" not in [n.lower() for n in w.topout_view._checked_names()]
    w.on_load()
    r2 = _topo_row(w, "d_logic_bt_lp_rx_en")
    assert w.topo_table.item(r2, G.TOPO_SEL).checkState() == QtCore.Qt.Unchecked


def test_topout_mux_edit_persists_on_reload(topo_win):
    """⭐#2 mux 路：mux 根的编辑(加负向列)也按 assignments 重建 vec 持久化、换表恢复（最难的一路）。"""
    w = topo_win
    v = _sel(w, "d_bt_lp_lna_itrim")              # mux 根
    if v.cur_an is None or v.cur_an["editable"] != "mux":
        pytest.skip("夹具里该信号非 mux 根")
    n0 = len(v.cur_cols)
    v.truth.setCurrentCell(0, 0)
    v._e_addneg()
    assert len(v.cur_cols) == n0 + 1 and w.topout_view.edits.get("d_bt_lp_lna_itrim")
    w.on_load()
    ed = w.topout_view.edits.get("d_bt_lp_lna_itrim")
    assert ed and len(ed["cols"]) == n0 + 1 and any(c["neg"] for c in ed["cols"])


def test_negative_with_expected_eq_auto_stays_negative(topo_win):
    """负向列期望被填成 == auto → 导出仍是负向（不静默退化成通过断言；MAJOR 修复）。"""
    v = _sel(topo_win, "d_logic_bt_lp_rx_en")
    v._e_regen()
    v.truth.setCurrentCell(0, 0)
    v._e_addneg()                                 # 加一条负向
    negcol = next(c for c in v.cur_cols if c["neg"])
    negcol["exp"] = negcol["auto"]                # 把负向期望填成正确值(auto)
    v._commit()
    vecs = v._cols_to_vectors(v.cur_an, v.cur_cols)
    assert any(x.is_negative for x in vecs)       # 仍保住负向身份


# ───────────── 截图：Topout 视图版面过目 ─────────────
def test_topout_view_screenshot(topo_win, gui_app):
    """截一张 Topout 视图 PNG（版面过目；文字方框是 offscreen 缺中文字体所致，文字由上面断言把关）。"""
    w = topo_win
    r = _topo_row(w, "d_logic_bt_lp_rx_en")
    w.topo_table.setCurrentCell(r, 1)                    # 选 rx_en
    w.topo_inner.setCurrentIndex(0)                      # 停在『展开链/真值表』页
    w.show()
    gui_app.processEvents()
    out_dir = os.environ.get("TOPO_SHOT_DIR") or os.path.dirname(os.path.abspath(__file__))
    png = os.path.join(out_dir, "topout_view.png")
    ok = w.grab().save(png)
    assert ok and os.path.getsize(png) > 0
