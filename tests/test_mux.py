# -*- coding: utf-8 -*-
"""mux 页验证功能测试（2026-06-03 第九轮）。

覆盖（按实现顺序逐层补充）：
  ① 数据层: excel_model.read_mux / strip_to_mux / MuxGroup/MuxCase / wb.mux 向后兼容
  ② case 字面量: expr.parse_case_literal / expand_case_values / case_matches（don't-care 位）
  后续: ③ resolver 数据寄存器/控制双路径 ④ make_mux_vectors ⑤ generator/sv_writer 端到端
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import fixtures                                     # noqa: E402
from dreg_verify import excel_model                 # noqa: E402
from dreg_verify import expr as E                   # noqa: E402
from dreg_verify import generator                   # noqa: E402
from dreg_verify import mux_gen                     # noqa: E402
from dreg_verify import resolver as R               # noqa: E402


# ───────────── fixtures ─────────────
@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    path = tmp_path_factory.mktemp("mux") / "synthetic_mux.xlsx"
    fixtures.build_workbook(str(path), with_mux=True)
    return excel_model.load_workbook(str(path))


@pytest.fixture(scope="module")
def wb_no_mux(tmp_path_factory):
    path = tmp_path_factory.mktemp("nomux") / "synthetic.xlsx"
    fixtures.build_workbook(str(path))
    return excel_model.load_workbook(str(path))


# ───────────── ① strip_to_mux ─────────────
def test_strip_to_mux():
    assert excel_model.strip_to_mux("d_x_to_mux") == "d_x"
    assert excel_model.strip_to_mux("d_x_TO_MUX") == "d_x"
    assert excel_model.strip_to_mux("d_x_to_logic") == "d_x_to_logic"   # 不剥 to_logic
    assert excel_model.strip_to_mux("d_x") == "d_x"
    # 反向：strip_to_logic 也不剥 _to_mux（两个后缀语义不同）
    assert excel_model.strip_to_logic("d_x_to_mux") == "d_x_to_mux"


# ───────────── ① read_mux ─────────────
def test_read_mux_groups(wb):
    assert len(wb.mux) == 2
    g1, g2 = wb.mux
    # 组1: rccal_i, 3 个 case, 控制=lna_agc
    assert g1.group_no == 1
    assert g1.out_name == "d_bt_lp_rccal_i[3:0]"
    assert g1.out_base == "d_bt_lp_rccal_i"
    assert g1.out_width == 4
    assert g1.ctrl_base == "d_logic_bt_lp_lna_agc"      # 剥 _to_mux + 位宽
    assert g1.ctrl_width == 3
    assert g1.owner == "Alice"
    assert g1.is_top
    assert [c.case_raw for c in g1.cases] == ["3'b010", "3'b011", "3'b10x"]
    assert g1.cases[0].input_base == "d_bt_lp_rccal_i_g1"   # 剥 _to_mux + 位宽
    assert g1.cases[0].input_width == 4
    # 组2
    assert g2.group_no == 2
    assert g2.out_base == "d_bt_lp_bias_q"
    assert g2.owner == "Bob"
    assert len(g2.cases) == 2


def test_read_mux_assert_id_and_rtl(wb):
    g1 = wb.mux[0]
    # 方案A: assert_mux<N>_T<n>，mux 前缀防与 logic R 号撞号
    assert g1.assert_id == "mux1"
    assert wb.mux[1].assert_id == "mux2"
    # mux 输出 = 顶层网名直接探（环境验证实证），不走 logic 的 _ls/_to_logic 变换
    assert g1.rtl_base == "d_bt_lp_rccal_i"
    assert g1.rtl_name == "d_bt_lp_rccal_i[3:0]"


def test_read_mux_filters_reserved(wb):
    # (reserved) 行 F 列为空 → 被过滤，且不会把组一分为二
    for g in wb.mux:
        for c in g.cases:
            assert "reserved" not in c.input_base.lower()


def test_no_mux_sheet_backward_compat(wb_no_mux):
    # 无 mux 页: wb.mux=[] 且 logic 流程完全不受影响（现有 236 个测试的前提）
    assert wb_no_mux.mux == []
    assert len(wb_no_mux.logic) == 3


def test_mux_data_registers_resolvable_in_tmm(wb):
    # mux 数据输入剥 _to_mux 后能在 tmm 查到地址（RF_WRITE 互异值的前提）
    g1 = wb.mux[0]
    for c in g1.cases:
        assert c.input_base in wb.tmm, "%s 应能在 tmm 查到" % c.input_base
        assert wb.tmm[c.input_base].reg_type == "RW"
    # g1/g2 同地址不同字段（镜像真表 t1..t4 同住一个 16bit 寄存器）→ 互异值分配要警惕合并裁剪
    assert wb.tmm["d_bt_lp_rccal_i_g1"].address == wb.tmm["d_bt_lp_rccal_i_g2"].address
    assert wb.tmm["d_bt_lp_rccal_i_g1"].bit_lsb != wb.tmm["d_bt_lp_rccal_i_g2"].bit_lsb


def test_mux_ctrl_links_to_logic_to_mux_row(wb):
    # 衔接点: 控制信号基名 == logic 页 M=to_mux 行的 K 基名
    g1 = wb.mux[0]
    logic_by_base = {s.out_base.lower(): s for s in wb.logic}
    assert g1.ctrl_base.lower() in logic_by_base
    ctrl_row = logic_by_base[g1.ctrl_base.lower()]
    assert ctrl_row.suffix == "to_mux"
    assert ctrl_row.is_top


# ───────────── ② parse_case_literal ─────────────
def test_parse_case_literal_plain():
    assert E.parse_case_literal("3'b010") == (0b010, 3, 0)
    assert E.parse_case_literal("3'b111") == (0b111, 3, 0)


def test_parse_case_literal_dontcare():
    assert E.parse_case_literal("4'b000x") == (0b0000, 4, 0b0001)
    assert E.parse_case_literal("4'b1x0x") == (0b1000, 4, 0b0101)
    assert E.parse_case_literal("3'b10x") == (0b100, 3, 0b001)
    # 真表 tsensor 形态全集 4'b000x..4'b111x
    for i in range(8):
        v, w, dc = E.parse_case_literal("4'b%s%s%sx" % ((i >> 2) & 1, (i >> 1) & 1, i & 1))
        assert (v >> 1, w, dc) == (i, 4, 1)


def test_parse_case_literal_hex_no_x():
    assert E.parse_case_literal("4'hA") == (0xA, 4, 0)


def test_parse_case_literal_errors():
    with pytest.raises(E.ExprError):
        E.parse_case_literal("not_a_literal")
    with pytest.raises(E.ExprError):
        E.parse_case_literal("4'hAx")       # 非二进制含 x
    with pytest.raises(E.ExprError):
        E.parse_case_literal("4'sb0000")    # 有符号


def test_parse_based_literal_unchanged():
    # 现有 parse_based_literal 行为不变（x 当 0）—— logic 路径不受任何影响
    assert E.parse_based_literal("4'b000x") == (0, 4)


# ───────────── ② expand_case_values / case_matches ─────────────
def test_expand_case_values():
    assert E.expand_case_values(0b010, 3, 0) == [0b010]                          # 无 x
    assert E.expand_case_values(0b0000, 4, 0b0001) == [0b0000, 0b0001]           # 1 个 x
    assert E.expand_case_values(0b1000, 4, 0b0101) == [0b1000, 0b1001, 0b1100, 0b1101]  # 2 个 x


def test_case_matches():
    # 4'b000x: 命中 0/1, 不命中 2
    assert E.case_matches(0b0000, 4, 0b0001, 0b0000)
    assert E.case_matches(0b0000, 4, 0b0001, 0b0001)
    assert not E.case_matches(0b0000, 4, 0b0001, 0b0010)
    # 精确 case
    assert E.case_matches(0b010, 3, 0, 0b010)
    assert not E.case_matches(0b010, 3, 0, 0b011)


def test_designer_sv_decoded_case_mapping():
    """复盘 designer .sv 的 assert_151 解码: tsensor=0..15 → t1..t8（每 case 命中 2 个值）。"""
    cases = ["4'b000x", "4'b001x", "4'b010x", "4'b011x",
             "4'b100x", "4'b101x", "4'b110x", "4'b111x"]
    for ctrl in range(16):
        hits = [i for i, c in enumerate(cases)
                if E.case_matches(*E.parse_case_literal(c), ctrl_value=ctrl)]
        assert hits == [ctrl >> 1], "tsensor=%d 应只命中 t%d" % (ctrl, (ctrl >> 1) + 1)


# ───────────── ③ 控制路径发现 + 驱动解析 ─────────────
@pytest.fixture(scope="module")
def expansion(wb):
    """组1 (rccal_i) 的完整解析结果。"""
    resolver = R.Resolver(wb)
    return mux_gen.expand_mux_group(wb, resolver, wb.mux[0])


def test_discover_ctrl_paths(wb):
    """A?C:B 透传发现: B(RO)=line 路径(A=0), C(RW)=local 路径(A=1)。"""
    resolver = R.Resolver(wb)
    ctrl_sig = next(s for s in wb.logic if s.out_base == "d_logic_bt_lp_lna_agc")
    node = E.parse(ctrl_sig.expr)
    bindings = resolver.resolve_signal_inputs(ctrl_sig)
    paths = mux_gen.discover_ctrl_paths(node, bindings)
    by_var = {p["var"]: p for p in paths}
    assert by_var["B"]["kind"] == "RO" and by_var["B"]["ctrl_assign"] == {"A": 0}   # line
    assert by_var["C"]["kind"] == "RW" and by_var["C"]["ctrl_assign"] == {"A": 1}   # local


def test_expand_mux_group(expansion):
    assert expansion["issues"] == []
    # 控制行 3 个输入 + 3 个数据寄存器
    assert set(expansion["used_vars"]) == {"c:A", "c:B", "c:C", "d:0", "d:1", "d:2"}
    # line=force 线控(RO), local=RF_WRITE 本地(RW)
    assert expansion["line"]["kind"] == "RO" and expansion["line"]["key"] == "c:B"
    assert expansion["local"]["kind"] == "RW" and expansion["local"]["key"] == "c:C"
    # 数据寄存器全部解析为 RW + 地址
    for k in expansion["data_keys"]:
        b = expansion["bindings"][k]
        assert b.kind == "RW" and b.address is not None
    # case 解析: 第三个带 don't-care
    assert expansion["parsed_cases"][0] == (0b010, 3, 0)
    assert expansion["parsed_cases"][2] == (0b100, 3, 0b001)


def test_alloc_distinct_values(wb):
    values, collision = mux_gen.alloc_distinct_values(wb.mux[0])
    assert not collision
    assert len(set(values)) == 3 and 0 not in values        # 互异且避开 0
    # 4-bit 寄存器 → designer 风格从 0xA 递减
    assert values == [0xA, 0x9, 0x8]


# ───────────── ④ 向量生成 ─────────────
def test_make_mux_vectors_min(wb, expansion):
    """精简: 每 case 一个值(x=0) + 1 个 local 路径测试 = 3+1。"""
    vecs, meta = mux_gen.make_mux_vectors(wb.mux[0], expansion, mode="min")
    assert len(vecs) == 4
    assert meta["scan_path"] == "line"
    # T0: case 3'b010 → ctrl 走 line 路径(c:A=0, c:B=0b010), 期望=第1个寄存器的互异值 0xA
    t0 = vecs[0]
    assert t0.assignments["c:A"] == 0 and t0.assignments["c:B"] == 0b010
    assert t0.assignments["c:C"] == (~0b010) & 0b111          # 非主动路径=干扰值(取反)
    assert t0.assignments["d:0"] == 0xA and t0.assignments["d:1"] == 0x9
    assert t0.exp_value == 0xA
    # T2: case 3'b10x 精简取 x=0 → ctrl=0b100, 期望=第3个寄存器 0x8
    assert vecs[2].assignments["c:B"] == 0b100 and vecs[2].exp_value == 0x8
    # T3: local 路径(c:A=1, 控制值从 c:C 进), 期望=第1个寄存器
    t3 = vecs[3]
    assert t3.assignments["c:A"] == 1 and t3.assignments["c:C"] == 0b010
    assert t3.exp_value == 0xA and "local" in t3.note


def test_make_mux_vectors_max(wb, expansion):
    """全面(2026-06-03 第十一轮): x 位展开(4) + 反码数据轮(3) + 1 个 local 抽测 = 8。"""
    vecs, meta = mux_gen.make_mux_vectors(wb.mux[0], expansion, mode="max")
    assert len(vecs) == 8
    ctrl_values = [v.assignments["c:B"] for v in vecs[:4]]
    assert ctrl_values == [0b010, 0b011, 0b100, 0b101]
    # x=0/x=1 两个值期望相同(都选中第 3 个寄存器)
    assert vecs[2].exp_value == vecs[3].exp_value == 0x8
    # 反码数据轮(T4..T6): 数据寄存器写反码互异值, 期望跟着变
    inv = [v for v in vecs if "inverted data" in (v.note or "")]
    assert len(inv) == 3
    # 第一轮 0xA/0x9/0x8 → 反码 0x5/0x6/0x7（有效位宽 4bit 内取反）
    assert inv[0].assignments["d:0"] == 0x5 and inv[0].exp_value == 0x5
    assert inv[1].assignments["d:1"] == 0x6 and inv[1].exp_value == 0x6
    assert inv[2].assignments["d:2"] == 0x7 and inv[2].exp_value == 0x7
    # 每个数据寄存器: 两轮值合起来每一位都翻转过(抓数据通路位坏死的前提)
    for i, k in enumerate(["d:0", "d:1", "d:2"]):
        r1, r2 = vecs[0].assignments[k], inv[0 if i == 0 else i].assignments[k]
        assert (r1 ^ r2) == 0xF, "寄存器 %s 两轮值 0x%X/0x%X 必须每位互补" % (k, r1, r2)
    assert meta["data_rounds"] == 2 and meta["other_path_scan"] == "probe"


def test_make_mux_vectors_exhaustive(wb, expansion):
    """穷举(2026-06-03 第十一轮): 全面的全部 + local 路径全扫(取代单条抽测)。"""
    vecs, meta = mux_gen.make_mux_vectors(wb.mux[0], expansion, mode="exhaustive")
    # line 扫描 4 + 反码轮 3 + local 全扫 4 = 11
    assert len(vecs) == 11
    assert meta["other_path_scan"] == "full"
    # local 全扫: 控制值从 c:C(local 寄存器)进, 模式位 c:A=1, 扫全部 case 值
    full = [v for v in vecs if "full scan" in (v.note or "")]
    assert len(full) == 4
    assert [v.assignments["c:C"] for v in full] == [0b010, 0b011, 0b100, 0b101]
    assert all(v.assignments["c:A"] == 1 for v in full)
    # 期望与 line 扫描一致(选中同一个寄存器)
    assert [v.exp_value for v in full] == [0xA, 0x9, 0x8, 0x8]


def test_coverage_mode_mapping():
    """覆盖度三档映射(generator/GUI 共用): 精简→min, 全面→max, 穷举→exhaustive。"""
    assert mux_gen.coverage_mode("min", False) == "min"
    assert mux_gen.coverage_mode("max", False) == "max"
    assert mux_gen.coverage_mode("min", True) == "exhaustive"
    assert mux_gen.coverage_mode("max", True) == "exhaustive"


def test_alloc_inverted_values(wb):
    """反码互异值: 取反后仍互异、避 0；与第一轮合起来每位都翻转过。"""
    grp = wb.mux[0]
    base, c1 = mux_gen.alloc_distinct_values(grp)
    inv, c2 = mux_gen.alloc_inverted_values(grp, base)
    assert not c1 and not c2
    assert len(set(inv)) == len(inv) and 0 not in inv
    assert inv == [0x5, 0x6, 0x7]                       # ~0xA/~0x9/~0x8 & 0xF


def test_three_levels_differ_without_x_bits(wb):
    """⭐ 用户问题回归: case 值没有 x 位的 mux 组, 三档测试数也必须不同。"""
    resolver = R.Resolver(wb)
    grp = wb.mux[1]                                     # bias_q: 2 个 case, 无 x 位
    exp = mux_gen.expand_mux_group(wb, resolver, grp)
    n = {m: len(mux_gen.make_mux_vectors(grp, exp, mode=m)[0])
         for m in ("min", "max", "exhaustive")}
    assert n["min"] < n["max"] < n["exhaustive"], "三档必须递增: %s" % n
    assert n["min"] == 3                                # 2 case + 1 local 抽测
    assert n["max"] == 5                                # + 2 反码轮
    assert n["exhaustive"] == 6                         # + 2 local 全扫 - 1 抽测(被全扫取代)


# ───────────── ⑤ generator/sv_writer 端到端 ─────────────
def test_build_with_mux(wb):
    """logic + mux 同文件混排(用户拍板): 3 logic + 2 mux = 5 blocks。"""
    res = generator.build(wb, generator.GenOptions())
    assert res["summary"]["n_mux_groups"] == 2
    assert res["summary"]["n_mux_generated"] == 2
    assert res["summary"]["n_generated"] == 5                # 3 logic + 2 mux
    assert res["summary"]["n_dup_labels"] == 0               # mux 前缀防撞号
    mux_stats = [st for _, st in res["blocks"] if st.get("is_mux")]
    assert len(mux_stats) == 2
    assert mux_stats[0]["assert_id"] == "mux1"
    assert not res["skipped"]


def test_build_mux_sv_content(wb):
    """生成的 .sv 内容正确性（对照 designer 配方逐项检查）。"""
    res = generator.build(wb, generator.GenOptions(signals=["d_bt_lp_rccal_i"]))
    text = generator.render(res)
    # ① assert 标签 = assert_mux<N>_T<n>（方案 A）
    assert "assert_mux1_T0:" in text and "assert_mux1_T3:" in text
    # ② 输出探针 = G 列名直探(顶层网名, 带位宽切片), 无 _ls/_to_logic 后缀
    assert "assert (`ENV_RF.d_bt_lp_rccal_i[3:0]==" in text
    assert "d_bt_lp_rccal_i_ls" not in text
    # ③ 控制线控输入 force 基名(designer 实证)
    assert "force `ENV_RF.d_bt_lp_lna_agc_line[2:0]=" in text
    # ④ 数据寄存器 RF_WRITE: g1(0xA)/g2(0x9) 同地址 h40 合并 → 16'h9A
    assert "`RF_WRITE(10'h40,16'h9A);" in text
    assert "`RF_WRITE(10'h41,16'h8);" in text                # g3 单独地址
    # ⑤ 模式位寄存器 RF_WRITE(line 路径 A=0)
    assert "`RF_WRITE(10'h31,16'h0);" in text
    # ⑥ 期望 = 被选中寄存器的互异值
    assert "==4'b1010)" in text                              # T0 选 g1=0xA
    assert "==4'b1000)" in text                              # T2 选 g3=0x8
    # ⑦ 驱动顺序: RF_WRITE 在 force 之前(与 logic 规则一致)
    first_force = text.index("force `ENV_RF")
    first_write = text.index("`RF_WRITE(")
    assert first_write < first_force


def test_build_mux_local_path_test(wb):
    """local 路径测试: 模式位=1 + 本地寄存器带控制值。"""
    res = generator.build(wb, generator.GenOptions(signals=["d_bt_lp_rccal_i"]))
    text = generator.render(res)
    # T3(local): c:A=1 → RF_WRITE(10'h31,16'h1); c:C=0b010 → RF_WRITE(10'h32,16'h2)
    t3_start = text.index("assert_mux1_T3:")
    t3_drives = text[text.rindex("`RF_WRITE(10'h31,", 0, t3_start):t3_start]
    assert "`RF_WRITE(10'h31,16'h1);" in t3_drives           # 模式位 → local
    assert "`RF_WRITE(10'h32,16'h2);" in t3_drives           # 本地寄存器 = 控制值 0b010


def test_build_mux_negative(wb):
    """反例: 自检式==错值(取反), 标签带 _NEG, 错值=互异值取反。"""
    res = generator.build(wb, generator.GenOptions(signals=["d_bt_lp_rccal_i"], neg_all=True))
    text = generator.render(res)
    assert "assert_mux1_T4_NEG:" in text
    # T0 期望 0xA=4'b1010 → 反例错值 = ~0xA & 0xF = 0x5 = 4'b0101
    assert "==4'b0101)" in text
    assert "NEG-EXPECTED-FAIL" in text


def test_build_gen_mux_off(wb):
    """gen_mux=False: 只出 logic, 与旧版产物一致。"""
    res = generator.build(wb, generator.GenOptions(gen_mux=False))
    assert res["summary"]["n_mux_generated"] == 0
    assert res["summary"]["n_generated"] == 3                # 仅 logic
    assert not any(st.get("is_mux") for _, st in res["blocks"])


def test_build_no_mux_sheet_unchanged(wb_no_mux):
    """无 mux 页的 Excel: 行为与第八轮完全一致(向后兼容的核心保证)。"""
    res = generator.build(wb_no_mux, generator.GenOptions())
    assert res["summary"]["n_mux_groups"] == 0
    assert res["summary"]["n_mux_generated"] == 0
    assert res["summary"]["n_generated"] == 3


def test_mux_type_filter(wb):
    """--type mux 只筛 mux 组; --type ls 不含 mux。"""
    res = generator.build(wb, generator.GenOptions(types=["mux"]))
    assert res["summary"]["n_generated"] == res["summary"]["n_mux_generated"] == 2
    res2 = generator.build(wb, generator.GenOptions(types=["ls"]))
    assert res2["summary"]["n_mux_generated"] == 0


def test_mux_owner_filter(wb):
    """owner 过滤对 mux 生效(组1=Alice, 组2=Bob)。"""
    res = generator.build(wb, generator.GenOptions(owners=["Bob"]))
    mux_stats = [st for _, st in res["blocks"] if st.get("is_mux")]
    assert len(mux_stats) == 1 and mux_stats[0]["out_name"] == "d_bt_lp_bias_q[1:0]"


def test_mux_sv_summary_counts(wb):
    """汇总块: mux 块的向量数计入 asserts 总数。"""
    res = generator.build(wb, generator.GenOptions(sv_summary=True))
    text = generator.render(res)
    assert "begin : dreg_rf_test" in text
    # 3 logic 信号 + 2 mux 组 = 5 signals
    assert "signals:5" in text


# ───────────── 对抗式审查修复回归（2026-06-03，6 个确认问题）─────────────
def _mini_mux_group(case_specs, input_width=4):
    """造一个最小 MuxGroup（不经 Excel），case_specs=[(case_raw, input_base), ...]。"""
    cases = [excel_model.MuxCase(row=i + 3, case_raw=cr, input_raw=ib + "_to_mux",
                                 input_base=ib, input_width=input_width,
                                 input_msb=input_width - 1, input_lsb=0)
             for i, (cr, ib) in enumerate(case_specs)]
    return excel_model.MuxGroup(group_no=1, out_name="d_x_out[%d:0]" % (input_width - 1),
                                out_width=input_width, ctrl_raw="d_logic_ctrl_to_mux[3:0]",
                                ctrl_base="d_logic_ctrl", ctrl_width=4,
                                owner="", top_output=1, cases=cases)


def test_fix_A_narrow_field_marker_method(wb):
    """[审查A→点名法] 互异值装不下(原假绿防护) 不再跳过，改『点名法』生成（被测=标记/其余=0）。
    窄位宽组从"跳过"变"真能验的测试"。"""
    # 1-bit 数据寄存器 × 3 个 case → 互异值装不下
    grp = _mini_mux_group([("4'b0000", "d_n1"), ("4'b0001", "d_n2"), ("4'b0010", "d_n3")],
                          input_width=1)
    values, collision = mux_gen.alloc_distinct_values(grp)
    assert collision                                   # 互异值确实装不下
    # 经 build 流程：碰撞组改点名法生成，不再落进 skipped
    import copy
    wb2 = copy.copy(wb)
    wb2.mux = [grp]
    res = generator.build(wb2, generator.GenOptions(types=["mux"], include_risky=True))
    assert res["summary"]["n_mux_generated"] == 1      # 点名法 → 生成
    assert not any("假绿" in str(reasons) for _n, _a, reasons in res["skipped"])


def test_fix_B_effective_width_allocation(wb):
    """[审查B] 互异值按有效位宽分配 = min(声明位宽, tmm 字段位宽)，截断后仍互异。"""
    resolver = R.Resolver(wb)
    grp = wb.mux[1]                                     # bias_q: 声明 2bit, tmm 字段 2bit
    exp = mux_gen.expand_mux_group(wb, resolver, grp)
    values, collision = mux_gen.alloc_distinct_values(grp, exp["bindings"], exp["data_keys"])
    assert not collision and len(set(values)) == len(values)
    # 所有值在字段位宽内（写入不会被截断改变）
    for i, v in enumerate(values):
        b = exp["bindings"][exp["data_keys"][i]]
        fw = b.reg_msb - b.reg_lsb + 1
        assert v == (v & E.mask(fw)), "互异值必须在字段位宽内"


def test_fix_C_case_wider_than_ctrl_dropped(wb):
    """[审查C] case 比控制信号宽 → 截断后命中错 case 的向量被丢弃(不是静默生成误报)。"""
    resolver = R.Resolver(wb)
    grp = wb.mux[0]
    exp = mux_gen.expand_mux_group(wb, resolver, grp)
    # 伪造一个比控制(3bit)宽的 case：4'b1010 截断成 3'b010 会命中 case0 而不是它自己
    exp2 = dict(exp)
    exp2["parsed_cases"] = list(exp["parsed_cases"])
    exp2["parsed_cases"][1] = E.parse_case_literal("4'b1010")    # 替换 case1
    vecs, meta = mux_gen.make_mux_vectors(grp, exp2, mode="min")
    # case1 的向量被丢弃（截断后不命中），其余照常 + local 测试
    assert meta["dropped"] >= 1 and meta["truncated"]
    assert all("4'b1010" not in (v.note or "") for v in vecs)


def test_fix_E_noncontiguous_rows_merged(tmp_path):
    """[审查E] 非连续的同 G 输出行 → 归并进同一组（不劈分 → 不撞号 → 互异值不重启）。"""
    import openpyxl
    path = tmp_path / "noncontig.xlsx"
    wbk = openpyxl.Workbook()
    ws = wbk.active
    ws.title = "mux"
    ws.append([""] * 14)
    ws.append(["mux_input", "ctrl", "", "", "", "case", "mux_out", "", "top", "", "", "Owner", "", "N"])
    rows = [
        ("d_a_g1_to_mux[3:0]", "3'b000", "d_a[3:0]", 1),
        ("d_b_t1_to_mux[3:0]", "3'b000", "d_b[3:0]", 2),     # 中间插另一组
        ("d_a_g2_to_mux[3:0]", "3'b001", "d_a[3:0]", 1),     # 同 G 不连续
    ]
    for inp, case, out, n in rows:
        ws.append([inp, "d_logic_c_to_mux[2:0]", "", "", "", case, out, "", 1, "", "", "", "", n])
    wbk.save(str(path))
    wbr = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    groups = excel_model.read_mux(wbr["mux"])
    wbr.close()
    assert len(groups) == 2                              # 不是 3
    by_base = {g.out_base: g for g in groups}
    assert len(by_base["d_a"].cases) == 2                # 非连续行归并
    # 组号唯一
    assert len({g.group_no for g in groups}) == 2


def test_fix_F_rtl_name_with_space(tmp_path):
    """[审查F] G 列名含内部空格 → 探针 LHS 用解析的 msb/lsb 重建，不带空格。"""
    grp = excel_model.MuxGroup(group_no=1, out_name="d_foo [1:0]", out_width=2,
                               ctrl_raw="", ctrl_base="", ctrl_width=1,
                               owner="", top_output=1, cases=[])
    assert grp.rtl_name == "d_foo[1:0]"                  # 无空格
    # 无位宽的输出
    grp2 = excel_model.MuxGroup(group_no=2, out_name="d_bar", out_width=1,
                                ctrl_raw="", ctrl_base="", ctrl_width=1,
                                owner="", top_output=1, cases=[])
    assert grp2.rtl_name == "d_bar"


def test_fix_D_report_has_mux(wb):
    """[审查D-critical] report() 必须与 build() 双轨同步：mux 组出现在 summary/tables/可验证性。"""
    opts = generator.GenOptions()
    rep = generator.report(wb, opts)
    res = generator.build(wb, opts)
    # summary 行数 = logic + mux
    mux_rows = [r for r in rep["summary"] if r["type"] == "mux"]
    assert len(mux_rows) == res["summary"]["n_mux_generated"] == 2
    # 报告的 T 数与 .sv 一致（双轨同步的核心）
    mux_stats = {st["out_name"]: st for _, st in res["blocks"] if st.get("is_mux")}
    for r in mux_rows:
        assert r["n_tests"] == mux_stats[r["signal"]]["n_vectors"]
        assert r["error"] == ""
    # tables 段有 kind='mux' 的 case 选择表
    mux_tables = [t for t in rep["tables"] if t.get("kind") == "mux"]
    assert len(mux_tables) == 2
    assert "case(" in mux_tables[0]["expr"]
    # 可验证性表含 mux 行且为 clean
    mux_verif = [v for v in rep["verifiability"]["signals"] if v["type"] == "mux"]
    assert len(mux_verif) == 2 and all(v["status"] == "clean" for v in mux_verif)


def test_fix_D_report_mux_skip_reason_visible(wb):
    """[审查D] 跳过的 mux 组在报告里以 error 原因呈现（不是消失）。"""
    import copy
    grp = _mini_mux_group([("4'b0000", "d_n1"), ("4'b0001", "d_n2")], input_width=1)
    wb2 = copy.copy(wb)
    wb2.mux = [grp]
    rep = generator.report(wb2, generator.GenOptions(types=["mux"]))
    mux_rows = [r for r in rep["summary"] if r["type"] == "mux"]
    assert len(mux_rows) == 1
    assert mux_rows[0]["error"] != "" and mux_rows[0]["n_tests"] == 0


# ───────────── ⑥ CLI ─────────────
def test_cli_mux_flags(wb, tmp_path):
    """CLI: --no-mux / --mux-only / 默认都生成（用户拍板）。"""
    from dreg_verify import cli
    excel = tmp_path / "mux_cli.xlsx"
    fixtures.build_workbook(str(excel), with_mux=True)

    # 默认: logic+mux 都生成
    out1 = tmp_path / "all.sv"
    cli.main(["--excel", str(excel), "--out", str(out1)])
    text1 = out1.read_text(encoding="utf-8")
    assert "assert_mux1_T0:" in text1 and "assert_108_T0:" in text1

    # --no-mux: 仅 logic
    out2 = tmp_path / "logic_only.sv"
    cli.main(["--excel", str(excel), "--out", str(out2), "--no-mux"])
    text2 = out2.read_text(encoding="utf-8")
    assert "assert_mux" not in text2 and "assert_108_T0:" in text2

    # --mux-only: 仅 mux
    out3 = tmp_path / "mux_only.sv"
    cli.main(["--excel", str(excel), "--out", str(out3), "--mux-only"])
    text3 = out3.read_text(encoding="utf-8")
    assert "assert_mux1_T0:" in text3 and "assert_108_T0:" not in text3


def test_cli_html_report_contains_mux(wb, tmp_path):
    """CLI HTML 报告: mux 组出现在汇总/真值表(case 选择表)/可验证性。"""
    from dreg_verify import cli
    excel = tmp_path / "mux_rep.xlsx"
    fixtures.build_workbook(str(excel), with_mux=True)
    html_path = tmp_path / "report.html"
    cli.main(["--excel", str(excel), "--report", str(html_path)])
    html = html_path.read_text(encoding="utf-8")
    assert "d_bt_lp_rccal_i[3:0]" in html              # mux 信号在报告里
    assert "case(d_logic_bt_lp_lna_agc)" in html       # case 选择表达式
    assert "mux1" in html                              # assert id
    # mux 真值表块标题不带 R 前缀（Rmux1 → mux1）
    assert "Rmux1" not in html


def test_cli_no_mux_sheet_unchanged(tmp_path):
    """CLI: 无 mux 页的 Excel + 新旗标不崩、行为与旧版一致。"""
    from dreg_verify import cli
    excel = tmp_path / "no_mux.xlsx"
    fixtures.build_workbook(str(excel))
    out = tmp_path / "out.sv"
    cli.main(["--excel", str(excel), "--out", str(out)])
    assert "assert_mux" not in out.read_text(encoding="utf-8")


# ───────────── ⑦ GUI（离屏冒烟）─────────────
@pytest.fixture(scope="module")
def gui_app():
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    pytest.importorskip("PySide6")
    from PySide6 import QtWidgets
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([])
    yield app


@pytest.fixture()
def gui_win(gui_app, tmp_path):
    from dreg_verify import gui as G
    excel = tmp_path / "gui_mux.xlsx"
    fixtures.build_workbook(str(excel), with_mux=True)
    w = G.MainWindow()
    w.path_edit.setText(str(excel))
    w.on_load()
    yield w
    w.close()


def test_gui_mux_in_signal_table(gui_win):
    """mux 信号混排进信号表：type=mux、R 列=mux<N>、表达式列=case 文本、状态=可验证。"""
    w = gui_win
    assert len(w.signals) == 5                            # 3 logic + 2 mux
    mux_rows = [i for i, s in enumerate(w.signals)
                if getattr(s, "suffix", "") == "mux"]
    assert len(mux_rows) == 2
    grp = w.signals[mux_rows[0]]
    assert grp.assert_id == "mux1"
    assert "case(" in grp.expr
    # 状态分析: mux 组应为 clean（fixture 数据完整可解析）
    assert w._analysis[mux_rows[0]]["status"] == "clean"


def test_gui_mux_test_items_readonly(gui_win):
    """点击 mux 信号 → case 表（_ti_sig=None 屏蔽列编辑按钮，_ti_mux_sig 记录当前 mux；
    输入行只读，期望行可手填——见 test_designer_expected.py 的 mux 专项测试）。"""
    w = gui_win
    grp = next(s for s in w.signals if getattr(s, "suffix", "") == "mux")
    w._load_test_items(grp)
    assert w._ti_sig is None                              # 编辑路径安全屏蔽
    assert w._ti_mux_sig is grp
    assert w.ti_table.columnCount() == 4                  # 3 case(min) + 1 local 路径
    assert w.ti_table.rowCount() == 8                     # 3 控制输入 + 3 数据寄存器 + auto_out + 期望
    assert "mux" in w.ti_header.text()


def test_gui_mux_inputs_box_populated(gui_win):
    """⭐ 用户问题回归: 点 mux 信号后『输入信号』表不再空白——控制输入+数据寄存器都列出来。"""
    w = gui_win
    grp = next(s for s in w.signals if getattr(s, "suffix", "") == "mux")
    w._load_test_items(grp)
    # 3 控制行输入(c:A/c:B/c:C) + 3 数据寄存器(d:0/d:1/d:2) = 6 行
    assert w.ti_inputs.rowCount() == 6
    roles = [w.ti_inputs.item(r, 2).text() for r in range(6)]
    # 控制输入细分 line/local/模式位；数据寄存器标"被该 case 选中"
    assert any("line路径" in x for x in roles)
    assert any("local路径" in x for x in roles)
    assert sum(1 for x in roles if "数据寄存器" in x) == 3
    # 数据寄存器的『字母』列显示对应的 case 值（不是 d:0 这种内部键）
    letters = [w.ti_inputs.item(r, 0).text() for r in range(6)]
    assert "case 3'b010" in letters and "case 3'b10x" in letters
    # 驱动列：数据寄存器 = RF_WRITE，line 线控 = force
    drives = [w.ti_inputs.item(r, 4).text() for r in range(6)]
    assert any(x.startswith("RF_WRITE") for x in drives)
    assert any(x.startswith("force ENV_RF") for x in drives)


def test_gui_mux_coverage_levels_differ(gui_win):
    """⭐ 用户问题回归: GUI 覆盖度三档下 mux 测试列数必须递增（精简<全面<穷举）。"""
    w = gui_win
    grp = next(s for s in w.signals if getattr(s, "suffix", "") == "mux")
    counts = {}
    for level in ("精简", "全面", "穷举"):
        w.coverage_mux.setCurrentText(level)              # mux 侧档位（第二十二轮起与 logic 解耦）
        w._load_test_items(grp)
        counts[level] = w.ti_table.columnCount()
    assert counts["精简"] < counts["全面"] < counts["穷举"], "三档必须递增: %s" % counts
    # 表头标明当前档位生成内容
    assert "覆盖度=穷举" in w.ti_header.text()
    w.coverage_mux.setCurrentText("精简")                 # 还原默认，不影响其它用例


def test_gui_mux_generate(gui_win, tmp_path):
    """GUI 勾选 mux 信号 → build 产出 mux 块（名字匹配路径）+ 负向勾选经 neg_signals 生效。"""
    w = gui_win
    from PySide6 import QtCore
    G_COL_SEL, G_COL_NEG = 0, 1
    # 勾选全部行（含 mux）+ 给 mux 行勾负向
    for r in range(w.table.rowCount()):
        w.table.item(r, G_COL_SEL).setCheckState(QtCore.Qt.Checked)
        sig = w._sig_of_row(r)
        if getattr(sig, "suffix", "") == "mux":
            w.table.item(r, G_COL_NEG).setCheckState(QtCore.Qt.Checked)
    sel = w._collect()
    assert len(sel) == 5
    assert len(w._mux_neg_checked()) == 2                 # 两个 mux 都勾了负向
    res = generator.build(w.wb, w._opts(sel))
    assert res["summary"]["n_mux_generated"] == 2
    # mux 负向真的生效（每组追加 1 条 _NEG）
    text = generator.render(res)
    assert "assert_mux1_T4_NEG:" in text
