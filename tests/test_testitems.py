# -*- coding: utf-8 -*-
"""测试项编辑能力的回归测试：
  · vectors.input_groups / vector_to_base_values / make_vector_from_base_values
  · generator.build 的 vector_overrides 回流(GUI 编辑 → .sv)
  · GUI 编辑流程冒烟(无 PySide6 则跳过)
"""

import os
import sys

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")   # GUI 测试用离屏后端
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dreg_verify import excel_model, generator     # noqa: E402
from dreg_verify import expr as E                   # noqa: E402
from dreg_verify import vectors as V                # noqa: E402
from dreg_verify import resolver as R               # noqa: E402
import fixtures                                      # noqa: E402


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    path = tmp_path_factory.mktemp("xl") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    return excel_model.load_workbook(str(path))


def _reserve(wb):
    sig = next(s for s in wb.logic if s.out_name == "d_logic_bt_lp_reserve")
    node = E.parse(sig.expr)               # (A?C:B)&(~J)
    bindings = R.Resolver(wb).resolve_signal_inputs(sig)
    groups = V.input_groups(node, bindings)
    return sig, node, bindings, groups


# ───────────── input_groups ─────────────
def test_input_groups_order_and_roles(wb):
    sig, node, bindings, groups = _reserve(wb)
    bases = [g["base"] for g in groups]
    # collect_vars 发现顺序 = A,C,B,J（三元先 cond 再 then/els，再门控位）
    assert bases == ["d_bt_lp_linelocal_mode_ctrl", "d_bt_lp_bt_mode_sel_local",
                     "d_bt_lp_bt_mode_sel", "d_bt_lp_iddq"]
    gd = {g["base"]: g for g in groups}
    assert gd["d_bt_lp_linelocal_mode_ctrl"]["is_control"]   # 三元条件 A
    assert gd["d_bt_lp_iddq"]["is_control"]                  # 门控位 J
    assert not gd["d_bt_lp_bt_mode_sel"]["is_control"]       # 分支数据 B
    assert all(g["width"] == 1 for g in groups)


def test_input_groups_label_keeps_width(wb):
    """显示用 label 保留位宽 [msb:lsb]，但 base(查表 key)仍不带位宽。"""
    sig = next(s for s in wb.logic if s.out_name == "d_logic_bt_lp_lna_agc[2:0]")
    node = E.parse(sig.expr)
    bindings = R.Resolver(wb).resolve_signal_inputs(sig)
    by_base = {g["base"]: g for g in V.input_groups(node, bindings)}
    assert by_base["d_bt_lp_lna_agc_local"]["label"] == "d_bt_lp_lna_agc_local[2:0]"
    assert by_base["d_bt_lp_lna_agc_line"]["label"] == "d_bt_lp_lna_agc_line[2:0]"
    assert by_base["d_bt_lp_lna_line_sel"]["label"] == "d_bt_lp_lna_line_sel"   # 1bit 无切片
    assert by_base["d_bt_lp_lna_agc_local"]["base"] == "d_bt_lp_lna_agc_local"  # base 仍纯净


def test_input_groups_shared_base(wb):
    """同一物理信号占多个字母时只占一列（直通信号 d_en_refbuf 只有 A）。"""
    sig = next(s for s in wb.logic if s.out_name == "d_en_refbuf")
    node = E.parse(sig.expr)
    bindings = R.Resolver(wb).resolve_signal_inputs(sig)
    groups = V.input_groups(node, bindings)
    assert [g["base"] for g in groups] == ["d_bt_lp_en_refbuf_cfg"]
    assert groups[0]["letters"] == ["A"]


# ───────────── make_vector_from_base_values ─────────────
def test_make_vector_expected_autocompute(wb):
    sig, node, bindings, groups = _reserve(wb)
    # A=1(选 C),C=1,B=0,J=0 → (1?1:0)&(~0)=1
    bv = {"d_bt_lp_linelocal_mode_ctrl": 1, "d_bt_lp_bt_mode_sel_local": 1,
          "d_bt_lp_bt_mode_sel": 0, "d_bt_lp_iddq": 0}
    vec = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width)
    assert vec.exp_value == 1 and not vec.is_negative
    assert vec.assignments["A"] == 1 and vec.assignments["C"] == 1
    assert vec.assignments["B"] == 0 and vec.assignments["J"] == 0
    # 门控 J=1 → 拉 0
    bv2 = dict(bv); bv2["d_bt_lp_iddq"] = 1
    assert V.make_vector_from_base_values(node, bindings, groups, bv2, sig.out_width).exp_value == 0


def test_make_vector_override_marks_negative(wb):
    sig, node, bindings, groups = _reserve(wb)
    bv = {"d_bt_lp_linelocal_mode_ctrl": 1, "d_bt_lp_bt_mode_sel_local": 1,
          "d_bt_lp_bt_mode_sel": 0, "d_bt_lp_iddq": 0}   # 正确期望=1
    neg = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width,
                                         expected_override=0)
    assert neg.is_negative and neg.neg_value == 0 and neg.exp_value == 1
    assert neg.asserted_value == 0
    # 手填的期望与算出值相同 → 不算负向
    ok = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width,
                                        expected_override=1)
    assert not ok.is_negative


def test_vector_to_base_values_roundtrip(wb):
    sig, node, bindings, groups = _reserve(wb)
    vecs, _meta = V.generate_vectors(node, bindings, sig.out_width)
    for vec in vecs:
        bv = V.vector_to_base_values(vec, groups)
        rt = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width)
        assert rt.exp_value == vec.exp_value          # 往返不改期望
        assert rt.assignments == vec.assignments       # 字母展开一致


def test_base_value_clamped_to_width(wb):
    """编辑值超过输入位宽时按位宽裁剪，不溢出。"""
    sig = next(s for s in wb.logic if s.out_name == "d_logic_bt_lp_lna_agc[2:0]")
    node = E.parse(sig.expr)                # A?C:B, B/C 为 3bit
    bindings = R.Resolver(wb).resolve_signal_inputs(sig)
    groups = V.input_groups(node, bindings)
    gd = {g["base"]: g for g in groups}
    assert gd["d_bt_lp_lna_agc_local"]["width"] == 3
    bv = {"d_bt_lp_lna_line_sel": 1, "d_bt_lp_lna_agc_local": 0xFF,
          "d_bt_lp_lna_agc_line": 0}
    vec = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width)
    # 0xFF 裁到 3bit=7；A=1 选 C → 期望=7
    assert vec.exp_value == 7


# ───────────── generator.build vector_overrides ─────────────
def test_vector_overrides_replaces_auto(wb):
    sig, node, bindings, groups = _reserve(wb)
    bv = {"d_bt_lp_linelocal_mode_ctrl": 1, "d_bt_lp_bt_mode_sel_local": 1,
          "d_bt_lp_bt_mode_sel": 0, "d_bt_lp_iddq": 0}
    custom = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width)
    opts = generator.GenOptions(signals=["d_logic_bt_lp_reserve"], top_output_only=False,
                                vector_overrides={"d_logic_bt_lp_reserve": [custom]})
    res = generator.build(wb, opts)
    assert res["summary"]["n_generated"] == 1
    assert res["summary"]["n_vectors"] == 1           # 只有我们这一条，不是自动那批
    text = "\n".join(res["blocks"][0][0])
    assert "==1'b1" in text                            # 期望 1
    assert "`RF_WRITE(10'h2D,16'h3);" in text          # A,C 同地址 0x2D 合并→bit0|bit1=0x3
    assert "force `ENV_RF.d_bt_lp_iddq=16'h0;" in text  # J(RO)→force


def test_vector_overrides_negative_to_sv(wb):
    sig, node, bindings, groups = _reserve(wb)
    bv = {"d_bt_lp_linelocal_mode_ctrl": 1, "d_bt_lp_bt_mode_sel_local": 1,
          "d_bt_lp_bt_mode_sel": 0, "d_bt_lp_iddq": 0}
    neg = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width,
                                         expected_override=0)
    opts = generator.GenOptions(signals=["d_logic_bt_lp_reserve"], top_output_only=False,
                                vector_overrides={"d_logic_bt_lp_reserve": [neg]})
    res = generator.build(wb, opts)
    text = "\n".join(res["blocks"][0][0])
    # 反例(自检式，2026-06-03)：断言语法与正例一模一样，== 故意填错的 0（正确应为 1）→
    # 仿真时必然 FAIL → uvm_report_error 正常触发(消息带 NEG-EXPECTED-FAIL 标签)
    assert "==1'b0" in text and "!=" not in text
    assert "NEG-EXPECTED-FAIL" in text and "NEG-BROKEN" in text
    assert res["blocks"][0][1]["n_negative"] == 1


def test_overrides_only_affect_named_signal(wb):
    """有 override 的信号用 override；其它信号仍走自动生成。"""
    sig, node, bindings, groups = _reserve(wb)
    bv = {"d_bt_lp_linelocal_mode_ctrl": 0, "d_bt_lp_bt_mode_sel_local": 0,
          "d_bt_lp_bt_mode_sel": 0, "d_bt_lp_iddq": 0}
    custom = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width)
    opts = generator.GenOptions(top_output_only=False,
                                vector_overrides={"d_logic_bt_lp_reserve": [custom]})
    res = generator.build(wb, opts)
    by_name = {st["out_name"]: (lines, st) for lines, st in res["blocks"]}
    assert by_name["d_logic_bt_lp_reserve"][1]["n_vectors"] == 1        # override
    assert by_name["d_logic_bt_lp_lna_agc[2:0]"][1]["n_vectors"] > 1     # 自动


# ───────────── 审查修复回归 ─────────────
def test_unequal_width_same_base_roundtrip():
    """#5: 同一 base 跨不等宽字母时，往返不丢高位（取最宽字母作代表）。"""
    from dreg_verify.resolver import InputBinding
    A = InputBinding("A", "sig", "sig", 2, "RO", None, None, None, "sig", "wire")
    B = InputBinding("B", "sig", "sig", 4, "RO", None, None, None, "sig", "wire")
    bindings = {"A": A, "B": B}
    node = E.parse("A|B")
    groups = V.input_groups(node, bindings)
    assert len(groups) == 1
    assert groups[0]["rep"] == "B" and groups[0]["width"] == 4   # 最宽字母作代表
    vecs, _ = V.generate_vectors(node, bindings, 4)
    for vec in vecs:
        bv = V.vector_to_base_values(vec, groups)
        rt = V.make_vector_from_base_values(node, bindings, groups, bv, 4)
        assert rt.exp_value == vec.exp_value, "往返丢位: %r" % vec.assignments


def test_empty_override_means_zero_tests(wb):
    """#1: 空 override 列表 = 该信号零用例(一键清空)，而非回退自动生成；
    且不静默——整组进 skipped 并给「用户已清空」原因(第二十六轮：与 mux 清空对称、可见)。"""
    opts = generator.GenOptions(signals=["d_logic_bt_lp_reserve"], top_output_only=False,
                                vector_overrides={"d_logic_bt_lp_reserve": []})
    res = generator.build(wb, opts)
    assert res["summary"]["n_generated"] == 0
    assert res["summary"]["n_vectors"] == 0
    skipped_names = {n for n, _aid, _r in res["skipped"]}
    assert "d_logic_bt_lp_reserve" in skipped_names
    reason = next(r for n, _aid, r in res["skipped"] if n == "d_logic_bt_lp_reserve")
    assert "清空" in "; ".join(str(w) for *_h, w in reason)
    # 报告侧同口径：summary 行 n_tests=0 且 error 带「清空」(双轨一致)
    rep = generator.report(wb, opts)
    row = next(r for r in rep["summary"] if r["signal"] == "d_logic_bt_lp_reserve")
    assert row["n_tests"] == 0 and "清空" in row["error"]


def test_negative_sv_has_neg_suffix(wb):
    """#4: 负向断言 id 带 _NEG 后缀，可与正向区分。"""
    sig, node, bindings, groups = _reserve(wb)
    bv = {"d_bt_lp_linelocal_mode_ctrl": 1, "d_bt_lp_bt_mode_sel_local": 1,
          "d_bt_lp_bt_mode_sel": 0, "d_bt_lp_iddq": 0}
    neg = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width,
                                         expected_override=0)
    res = generator.build(wb, generator.GenOptions(
        signals=["d_logic_bt_lp_reserve"], top_output_only=False,
        vector_overrides={"d_logic_bt_lp_reserve": [neg]}))
    text = "\n".join(res["blocks"][0][0])
    assert "assert_106_T0_NEG:" in text


def test_override_plus_left_neg_appends_auto(wb):
    """#3: 既 override 又在 neg 列表里的信号，正向行仍会补自动负向(不被静默丢)。"""
    sig, node, bindings, groups = _reserve(wb)
    bv = {"d_bt_lp_linelocal_mode_ctrl": 1, "d_bt_lp_bt_mode_sel_local": 1,
          "d_bt_lp_bt_mode_sel": 0, "d_bt_lp_iddq": 0}
    pos = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width)  # 正向
    res = generator.build(wb, generator.GenOptions(
        signals=["d_logic_bt_lp_reserve"], top_output_only=False,
        neg_all=True, neg_which="first",
        vector_overrides={"d_logic_bt_lp_reserve": [pos]}))
    st = res["blocks"][0][1]
    assert st["n_vectors"] == 2          # 1 正向 + 1 追加的自动负向
    assert st["n_negative"] == 1


def test_report_tables_structure(wb):
    """report() 返回 tables（每信号真值表）：输入做行、各测试做列，负向列有 _NEG。"""
    rep = generator.report(wb, generator.GenOptions(
        signals=["d_logic_bt_lp_reserve"], top_output_only=False,
        neg_all=True, neg_which="first"))
    assert "tables" in rep
    t = next(x for x in rep["tables"] if x["signal"] == "d_logic_bt_lp_reserve")
    assert t["inputs"] and t["tests"]
    names = [tc["name"] for tc in t["tests"]]
    assert any(n.endswith("_NEG") for n in names)              # 含负向列
    assert all(len(tc["values"]) == len(t["inputs"]) for tc in t["tests"])  # 每列值数 = 输入数
    neg_tc = next(tc for tc in t["tests"] if tc["neg"])
    assert neg_tc["expected"] != neg_tc["correct"]             # 负向期望 != 正确值


def test_report_tables_bitwidth_label(wb):
    """真值表的输入/输出行标签保留位宽。"""
    rep = generator.report(wb, generator.GenOptions(
        signals=["d_logic_bt_lp_lna_agc"], top_output_only=False))
    t = next(x for x in rep["tables"] if x["signal"] == "d_logic_bt_lp_lna_agc[2:0]")
    labels = [i["label"] for i in t["inputs"]]
    assert "d_bt_lp_lna_agc_local[2:0]" in labels
    assert t["exp_label"] == "期望(out)[2:0]"


def test_custom_test_name_flows_to_sv(wb):
    """自定义 name 贯穿到 .sv 断言 id；默认 None 时仍是 T<index>；负向仍带 _NEG。"""
    import dreg_verify.sv_writer as W
    sig, node, bindings, groups = _reserve(wb)
    bv = {g["base"].lower(): 0 for g in groups}
    v_named = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width,
                                             index=0, name="boundary_case")
    v_default = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width, index=1)
    v_neg = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width,
                                           index=2, expected_override=1, name="bad_iddq")
    assert W.test_label(v_named) == "boundary_case"
    assert W.test_label(v_default) == "T1"               # 默认仍 T<index>
    assert W.test_label(v_neg) == "bad_iddq_NEG"          # 负向自动带 _NEG
    res = generator.build(wb, generator.GenOptions(
        signals=["d_logic_bt_lp_reserve"], top_output_only=False,
        vector_overrides={"d_logic_bt_lp_reserve": [v_named, v_neg]}))
    text = "\n".join(res["blocks"][0][0])
    assert "assert_106_boundary_case:" in text
    assert "assert_106_bad_iddq_NEG:" in text


def test_gui_rename_only_user_added(qapp, wb, tmp_path_factory):
    """自动列不可改名；用户新增列可改名，且贯穿到生成的 .sv。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "g7")
    w._load_test_items(sig)
    n_auto = len(w._ti_rows)
    # 自动列(col 0)拒绝改名
    ok, msg = w._ti_set_test_name(0, "should_fail")
    assert not ok
    # 新增一正向列 → 可改名
    w.on_ti_add()
    new_col = len(w._ti_rows) - 1
    ok, final = w._ti_set_test_name(new_col, "my special case!")   # 非法字符会被清成 _
    assert ok and final == "my_special_case_"
    assert w._ti_rows[new_col]["name"] == "my_special_case_"
    # 列头显示自定义名
    assert w.ti_table.horizontalHeaderItem(new_col).text() == "my_special_case_"
    # 改名贯穿 .sv
    sel = [sig.out_name]
    res = generator.build(w.wb, w._opts(sel))
    text = generator.render(res)
    assert "_my_special_case_:" in text
    # 重名拒绝
    w.on_ti_add()
    ok2, msg2 = w._ti_set_test_name(len(w._ti_rows) - 1, "my special case!")
    assert not ok2 and "重复" in msg2


def test_gui_reject_reserved_tname(qapp, wb, tmp_path_factory):
    """禁止把列改名成 T<编号>(自动测试保留命名)，防后续位移撞名→重复标号(审查#1/#5)。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "g8")
    w._load_test_items(sig)
    w.on_ti_add()
    col = len(w._ti_rows) - 1
    ok, msg = w._ti_set_test_name(col, "T5")
    assert not ok and "冲突" in msg
    ok2, _ = w._ti_set_test_name(col, "t9_neg")     # 大小写/带 _neg 也拒绝
    assert not ok2
    ok3, final = w._ti_set_test_name(col, "boundary")   # 正常名可改
    assert ok3 and final == "boundary"


def test_build_no_dup_labels_normal(wb):
    """正常情况(R 各异)无重复标号。"""
    res = generator.build(wb, generator.GenOptions(top_output_only=False))
    assert res["summary"]["n_dup_labels"] == 0
    assert res["dup_labels"] == []


def test_build_dup_labels_same_R(wb, tmp_path_factory):
    """构造一张两信号同 R 的表，确认 build 抓到重复标号。"""
    import openpyxl
    from openpyxl.utils import column_index_from_string
    path = tmp_path_factory.mktemp("dup") / "dup.xlsx"
    import fixtures
    fixtures.build_workbook(str(path))
    # 把 lna_agc(R=108) 的 R 改成和 reserve 一样的 106
    wbx = openpyxl.load_workbook(str(path))
    ws = wbx["logic"]
    for row in ws.iter_rows():
        if ws.cell(row=row[0].row, column=column_index_from_string("K")).value == "d_logic_bt_lp_lna_agc[2:0]":
            ws.cell(row=row[0].row, column=column_index_from_string("R")).value = 106
    wbx.save(str(path))
    wb2 = excel_model.load_workbook(str(path))
    res = generator.build(wb2, generator.GenOptions(top_output_only=False))
    assert res["summary"]["n_dup_labels"] > 0          # 106_T0 等被两信号共用
    assert any("106_T0" == lbl for lbl, _a, _b in res["dup_labels"])


def test_gui_neg_rebuild_keeps_name_and_wrong(qapp, wb, tmp_path_factory):
    """重建负向时保留用户对负向列的改名与手填错值(审查#2/#3)。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "g9")
    w._load_test_items(sig)
    w.on_ti_add_neg_all()                                  # 每条正向各追加一条负向
    neg_col = next(i for i, rd in enumerate(w._ti_rows) if rd.get("kind") == "neg")
    # 改名 + 手填错值
    ok, _ = w._ti_set_test_name(neg_col, "my_special_neg")
    assert ok
    w._ti_rows[neg_col]["wrong_value"] = 1
    w._edited[sig.out_name.lower()] = {"sig": sig, "rows": w._ti_rows}
    # 再次"加负向"(重建路径)
    w.on_ti_add_neg_all()
    negs = [rd for rd in w._ti_rows if rd.get("kind") == "neg"]
    assert any(rd.get("name") == "my_special_neg" for rd in negs), "重建丢了负向列改名"
    assert any(rd.get("wrong_value") == 1 for rd in negs), "重建丢了手填错值"


def test_report_honors_overrides(wb):
    """#8: report() 也用 override，与 .sv 一致。"""
    sig, node, bindings, groups = _reserve(wb)
    bv = {"d_bt_lp_linelocal_mode_ctrl": 0, "d_bt_lp_bt_mode_sel_local": 0,
          "d_bt_lp_bt_mode_sel": 0, "d_bt_lp_iddq": 0}
    custom = V.make_vector_from_base_values(node, bindings, groups, bv, sig.out_width)
    rep = generator.report(wb, generator.GenOptions(
        signals=["d_logic_bt_lp_reserve"], top_output_only=False,
        vector_overrides={"d_logic_bt_lp_reserve": [custom]}))
    row = next(r for r in rep["summary"] if r["signal"] == "d_logic_bt_lp_reserve")
    assert row["n_tests"] == 1           # override 那一条，不是自动那批


def test_no_overrides_unchanged(wb):
    """不传 vector_overrides 时行为与以前完全一致（默认 None）。"""
    a = generator.build(wb, generator.GenOptions(signals=["d_logic_bt_lp_reserve"],
                                                 top_output_only=False))
    b = generator.build(wb, generator.GenOptions(signals=["d_logic_bt_lp_reserve"],
                                                 top_output_only=False, vector_overrides=None))
    assert a["summary"]["n_vectors"] == b["summary"]["n_vectors"]
    assert generator.render(a) == generator.render(b)


# ───────────── GUI 冒烟（无 PySide6 跳过） ─────────────
@pytest.fixture(scope="module")
def qapp():
    pytest.importorskip("PySide6")
    from PySide6 import QtWidgets
    return QtWidgets.QApplication.instance() or QtWidgets.QApplication([])


def test_gui_testitem_edit_flow(qapp, wb, tmp_path_factory):
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("g") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    w = gui.MainWindow()
    w.path_edit.setText(str(path))
    w.on_load()
    assert len(w.signals) == 3
    sig = next(s for s in w.signals if s.out_name == "d_logic_bt_lp_reserve")
    w._load_test_items(sig)
    assert w._ti_sig is sig and len(w._ti_rows) >= 1
    n0 = len(w._ti_rows)

    # 纵向布局：行=输入/期望/负向，列=测试。编辑 T0(列0)第一个输入(行0)的单元格
    w.ti_table.item(0, 0).setText("0x1")
    assert "d_logic_bt_lp_reserve" in w._customized

    # 加一列(测试)
    w.on_ti_add()
    assert len(w._ti_rows) == n0 + 1

    # 编辑过的测试项经 vector_overrides 进入 build
    ov = w._vector_overrides()
    assert ov and "d_logic_bt_lp_reserve" in ov
    assert len(ov["d_logic_bt_lp_reserve"]) == len(w._ti_rows)
    res = generator.build(w.wb, w._opts(["d_logic_bt_lp_reserve"], None))
    assert res["summary"]["n_vectors"] == len(w._ti_rows)

    # 重新生成 → 丢弃自定义
    w.on_ti_regen()
    assert "d_logic_bt_lp_reserve" not in w._customized


def _reserve_window(tmp_path_factory, sub):
    from dreg_verify import gui
    path = tmp_path_factory.mktemp(sub) / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    w = gui.MainWindow()
    w.path_edit.setText(str(path)); w.on_load()
    sig = next(s for s in w.signals if s.out_name == "d_logic_bt_lp_reserve")
    return gui, w, sig


def test_gui_add_negative_appends_not_modifies(qapp, wb, tmp_path_factory):
    """加负向 = 追加专门的负向测试，原有正向测试一条不改(#2 语义)。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "g2")
    w._load_test_items(sig)
    n_pos = len(w._ti_rows)
    pos_snapshot = [(dict(rd["base_values"]), rd["expected"]) for rd in w._ti_rows]
    w.on_ti_add_neg_all()
    pos = [rd for rd in w._ti_rows if rd.get("kind") != "neg"]
    neg = [rd for rd in w._ti_rows if rd.get("kind") == "neg"]
    assert len(pos) == n_pos and len(neg) == n_pos      # 正向数不变，等量追加负向
    # 正向测试完全没被改：期望仍 = 正确值
    assert [(dict(rd["base_values"]), rd["expected"]) for rd in pos] == pos_snapshot
    assert all(rd["expected"] == rd["correct"] for rd in pos)
    # 负向测试：期望 != 正确值，输入复制自某条正向
    assert all(rd["is_negative"] and rd["expected"] != rd["correct"] for rd in neg)
    # 删负向 → 只剩正向
    w.on_ti_del_neg()
    assert all(rd.get("kind") != "neg" for rd in w._ti_rows) and len(w._ti_rows) == n_pos


def test_gui_left_neg_adds_one(qapp, wb, tmp_path_factory):
    """左侧'负向'勾选 → 默认只加 1 条负向自检(正向保留)；取消 → 撤销定制。"""
    from PySide6 import QtCore
    gui, w, sig = _reserve_window(tmp_path_factory, "g5")
    row = next(r for r in range(w.table.rowCount())
               if w._sig_of_row(r).out_name == "d_logic_bt_lp_reserve")
    w.on_row_focus(row, gui.COL_K, -1, -1)
    n_pos = len(w._ti_rows)
    w.table.item(row, gui.COL_NEG).setCheckState(QtCore.Qt.Checked)
    pos = [rd for rd in w._ti_rows if rd.get("kind") != "neg"]
    neg = [rd for rd in w._ti_rows if rd.get("kind") == "neg"]
    assert len(pos) == n_pos and len(neg) == 1           # 默认仅 1 条
    assert "d_logic_bt_lp_reserve" in w._neg_only        # 仅负向定制(可随覆盖度重算)
    # 取消 → 该信号定制被整体撤销(纯自动，不留 override)
    w.table.item(row, gui.COL_NEG).setCheckState(QtCore.Qt.Unchecked)
    assert "d_logic_bt_lp_reserve" not in w._customized


def test_gui_add_neg_selected_precise(qapp, wb, tmp_path_factory):
    """编辑器'加负向(选中)' → 只给选中的正向列各加一条负向(精确控制)。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "gsel")
    w._load_test_items(sig)
    n_pos = len([rd for rd in w._ti_rows if rd.get("kind") != "neg"])
    assert n_pos >= 2                                     # 需要 ≥2 条正向才能验证"只选其一"
    # 选中第 1 列(T0) → 只加 1 条负向，且 base_values 复制自 T0
    w.ti_table.setCurrentCell(0, 0)
    w.on_ti_add_neg_selected()
    neg = [rd for rd in w._ti_rows if rd.get("kind") == "neg"]
    assert len(neg) == 1
    assert neg[0]["base_values"] == w._ti_rows[0]["base_values"]
    assert sig.out_name.lower() in w._customized and sig.out_name.lower() not in w._neg_only
    # 再选中 T1 → 再加 1 条(累加，且这条复制自 T1，而非又一条 T0)
    t1_bv = dict(w._ti_rows[1]["base_values"])
    w.ti_table.setCurrentCell(0, 1)
    w.on_ti_add_neg_selected()
    negs = [rd for rd in w._ti_rows if rd.get("kind") == "neg"]
    assert len(negs) == 2
    assert any(rd["base_values"] == t1_bv for rd in negs)   # 精确命中 T1，不是兜底回 T0


def test_gui_all_signals_negative(qapp, wb, tmp_path_factory):
    """'全部加负向' → 每个可见信号各拿到【恰好 1 条】负向自检(满足 R 全部负向、默认1条)。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "g6")
    w.on_all_signals_neg(True)
    ov = w._vector_overrides()
    assert ov is not None
    # 每个有 override 的信号恰好 1 条负向(不是每条正向各一条)
    assert all(sum(1 for v in vs if v.is_negative) == 1 for vs in ov.values() if vs)
    assert len(ov) == len(w.signals)            # 3 个信号都加了
    # 列头 _NEG 检查(打开 reserve)
    w._load_test_items(sig)
    neg_cols = [i for i, rd in enumerate(w._ti_rows) if rd["is_negative"]]
    assert neg_cols and w.ti_table.horizontalHeaderItem(neg_cols[0]).text().endswith("_NEG")
    # 清除负向 → 撤销全部定制(恢复默认，不残留 override 绕过 risky-skip)
    w.on_all_signals_neg(False)
    assert not w._customized


def test_gui_bulk_neg_nondestructive(qapp, wb, tmp_path_factory):
    """审查#1：'全部加负向'不能把已有多条负向的信号塌成 1 条。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "gbulk")
    w._load_test_items(sig)
    w.on_ti_add_neg_all()                                  # reserve 拿到 N 条(每条正向各一)
    n_neg_before = sum(1 for rd in w._ti_rows if rd.get("kind") == "neg")
    assert n_neg_before >= 2
    w.on_all_signals_neg(True)                            # 批量加负向：已有的应保持不动
    ov = w._vector_overrides()
    n_neg_after = sum(1 for v in ov[sig.out_name.lower()] if v.is_negative)
    assert n_neg_after == n_neg_before                    # 没被塌成 1 条


def test_gui_add_neg_selected_no_duplicate(qapp, wb, tmp_path_factory):
    """审查#2：对'已经有负向'的正向列再'加负向(选中)'不产生重复负向。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "gdup")
    w._load_test_items(sig)
    w.on_ti_add_neg_all()                                 # 每条正向各 1 条
    n_neg = sum(1 for rd in w._ti_rows if rd.get("kind") == "neg")
    w.ti_table.setCurrentCell(0, 0)                       # T0 已有负向
    w.on_ti_add_neg_selected()
    assert sum(1 for rd in w._ti_rows if rd.get("kind") == "neg") == n_neg   # 没增加


def test_gui_default_one_survives_coverage_change(qapp, wb, tmp_path_factory):
    """审查#5(头号)：左侧勾选的 1 条负向，切覆盖度后仍是 1 条(不被炸成每条一条)。"""
    from PySide6 import QtCore
    gui, w, sig = _reserve_window(tmp_path_factory, "gcov1")
    row = next(r for r in range(w.table.rowCount())
               if w._sig_of_row(r).out_name == sig.out_name)
    w.coverage.setCurrentText("全面")
    w.on_row_focus(row, gui.COL_K, -1, -1)
    w.table.item(row, gui.COL_NEG).setCheckState(QtCore.Qt.Checked)   # 默认 1 条
    assert sum(1 for rd in w._ti_rows if rd.get("kind") == "neg") == 1
    assert w._neg_only.get(sig.out_name.lower()) == "first"           # 规则被记住
    w.coverage.setCurrentText("精简")                                  # 切覆盖度 → 重算
    n_neg = sum(1 for rd in w._ti_rows if rd.get("kind") == "neg")
    assert n_neg == 1                                                  # 仍是 1 条，不是每条一条！
    assert w._neg_only.get(sig.out_name.lower()) == "first"


def test_gui_positive_only_strips_negatives(qapp, wb, tmp_path_factory):
    """#2: positive_only 把自定义负向向量从正向文件剔除。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "g4")
    w._load_test_items(sig)
    w.on_ti_add_neg_all()
    name = "d_logic_bt_lp_reserve"
    ov_all = w._vector_overrides()
    ov_pos = w._vector_overrides(positive_only=True)
    n_neg = sum(1 for v in ov_all[name] if v.is_negative)
    assert n_neg >= 1
    assert all(not v.is_negative for v in ov_pos[name])
    assert len(ov_pos[name]) == len(ov_all[name]) - n_neg


# ───────────── 覆盖度（精简/全面/穷举 合一） ─────────────
def test_gui_coverage_mapping(qapp, wb, tmp_path_factory):
    """覆盖度下拉 → (mode, exhaustive)：精简=min, 全面=max, 穷举=max+exhaustive。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "cov1")
    w.coverage.setCurrentText("精简"); assert w._coverage() == ("min", False)
    w.coverage.setCurrentText("全面"); assert w._coverage() == ("max", False)
    w.coverage.setCurrentText("穷举"); assert w._coverage() == ("max", True)


def test_gui_coverage_decouple_mapping(qapp, wb, tmp_path_factory):
    """第二十二轮解耦：logic 走 _coverage()、mux 走 _mux_coverage()/_mux_cov_mode()，互不影响。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "covdec")
    w.coverage.setCurrentText("精简"); w.coverage_mux.setCurrentText("穷举")
    assert w._coverage() == ("min", False)              # logic 侧
    assert w._mux_coverage() == ("max", True)           # mux 侧独立
    assert w._mux_cov_mode() == "exhaustive"
    w.coverage.setCurrentText("全面"); w.coverage_mux.setCurrentText("精简")
    assert w._coverage() == ("max", False)
    assert w._mux_cov_mode() == "min"


def test_gui_coverage_no_restore_under_pytest(qapp, wb, tmp_path_factory, monkeypatch):
    """⭐ 审查修复(wf_b5a361e3 confirmed)：pytest 下【不】从真实 settings 恢复覆盖档——即便磁盘有
    coverage_mux / legacy coverage 也默认精简。否则用户真机持久化的档位会污染断言覆盖版面的 GUI 测试。"""
    from dreg_verify import gui as G
    monkeypatch.setattr(G, "_load_settings",
                        lambda: {"coverage_mux": "穷举", "coverage_logic": "全面", "coverage": "穷举"})
    _gui, w, _sig = _reserve_window(tmp_path_factory, "covnorestore")
    assert w.coverage.currentText() == "精简"
    assert w.coverage_mux.currentText() == "精简"


def test_gui_coverage_live_update(qapp, wb, tmp_path_factory):
    """切换覆盖度即时刷新'未自定义'信号的测试项；'全面'用例数 ≥ '精简'。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "cov2")
    w.on_row_focus(next(r for r in range(w.table.rowCount())
                        if w._sig_of_row(r).out_name == sig.out_name),
                   gui.COL_K, -1, -1)
    w.coverage.setCurrentText("精简"); n_min = len(w._ti_rows)
    w.coverage.setCurrentText("全面"); n_max = len(w._ti_rows)
    assert n_min >= 1 and n_max >= n_min          # 切换即时生效(无需重新点信号)


def test_gui_coverage_keeps_customized(qapp, wb, tmp_path_factory):
    """已自定义的信号：切换覆盖度不应被覆盖掉编辑。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "cov3")
    w._load_test_items(sig)
    w.ti_table.item(0, 0).setText("0x1")          # 触发自定义
    assert sig.out_name.lower() in w._customized
    rows_before = len(w._ti_rows)
    w.coverage.setCurrentText("穷举")
    assert sig.out_name.lower() in w._customized   # 仍是自定义
    assert len(w._ti_rows) == rows_before          # 编辑没被重算冲掉


def test_gui_max_tests_live_update(qapp, wb, tmp_path_factory):
    """改'上限'spinbox 也即时重算当前非自定义信号(请求#1的另一半：max_tests 实时生效)。"""
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("cap") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    sig = next(s for s in w.signals if s.out_name == "d_logic_bt_lp_lna_agc[2:0]")
    row = next(r for r in range(w.table.rowCount())
               if w._sig_of_row(r).out_name == sig.out_name)
    w.coverage.setCurrentText("穷举")            # 穷举下用例较多
    w.on_row_focus(row, gui.COL_K, -1, -1)
    n0 = len(w._ti_rows)
    assert n0 > 2
    w.max_tests.setValue(2)                      # 改上限 → valueChanged 即时重算
    assert len(w._ti_rows) <= 2 and len(w._ti_rows) < n0


def test_gui_coverage_live_update_neg_only(qapp, wb, tmp_path_factory):
    """仅靠'负向'定制(无手改)的信号：切换覆盖度应重算正向并补回负向(审查#1修复)。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "covn")
    row = next(r for r in range(w.table.rowCount())
               if w._sig_of_row(r).out_name == sig.out_name)
    w.coverage.setCurrentText("全面")
    w.on_row_focus(row, gui.COL_K, -1, -1)
    w.on_ti_add_neg_all()                            # 仅负向定制(进入 _neg_only)
    assert sig.out_name.lower() in w._neg_only
    n_pos_max = sum(1 for rd in w._ti_rows if rd.get("kind") != "neg")
    # 切回精简：正向应按新覆盖度重算(更少)，且负向仍在(每条正向各一)
    w.coverage.setCurrentText("精简")
    pos = [rd for rd in w._ti_rows if rd.get("kind") != "neg"]
    neg = [rd for rd in w._ti_rows if rd.get("kind") == "neg"]
    assert len(pos) <= n_pos_max and len(neg) == len(pos) and len(neg) >= 1
    assert sig.out_name.lower() in w._neg_only   # 仍是仅负向定制


def test_gui_neg_only_switch_follows_global_coverage(qapp, tmp_path_factory, monkeypatch):
    """⭐用户实测bug：全选+加负向(在精简下)后，改全局覆盖度只对【当前看的】信号生效，
    切到【别的】neg_only信号还是精简。修=neg_only信号每次加载都按当前全局覆盖度重算正向+补负向。
    覆盖①会话内切换 ②退出重开 两条路径。"""
    from dreg_verify import gui
    monkeypatch.setattr(gui, "EDITS_PATH", str(tmp_path_factory.mktemp("negp") / "edits.json"))
    path = tmp_path_factory.mktemp("negsw") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    A = "d_logic_bt_lp_lna_agc[2:0]"; B = "d_logic_bt_lp_reserve"

    def pos_neg(w):
        pos = sum(1 for rd in w._ti_rows if rd.get("kind") != "neg")
        neg = sum(1 for rd in w._ti_rows if rd.get("kind") == "neg")
        return pos, neg

    def select(w, name):
        w._ti_loaded_idx = None
        r = next(i for i in range(w.table.rowCount()) if w._sig_of_row(i).out_name == name)
        w.on_row_focus(r, gui.COL_K, -1, -1)

    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    # 全局精简 + 全选 + 加负向（用户的确切操作）
    w.coverage.setCurrentText("精简")
    w.set_all_visible(True)
    w.on_all_signals_neg(True)
    assert A.lower() in w._neg_only and B.lower() in w._neg_only
    # 看着 A 改全局到全面（A 跟着变）
    select(w, A)
    w.coverage.setCurrentText("全面")
    a_pos, a_neg = pos_neg(w)
    # 基准：B 纯自动全面的正向数
    w2 = gui.MainWindow(); w2.path_edit.setText(str(path)); w2.on_load()
    w2.coverage.setCurrentText("全面"); select(w2, B)
    b_auto_full = pos_neg(w2)[0]
    # 关键：在原窗口切到 B（没在改全局时看它），应=全面、不是精简，且负向仍在
    select(w, B)
    b_pos, b_neg = pos_neg(w)
    assert b_pos == b_auto_full, "切到未在改全局时看的 neg_only 信号 B 应跟随全局全面(%d)，实得 %d" % (b_auto_full, b_pos)
    assert b_neg >= 1, "负向应保留"
    # ② 退出重开：global 全面恢复后，B 选中应仍是全面（不被磁盘冻结行钉死）
    w.coverage.setCurrentText("精简")              # 故意改回精简再重开，验证重开按重开时的全局走
    w3 = gui.MainWindow(); w3.path_edit.setText(str(path)); w3.on_load()
    w3.coverage.setCurrentText("全面"); select(w3, B)
    assert pos_neg(w3)[0] == b_auto_full, "重开后 B 选中应跟随当前全局全面"
    assert B.lower() in w3._neg_only and pos_neg(w3)[1] >= 1, "重开后负向定制应恢复"


def test_gui_neg_only_build_follows_global_coverage(qapp, tmp_path_factory):
    """⭐用户实测【设计哲学】bug(R27)：全选+加负向后，全局『精简/全面/穷举』对已加负向信号在
    【生成/报告/导出】里失效——build 用 _edited 冻结行(加负向那刻的档)而非按当前全局覆盖度重算。
    R25 只修了 GUI 显示侧(_load_test_items 切到信号才 reflow)；本测试盯的是 build 侧
    (generator.report / _opts / _vector_overrides)：不点开任一信号、直接切全局覆盖度再生成，
    报告里正向数必须跟随全局。修=_vector_overrides 对 neg_only 信号调 _neg_only_rows_now 重算。"""
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("negbuild") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path), with_pll_chain=True)   # 多输入信号 → 覆盖度差异明显
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    w.set_all_visible(True)

    def build_pos_neg():
        rep = generator.report(w.wb, w._opts(w._collect() or None))
        pos = sum((r.get("n_tests", 0) or 0) - (r.get("n_neg", 0) or 0) for r in rep["summary"])
        neg = sum(r.get("n_neg", 0) or 0 for r in rep["summary"])
        return pos, neg

    # 基准：纯自动两档，穷举正向应明显多于精简
    w.coverage.setCurrentText("精简"); clean_min, _ = build_pos_neg()
    w.coverage.setCurrentText("穷举"); clean_exh, _ = build_pos_neg()
    assert clean_exh > clean_min, "纯自动：穷举(%d)应>精简(%d)" % (clean_exh, clean_min)

    # 精简下全选加负向（用户确切动作）
    w.coverage.setCurrentText("精简")
    w.on_all_signals_neg(True)
    neg_min_pos, neg_min_neg = build_pos_neg()
    assert neg_min_pos == clean_min, "加负向后精简正向应=纯自动精简(%d)，实得 %d" % (clean_min, neg_min_pos)
    assert neg_min_neg >= 1, "确实加了负向"

    # 关键：不点开任一信号，直接切穷举再 build → 正向应跟随全局穷举(=clean_exh)，不是冻结的精简
    w.coverage.setCurrentText("穷举")
    neg_exh_pos, neg_exh_neg = build_pos_neg()
    assert neg_exh_pos == clean_exh, \
        "加负向后切穷举，build 正向应=clean穷举(%d)，实得 %d(冻结bug)" % (clean_exh, neg_exh_pos)
    assert neg_exh_neg == neg_min_neg, "负向数(first 规则=每信号1条)不应因切覆盖度丢失"


def test_gui_sig_cov_per_signal(qapp, tmp_path_factory):
    """⭐ 单点覆盖度：本信号下拉只改该信号、压过全局；存进 _sig_cov；_opts/build 带出 sig_cov。"""
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("sigcov") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    sig = next(s for s in w.signals if s.out_name == "d_logic_bt_lp_lna_agc[2:0]")
    row = next(r for r in range(w.table.rowCount())
               if w._sig_of_row(r).out_name == sig.out_name)
    name_low = sig.out_name.lower()
    # 全局精简，选中信号
    w.coverage.setCurrentText("精简")
    w.on_row_focus(row, gui.COL_K, -1, -1)
    assert w.sig_cov_combo.isEnabled()
    assert w.sig_cov_combo.currentText() == "跟随全局"
    n_min = len(w._ti_rows)
    # 本信号设单点穷举 → 即时重算变多；全局下拉没动
    w.sig_cov_combo.setCurrentText("穷举")
    assert w._sig_cov.get(name_low) == "exhaustive"
    assert w.coverage.currentText() == "精简"
    n_one = len(w._ti_rows)
    assert n_one > n_min, "单点穷举应比全局精简用例多"
    # _opts 把单点覆盖带进 GenOptions；build 该信号用例数 == 单点穷举数
    opts = w._opts([sig.out_name])
    assert opts.sig_cov.get(name_low) == "exhaustive"
    res = generator.build(w.wb, opts)
    nv = next(st["n_vectors"] for _l, st in res["blocks"] if st["out_name"] == sig.out_name)
    assert nv == n_one
    # 切回"跟随全局" → 从 _sig_cov 删除、用例数回到精简
    w.sig_cov_combo.setCurrentText("跟随全局")
    assert name_low not in w._sig_cov
    assert len(w._ti_rows) == n_min


def test_gui_sig_cov_isolated_to_signal(qapp, tmp_path_factory):
    """⭐ 用户报的 bug：切到别的信号，单点档不应跨信号串——A 设穷举，B 仍跟随全局精简。"""
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("sigcoviso") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    sigA = next(s for s in w.signals if s.out_name == "d_logic_bt_lp_lna_agc[2:0]")
    sigB = next(s for s in w.signals if s.out_name != sigA.out_name
                and not isinstance(s, excel_model.MuxGroup))
    rowA = next(r for r in range(w.table.rowCount())
                if w._sig_of_row(r).out_name == sigA.out_name)
    rowB = next(r for r in range(w.table.rowCount())
                if w._sig_of_row(r).out_name == sigB.out_name)
    w.coverage.setCurrentText("精简")
    # A 设单点穷举
    w.on_row_focus(rowA, gui.COL_K, -1, -1)
    w.sig_cov_combo.setCurrentText("穷举")
    assert w._sig_cov.get(sigA.out_name.lower()) == "exhaustive"
    # 切到 B：下拉回到"跟随全局"，B 不被 A 的单点档带动
    w.on_row_focus(rowB, gui.COL_K, -1, -1)
    assert w.sig_cov_combo.currentText() == "跟随全局"
    assert sigB.out_name.lower() not in w._sig_cov
    # 改全局到全面 → B 跟着变、A 仍是单点穷举（回 A 验证）
    res_min = generator.build(w.wb, w._opts([sigA.out_name, sigB.out_name],
                                            None))  # 全局仍精简
    # A 是单点穷举、B 跟随全局精简
    cmap = {st["out_name"]: st["n_vectors"] for _l, st in res_min["blocks"]}
    # A 用例数 = 穷举（独立重算对照）
    a_ex = generator.build(w.wb, generator.GenOptions(mode="max", exhaustive=True,
                                                      signals=[sigA.out_name]))
    a_ex_n = next(st["n_vectors"] for _l, st in a_ex["blocks"])
    assert cmap[sigA.out_name] == a_ex_n


def test_gui_sig_cov_global_clears_displayed_override(qapp, tmp_path_factory):
    """⭐用户报的bug + 拍板修法(Option A)：单点档会暗中盖过全局下拉(还从磁盘恢复)→全局看着'失效'。
    修=改【与当前信号同类】的全局下拉时,清掉【当前正在看】那个信号的单点档,让全局立刻生效；
    其它没在看的信号保留各自单点档；max_tests / 另一类下拉 不清档。"""
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("sigcovglobal") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    A = "d_logic_bt_lp_lna_agc[2:0]"; B = "d_logic_bt_lp_reserve"
    def select(name):
        r = next(i for i in range(w.table.rowCount()) if w._sig_of_row(i).out_name == name)
        w.on_row_focus(r, gui.COL_K, -1, -1)
    w.coverage.setCurrentText("精简")
    # A、B 都设单点穷举（模拟探索/磁盘恢复后两个信号都有 override）
    select(A); w.sig_cov_combo.setCurrentText("穷举")
    select(B); w.sig_cov_combo.setCurrentText("穷举")
    select(A)
    n_exh = len(w._ti_rows)
    # 看着 A 拧全局下拉到【不同】档位（同值 Qt 不发信号；必须 setCurrentText 真改变，self.sender() 才辨控件）
    w.coverage.setCurrentText("全面")
    assert A.lower() not in w._sig_cov, "看A时拧全局→应清掉A的单点档"
    assert len(w._ti_rows) < n_exh, "清档后A应跟随全局全面、用例变少"
    assert B.lower() in w._sig_cov, "没在看的B应保留单点档"
    # 点开 B：单点档仍在，下拉回显穷举
    select(B)
    assert w._sig_cov.get(B.lower()) == "exhaustive"
    assert w.sig_cov_combo.currentText() == "穷举"
    # max_tests 变化不清档
    select(A); w.coverage.setCurrentText("精简"); w.sig_cov_combo.setCurrentText("穷举")
    w.max_tests.setValue(64)
    assert w._sig_cov.get(A.lower()) == "exhaustive", "改 max_tests 不该清单点档"


def test_gui_sig_cov_session_only_not_persisted(qapp, tmp_path_factory, monkeypatch):
    """⭐用户报的真bug根因+修法：单点档是【会话内临时档】，不存盘、不恢复。
    否则上次留的单点档(如设成精简)会被静默恢复、暗中盖过全局下拉→'刚开GUI改全局对某些信号无效'。"""
    from dreg_verify import gui
    monkeypatch.setattr(gui, "EDITS_PATH", str(tmp_path_factory.mktemp("sigcovp") / "edits.json"))
    path = tmp_path_factory.mktemp("sigcovs") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    A = "d_logic_bt_lp_lna_agc[2:0]"
    # 会话1：全局穷举，把 A 单点设成精简（= 用户上一会话往下调的残留），存盘
    w1 = gui.MainWindow(); w1.path_edit.setText(str(path)); w1.on_load()
    w1.coverage.setCurrentText("穷举")
    r = next(i for i in range(w1.table.rowCount()) if w1._sig_of_row(i).out_name == A)
    w1.on_row_focus(r, gui.COL_K, -1, -1)
    w1.sig_cov_combo.setCurrentText("精简")
    assert w1._sig_cov.get(A) == "min"
    w1._persist_edits()
    # 存盘文件里不该有 sig_cov（哪怕被写了别的编辑）
    import json, os
    if os.path.exists(gui.EDITS_PATH):
        bucket = json.load(open(gui.EDITS_PATH)).get(str(path), {})
        assert "sig_cov" not in bucket, "单点档不该进存盘桶"
    # 会话2：全新打开（什么都没干）→ 不该恢复任何单点档；全局穷举对 A 生效
    w2 = gui.MainWindow(); w2.path_edit.setText(str(path)); w2.on_load()
    assert w2._sig_cov == {}, "单点档不该从磁盘恢复（会话内临时）"
    w2.coverage.setCurrentText("穷举")
    r = next(i for i in range(w2.table.rowCount()) if w2._sig_of_row(i).out_name == A)
    w2.on_row_focus(r, gui.COL_K, -1, -1)
    n_exh = len(w2._ti_rows)
    w2.coverage.setCurrentText("精简")
    w2.on_row_focus(r, gui.COL_K, -1, -1)   # 重新读
    # 验证 A 现在完全跟随全局（穷举 > 精简），没有被残留单点档钉死
    assert n_exh > len(w2._ti_rows), "新会话 A 应跟随全局(穷举>精简)，不被上次残留单点档盖住"


def test_gui_sig_cov_bucket_sig_cov_ignored(qapp, tmp_path_factory):
    """旧桶里残留的 sig_cov 段一律忽略（会话内临时档不恢复）。"""
    from dreg_verify import gui
    path = tmp_path_factory.mktemp("sigcovr") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    real = next(s for s in w.signals if s.out_name == "d_logic_bt_lp_lna_agc[2:0]")
    bucket = {"sig_cov": {real.out_name.lower(): "exhaustive"}}   # 旧版残留
    w._apply_edits_bucket(bucket)
    assert w._sig_cov == {}, "旧桶 sig_cov 应被忽略"


# ───────────── 导出范围：仅负向 / 仅正向 ─────────────
def test_gui_export_negative_only(qapp, wb, tmp_path_factory):
    """'仅负向'导出：override 只剩负向向量、无负向的信号被略过；build 出的全是负向。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "exp1")
    w._load_test_items(sig)
    w.on_ti_add_neg_all()
    name = sig.out_name
    ov_neg = w._vector_overrides(negative_only=True)
    assert name.lower() in ov_neg
    assert ov_neg[name.lower()] and all(v.is_negative for v in ov_neg[name.lower()])
    # _negative_signal_names 从勾选里挑出含负向的
    assert w._negative_signal_names([name]) == [name]
    # build(仅负向) → 渲染出的向量全为负向
    res = generator.build(w.wb, w._opts([name], negative_only=True))
    s = res["summary"]
    assert s["n_vectors"] > 0 and s["n_vectors"] == s["n_negative"]


def test_gui_export_negative_only_signal_without_neg_skipped(qapp, wb, tmp_path_factory):
    """没有负向的信号不会出现在'仅负向'范围里。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "exp2")
    w._load_test_items(sig)                        # 只看，不加负向
    w.ti_table.item(0, 0).setText("0x1")           # 自定义但无负向
    assert w._vector_overrides(negative_only=True) is None
    assert w._negative_signal_names([sig.out_name]) == []


# ───────────── 可验证性报告（取代旧'覆盖诊断'按钮） ─────────────
def test_report_verifiability_structure(wb):
    """report() 返回 verifiability：counts 四类齐全，且与逐信号 status 计数一致。"""
    from collections import Counter
    rep = generator.report(wb, generator.GenOptions(top_output_only=False))
    v = rep["verifiability"]
    assert set(v["counts"]) >= {"clean", "wire-fallback", "unresolved", "parse-err"}
    assert sum(v["counts"].values()) == len(v["signals"]) == len(wb.logic)
    assert all(s["status"] in v["counts"] for s in v["signals"])
    c = Counter(s["status"] for s in v["signals"])
    for k, n in c.items():
        assert v["counts"][k] == n


def test_html_report_has_verifiability_tab(wb, tmp_path):
    """HTML 报告含第④'可验证性'标签页与对应行。"""
    from dreg_verify import cli
    rep = generator.report(wb, generator.GenOptions(top_output_only=False))
    path = tmp_path / "r.html"
    cli.write_report(str(path), rep, "synthetic.xlsx")
    html = path.read_text(encoding="utf-8")
    assert "可验证性" in html and 'data-tab="ver"' in html
    assert 'id="ver"' in html and "vrow" in html


def test_csv_report_writes_verifiability(wb, tmp_path):
    """CSV 报告额外写出 *_verifiability.csv。"""
    from dreg_verify import cli
    rep = generator.report(wb, generator.GenOptions(top_output_only=False))
    path = tmp_path / "r.csv"
    written = cli.write_report(str(path), rep, "synthetic.xlsx")
    verif = [p for p in written if p.endswith("_verifiability.csv")]
    assert verif, "应生成可验证性 CSV"
    content = open(verif[0], encoding="utf-8-sig").read()
    assert "可验证性" in content


# ───────────── 真值表标注表达式字母 A/B/C（#1） ─────────────
def test_gui_truthtable_short_headers_and_inputs_table(qapp, wb, tmp_path_factory):
    """真值表行表头=信号名(2026-06-03 用户拍板：不用字母)；字母→信号/角色/类型/驱动 在上方『输入信号』表。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "g_letters")   # (A?C:B)&(~J)
    w._load_test_items(sig)
    # 真值表行表头：直接用信号名，控制位带 "(控制)"；不再是字母
    headers = [w.ti_table.verticalHeaderItem(i).text() for i in range(len(w._ti_groups))]
    assert "d_bt_lp_linelocal_mode_ctrl (控制)" in headers      # A=三元条件 → 控制位
    assert "d_bt_lp_iddq (控制)" in headers                     # J=门控 → 控制位
    assert "d_bt_lp_bt_mode_sel_local" in headers               # C=数据位无标记
    assert "d_bt_lp_bt_mode_sel" in headers                     # B=数据位无标记
    assert all(h.split()[0] not in ("A", "B", "C", "J") for h in headers)   # 不再是裸字母
    # 字母对照在 tooltip 里(对照表达式用)
    tips = [w.ti_table.verticalHeaderItem(i).toolTip() for i in range(len(w._ti_groups))]
    assert any("表达式里的 A" in t for t in tips)
    assert any("表达式里的 J" in t for t in tips)
    # 加粗行=控制位
    bold = {w._group_letters(g) for i, g in enumerate(w._ti_groups)
            if w.ti_table.verticalHeaderItem(i).font().bold()}
    assert bold == {"A", "J"}
    # 头部精简：等式在，难读的字母图例没了
    assert "(A?C:B)&(~J)" in w.ti_header.text()
    assert "字母对应" not in w.ti_header.text()
    # 『输入信号』表：每输入一行，列= 字母/信号(位宽)/角色/类型/驱动
    ti = w.ti_inputs
    assert ti.columnCount() == 5
    assert ti.rowCount() == len(w._ti_groups)
    rowmap = {ti.item(r, 0).text(): r for r in range(ti.rowCount())}
    assert set(rowmap) >= {"A", "B", "C", "J"}
    assert "d_bt_lp_linelocal_mode_ctrl" in ti.item(rowmap["A"], 1).text()
    assert "控制" in ti.item(rowmap["A"], 2).text() and "控制" in ti.item(rowmap["J"], 2).text()
    assert "数据" in ti.item(rowmap["C"], 2).text() and "数据" in ti.item(rowmap["B"], 2).text()
    # 类型↔驱动机制自洽：RO→force ENV_RF.，RW→RF_WRITE
    for L in ("A", "B", "C", "J"):
        kind, drive = ti.item(rowmap[L], 3).text(), ti.item(rowmap[L], 4).text()
        assert kind in ("RO", "RW", "UNKNOWN", "?")
        if kind == "RO":
            assert drive.startswith("force ENV_RF.")
        elif kind == "RW":
            assert drive.startswith("RF_WRITE")


def test_report_table_inputs_carry_letters(wb):
    """report() 的真值表 inputs 每项带 letters（A/B/C…），供 HTML/CSV 对照表达式。
    2026-06-10 起行序=寄存器地址+bit 位（for_test 生成规则），不再是表达式字母序。"""
    rep = generator.report(wb, generator.GenOptions(top_output_only=False))
    t = next(x for x in rep["tables"] if x["signal"] == "d_logic_bt_lp_reserve")
    assert sorted(i["letters"] for i in t["inputs"]) == ["A", "B", "C", "J"]   # 不丢输入
    assert [i["letters"] for i in t["inputs"]] == ["B", "A", "C", "J"]          # 地址序
    assert all("label" in i for i in t["inputs"])   # 原 label 不破坏


def test_html_report_maps_letters_to_signals(wb, tmp_path):
    """HTML 报告真值表行表头出现 '字母 → 信号' 的映射。"""
    from dreg_verify import cli
    rep = generator.report(wb, generator.GenOptions(top_output_only=False))
    path = tmp_path / "r.html"
    cli.write_report(str(path), rep, "synthetic.xlsx")
    html = path.read_text(encoding="utf-8")
    assert "A → d_bt_lp_linelocal_mode_ctrl" in html
    assert "J → d_bt_lp_iddq" in html


def test_html_report_truth_table_value_filter(wb, tmp_path):
    """⭐真值表「按值筛选」(2026-06-12)：穷举真值表列多时，designer 在第0列每个『变化的输入』旁的输入框里
    像 Excel 一样手填值 → 只留该值匹配的测试列（整列 display:none）。验证输入框/列标 data-c/计数+清除/
    CSS+JS(含边打边筛 input)通道都在，且真值表块仍只在 JSON blob 里（不预渲进静态壳）。"""
    import json
    import re
    from dreg_verify import cli
    rep = generator.report(wb, generator.GenOptions(top_output_only=False))
    path = tmp_path / "r.html"
    cli.write_report(str(path), rep, "synthetic.xlsx")
    raw = path.read_text(encoding="utf-8")

    # ②③ 共用的真值表数据 blob：解析出来逐项查（HTML 在 item.h 里，未转义）
    m = re.search(r'<script type="application/json" id="tt-data">(.*?)</script>', raw, re.S)
    assert m, "缺少 tt-data blob"
    items = json.loads(m.group(1))
    blk = next((it["h"] for it in items if 'class="ttf"' in it["h"]), None)
    assert blk, "应至少有一个真值表块（多列+变化输入）挂上按值筛选输入框"
    # 是手填输入框（不是下拉），带 data-ri，varying 输入才有
    assert re.search(r'<input class="ttf" type="text" data-ri="\d+"', blk)
    assert "<select" not in blk and "(全部)" not in blk     # 已不是下拉
    assert 'data-c="0"' in blk                             # 每列每格带列标（整列一起显隐）
    assert 'class="ttfcount"' in blk and 'class="ttfclr"' in blk   # 计数 + 清除
    assert 'class="inrow" data-ri=' in blk                 # 输入行可被 JS 定位读取本列取值

    # 真值表块仍只在 blob、不在静态壳；CSS/JS 通道齐
    shell = re.sub(r'<script type="application/json"[^>]*>.*?</script>', "", raw, flags=re.S)
    assert '<div class="ttblock">' not in shell
    assert '.tt .ttfx{display:none}' in shell               # 整列隐藏的 CSS
    assert 'function filterBlock' in shell                  # 筛选核心
    assert "contains('ttf')" in shell                       # 委托接上 .ttf 输入框
    assert "contains('ttfclr')" in shell                    # click 委托接上「清除」
    assert "addEventListener('input'" in shell              # 边打边筛（live 输入）


# ───────────── 面板可自由拖动大小（#2：FlowLayout + 不可塌陷） ─────────────
def test_gui_panes_are_resizable(qapp, wb, tmp_path_factory):
    """两个工具条用 FlowLayout（最小宽度≈单按钮、可换行），主分割条不可把面板拖没。"""
    from PySide6 import QtWidgets
    gui, w, _sig = _reserve_window(tmp_path_factory, "g_resize")
    sp = w.centralWidget().findChild(QtWidgets.QSplitter)
    assert not sp.childrenCollapsible()              # 拖到底也不会被吞掉
    assert sp.widget(0).minimumWidth() <= 260        # 左面板能缩到较窄
    flows = [c.layout() for c in w.findChildren(QtWidgets.QWidget)
             if isinstance(c.layout(), gui.FlowLayout)]
    # 排查(旧)的左批量条 + 右编辑条 + 各 SignalView(Topout/子视图) 的批量条/编辑条都用 FlowLayout；
    # 至少有这 2 条，且【每一条】都要可缩到单按钮宽、窄屏换行（断言更强=覆盖所有工具条，不只 2 条）。
    assert len(flows) >= 2
    for fl in flows:
        # 最小宽度≈单个最宽按钮，而不是所有按钮之和（否则 splitter 拖不动）
        assert 0 < fl.minimumSize().width() < 260
        # 窄到 180px 时会换行（高度 > 单行）
        assert fl.heightForWidth(180) >= fl.heightForWidth(2000)


# ───────────── Batch A/B/C：安全 / 清晰 / 效率 回归 ─────────────
def test_gui_cell_text_binary_roundtrip(qapp):
    """B3：小字段显二进制(0bXXXX)，且能被 _parse_int 原样读回(round-trip)。"""
    from dreg_verify import gui
    M = gui.MainWindow
    assert M._cell_text(5, 3) == "0b101"
    assert M._cell_text(1, 1) == "1"
    assert M._cell_text(0xAB, 8) == "0xAB"
    assert M._parse_int("0b101") == 5
    assert M._parse_int("0b0011") == 3
    assert M._parse_int("0xAB") == 0xAB        # 0b 之后仍能解析裸/带前缀 hex
    # 审查修复#3：'0bc' 这类裸 hex(0b 后非全 0/1)仍按 16 进制，不被新 0b 分支吃掉
    assert M._parse_int("0bc") == 0xBC
    assert M._parse_int("0bf") == 0xBF
    assert M._parse_int("0b11") == 3           # 全 0/1 → 二进制


def test_gui_header_tooltips_present(qapp, wb, tmp_path_factory):
    """B2：'负向'/'状态'表头与状态单元格带说明 tooltip。"""
    from dreg_verify import gui
    _g, w, _sig = _reserve_window(tmp_path_factory, "g_tips")
    assert "负向" in w.table.horizontalHeaderItem(gui.COL_NEG).toolTip()
    assert w.table.horizontalHeaderItem(gui.COL_STATUS).toolTip()
    assert w.table.item(0, gui.COL_STATUS).toolTip()


def test_gui_check_selected_rows_and_scope(qapp, wb, tmp_path_factory):
    """C1/C2：多选行→『勾选选中行』一次勾上；有勾选时批量作用域取勾选行。"""
    from PySide6 import QtCore
    gui, w, _sig = _reserve_window(tmp_path_factory, "g_check")
    w.table.clearSelection()
    sm = w.table.selectionModel()
    for r in (0, 2):
        sm.select(w.table.model().index(r, 0),
                  QtCore.QItemSelectionModel.Select | QtCore.QItemSelectionModel.Rows)
    w.on_check_selected_rows()
    checked = w._checked_rows()
    assert set(checked) >= {0, 2}
    assert sorted(w._scope_rows()) == sorted(checked)


def test_gui_bulk_neg_scoped_to_checked(qapp, wb, tmp_path_factory):
    """C2：勾了某信号后『全部加负向』只作用于已勾选信号，不波及其它。"""
    from PySide6 import QtCore
    gui, w, _sig = _reserve_window(tmp_path_factory, "g_scope")
    target = next(r for r in range(w.table.rowCount())
                  if w._sig_of_row(r).out_name == "d_logic_bt_lp_reserve")
    w.table.item(target, gui.COL_SEL).setCheckState(QtCore.Qt.Checked)
    w.on_all_signals_neg(True)
    assert w._signal_has_negative("d_logic_bt_lp_reserve")
    others = [w._sig_of_row(r).out_name.lower() for r in range(w.table.rowCount()) if r != target]
    assert not any(w._signal_has_negative(o) for o in others)


def test_gui_del_neg_confirms_named(qapp, wb, tmp_path_factory, monkeypatch):
    """A3：删除含命名负向时先确认——选 No 不删、选 Yes 才删。"""
    from PySide6 import QtWidgets
    gui, w, sig = _reserve_window(tmp_path_factory, "g_delneg")
    w._load_test_items(sig)
    w.on_ti_add_neg_all()
    negcol = next(c for c, rd in enumerate(w._ti_rows) if rd.get("is_negative"))
    w._ti_rows[negcol]["name"] = "MYNEG"
    assert w._signal_has_named_negative(w._ti_name_low)
    monkeypatch.setattr(QtWidgets.QMessageBox, "question",
                        staticmethod(lambda *a, **k: QtWidgets.QMessageBox.No))
    w.on_ti_del_neg()
    assert any(rd.get("is_negative") for rd in w._ti_rows)     # No → 保住
    monkeypatch.setattr(QtWidgets.QMessageBox, "question",
                        staticmethod(lambda *a, **k: QtWidgets.QMessageBox.Yes))
    w.on_ti_del_neg()
    assert not any(rd.get("is_negative") for rd in w._ti_rows)  # Yes → 删掉


def test_gui_confirm_dup_labels_no_dups_is_silent(qapp):
    """A1：无重复标号时 _confirm_dup_labels 直接放行(不弹框)。"""
    from dreg_verify import gui
    w = gui.MainWindow()
    assert w._confirm_dup_labels({"dup_labels": []}) is True
    assert w._confirm_dup_labels({}) is True


def test_skipped_detail_text_lists_names_and_reasons():
    """生成完成弹窗：跳过的信号在『显示详情』里逐个列出名字 + 不可驱动原因。"""
    from dreg_verify.gui import _skipped_detail_text
    skipped = [
        ("d_logic_pll_gear_shift_mux_sel", "130",
         [("A", "close_ready_flag", "表中查无字段，按 wire 处理(force 信号名)")]),
        ("d_logic_pll_n2", "131",
         [("B", "pll_n1", "输入是内部信号(top_output=0)"), ("C", "int_n", "")]),
    ]
    text = _skipped_detail_text(skipped)
    assert "d_logic_pll_gear_shift_mux_sel" in text
    assert "close_ready_flag" in text and "wire" in text
    assert "d_logic_pll_n2" in text and "pll_n1" in text
    assert "C = int_n：不可驱动" in text       # 空原因 → 兜底文案
    assert "--include-risky" in text           # 告诉用户如何强制生成


def test_gui_bulk_scope_excludes_hidden_checked(qapp, wb, tmp_path_factory):
    """审查修复#2：勾了但被筛掉(隐藏)的行不进批量负向作用域，避免改到看不见的信号。"""
    from PySide6 import QtCore
    gui, w, _sig = _reserve_window(tmp_path_factory, "g_hidden")
    r0 = 0
    w.table.item(r0, gui.COL_SEL).setCheckState(QtCore.Qt.Checked)
    assert r0 in w._scope_rows()                 # 可见时在作用域内
    w.table.setRowHidden(r0, True)
    assert r0 not in w._checked_rows()           # 隐藏后不算"已勾选可见"
    assert r0 not in w._scope_rows()             # 也不进批量作用域(回退到其它可见行)


def test_gui_truthtable_active_column_highlight(qapp, wb, tmp_path_factory):
    """选中某测试列→该列高亮(淡蓝)；换列旧列恢复；负向列被选=琥珀+高亮叠加(列多横滚时不丢位置)。"""
    gui, w, sig = _reserve_window(tmp_path_factory, "g_hl")
    w._load_test_items(sig)
    w.on_ti_add_neg_all()                       # 让后面有负向列
    def bg(c):
        return w.ti_table.item(0, c).background().color().name().lower()
    w.ti_table.setCurrentCell(0, 0)             # 选中正向列 0
    assert w._ti_hl_col == 0
    assert bg(0) == gui.HL_BG.name().lower()
    negc = next(c for c, rd in enumerate(w._ti_rows) if rd.get("is_negative"))
    w.ti_table.setCurrentCell(0, negc)          # 选中负向列 → 叠加色
    assert w._ti_hl_col == negc
    assert bg(negc) == gui.HL_NEG_BG.name().lower()
    assert bg(0) != gui.HL_BG.name().lower()    # 旧正向列已恢复
    others = [c for c, rd in enumerate(w._ti_rows) if rd.get("is_negative") and c != negc]
    if others:
        assert bg(others[0]) == gui.NEG_BG.name().lower()   # 未选的负向列仍普通琥珀


def test_gui_copy_shortcut_is_widget_scoped(qapp, wb, tmp_path_factory):
    """审查修复#1：复制列快捷键挂在真值表上且为 WidgetShortcut(编辑单元格时不触发)，
    而不是挂按钮的 WindowShortcut。"""
    from PySide6 import QtCore, QtGui
    gui, w, _sig = _reserve_window(tmp_path_factory, "g_copysc")
    scs = [s for s in w.ti_table.findChildren(QtGui.QShortcut)
           if s.key() == QtGui.QKeySequence("Ctrl+D")]
    assert scs, "应在 ti_table 上挂 Ctrl+D 的 QShortcut"
    assert any(s.context() == QtCore.Qt.WidgetShortcut for s in scs)


def test_gui_cascade_help_is_builtin_dialog(qapp, wb, tmp_path_factory):
    """级联 ? → 程序内置帮助窗(QTextBrowser 渲染 .md)，不再调外部编辑器(如 VS Code)打开文件。"""
    from PySide6 import QtWidgets
    gui, w, _sig = _reserve_window(tmp_path_factory, "g_chelp")
    w._open_cascade_doc()
    dlg = w._cascade_doc_dlg
    assert dlg is not None and dlg.isVisible()           # GUI 内弹窗，而不是 QDesktopServices.openUrl
    view = dlg.findChild(QtWidgets.QTextBrowser)
    assert view is not None
    text = view.toPlainText()
    assert "展开上游" in text and "force" in text         # 内容已渲染(完整 .md 或内置摘要兜底)
    # 再点一次 → 复用同一个窗口(重读文件刷新内容)，不堆窗口
    w._open_cascade_doc()
    assert w._cascade_doc_dlg is dlg
    dlg.close()


# ───────────── 第二十六轮：一键清空(logic) + 完整配置导入导出 + 信号勾选/上限持久化 ─────────────
def test_gui_clear_logic_zero_tests(qapp, wb, tmp_path_factory, monkeypatch):
    """logic「一键清空」= 空 override → build 跳过给「清空」原因；「重新生成」可恢复。"""
    from PySide6 import QtWidgets
    monkeypatch.setattr(QtWidgets.QMessageBox, "question",
                        staticmethod(lambda *a, **k: QtWidgets.QMessageBox.Yes))
    gui, w, sig = _reserve_window(tmp_path_factory, "clr")
    w._load_test_items(sig)
    assert len(w._ti_rows) >= 1
    w.on_ti_clear()
    nl = sig.out_name.lower()
    assert w._ti_rows == []
    assert nl in w._edited and w._edited[nl]["rows"] == []
    res = generator.build(w.wb, w._opts([sig.out_name]))
    assert sig.out_name not in {st.get("out_name") for _l, st in res["blocks"]}
    assert any(n == sig.out_name and "清空" in "; ".join(str(x) for *_h, x in rs)
               for n, _a, rs in res["skipped"])
    w.on_ti_regen()
    assert len(w._ti_rows) >= 1


def test_gui_signal_selection_persisted(qapp, tmp_path_factory, monkeypatch):
    """信号勾选(COL_SEL)跨 GUI 重开自动恢复（第二十六轮：进 EDITS 桶）。"""
    from PySide6 import QtCore
    from dreg_verify import gui
    monkeypatch.setattr(gui, "EDITS_PATH", str(tmp_path_factory.mktemp("selp") / "edits.json"))
    path = tmp_path_factory.mktemp("selw") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    names = ["d_logic_bt_lp_reserve", "d_logic_bt_lp_lna_agc[2:0]"]
    w1 = gui.MainWindow(); w1.path_edit.setText(str(path)); w1.on_load()
    for nm in names:
        r = next(i for i in range(w1.table.rowCount()) if w1._sig_of_row(i).out_name == nm)
        w1.table.item(r, gui.COL_SEL).setCheckState(QtCore.Qt.Checked)
    assert set(w1._collect_checked()) == set(names)
    w1.close()
    w2 = gui.MainWindow(); w2.path_edit.setText(str(path)); w2.on_load()
    assert set(w2._collect_checked()) == set(names)
    w2.close()


def test_gui_max_tests_persisted(qapp, tmp_path_factory, monkeypatch):
    """用例上限 max_tests 改动 → 写进 settings（第二十六轮；__init__ 恢复走与 coverage 同款
    pytest 守卫，恢复路径由 coverage 那套已验证的 settings 机制保证）。"""
    from dreg_verify import gui
    store = {}
    monkeypatch.setattr(gui, "_load_settings", lambda: dict(store))
    monkeypatch.setattr(gui, "_save_settings", lambda d: store.update(d))
    path = tmp_path_factory.mktemp("mtw") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    w1 = gui.MainWindow(); w1.path_edit.setText(str(path)); w1.on_load()
    w1.max_tests.setValue(99)
    assert store.get("max_tests") == 99
    w1.close()


def test_full_config_export_import_roundtrip(qapp, tmp_path_factory, monkeypatch):
    """完整配置导出→导入：信号勾选/全局档/上限/探针前缀/强制force 全部还原（第二十六轮）。"""
    import json
    from PySide6 import QtCore
    from dreg_verify import gui
    monkeypatch.setattr(gui, "EDITS_PATH", str(tmp_path_factory.mktemp("cfgp") / "edits.json"))
    monkeypatch.setattr(gui.QtWidgets.QMessageBox, "information", staticmethod(lambda *a, **k: None))
    path = tmp_path_factory.mktemp("cfgw") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    nm = "d_logic_bt_lp_reserve"
    r = next(i for i in range(w.table.rowCount()) if w._sig_of_row(i).out_name == nm)
    w.table.item(r, gui.COL_SEL).setCheckState(QtCore.Qt.Checked)
    w.coverage.setCurrentText("全面"); w.coverage_mux.setCurrentText("穷举")
    w.max_tests.setValue(77)
    w.include_risky_chk.setChecked(True)        # 缺前缀强制生成（pytest 默认不勾）
    w._probe_prefixes = {"d_foo": "U_X.U_Y"}
    w._force_signals = {"d_bar"}
    cfg = tmp_path_factory.mktemp("cfgout") / "cfg.json"
    monkeypatch.setattr(gui.QtWidgets.QFileDialog, "getSaveFileName",
                        staticmethod(lambda *a, **k: (str(cfg), "")))
    w.on_export_edits()
    saved = json.load(open(str(cfg), encoding="utf-8"))
    assert saved["dreg_verify_config"] == 2
    assert saved["global"]["max_tests"] == 77
    assert saved["global"]["include_risky"] is True       # 随配置带走
    assert saved["excel"] == "synthetic_dreg.xlsx"        # 源表文件名
    assert saved["excel_path"] == str(path)               # 源表全路径
    assert nm in saved["signals_checked"]
    w.close()
    # 新窗口导入 → 全部还原
    w2 = gui.MainWindow(); w2.path_edit.setText(str(path)); w2.on_load()
    assert not w2.include_risky_chk.isChecked()            # 导入前=默认不勾
    monkeypatch.setattr(gui.QtWidgets.QFileDialog, "getOpenFileName",
                        staticmethod(lambda *a, **k: (str(cfg), "")))
    w2.on_import_edits()
    assert set(w2._collect_checked()) == {nm}
    assert w2.max_tests.value() == 77
    assert w2.coverage.currentText() == "全面" and w2.coverage_mux.currentText() == "穷举"
    assert w2._probe_prefixes.get("d_foo") == "U_X.U_Y"
    assert "d_bar" in w2._force_signals
    assert w2.include_risky_chk.isChecked()                # include_risky 已套用
    w2.close()


def test_config_excel_mismatch_warns_but_imports(qapp, tmp_path_factory, monkeypatch):
    """配置记的源表文件名与当前加载的不一致 → 导入完成弹窗带 ⚠ 提示，但仍照常导入。
    缺前缀强制生成『缺键时保持当前』——旧配置不该翻动本机选择。"""
    import json
    from dreg_verify import gui
    monkeypatch.setattr(gui, "EDITS_PATH", str(tmp_path_factory.mktemp("mmp") / "edits.json"))
    captured = {}
    monkeypatch.setattr(gui.QtWidgets.QMessageBox, "information",
                        staticmethod(lambda parent, title, text, *a, **k: captured.update(msg=text)))
    path = tmp_path_factory.mktemp("mmw") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    # 手写一份『为别的表』导出的配置（excel 文件名不同）；含 max_tests，故缺 include_risky 时不翻动
    cfg = tmp_path_factory.mktemp("mmf") / "other.json"
    json.dump({"dreg_verify_config": 2, "excel": "some_other_chip.xlsx",
               "excel_path": "D:/elsewhere/some_other_chip.xlsx",
               "global": {"coverage_logic": "全面", "max_tests": 42},
               "signals_checked": []}, open(str(cfg), "w", encoding="utf-8"))
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    risky_before = w.include_risky_chk.isChecked()
    monkeypatch.setattr(gui.QtWidgets.QFileDialog, "getOpenFileName",
                        staticmethod(lambda *a, **k: (str(cfg), "")))
    w.on_import_edits()
    assert "some_other_chip.xlsx" in captured.get("msg", "")     # 弹窗提示了配错表
    assert "synthetic_dreg.xlsx" in captured.get("msg", "")
    assert w.max_tests.value() == 42                              # 仍照常导入
    assert w.include_risky_chk.isChecked() == risky_before        # 缺键 → 保持当前，不翻动
    w.close()


def test_legacy_edits_import_still_merges(qapp, tmp_path_factory, monkeypatch):
    """向后兼容：旧版 v1『测试项编辑』文件仍能导入(合并语义)，不碰勾选/全局。"""
    import json
    from dreg_verify import gui
    monkeypatch.setattr(gui, "EDITS_PATH", str(tmp_path_factory.mktemp("legp") / "edits.json"))
    monkeypatch.setattr(gui.QtWidgets.QMessageBox, "information", staticmethod(lambda *a, **k: None))
    path = tmp_path_factory.mktemp("legw") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    legacy = tmp_path_factory.mktemp("legf") / "old.json"
    json.dump({"dreg_verify_edits": 1,
               "edits": {"d_logic_bt_lp_reserve": [{"base_values": {}, "kind": "pos"}]}},
              open(str(legacy), "w", encoding="utf-8"))
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    monkeypatch.setattr(gui.QtWidgets.QFileDialog, "getOpenFileName",
                        staticmethod(lambda *a, **k: (str(legacy), "")))
    w.on_import_edits()
    assert "d_logic_bt_lp_reserve" in w._edited
    w.close()


def test_full_config_import_clears_stale_negatives(qapp, tmp_path_factory, monkeypatch):
    """导入完整配置 = 加载该状态：上一会话残留的负向勾选被清掉(权威同步，第二十六轮自审修)。"""
    from PySide6 import QtCore
    from dreg_verify import gui
    monkeypatch.setattr(gui, "EDITS_PATH", str(tmp_path_factory.mktemp("stp") / "edits.json"))
    monkeypatch.setattr(gui.QtWidgets.QMessageBox, "information", staticmethod(lambda *a, **k: None))
    path = tmp_path_factory.mktemp("stw") / "synthetic_dreg.xlsx"
    fixtures.build_workbook(str(path))
    # 导出一份"干净无负向"的配置
    w = gui.MainWindow(); w.path_edit.setText(str(path)); w.on_load()
    cfg = tmp_path_factory.mktemp("stout") / "clean.json"
    monkeypatch.setattr(gui.QtWidgets.QFileDialog, "getSaveFileName",
                        staticmethod(lambda *a, **k: (str(cfg), "")))
    w.on_export_edits()
    w.close()
    # 新会话：给某信号加负向，再导入干净配置 → 负向被清
    w2 = gui.MainWindow(); w2.path_edit.setText(str(path)); w2.on_load()
    nm = "d_logic_bt_lp_reserve"
    r = next(i for i in range(w2.table.rowCount()) if w2._sig_of_row(i).out_name == nm)
    w2.table.item(r, gui.COL_NEG).setCheckState(QtCore.Qt.Checked)
    assert w2._signal_has_negative(nm.lower())
    monkeypatch.setattr(gui.QtWidgets.QFileDialog, "getOpenFileName",
                        staticmethod(lambda *a, **k: (str(cfg), "")))
    w2.on_import_edits()
    rr = next(i for i in range(w2.table.rowCount()) if w2._sig_of_row(i).out_name == nm)
    assert w2.table.item(rr, gui.COL_NEG).checkState() == QtCore.Qt.Unchecked
    assert not w2._signal_has_negative(nm.lower())
    w2.close()


def test_cleared_logic_stays_zero_in_negative_only_export(qapp, tmp_path_factory, monkeypatch):
    """自审 Finding2：清空的 logic 信号在『仅负向』导出里仍是零用例(空 override 保留，不回退自动重生)。"""
    from PySide6 import QtWidgets
    monkeypatch.setattr(QtWidgets.QMessageBox, "question",
                        staticmethod(lambda *a, **k: QtWidgets.QMessageBox.Yes))
    gui, w, sig = _reserve_window(tmp_path_factory, "negclr")
    w._load_test_items(sig)
    w.on_ti_clear()
    nl = sig.out_name.lower()
    ov = w._vector_overrides(negative_only=True)
    assert ov is not None and ov.get(nl) == []          # 保留空 override，不在仅负向导出里略过
    res = generator.build(w.wb, w._opts([sig.out_name], negative_only=True))
    assert sig.out_name not in {st.get("out_name") for _l, st in res["blocks"]}
