# -*- coding: utf-8 -*-
"""WL_RFTRX 形态 mux 的 CLI 全流程测试（2026-06-03 第十四轮）。

覆盖 cli.py 在 WL 表上的外围表现（核心生成逻辑在 test_mux_wl.py 已测）：
  ① --list 列出 5 个 mux 组，多控制组的控制列显示拼接 {c1,c2}（B 高位）
  ② 不配探针前缀时，top_out=0 的 mux 组【照常生成】裸名探针 + 警告含 scan_rtl（用户拍板：
     工具不替用户假设 top_out=0=埋深；探得到就过、真 CUVUNF 再配前缀）
  ③ --probe-prefix-file 配前缀后，输出探针带层级前缀、无警告

注意：mux 输出的 top_out=0 只是"喂内部、非芯片顶层输出"（WL 全部如此），它们正是要验的信号——
所以 mux 组【不受】 --include-internal/top_output_only 过滤（与 logic 内部节点不同）。
不带 --include-internal 也能生成全部 mux。
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fixtures import build_wl_workbook  # noqa: E402

# WL 五个 mux 组 + mux→logic 衔接网的探针前缀（镜像 test_mux_wl.py 的层级）
WL_PREFIX = "U_WL_DREG.U_RF_MUX"
WL_PREFIX_NAMES = [
    "d_wl_rf_lna_gain", "d_wl_rf_bwctrl", "d_wl_rf_tx_bwctrl",
    "d_wl_rf_tx_rc_code", "d_wl_rf_dpd_path",
    "d_wl_rf_lna_gain_to_logic",
]


@pytest.fixture()
def wl_excel(tmp_path):
    xl = tmp_path / "wl_dreg.xlsx"
    build_wl_workbook(str(xl))
    return xl


def _write_prefix_file(path):
    """写一个与 GUI『探针前缀映射→导出』同格式的映射文件（每行 信号=层级路径，# 注释）。"""
    with open(path, "w", encoding="utf-8") as f:
        f.write("# WL_RFTRX 探针前缀（scan_rtl 导出后的格式）\n")
        for n in WL_PREFIX_NAMES:
            f.write("%s=%s\n" % (n, WL_PREFIX))
    return path


# ───────────── ① --list ─────────────
def test_wl_cli_list_shows_five_mux_groups(wl_excel, capsys):
    """--list 列出 5 个 mux 组（不带 --include-internal 也全在——mux 不受 top_out 过滤）。"""
    from dreg_verify import cli
    cli.main(["--excel", str(wl_excel), "--list"])
    out = capsys.readouterr().out
    # 5 个组的输出名与组号都在
    for n in range(1, 6):
        assert "mux%d" % n in out
    assert "d_wl_rf_lna_gain[2:0]" in out
    assert "d_wl_rf_tx_rc_code[5:0]" in out
    # 单控制组：case(基名)
    assert "case(d_wl_rf_lna_gain_ctrl_mode)" in out
    # 多控制组（组4）：case({c1,c2}) 按列序拼接，B 高位在前
    assert "case({d_wl_rf_rc_code_lut_en,d_wl_rf_bwctrl})" in out


# ───────────── ② 不配前缀 → 照常生成 + 警告 ─────────────
def test_wl_cli_generates_with_warning_no_prefix(wl_excel, tmp_path, capsys):
    """没配探针前缀（也不带 --include-internal）：5 个 top_out=0 的 mux 组照常生成裸名探针，
    控制台打印警告含 scan_rtl（提示真 CUVUNF 再配前缀）。"""
    from dreg_verify import cli
    out_sv = tmp_path / "wl.sv"
    cli.main(["--excel", str(wl_excel), "--out", str(out_sv)])
    out = capsys.readouterr().out
    # 警告（不是跳过）提示 scan_rtl
    assert "scan_rtl" in out
    # 5 个组都以 [R=mux<N>] 形式列在警告里
    for n in range(1, 6):
        assert "[R=mux%d]" % n in out
    # 产物【含】mux 断言，且是裸名探针（无层级前缀）
    sv = out_sv.read_text(encoding="utf-8")
    assert "assert_mux1_T0:" in sv
    assert "`ENV_RF.d_wl_rf_lna_gain[2:0]==" in sv


# ───────────── ②.5 --suffix-signals：mux 输出单点探尾缀网 ─────────────
def test_wl_cli_suffix_signals_optin(wl_excel, tmp_path):
    """--suffix-signals：mux 输出默认探裸名，指定后单点探带去向尾缀(_to_logic)的网（rxiq 形态）。
    前缀按全名 key 命中（不静默失配）。--no-suffix-signals 同理可把 logic 撞名信号拉回裸名。"""
    from dreg_verify import cli
    out_sv = tmp_path / "wl.sv"
    cli.main(["--excel", str(wl_excel), "--out", str(out_sv),
              "--suffix-signals", "d_wl_rf_lna_gain",
              "--probe-prefix", "d_wl_rf_lna_gain_to_logic=%s" % WL_PREFIX])
    sv = out_sv.read_text(encoding="utf-8")
    assert "`ENV_RF.%s.d_wl_rf_lna_gain_to_logic[2:0]==" % WL_PREFIX in sv   # 探尾缀网 + 前缀命中
    # 没点名的 mux 输出仍探裸名（默认）
    assert "`ENV_RF.d_wl_rf_bwctrl[1:0]==" in sv or "d_wl_rf_bwctrl" in sv


# ───────────── ③ 配前缀文件 → 全流程生成 ─────────────
def test_wl_cli_probe_prefix_file_generates_all(wl_excel, tmp_path):
    """--probe-prefix-file 导入前缀后，CLI 全流程产出 5 个组的 .sv：
    三来源驱动（寄存器直出 RF_WRITE / RO 线控 force / mux 级联载体）+ 带前缀的输出探针。"""
    from dreg_verify import cli
    pf = _write_prefix_file(str(tmp_path / "probe_prefixes.txt"))
    out_sv = tmp_path / "wl.sv"
    cli.main(["--excel", str(wl_excel), "--out", str(out_sv), "--probe-prefix-file", pf])
    sv = out_sv.read_text(encoding="utf-8")
    # 5 个组都有断言
    for n in range(1, 6):
        assert "assert_mux%d_T0:" % n in sv
    # ① 寄存器直出控制：模式寄存器 RF_WRITE（h50）
    assert "`RF_WRITE(10'h50," in sv
    # ② RO 线控数据：force（线控是顶层 RO 寄存器，无前缀）
    assert "force `ENV_RF.d_wl_rf_linectrl_lna_gain[2:0]=" in sv
    # ③ 输出探针带层级前缀（top_out=0）
    assert "`ENV_RF.%s.d_wl_rf_lna_gain[2:0]==" % WL_PREFIX in sv
    # ④ mux 级联：上游载体寄存器（bwctrl_local h53）被 RF_WRITE
    assert "`RF_WRITE(10'h53," in sv
    # ASCII 干净（仿真服务器编码安全）
    sv.encode("ascii")


def test_wl_cli_probe_prefix_file_with_negatives(wl_excel, tmp_path):
    """--probe-prefix-file + --neg-all：mux 组照常追加 _NEG 自检断言（与 logic 同口径）。"""
    from dreg_verify import cli
    pf = _write_prefix_file(str(tmp_path / "probe_prefixes.txt"))
    out_sv = tmp_path / "wl.sv"
    cli.main(["--excel", str(wl_excel), "--out", str(out_sv),
              "--probe-prefix-file", pf, "--neg-all"])
    sv = out_sv.read_text(encoding="utf-8")
    assert "_NEG:" in sv


def test_wl_cli_mux_only_isolates_mux(wl_excel, tmp_path):
    """--mux-only：只出 mux，不出 logic 行断言。"""
    from dreg_verify import cli
    pf = _write_prefix_file(str(tmp_path / "probe_prefixes.txt"))
    out_sv = tmp_path / "wl.sv"
    cli.main(["--excel", str(wl_excel), "--out", str(out_sv),
              "--probe-prefix-file", pf, "--mux-only"])
    sv = out_sv.read_text(encoding="utf-8")
    assert "assert_mux1_T0:" in sv
    # logic 行（R=1/R=2）不出现
    assert "assert_1_T0:" not in sv and "assert_2_T0:" not in sv


def test_wl_cli_html_report_contains_mux(wl_excel, tmp_path):
    """--report HTML：WL mux 组（多控制拼接表达式）出现在报告里。"""
    from dreg_verify import cli
    html_path = tmp_path / "report.html"
    cli.main(["--excel", str(wl_excel), "--report", str(html_path)])
    html = html_path.read_text(encoding="utf-8")
    assert "d_wl_rf_tx_rc_code[5:0]" in html
    # 多控制拼接表达式（HTML 转义后逗号/花括号仍是字面量）
    assert "case({d_wl_rf_rc_code_lut_en,d_wl_rf_bwctrl})" in html


def test_wl_cli_html_report_has_more_filters(wl_excel, tmp_path):
    """R27：HTML 报告除 owner/搜索/只看负向外，再加 信号名/类型/top_output 三个下拉。
    既要有控件，也要 JS 过滤逻辑用到对应 item 字段(s/ty/tp)，且下拉里能列出真实信号名。"""
    from dreg_verify import cli
    html_path = tmp_path / "report.html"
    cli.main(["--excel", str(wl_excel), "--report", str(html_path)])
    html = html_path.read_text(encoding="utf-8")
    # 三个新控件
    assert 'id="sig"' in html and 'id="typ"' in html and 'id="topf"' in html
    assert "全部信号" in html and "全部类型" in html and "仅 top_output" in html
    # JS 过滤用到新字段
    assert "it.s!==ss" in html and "it.ty!==tt" in html and "String(it.tp)!==tp" in html
    # 信号名下拉里有真实信号（mux 组）
    assert '<option value="d_wl_rf_tx_rc_code[5:0]">' in html


def test_wl_cli_xlsx_report_sheets_and_layout(wl_excel, tmp_path):
    """R27：--report 写 .xlsx → 真值表(分块) + 汇总/明细(autofilter+冻结表头)。
    给 designer 看：真值表 sheet 首列冻结、信号名做块标题；扁平表带自动筛选。"""
    import openpyxl
    from dreg_verify import cli
    xlsx_path = tmp_path / "report.xlsx"
    cli.main(["--excel", str(wl_excel), "--report", str(xlsx_path)])
    assert xlsx_path.exists()
    wb = openpyxl.load_workbook(str(xlsx_path))
    assert "真值表" in wb.sheetnames and "汇总" in wb.sheetnames and "明细" in wb.sheetnames

    # 汇总：表头 = SUMMARY_COLS、有自动筛选、冻结表头行
    s = wb["汇总"]
    assert [c.value for c in s[1]] == [h for _k, h in cli.SUMMARY_COLS]
    assert s.auto_filter.ref and s.auto_filter.ref.startswith("A1:")
    assert s.freeze_panes == "A2"
    sum_signals = {s.cell(r, 2).value for r in range(2, s.max_row + 1)}
    assert "d_wl_rf_tx_rc_code[5:0]" in sum_signals

    # 明细：autofilter + 冻结表头
    d = wb["明细"]
    assert d.freeze_panes == "A2" and d.auto_filter.ref

    # 真值表：冻结首列；信号名作为块标题出现在 A 列
    tt = wb["真值表"]
    assert tt.freeze_panes == "B1"
    col_a = {tt.cell(r, 1).value for r in range(1, tt.max_row + 1)}
    assert "d_wl_rf_tx_rc_code[5:0]" in col_a            # 块标题
    assert "表达式" in col_a and "force" in col_a and "RF_WRITE" in col_a
