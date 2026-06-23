# -*- coding: utf-8 -*-
"""生成期"自检闸门"回归(2026-06-16 R42)：把假绿做成块顶 ⚠。

检查1=探针读回自身输入(纵深防御，正常恒 0) / 检查2=K 列漏标位宽 / 检查3=RW 写值截断 /
SUF-2=mux 输出回填过自引用中心闸门(补上 _apply_mux_output_ref_suffix 这条原本裸赋值的路)。
"""

import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dreg_verify import expr as E             # noqa: E402
from dreg_verify import excel_model as M      # noqa: E402
from dreg_verify import generator as G        # noqa: E402
from dreg_verify.excel_model import (         # noqa: E402
    LogicSignal, MuxGroup, MuxCase, MuxCtrl, DregWorkbook)
from dreg_verify.resolver import InputBinding  # noqa: E402


def _logic_ws(rows, header_row=2):
    from openpyxl.utils import column_index_from_string
    wbx = openpyxl.Workbook()
    ws = wbx.active
    ws.title = "logic"
    ws.cell(row=header_row, column=column_index_from_string("K"), value="output")
    ws.cell(row=header_row, column=column_index_from_string("L"), value="expr")
    for i, rd in enumerate(rows, start=header_row + 1):
        for col, val in rd.items():
            ws.cell(row=i, column=column_index_from_string(col), value=val)
    return ws


def _logic(out_name, expr, inputs, out_width=1, suffix="", top_output=0):
    return LogicSignal(row=1, out_name=out_name, out_width=out_width, expr=expr,
                       suffix=suffix, top_output=top_output, notes="", owner="C",
                       assert_id="1", inputs=inputs)


def _mux_group(out_name, cases, ctrls, top_output=0):
    return MuxGroup(group_no="1", out_name=out_name, out_width=1,
                    ctrl_raw="", ctrl_base="", ctrl_width=1, owner="C",
                    top_output=top_output, cases=cases, ctrls=ctrls)


def _rw(base, width, slice_lsb, reg_lsb, reg_msb):
    return InputBinding(letter="A", raw=base, base=base, width=width, kind="RW",
                        address=0x40, reg_lsb=reg_lsb, reg_msb=reg_msb,
                        wire=base, found_in="tmm", slice_msb=None, slice_lsb=slice_lsb)


# ───────────────────────── SUF-2：mux 输出回填过自引用闸门 ─────────────────────────
def test_mux_output_self_ref_suffix_suppressed():
    """X 的 mux 输出在自己的 case 列引用 <X>_to_mux（= X 的输入前级网）→ 回填 ref_suffix 必须被
    自引用闸门挡住(否则探针=自己输入网=假绿/假红)。此前 _apply_mux_output_ref_suffix 裸赋值、
    MuxGroup 又没 _self_ref_suffixes 字段 → mux 路径与 R40 logic 自引用同盲区(SUF-2)。"""
    g = _mux_group("d_x",
                   cases=[MuxCase(row=3, case_raw="1'b1", input_raw="d_x_to_mux",
                                  input_base="d_x", input_width=1,
                                  input_msb=None, input_lsb=None)],
                   ctrls=[MuxCtrl("B", "sel", "sel", 1)])
    M._apply_mux_output_ref_suffix([], [g])
    assert g._self_ref_suffixes == {"_to_mux"}      # 标到自读后缀
    assert g.ref_suffix == ""                        # ⭐过闸门后被抑制，没回填成自己的输入网
    # 即便单点勾"探尾缀网"(append_to_mux 语义)，ref_suffix 已空 → 探针仍 bare、不撞自身输入
    g._append_to_logic = True
    assert g.rtl_base == "d_x"
    # 闸门后检查1 不报(纵深防御恒 0)
    assert G._selfaudit_probe_self_ref(g) is None


def test_mux_output_cross_ref_suffix_still_backfilled():
    """反面对照：mux 输出 Y 被【别的 mux】以 Y_to_mux 引用(真·下游 mux→mux 级联) → 照常回填 _to_mux，
    没误伤正常跨页网。"""
    y = _mux_group("d_y",
                   cases=[MuxCase(row=3, case_raw="1'b1", input_raw="cfg",
                                  input_base="cfg", input_width=1,
                                  input_msb=None, input_lsb=None)],
                   ctrls=[MuxCtrl("B", "sel", "sel", 1)])
    other = _mux_group("d_z",
                       cases=[MuxCase(row=4, case_raw="1'b1", input_raw="d_y_to_mux",
                                      input_base="d_y", input_width=1,
                                      input_msb=None, input_lsb=None)],
                       ctrls=[MuxCtrl("B", "sel2", "sel2", 1)])
    M._apply_mux_output_ref_suffix([], [y, other])
    assert y._self_ref_suffixes == set()
    assert y.ref_suffix == "_to_mux"


# ───────────────────────── 检查1：探针读回自身输入(纵深防御) ─────────────────────────
def test_check1_detects_probe_equals_own_input():
    """模拟"未来某条 ref_suffix 赋值路径绕过中心闸门"：探针全名 == <基名>_to_logic = 自己的输入前级网。
    正常路径恒 0，此处手工制造泄漏验证闸门能抓住。"""
    sig = _logic("d_x", "A", {"A": {"raw": "d_x_to_logic", "base": "d_x", "width": 1,
                                    "msb": None, "lsb": None}})
    sig._self_ref_suffixes = {"_to_logic"}
    sig.ref_suffix = "_to_logic"           # 泄漏：被错误回填成自读后缀
    sig._append_to_logic = True
    assert sig.rtl_base == "d_x_to_logic"  # 探针 == 自己输入网
    assert G._selfaudit_probe_self_ref(sig) is not None


def test_check1_read_logic_self_ref_gate_stays_zero():
    """真·自引用结构经 read_logic → ref_suffix 被闸门抑制成 bare → 探针≠自身输入 → 检查1 恒 0。"""
    ws = _logic_ws([
        {"K": "d_x", "L": "A?B:1'b0", "M": "", "N": 0,
         "A": "d_x_to_logic", "B": "d_y_to_logic"},   # 自引用 _to_logic
    ])
    sig = M.read_logic(ws)[0]
    assert sig.ref_suffix == ""                        # 闸门已抑制
    assert sig._self_ref_suffixes == {"_to_logic"}
    sig._append_to_logic = True
    assert G._selfaudit_probe_self_ref(sig) is None    # 探针 bare、不撞自身输入 → 恒 0


# ───────────────────────── 检查2：K 列漏标位宽 ─────────────────────────
def test_check2_missing_width_multibit_flagged():
    sig = _logic("d_sel", "A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                      "msb": 3, "lsb": 0}}, out_width=1)
    assert G._selfaudit_output_width(sig, E.parse("A"), E.Env({"A": 4})) is not None


def test_check2_explicit_width_not_flagged():
    sig = _logic("d_sel[3:0]", "A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                           "msb": 3, "lsb": 0}}, out_width=4)
    assert G._selfaudit_output_width(sig, E.parse("A"), E.Env({"A": 4})) is None


def test_check2_reduction_one_bit_not_flagged():
    """X = |A：自决宽 1 bit、K 列裸名也对 → 不误报(self_width 已把约简算成 1)。"""
    sig = _logic("d_any", "|A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                       "msb": 3, "lsb": 0}}, out_width=1)
    assert G._selfaudit_output_width(sig, E.parse("|A"), E.Env({"A": 4})) is None


def test_check2_ls_name_skipped():
    """_ls 顶层口宽度由 level_shift 页给名、不走 K 列推断 → 跳过、不误报。"""
    sig = _logic("d_sel", "A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                      "msb": 3, "lsb": 0}}, out_width=1)
    sig._ls_name = "d_sel_ls"
    assert G._selfaudit_output_width(sig, E.parse("A"), E.Env({"A": 4})) is None


# ───────────────────────── 检查3：RW 写值截断 ─────────────────────────
def test_check3_rw_overflow_flagged():
    # 输入占 [1..15]=15bit@lsb1，字段只有 8bit → 高位被 & mask(fw) 丢掉
    assert G._selfaudit_rw_truncation({"A": _rw("f", 15, 1, 0, 7)}) is not None


def test_check3_rw_fits_not_flagged():
    assert G._selfaudit_rw_truncation({"A": _rw("f", 8, 0, 0, 7)}) is None   # 正好装下


def test_check3_ro_not_flagged():
    ro = InputBinding(letter="A", raw="f", base="f", width=15, kind="RO", address=None,
                      reg_lsb=None, reg_msb=None, wire="f", found_in="tmm")
    assert G._selfaudit_rw_truncation({"A": ro}) is None


def test_check3_unknown_bounds_default16_overflow_flagged():
    """字段边界未知(reg_msb/lsb=None)→ compute_drives 用默认 16 位且真会 & mask(16) 截断；
    check3 必须同口径用 16、不能漏报(对抗审查抓出的盲区)。20bit 输入 > 16 → 报。"""
    assert G._selfaudit_rw_truncation({"A": _rw("f", 20, 0, None, None)}) is not None


def test_check3_unknown_bounds_within16_not_flagged():
    """边界未知但 ≤16 位 → compute_drives 不丢位 → 不报(默认 16 口径下装得下)。"""
    assert G._selfaudit_rw_truncation({"A": _rw("f", 8, 0, None, None)}) is None


# ───────────────────────── e2e：闸门进生成管线(块顶 ⚠ + 汇总) ─────────────────────────
def test_check2_e2e_block_top_warning_and_summary():
    sig = _logic("d_sel", "A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                      "msb": 3, "lsb": 0}}, out_width=1, top_output=1)
    wb = DregWorkbook(logic=[sig], regmap={}, tmm={}, sheet_names=[])
    res = G.build(wb, G.GenOptions(include_risky=True))
    assert res["summary"]["n_selfaudit_warnings"] >= 1
    assert res["selfaudit_warnings"]
    text = "\n".join("\n".join(lines) for lines, _st in res["blocks"])
    assert "// ⚠" in text and "K 列可能漏标位宽" in text


def test_clean_table_no_selfaudit_warnings():
    """干净表(K 列标宽匹配、无自引用泄漏、RW 装得下)→ 零闸门命中(产物逐字节不变的前提)。"""
    sig = _logic("d_sel[3:0]", "A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                           "msb": 3, "lsb": 0}}, out_width=4, top_output=1)
    wb = DregWorkbook(logic=[sig], regmap={}, tmm={}, sheet_names=[])
    res = G.build(wb, G.GenOptions(include_risky=True))
    assert res["summary"]["n_selfaudit_warnings"] == 0


# ───────────────────────── make_supplement_signal 自引用继承（对抗审查 #4） ─────────────────────────
def test_supplement_renamed_output_does_not_inherit_base_self_ref():
    """补充用 out_name 改了输出基名 → 不继承原信号绑 base out_base 的自引用集(那是另一身份的事实)。"""
    base = _logic("d_base", "A", {"A": {"raw": "d_base_to_logic", "base": "d_base", "width": 1,
                                        "msb": None, "lsb": None}})
    base._self_ref_suffixes = {"_to_logic"}
    base.ref_suffix = ""
    spec = {"expr": "A", "inputs": [{"var": "A", "raw": "cfg_to_logic"}], "out_name": "d_renamed"}
    sig = G.make_supplement_signal("d_base", spec, base)
    assert sig.out_base == "d_renamed"
    assert sig._self_ref_suffixes == set()


def test_supplement_same_output_inherits_base_self_ref():
    """补充不改 out_name(默认用法：只改逻辑、继承探针身份)→ 继承原信号自引用集。"""
    base = _logic("d_base", "A", {"A": {"raw": "d_base_to_logic", "base": "d_base", "width": 1,
                                        "msb": None, "lsb": None}})
    base._self_ref_suffixes = {"_to_logic"}
    base.ref_suffix = ""
    spec = {"expr": "A", "inputs": [{"var": "A", "raw": "cfg_to_logic"}]}
    sig = G.make_supplement_signal("d_base", spec, base)
    assert sig.out_base == "d_base"
    assert sig._self_ref_suffixes == {"_to_logic"}


# ───────────────────────── claim 导出器（红区数据契约） ─────────────────────────
def test_claims_exported_for_probe_and_force():
    sig = _logic("d_sel[3:0]", "A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                           "msb": 3, "lsb": 0}}, out_width=4, top_output=1)
    wb = DregWorkbook(logic=[sig], regmap={}, tmm={}, sheet_names=[])
    res = G.build(wb, G.GenOptions(include_risky=True))
    claims = res["claims"]
    probes = [c for c in claims if c["kind"] == "probe"]
    forces = [c for c in claims if c["kind"] == "force"]
    assert len(probes) == 1
    assert probes[0]["net_base"] == "d_sel" and probes[0]["identity"] == "output"
    assert probes[0]["slice"] == "[3:0]" and probes[0]["top_output"] is True
    assert probes[0]["self_ref_suffixes"] == []
    # force claim 的 full = .sv 里真正 force 的 LHS(多位带切片)，slice 也带上(对抗审查 #2 修复)
    f = next(c for c in forces if c["net_base"] == "a_in")
    assert f["full"] == "a_in[3:0]" and f["slice"] == "[3:0]"
    # schema 固定：每条 claim 字段完全一致
    for c in claims:
        assert set(c.keys()) == set(G._CLAIM_KEYS)


def test_export_claims_writes_json(tmp_path):
    from dreg_verify import cli
    import json
    sig = _logic("d_sel[3:0]", "A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                           "msb": 3, "lsb": 0}}, out_width=4, top_output=1)
    wb = DregWorkbook(logic=[sig], regmap={}, tmm={}, sheet_names=[])
    res = G.build(wb, G.GenOptions(include_risky=True))
    p = tmp_path / "claims.json"
    cli._export_claims(res, str(p), "fake.xlsx")
    data = json.loads(p.read_text(encoding="utf-8"))
    assert data["source_excel"] == "fake.xlsx"
    assert data["n_claims"] == len(res["claims"])
    assert data["schema"] == list(G._CLAIM_KEYS)
    assert any(c["kind"] == "probe" for c in data["claims"])
    # C2(2026-06-23)：契约版本号 + 命名模型，跨空气墙防静默不兼容
    assert data["schema_version"] == G.CLAIM_SCHEMA_VERSION
    assert data["naming_model"] == G.NAMING_MODEL_LOGIC      # 默认旧路径
    assert "logic-rooted" in data["note"]


def test_export_claims_naming_model_topout(tmp_path):
    """显式 naming_model='topout' → JSON 标 topout + note 改写成顶层真名口径。"""
    from dreg_verify import cli
    import json
    sig = _logic("d_sel[3:0]", "A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                           "msb": 3, "lsb": 0}}, out_width=4, top_output=1)
    wb = DregWorkbook(logic=[sig], regmap={}, tmm={}, sheet_names=[])
    res = G.build(wb, G.GenOptions(include_risky=True))
    p = tmp_path / "claims.json"
    cli._export_claims(res, str(p), "fake.xlsx", naming_model=G.NAMING_MODEL_TOPOUT)
    data = json.loads(p.read_text(encoding="utf-8"))
    assert data["naming_model"] == "topout"
    assert "topout" in data["note"]


# ── M0(goal-redzone-binder)：claims 补 input_nets + on_missing 两字段 ──
def _sel_wb():
    sig = _logic("d_sel[3:0]", "A", {"A": {"raw": "a_in[3:0]", "base": "a_in", "width": 4,
                                           "msb": 3, "lsb": 0}}, out_width=4, top_output=1)
    return DregWorkbook(logic=[sig], regmap={}, tmm={}, sheet_names=[])


def test_probe_claim_carries_input_nets():
    """探针 claim 的 input_nets = 该信号各输入绑定的网基名(红区 binder 拿 assign RHS 做指纹用)；
    非探针 claim 不挂(None)。"""
    res = G.build(_sel_wb(), G.GenOptions(include_risky=True))
    probe = next(c for c in res["claims"] if c["kind"] == "probe")
    assert "a_in" in (probe["input_nets"] or [])
    for c in res["claims"]:
        if c["kind"] != "probe":
            assert c["input_nets"] is None


def test_on_missing_defaults_warn_and_per_signal_override():
    """on_missing 挂在探针 claim：缺省 warn(甲)；GenOptions.on_missing 单点可设 fallback(乙)。"""
    p0 = next(c for c in G.build(_sel_wb(), G.GenOptions(include_risky=True))["claims"]
              if c["kind"] == "probe")
    assert p0["on_missing"] == "warn"
    res = G.build(_sel_wb(), G.GenOptions(include_risky=True, on_missing={"d_sel": "fallback"}))
    p1 = next(c for c in res["claims"] if c["kind"] == "probe")
    assert p1["on_missing"] == "fallback"
    # 非法值被规整回 warn，不污染契约
    res2 = G.build(_sel_wb(), G.GenOptions(include_risky=True, on_missing={"d_sel": "bogus"}))
    p2 = next(c for c in res2["claims"] if c["kind"] == "probe")
    assert p2["on_missing"] == "warn"
