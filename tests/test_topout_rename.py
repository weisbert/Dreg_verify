# -*- coding: utf-8 -*-
"""dft 改名桥接测试（2026-06-24）：Topout 顶层名 A 在 dft 页被观测改名(源=logic A_sm) →
桥到源拿逻辑/驱动，但断言探针贴顶层真名 A（不是 A_sm）。

无改名的表(mirror)零影响=逐字节不变(主套件 e2e 已覆盖)；本文件只验改名场景。
"""

import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import make_mirror_btlp                            # noqa: E402
from dreg_verify import excel_model as M           # noqa: E402
from dreg_verify import topout as T                # noqa: E402


@pytest.fixture(scope="module")
def renamed_path(tmp_path_factory):
    """mirror + 一条 dft 改名信号：logic a_sm=(A?C:B)&(~D) → dft 观测改名成顶层 a_top → Topout a_top。"""
    p = str(tmp_path_factory.mktemp("rn") / "renamed.xlsx")
    make_mirror_btlp.build(p)
    wb = openpyxl.load_workbook(p)
    lg, dft, top = wb["logic"], wb["dft"], wb["Topout"]
    r = lg.max_row + 1
    lg.cell(r, 1, "d_bt_lp_linelocal_mode_ctrl_to_logic")
    lg.cell(r, 2, "d_bt_lp_linectrl_rx_en_to_logic")
    lg.cell(r, 3, "d_bt_lp_rx_en_local_to_logic")
    lg.cell(r, 4, "d_bt_lp_pll_dig_dft_iddq_mode_to_logic")
    lg.cell(r, 11, "a_sm"); lg.cell(r, 12, "(A?C:B)&(~D)"); lg.cell(r, 16, "Test Owner")
    r = dft.max_row + 1
    dft.cell(r, 1, "a_sm_to_dft"); dft.cell(r, 4, "a_top"); dft.cell(r, 5, "A")
    dft.cell(r, 6, "16"); dft.cell(r, 8, "Test Owner")
    r = top.max_row + 1
    top.cell(r, 1, "Test Owner"); top.cell(r, 2, "a_top")
    wb.save(p)
    return p


def test_dft_rename_resolves_to_logic_source(renamed_path):
    wb = M.load_workbook(renamed_path)
    root = T.resolve_root(wb, "a_top")
    assert root.kind == T.LOGIC
    assert root.renamed is True
    assert root.probe_name == "a_top"      # 断言贴顶层真名
    assert root.source_name == "a_sm"      # 逻辑/驱动来自源
    # 源名仍能直接解析（桥接不影响直接路径）
    assert T.resolve_root(wb, "a_sm").kind == T.LOGIC


def test_dft_rename_view_model_has_truth_table(renamed_path):
    wb = M.load_workbook(renamed_path)
    m = next(x for x in T.topout_view_models(wb, mode="max", max_tests=32) if x["name"] == "a_top")
    assert m["kind"] == "logic" and m["status"] == "ok"
    assert m["n_vectors"] > 0 and len(m["inputs"]) == 4      # 真值表来自 a_sm 的 4 输入逻辑


def test_dft_rename_sv_asserts_top_name_not_source(renamed_path):
    wb = M.load_workbook(renamed_path)
    text, b = T.render_topout_sv(wb, mode="max", max_tests=32, only=["a_top"])
    assert "`ENV_RF.a_top==" in text          # 探针贴顶层真名 a_top
    assert "`ENV_RF.a_sm==" not in text       # 不是源名 a_sm
    assert text.count("assert (") > 0


def test_no_rename_table_has_zero_renames(renamed_path, tmp_path_factory):
    """对照：无改名的 mirror → 改名表为空（旧表零影响）。"""
    plain = str(tmp_path_factory.mktemp("plain") / "mirror.xlsx")
    make_mirror_btlp.build(plain)
    wb = M.load_workbook(plain)
    assert T._dft_rename_map(wb) == {}
    # renamed 表里恰好 1 条改名
    assert T._dft_rename_map(M.load_workbook(renamed_path)).get("a_top") == "a_sm"


@pytest.fixture(scope="module")
def chain_path(tmp_path_factory):
    """真表实况：链式改名 level_shift → dft → logic（211 信号里唯一未解析的那个）。
    logic d_vco_en_faston_fsm → dft 观测改名 d_vco_en_faston → level_shift 顶层口 d_vco_en_faston_ls。"""
    p = str(tmp_path_factory.mktemp("ch") / "chain.xlsx")
    make_mirror_btlp.build(p)
    wb = openpyxl.load_workbook(p)
    lg, dft, ls, top = wb["logic"], wb["dft"], wb["level_shift"], wb["Topout"]
    r = lg.max_row + 1
    for c, v in [(1, "d_bt_lp_linelocal_mode_ctrl_to_logic"), (2, "d_bt_lp_linectrl_rx_en_to_logic"),
                 (3, "d_bt_lp_rx_en_local_to_logic"), (4, "d_bt_lp_pll_dig_dft_iddq_mode_to_logic"),
                 (11, "d_vco_en_faston_fsm"), (12, "(A?C:B)&(~D)"), (16, "Yao Wang")]:
        lg.cell(r, c, v)
    r = dft.max_row + 1
    for c, v in [(1, "d_vco_en_faston_fsm_to_dft"), (4, "d_vco_en_faston"), (5, "A"), (6, "16"),
                 (8, "Yao Wang")]:
        dft.cell(r, c, v)
    r = ls.max_row + 1
    for c, v in [(1, "d_vco_en_faston_to_ls"), (2, "STD_SR_L2H"), (3, "d_vco_en_faston_ls"),
                 (5, "1"), (6, "Yao Wang")]:
        ls.cell(r, c, v)
    r = top.max_row + 1
    top.cell(r, 1, "Yao Wang"); top.cell(r, 2, "d_vco_en_faston_ls")
    wb.save(p)
    return p


def test_chained_rename_level_shift_then_dft_to_logic(chain_path):
    """链式回溯：Topout d_vco_en_faston_ls ←(level_shift) d_vco_en_faston ←(dft) logic _fsm。"""
    wb = M.load_workbook(chain_path)
    root = T.resolve_root(wb, "d_vco_en_faston_ls")
    assert root.kind == T.LOGIC and root.renamed
    assert root.probe_name == "d_vco_en_faston_ls"       # 探针贴顶层 _ls 真名
    assert root.source_name == "d_vco_en_faston_fsm"     # 链尾的 logic 源
    # 真值表建出来 + .sv 探顶层名(非源名、非中间名)
    m = next(x for x in T.topout_view_models(wb, mode="max", max_tests=32)
             if x["name"] == "d_vco_en_faston_ls")
    assert m["kind"] == "logic" and m["status"] == "ok" and m["n_vectors"] > 0
    text, _ = T.render_topout_sv(wb, mode="max", max_tests=32, only=["d_vco_en_faston_ls"])
    assert "`ENV_RF.d_vco_en_faston_ls==" in text
    assert "`ENV_RF.d_vco_en_faston_fsm==" not in text
    assert "`ENV_RF.d_vco_en_faston==" not in text       # 中间名也不探


@pytest.fixture(scope="module")
def collision_path(tmp_path_factory):
    """真表 d_vco_en_faston_ls 实况(2026-06-24)：dft 观测【输出名】d_vco_en_faston 恰与一个
    【同名 RW 输入寄存器】撞车，且该 dft 观测带 iddq 门(B?0:A)。三层级联：
        d_vco_en_faston_ls ←(level_shift) d_vco_en_faston(=iddq?0:fsm；又有同名 RW 寄存器=FSM 的输入位)
                           ←(dft 门控)    d_vco_en_faston_fsm(logic: reg[1]?reg[0]:fll_active)
    旧版桥接走到 d_vco_en_faston 时撞上同名寄存器即停→误判『直连寄存器』(1 输入、按字段宽 2 验=假绿)，
    漏掉 logic 选路 + iddq 门 + fll_active。修后应桥到 logic 源(reg+fll_active+iddq 门、1 bit)。"""
    from openpyxl.utils import column_index_from_string as ci
    p = str(tmp_path_factory.mktemp("col") / "collision.xlsx")
    make_mirror_btlp.build(p)
    wb = openpyxl.load_workbook(p)
    lg, dft, ls, rm, top = wb["logic"], wb["dft"], wb["level_shift"], wb["regmap"], wb["Topout"]
    # logic: d_vco_en_faston_fsm = A?B:C （A,B=同名寄存器两位 _to_logic；C=fll_active）
    r = lg.max_row + 1
    for c, v in [(1, "d_vco_en_faston_to_logic[1]"), (2, "d_vco_en_faston_to_logic[0]"),
                 (3, "fll_active_to_logic"), (11, "d_vco_en_faston_fsm"),
                 (12, "A?B:C"), (13, "to_logic,to_dft"), (16, "Yao Wang")]:
        lg.cell(r, c, v)
    # dft 门控观测：out=d_vco_en_faston, A=fsm_to_dft, B=iddq_to_dft(=mirror 既有 RO), E=B?0:A
    r = dft.max_row + 1
    for c, v in [(1, "d_vco_en_faston_fsm_to_dft"), (2, "d_bt_lp_pll_dig_dft_iddq_mode_to_dft"),
                 (4, "d_vco_en_faston"), (5, "B?0:A"), (6, "16"), (8, "Yao Wang")]:
        dft.cell(r, c, v)
    # level_shift: d_vco_en_faston_to_ls → 顶层口 d_vco_en_faston_ls
    r = ls.max_row + 1
    for c, v in [(1, "d_vco_en_faston_to_ls"), (2, "STD_SR_L2H"), (3, "d_vco_en_faston_ls"),
                 (5, 1), (6, "Yao Wang")]:
        ls.cell(r, c, v)
    # regmap: 撞名 RW 寄存器 d_vco_en_faston[1:0]@0x13 bits1:0 + fll_active(RO)
    r = rm.max_row + 1
    for col, v in {"D": "VCO", "F": "RW", "G": "d_vco_en_faston[1:0]", "H": "d19",
                   "Z": "to_logic", "AB": 0, "AE": "Yao Wang"}.items():
        rm.cell(r, ci(col), v)
    for b in (0, 1):
        rm.cell(r, ci("Y") - b, 1)
    r = rm.max_row + 1
    for col, v in {"D": "CT_read_back", "F": "RO", "G": "fll_active", "H": "d59",
                   "Z": "to_logic", "AB": 0, "AE": "Yao Wang"}.items():
        rm.cell(r, ci(col), v)
    rm.cell(r, ci("Y") - 1, 1)
    # Topout: 要验 d_vco_en_faston_ls
    r = top.max_row + 1
    top.cell(r, 1, "Yao Wang"); top.cell(r, 2, "d_vco_en_faston_ls")
    wb.save(p)
    return p


def test_dft_observed_name_colliding_with_register_resolves_to_logic(collision_path):
    """⭐核心修复：dft 观测输出名撞同名 RW 输入寄存器时，桥到真正的 logic 源（非误判直连寄存器）。"""
    wb = M.load_workbook(collision_path)
    root = T.resolve_root(wb, "d_vco_en_faston_ls")
    assert root.kind == T.LOGIC                          # 修前=REGISTER（误判）
    assert root.renamed
    assert root.source_name == "d_vco_en_faston_fsm"     # 链尾 logic 源（非撞名寄存器）
    assert root.probe_name == "d_vco_en_faston_ls"       # 断言贴顶层真名
    assert root.dft_obs_name == "d_vco_en_faston"        # 途经的带 iddq 门 dft 观测名


def test_collision_cone_has_reg_fll_iddq_and_is_one_bit(collision_path):
    """cone = 寄存器 + fll_active + iddq 门(force)；输出 1 bit（修前=按字段宽 2 验=假绿黄字）。"""
    from dreg_verify import resolver as R
    wb = M.load_workbook(collision_path)
    topo = next(t for t in wb.topout if t.name == "d_vco_en_faston_ls")
    res = T.analyze_signal(wb, R.Resolver(wb), topo, mode="max", max_tests=64)
    assert res.status == "ok"
    assert res.out_width == 1                            # 顶层 1 bit（修前误判=2）
    assert not any("字段宽" in s for s in res.issues)    # 不再有误诊『无位宽切片但寄存器字段宽』黄字
    bases = {b.base.lower() for b in res.bindings.values() if getattr(b, "base", None)}
    assert "d_vco_en_faston" in bases and "fll_active" in bases   # logic 选路两输入展进 cone
    forced = {wl.lower() for v in res.vectors
              for (wl, _x, _w) in (getattr(v, "extra_forces", None) or [])}
    assert any("iddq" in f for f in forced)              # iddq 门叠成每条向量的 force（透传值）


def test_collision_sv_asserts_top_name_one_bit_and_forces_iddq(collision_path):
    wb = M.load_workbook(collision_path)
    text, _ = T.render_topout_sv(wb, mode="max", max_tests=64, only=["d_vco_en_faston_ls"])
    assert "`ENV_RF.d_vco_en_faston_ls==1'b" in text     # 探顶层真名 + 1-bit 断言
    assert "==2'b" not in text                           # 修前=2'b（按撞名寄存器字段宽误验）
    assert "`ENV_RF.d_vco_en_faston_fsm==" not in text   # 不探源名
    assert "iddq" in text.lower()                        # iddq 门被 force（与 for_test 金标准同口径）


def test_renamed_signal_gui_edit_threads_to_export(renamed_path):
    """GUI 编辑改名信号 → 走 reg 路按顶层名键回流（否则编辑落源名键被改名路忽略=静默不生效）。"""
    import os
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    pytest.importorskip("PySide6")
    from PySide6 import QtWidgets
    from dreg_verify import gui as G
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([])
    w = G.MainWindow(); w.path_edit.setText(renamed_path); w.on_load()
    try:
        v = w.topout_view
        r = next(i for i in range(v.sig_table.rowCount())
                 if v.sig_table.item(i, G.TOPO_NAME).text() == "a_top")
        v.sig_table.setCurrentCell(r, G.TOPO_NAME)
        assert v.cur_an["renamed"] is True and v.cur_an["editable"] == "logic"
        base, _ = v.provider.render_sv(None, "max", 32, False, {})
        assert "`ENV_RF.a_top==" in base
        v._e_clear()                                 # 清零改名信号
        text, _ = v.provider.render_sv(None, "max", 32, False, v._compute_edited())
        assert "`ENV_RF.a_top==" not in text         # 编辑生效：断言消失
        assert "用户已清空" in text                   # 记账(不静默丢)
    finally:
        w.close()
