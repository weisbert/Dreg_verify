# -*- coding: utf-8 -*-
"""
inspect_dft.py — 摸清真表里 iddq/DFT 门控的真实写法（2026-06-10）。

背景：d_wl_rf_lo2g5g_mixer2g_trim 等输出受 d_wl_rf_trx_reg_dft_iddq_mode 门控，
但工具在真表上没认出来（GUI 输入表无 DFT 门行 / .sv 无 IDDQ 拍）。read_dft 的几个
静默失败点——页名必须叫 dft、列位假设 B=门/D=被观测输出/E=表达式、表达式必须是
单条件门 cond?0:x 或 cond?x:0（带位宽/括号/取反都不匹配）、表头在第 2 行——真表
偏离任何一条，wb.dft 就是空的且无任何提示。

这个脚本只做一件事：把真表里所有和 iddq/dft 相关的格子原样挖出来 ——
  ① 工具版本 + 全部页名 + 哪个页会被 _find_sheet(wb,'dft') 命中；
  ② 名字含 dft 的每个页：前 N 行 × A..J 列【原始逐格】dump（带列字母，None 跳过）；
  ③ 全工作簿扫描：任何格子内容含『iddq』的，打印 页名!单元格 = 值，并把该行
     A..J 列整行带出（门控关系藏在哪一页、哪几列，一次看全）；
  ④ 用产品代码 excel_model.load_workbook 实际装载一次，打印 wb.dft 解析结果
     + mux/logic 里 out_base 含指定子串(默认 mixer2g_trim)的条目（对照键是否匹配）。

只读，不碰任何产品代码。输出写 UTF-8 文件（Windows 控制台是 GBK，直接 print
中文会变 ?），跑完把 .txt 全文贴回来即可。

用法（装了 openpyxl 的 python 环境）：
    python inspect_dft.py 真表.xlsx
        [--rows 60]        每个 dft 页最多 dump 多少行（默认 60）
        [--cols 10]        dump 到第几列（默认 10 = A..J）
        [--sig mixer2g_trim]  对照的输出名子串（默认 mixer2g_trim）
        [--scan-rows 20000]   全簿 iddq 扫描每页最多看多少行（默认 20000）
        [--out 文件]       默认 <excel名>_dft.txt

依赖：openpyxl。仅支持 .xlsx。
"""
import argparse
import os
import sys

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少 openpyxl，请先运行:  pip install openpyxl")


def fmt_cell(v):
    if v is None:
        return None
    s = str(v).replace("\n", "⏎")
    return s if len(s) <= 80 else s[:77] + "..."


def dump_row(ws_title, row_idx, row_vals, max_cols, out):
    cells = []
    for ci, v in enumerate(row_vals[:max_cols]):
        fv = fmt_cell(v)
        if fv is not None and fv != "":
            cells.append("%s=%s" % (get_column_letter(ci + 1), fv))
    if cells:
        out.append("  %s!行%d:  %s" % (ws_title, row_idx, "   ".join(cells)))


def main():
    ap = argparse.ArgumentParser(description="摸清真表 iddq/DFT 门控的真实写法（只读）")
    ap.add_argument("excel")
    ap.add_argument("--rows", type=int, default=60)
    ap.add_argument("--cols", type=int, default=10)
    ap.add_argument("--sig", default="mixer2g_trim")
    ap.add_argument("--scan-rows", type=int, default=20000)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    out_path = args.out or (os.path.splitext(os.path.basename(args.excel))[0] + "_dft.txt")
    L = []

    # ── ① 工具版本（确认跑的是新代码）+ 页名清单 ──
    here = os.path.dirname(os.path.abspath(__file__))
    try:
        import subprocess
        head = subprocess.run(["git", "-C", here, "rev-parse", "--short", "HEAD"],
                              capture_output=True, text=True, timeout=10).stdout.strip()
    except Exception:  # noqa: BLE001
        head = "(git 不可用)"
    L.append("工具目录 HEAD: %s" % head)
    L.append("Excel: %s" % args.excel)

    wb = openpyxl.load_workbook(args.excel, data_only=True, read_only=True)
    L.append("\n══ ① 全部页名（×=名字恰为'dft'会被工具当 dft 页；~=名字含 dft 但不会命中）══")
    hit = None
    for s in wb.sheetnames:
        mark = " "
        if s.lower() == "dft":
            mark, hit = "×", s
        elif "dft" in s.lower():
            mark = "~"
        L.append("  [%s] %r" % (mark, s))
    if hit is None:
        L.append("  ⛔ 没有页名恰好是 'dft' —— 工具 _find_sheet 找不到 dft 页，wb.dft 必为空（一类根因）")

    # ── ② 名字含 dft 的页逐格 dump ──
    dft_sheets = [s for s in wb.sheetnames if "dft" in s.lower()]
    for s in dft_sheets:
        ws = wb[s]
        L.append("\n══ ② 页 %r 前 %d 行原始逐格（A..%s；空格子跳过）══"
                 % (s, args.rows, get_column_letter(args.cols)))
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=args.rows,
                                              values_only=True), start=1):
            dump_row(s, ri, row, args.cols, L)
    if not dft_sheets:
        L.append("\n══ ② 没有名字含 dft 的页 ══")

    # ── ③ 全簿扫描含 iddq 的格子（门控关系藏在哪，一次看全）──
    L.append("\n══ ③ 全簿扫描：内容含 'iddq' 的格子（每页最多看 %d 行；整行 A..%s 带出）══"
             % (args.scan_rows, get_column_letter(args.cols)))
    n_hits = 0
    for s in wb.sheetnames:
        ws = wb[s]
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=args.scan_rows,
                                              values_only=True), start=1):
            if any(v is not None and "iddq" in str(v).lower() for v in row):
                n_hits += 1
                if n_hits <= 120:
                    dump_row(s, ri, row, args.cols, L)
    L.append("  共 %d 行命中%s" % (n_hits, "（只显示前 120 行）" if n_hits > 120 else ""))

    # ── ⑤ for_test 顺序探查：--sig 命中组的 D/F 原始格（输入行序对照真值表排序用）──
    L.append("\n══ ⑤ for_test 页：内容含 %r 的格子所在【组】的 D/F 列原样（行序=排序依据）══"
             % args.sig)
    ftn = next((s for s in wb.sheetnames if s.lower() == "for_test"), None)
    if ftn is None:
        L.append("  ⛔ 没有 for_test 页 —— 排序无来源，回退原序(门殿后)")
    else:
        ws = wb[ftn]
        rows_df = []        # [(行号, D, F)]
        hit_rows = []
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_col=6, values_only=True),
                                 start=1):
            d = "" if len(row) < 4 or row[3] is None else str(row[3])
            fv = "" if len(row) < 6 or row[5] is None else str(row[5])
            rows_df.append((ri, d, fv))
            if args.sig.lower() in (d + " " + fv).lower():
                hit_rows.append(ri)
        if not hit_rows:
            L.append("  ⛔ for_test 里没有含 %r 的 D/F 格 —— 该输出不在 for_test，回退原序" % args.sig)
        else:
            lo = max(0, min(hit_rows) - 16)
            hi = min(len(rows_df), max(hit_rows) + 4)
            for (ri, d, fv) in rows_df[lo:hi]:
                if d or fv:
                    L.append("  for_test!行%-5d D=%-50s F=%s" % (ri, d, fv))
    wb.close()

    # ── ④ 产品代码实际装载：wb.dft 解析结果 + 对照输出名 ──
    L.append("\n══ ④ 产品代码 load_workbook 实测 ══")
    sys.path.insert(0, here)
    try:
        from dreg_verify import excel_model
        m = excel_model.load_workbook(args.excel)
        L.append("wb.dft 解析到 %d 条:" % len(m.dft))
        for k, v in sorted(m.dft.items()):
            L.append("  %s -> 门=%s 透传=%s (原文 %s)"
                     % (k, v["gate_base"], v["transparent"], v["gate_raw"]))
        L.append("out_base 含 %r 的信号:" % args.sig)
        for g in m.mux:
            if args.sig.lower() in g.out_base.lower():
                L.append("  mux: out_name=%s out_base=%s（dft键须=%s）"
                         % (g.out_name, g.out_base, g.out_base.lower()))
        for sig in m.logic:
            if args.sig.lower() in (sig.out_base or "").lower():
                L.append("  logic: out_name=%s out_base=%s" % (sig.out_name, sig.out_base))
        # for_test 顺序解析结果（排序生效与否就看这里有没有该输出的键）
        L.append("fortest_order 解析到 %d 个输出组；键含 %r 的："
                 % (len(getattr(m, "fortest_order", {}) or {}), args.sig))
        for k, order in sorted((getattr(m, "fortest_order", {}) or {}).items()):
            if args.sig.lower() in k:
                L.append("  %s -> %s" % (k, order))
    except Exception as ex:  # noqa: BLE001
        L.append("  ⛔ 装载失败: %s" % ex)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print("已写出: %s（共 %d 行）—— 把这个文件全文贴回来" % (out_path, len(L)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
