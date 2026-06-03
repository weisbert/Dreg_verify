# -*- coding: utf-8 -*-
"""
rtl_scan.py — Excel 侧的 RTL 扫描配套：把 Excel 信号转换成"需要在 ENV_RF 层级存在的网"清单。

RTL 解析的全部实现在仓库根目录的 scan_rtl.py（单文件、零第三方依赖，可直接拷到仿真服务器）。
本模块复用其实现，并补充需要 dreg_verify/openpyxl 的部分：
    collect_excel_nets(wb)  Excel → {网名: 用途}
    match_excel(wb, sigmap) 对照 Excel 与 RTL 层级

跨机器两段式工作流（Excel 在 Windows、RTL 在 Linux 服务器）见 scan_rtl.py 文件头。
"""

import os
import sys

# 仓库根目录加入 path，导入单文件版 scan_rtl 的解析实现
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scan_rtl import (KEYWORDS, build_signal_map, find_verilog_files,            # noqa: F401,E402
                      match_nets, parse_modules, parse_nets_text,
                      render_nets_text, render_prefix_file, scan_files,
                      strip_comments)


def collect_excel_nets(wb):
    """Excel → 需要在 ENV_RF 层级存在的网名集合：

    ① 每个输出的 RTL 网名 —— assert 探针：
         top_output=1 → K 列名（ls 行带 _ls 后缀）
         top_output=0 → K 列名 + _to_logic（在 sig_logic 模块内部，需要探针前缀才能验）
    ② 它们的 force 输入（RO/wire 兜底/级联）—— force 路径
    RW 寄存器输入走 RF_WRITE，不需要层级。返回 {网名: 用途说明}。

    ⭐ 两种级联模式(cone 展开上游 / force 级联网)需要的网都导出——
    一次 RTL 扫描同时覆盖两种模式，之后在 GUI/CLI 里切换模式不用重新扫。
    """
    from . import expr as E
    from . import generator
    from . import resolver as R

    nets = {}
    for sig in wb.logic:
        kind = "" if sig.is_top else "（内部信号）"
        nets.setdefault(sig.rtl_base, "输出 %s 的 assert 探针%s" % (sig.out_name, kind))
    # 两种级联模式各跑一遍解析，导出网取并集
    for mode in ("cone", "force"):
        resolver = R.Resolver(wb, cascade_mode=mode)
        tag = "" if mode == "cone" else "（force级联网模式）"
        for sig in wb.logic:
            try:
                node, bindings, _ = generator.expand_signal(wb, resolver, sig)
            except Exception:  # noqa: BLE001  cone 失败时退回原始绑定
                try:
                    node = E.parse(sig.expr)
                    bindings = resolver.resolve_signal_inputs(sig)
                except Exception:  # noqa: BLE001
                    continue
            for b in bindings.values():
                if b.kind == "RO":
                    base = b.wire.split(".")[-1]
                    nets.setdefault(base, "输出 %s 的输入 %s (force)%s"
                                    % (sig.out_name, b.base, tag))
    return nets


def collect_mux_nets(excel_path):
    """mux 页 → 需要在 ENV_RF 层级核对的网（2026-06-03 第九轮：mux 验证环境核查）。

    直接读 Excel 的 mux 页（不依赖 excel_model 的 mux 支持——那是后续功能；
    本函数只为"动手写 mux 功能前先验证 RTL 环境"服务）。导出两类：
      ① G 列 mux 输出基名 —— assert 探针（designer .sv 实证顶层可探，扫描确认层级）
      ② B 列控制信号网（_to_mux 名）—— logic→mux 衔接核对（测试不直接 force，仅核对存在）
    数据输入（A 列）是 RW 寄存器走 RF_WRITE，无需网名核对。
    控制信号的线控 force 输入（linectrl_*）已由 collect_excel_nets 的 logic 循环导出，不重复。

    mux 页不存在 / 读取失败 → 返回空 dict，纯 logic 流程完全不受影响。
    """
    import re
    try:
        import openpyxl
    except ImportError:
        return {}
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return {}
    ws = None
    for s in wb.sheetnames:
        if "mux" in s.lower():
            ws = wb[s]
            break
    if ws is None:
        wb.close()
        return {}

    def base_name(v):
        """去位宽尾巴；不是合法信号名（如 '(reserved)'）→ None。"""
        s = re.sub(r"\[[^\]]*\]\s*$", "", str(v).strip())
        return s if re.match(r"^[A-Za-z_]\w*$", s) else None

    nets = {}
    upto = min(ws.max_row or 0, 5000)
    # mux 页表头在第 2 行（与 logic 页同套路），数据从第 3 行起
    for row in ws.iter_rows(min_row=3, max_row=upto, values_only=True):
        g = row[6] if len(row) > 6 else None     # G 列 = mux_out（被验证输出）
        b = row[1] if len(row) > 1 else None     # B 列 = mux_ctrl_sig1（控制信号）
        if g is not None and str(g).strip():
            gb = base_name(g)
            if gb:
                nets.setdefault(gb, "mux 输出 %s 的 assert 探针" % str(g).strip())
        if b is not None and str(b).strip():
            bb = base_name(b)
            if bb:
                nets.setdefault(bb, "mux 控制网 %s（logic→mux 衔接核对）" % str(b).strip())
    wb.close()
    return nets


def match_excel(wb, sigmap):
    """对照：Excel 需要的网 vs RTL 层级。返回 (prefixes, at_top, missing)，见 scan_rtl.match_nets。"""
    return match_nets(collect_excel_nets(wb), sigmap)
