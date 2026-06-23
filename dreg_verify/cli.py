# -*- coding: utf-8 -*-
"""
cli.py — 命令行入口：从 Dreg 核心 Excel 生成 wr_rf_tc.sv。

用法示例:
  # 生成全部信号
  python -m dreg_verify.cli --excel core.xlsx --out wr_rf_tc.sv

  # 只生成某 owner 的信号
  python -m dreg_verify.cli --excel core.xlsx --owner Alice --out wr_rf_tc.sv

  # 给某些信号加负向(异常)用例，取反造错，单独出文件
  python -m dreg_verify.cli --excel core.xlsx --neg-signals d_logic_bt_lp_reserve \
      --neg-mode invert --neg-file separate --out wr_rf_tc.sv

  # 列出信号清单(不生成)
  python -m dreg_verify.cli --excel core.xlsx --list
"""

import argparse
import os
import sys

# Windows 控制台常是 GBK，打印中文会崩；统一改 UTF-8（文件输出本就 UTF-8）
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dreg_verify import excel_model, generator  # noqa: E402


def _split(s):
    if not s:
        return None
    return [x.strip() for x in s.replace(";", ",").split(",") if x.strip()]


def build_argparser():
    p = argparse.ArgumentParser(
        prog="dreg_verify",
        description="从 Dreg 核心 Excel 的 logic 真值表达式生成 wr_rf_tc.sv 验证文件。")
    p.add_argument("--excel", required=True, help="核心 Excel (.xlsx) 路径")
    p.add_argument("--out", default=None, help="输出 .sv 路径 (默认 wr_rf_tc.sv；只想出报告时不传它)")
    p.add_argument("--out-file", default=None,
                   help="把屏幕报告(--account/--list/--diagnose 及生成小结)直接写到此文件(UTF-8 带 BOM)。"
                        "比 PowerShell 的 `> x.txt` 稳：不会被 GBK 控制台编码弄成乱码")
    p.add_argument("--report", default=None,
                   help="导出'给人看'的测试用例表格(按扩展名: .csv 用 Excel 打开 / .html 网页)。"
                        "可与 --out 同时用(同时出 .sv 和报告)；只传 --report 则只出报告")
    p.add_argument("--export-claims", default=None, metavar="claims.json",
                   help="导出探针/force 网名声明清单(JSON)——红区 scan_rtl 校验器的输入契约：带进红区"
                        "校验'探针是真·RTL 输出网、force 网真实存在'。可与 --out 同时用")
    p.add_argument("--list", action="store_true", help="只列出可生成信号清单，不生成")
    p.add_argument("--account", action="store_true",
                   help="完整账目：列出每个 logic 信号 + mux 组的去向（生成/跳过原因/过滤原因），"
                        "一个不漏——核对'有没有谁被悄悄漏掉'")
    p.add_argument("--comments", action="store_true",
                   help="在 .sv 里加少量导航注释(文件头 + 每信号 1 行 // 名)；默认零注释(对齐真实模板)")
    p.add_argument("--owner-in-msg", action="store_true",
                   help="断言消息尾部追加 ', owner:<P列>'(英文)：仿真 log 里直接看出 fail 的是谁的信号。"
                        "消息前半段格式不变")
    p.add_argument("--sv-summary", action="store_true",
                   help="产物末尾加测试汇总：统计断言总数/正反例数 + 运行时真 FAIL 数。"
                        "会把整个语句体包进一层命名 begin/end 块(任何 task/initial 里贴入都合法)")
    p.add_argument("--diagnose", action="store_true",
                   help="覆盖诊断: 实测各输入被解析成 force(RO)/RF_WRITE(RW)/未知, "
                        "类型列有哪些写法, 有无 >16bit 输入(驱动会截断), 不生成")
    p.add_argument("--topout", action="store_true",
                   help="【Topout-rooted 路径，2026-06-23 新模型】要验信号 = Topout 页 B 列(顶层真名)，"
                        "cone 展到源寄存器、断言贴顶层真名(前后缀整类问题消失)。配合 --list(列清单+分类)/"
                        "--account(账目)/--report(报告)/默认(出 .sv)。clean cone 默认——级联/尾缀/前缀属"
                        "旧 logic-rooted 路径(排查/隔离用)，此路径不需要。")

    g = p.add_argument_group("信号筛选")
    g.add_argument("--owner", help="按 owner 筛选(逗号分隔，匹配 logic P 列)")
    g.add_argument("--signals", help="按信号名筛选(逗号分隔，K 全名或去位宽基名)")
    g.add_argument("--regex", help="按信号名正则筛选")
    g.add_argument("--exclude", help="排除这些信号(逗号分隔，K 全名或去位宽基名)")
    g.add_argument("--exclude-regex", help="按正则排除信号(如 'pll_n|_to_dsm' 排掉 datapath 中间信号)")
    g.add_argument("--type", help="按 type/suffix(M 列)筛选，如 to_mux,ls；mux 页的组类型固定为 mux"
                                  "（--type mux 即只出 mux 验证）")
    g.add_argument("--include-internal", action="store_true",
                   help="连 top_output=0 的内部信号也生成（默认只生成 top_output=1 的可验证输出；"
                        "内部信号在 RTL/ENV_RF 层探不到，会导致 elaboration 层级查找失败）")
    g.add_argument("--top-output-only", action="store_true",
                   help="（已是默认行为，保留兼容）只取 top_output=1")
    g.add_argument("--no-mux", action="store_true",
                   help="不生成 mux 页验证（默认 logic+mux 都生成；Excel 无 mux 页时此开关无意义）")
    g.add_argument("--mux-only", action="store_true",
                   help="只生成 mux 页验证（等价 --type mux）；与 --no-mux 互斥")

    g2 = p.add_argument_group("测试向量")
    g2.add_argument("--mode", choices=["min", "max"], default="min",
                    help="向量密度: min(默认)=控制全组合×1数据特征; max=多数据模式")
    g2.add_argument("--max-tests", type=int, default=256, help="单信号向量上限(默认256)")
    g2.add_argument("--exhaustive", action="store_true",
                    help="总输入位很少时做真·全穷举")
    g2.add_argument("--logic-mode", choices=["min", "max", "exhaustive"], default=None,
                    help="只设 logic 侧覆盖档(精简/全面/穷举)，与 mux 侧解耦；不传则跟随 --mode/--exhaustive")
    g2.add_argument("--mux-mode", choices=["min", "max", "exhaustive"], default=None,
                    help="只设 mux 侧覆盖档(精简/全面/穷举)，与 logic 侧解耦；不传则跟随 --mode/--exhaustive")

    g3 = p.add_argument_group("负向(异常)用例")
    g3.add_argument("--neg-signals", help="对这些信号追加负向用例(逗号分隔)")
    g3.add_argument("--neg-all", action="store_true", help="对所有选中信号加负向用例")
    g3.add_argument("--neg-mode", choices=["invert", "inc", "value"], default="invert",
                    help="造错方式: invert(按位取反,默认)/inc(+1)/value(固定值)")
    g3.add_argument("--neg-value", type=lambda x: int(x, 0), default=None,
                    help="--neg-mode value 时的固定错误值(支持0x前缀)")
    g3.add_argument("--neg-which", choices=["first", "all"], default="first",
                    help="对每个选中信号: first(仅第一个向量,默认)/all(每个向量)")
    g3.add_argument("--neg-file", choices=["inline", "separate"], default="inline",
                    help="负向用例放同文件(inline,默认)还是单独 *_neg.sv(separate)")

    g4 = p.add_argument_group("类型覆盖(解决名称/RO-RW判定问题)")
    g4.add_argument("--force-signals", help="强制按 RO(force) 处理的基名(逗号分隔)")
    g4.add_argument("--rfwrite-signals", help="强制按 RW(RF_WRITE) 处理的基名(逗号分隔)")
    g4.add_argument("--default-kind", choices=["RO", "RW"], default=None,
                    help="类型判不出时的兜底(默认保持未解析并报告)")
    g4.add_argument("--no-wire-fallback", action="store_true",
                    help="关闭 wire 兜底：非 RW 寄存器且查不到的输入不再默认 force，而是标 UNKNOWN 交人工")
    g4.add_argument("--include-risky", action="store_true",
                    help="强制生成含'不可驱动输入'(wire兜底/未解析)的信号（默认跳过，因为 force 不存在的 net "
                         "会导致 elaboration CUVUNF 失败；与 VBA 一致默认跳过这类信号）")
    g4.add_argument("--match-fortest", action="store_true",
                    help="复刻 for_test 覆盖面: = --include-internal + --include-risky, 不跳过任何信号"
                         "(含 top_output=0 内部信号、含 close_ready_flag 等不可驱动 wire 的信号)。"
                         "用于和 for_test 逐信号对照；产物可能 CUVUNF 跑不起来(预期内)。")
    g4.add_argument("--no-ref-suffix", "--no-to-logic-suffix", dest="append_to_logic",
                    action="store_false",
                    help="关闭输出【引用尾缀】：top_output=0 的输出被下游引用时，探针网名默认补尾缀(随 "
                         "Excel：logic 行引用→_to_logic、mux 页引用→_to_mux)；本旗标让它直接探基名网——"
                         "用于 <名>_to_logic/_to_mux 恰好撞了某个真实输入网的设计(如 Hi1108)。"
                         "_ls 与顶层输出不受影响。(--no-to-logic-suffix 为兼容别名)")
    g4.add_argument("--mux-ref-suffix", dest="append_to_mux", action="store_true",
                    help="【整设计】给 mux 输出也补去向尾缀(_to_mux/_to_logic)。mux 输出默认探裸名(端口尾缀"
                         "设计相关、Excel 推不出，WL 实证裸名)；本旗标对所有被引用的 mux 输出统一补——用于"
                         "端口真名带尾缀的设计(如 Hi1108 rxiq)。个别例外用 --no-suffix-signals 单独拉回裸名。")
    g4.add_argument("--no-suffix-signals",
                    help="单点【探裸名】：这些信号(基名/全名，逗号分隔)单独不补尾缀、直接探基名网——用于"
                         "logic 撞名信号(其 <名>_to_logic 恰是另一根真实输入网，如 lo2g5g)。压过全局开关。")
    g4.add_argument("--suffix-signals",
                    help="单点【探尾缀网】：这些信号(逗号分隔)单独探带去向尾缀(_to_logic/_to_mux)的网——"
                         "主要用于 mux 输出(默认探裸名，因端口尾缀设计相关)：其端口本身就叫 <名>_to_logic 时"
                         "(如 Hi1108 rxiq=2:1 mux 喂 sig_logic)单独开。压过类型默认。")
    g4.add_argument("--cascade-mode", choices=["cone", "force"], default="cone",
                    help="级联(输入/mux控制引用上游算出来的网)的【全局】驱动模式，详见 级联模式说明.md："
                         "cone(默认)=展开上游表达式、驱动其源头寄存器(纯 Excel，不需要探针前缀)；"
                         "force=直接 force 字面 _to_logic/_to_mux 网(隔离验证；需要 scan_rtl 前缀)。"
                         "logic/mux 可用下面两参单独覆盖。")
    g4.add_argument("--logic-cascade", choices=["cone", "force"], default=None,
                    help="【logic 信号】级联模式，覆盖 --cascade-mode（缺省=跟随 --cascade-mode）。")
    g4.add_argument("--mux-cascade", choices=["cone", "force"], default=None,
                    help="【mux 信号】级联模式，覆盖 --cascade-mode（缺省=跟随 --cascade-mode）。"
                         "mux 控制是深级联(另一 mux 的输出)时常用 force 直接 force 控制衔接网更稳。")
    g4.add_argument("--probe-prefix", action="append", default=[], metavar="信号=层级路径",
                    help="信号网在 ENV_RF 子模块里时的探针前缀，可多次。"
                         "如 --probe-prefix pll_n=U_BT_LP_PLL_DIG → 断言写 "
                         "`ENV_RF.U_BT_LP_PLL_DIG.pll_n[31:0]；force 输入 wire 同理")
    g4.add_argument("--probe-prefix-file", default=None, metavar="映射.txt",
                    help="从映射文件读探针前缀（每行 信号名=层级路径，# 为注释）。"
                         "与 GUI『探针前缀映射→导出』的文件格式一致，可直接复用")
    return p


def _parse_probe_prefixes(items, prefix_file=None):
    """['pll_n=U_BT_LP_PLL_DIG', ...] + 映射文件 → {'pll_n': 'U_BT_LP_PLL_DIG'}。
    命令行与文件同名时命令行优先。格式错给出明确报错。"""
    out = {}
    if prefix_file:
        if not os.path.isfile(prefix_file):
            sys.exit("--probe-prefix-file 找不到文件: %s" % prefix_file)
        with open(prefix_file, "r", encoding="utf-8") as f:
            out.update(generator.parse_probe_prefix_lines(f.read()))
    for it in items or []:
        if "=" not in it:
            sys.exit("--probe-prefix 格式应为 信号名=层级路径，收到: %r" % it)
        name, prefix = it.split("=", 1)
        if not name.strip() or not prefix.strip():
            sys.exit("--probe-prefix 信号名与层级路径都不能为空: %r" % it)
        out[name.strip()] = prefix.strip()
    return out


def cmd_list(wb, opts):
    sigs = generator.select_signals(wb, opts)
    muxes = generator.select_mux_groups(wb, opts)
    print("可生成信号 %d (logic %d + mux %d) / logic 总行 %d, mux 组 %d (tmm字段=%d, regmap信号=%d)"
          % (len(sigs) + len(muxes), len(sigs), len(muxes),
             len(wb.logic), len(wb.mux), len(wb.tmm), len(wb.regmap)))
    print("%-5s %-40s %-12s %-12s %-3s %s" % ("R", "输出名(K)", "owner", "type", "top", "表达式"))
    print("-" * 100)
    for s in sigs:
        print("%-5s %-40s %-12s %-12s %-3s %s"
              % (s.assert_id, s.out_name[:40], (s.owner or "")[:12],
                 (s.suffix or "")[:12], s.top_output, s.expr[:50]))
    for g in muxes:
        # 控制信号描述：单控制=基名；多控制=按列序拼接 {c1,c2}（B 高位），与真值表/报告同口径
        expr_text = "case(%s) %d 选 1" % (generator._mux_ctrl_desc(g), len(g.cases))
        print("%-5s %-40s %-12s %-12s %-3s %s"
              % (g.assert_id, g.out_name[:40], (g.owner or "")[:12],
                 "mux", g.top_output, expr_text[:50]))


def cmd_account(wb, opts):
    """完整账目：跑一遍 build，把每个 logic 信号 + mux 组的去向列全——一个不漏。"""
    res = generator.build(wb, opts)
    items = generator.compose_account(wb, opts, res)
    # 按去向归类统计
    from collections import OrderedDict
    by_disp = OrderedDict()
    for it in items:
        by_disp.setdefault(it["disposition"], []).append(it)
    print("完整账目：%d 个 logic 信号 + %d 个 mux 组，逐个去向（一个不漏）"
          % (len(wb.logic), len(wb.mux)))
    print("  小结：" + " | ".join("%s %d" % (d, len(v)) for d, v in by_disp.items()))
    print("-" * 100)
    print("%-7s %-10s %-44s %s" % ("kind", "去向", "名字(R)", "原因/说明"))
    print("-" * 100)
    for it in items:
        tag = "%s(%s)" % (it["name"][:38], it["aid"])
        print("%-7s %-10s %-44s %s"
              % (it["kind"], it["disposition"], tag, (it["reason"] or "")[:60]))
    print("-" * 100)
    print("提示：'过滤'=被筛选条件/默认 top_output 滤掉（--include-internal 纳入内部节点）；"
          "'跳过'=解析问题/缺前缀（原因见上，--include-risky 可强制）；'生成(裸名探针)'=已进 .sv。")


def cmd_topout(args, wb, opts):
    """【Topout-rooted 路径】要验信号 = Topout 页 B 列；cone 展到源寄存器、断言贴顶层真名。

    --list 列清单+分类 / --account 账目(默认不空、不刷 top_out=0 假警告) / --report 出报告 /
    默认出 .sv。clean cone 默认（前后缀/级联/top_output 筛属旧 logic-rooted 路径，此路径不需要）。"""
    from . import topout as T
    mode, mt, exh = opts.mode, opts.max_tests, opts.exhaustive
    if not wb.topout:
        print("⚠ 该 Excel 无 Topout 页（B 列要验信号清单为空）——无法走 Topout 路径。"
              "请确认表里有 Topout 页；不带 --topout 则走旧 logic-rooted 默认路径。")
        return 0

    if args.list:
        ms = T.topout_view_models(wb, mode=mode, max_tests=mt, exhaustive=exh)
        by_kind = {}
        for m in ms:
            by_kind[m["kind"]] = by_kind.get(m["kind"], 0) + 1
        print("Topout 要验信号 %d 个（B 列锚点）：%s"
              % (len(ms), " | ".join("%s %d" % (k, v) for k, v in sorted(by_kind.items()))))
        print("%-40s %-12s %-12s %-9s %-5s %s"
              % ("Topout信号(B列)", "owner", "分类", "状态", "用例", "说明"))
        print("-" * 112)
        for m in ms:
            note = (m["note"] or "; ".join(m["issues"]) or "").replace("\n", " ")[:38]
            print("%-40s %-12s %-12s %-9s %-5d %s"
                  % (m["name"][:40], (m["owner"] or "")[:12], m["kind"], m["status"],
                     m["n_vectors"], note))
        return 0

    if args.account:
        from . import resolver as R
        acc = T.compose_topout_account(wb, R.Resolver(wb), mode=mode, max_tests=mt, exhaustive=exh)
        s = acc["summary"]
        print("Topout 账目：%d 个要验信号（默认不空，不套 top_output_only）" % s["n_total"])
        print("  小结：可建 %d | 跳过(RO) %d | 未解析 %d | error %d | 有 issues %d"
              % (s["n_ok"], s["n_skip"], s["n_unresolved"], s["n_error"], s["n_with_issues"]))
        print("-" * 112)
        print("%-40s %-12s %-12s %-9s %-5s %s"
              % ("信号", "owner", "分类", "状态", "叶子", "原因/说明"))
        print("-" * 112)
        for r in acc["rows"]:
            why = ("; ".join(r["issues"]) or r["note"] or "").replace("\n", " ")[:42]
            print("%-40s %-12s %-12s %-9s %-5d %s"
                  % (r["name"][:40], (r["owner"] or "")[:12], r["kind"], r["status"],
                     r["n_leaves"], why))
        return 0

    if args.report:
        rep = T.topout_report(wb, mode=mode, max_tests=mt, exhaustive=exh)
        written = write_report(args.report, rep, args.excel)
        print("Topout 报告已写出: %s" % "  ".join(written))
        print("  Topout 信号 %d 个（真值表 %d 个）" % (len(rep["summary"]), len(rep["tables"])))
        if args.out is None:
            return 0

    out = args.out or "wr_rf_tc.sv"
    text, b = T.render_topout_sv(wb, mode=mode, max_tests=mt, exhaustive=exh,
                                 comments=opts.comments, sv_summary=opts.sv_summary,
                                 owner_in_msg=opts.owner_in_msg)
    _write(out, text)
    s = b["summary"]
    print("已写出(Topout 路径): %s" % out)
    print("  Topout 信号 %d；产出断言块 %d；向量 %d（负向 %d）；记账(不产断言) %d"
          % (s["n_total"], s["n_emitted"], s["n_vectors"], s["n_negative"], s["n_accounted"]))
    if b["accounted"]:
        print("  记账（解析不了/跳过的信号也不静默丢）：")
        for a in b["accounted"]:
            print("    - %s [%s/%s]: %s"
                  % (a["name"], a["kind"], a["status"], (a["reason"] or "").replace("\n", " ")[:60]))
    return 0


def cmd_diagnose(wb, opts):
    d = generator.diagnose(wb, opts)
    print("\n===== 覆盖诊断 =====")
    print("参与诊断信号: %d" % d["n_signals"])
    print("\n[total_memory_map H 列(类型)原文分布]  ← 看是否只有 RO/RW 系，有无没覆盖的写法")
    for k, v in d["tmm_type_raw"].items():
        print("   %-16s x %d" % (k, v))
    print("[regmap F 列(Reg Type)原文分布]")
    for k, v in d["regmap_type_raw"].items():
        print("   %-16s x %d" % (k, v))
    print("[total_memory_map D 列(DIG TOP PIN)分布]")
    for k, v in d["tmm_dig_top_pin"].items():
        print("   %-16s x %d" % (k, v))

    c = d["cats"]
    total = sum(c.values())
    print("\n[输入驱动方式分类]")
    print("   RF_WRITE (RW 寄存器)       : %d" % c["rfwrite"])
    print("   force - RO 寄存器/管脚      : %d" % c["force_ro"])
    print("   force - 级联/自引用前级信号(logic 输出): %d" % c["force_chained"])
    print("   force - wire 兜底(表中查无)  : %d  ← 需你确认这些确实是 wire" % c["force_wire"])
    print("   UNKNOWN (仍无法处理)        : %d" % c["unknown"])
    if total:
        ok = total - c["unknown"]
        print("   → 可生成(force/RF_WRITE)覆盖率: %.1f%% (%d/%d)" % (100.0 * ok / total, ok, total))

    if d["wide_inputs"]:
        print("\n[>16bit 输入] force 会按位宽自适应(如 32'h)不截断；但若它是 RW 寄存器，RF_WRITE 仍 16'h 受限:")
        for name, ltr, base, w, kind, src in d["wide_inputs"][:30]:
            print("   %s 的 %s=%s (%dbit, %s, %s)" % (name, ltr, base, w, kind, src))
        if len(d["wide_inputs"]) > 30:
            print("   ...(共 %d 条)" % len(d["wide_inputs"]))
    else:
        print("\n（无 >16bit 输入）")

    if d["fallback_wires"]:
        print("\n⚠ [wire 兜底] 表中查无、按 wire 直接 force 的输入——请确认它们确实是 wire/管脚/中间信号；")
        print("   若其实是寄存器，用 --rfwrite-signals 指定，否则会 force 而非 RF_WRITE:")
        for name, ltr, base, w in d["fallback_wires"][:40]:
            print("   %s 的 %s=%s (%dbit)" % (name, ltr, base, w))
        if len(d["fallback_wires"]) > 40:
            print("   ...(共 %d 条)" % len(d["fallback_wires"]))

    if d["unknown"]:
        print("\n⚠ [UNKNOWN] 仍无法处理(多为命名歧义)，需人工核对:")
        for name, ltr, base, note in d["unknown"][:40]:
            print("   %s 的 %s=%s: %s" % (name, ltr, base, note))
        if len(d["unknown"]) > 40:
            print("   ...(共 %d 条)" % len(d["unknown"]))
    else:
        print("\n✅ 无 UNKNOWN：所有被引用输入都已归类为 force 或 RF_WRITE。")


SUMMARY_COLS = [("R", "R/序号"), ("signal", "信号(K)"), ("owner", "owner"), ("type", "类型"),
                ("top", "top_output"), ("n_tests", "用例数"), ("n_neg", "负向数"),
                ("control", "控制位"), ("data", "数据位"), ("expr", "表达式"),
                ("unresolved", "未解析输入"), ("supplement", "RTL补充⚠"), ("error", "错误")]
DETAIL_COLS = [("R", "R"), ("signal", "信号(K)"), ("owner", "owner"), ("type", "类型"),
               ("test", "用例"), ("neg", "负向"), ("auto_out", "auto_out(表达式计算)"),
               ("expected", "期望(断言对比值)"), ("exp_src", "期望来源"),
               ("force", "force 驱动"),
               ("rfwrite", "RF_WRITE 驱动"), ("expr", "表达式"), ("note", "备注")]
VERIF_COLS = [("R", "R"), ("signal", "信号(K)"), ("owner", "owner"), ("type", "类型"),
              ("top", "top_output"), ("status_label", "可验证性"),
              ("detail", "风险输入 / 原因"), ("out_net", "断言输出 net")]
# 可验证性状态 → 给人看的标签 + 颜色等级(用于 HTML 高亮/CSV 文字)
VERIF_STATUS = {
    "clean":         ("✅ 可验证", "ok"),
    "wire-fallback": ("⚠ 存疑(按名 force，elaboration 最易 CUVUNF)", "warn"),
    "unresolved":    ("✗ 不可验证(输入未解析)", "bad"),
    "parse-err":     ("✗ 表达式解析失败", "bad"),
    "spec-collision": ("✗ 规格冲突·待 designer 核对(同选择值选不同源)", "bad"),
}


def write_report(path, rep, excel):
    base, ext = os.path.splitext(path)
    if ext.lower() in (".html", ".htm"):
        _write_report_html(path, rep, excel)
        return [path]
    if ext.lower() == ".xlsx":
        _write_report_xlsx(path, rep, excel)
        return [path]
    # CSV：明细写到给定路径，汇总写到 *_summary.csv
    import csv
    detail_path = path if ext.lower() == ".csv" else base + ".csv"
    summary_path = base + "_summary.csv"
    with open(summary_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([h for _k, h in SUMMARY_COLS])
        for r in rep["summary"]:
            w.writerow([r.get(k, "") for k, _h in SUMMARY_COLS])
    with open(detail_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([h for _k, h in DETAIL_COLS])
        for r in rep["detail"]:
            w.writerow([r.get(k, "") for k, _h in DETAIL_COLS])
    written = [detail_path, summary_path]
    verif = rep.get("verifiability")
    if verif and verif.get("signals"):
        verif_path = base + "_verifiability.csv"
        with open(verif_path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow([h for _k, h in VERIF_COLS])
            for r in verif["signals"]:
                lbl = VERIF_STATUS.get(r.get("status", ""), (r.get("status", ""), ""))[0]
                row = dict(r, status_label=lbl)
                w.writerow([row.get(k, "") for k, _h in VERIF_COLS])
        written.append(verif_path)
    return written


_REPORT_CSS = """
:root{--bd:#ccc;--hd:#f0f3f7}
*{box-sizing:border-box}
body{font-family:"Segoe UI",Microsoft YaHei,sans-serif;margin:0;color:#222}
header{position:sticky;top:0;background:#fff;border-bottom:1px solid var(--bd);
 padding:14px 24px 0;z-index:10;box-shadow:0 2px 6px rgba(0,0,0,.04)}
h1{font-size:19px;margin:0 0 4px} h3{font-size:13px;margin:22px 0 0}
.sum{color:#555;margin:2px 0;font-size:12px}
.toolbar{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin:10px 0}
.toolbar input[type=text]{padding:5px 9px;border:1px solid var(--bd);border-radius:5px;
 width:300px;font-size:13px}
.toolbar select{padding:5px;border:1px solid var(--bd);border-radius:5px}
.toolbar label{font-size:13px;color:#444;user-select:none}
/* owner 多选下拉：按钮 + 浮层复选列表（任一勾中即显示，OR） */
.msel{position:relative;display:inline-block}
.mselbtn{padding:5px 9px;border:1px solid var(--bd);border-radius:5px;background:#fff;cursor:pointer;
 font-size:13px;color:#333}
.mselbtn:hover{background:#f4f6f9}
.mselpop{display:none;position:absolute;z-index:50;top:100%;left:0;margin-top:3px;max-height:260px;
 overflow:auto;background:#fff;border:1px solid var(--bd);border-radius:6px;
 box-shadow:0 3px 10px rgba(0,0,0,.15);padding:4px 0;min-width:170px}
.msel.open .mselpop{display:block}
.mselpop label{display:block;padding:4px 12px;font-size:13px;white-space:nowrap;cursor:pointer;color:#333}
.mselpop label:hover{background:#eef3fb}
.mselpop input{margin-right:7px;vertical-align:middle}
.mselclr{border-top:1px solid var(--bd);color:#06c;margin-top:2px;padding-top:6px}
#count{color:#888;font-size:12px;margin-left:auto}
.tabs{display:flex;gap:4px;margin-top:8px}
.tabbtn{border:1px solid var(--bd);border-bottom:none;background:#f4f6f9;cursor:pointer;
 padding:7px 16px;font-size:13px;border-radius:6px 6px 0 0;color:#555}
.tabbtn.active{background:#fff;color:#1558d6;font-weight:600;box-shadow:0 -2px 0 #1558d6 inset}
main{padding:6px 24px 40px}
.tab{display:none} .tab.active{display:block}
table{border-collapse:collapse;font-size:12px;margin-top:8px}
table.full,table.sum{width:100%}
th,td{border:1px solid var(--bd);padding:4px 6px;text-align:left;vertical-align:top}
th{background:var(--hd)} thead th{position:sticky;top:128px}
/* 真值表每块表头不 sticky：一页几十张表 = 几百个 sticky 元素，鼠标一动就全量重算→卡。
   每块独立、滚动时钉住价值不大；扁平表(汇总/明细/可验证性)仍保留单一 sticky 表头。 */
.tt thead th{position:static;top:auto}
tbody tr:nth-child(even){background:#fafbfc}
tr.neg{background:#fff3f3} tr.err{background:#ffe9e9}
.tt td,.tt th{text-align:center} .tt th.rowhdr{text-align:left;background:#eef1f5;white-space:nowrap}
.tt td.neg,.tt th.negh{background:#ffe3e3;color:#a40000}
.tt tr.exprow th,.tt tr.exprow td{font-weight:600;border-top:2px solid #999}
.tt td.drv{font-family:Consolas,monospace;font-size:11px;text-align:left;color:#555}
/* 屏幕外的真值表块跳过布局/绘制：一页几十张表，窗口缩放/滚动只算可见的几张，不再全量重排。
   contain-intrinsic-size 给屏外块占位高度(auto=记住实测值)，滚动条稳定。检查切换的重排也跟着变轻。 */
.ttblock{margin-bottom:6px;content-visibility:auto;contain-intrinsic-size:auto 320px}
.ex{color:#888} code{background:#f5f5f5;padding:0 3px}
pre.chain{background:#f4f8ff;border:1px solid #d4e0f5;border-radius:4px;padding:8px 10px;
 font-family:Consolas,monospace;font-size:12px;color:#234;margin:6px 0 2px;overflow-x:auto}
.empty{color:#999;padding:20px}
/* auto_out 行(表达式计算值，只读参考) 与 期望 行(进 .sv 的对比值) */
.tt tr.autorow th,.tt tr.autorow td{color:#666;font-style:italic;border-top:2px solid #999}
.tt tr.exprow{border-top:none}
.tt td.dsgn{background:#e2f4e2}        /* designer 手填期望(且非负向) */
.tt td.fb{color:#999}                  /* 未手填 -> auto_out 兜底(灰) */
.tt td.dsgndiff{background:#ffd9d9}    /* 手填期望 != auto_out: 表达式可能与 designer 意图不符 */
/* ── 真值表检查 tab：遮盖/填空/判定 ── */
.chkbar{display:flex;gap:14px;align-items:center;margin:10px 0;padding:10px;
 background:#f4f7ff;border:1px solid #c8d6f5;border-radius:6px;flex-wrap:wrap}
.chkbar button{padding:7px 18px;font-size:13px;cursor:pointer;border:1px solid #1558d6;
 border-radius:5px;background:#1558d6;color:#fff;font-weight:600}
.chkbar button.off{background:#fff;color:#1558d6}
#chkscore{font-size:13px;color:#333}
#chkscore b.ok{color:#1a7f37} #chkscore b.bad{color:#c00}
.chkhint{font-size:12px;color:#666;flex-basis:100%}
/* 遮盖格：文本由 JS 换成 "?"（不靠 CSS 透明——会被 tr.autorow 的颜色规则按优先级盖回去），
   这里只负责底色与禁选 */
.tt td.masked{background:#e3e3e3 !important;color:#999 !important;font-style:normal;user-select:none}
.cin{width:76px;font-family:Consolas,monospace;font-size:12px;padding:2px 5px;
 border:1px solid #1558d6;border-radius:3px;text-align:right}
.cin.okin{border-color:#1a7f37;background:#e2f4e2}
.cin.badin{border-color:#c00;background:#ffd9d9}
.tt td.okc{background:#d9f0d9 !important}
.tt td.badc{background:#ffd2d2 !important}
.tt th.okh{background:#1a7f37;color:#fff}
.tt th.badh{background:#c00;color:#fff}
.vstat{font-weight:600;white-space:nowrap}
.vstat.ok{color:#1a7f37} .vstat.warn{color:#b26a00} .vstat.bad{color:#c00}
.vsum{margin:10px 0;font-size:13px} .vsum b{margin-right:14px}
tr.vrow.warn{background:#fff8ec} tr.vrow.bad{background:#ffecec}
/* ── 分页器（窗口化懒渲染：每页只把当前窗注入 DOM，避免 6000 组一次性全进页面卡死） ── */
.pagenav{display:flex;gap:10px;align-items:center;margin:10px 0 4px;font-size:13px;
 flex-wrap:wrap;background:#f7f9fc;border:1px solid var(--bd);border-radius:6px;padding:7px 10px}
.pagenav.pgbot{margin-top:12px}        /* 每页底部再放一个，读完不必滚回顶 */
.pgb{padding:5px 13px;border:1px solid var(--bd);border-radius:5px;background:#fff;
 cursor:pointer;font-size:13px;color:#444}
.pgb:hover{background:#e9eef6} .pgb:active{background:#dde6f3}
.pgi{color:#555} .pgper{color:#666;margin-left:auto}
.pgsel{padding:3px 5px;border:1px solid var(--bd);border-radius:4px}
.pgbody.ttwrap{min-height:40px}
/* ── 真值表「按值筛选」：第0列每个变化的输入挂下拉；选定取值后只留匹配的测试列(隐藏其余整列) ── */
.tt .ttfx{display:none}
.tt th.rowhdr input.ttf{margin-left:6px;font-size:11px;padding:1px 4px;border:1px solid #9fb4d6;
 border-radius:3px;background:#fff;vertical-align:middle;width:48px;text-align:center}
.tt th.rowhdr input.ttf.set{border-color:#1558d6;background:#eaf1ff;font-weight:700;color:#1558d6}
.tt th.rowhdr input.ttf::placeholder{color:#aab;font-weight:400}
.ttfbar{margin-left:14px;font-size:12px;color:#666;font-weight:400;white-space:nowrap}
.ttfbar .ttfcount{font-family:Consolas,monospace}
.ttfbar .ttfcount.on{color:#1558d6;font-weight:700}
.ttfclr{margin-left:6px;font-size:11px;padding:1px 8px;border:1px solid var(--bd);
 border-radius:4px;background:#fff;cursor:pointer;color:#555}
.ttfclr:hover{background:#eef2f8}
.suppbar{margin:4px 0 8px;padding:5px 10px;border:1px solid #e6a23c;border-left:4px solid #e6a23c;
 border-radius:4px;background:#fdf6ec;color:#a05a00;font-size:12px;line-height:1.5}
"""

_REPORT_JS = """
(function(){
  function $(id){return document.getElementById(id);}
  function jget(id){var el=$(id);return el?JSON.parse(el.textContent):[];}
  var q=$('q'),no=$('negonly'),cnt=$('count');
  var sig=$('sig'),typ=$('typ'),topf=$('topf');
  var ownerSel=[];        /* owner 多选：勾中的值数组（含 '__no_owner__' 哨兵），空=全部 */
  function ownerLabel(v){return v==='__no_owner__'?'（无 owner）':v;}
  var PER=[30,50,100,200];
  /* 每个 tab 一份数据 + 当前页/每页。重的真值表(②③)共用一份(check=True 标记)，③ 不再复制第二份 DOM。
     渲染只把"当前窗"注入 body，DOM 节点数恒定 ⇒ 6000 组也不卡。 */
  var SEC={
    sum:{data:jget('sum-data'),per:200,page:0},
    tt :{data:jget('tt-data'), per:50, page:0},
    chk:{data:null,            per:50, page:0,ischk:true},
    det:{data:jget('det-data'),per:200,page:0},
    ver:{data:jget('ver-data'),per:200,page:0}
  };
  SEC.chk.data=SEC.tt.data;
  var KEYS=['sum','tt','chk','det','ver'];

  function flt(){
    var qq=(q.value||'').toLowerCase(),nn=no.checked,
        ss=sig.value,tt=typ.value,tp=topf.value,osel=ownerSel;
    return function(it){
      if(qq&&it.t.indexOf(qq)<0)return false;
      if(osel.length){                                      /* owner 多选 OR：任一命中即留 */
        var hit=false,k;
        for(k=0;k<osel.length;k++){
          if(osel[k]==='__no_owner__'){if(it.o==='')hit=true;}   /* 「（无 owner）」 */
          else if(it.o===osel[k])hit=true;
          if(hit)break;
        }
        if(!hit)return false;
      }
      if(ss&&it.s!==ss)return false;                        /* 信号名下拉(逐字匹配) */
      if(tt&&it.ty!==tt)return false;                       /* 类型下拉 */
      if(tp!==''&&String(it.tp)!==tp)return false;          /* top_output 下拉 */
      if(nn&&it.n!==1)return false;
      return true;
    };
  }
  function filtered(key){return SEC[key].data.filter(flt());}

  function navHtml(key,page,np,total,per){
    var opts='',i;
    for(i=0;i<PER.length;i++)opts+='<option'+(PER[i]===per?' selected':'')+'>'+PER[i]+'</option>';
    return '<button class="pgb" data-k="'+key+'" data-d="-99">«</button>'+
      '<button class="pgb" data-k="'+key+'" data-d="-1">◀ 上一页</button>'+
      '<span class="pgi">第 '+(total?page+1:0)+' / '+np+' 页（共 '+total+' 条）</span>'+
      '<button class="pgb" data-k="'+key+'" data-d="1">下一页 ▶</button>'+
      '<button class="pgb" data-k="'+key+'" data-d="99">»</button>'+
      '<label class="pgper">每页 <select class="pgsel" data-k="'+key+'">'+opts+'</select></label>';
  }

  function render(key){
    var sec=SEC[key],body=$(key+'-body'),nav=$(key+'-nav');
    if(!body)return;
    var list=filtered(key),per=sec.per,np=Math.max(1,Math.ceil(list.length/per));
    if(sec.page>=np)sec.page=np-1; if(sec.page<0)sec.page=0;
    if(nav)nav.innerHTML=navHtml(key,sec.page,np,list.length,per);
    if(!list.length){body.innerHTML='<p class="empty">（无匹配）</p>';return;}
    var s=sec.page*per,win=list.slice(s,s+per),h='',i;
    for(i=0;i<win.length;i++)h+=win[i].h;
    /* 页尾再放一个分页条：读完本页不必滚回顶（按钮走文档级委托，重建 innerHTML 也有效） */
    body.innerHTML=h+'<div class="pagenav pgbot">'+navHtml(key,sec.page,np,list.length,per)+'</div>';
    if(sec.ischk&&chkon){applyChk(body,true);resetScore();}   /* 翻到新一页时按本页重置自测 */
  }

  var active='sum';
  function tab(name){
    active=name;
    var bs=document.querySelectorAll('.tabbtn'),ss=document.querySelectorAll('.tab'),i;
    for(i=0;i<bs.length;i++)bs[i].classList.toggle('active',bs[i].getAttribute('data-tab')===name);
    for(i=0;i<ss.length;i++)ss[i].classList.toggle('active',ss[i].id===name);
    render(name);                       /* 切到才渲染：未访问的 tab 不进 DOM */
  }
  var btns=document.querySelectorAll('.tabbtn');
  for(var bi=0;bi<btns.length;bi++)(function(b){b.onclick=function(){tab(b.getAttribute('data-tab'));};})(btns[bi]);

  /* ===== 真值表「按值筛选」：第0列下拉选定输入取值 → 只留匹配的测试列（整列 display:none）。
     只在 change/清除 时跑一次、且只扫被操作的那一块(closest .ttblock)；翻页重渲后下拉自动复位(全部)。
     用 data-c 给每列每格打标，靠一次 classList 切换整列显隐，不逐格轮询 → 128 列也不卡。 ===== */
  function ttBlock(el){while(el&&!(el.classList&&el.classList.contains('ttblock')))el=el.parentNode;return el;}
  function filterBlock(blk){
    if(!blk)return;
    var sels=blk.querySelectorAll('.ttf'),i,j,cons=[];
    for(i=0;i<sels.length;i++){
      var on=(sels[i].value||'').trim()!=='';
      sels[i].classList.toggle('set',on);
      if(on)cons.push(sels[i]);
    }
    var hide={};                                   /* 任一手填输入不匹配 → 该列隐藏(AND 语义) */
    for(i=0;i<cons.length;i++){
      var ri=cons[i].getAttribute('data-ri'),want=(cons[i].value||'').trim();
      var row=blk.querySelector('tr.inrow[data-ri="'+ri+'"]');
      if(!row)continue;
      var vc=row.querySelectorAll('td[data-c]');
      for(j=0;j<vc.length;j++){
        if((vc[j].textContent||'').trim()!==want)hide[vc[j].getAttribute('data-c')]=1;
      }
    }
    var all=blk.querySelectorAll('[data-c]');
    for(i=0;i<all.length;i++)all[i].classList.toggle('ttfx',hide[all[i].getAttribute('data-c')]===1);
    var heads=blk.querySelectorAll('thead [data-c]'),total=heads.length,shown=0;
    for(i=0;i<heads.length;i++)if(hide[heads[i].getAttribute('data-c')]!==1)shown++;
    var cs=blk.querySelector('.ttfcount');
    if(cs){cs.textContent=shown+'/'+total;cs.classList.toggle('on',shown!==total);}
  }

  /* 翻页/每页：文档级委托，data-k 指明属于哪个 tab（顶部条与页尾条共用一套处理） */
  document.addEventListener('click',function(e){
    var b=e.target;
    if(b&&b.classList&&b.classList.contains('ttfclr')){      /* 真值表「清除筛选」：清空本块所有输入框 */
      var blk=ttBlock(b),sels=blk?blk.querySelectorAll('.ttf'):[],i;
      for(i=0;i<sels.length;i++)sels[i].value='';
      filterBlock(blk);return;
    }
    while(b&&b.nodeType===1&&!(b.classList&&b.classList.contains('pgb')))b=b.parentNode;
    if(!b||b.nodeType!==1||!b.classList.contains('pgb'))return;
    var sec=SEC[b.getAttribute('data-k')];if(!sec)return;
    var d=parseInt(b.getAttribute('data-d'),10);
    if(d===-99)sec.page=0;else if(d===99)sec.page=1e9;else sec.page+=d;
    render(b.getAttribute('data-k'));
  });
  document.addEventListener('change',function(e){
    var t=e.target;
    if(t.classList&&t.classList.contains('ttf')){filterBlock(ttBlock(t));return;}   /* 真值表按值筛选 */
    if(!t.classList||!t.classList.contains('pgsel'))return;
    var key=t.getAttribute('data-k'),sec=SEC[key];if(!sec)return;
    sec.per=parseInt(t.value,10);sec.page=0;render(key);
  });
  document.addEventListener('input',function(e){      /* 真值表按值筛选：边打边筛（手填输入框） */
    var t=e.target;
    if(t.classList&&t.classList.contains('ttf'))filterBlock(ttBlock(t));
  });

  function applyAll(){
    for(var i=0;i<KEYS.length;i++)SEC[KEYS[i]].page=0;
    render(active);                     /* 只重渲当前 tab；其余切过去时再渲 */
    cnt.textContent='匹配信号 '+filtered('sum').length;
  }
  q.oninput=applyAll;no.onchange=applyAll;
  sig.onchange=applyAll;typ.onchange=applyAll;topf.onchange=applyAll;

  /* ===== owner 多选下拉：按钮开合 + 复选项勾选(OR) + 清除 + 点外部收起 ===== */
  var ownerBox=$('ownerbox'),ownerBtn=$('ownerbtn'),ownerPop=$('ownerpop'),ownerClr=$('ownerclr');
  function readOwnerSel(){
    var cks=document.querySelectorAll('.ownerck'),arr=[],i;
    for(i=0;i<cks.length;i++)if(cks[i].checked)arr.push(cks[i].value);
    ownerSel=arr;
    if(!arr.length)ownerBtn.textContent='全部 owner ▾';
    else if(arr.length===1)ownerBtn.textContent=ownerLabel(arr[0])+' ▾';
    else ownerBtn.textContent='owner ×'+arr.length+' ▾';
  }
  if(ownerBtn){
    ownerBtn.onclick=function(e){e.stopPropagation();ownerBox.classList.toggle('open');};
    ownerPop.onclick=function(e){e.stopPropagation();};        /* 浮层内点击不冒泡(不收起) */
    var ocks=document.querySelectorAll('.ownerck'),j;
    for(j=0;j<ocks.length;j++)ocks[j].onchange=function(){readOwnerSel();applyAll();};
    if(ownerClr)ownerClr.onclick=function(e){
      e.preventDefault();e.stopPropagation();
      var cs=document.querySelectorAll('.ownerck'),k;
      for(k=0;k<cs.length;k++)cs[k].checked=false;
      readOwnerSel();applyAll();
    };
    document.addEventListener('click',function(){ownerBox.classList.remove('open');});
  }

  /* ===== 真值表检查（designer 自测：遮 auto_out、期望变填空、回车判定）。按当前页进行。 ===== */
  var chkbtn=$('chkbtn'),chkscore=$('chkscore');
  var chkon=false,okn=0,badn=0,answered=0;
  /* 宽容解析 designer 输入的数值：0x../0b../16'h../'b../'d../十进制/裸hex（与 GUI 同语义，十进制优先于裸hex） */
  function parseVal(s){
    s=String(s||'').trim().toLowerCase().replace(/[_\\s]/g,'');
    if(!s)return null;
    var m;
    if(m=s.match(/^(?:\\d+)?'h([0-9a-f]+)$/))return parseInt(m[1],16);
    if(m=s.match(/^(?:\\d+)?'b([01]+)$/))return parseInt(m[1],2);
    if(m=s.match(/^(?:\\d+)?'d(\\d+)$/))return parseInt(m[1],10);
    if(/^0x[0-9a-f]+$/.test(s))return parseInt(s.slice(2),16);
    if(/^0b[01]+$/.test(s))return parseInt(s.slice(2),2);
    if(/^\\d+$/.test(s))return parseInt(s,10);
    if(/^[0-9a-f]+$/.test(s))return parseInt(s,16);
    return NaN;
  }
  function chkBody(){return $('chk-body');}
  function updScore(){
    var cb=chkBody(),total=cb?cb.querySelectorAll('td.cquiz').length:0;
    chkscore.innerHTML='本页已检查 '+answered+'/'+total+' · <b class="ok">一致 '+okn+'</b> · <b class="bad">不一致 '+badn+'</b>';
  }
  function resetScore(){okn=0;badn=0;answered=0;updScore();}
  /* 遮盖/揭示 = 直接换单元格文本（CSS 透明会被 tr.autorow 的颜色规则按优先级盖回去，靠不住）。
     原文存在 data-t 里，揭示时取回。 */
  function maskCell(cell){
    if(cell.getAttribute('data-t')===null)cell.setAttribute('data-t',cell.textContent);
    cell.textContent='?';
    cell.classList.add('masked');
  }
  function unmaskCell(cell){
    var t=cell.getAttribute('data-t');
    if(t!==null)cell.textContent=t;
    cell.classList.remove('masked');
  }
  /* 分帧调度：一次性显示整页几百个 <input> 会触发一次巨大同步重排、主线程冻结(鼠标卡)。
     改成每帧处理一批、帧间让步，把"一次长冻结"摊成几帧，鼠标全程不卡。 */
  var chkJob=0;
  var raf=window.requestAnimationFrame?function(f){return window.requestAnimationFrame(f);}
                                       :function(f){return setTimeout(f,16);};
  var caf=window.cancelAnimationFrame?function(h){window.cancelAnimationFrame(h);}
                                      :function(h){clearTimeout(h);};
  /* 对一个容器(当前渲染的 chk 窗)套用/撤销检查态：遮 auto_out、期望格隐值露填空、清旧判定色。
     done 在全部处理完后回调。 */
  function applyChk(container,on,done){
    if(chkJob){caf(chkJob);chkJob=0;}
    if(!container){if(done)done();return;}
    /* 旧判定色量小，立即清掉 */
    var marked=container.querySelectorAll('.okc,.badc,.okh,.badh'),i;
    for(i=0;i<marked.length;i++)marked[i].classList.remove('okc','badc','okh','badh');
    var autos=container.querySelectorAll('td.cauto'),qz=container.querySelectorAll('td.cquiz');
    var ai=0,qi=0,BATCH=80;
    function step(){
      var budget=BATCH;
      while(ai<autos.length&&budget>0){if(on)maskCell(autos[ai]);else unmaskCell(autos[ai]);ai++;budget--;}
      while(qi<qz.length&&budget>0){
        var sp=qz[qi].querySelector('.cval'),inp=qz[qi].querySelector('.cin');
        if(sp)sp.style.display=on?'none':'';
        if(inp){inp.style.display=on?'':'none';inp.value='';inp.className='cin';inp.removeAttribute('data-hit');}
        qi++;budget--;
      }
      if(ai<autos.length||qi<qz.length){chkJob=raf(step);}     /* 还有 → 下一帧继续(让步) */
      else{chkJob=0;if(done)done();}
    }
    step();
  }
  function setChk(on){
    chkon=on;
    chkbtn.textContent=on?'结束检查（显示全部 auto_out）':'开始检查（遮盖所有 auto_out）';
    chkbtn.classList.toggle('off',on);
    applyChk(chkBody(),on,function(){if(!on)chkscore.textContent='';});
    if(on)resetScore();else chkscore.textContent='';     /* 计分总数靠 cquiz 数量，立即可显示 */
  }
  if(chkbtn){
    chkbtn.onclick=function(){setChk(!chkon);};
    document.addEventListener('keydown',function(ev){
      if(!chkon||ev.key!=='Enter'||!ev.target.classList||!ev.target.classList.contains('cin'))return;
      var inp=ev.target,td=inp.parentNode;
      var v=parseVal(inp.value);
      if(v===null)return;                               /* 空着回车不判 */
      var want=parseInt(td.getAttribute('data-v'),10);
      var hit=!isNaN(v)&&v===want;
      /* 揭示同列 auto_out + 整列染色 + 列头染色（重答时先清旧色再上新色） */
      var table=td.closest('table'),ci=td.cellIndex,rows=table.tBodies[0].rows,j;
      for(j=0;j<rows.length;j++){
        var cell=rows[j].cells[ci];
        if(!cell)continue;
        if(cell.classList.contains('cauto'))unmaskCell(cell);
        cell.classList.remove('okc','badc');
        cell.classList.add(hit?'okc':'badc');
      }
      var hdr=table.tHead.rows[0].cells[ci];
      if(hdr){hdr.classList.remove('okh','badh');hdr.classList.add(hit?'okh':'badh');}
      /* 计分：首答 answered++；重答(允许改答案)则把旧结果从计数里挪走 */
      var prev=inp.getAttribute('data-hit');
      if(prev===null)answered++;
      else if(prev==='1')okn--;
      else badn--;
      if(hit)okn++;else badn++;
      inp.setAttribute('data-hit',hit?'1':'0');
      inp.className='cin '+(hit?'okin':'badin');
      updScore();
    });
  }

  function boot(){     /* 首屏初始化：active=①汇总 必须主动渲一次，否则首屏空白、要切 tab 才出来 */
    render(active);
    cnt.textContent='匹配信号 '+filtered('sum').length;
  }
  boot();
})();
"""


def _write_report_xlsx(path, rep, excel):
    """导出为 Excel（给 designer 看、复制粘贴）：
      ① 真值表 sheet —— 每信号一块(输入做行、各测试做列；auto_out / 期望(进.sv) / force / RF_WRITE
         分行)，块间空行；冻结首列让信号/输入标签随右滚恒可见。
      ② 汇总 ③ 明细 (④ 可验证性，有则出) —— 扁平表 + Excel 自动筛选(autofilter) + 冻结表头，
         designer 可在 Excel 内按 owner / 信号名 / 类型直接筛。
    期望格配色与 HTML 报告一致：绿=手填且=auto_out，红=手填但≠auto_out(表达式存疑)，灰=auto_out 兜底，
    橙=负向(故意填错)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEAD_FILL = PatternFill("solid", fgColor="DCE6F1")   # 表头淡蓝
    SIG_FILL = PatternFill("solid", fgColor="C6D9F1")    # 真值表信号/测试标题行
    NEG_FILL = PatternFill("solid", fgColor="F8CBAD")    # 负向列/格 橙
    DSGN_FILL = PatternFill("solid", fgColor="C6EFCE")   # 手填且与 auto_out 一致 绿
    DIFF_FILL = PatternFill("solid", fgColor="FFC7CE")   # 手填但不一致 红
    FB_FILL = PatternFill("solid", fgColor="F2F2F2")     # auto_out 兜底 灰
    BOLD = Font(bold=True)

    wb = Workbook()

    # ───────── ① 真值表 sheet ─────────
    ws = wb.active
    ws.title = "真值表"
    r = 1
    max_test_col = 1
    for t in rep.get("tables", []):
        tests = t["tests"]
        max_test_col = max(max_test_col, 1 + len(tests))
        # 信号标题行：A=信号名(粗)，B+=测试名(负向列橙)
        c = ws.cell(r, 1, t["signal"]); c.font = BOLD; c.fill = SIG_FILL
        for ci, tc in enumerate(tests):
            cc = ws.cell(r, 2 + ci, tc["name"]); cc.font = BOLD
            cc.fill = NEG_FILL if tc["neg"] else SIG_FILL
        if not tests:
            ws.cell(r, 2, "(无测试用例)")
        r += 1
        # 表达式行
        ws.cell(r, 1, "表达式").font = BOLD
        ws.cell(r, 2, t.get("expr", ""))
        r += 1
        # 输入行（A=「字母 → 信号名[位宽]」，B+=各测试的输入取值）
        for ii, inp in enumerate(t["inputs"]):
            ltr = inp.get("letters") or ""
            rh = ("%s → %s" % (ltr, inp["label"])) if ltr else inp["label"]
            ws.cell(r, 1, rh).font = BOLD
            for ci, tc in enumerate(tests):
                ws.cell(r, 2 + ci, tc["values"][ii] if ii < len(tc["values"]) else "")
            r += 1
        # auto_out 行（表达式计算值，参考）
        c = ws.cell(r, 1, t.get("auto_label", "auto_out")); c.font = BOLD; c.fill = FB_FILL
        for ci, tc in enumerate(tests):
            cc = ws.cell(r, 2 + ci, tc.get("auto_out", tc.get("correct", "")))
            cc.fill = NEG_FILL if tc["neg"] else FB_FILL
        r += 1
        # 期望 行（.sv 断言对比值；配色同 HTML）
        ws.cell(r, 1, t.get("exp_label", "期望(out)")).font = BOLD
        for ci, tc in enumerate(tests):
            cc = ws.cell(r, 2 + ci, tc.get("expected", ""))
            if tc["neg"]:
                cc.fill = NEG_FILL
            elif tc.get("designer_filled"):
                same = tc.get("expected") == tc.get("auto_out", tc.get("correct"))
                cc.fill = DSGN_FILL if same else DIFF_FILL
            else:
                cc.fill = FB_FILL
        r += 1
        # force / RF_WRITE 驱动行
        for label, key in (("force", "force"), ("RF_WRITE", "rfwrite")):
            ws.cell(r, 1, label).font = BOLD
            for ci, tc in enumerate(tests):
                ws.cell(r, 2 + ci, tc.get(key, ""))
            r += 1
        r += 1   # 块间空行
    ws.freeze_panes = "B1"               # 冻结首列：右滚看测试列时，信号/输入标签恒可见
    ws.column_dimensions["A"].width = 40
    for ci in range(2, max_test_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # ───────── 扁平表（汇总/明细/可验证性）：autofilter + 冻结表头 ─────────
    def flat_sheet(title, cols, rows, value_of=None):
        s = wb.create_sheet(title)
        for ci, (_k, h) in enumerate(cols, 1):
            cc = s.cell(1, ci, h); cc.font = BOLD; cc.fill = HEAD_FILL
        for ri, row in enumerate(rows, 2):
            for ci, (k, _h) in enumerate(cols, 1):
                s.cell(ri, ci, value_of(row, k) if value_of else row.get(k, ""))
        ncol = len(cols)
        s.auto_filter.ref = "A1:%s%d" % (get_column_letter(ncol), max(len(rows) + 1, 1))
        s.freeze_panes = "A2"
        for ci, (k, _h) in enumerate(cols, 1):
            wide = k in ("expr", "signal", "detail", "force", "rfwrite", "unresolved", "out_net")
            s.column_dimensions[get_column_letter(ci)].width = 44 if wide else 14
        return s

    flat_sheet("汇总", SUMMARY_COLS, rep["summary"])
    flat_sheet("明细", DETAIL_COLS, rep["detail"])
    verif = rep.get("verifiability")
    if verif and verif.get("signals"):
        def _vstat(row, k):
            if k == "status_label":
                return VERIF_STATUS.get(row.get("status", ""), (row.get("status", ""), ""))[0]
            return row.get(k, "")
        flat_sheet("可验证性", VERIF_COLS, verif["signals"], value_of=_vstat)

    wb.save(path)


def _write_report_html(path, rep, excel):
    import html
    import json

    def esc(x):
        return html.escape(str(x))

    # 信号 → top_output（明细/真值表 item 自身没有 top 字段，从汇总查；汇总/可验证性各有自带字段）
    _top_by_sig = {r.get("signal"): (1 if r.get("top") else 0) for r in rep["summary"]}

    def _top_of(sig):
        return _top_by_sig.get(sig, 0)

    def flat_rows(rows, cols, kind):
        """① 汇总 / ④ 明细：返回 (表头 HTML, 行 item 列表)。每个 item =
        {h:<tr>串, t:搜索文本(小写), o:owner, n:负向0/1, s:信号名, ty:类型, tp:top0/1}，供 JS 过滤+窗口化渲染。"""
        th = "".join("<th>%s</th>" % esc(h) for _k, h in cols)
        items = []
        for r in rows:
            sname = r.get("signal", "") or ""
            ty = r.get("type", "") or ""
            if kind == "sum":
                neg = bool(r.get("n_neg"))
                text = "%s %s %s" % (sname, r.get("owner", ""), r.get("expr", ""))
                cls = "err" if r.get("error") else ("neg" if neg else "")
                tp = 1 if r.get("top") else 0
            else:
                neg = r.get("neg") == "是"
                text = "%s %s %s %s" % (sname, r.get("owner", ""),
                                        r.get("test", ""), r.get("expr", ""))
                cls = "neg" if neg else ""
                tp = _top_of(sname)
            tds = "".join("<td>%s</td>" % esc(r.get(k, "")) for k, _h in cols)
            items.append({"h": '<tr class="%s">%s</tr>' % (cls, tds),
                          "t": text.lower(), "o": r.get("owner", "") or "", "n": 1 if neg else 0,
                          "s": sname, "ty": ty, "tp": tp})
        return th, items

    def _exp_cell(tc, check, col=None):
        """期望行单元格：负向=红(错值)；手填=绿/红(与 auto_out 一致/不一致)；未填=灰(auto_out 兜底)。
        check=True 时正向格变成『可填空自测』结构(span 显示值 + 隐藏的 input，JS 切换)。
        col=本列序号(给『按值筛选』用 data-c，整列一起显隐)。"""
        dc = (' data-c="%d"' % col) if col is not None else ""
        if tc["neg"]:
            title = ' title="负向：故意填错(预期 FAIL，自检 checker)。正确(auto_out)应为 %s"' % esc(tc["correct"])
            return '<td class="neg"%s%s>%s</td>' % (title, dc, esc(tc["expected"]))
        filled = tc.get("designer_filled")
        same = tc.get("expected") == tc.get("auto_out", tc.get("correct"))
        if filled:
            cls = "dsgn" if same else "dsgndiff"
            title = (' title="designer 手填期望，与 auto_out 一致"' if same else
                     ' title="⚠ designer 手填期望与 auto_out 不一致——表达式可能与 designer 意图不符，'
                     '仿真该测试预期 FAIL(这正是 Dreg 要抓的 bug)"')
            shown = esc(tc["expected"])
        else:
            cls = "fb"
            title = ' title="期望未手填，生成 .sv 时用 auto_out 兜底(未经 designer 人工核对)"'
            shown = esc(tc["expected"])
        if not check:
            return '<td class="%s"%s%s>%s</td>' % (cls, title, dc, shown)
        # 检查模式格：data-v=auto_out 数值(回车后 JS 比对)；input 平时隐藏
        return ('<td class="cquiz %s" data-v="%d"%s%s><span class="cval">%s</span>'
                '<input class="cin" style="display:none" placeholder="?"></td>'
                % (cls, tc.get("auto_num", 0), title, dc, shown))

    def truth_blocks(tabs):
        """②真值表 / ③真值表检查 共用：输入(带位宽)做行、各测试做列；auto_out 与 期望 分两行。
        统一用『检查态』标记(cauto/cquiz)——②不开检查时与只读表外观一致，③开检查才遮盖/填空，
        故只生成一份(不再为 ③ 复制第二份 DOM)。返回 item 列表 {h,t,o,n}。"""
        items = []
        for t in tabs:
            tests = t["tests"]
            # 每个输入在各测试列中实际出现过的取值(去重、保持出现序)；>1 个的才是「变化的输入」，挂筛选下拉。
            distinct = []
            for ri in range(len(t["inputs"])):
                seen, vals = set(), []
                for tc in tests:
                    sv = str(tc["values"][ri]) if ri < len(tc["values"]) else ""
                    if sv not in seen:
                        seen.add(sv); vals.append(sv)
                distinct.append(vals)
            has_filter = len(tests) >= 2 and any(len(v) > 1 for v in distinct)
            hdr = ['<th class="rowhdr">信号\\测试</th>']
            for j, tc in enumerate(tests):
                hdr.append('<th class="%s" data-c="%d">%s</th>'
                           % ("negh" if tc["neg"] else "", j, esc(tc["name"])))
            body = []
            for ri, inp in enumerate(t["inputs"]):
                ltr = inp.get("letters") or ""               # 表达式变量(A/B/C…) → 物理信号
                rh = ("%s → %s" % (ltr, inp["label"])) if ltr else inp["label"]
                # 第0列：变化的输入挂「按值筛选」输入框（像 Excel 一样手填值，边打边筛，留匹配列）
                sel = ""
                if has_filter and len(distinct[ri]) > 1:
                    sel = ('<input class="ttf" type="text" data-ri="%d" placeholder="筛值"'
                           ' title="手填该输入的值(如 0 / 1)，只保留该值匹配的测试列；留空=全部">' % ri)
                cells = ['<th class="rowhdr">%s%s</th>' % (esc(rh), sel)]
                for j, tc in enumerate(tests):
                    v = tc["values"][ri] if ri < len(tc["values"]) else ""
                    cells.append('<td class="%s" data-c="%d">%s</td>'
                                 % ("neg" if tc["neg"] else "", j, esc(v)))
                body.append('<tr class="inrow" data-ri="%d">%s</tr>' % (ri, "".join(cells)))
            # ── auto_out 行（表达式计算值，只读参考）。检查模式下负向列不参与遮盖(不是自测对象) ──
            auto_cells = ['<th class="rowhdr" title="程序按表达式算出的输出值(参考)。'
                          '用它当期望验证表达式有自证嫌疑，.sv 断言对比的是下面的期望">%s</th>'
                          % esc(t.get("auto_label", "auto_out"))]
            for j, tc in enumerate(tests):
                cls = "neg" if tc["neg"] else "cauto"
                auto_cells.append('<td class="%s" data-v="%d" data-c="%d">%s</td>'
                                  % (cls, tc.get("auto_num", 0), j,
                                     esc(tc.get("auto_out", tc.get("correct", "")))))
            body.append('<tr class="autorow">%s</tr>' % "".join(auto_cells))
            # ── 期望 行（designer 手填 > auto_out 兜底 > 负向错值；.sv 断言用这一行） ──
            exp_cells = ['<th class="rowhdr" title="designer 手填的期望，.sv 断言用它对比；'
                         '未填的列用 auto_out 兜底(灰)。绿=手填且与 auto_out 一致；红=手填但不一致">%s</th>'
                         % esc(t.get("exp_label", "期望(out)"))]
            for j, tc in enumerate(tests):
                exp_cells.append(_exp_cell(tc, True, j))
            body.append('<tr class="exprow">%s</tr>' % "".join(exp_cells))
            for label, key in (("force", "force"), ("RF_WRITE", "rfwrite")):
                cells = ['<th class="rowhdr">%s</th>' % label]
                for j, tc in enumerate(tests):
                    cells.append('<td class="drv %s" data-c="%d">%s</td>'
                                 % ("neg" if tc["neg"] else "", j, esc(tc.get(key, ""))))
                body.append("<tr>%s</tr>" % "".join(cells))
            neg_block = any(tc["neg"] for tc in tests)
            text = "%s %s %s" % (t["signal"], t.get("owner", ""), t["expr"])
            # cone 展开链：本行 + 逐层代入的上游行（Excel 原式 = 代入信号名的等价形式），真值表上方
            chain = t.get("chain") or []
            chain_html = ""
            if len(chain) >= 2:
                marks = "①②③④⑤⑥⑦⑧⑨"
                cl = []
                for ci, c in enumerate(chain):
                    mk = marks[ci] if ci < len(marks) else "(%d)" % (ci + 1)
                    head = "%s %s" % (mk, c["out"])
                    cl.append("%s = %s" % (esc(head), esc(c["expr"])))
                    cl.append("%s = %s" % (" " * len(head), esc(c["subst"])))
                chain_html = ('<pre class="chain">展开链（输入引用内部信号/上游计算网，已展开上游；'
                              '①=本行，往下逐层代入）:\n%s</pre>' % "\n".join(cl))
            # 标题序号: logic 显示 "R<序号>"; mux 的 R 已是 "mux<N>"，不再加 R 前缀
            rid = str(t["R"])
            rlabel = rid if rid.startswith("mux") else ("R" + rid)
            ttfbar = (('<span class="ttfbar">显示 <span class="ttfcount">%d/%d</span> 列'
                       '<button class="ttfclr" type="button">清除</button></span>')
                      % (len(tests), len(tests))) if has_filter else ""
            # RTL 补充逻辑(Excel 真表缺、手工补)：块内显眼红条 banner，提示本张真值表非纯 Excel 推导
            supp = t.get("supplement") or ""
            supp_html = ('<div class="suppbar">⚠ %s</div>' % esc(supp)) if supp else ""
            h = ('<div class="ttblock">'
                 '<h3>%s　<code>%s</code>　<span class="ex">%s</span>%s</h3>%s%s'
                 '<table class="tt"><thead><tr>%s</tr></thead><tbody>%s</tbody></table></div>'
                 % (esc(rlabel), esc(t["signal"]), esc(t["expr"]), ttfbar, supp_html, chain_html,
                    "".join(hdr), "".join(body)))
            items.append({"h": h, "t": text.lower(), "o": t.get("owner", "") or "",
                          "n": 1 if neg_block else 0,
                          "s": t["signal"], "ty": t.get("type", "") or "", "tp": _top_of(t["signal"])})
        return items

    def verif_rows(verif):
        """⑤ 可验证性：返回 (顶部健康度概览 HTML, 表头 HTML, 行 item 列表)。"""
        if not verif or not verif.get("signals"):
            return '<p class="empty">（无数据）</p>', "", []
        c = verif["counts"]
        head = ('<p class="vsum">'
                '<b class="vstat ok">✅ 可验证 %d</b>'
                '<b class="vstat warn">⚠ 存疑(按名 force) %d</b>'
                '<b class="vstat bad">✗ 不可验证 %d</b>'
                '<b class="vstat bad">✗ 表达式错误 %d</b>'
                '<b class="vstat bad">✗ 规格冲突(待 designer 核对) %d</b></p>'
                '<p class="ex">⚠/✗ 的信号最可能在 elaboration 阶段 CUVUNF 失败：'
                '"存疑"=表里查无、按信号名直接 force；"不可验证"=输入未能解析到 force/RF_WRITE；'
                '"规格冲突"=mux 同选择值却选不同源(撞车行号/两源/owner 见下表)，需 designer 核对改表'
                '——可能是真矛盾(改数据源)，也可能是两个 mux 撞了同一输出名(『一个控制管多个 mux』合法，改输出名即可)。</p>'
                % (c.get("clean", 0), c.get("wire-fallback", 0),
                   c.get("unresolved", 0), c.get("parse-err", 0),
                   c.get("spec-collision", 0)))
        th = "".join("<th>%s</th>" % esc(h) for _k, h in VERIF_COLS)
        items = []
        for r in verif["signals"]:
            lbl, lvl = VERIF_STATUS.get(r.get("status", ""), (r.get("status", ""), ""))
            text = "%s %s %s" % (r.get("signal", ""), r.get("owner", ""), r.get("detail", ""))
            cells = []
            for k, _h in VERIF_COLS:
                if k == "status_label":
                    cells.append('<td><span class="vstat %s">%s</span></td>' % (lvl, esc(lbl)))
                else:
                    cells.append("<td>%s</td>" % esc(r.get(k, "")))
            items.append({"h": '<tr class="vrow %s">%s</tr>' % (lvl, "".join(cells)),
                          "t": text.lower(), "o": r.get("owner", "") or "", "n": 0,
                          "s": r.get("signal", "") or "", "ty": r.get("type", "") or "",
                          "tp": 1 if r.get("top") else 0})
        return head, th, items

    def js_blob(el_id, items):
        # JSON 嵌进 <script type=application/json>：把 "</" 转义成 "<\\/" 防止意外闭合 script 标签
        # （JSON.parse 时 \\/ 还原成 /）。ensure_ascii=False 让中文可读、体积更小。
        payload = json.dumps(items, ensure_ascii=False).replace("</", "<\\/")
        return '<script type="application/json" id="%s">%s</script>' % (el_id, payload)

    owners = sorted({(r.get("owner") or "") for r in rep["summary"] if r.get("owner")})
    # owner 多选下拉（复选浮层）：勾多个=OR，任一命中即显示；不勾=全部。
    # owner 列(P/L/AE)留空的信号给一个「（无 owner）×N」复选项（哨兵值 __no_owner__），
    # 仅当真有空 owner 信号时才出现（都填好就不打扰）。
    n_no_owner = sum(1 for r in rep["summary"] if not r.get("owner"))
    _ck = '<label><input type="checkbox" class="ownerck" value="%s">%s</label>'
    no_owner_ck = (_ck % ("__no_owner__", "（无 owner） ×%d" % n_no_owner)) if n_no_owner else ""
    owner_opts = (no_owner_ck + "".join(_ck % (esc(o), esc(o)) for o in owners)
                  + '<label class="mselclr"><a href="#" id="ownerclr">清除选择</a></label>')
    # 信号名下拉：列全部信号(含 mux 组)，按名排序；选项 value=原始信号名(与 item.s 逐字匹配)
    sig_names = sorted({(r.get("signal") or "") for r in rep["summary"] if r.get("signal")})
    sig_opts = ('<option value="">全部信号</option>' + "".join(
        '<option value="%s">%s</option>' % (esc(s), esc(s)) for s in sig_names))
    # 类型下拉：to_logic / to_mux / ls / mux 等(去重排序)
    types = sorted({(r.get("type") or "") for r in rep["summary"] if r.get("type")})
    typ_opts = ('<option value="">全部类型</option>' + "".join(
        '<option value="%s">%s</option>' % (esc(t), esc(t)) for t in types))
    top_opts = ('<option value="">全部 top</option>'
                '<option value="1">仅 top_output</option>'
                '<option value="0">仅非 top</option>')
    n_sig = len(rep["summary"])
    n_tc = len(rep["detail"])
    n_neg = sum(1 for r in rep["detail"] if r.get("neg") == "是")

    # 期望来源统计：手填(designer 审过) vs auto_out 兜底(有自证嫌疑)
    n_designer = sum(1 for t in rep.get("tables", []) for tc in t["tests"]
                     if tc.get("designer_filled"))
    n_pos_tc = sum(1 for t in rep.get("tables", []) for tc in t["tests"] if not tc["neg"])

    # 各 tab 的数据收集成 item 列表（HTML 串 + 元数据），下面嵌成 JSON 给前端窗口化渲染。
    sum_th, sum_items = flat_rows(rep["summary"], SUMMARY_COLS, "sum")
    det_th, det_items = flat_rows(rep["detail"], DETAIL_COLS, "det")
    tt_items = truth_blocks(rep.get("tables", []))            # ②③ 共用这一份
    ver_head, ver_th, ver_items = verif_rows(rep.get("verifiability"))

    # 真值表检查 tab 顶部说明条（静态，不进窗口化）
    chkbar = (
        '<div class="chkbar">'
        '<button id="chkbtn">开始检查（遮盖所有 auto_out）</button>'
        '<span id="chkscore"></span>'
        '<span class="chkhint">这是 designer 的防自证自测：auto_out 是表达式算出来的——用它当期望验证表达式'
        '等于自己证明自己。点「开始检查」后 auto_out 全部遮住(显示 ?)、期望变成空格：只看上面的输入，'
        '自己算输出填进去按回车 → 立即揭晓 auto_out 并比对。'
        '<b style="color:#1a7f37">绿=一致</b>；<b style="color:#c00">红=不一致</b>'
        '（要么你算错了，要么表达式有 bug——后者正是 Dreg 要抓的）。'
        '答完可以改值再回车重判（计分随之修正）。检查按「当前页」进行（翻页即换一组）。'
        '检查结果只在本页面，不回写 .sv；负向列(_NEG)是故意填错的自检用例，不参与检查。</span>'
        '</div>'
    )

    # 每个 section 只留『分页条 + 空挂载点』，真正内容由 JS 按当前窗注入；行表的表头静态保留。
    sum_sec = ('<div class="pagenav" id="sum-nav"></div>'
               '<table class="sum"><thead><tr>%s</tr></thead>'
               '<tbody class="pgbody" id="sum-body"></tbody></table>' % sum_th)
    tt_sec = '<div class="pagenav" id="tt-nav"></div><div class="pgbody ttwrap" id="tt-body"></div>'
    chk_sec = (chkbar + '<div class="pagenav" id="chk-nav"></div>'
               '<div class="pgbody ttwrap" id="chk-body"></div>')
    det_sec = ('<div class="pagenav" id="det-nav"></div>'
               '<table class="full"><thead><tr>%s</tr></thead>'
               '<tbody class="pgbody" id="det-body"></tbody></table>' % det_th)
    ver_sec = (ver_head + '<div class="pagenav" id="ver-nav"></div>'
               '<table class="full"><thead><tr>%s</tr></thead>'
               '<tbody class="pgbody" id="ver-body"></tbody></table>' % ver_th)

    # 仅对 body 模板做 % 替换；CSS/JS 含字面 % 与 {}，单独拼接(不参与格式化)。
    body = (
        '<header>'
        '<h1>Dreg 测试用例报告</h1>'
        '<p class="sum">源 Excel: <code>%s</code>　信号 %d 个　用例 %d 条（其中负向 %d 条）　'
        '正向期望: designer 手填 %d / auto_out 兜底 %d　'
        '负向(红/_NEG)=故意填错期望, 预期应 FAIL。</p>'
        '<p class="sum">auto_out=程序按表达式算的值(参考)；期望=designer 手填、.sv 断言用它对比'
        '（未填→auto_out 兜底）。用「③ 真值表检查」可自测期望而不被 auto_out 影响。'
        '<span class="ex">　大表分页显示：用搜索/owner 过滤后翻页，DOM 只渲当前页，再多组也不卡。</span></p>'
        '<div class="toolbar">'
        '<input type="text" id="q" placeholder="搜索 信号名 / owner / 表达式…">'
        '<div class="msel" id="ownerbox">'
        '<button type="button" class="mselbtn" id="ownerbtn">全部 owner ▾</button>'
        '<div class="mselpop" id="ownerpop">%s</div></div>'
        '<select id="sig" title="按信号名筛选">%s</select>'
        '<select id="typ" title="按类型(to_logic/to_mux/ls/mux)筛选">%s</select>'
        '<select id="topf" title="按 top_output 筛选">%s</select>'
        '<label><input type="checkbox" id="negonly"> 只看负向</label>'
        '<span id="count"></span>'
        '</div>'
        '<div class="tabs">'
        '<button class="tabbtn active" data-tab="sum">① 汇总</button>'
        '<button class="tabbtn" data-tab="tt">② 真值表</button>'
        '<button class="tabbtn" data-tab="chk">③ 真值表检查</button>'
        '<button class="tabbtn" data-tab="det">④ 明细</button>'
        '<button class="tabbtn" data-tab="ver">⑤ 可验证性</button>'
        '</div></header>'
        '<main>'
        '<section id="sum" class="tab active">%s</section>'
        '<section id="tt" class="tab">%s</section>'
        '<section id="chk" class="tab">%s</section>'
        '<section id="det" class="tab">%s</section>'
        '<section id="ver" class="tab">%s</section>'
        '</main>'
    ) % (
        esc(os.path.basename(excel)), n_sig, n_tc, n_neg, n_designer, n_pos_tc - n_designer,
        owner_opts, sig_opts, typ_opts, top_opts, sum_sec, tt_sec, chk_sec, det_sec, ver_sec,
    )
    # 各 tab 数据：JSON 嵌在页面里，由前端 JS 解析后窗口化渲染（②③ 共用 tt-data）。
    blobs = (js_blob("sum-data", sum_items) + js_blob("tt-data", tt_items)
             + js_blob("det-data", det_items) + js_blob("ver-data", ver_items))
    doc = ('<!doctype html><html lang="zh"><head><meta charset="utf-8">'
           '<title>Dreg 测试用例报告</title><style>' + _REPORT_CSS + '</style></head><body>'
           + body + blobs + '<script>' + _REPORT_JS + '</script></body></html>')
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)


def main(argv=None):
    args = build_argparser().parse_args(argv)
    if not os.path.isfile(args.excel):
        sys.exit("找不到 Excel: %s" % args.excel)

    # --match-fortest: 复刻 for_test 覆盖面——不跳过任何信号(含内部 top_output=0 与含不可驱动
    # wire输入的)，与 VBA for_test 一样照单全收。注意产物可能 CUVUNF 跑不起来(预期内)。
    if args.match_fortest:
        args.include_internal = True
        args.include_risky = True

    # mux 范围旗标（2026-06-03 第九轮）
    if args.no_mux and args.mux_only:
        sys.exit("--no-mux 与 --mux-only 互斥")
    arg_types = _split(args.type)
    if args.mux_only:
        arg_types = {"mux"}

    opts = generator.GenOptions(
        owners=_split(args.owner),
        signals=_split(args.signals),
        signal_regex=args.regex,
        exclude=_split(args.exclude),
        exclude_regex=args.exclude_regex,
        types=arg_types,
        top_output_only=not args.include_internal,   # 默认只生成 top_output=1（可验证输出）
        gen_mux=not args.no_mux,                     # 默认 logic+mux 都生成（用户拍板）
        mode=args.mode,
        max_tests=args.max_tests,
        exhaustive=args.exhaustive,
        logic_mode=args.logic_mode,
        mux_mode=args.mux_mode,
        neg_signals=_split(args.neg_signals),
        neg_all=args.neg_all,
        neg_mode=args.neg_mode,
        neg_which=args.neg_which,
        neg_value=args.neg_value,
        force_overrides=_split(args.force_signals),
        rfwrite_overrides=_split(args.rfwrite_signals),
        default_kind=args.default_kind,
        wire_fallback=not args.no_wire_fallback,
        comments=args.comments,
        include_risky=args.include_risky,
        probe_prefixes=_parse_probe_prefixes(args.probe_prefix, args.probe_prefix_file),
        owner_in_msg=args.owner_in_msg,
        sv_summary=args.sv_summary,
        cascade_mode=args.cascade_mode,
        logic_cascade=args.logic_cascade, mux_cascade=args.mux_cascade,
        append_to_logic=args.append_to_logic,
        append_to_mux=args.append_to_mux,
        suffix_override=dict({n: False for n in (_split(args.no_suffix_signals) or [])},
                             **{n: True for n in (_split(args.suffix_signals) or [])}),
    )

    if args.out_file:
        return _run_with_outfile(args, opts)
    return _dispatch(args, opts)


def _run_with_outfile(args, opts):
    """把屏幕报告直接写盘（UTF-8 带 BOM），绕开 PowerShell `>` 的 GBK 重定向乱码。"""
    path = args.out_file
    saved = sys.stdout
    try:
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            sys.stdout = f
            rc = _dispatch(args, opts)
    finally:
        sys.stdout = saved
    print("报告已写入: %s（UTF-8 带 BOM，可直接用记事本/Excel 打开）" % path)
    return rc


def _dispatch(args, opts):
    """list/account/diagnose/report/生成 的分派；所有屏幕输出走当前 sys.stdout。"""
    print("装载 Excel: %s ..." % args.excel)
    wb = excel_model.load_workbook(args.excel)
    print("  sheets: %s" % wb.sheet_names)

    # Topout-rooted 路径（2026-06-23 新模型）：要验信号=Topout B 列，cone 展到源、断言贴顶层真名。
    # 与旧 logic-rooted 路径完全并存：默认仍走旧路径；--topout 才切到新路径（旧 CLI 行为逐字节不变）。
    if args.topout:
        return cmd_topout(args, wb, opts)

    if args.list:
        cmd_list(wb, opts)
        return 0

    if args.account:
        cmd_account(wb, opts)
        return 0

    if args.diagnose:
        cmd_diagnose(wb, opts)
        return 0

    # 给人看的报告（可与 .sv 生成并存；只传 --report 不传 --out 则只出报告）
    if args.report:
        rep = generator.report(wb, opts)
        written = write_report(args.report, rep, args.excel)
        print("测试用例报告已写出: %s" % "  ".join(written))
        print("  信号 %d 个，用例 %d 条（负向 %d 条）"
              % (len(rep["summary"]), len(rep["detail"]),
                 sum(1 for r in rep["detail"] if r.get("neg") == "是")))
        vc = rep.get("verifiability", {}).get("counts", {})
        if vc:
            print("  可验证性：可验证 %d / 存疑(按名force) %d / 不可验证 %d / 表达式错误 %d / 规格冲突 %d"
                  % (vc.get("clean", 0), vc.get("wire-fallback", 0),
                     vc.get("unresolved", 0), vc.get("parse-err", 0),
                     vc.get("spec-collision", 0)))
        if args.out is None and not args.export_claims:
            return 0   # 只要报告

    # 只导 claims(无 --out)：build 一次拿 res 即可，不出 .sv
    if args.out is None and args.export_claims:
        res = generator.build(wb, opts)
        _export_claims(res, args.export_claims, args.excel)
        return 0

    out = args.out or "wr_rf_tc.sv"

    # 负向用例单独出文件：分两次生成
    if args.neg_file == "separate" and (opts.neg_all or opts.neg_signals):
        # 正常文件：不含负向
        pos_opts = _copy_opts(opts, neg_all=False, neg_signals=None)
        pos_res = generator.build(wb, pos_opts)
        _write(out, generator.render(pos_res, comments=opts.comments))
        _report(pos_res, out)
        _export_claims(pos_res, args.export_claims, args.excel)   # 设了 --export-claims 才导
        # 负向文件：真·仅负向(每信号只留负向向量，无负向的信号不出现)；
        # 汇总命名块加 _neg 后缀 → 与主文件贴进同一作用域也不重名
        neg_path = _neg_path(out)
        neg_res = generator.build(wb, _copy_opts(opts, negative_vectors_only=True))
        _write(neg_path, generator.render(neg_res, comments=opts.comments, block_suffix="_neg"))
        print("负向用例已单独写入: %s" % neg_path)
        return 0

    res = generator.build(wb, opts)
    text = generator.render(res, comments=opts.comments)
    _write(out, text)
    _report(res, out)
    _export_claims(res, args.export_claims, args.excel)          # 设了 --export-claims 才导
    return 0


def _export_claims(res, path, excel, naming_model=None):
    """把 build 结果里的 claim 清单写成 JSON——红区 scan_rtl 校验器的输入契约。path 为空则不导。

    naming_model(C2,2026-06-23)：探针网名的命名模型，随 claims 进红区让 binder 知道名字怎么来的——
      'logic-rooted'(默认,旧路径,探针名按命名约定猜) / 'topout'(新路径,Topout 顶层真名,近零前缀)。
    """
    if not path:
        return
    import json
    claims = res.get("claims", [])
    nm = naming_model or generator.NAMING_MODEL_LOGIC
    note_by_model = {
        generator.NAMING_MODEL_LOGIC: (
            "命名模型=logic-rooted：探针名按命名约定+尾缀【猜】出(_to_logic/_to_mux/_ls/前缀)，可能假绿——"
            "红区 binder 须逐探针核对真名(R41/R42)。校验器契约：probe 须为真·RTL 输出网(不是某 assign 的"
            " RHS 输入网)、force 网须真实存在；self_ref_suffixes 非空且 == 探针后缀 = 自引用假绿(应为空)。"),
        generator.NAMING_MODEL_TOPOUT: (
            "命名模型=topout：探针名是 Topout 顶层真名(无路由后缀)，cone 已展到源寄存器→近零前缀——"
            "红区 binder 顶层口直接核存在性即可。校验器契约同 logic-rooted(probe 须真·输出网、force 须存在、"
            "self_ref_suffixes 应空)，但顶层真名理应直接命中、需前缀的属意外埋件。"),
    }
    payload = {
        "source_excel": os.path.basename(excel),
        "schema_version": generator.CLAIM_SCHEMA_VERSION,
        "naming_model": nm,
        "n_claims": len(claims),
        "schema": list(generator._CLAIM_KEYS),
        "note": note_by_model.get(nm, note_by_model[generator.NAMING_MODEL_LOGIC]),
        "claims": claims,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    n_probe = sum(1 for c in claims if c.get("kind") == "probe")
    n_force = sum(1 for c in claims if c.get("kind") == "force")
    n_rfw = sum(1 for c in claims if c.get("kind") == "rfwrite")
    print("claim 清单已导出: %s（%d 条：探针 %d / force %d / rfwrite %d）"
          % (path, len(claims), n_probe, n_force, n_rfw))


def _copy_opts(opts, **overrides):
    import copy
    o = copy.copy(opts)
    for k, v in overrides.items():
        setattr(o, k, v)
    return o


def _neg_path(out):
    root, ext = os.path.splitext(out)
    return root + "_neg" + (ext or ".sv")


def _header(excel, opts, kind):
    return {
        "源Excel": os.path.basename(excel),
        "类型": kind,
        "向量模式": opts.mode,
        "筛选owner": sorted(opts.owners) if opts.owners else "全部",
    }


def _write(path, text):
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)


def _report(res, out):
    s = res["summary"]
    print("已写出: %s" % out)
    mux_part = ""
    if s.get("n_mux_groups"):
        mux_part = ("（logic %d + mux %d，mux 组共 %d 个）"
                    % (s["n_generated"] - s.get("n_mux_generated", 0),
                       s.get("n_mux_generated", 0), s["n_mux_groups"]))
    print("  选中信号: %d / logic总行 %d；生成块: %d%s；跳过: %d；向量: %d（负向 %d）"
          % (s["n_selected"], s["n_logic_rows"], s["n_generated"], mux_part,
             s.get("n_skipped", 0), s["n_vectors"], s["n_negative"]))
    if s["n_parse_errors"]:
        print("  ⚠ 表达式解析失败 %d 个:" % s["n_parse_errors"])
        for name, aid, msg in res["errors"]:
            print("    - [R=%s] %s: %s" % (aid, name, msg))
    if s.get("n_dup_labels"):
        print("  ⛔ 发现 %d 处重复 assert 标号（同一作用域内重复=非法 SV，会 elaboration 失败）:"
              % s["n_dup_labels"])
        for lbl, sig1, sig2 in res.get("dup_labels", [])[:20]:
            print("    - assert_%s: 同时来自 %s 与 %s" % (lbl, sig1, sig2))
        print("    多因两信号共用同一序号；请核对 logic R 列 / mux 页 N 列唯一性，或改掉自定义测试名。")
    if s.get("n_skipped"):
        print("  ↷ 跳过 %d 个含'不可驱动输入'的信号（force 不存在的 net 会 elaboration 失败；VBA 也跳过这类）:"
              % s["n_skipped"])
        for name, aid, risky in res.get("skipped", [])[:30]:
            rs = ", ".join("%s=%s(%s)" % (l, b, why) for (l, b, why) in risky)
            print("    - [R=%s] %s ← %s" % (aid, name, rs))
        if len(res.get("skipped", [])) > 30:
            print("    ...(共 %d 个)" % len(res["skipped"]))
        print("    如确认这些 net 可驱动，用 --include-risky 强制生成，或 --force-signals/--rfwrite-signals 指定。")
    if s.get("n_mux_warnings"):
        # top_out=0 的 mux 输出：已照常生成（裸名探针），仅提示——若仿真 CUVUNF 再配前缀
        print("  ⚠ %d 个 mux 组用裸名探针生成（输出 top_out=0，非芯片顶层输出）："
              % s["n_mux_warnings"])
        for name, aid, why in res.get("mux_warnings", [])[:30]:
            print("    - [R=%s] %s" % (aid, name))
        if len(res.get("mux_warnings", [])) > 30:
            print("    ...(共 %d 个)" % len(res["mux_warnings"]))
        print("    这些已经生成进 .sv；若仿真 elaboration 报 CUVUNF 说明它们埋在子模块，"
              "跑 scan_rtl 配 --probe-prefix-file 后重生成即可。")
    if s.get("n_regmap_warnings"):
        # regmap 同名重复定义：已按【首个】采纳(与表 VLOOKUP/for_test 一致)，但首个未必是设计意图
        print("  ⚠ %d 个信号用到 regmap 同名重复字段（已按首个采纳；如不符请核对 regmap 或由 SE 消重）："
              % s["n_regmap_warnings"])
        for name, aid, why in res.get("regmap_warnings", [])[:30]:
            print("    - [R=%s] %s ← %s" % (aid, name, why))
        if len(res.get("regmap_warnings", [])) > 30:
            print("    ...(共 %d 个)" % len(res["regmap_warnings"]))
    if s.get("n_supplement"):
        # RTL 补充逻辑：Excel 真表缺此级(如 ECO 加的 mux/iddq)、由人工补一条等价 logic 式扫真值表。
        # 偏离纯 Excel 推导 → 显式列出供 SE review（块顶也已有 // ⚠）。
        print("  ⚠ %d 个信号用了【RTL 补充逻辑】（Excel 真表缺、手工补的等价式；请 SE 核对）："
              % s["n_supplement"])
        for name, aid, why in res.get("supplement_warnings", [])[:30]:
            print("    - [R=%s] %s ← %s" % (aid, name, why))
        if len(res.get("supplement_warnings", [])) > 30:
            print("    ...(共 %d 个)" % len(res["supplement_warnings"]))
    if s.get("n_selfaudit_warnings"):
        # 生成期"自检闸门"命中(假绿可疑)：探针读回自身输入 / K列漏标位宽 / RW写值截断。
        # 这些是"仿真可能静默通过但其实没真验"的隐患，块顶也已有 // ⚠，务必逐条核对。
        print("  ⚠ %d 处生成期自检闸门命中（假绿可疑——探针读回自身输入 / K列漏标位宽 / RW写值截断）："
              % s["n_selfaudit_warnings"])
        for name, aid, why in res.get("selfaudit_warnings", [])[:30]:
            print("    - [R=%s] %s ← %s" % (aid, name, why))
        if len(res.get("selfaudit_warnings", [])) > 30:
            print("    ...(共 %d 个)" % len(res["selfaudit_warnings"]))
    if s.get("n_filtered_internal"):
        # 唯一'默认静默减少'的一类——把名字也亮出来，别让它无声无息
        fi = res.get("filtered_internal", [])
        print("  ℹ %d 个 logic 内部节点(top_output=0)默认未生成（不是 RTL 端口、ENV_RF 探不到）："
              % s["n_filtered_internal"])
        for sig in fi[:30]:
            print("    - [R=%s] %s" % (sig.assert_id, sig.out_name))
        if len(fi) > 30:
            print("    ...(共 %d 个)" % len(fi))
        print("    要把它们也纳入用 --include-internal；要逐个核对全部信号去向用 --account。")


if __name__ == "__main__":
    sys.exit(main())
