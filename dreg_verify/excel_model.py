# -*- coding: utf-8 -*-
"""
excel_model.py — 读取 Dreg 核心 Excel 的 logic / regmap / total_memory_map 三页，
构建生成器所需的内存数据结构。

约定（来自对真实 Excel 的结构分析，表头多在第 2 行）：
  logic 页：A..J 列 = 输入信号（变量字母 = 列字母），K=输出名，L=表达式，M=类型(suffix),
            N=top_output, O=Notes, P=owner, R=序号(assert id)
  regmap 页：D=Reg_Name, F=Reg Type, G=Signal_Name, I=default, J..Y=bit15..bit0, Z=suffix, AE=owner
  total_memory_map 页：无表头、寄存器块结构；字段行 A=字段名 B=Bit Field(8 或 15:7)
            C=复位 D=DIG TOP PIN(Y/N) E=功能 F=该字段所属寄存器地址(hex) H=类型(RO/RW)
"""

import re
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:  # pragma: no cover
    raise SystemExit("缺少 openpyxl，请先运行:  pip install openpyxl")


# ───────────────────────────── 数据结构 ─────────────────────────────
class LogicSignal:
    """logic 页的一行：一个被验证输出信号。"""

    def __init__(self, row, out_name, out_width, expr, suffix, top_output,
                 notes, owner, assert_id, inputs):
        self.row = row                  # Excel 行号（便于回查）
        self.out_name = out_name        # K 列原文（含位宽如 [2:0]）
        self.out_base = _strip_width(out_name)[0]
        self.out_width = out_width       # 解析出的输出位宽
        self.expr = expr                 # L 列表达式
        self.suffix = suffix             # M 列类型
        self.top_output = top_output     # N 列
        self.notes = notes               # O 列
        self.owner = owner               # P 列
        self.assert_id = assert_id       # R 列序号
        # inputs: dict 字母 -> {'raw','base','width','msb','lsb'}
        self.inputs = inputs
        # 该内部信号在 Excel 下游被怎样引用 → 决定 RTL 探针网名【尾缀】（尾缀随 Excel、不写死）：
        #   被下游 logic 行以 <名>_to_logic 引用 → "_to_logic"（read_logic 回填）
        #   被 mux 页(数据/控制)以 <名>_to_mux 引用 → "_to_mux"（load_workbook 回填）
        #   未被引用(顶层输出/reserve) → ""（用基名）
        self.ref_suffix = ""
        # 尾缀总开关（2026-06-11，Resolver 据 GenOptions.append_to_logic 回填）：False=即便有
        # ref_suffix 也不补、直接探基名网（某些设计 <名>_to_xxx 撞了真实输入网时用）。ref_suffix 保持
        # 不变(=客观事实)，只由本开关 AND 决定是否真正追加。默认 True=随 Excel 补尾缀（旧行为）。
        self._append_to_logic = True
        # level_shift 页给出的【顶层电平移位输出口】真名（2026-06-12 Hi1108 Pilot）：新表把 logic 输出
        # (M=to_dft、top_output=0)的可探顶层口拆到 level_shift 页声明(d_en_refbuf → d_en_refbuf_ls，
        # lpbt_dig_top.v output)。非空时 rtl_base 直接返回它，压过 M 列/ref_suffix/加尾缀开关——
        # 因为这就是 RTL 顶层口真名。None=该信号不过 level_shift 或无此页(旧表 M=ls 直给后缀，不受影响)。
        self._ls_name = None
        self._ls_is_top = False        # level_shift 页 E 列 top_out=1：顶层网，探针绝不套层级前缀

    @property
    def is_top(self):
        """top_output==1：RTL/ENV_RF 顶层可见输出。"""
        return str(self.top_output).strip() in ("1", "1.0", "True", "true")

    @property
    def to_logic_ref(self):
        """兼容旧名：是否被下游以 _to_logic 引用（= ref_suffix == "_to_logic"）。"""
        return self.ref_suffix == "_to_logic"

    @property
    def rtl_base(self):
        """RTL 真实网名(去位宽)。level_shift 页给的顶层 _ls 口最优先(_ls_name)；否则 ls 行带 _ls 后缀；
        被下游引用的内部信号带其 Excel 引用尾缀(_to_logic / _to_mux)。_append_to_logic 关时不补尾缀、
        直接探基名；_ls 不受此开关影响。"""
        if getattr(self, "_ls_name", None):
            return self._ls_name
        rs = self.ref_suffix if getattr(self, "_append_to_logic", True) else ""
        return rtl_net_name(self.out_base, self.suffix, is_top=self.is_top, ref_suffix=rs)

    @property
    def rtl_name(self):
        """RTL 真实网名(含位宽切片)。assert 探针 LHS 用它，不能直接用 K 列原文。"""
        return self.rtl_base + self.out_name[len(self.out_base):]

    @property
    def rtl_base_full(self):
        """rtl_base 但【无视 _append_to_logic 开关】——总是按 Excel ref_suffix 补尾缀。
        scan_rtl 导出/用户手配的探针前缀 key 往往是这个【全名】；开关关时 rtl_base 退成裸名，
        前缀查找若只认 rtl_base 就和 _to_logic key 对不上、把用户配的前缀静默丢掉
        (2026-06-11 Hi1108 rxiq_phase_ctrl CUVUNF 实证)。前缀查找两个名都试 → 不再失配。"""
        if getattr(self, "_ls_name", None):
            return self._ls_name
        return rtl_net_name(self.out_base, self.suffix, is_top=self.is_top,
                            ref_suffix=self.ref_suffix)

    @property
    def rtl_name_full(self):
        """rtl_name 的全名版(含位宽切片，无视开关补尾缀)——仅供探针前缀 key 匹配，不进断言。"""
        return self.rtl_base_full + self.out_name[len(self.out_base):]

    def __repr__(self):
        return "LogicSignal(R=%s, %s, expr=%r)" % (self.assert_id, self.out_name, self.expr)


class RegmapEntry:
    def __init__(self, signal, reg_name, reg_type, default, bit_lsb, bit_msb, owner,
                 address=None):
        self.signal = signal
        self.reg_name = reg_name
        self.reg_type = reg_type        # 原文（可能 RW/RO/R/W）
        self.default = default
        self.bit_lsb = bit_lsb
        self.bit_msb = bit_msb
        self.owner = owner
        self.address = address          # regmap H 列地址（'d13'→13），tmm 缺失时的兜底


class TmmField:
    def __init__(self, name, bit_msb, bit_lsb, address, reg_type, dig_top_pin, reg_name,
                 reg_type_raw=""):
        self.name = name
        self.bit_msb = bit_msb
        self.bit_lsb = bit_lsb
        self.address = address          # int
        self.reg_type = reg_type        # 归一化后 'RO'/'RW'/None
        self.reg_type_raw = reg_type_raw  # H 列原文（诊断用）
        self.dig_top_pin = dig_top_pin  # 'Y'/'N'/None
        self.reg_name = reg_name        # 所属寄存器名（B 列里寄存器行给出）


class MuxCase:
    """mux 页的一行：一个 case 值 → 数据输入 的映射。"""

    def __init__(self, row, case_raw, input_raw, input_base, input_width,
                 input_msb, input_lsb):
        self.row = row                  # Excel 行号（便于回查）
        self.case_raw = case_raw        # F 列原文（如 3'b010 / 4'b000x，x=don't care 位）
        self.input_raw = input_raw      # A 列原文（含 _to_mux 后缀与位宽）
        self.input_base = input_base    # 剥位宽+_to_mux 后的基名 = 查 regmap/tmm 的 key
        self.input_width = input_width
        self.input_msb = input_msb
        self.input_lsb = input_lsb

    def __repr__(self):
        return "MuxCase(%s -> %s)" % (self.case_raw, self.input_base)


class MuxCtrl:
    """mux 页 B/C/D/E 列中的一个控制信号。

    WL_RFTRX 真表实证（2026-06-03 第十四轮）：一个 mux 组最多 4 个控制信号联合选择，
    case 值（F 列）= 各控制信号按列序拼接（**B 在最高位**）。LPBT 真表只用 B 列（单控制）。
    """

    def __init__(self, letter, raw, base, width, msb=None, lsb=None):
        self.letter = letter            # Excel 列字母 B/C/D/E
        self.raw = raw                  # 列原文（含 _to_mux 后缀与位宽）
        self.base = base                # 剥位宽+_to_mux 的基名（查 regmap/tmm/logic K 名的 key）
        self.width = width
        self.msb = msb
        self.lsb = lsb

    def __repr__(self):
        return "MuxCtrl(%s=%s[%d])" % (self.letter, self.base, self.width)

    @property
    def label(self):
        """显示用控制名：lsb>0 的切片(如 temp_code[3:1])显出来——非零偏移一眼可见(A1 隐患处，
        case 选值落在 [msb:lsb] 段而非低位)；lsb==0/无切片显裸基名(与历史显示一致，不扰动)。"""
        if self.lsb is not None and self.lsb > 0 and self.msb is not None:
            return "%s[%d:%d]" % (self.base, self.msb, self.lsb)
        return self.base


class MuxGroup:
    """mux 页的一组（同一个被验证输出）：G = case({B,C,D,E}) { F行: A行; ... }

    与 LogicSignal 的本质区别：N 选 1 的 case 结构化列表，不是表达式 AST。
    属性接口尽量与 LogicSignal 对齐（out_name/out_base/out_width/owner/assert_id/
    rtl_base/rtl_name/is_top），方便 GUI/报告复用。

    控制信号（2026-06-03 WL 多控制支持）：
      ctrls = list[MuxCtrl]，按 B/C/D/E 列序，B 在首 = case 值最高位段。
      ctrl_raw/ctrl_base/ctrl_width 是单控制兼容别名（= ctrls[0]），LPBT 路径/旧调用方零改动。
    """

    def __init__(self, group_no, out_name, out_width, ctrl_raw, ctrl_base, ctrl_width,
                 owner, top_output, cases, ctrls=None):
        self.group_no = group_no        # N 列组号（assert_mux<N>_T<n> 的 N，用户拍板方案 A）
        self.out_name = out_name        # G 列原文（含位宽；顶层网名，直接探）
        out_base, _w, out_msb, out_lsb = _strip_width(out_name)
        self.out_base = out_base
        self.out_msb = out_msb          # G 列位宽切片（探针 LHS 重建用，比字符串索引可靠）
        self.out_lsb = out_lsb
        self.out_width = out_width
        # ── 控制信号列表（B 在首 = case 高位）──
        if ctrls is None:
            # 兼容旧构造（单控制三标量；ctrl_base 为空 = 无控制信号）
            ctrls = [MuxCtrl("B", ctrl_raw, ctrl_base, ctrl_width)] if ctrl_base else []
        self.ctrls = ctrls
        # 单控制兼容别名（mux_gen/generator/gui/cli/旧测试大量引用；多控制时 = B 列那个）
        self.ctrl_raw = ctrls[0].raw if ctrls else ctrl_raw
        self.ctrl_base = ctrls[0].base if ctrls else ctrl_base
        self.ctrl_width = ctrls[0].width if ctrls else ctrl_width
        self.owner = owner              # L 列
        self.top_output = top_output    # I 列
        self.cases = cases              # list[MuxCase]
        # 同组各行控制列与首行不一致的行号（read_mux 填；expand_mux_group 转 issue）
        self.ctrl_mismatch_rows = []
        # 嵌套 mux 自动归一化提示（read_mux 把"一个块塞两级"折叠成多控制拼接时填，非空=已自动合并，
        # 给 designer 复核用；空=没动过）。详见 _normalize_nested_mux。
        self.normalized_note = ""
        # 与 LogicSignal 对齐的占位字段（expr 是只读属性=case 描述文本）
        self.suffix = "mux"
        # mux 输出被下游引用的去向尾缀（load_workbook 跨页回填，仅记录【客观事实】=被怎么引用）：
        #   被 logic 行以 <名>_to_logic 引用 → "_to_logic"；被别的 mux 以 <名>_to_mux 引用 → "_to_mux"。
        #   未被引用/顶层输出 = ""。注意：mux 输出端口名带不带这尾缀是【设计相关】(WL 的 mux 输出是裸名
        #   +另有 _to_logic 衔接网；Hi1108 rxiq 的 mux 输出端口本身叫 _to_logic)，Excel 推不出——故仅记录。
        self.ref_suffix = ""
        # 是否真把 ref_suffix 补到探针网名上。mux【默认 False=探裸名】(与 logic 默认 True 相反，因端口尾缀
        # 设计相关、不默认套)；Resolver 据 GenOptions.suffix_override 单点回填(rxiq 这类单点置 True)。
        self._append_to_logic = False
        # level_shift 页给的顶层 _ls 输出口真名（同 LogicSignal._ls_name；mux 输出若也过电平移位则回填）。
        self._ls_name = None
        self._ls_is_top = False

    @property
    def ctrl_total_width(self):
        """case 值（F 列）的总位宽 = 各控制信号位宽之和（多控制拼接）。单控制 = ctrl_width。"""
        return sum(c.width for c in self.ctrls) if self.ctrls else self.ctrl_width

    @property
    def is_multi_ctrl(self):
        return len(self.ctrls) > 1

    @property
    def expr(self):
        """人读的 case 描述（GUI 表达式列 / 报告 / 搜索过滤用）。
        注意这不是可求值表达式——mux 不走 expr.parse 路径。"""
        items = "; ".join("%s:%s" % (c.case_raw, c.input_base) for c in self.cases)
        if len(self.ctrls) > 1:
            sel = "{%s}" % ",".join(c.label for c in self.ctrls)   # 多控制拼接（B 在高位，带切片）
        else:
            sel = self.ctrls[0].label if self.ctrls else self.ctrl_base
        return "case(%s){%s}" % (sel, items)

    @property
    def is_top(self):
        """I 列 top_out==1（真表全 1）。"""
        return str(self.top_output).strip() in ("1", "1.0", "True", "true")

    @property
    def assert_id(self):
        """assert 标签前缀：assert_mux<N>_T<n>（mux 前缀防与 logic R 号撞号）。"""
        return "mux%s" % self.group_no

    @property
    def rtl_base(self):
        """mux 输出 RTL 网名。默认 = G 列基名【裸名】(顶层输出、以及绝大多数 mux 输出)。仅当本组被单点
        置「探尾缀网」(_append_to_logic=True，如 Hi1108 rxiq=2:1 mux 喂 sig_logic、端口真名带 _to_logic)
        才补其 Excel 去向尾缀(_to_logic/_to_mux)。mux 默认裸名——端口尾缀设计相关，Excel 推不出(WL 的 mux
        输出就是裸名)。level_shift 页给的顶层 _ls 口(_ls_name)最优先。"""
        if getattr(self, "_ls_name", None):
            return self._ls_name
        rs = self.ref_suffix if getattr(self, "_append_to_logic", False) else ""
        return rtl_net_name(self.out_base, self.suffix, is_top=self.is_top, ref_suffix=rs)

    @property
    def rtl_base_full(self):
        """rtl_base 但【无视 _append_to_logic 开关】，总按 ref_suffix 补尾缀——仅供探针前缀 key 匹配
        (与 LogicSignal.rtl_base_full 同口径：scan_rtl/用户配的前缀 key 常是带尾缀的全名)。"""
        if getattr(self, "_ls_name", None):
            return self._ls_name
        return rtl_net_name(self.out_base, self.suffix, is_top=self.is_top,
                            ref_suffix=self.ref_suffix)

    def _with_slice(self, base):
        """按解析出的 msb/lsb 重建位宽切片——不要用字符串索引拼接（G 列名含内部空格时会拼出非法 SV）。"""
        if self.out_msb is None or self.out_lsb is None:
            return base
        if self.out_msb == self.out_lsb:
            return "%s[%d]" % (base, self.out_msb)
        return "%s[%d:%d]" % (base, self.out_msb, self.out_lsb)

    @property
    def rtl_name(self):
        """含位宽切片的探针 LHS（assert 用）。"""
        return self._with_slice(self.rtl_base)

    @property
    def rtl_name_full(self):
        """rtl_name 的全名版(无视开关补尾缀)——仅供探针前缀 key 匹配，不进断言。"""
        return self._with_slice(self.rtl_base_full)

    def __repr__(self):
        return "MuxGroup(mux%s, %s, ctrl=%s, %d cases)" % (
            self.group_no, self.out_name, self.ctrl_base, len(self.cases))


# ───────────────────────────── 工具 ─────────────────────────────
def _s(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


_WIDTH_RANGE = re.compile(r"\[(\d+)\s*:\s*(\d+)\]\s*$")
_WIDTH_BIT = re.compile(r"\[(\d+)\]\s*$")


def _strip_width(text):
    """返回 (去掉位宽后的名字, width, msb, lsb)。无位宽则 width=1, msb=lsb=None。"""
    text = _s(text)
    m = _WIDTH_RANGE.search(text)
    if m:
        msb, lsb = int(m.group(1)), int(m.group(2))
        return text[:m.start()].strip(), abs(msb - lsb) + 1, msb, lsb
    m = _WIDTH_BIT.search(text)
    if m:
        b = int(m.group(1))
        return text[:m.start()].strip(), 1, b, b
    return text, 1, None, None


def strip_to_logic(name):
    """去掉 _to_logic 后缀（rule 2：输入 wire 名 = 输入名去 _to_logic）。"""
    name = name.strip()
    for suf in ("_to_logic", "_TO_LOGIC"):
        if name.endswith(suf):
            return name[: -len(suf)]
    return name


def strip_to_mux(name):
    """去掉 _to_mux 后缀（mux 页 A/B 列都带；剥掉后才能查 regmap/tmm/logic K 名）。
    注意 strip_to_logic 不剥 _to_mux —— 两个后缀语义不同，分开两个函数。"""
    name = name.strip()
    for suf in ("_to_mux", "_TO_MUX"):
        if name.endswith(suf):
            return name[: -len(suf)]
    return name


def rtl_net_name(base, suffix, is_top=True, to_logic_ref=False, ref_suffix=None):
    """K 列名 → RTL 真实网名。

    实证（2026-06-02 公司 lpbt_dig_top.v / BT_LP_DREG_sig_logic.v）：
      - M=ls 行：RTL 端口 = K 列名 + "_ls"（d_en_refbuf → d_en_refbuf_ls）
      - 内部信号(top_output=0)且被下游引用：RTL wire = K 列名 + 【引用尾缀 ref_suffix】，
        尾缀由 Excel 下游引用决定——被 logic 行以 <名>_to_logic 引用→"_to_logic"
        (pll_n1 → pll_n1_to_logic)；被 mux 页以 <名>_to_mux 引用→"_to_mux"。
        （在子模块内部，配合探针前缀即可验证）
      - 其余（顶层输出、未被引用的内部信号如 reserve）：K 列已是全名，原样用
    直接用 K 列名探针会 elaboration CUVUNF。
    ref_suffix=None 时退回旧入参 to_logic_ref（True→"_to_logic"），保证老调用方不变。
    """
    sfx = _s(suffix).lower()
    low = base.lower()
    if sfx == "ls" and not low.endswith("_ls"):
        return base + "_ls"
    rs = _s(ref_suffix if ref_suffix is not None else ("_to_logic" if to_logic_ref else ""))
    if not is_top and rs and not low.endswith(rs.lower()):
        return base + rs
    return base


def parse_hex_addr(text):
    """地址解析。本 Excel 约定：tmm 用 'h' 前缀(hex，如 hD=13)，regmap 用 'd' 前缀(十进制，如 d13=13)。
    'h1'/'h2D'/'0x2D'/'d13'/'2D' → int；无法解析返回 None。"""
    t = _s(text)
    if t == "":
        return None
    low = t.lower()
    try:
        if low.startswith("0x"):
            return int(low, 16)
        if low.startswith("'h"):
            return int(low[2:], 16)
        if low.startswith("h"):
            return int(low[1:], 16)
        if low.startswith("d") and low[1:].isdigit():
            return int(low[1:], 10)        # 'd13' = 十进制 13（regmap 约定）
        # 纯数字：可能是十进制行号也可能是 hex；按 hex 解释更贴近寄存器地址语义
        if re.fullmatch(r"[0-9a-fA-F]+", t):
            return int(t, 16)
    except ValueError:
        return None
    return None


def parse_bitfield(text):
    """'8' → (8,8)；'15:7' → (15,7)；无法解析返回 (None,None)。"""
    t = _s(text).replace(" ", "")
    m = re.fullmatch(r"(\d+):(\d+)", t)
    if m:
        return int(m.group(1)), int(m.group(2))
    if re.fullmatch(r"\d+", t):
        return int(t), int(t)
    return None, None


def _col(row, letter):
    """按列字母取值（row 是 values_only 的 tuple，0-based）。"""
    idx = column_index_from_string(letter) - 1
    return row[idx] if idx < len(row) else None


# ───────────────────────────── 读取 logic 页 ─────────────────────────────
LOGIC_INPUT_LETTERS = list("ABCDEFGHIJ")   # A..J


def read_logic(ws, header_row=2):
    """返回 list[LogicSignal]。只保留 K(输出名) 与 L(表达式) 同时非空的行。"""
    signals = []
    for ri, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        out_raw = _s(_col(row, "K"))
        expr = _s(_col(row, "L"))
        if out_raw == "" or expr == "":
            continue
        out_base, out_width, _, _ = _strip_width(out_raw)
        inputs = {}
        for letter in LOGIC_INPUT_LETTERS:
            cell = _s(_col(row, letter))
            if cell == "":
                continue
            base_with_tl, width, msb, lsb = _strip_width(cell)
            base = strip_to_logic(base_with_tl)
            inputs[letter] = {
                "raw": cell, "base": base, "width": width, "msb": msb, "lsb": lsb,
            }
        signals.append(LogicSignal(
            row=ri,
            out_name=out_raw,
            out_width=out_width,
            expr=expr,
            suffix=_s(_col(row, "M")),
            top_output=_s(_col(row, "N")),
            notes=_s(_col(row, "O")),
            owner=_s(_col(row, "P")),
            assert_id=_s(_col(row, "R")),
            inputs=inputs,
        ))

    # ── 标注"被下游引用"的内部信号 + 其引用尾缀（决定 RTL 探针网名，尾缀随 Excel 不写死）──
    # 实证(2026-06-02 BT_LP_DREG_sig_logic.v)：被下游 logic 行以 <名>_to_logic 引用的内部信号 RTL
    # wire 名 = K列名 + "_to_logic"(pll_n1 → pll_n1_to_logic)；被 mux 页以 <名>_to_mux 引用的则带
    # "_to_mux"(由 load_workbook 跨页回填)。reserve 这类没被引用的内部信号 RTL 名 = K 原文。
    # 这里先扫 logic 输入列（_to_logic 优先；个别 logic 行也可能引用 _to_mux 衔接网）。
    #
    # ⭐ 排除【自引用】(2026-06-11 Hi1108 d_wl_rf_dpd_s2d_bias_en 实证)：输出 X 自己读 <X>_to_logic 当
    #   输入时，那根 <X>_to_logic 是 X 的【前级原始信号】(RTL: 端口 X → regfile → X_to_logic)，不是下游
    #   消费 X 输出网的引用。若据此给 X 补 _to_logic，X 的输出探针网名就 = 它自己的输入网名 → 撞名探错。
    #   故：base==本行输出基名(自引用)时不计入。只有【别的】信号引用 <X>_to_logic 才算真·下游引用。
    #   (正常级联如 pll_n1 不受影响：pll_n1 的 _to_logic 来自 pll_n2 的引用，非 pll_n1 自己的输入。)
    #
    # ⭐⭐ 自引用【抑制输出尾缀】(2026-06-16 Hi1108 WL_C0C1 d_wl_rf_5g_tx_lodiv_en 实证)：
    #   若 X 自己读 <X>_to_logic 当输入（self_ref_logic），那根 <X>_to_logic 是 X 的【前级输入网】
    #   (端口/寄存器 X → regfile → X_to_logic → logic)，**不是 X 的输出网**；X 的输出是 bare 基名
    #   (顶层 pin / 计算结果)。此时即便【别的行】也引用 <X>_to_logic（如 trxbuf 行引用
    #   d_wl_rf_5g_tx_lodiv_en_to_logic），它们读的也是【同一根前级网】，不能据此把 X 的【输出探针】
    #   补成 X_to_logic —— 那会让 assert 探到 X 自己的输入(=直接读输入信号，仿真假绿/假红)。
    #   故：X ∈ self_ref_logic 时，其输出尾缀强制清空（探 bare 基名）。
    #   resolver 输入侧早已同口径(把 <X>_to_logic 自输入 force 成 bare 基名)，这里把输出侧对齐。
    #   仅当自引用后缀==下游引用后缀时抑制（X 自读 _to_logic 但被别人按 _to_mux 引用 → _to_mux 是
    #   另一根真·输出网，不抑制）。M=ls 行不受影响(rtl_net_name 的 ls 分支优先于 ref_suffix)。
    ref_logic, ref_mux = set(), set()
    self_ref_logic, self_ref_mux = set(), set()
    for sig in signals:
        ob = sig.out_base.lower()
        for info in sig.inputs.values():
            rb = _strip_width(info["raw"])[0].lower()
            if rb.endswith("_to_logic"):
                base = rb[: -len("_to_logic")]
                if base != ob:                       # 排除自引用（前级信号，非下游引用）
                    ref_logic.add(base)
                else:
                    self_ref_logic.add(ob)           # X 读自己的 X_to_logic → 那是输入前级网
            elif rb.endswith("_to_mux"):
                base = rb[: -len("_to_mux")]
                if base != ob:
                    ref_mux.add(base)
                else:
                    self_ref_mux.add(ob)
    for sig in signals:
        b = sig.out_base.lower()
        cand = "_to_logic" if b in ref_logic else ("_to_mux" if b in ref_mux else "")
        if cand == "_to_logic" and b in self_ref_logic:
            cand = ""                                # 自引用：输出探 bare 基名，不补 _to_logic
        elif cand == "_to_mux" and b in self_ref_mux:
            cand = ""
        sig.ref_suffix = cand
    return signals


# ───────────────────────────── 读取 mux 页 ─────────────────────────────
MUX_CTRL_LETTERS = list("BCDE")     # 控制信号 1..4（LPBT 只用 B；WL 最多 4 个联合选择）


def _row_ctrls(row):
    """读一行的 B/C/D/E 控制列 → list[MuxCtrl]（按列序，B 在首 = case 高位）。"""
    ctrls = []
    for letter in MUX_CTRL_LETTERS:
        v = _s(_col(row, letter))
        if v:
            c_base_tm, c_width, c_msb, c_lsb = _strip_width(v)
            ctrls.append(MuxCtrl(letter, v, strip_to_mux(c_base_tm), c_width, c_msb, c_lsb))
    return ctrls


def _case_bitstr(case_raw, expected_width):
    """case 字面量 → 定宽 '0'/'1'/'x' 位串（MSB 在左）。只认二进制；非二进制/位宽不符 → ValueError。

    不依赖 expr（避免 excel_model→expr 依赖）；归一化拼接 case 用，遇到八/十/十六进制直接 ValueError
    让归一化放弃（保守：不认就不动，维持原报错）。
    """
    m = re.match(r"\s*(\d+)?'[bB]([01xXzZ?_]+)\s*$", case_raw or "")
    if not m:
        raise ValueError("非二进制 case，不归一化: %r" % case_raw)
    bits = m.group(2).replace("_", "")
    width = int(m.group(1)) if m.group(1) else len(bits)
    if len(bits) < width:
        bits = "0" * (width - len(bits)) + bits      # 左补 0 到声明位宽（4'b01 = 0001）
    elif len(bits) > width:
        bits = bits[-width:]
    for ch in "XZz?":
        bits = bits.replace(ch, "x")
    if width != expected_width:
        raise ValueError("case 位宽 %d != 控制位宽 %d" % (width, expected_width))
    return bits


def _remake_case(src_case, bits, total_w):
    """以 src_case 的数据输入造一个新 MuxCase，case_raw = 新拼接位串（数据/位宽切片全继承）。"""
    return MuxCase(row=src_case.row, case_raw="%d'b%s" % (total_w, bits),
                   input_raw=src_case.input_raw, input_base=src_case.input_base,
                   input_width=src_case.input_width, input_msb=src_case.input_msb,
                   input_lsb=src_case.input_lsb)


def _normalize_nested_mux(grp, per_case_ctrls):
    """把『一个输出块塞了两级 mux』(不同行用不同控制 + 一个自引用)折叠成多控制拼接。

    designer 实证形态（envdet5g cornercode，2026-06-04 真表 1607~1612 行）：
        里层  case(r_code[2:0]) {00x:lut0; 01x:lut1; 10x:lut2; 11x:lut3}
        外层  case(sel)         {0: <里层结果=输出自己>; 1: local}
      → 合并成 case({sel, r_code}) {1xxx:local; 000x:lut0; 001x:lut1; 010x:lut2; 011x:lut3}
        （sel 高位、r_code 低位；外层非自引用 case 的里层位取 don't-care）。

    ⭐只在【唯一安全形态】触发，否则返回 False 保持原样（→ 维持 ctrl_mismatch_rows 报错，绝不猜）：
      ① 正好 2 个控制 regime（按控制基名分子块），且每个子块都是单控制；
      ② 全组正好 1 个自引用（数据基名==输出基名），它所在子块=外层，另一个=里层；
      ③ 各子块 case 都是二进制、位宽==其控制位宽（算得通）。
    成功 → 改写 grp.ctrls/grp.cases、清 ctrl_mismatch_rows、挂 grp.normalized_note；返回 True。

    安全性：这条路径只在【今天本就报"控制列不一致"错】的块上跑（read_mux 仅对 ctrl_mismatch
    的组调它）——能跑通的正常表每个块控制一致、永不进来；折叠产物是工具早支持的多控制拼接，
    resolver/mux_gen/generator/gui 全不改。
    """
    cases = grp.cases
    if len(per_case_ctrls) != len(cases) or len(cases) < 3:
        return False
    # 按控制基名分子块（保序）；任一行非单控制 → 放弃
    blocks = OrderedDict()
    for case, ctrls in zip(cases, per_case_ctrls):
        if len(ctrls) != 1:
            return False
        bkey = ctrls[0].base.lower()
        if bkey not in blocks:
            blocks[bkey] = {"ctrl": ctrls[0], "cases": []}
        blocks[bkey]["cases"].append(case)
    if len(blocks) != 2:
        return False
    # 找自引用：数据基名 == 输出基名；全组必须正好 1 个，且只落在一个子块
    out_low = grp.out_base.lower()
    self_keys = [bk for bk, b in blocks.items()
                 if any(c.input_base.lower() == out_low for c in b["cases"])]
    if len(self_keys) != 1:
        return False
    outer_key = self_keys[0]
    inner_key = next(bk for bk in blocks if bk != outer_key)
    outer, inner = blocks[outer_key], blocks[inner_key]
    self_cases = [c for c in outer["cases"] if c.input_base.lower() == out_low]
    if len(self_cases) != 1:
        return False
    self_case = self_cases[0]
    outer_ctrl, inner_ctrl = outer["ctrl"], inner["ctrl"]
    inner_w = inner_ctrl.width
    total_w = outer_ctrl.width + inner_w
    try:
        outer_bits = {id(c): _case_bitstr(c.case_raw, outer_ctrl.width) for c in outer["cases"]}
        inner_bits = {id(c): _case_bitstr(c.case_raw, inner_w) for c in inner["cases"]}
    except ValueError:
        return False                          # 非二进制/位宽不通：不猜，维持原报错
    # 构造拼接 case：外层高位、里层低位
    new_cases = []
    for c in outer["cases"]:
        if c is self_case:
            continue
        new_cases.append(_remake_case(c, outer_bits[id(c)] + "x" * inner_w, total_w))
    self_hi = outer_bits[id(self_case)]
    for c in inner["cases"]:
        new_cases.append(_remake_case(c, self_hi + inner_bits[id(c)], total_w))
    # 改写组：多控制 [外层(高,B), 里层(低,C)]
    outer_ctrl.letter, inner_ctrl.letter = "B", "C"
    grp.ctrls = [outer_ctrl, inner_ctrl]
    grp.ctrl_raw, grp.ctrl_base, grp.ctrl_width = outer_ctrl.raw, outer_ctrl.base, outer_ctrl.width
    grp.cases = new_cases
    grp.ctrl_mismatch_rows = []
    grp.normalized_note = (
        "自动识别为嵌套 mux：外层 %s 选 {里层 %s 的结果, %s} → 已合并成多控制拼接 {%s,%s}（请核对）"
        % (outer_ctrl.base, inner_ctrl.base, self_case.input_base,
           outer_ctrl.base, inner_ctrl.base))
    return True


def read_mux(ws, header_row=2):
    """读 mux 页 → list[MuxGroup]。

    真表排版（LPBT 2026-06-03 / WL_RFTRX 2026-06-03 两轮 inspect_mux 实证，表头第2行）：
      A=mux_input（寄存器字段+_to_mux）   B~E=mux_ctrl_sig1~4（控制信号，LPBT 只用 B）
      F=case（控制值=各控制按列序拼接，B 高位；可含 don't-care 位 x，如 5'b0xxxx）
      G=mux_out（被验证输出）  H=去向(to_logic/to_mux/to_dft)  I=top_out  L=Owner  N=组号
    语义：G = case({B,C,D,E}) { F行: A行; ... }

    只保留 A/F/G 同时非空的行（'(reserved)' 行 F 为空被自然过滤）。
    分组按 **G 列基名全局归并**（同一输出的行即使不连续也归同一组）——
    劈成多组会导致 assert_id 撞号(非法 SV) + 互异值各自重启(撞值=假绿)。
    控制信号取自组的首行；后续行控制列与首行不一致时记入 ctrl_mismatch_rows
    （expand_mux_group 转 issue，不静默吞掉）。
    不依赖 N 列本身，防 read_only 模式下行尾留空导致 N 列读丢（232 列宽表的真实风险）。
    组号优先用首行 N 列值；N 列缺失/非法时顺延（上一组+1）。
    """
    groups = []
    by_base = {}        # G 基名(小写) -> MuxGroup（全局归并，非连续行也并入同一组）
    rowctrls = {}       # G 基名(小写) -> [本行有效控制列]，与 grp.cases 一一对应（嵌套归一化用）
    for ri, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                             start=header_row + 1):
        in_raw = _s(_col(row, "A"))
        case_raw = _s(_col(row, "F"))
        out_raw = _s(_col(row, "G"))
        if in_raw == "" or case_raw == "" or out_raw == "":
            continue

        in_base_tm, in_width, in_msb, in_lsb = _strip_width(in_raw)
        case = MuxCase(
            row=ri, case_raw=case_raw, input_raw=in_raw,
            input_base=strip_to_mux(in_base_tm),
            input_width=in_width, input_msb=in_msb, input_lsb=in_lsb,
        )

        out_base, out_width, _, _ = _strip_width(out_raw)
        key = out_base.lower()
        if key in by_base:
            grp = by_base[key]
            # 同组各行控制列应与首行一致（基名比较）。续行控制列【全空】视为合法——继承首行控制
            # （Excel 合并单元格 / 控制只写组首行是 spec 表常见排版；read_only 模式下合并续行读到空）。
            # 只有续行【写了非空但不同】的控制信号才记 mismatch（expand 时转 issue，抓真错误）。
            this_ctrls = _row_ctrls(row)
            row_bases = [c.base.lower() for c in this_ctrls]
            if row_bases and row_bases != [c.base.lower() for c in grp.ctrls]:
                grp.ctrl_mismatch_rows.append(ri)
            grp.cases.append(case)
            rowctrls[key].append(this_ctrls or list(grp.ctrls))   # 空控制列继承首行
            continue

        # ── 新组 ──
        ctrls = _row_ctrls(row)                 # 收集全部非空控制列（B 在首 = case 高位）
        n_raw = _s(_col(row, "N"))
        try:
            group_no = int(float(n_raw))
        except (ValueError, TypeError):
            group_no = groups[-1].group_no + 1 if groups else 1
        grp = MuxGroup(
            group_no=group_no,
            out_name=out_raw, out_width=out_width,
            ctrl_raw="", ctrl_base="", ctrl_width=1,    # 占位；实际控制来自 ctrls
            owner=_s(_col(row, "L")),
            top_output=_s(_col(row, "I")),
            cases=[case],
            ctrls=ctrls,
        )
        by_base[key] = grp
        rowctrls[key] = [list(ctrls)]
        groups.append(grp)

    # 嵌套 mux 归一化：仅对【控制列不一致】的组试折叠（两级嵌套+单自引用 → 多控制拼接）；
    # 不匹配则维持 ctrl_mismatch_rows 原报错。正常表每组控制一致，永不进这里。
    for grp in groups:
        if grp.ctrl_mismatch_rows:
            _normalize_nested_mux(grp, rowctrls.get(grp.out_base.lower(), []))

    # 组号冲突兜底：N 列读丢/重复导致两组同号时，后者顺延到未用号（撞号=非法 SV 标签）
    seen_no = set()
    for grp in groups:
        while grp.group_no in seen_no:
            grp.group_no += 1
        seen_no.add(grp.group_no)
    return groups


# ───────────────────────────── 读取 regmap 页 ─────────────────────────────
def read_regmap(ws, header_row=2):
    """返回 (out, dups)。bit 位置由 J..Y(=bit15..bit0) 哪列非空推出。
      out:  dict signal_name -> RegmapEntry。
      dups: dict signal_low -> [RegmapEntry,…]，仅含同名出现 >1 次者(按出现序，含被采纳的首个)，供告警。

    ⭐同名信号重复定义时【保留首个】(setdefault 语义)——与表自身 for_test 的 VLOOKUP(取首个匹配)
      及 read_tmm(setdefault 保留首个) 一致。旧版 `out[signal]=…` 让【后者覆盖前者】是 bug：真表
      d_bt_lp_pmu_test_en 同名出现于 0x32/bit12(先) 与 0x36/bit1(后)，后者覆盖→工具错绑 0x36/bit1
      →仿真 T2 挂；而设计师 for_test 用的真值是首个 0x32/bit12(2026-06-12 实证)。重复不静默吞，
      经 dups 透出告警(见 generator._regmap_dup_warning)。"""
    out = {}
    seen_lower = set()      # 已采纳(首个)的小写名
    dup_lower = {}          # 小写名 -> [所有同名条目]，用于检出重复并告警
    # J..Y 共 16 列，分别代表 bit15..bit0
    bit_cols = [openpyxl.utils.get_column_letter(c) for c in
                range(column_index_from_string("J"), column_index_from_string("Y") + 1)]
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        raw_signal = _s(_col(row, "G"))
        if raw_signal == "":
            continue
        # Signal_Name 带位宽如 int_n[8:0]，去掉 [msb:lsb] 后才能与 logic 输入基名匹配
        signal, _w, _m, _l = _strip_width(raw_signal)
        occupied = []
        for offset, col in enumerate(bit_cols):
            val = _s(_col(row, col))
            if val not in ("", "0"):
                occupied.append(15 - offset)   # J=bit15 ... Y=bit0
        if occupied:
            bit_msb, bit_lsb = max(occupied), min(occupied)
        else:
            bit_msb = bit_lsb = None
        entry = RegmapEntry(
            signal=signal,
            reg_name=_s(_col(row, "D")),
            reg_type=_s(_col(row, "F")),
            default=_s(_col(row, "I")),
            bit_lsb=bit_lsb,
            bit_msb=bit_msb,
            owner=_s(_col(row, "AE")),
            address=parse_hex_addr(_col(row, "H")),   # regmap H 列 = 地址（'d13'）
        )
        low = signal.lower()
        dup_lower.setdefault(low, []).append(entry)
        if low not in seen_lower:                      # 保留首个(与 VLOOKUP/read_tmm 一致)
            out[signal] = entry
            seen_lower.add(low)
    dups = {low: lst for low, lst in dup_lower.items() if len(lst) > 1}
    return out, dups


# ───────────────────────────── 读取 total_memory_map 页 ─────────────────────────────
def read_tmm(ws):
    """
    状态机解析：跳过 # / 'Register Name' / 'Field Name' 标题行；
    字段行判定 = B 像 bit 号/范围 且 F 像 hex 地址。
    返回 dict: field_name -> TmmField（同名取首个；地址不同会各保留，键带后缀）。
    """
    fields = {}
    cur_reg_name = ""
    for row in ws.iter_rows(values_only=True):
        a = _s(_col(row, "A"))
        b = _s(_col(row, "B"))
        f = _s(_col(row, "F"))
        if a == "":
            continue
        low = a.lower()
        if low.startswith("#") or low in ("register name", "field name"):
            continue
        # 寄存器定义行：A=寄存器名, B=地址hex, D=类型, 记录当前寄存器名
        reg_addr = parse_hex_addr(b)
        bit_msb, bit_lsb = parse_bitfield(b)
        field_addr = parse_hex_addr(f)
        if bit_msb is not None and field_addr is not None:
            # 字段行。字段名(A列)带位宽标注如 int_n[8:0] / d_pfd_en_lnmode[1:0]，
            # 去掉 [msb:lsb] 后才能与 logic 输入基名(已去位宽)匹配。
            name, _w, _m, _l = _strip_width(a)
            pin = _s(_col(row, "D")).upper()
            dig = "Y" if pin in ("Y", "YES") else ("N" if pin in ("N", "NO") else None)
            raw_type = _s(_col(row, "H"))
            entry = TmmField(
                name=name,
                bit_msb=bit_msb,
                bit_lsb=bit_lsb,
                address=field_addr,
                reg_type=_normalize_type(raw_type),
                reg_type_raw=raw_type,
                dig_top_pin=dig,        # 精确匹配，避免 'NA'/'N/A' 被首字符截断误判为 'N'
                reg_name=cur_reg_name,
            )
            # 同名字段：保留首个，但记录所有以便诊断
            fields.setdefault(name, entry)
        elif reg_addr is not None:
            # 寄存器定义行
            cur_reg_name = a
    return fields


def _normalize_type(v):
    """归一化寄存器类型为 RO/RW。基于关键词，兼容 RW/RO/R/W/R/W/READ WRITE/READ ONLY/WO 等写法。"""
    t = _s(v).upper()
    if not t:
        return None
    has_w = ("W" in t) or ("WRITE" in t)   # RW/W/WO/W/O/READ WRITE/...
    has_r = ("R" in t) or ("READ" in t)
    if has_w:                              # 可写（含只写 WO）→ 按 RW 处理
        return "RW"
    if has_r:                              # 只读
        return "RO"
    return t or None


# ───────────────────────────── 顶层装载 ─────────────────────────────
class DregWorkbook:
    def __init__(self, logic, regmap, tmm, sheet_names, mux=None, dft=None,
                 fortest_order=None, regmap_dups=None):
        self.logic = logic              # list[LogicSignal]
        self.regmap = regmap            # dict
        self.tmm = tmm                  # dict
        self.sheet_names = sheet_names
        self.mux = mux if mux is not None else []   # list[MuxGroup]（无 mux 页 = 空列表）
        # {signal_low: [RegmapEntry,…]} regmap 同名重复定义(已按首个采纳)，供 generator 告警。
        # 空 = 无重复(绝大多数表)。来自 read_regmap；手构 wb 不传=空。
        self.regmap_dups = regmap_dups if regmap_dups is not None else {}
        # {输出基名low: 门控信息}（dft 页；无 dft 页=空）。iddq DFT 拍 / pin_dft_gate 用，见 read_dft。
        self.dft = dft if dft is not None else {}
        # {输出基名low: [输入基名low顺序]}（for_test 页；无/空页=空）。真值表输入行排序对照用。
        self.fortest_order = fortest_order if fortest_order is not None else {}


def _dft_strip(name):
    """dft 门/输出名 → 去位宽 + 去 _to_dft 后缀的基名。"""
    s = _strip_width(name)[0]
    if s.lower().endswith("_to_dft"):
        s = s[:-len("_to_dft")]
    return s


_DFT_GATE_LO = re.compile(r"^\s*\w+\s*\?\s*0\s*:\s*\w+\s*$")    # cond?0:A → 门=0 时透传(取功能值 A)
_DFT_GATE_HI = re.compile(r"^\s*\w+\s*\?\s*\w+\s*:\s*0\s*$")    # cond?A:0 → 门=1 时透传


def read_dft(ws, header_row=2):
    """读 dft 页：被 DFT mux 门控的输出 + 其门控（Hi1108 实证，全表一致）。

    列：B=门信号(to_dft，恒=iddq_mode)、D=被观测输出(=mux/logic 的 G/K 基名)、E=门控表达式 B?0:A。
    `B?0:A` 含义：门 B=0 → 输出取功能值 A(透传)；B=1 → 输出强制 0(IDDQ 漏电测试)。
    返回 {输出基名low: {"gate_base": 门基名low, "gate_raw": 门原文, "transparent": 0/1}}。
    只收能识别成"单条件门 cond?0:x / cond?x:0"的行（其余跳过，不乱猜）。
    用途：iddq DFT 拍把门 force 到非透传支(=1，输出压常量 0)、pin_dft_gate 把门当显式输入 force 到
    透传值，对照 designer 的 for_test（iddq=0 做法）。
    """
    out = {}
    if ws is None:
        return out
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        out_raw = _s(r[3]) if len(r) > 3 else ""       # D 列：被观测输出
        gate_raw = _s(r[1]) if len(r) > 1 else ""      # B 列：门信号(iddq_mode)
        expr = _s(r[4]) if len(r) > 4 else ""          # E 列：门控表达式
        if not out_raw or not gate_raw or not expr:
            continue
        if _DFT_GATE_LO.match(expr):
            transparent = 0
        elif _DFT_GATE_HI.match(expr):
            transparent = 1
        else:
            continue                                   # 非"单条件门"形态：不处理
        ob = _strip_width(out_raw)[0].lower()
        out[ob] = {"gate_base": _dft_strip(gate_raw).lower(),
                   "gate_raw": gate_raw, "transparent": transparent}
    return out


_FT_WIRE_SUFFIXES = ("_to_dft", "_to_logic", "_to_mux")


def _ft_norm(name):
    """for_test 顺序对照用的名字归一：剥位宽 + 小写 + 剥接线后缀(_to_dft/_to_logic/_to_mux)。

    designer 的 for_test 观测点在 DFT mux 之后，输出/输入名可能写衔接网名（带后缀）；
    我们这边对照用的是基名——归一后两边才能对上（_ls 是真网名的一部分，不剥）。"""
    s = _strip_width(name)[0].lower()
    for suf in _FT_WIRE_SUFFIXES:
        if s.endswith(suf):
            return s[:-len(suf)]
    return s


def read_fortest_order(ws):
    """读 for_test 页的【输入行顺序】：{输出基名low: [输入基名low(按 for_test 行序)]}。

    2026-06-10 用户要求：我们真值表的输入行次序要与 designer for_test 一致。
    真表排版(R21/R27 实证，我们的回填也同构)：每组 = 若干输入行(F 列=输入名)
    + 输出行(D 列=输出名)；以"遇到 D 行"给当前组封口。
    只做顺序对照：F/D 都走 _ft_norm 归一（剥位宽/接线后缀）；解析不出/无 for_test 页
    → 空 dict(回退原序)。表头等杂行会进无人查询的键，无害。
    """
    out = {}
    if ws is None:
        return out
    cur = []
    # 只读 A..F 列（for_test 动辄上万行、T 向量几十列——读全宽会拖慢装载）
    for r in ws.iter_rows(min_row=1, max_col=6, values_only=True):
        fval = _s(r[5]) if len(r) > 5 else ""      # F 列：输入信号名
        dval = _s(r[3]) if len(r) > 3 else ""      # D 列：输出信号名(组封口)
        if fval:
            cur.append(_ft_norm(fval))
        if dval:
            ob = _ft_norm(dval)
            if cur and ob:
                out.setdefault(ob, list(cur))
            cur = []
    return out


def _find_sheet(wb, *candidates):
    lowered = {s.lower(): s for s in wb.sheetnames}
    for c in candidates:
        if c.lower() in lowered:
            return wb[lowered[c.lower()]]
    return None


def _apply_mux_ref_suffix(logic, mux):
    """跨页回填：mux 页(数据 A 列 / 控制 B-E 列)以 <名>_to_mux 引用某个 logic 输出时，该 logic 输出的
    RTL wire 就叫 <名>_to_mux —— 给同名 logic 输出的 ref_suffix 置 "_to_mux"（探针网名随之带尾缀）。

    只认 raw 真带 _to_mux 的引用（input_base/ctrl.base 已剥 _to_mux，故查 raw 后缀确认）；已被 logic
    以 _to_logic 引用的不覆盖（_to_logic 优先，保持 LPBT 行为）。被引用基名若是寄存器(非 logic 行)则
    不在 logic 列表里、自然不匹配。尾缀是 Excel 决定的，工具忠实照搬、不写死成 _to_logic。"""
    mux_refs = set()
    for g in (mux or []):
        for c in g.cases:
            if _strip_width(c.input_raw)[0].lower().endswith("_to_mux"):
                mux_refs.add(c.input_base.lower())
        for ct in g.ctrls:
            if _strip_width(ct.raw)[0].lower().endswith("_to_mux"):
                mux_refs.add(ct.base.lower())
    for sig in logic:
        if not sig.ref_suffix and sig.out_base.lower() in mux_refs:
            sig.ref_suffix = "_to_mux"


def _apply_mux_output_ref_suffix(logic, mux):
    """跨页回填【mux 输出】的引用尾缀（2026-06-11 Hi1108 rxiq_phase_ctrl 实证）：

    top_output=0 的 mux 输出若被下游引用，RTL wire 名 = G列基名 + 去向尾缀，和 logic 输出同一套
    net flows-to 命名约定——被 logic 行以 <名>_to_logic 引用 → "_to_logic"(rxiq_phase_ctrl 是 2:1
    mux 输出、喂 sig_logic，RTL: WL_TRX_C0C1_DREG_sig_mux.v 里 d_wl_rf_rxiq_phase_ctrl_to_logic)；
    被别的 mux 以 <名>_to_mux 引用 → "_to_mux"(mux→mux 级联)。顶层输出(top_output=1，原 7 个 WL
    mux)不带尾缀——只动 top_output=0 且被引用的 mux 输出。_to_logic 优先。

    与 read_logic / _apply_mux_ref_suffix 同源：扫 logic 输入列的 _to_logic/_to_mux 引用 + mux 页
    数据/控制列的 _to_mux 引用。被引用基名若不是 mux 输出(是寄存器/logic 行)自然不匹配。"""
    ref_logic, ref_mux = set(), set()
    for sig in logic:
        for info in sig.inputs.values():
            rb = _strip_width(info["raw"])[0].lower()
            if rb.endswith("_to_logic"):
                ref_logic.add(rb[: -len("_to_logic")])
            elif rb.endswith("_to_mux"):
                ref_mux.add(rb[: -len("_to_mux")])
    for g in (mux or []):
        for c in g.cases:
            if _strip_width(c.input_raw)[0].lower().endswith("_to_mux"):
                ref_mux.add(c.input_base.lower())
        for ct in g.ctrls:
            if _strip_width(ct.raw)[0].lower().endswith("_to_mux"):
                ref_mux.add(ct.base.lower())
    for g in (mux or []):
        if g.is_top or g.ref_suffix:
            continue
        b = g.out_base.lower()
        if b in ref_logic:
            g.ref_suffix = "_to_logic"
        elif b in ref_mux:
            g.ref_suffix = "_to_mux"


def read_level_shift(ws, header_row=2):
    """读 level_shift 页 → {数字域输出基名low: {"out": 顶层电平移位输出口真名, "is_top": bool}}。

    新表(Hi1108 Pilot, 2026-06)把『level shift → 顶层输出口』拆成独立页。实证列：
      A=level_shift_input(<base>_to_ls)  C=level_shift_output_name(顶层口真名)  E=top_out
    例：d_en_refbuf_to_ls →(STD_SR_L2H)→ d_en_refbuf_ls(E=1，lpbt_dig_top.v:95 output)。
    输出口真名【按设计而异】：d_* 类带 _ls(d_en_refbuf_ls)，en_pll* 类是裸名(en_pll_ld)——
    故 out 取 C 列原值(不自己拼后缀)、key 取 A 列剥 _to_ls 的基名。
    无此页 / 任一行缺 A 或 C → 跳过；整页空 → 空 dict（旧表 M=ls 直给后缀的路径完全不受影响）。
    """
    out = {}
    if ws is None:
        return out
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        in_raw = _s(r[0]) if len(r) > 0 else ""     # A 列：level_shift 输入(<base>_to_ls)
        out_raw = _s(r[2]) if len(r) > 2 else ""    # C 列：顶层输出口真名
        top_raw = _s(r[4]) if len(r) > 4 else ""    # E 列：top_out
        if not in_raw or not out_raw:
            continue
        low = _strip_width(in_raw)[0].lower()
        for suf in ("_to_ls", "_ls"):               # 从输入名剥得数字域基名(d_en_refbuf)
            if low.endswith(suf):
                low = low[: -len(suf)]
                break
        is_top = top_raw.strip() in ("1", "1.0", "True", "true")
        out.setdefault(low, {"out": _strip_width(out_raw)[0], "is_top": is_top})
    return out


def _apply_level_shift(logic, mux, ls_map):
    """据 level_shift 页把 logic/mux 输出的探针网名定到顶层电平移位口（2026-06-12 Hi1108 Pilot）。

    新表 logic 输出 M=to_dft、top_output=0，裸名在 RTL 不存在——真顶层口是 d_en_refbuf_ls
    (level_shift 页 + lpbt_dig_top.v output 双证)。给输出基名命中 level_shift 输入基名的信号设
    _ls_name=顶层口真名，rtl_base/rtl_name 改返回它(压过 M 列/ref_suffix/加尾缀开关)。
    无 level_shift 页 → ls_map 空 → 一个信号都不动（旧 LPBT/WL 表逐字节不变）。"""
    if not ls_map:
        return
    for sig in logic:
        info = ls_map.get(sig.out_base.lower())
        if info and info.get("out"):
            sig._ls_name = info["out"]
            sig._ls_is_top = bool(info.get("is_top"))
    for g in (mux or []):
        info = ls_map.get(g.out_base.lower())
        if info and info.get("out"):
            g._ls_name = info["out"]
            g._ls_is_top = bool(info.get("is_top"))


def load_workbook(path, logic_header_row=2, regmap_header_row=2):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws_logic = _find_sheet(wb, "logic")
    if ws_logic is None:
        wb.close()
        raise SystemExit("Excel 中找不到 'logic' 页；现有页：%s" % wb.sheetnames)
    ws_regmap = _find_sheet(wb, "regmap")
    ws_tmm = _find_sheet(wb, "total_memory_map", "total memory map", "memory_map")
    ws_mux = _find_sheet(wb, "mux")
    ws_dft = _find_sheet(wb, "dft")
    ws_ls = _find_sheet(wb, "level_shift", "level shift", "levelshift")
    ws_ft = _find_sheet(wb, "for_test")

    logic = read_logic(ws_logic, logic_header_row)
    regmap, regmap_dups = (read_regmap(ws_regmap, regmap_header_row)
                           if ws_regmap is not None else ({}, {}))
    tmm = read_tmm(ws_tmm) if ws_tmm is not None else {}
    mux = read_mux(ws_mux) if ws_mux is not None else []
    _apply_mux_ref_suffix(logic, mux)        # 跨页：mux 以 _to_mux 引用的 logic 输出 → 回填 _to_mux 尾缀
    _apply_mux_output_ref_suffix(logic, mux) # 跨页：被 logic/mux 引用的 top_out=0 mux 输出 → 回填去向尾缀
    # 跨页：level_shift 页声明的顶层电平移位口 → 把 logic/mux 输出探针定到 _ls 顶层口(_ls_name 最优先)
    _apply_level_shift(logic, mux, read_level_shift(ws_ls) if ws_ls is not None else {})
    dft = read_dft(ws_dft) if ws_dft is not None else {}
    fortest_order = read_fortest_order(ws_ft) if ws_ft is not None else {}
    names = list(wb.sheetnames)
    wb.close()
    return DregWorkbook(logic, regmap, tmm, names, mux=mux, dft=dft,
                        fortest_order=fortest_order, regmap_dups=regmap_dups)
