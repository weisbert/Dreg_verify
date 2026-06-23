# -*- coding: utf-8 -*-
"""
fortest_writer.py — 把 GUI/generator 的测试项【回填】到 Excel 的 for_test 页排版。

产物=新 .xlsx：**完整复制源 Excel 全部 sheet，只把 for_test 页换成我们生成的内容**
（源文件不动）。for_test 列排版逐格复刻真表(Hi1107C WL，2026-06-09 用 inspect_fortest
--raw 摸清)，给 designer 看 + 复制粘贴：

  每组(每输出一个) = [输入行 × n] + [输出行] + [per-test 名字行]
  col A  每组首行=case 组名(输出名+_g序号；真表里是 case/aa 这种占位垃圾，我们给真名)
  col B/C  RF_WRITE 地址/值——放在每寄存器【最后一个字段行】(C=该寄存器合并写值)；
           RO/wire 输入则每行 B=force 网名、C=16'h0
  col D    输出信号名(仅输出行)
  col E    期望(输出行；=末列那拍的期望，十进制)
  col F    输入信号名(每输入一行)
  col G    末列(Tn)那拍的值拷贝(VBA scratch；输入=二进制串、输出=十进制)
  col H    运行态寄存器累加值(VBA scratch；同寄存器逐字段 H+=(末列值<<lsb)，RO 行=0)
  col I    每行地址：RW=10'h..；RO/wire=force 网名
  col J    RO True/False
  col K/L/M  bits(如 7:4) / MSB / LSB
  col N    组号
  col O 起  T1..Tn 真值表向量：输入=二进制(补零到位宽)，输出=十进制

数据来源=generator.report() 的 tables(只取 is_logic 的 logic cone 表；mux 表结构不同，跳过)。
report 的 inputs[] 已带 addr/RO/bit/wire 元数据、tests[] 已带 raw(逐输入整数)/exp_num。

依赖 openpyxl（项目已用它读 Excel）。
"""
import re

from . import sv_writer as W


def _strip_width(name):
    """去掉信号名末尾的 [msb:lsb] 位宽标注，作 case 组名基。"""
    return re.sub(r"\[[^\]]*\]\s*$", "", str(name or "")).strip()


def _bin(val, width):
    """输入向量值→二进制串(补零到位宽，无前缀)，与真表一致：1bit='0'/'1'，多bit='1010'。"""
    w = max(int(width or 1), 1)
    return format(int(val) & ((1 << w) - 1), "0%db" % w)


def _reg_bits(inp):
    """该输入实际占的寄存器 bit 范围 (msb, lsb)，供 K/L/M 列：
    有 slice(如 x[15:1]) → 按 reg_lsb+slice 偏移显示真实寄存器位(拆分字段两片各显其位)；
    否则用字段 reg_msb/reg_lsb；都缺失则退化用位宽(msb=width-1, lsb=0)。"""
    rl = inp.get("reg_lsb")
    sl, sm = inp.get("slice_lsb"), inp.get("slice_msb")
    if sl is not None and sm is not None:
        base = rl or 0
        return base + sm, base + sl
    width = max(int(inp.get("width") or 1), 1)
    lsb = rl if rl is not None else 0
    rm = inp.get("reg_msb")
    msb = rm if rm is not None else lsb + width - 1
    return msb, lsb


def _logic_tables(rep, include_mux=False):
    """回填用真值表：默认只 is_logic(逐字节不变)。include_mux=True(Topout 块B 陷阱③)→ 也含 mux 表，
    前提是 mux 表已带回填键(generator.report 给 mux 表补了 raw/exp_num/writes + 输入元数据)。"""
    out = []
    for t in rep.get("tables", []):
        if not t.get("tests"):
            continue
        if t.get("is_logic") or (include_mux and t.get("kind") == "mux"):
            out.append(t)
    return out


def build_fortest_rows(rep, include_mux=False):
    """把 report 的真值表 → for_test 行模型(list[group])，便于单测/复用。
    每个 group = {"name": colA, "rows": [rowdict...]}，rowdict 是 {列字母小写: 值}(只放非空格)。
    列字母用 a..n + t(列表,O 起的逐拍值)。row["kind"] ∈ {input,output,tname}。
    include_mux(Topout 块B,2026-06-23)：默认 False=只回填 logic(旧行为)；True=mux 表也回填。"""
    groups = []
    for gi, t in enumerate(_logic_tables(rep, include_mux=include_mux), start=1):
        inputs = t["inputs"]
        tests = t["tests"]
        ntest = len(tests)
        case_name = "%s_g%d" % (_strip_width(t["signal"]), gi)

        last_raw = (tests[-1]["raw"] if ntest else [])

        def _lastval(i):
            return last_raw[i] if i < len(last_raw) else 0

        # 权威 RF_WRITE(末列那拍)：compute_drives 已把 bit 拆分字段(如 x[15:1]+x[0])拼对、按位宽裁剪，
        # B/C(寄存器写值)/H 直接用它 → 与生成的 .sv 逐字节一致(自己只按 reg_lsb 重算会漏 slice 偏移算错)。
        nom_writes = {}          # addr_str -> {"hex": 合并写值串, "contrib": {字段基名小写: 该字段寄存器贡献}}
        for w in ((tests[-1].get("writes") or []) if ntest else []):
            contrib = {}
            for fld in w.get("fields", []):
                fval = int(str(fld.get("hex", "16'h0")).split("'h")[-1], 16)
                contrib[str(fld.get("base", "")).lower()] = fval << (fld.get("lsb") or 0)
            nom_writes[w["addr"]] = {"hex": w["hex"], "contrib": contrib}

        # B/C 放寄存器最后字段行；H 在【字段最后一片行】并入该字段的权威贡献(拆分字段多片只计一次)
        last_row_of_reg, last_row_of_field = {}, {}
        for i, inp in enumerate(inputs):
            if not inp.get("ro") and inp.get("addr") is not None:
                a = W.fmt_addr(inp["addr"])
                last_row_of_reg[a] = i
                last_row_of_field[(a, str(inp.get("base", "")).lower())] = i
        running = {}             # addr_str -> H 运行累加

        rows = []
        for i, inp in enumerate(inputs):
            width = inp.get("width", 1)
            r = {"kind": "input", "f": inp["label"], "n": gi}
            if i == 0:
                r["a"] = case_name
            ro = bool(inp.get("ro"))
            r["j"] = "True" if ro else "False"
            # K/L/M：按 reg_lsb + slice 偏移显示真实寄存器 bit 位(拆分字段两片各显其位，非整字段)
            msb, lsb = _reg_bits(inp)
            r["k"] = str(lsb) if msb == lsb else "%d:%d" % (msb, lsb)
            r["l"], r["m"] = msb, lsb
            addr = inp.get("addr")
            if ro or addr is None:
                wire = inp.get("wire") or inp.get("base") or inp["label"]
                r["i"] = wire
                r["b"] = wire           # RO/wire：每行 B=网名, C=16'h0, H=0
                r["c"] = W.fmt_hex(0)
                r["h"] = 0
            else:
                a = W.fmt_addr(addr)
                r["i"] = a
                base_low = str(inp.get("base", "")).lower()
                if i == last_row_of_field.get((a, base_low)):    # 字段最后一片 → 并入权威贡献
                    nw = nom_writes.get(a)
                    if nw:
                        running[a] = running.get(a, 0) | nw["contrib"].get(base_low, 0)
                r["h"] = running.get(a, 0)
                if i == last_row_of_reg.get(a):                  # 寄存器最后字段行：放权威合并写 B/C
                    r["b"] = a
                    r["c"] = (nom_writes.get(a, {}).get("hex") or W.fmt_hex(running.get(a, 0)))
            # G：末列(Tn)那拍值(二进制)
            r["g"] = _bin(_lastval(i), width) if ntest else ""
            # T 向量列：逐拍输入值(二进制)
            r["t"] = [_bin(tests[j]["raw"][i] if i < len(tests[j]["raw"]) else 0, width)
                      for j in range(ntest)]
            rows.append(r)

        # 输出行：D=输出名, E/G=末拍期望(十进制), T=逐拍期望(十进制)
        exp_nums = [int(tests[j].get("exp_num", 0)) for j in range(ntest)]
        out_row = {"kind": "output", "d": t["signal"], "n": gi,
                   "e": exp_nums[-1] if exp_nums else "",
                   "g": (str(exp_nums[-1]) if exp_nums else ""),
                   "t": [str(v) for v in exp_nums]}
        rows.append(out_row)

        # per-test 名字行：O 起逐列填测试名(真表里此行空，我们填 GUI 测试名)
        rows.append({"kind": "tname", "n": gi, "t": [tests[j]["name"] for j in range(ntest)]})

        groups.append({"name": case_name, "rows": rows})
    return groups


# 列字母 → openpyxl 列号(1 基)：A..N = 1..14，T 向量从 O=15 起
_COL = {"a": 1, "b": 2, "c": 3, "d": 4, "e": 5, "f": 6, "g": 7, "h": 8,
        "i": 9, "j": 10, "k": 11, "l": 12, "m": 13, "n": 14}
_T_START = 15


def write_fortest(src_path, out_path, rep, sheet_name="for_test"):
    """复制源 Excel 全部 sheet → 新文件，只把 for_test 页替换为回填内容。返回写出的组数。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(src_path)     # 保留全部 sheet/公式；不 data_only(我们只写不读)
    # 定位旧 for_test：记下位置后删掉，原位重建一张干净的(丢掉 designer 旧内容/合并)
    idx = None
    target = next((s for s in wb.sheetnames if s.lower() == sheet_name.lower()), None)
    if target is not None:
        idx = wb.sheetnames.index(target)
        wb.remove(wb[target])
    ws = wb.create_sheet(sheet_name, idx if idx is not None else None)

    HEAD = Font(bold=True)
    SIG_FILL = PatternFill("solid", fgColor="C6D9F1")
    OUT_FILL = PatternFill("solid", fgColor="FFF2CC")
    NAME_FILL = PatternFill("solid", fgColor="E2EFDA")

    # row1 = 列图例(给人看；inspect/消费端从 col A 非空切组，图例行 col A 留空不被误当组起)
    ws.cell(1, 2, "for_test 回填  列: A=case名 B/C=regaddr/val D=输出 E=期望 F=输入 "
                  "G/H=scratch I=地址/wire J=RO K/L/M=bits N=组号 O起=T1..Tn 向量").font = HEAD

    groups = build_fortest_rows(rep)
    r = 3                                      # 组从第 3 行起(inspect build_groups 从 row idx2 扫)
    for grp in groups:
        for row in grp["rows"]:
            for key, col in _COL.items():
                if key in row and row[key] != "" and row[key] is not None:
                    ws.cell(r, col, row[key])
            for j, tv in enumerate(row.get("t", [])):
                if tv != "" and tv is not None:
                    ws.cell(r, _T_START + j, tv)
            # 配色：信号标题行(有 col A)/输出行/名字行
            if row.get("kind") == "output":
                ws.cell(r, _COL["d"]).fill = OUT_FILL
            elif row.get("kind") == "tname":
                for j in range(len(row.get("t", []))):
                    ws.cell(r, _T_START + j).fill = NAME_FILL
            elif row.get("a"):
                ws.cell(r, _COL["a"]).font = HEAD
                ws.cell(r, _COL["a"]).fill = SIG_FILL
            r += 1
        r += 1                                 # 组间空行

    ws.freeze_panes = "G3"                      # 冻结 A–F(含输入名 col F)+图例行，右滚看 T 列时标签恒在
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["F"].width = 34
    ws.column_dimensions["I"].width = 22
    max_t = max((len(row.get("t", [])) for grp in groups for row in grp["rows"]), default=0)
    for j in range(max_t):
        ws.column_dimensions[get_column_letter(_T_START + j)].width = 8

    wb.save(out_path)
    return len(groups)
