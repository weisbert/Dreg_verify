# -*- coding: utf-8 -*-
"""make_mirror_btlp.py — 镜像【新模型】Hi1108_Pilot_BT_LP_DREG 表的开发夹具 Excel。

⭐ Topout-rooted 重构(2026-06-23)用的测试夹具。与老的 make_mirror_excel.py(WL_RFTRX、
logic-rooted、喂 mirror_wl_dreg.xlsx 给 test_cross_boundary_mux_expand)**并行存在、互不影响**。

⭐ 2026-06-23 rebuild：for_test 金标准从【1 条手验】升级成【7 条真族】——
   值逐字来自 refactor_notes/extracted_validation_examples.md(真表 for_test/SignalPath dump 实证)。
   族 = BT_LP RX 选路(mode? local : line [& ~iddq])，覆盖 选路 + iddq 门 + 1/3/4/7 bit + RW写/RO force。
   rx_en 表达式由真表 SignalPath 实证((A?C:B)&(~J), logic:109)，同族其余由 for_test line/local 标签+值实证。

新模型要点(夹具忠实体现)：
  · 要验信号 = Topout 页 B 列；logic N(top_output) 全 0、mux I(top_out) 全 0。
  · 流水线：regmap 字段(RW 主输入) →(_to_logic/_to_mux/_to_dft/_to_ls 路由)→ logic/mux/dft/ls 组合 → Topout。
  · iddq 页空、ISO 页空(只 schema)；dft 多为恒等观测(E=A)；level_shift X_to_ls→X_ls。
  · SignalPath = VBA 一次性查询(开发期对照 oracle)。

夹具里的 cone 类型(对应 Topout B)：
  ① logic 选路族(7 条真族,for_test 金标准)：rx_en / reserve / lna_agc / lpf_agc / rx_dcoc_i / rx_dcoc_q / tsensor
  ② mux 选路(控制=logic 输出 tsensor)：d_bt_lp_lna_itrim[3:0]
  ③ 直连寄存器(dft 恒等观测)：clk_force_on / en_dig_clk
  ④ 电平移位输出(_ls)：d_en_refbuf_ls ← level_shift ← logic ← 寄存器
  ⑤ RO 回读(无 cone,跳过+记账)：pll_lock_indicator

用法:  python make_mirror_btlp.py [输出路径=mirror_btlp_dreg.xlsx]
"""

import sys

import openpyxl
from openpyxl.utils import column_index_from_string


def _set(ws, row, mapping):
    for col, val in mapping.items():
        ws.cell(row=row, column=column_index_from_string(col), value=val)


def _d(n):
    return "d%d" % n


def _ws(w):
    """位宽串：w>1 → '[w-1:0]'，否则空。"""
    return "[%d:0]" % (w - 1) if w > 1 else ""


def _bin(val, w):
    """无符号二进制串(定宽)，输入行/数据用。"""
    return format(val & ((1 << w) - 1), "0%db" % w)


def _sized(val, w):
    """Verilog 定宽字面量 1'b1 / 3'b101，OUT 行用。"""
    return "%d'b%s" % (w, _bin(val, w))


# ───────────── 地址(十进制 d-号，RW 写地址与真表对齐：2D/2E/2F/32 = 45/46/47/50)─────────────
ADDR_TOP_EN = 1       # en_pll[0] / en_dig_clk[8] / clk_force_on[9]
ADDR_TESTMODE = 9     # testmode[15] / d_en_refbuf[14]
ADDR_PLL_LINE = 23    # 0x17: d_pll_line_ctrl_mode[14]
ADDR_RO41 = 41        # 0x29 readro: bt_mode_sel[0]/linectrl_rx_en[1]/iddq_mode[2]/pll_en[3]
ADDR_RO42 = 42        # 0x2A readro: linectrl_lna_agc[2:0]/linectrl_lpf_agc[6:4]
ADDR_RO43 = 43        # 0x2B readro: linectrl_rx_dcoc_i[6:0]/linectrl_rx_dcoc_q[14:8]
ADDR_RO44 = 44        # 0x2C readro: linectrl_tsensor[3:0]
ADDR_RX_LOCAL = 45    # 0x2D: linelocal_mode_ctrl[0]/bt_mode_sel_local[1]/linelocal_tsensor_ctrl[2]/rx_en_local[4]
ADDR_DCOC_LOCAL = 46  # 0x2E: local_rx_dcoc_i[6:0]/local_rx_dcoc_q[14:8]
ADDR_AGC_LOCAL = 47   # 0x2F: local_lna_agc[2:0]/local_lpf_agc[6:4]
ADDR_TSENS_LOCAL = 50 # 0x32: tsensor_local[3:0]
ADDR_ITRIM_A = 96     # lna_itrim t1..t4 (RW, _to_mux 数据)
ADDR_ITRIM_B = 97     # lna_itrim t5..t8
ADDR_RO38 = 56        # READro_reg_38: pll_lock_indicator[15] (RO 回读)


# ───────────── BT_LP RX 选路族(7 条真例，值逐字来自 extracted_validation_examples.md)─────────────
# 每条：out 基名 / 位宽 / 表达式(字母变量) / inputs[每个输入一 dict] / cols[每列=(标签,{字母:值},输出值)]
# input dict：L=字母 base=信号基名 w=位宽 kind=RW/RO addr=RW写地址(RO为None) bit=寄存器bit串
FAMILY = [
    dict(out="d_logic_bt_lp_rx_en", w=1, expr="(A?C:B)&(~D)", note="选路&iddq门(SignalPath实证)",
         ins=[dict(L="A", base="d_bt_lp_linelocal_mode_ctrl", w=1, kind="RW", addr="10'h2D", bit="0"),
              dict(L="B", base="d_bt_lp_linectrl_rx_en", w=1, kind="RO", addr=None, bit="1"),
              dict(L="C", base="d_bt_lp_rx_en_local", w=1, kind="RW", addr="10'h2D", bit="4"),
              dict(L="D", base="d_bt_lp_pll_dig_dft_iddq_mode", w=1, kind="RO", addr=None, bit="2")],
         cols=[("Line1", dict(A=0, B=1, C=0, D=0), 1), ("Line0", dict(A=0, B=0, C=1, D=0), 0),
               ("Local1", dict(A=1, B=0, C=1, D=0), 1), ("Local0", dict(A=1, B=1, C=0, D=0), 0),
               ("iddq", dict(A=0, B=1, C=0, D=1), 0)]),
    dict(out="d_logic_bt_lp_reserve", w=1, expr="(A?C:B)&(~D)", note="选路&iddq门",
         ins=[dict(L="A", base="d_bt_lp_linelocal_mode_ctrl", w=1, kind="RW", addr="10'h2D", bit="0"),
              dict(L="B", base="d_bt_lp_bt_mode_sel", w=1, kind="RO", addr=None, bit="0"),
              dict(L="C", base="d_bt_lp_bt_mode_sel_local", w=1, kind="RW", addr="10'h2D", bit="1"),
              dict(L="D", base="d_bt_lp_pll_dig_dft_iddq_mode", w=1, kind="RO", addr=None, bit="2")],
         cols=[("Line1", dict(A=0, B=1, C=0, D=0), 1), ("Line0", dict(A=0, B=0, C=1, D=0), 0),
               ("Local1", dict(A=1, B=0, C=1, D=0), 1), ("Local0", dict(A=1, B=1, C=0, D=0), 0),
               ("iddq", dict(A=0, B=1, C=0, D=1), 0)]),
    dict(out="d_logic_bt_lp_lna_agc", w=3, expr="A?C:B", note="选路(bus)",
         ins=[dict(L="A", base="d_bt_lp_linelocal_mode_ctrl", w=1, kind="RW", addr="10'h2D", bit="0"),
              dict(L="B", base="d_bt_lp_linectrl_lna_agc", w=3, kind="RO", addr=None, bit="2:0"),
              dict(L="C", base="d_bt_lp_local_lna_agc", w=3, kind="RW", addr="10'h2F", bit="2:0")],
         cols=[("line", dict(A=0, B=5, C=0), 5), ("local", dict(A=1, B=0, C=6), 6)]),
    dict(out="d_logic_bt_lp_lpf_agc", w=3, expr="A?C:B", note="选路(bus)",
         ins=[dict(L="A", base="d_bt_lp_linelocal_mode_ctrl", w=1, kind="RW", addr="10'h2D", bit="0"),
              dict(L="B", base="d_bt_lp_linectrl_lpf_agc", w=3, kind="RO", addr=None, bit="6:4"),
              dict(L="C", base="d_bt_lp_local_lpf_agc", w=3, kind="RW", addr="10'h2F", bit="6:4")],
         cols=[("line", dict(A=0, B=5, C=0), 5), ("local", dict(A=1, B=0, C=6), 6)]),
    dict(out="d_logic_bt_lp_rx_dcoc_i", w=7, expr="A?C:B", note="选路(bus,7bit)",
         ins=[dict(L="A", base="d_bt_lp_linelocal_mode_ctrl", w=1, kind="RW", addr="10'h2D", bit="0"),
              dict(L="B", base="d_bt_lp_linectrl_rx_dcoc_i", w=7, kind="RO", addr=None, bit="6:0"),
              dict(L="C", base="d_bt_lp_local_rx_dcoc_i", w=7, kind="RW", addr="10'h2E", bit="6:0")],
         cols=[("line", dict(A=0, B=25, C=0), 25), ("local", dict(A=1, B=0, C=115), 115)]),
    dict(out="d_logic_bt_lp_rx_dcoc_q", w=7, expr="A?C:B", note="选路(bus,7bit)",
         ins=[dict(L="A", base="d_bt_lp_linelocal_mode_ctrl", w=1, kind="RW", addr="10'h2D", bit="0"),
              dict(L="B", base="d_bt_lp_linectrl_rx_dcoc_q", w=7, kind="RO", addr=None, bit="14:8"),
              dict(L="C", base="d_bt_lp_local_rx_dcoc_q", w=7, kind="RW", addr="10'h2E", bit="14:8")],
         cols=[("line", dict(A=0, B=25, C=0), 25), ("local", dict(A=1, B=0, C=115), 115)]),
    dict(out="d_logic_bt_lp_tsensor", w=4, expr="A?C:B", note="选路(bus,4bit;也当 lna_itrim mux 控制)",
         ins=[dict(L="A", base="d_bt_lp_linelocal_tsensor_ctrl", w=1, kind="RW", addr="10'h2D", bit="2"),
              dict(L="B", base="d_bt_lp_linectrl_tsensor", w=4, kind="RO", addr=None, bit="3:0"),
              dict(L="C", base="d_bt_lp_tsensor_local", w=4, kind="RW", addr="10'h32", bit="3:0")],
         cols=[("line", dict(A=0, B=3, C=0), 3), ("local", dict(A=1, B=0, C=12), 12)]),
]


# ───────────── regmap：D=Reg F=Type G=Signal H=addr J..Y=bit15..bit0 Z=suffix AA=suffix2 AB=top_output AE=Owner ─────
def _regmap(ws, r, reg, typ, signal, bit_msb, bit_lsb, addr_dec, suffix="to_logic", suffix2="", owner="Yao Wang"):
    _set(ws, r, {"D": reg, "F": typ, "G": signal, "H": _d(addr_dec),
                 "Z": suffix, "AA": suffix2, "AB": 0, "AE": owner})
    for b in range(bit_lsb, bit_msb + 1):
        col = column_index_from_string("Y") - b
        ws.cell(row=r, column=col, value=1)
    return r + 1


def _tmm_reg(ws, r, name, addr_dec, reset="h0"):
    _set(ws, r, {"A": name, "B": "h%X" % addr_dec, "C": reset, "D": "RW", "E": name})
    r += 1
    _set(ws, r, {"A": "Field Name", "B": "Bit Field", "C": "Reset", "D": "DIG TOP PIN", "E": "Function"})
    return r + 1


def _tmm_field(ws, r, name, bit, typ="N", desc=""):
    _set(ws, r, {"A": name, "B": bit, "C": "d0", "D": typ, "E": desc or name})
    return r + 1


def _logic(ws, r, out, expr, inputs, suffix="to_dft", owner="Wan Xu", note=""):
    m = {"K": out, "L": expr, "M": suffix, "N": 0, "O": note or out, "P": owner, "R": r - 2}
    for i, name in enumerate(inputs):
        m[chr(ord("A") + i)] = name
    _set(ws, r, m)
    return r + 1


def _mux(ws, r, mux_input, ctrls, case, out, group_no, dest="to_dft", owner="Jiao Yexiang"):
    if isinstance(ctrls, str):
        ctrls = [ctrls]
    m = {"A": mux_input, "F": case, "G": out, "H": dest, "I": 0, "L": owner, "N": group_no}
    for i, c in enumerate(ctrls):
        m[chr(ord("B") + i)] = c
    _set(ws, r, m)
    return r + 1


def _dft(ws, r, out, in_net, expr="A", observe="16", owner="Yao Wang"):
    _set(ws, r, {"A": in_net, "D": out, "E": expr, "F": observe, "H": owner, "I": in_net})
    return r + 1


def _ls(ws, r, in_net, out_name, owner="Yao Wang"):
    _set(ws, r, {"A": in_net, "B": "STD_SR_L2H", "C": out_name, "E": 1, "F": owner, "G": in_net})
    return r + 1


def _ft_block(ws, r, case_no, mem):
    """写一个 for_test case 块：各输入行(每列 T 值) + OUT 行 + 标签行。布局忠实真表竖向 for_test。"""
    out_base, w, ins, cols = mem["out"], mem["w"], mem["ins"], mem["cols"]
    for i, inp in enumerate(ins):
        iw = _ws(inp["w"])
        fname = inp["base"] + iw
        addr = inp["addr"] if inp["kind"] == "RW" else (inp["base"] + iw)
        m = {"F": fname, "I": addr, "J": "True" if inp["kind"] == "RO" else "False",
             "K": inp["bit"], "N": case_no}
        if i == 0:
            m["A"] = out_base            # case 名只放首输入行
        for j, (_lab, vals, _o) in enumerate(cols):
            m[chr(ord("O") + j)] = _bin(vals[inp["L"]], inp["w"])
        _set(ws, r, m)
        r += 1
    # OUT 行(待验输出 + 每列定宽期望值)
    m = {"D": out_base + _ws(w), "N": case_no}
    for j, (_lab, _vals, outv) in enumerate(cols):
        m[chr(ord("O") + j)] = _sized(outv, w)
    _set(ws, r, m)
    r += 1
    # 标签行(line/local/iddq)
    m = {"G": "label", "N": case_no}
    for j, (lab, _vals, _o) in enumerate(cols):
        m[chr(ord("O") + j)] = lab
    _set(ws, r, m)
    return r + 1


def build(path):
    wb = openpyxl.Workbook()

    # ========================= logic 页 =========================
    ws = wb.active
    ws.title = "logic"
    _set(ws, 1, {"A": "A", "K": "<<输出>>", "L": "<<表达式>>"})
    _set(ws, 2, {"A": "A", "B": "B", "C": "C", "D": "D", "E": "E",
                 "K": "logic_out_signal_name", "L": "logic expression", "M": "suffix",
                 "N": "top_output", "O": "Notes", "P": "Owner", "Q": "export_from_regmap", "R": "no"})
    r = 3
    # ① 7 条选路真族
    for mem in FAMILY:
        out = mem["out"] + _ws(mem["w"])
        inputs = [inp["base"] + "_to_logic" + _ws(inp["w"]) for inp in mem["ins"]]
        suffix = "to_mux,to_dft" if mem["out"].endswith("tsensor") else "to_dft"
        r = _logic(ws, r, out, mem["expr"], inputs, suffix=suffix, note=mem["note"])
    # ④ d_en_refbuf（经 level_shift 出顶层 _ls）
    r = _logic(ws, r, "d_en_refbuf", "D?(B?A:(A&C)):E",
               ["d_en_refbuf_to_logic", "testmode_to_logic", "en_pll_to_logic",
                "d_pll_line_ctrl_mode_to_logic", "d_bt_lp_pll_en_to_logic"],
               suffix="to_dft", owner="Yao Wang", note="refbuf 使能(经 ls 出顶层 _ls)")

    # ========================= mux 页（② lna_itrim，控制=logic 输出 tsensor）=========================
    mx = wb.create_sheet("mux")
    _set(mx, 2, {"A": "mux_input", "B": "mux_ctrl_sig1", "C": "mux_ctrl_sig2", "F": "case",
                 "G": "mux_out", "H": "suffix", "I": "top_out", "J": "Notes",
                 "K": "输入是否Input端口", "L": "Owner", "M": "export_from_regmap", "N": "组号"})
    r = 3
    cases = ["4'b000x", "4'b001x", "4'b010x", "4'b011x", "4'b100x", "4'b101x", "4'b110x", "4'b111x"]
    for k, case in enumerate(cases, start=1):
        r = _mux(mx, r, "d_bt_lp_lna_itrim_t%d_to_mux[3:0]" % k,
                 "d_logic_bt_lp_tsensor_to_mux[3:0]", case,
                 "d_bt_lp_lna_itrim[3:0]", 6, dest="to_dft", owner="Jiao Yexiang")

    # ========================= dft 页（③ 恒等观测）=========================
    dft = wb.create_sheet("dft")
    _set(dft, 2, {"A": "A", "B": "B", "C": "C", "D": "out put signal", "E": "logic expression",
                  "F": "observe", "G": "suffix", "H": "owner", "I": "Exported Signals",
                  "J": "Owner of exported Signals"})
    r = 3
    r = _dft(dft, r, "clk_force_on", "clk_force_on_to_dft")
    r = _dft(dft, r, "en_dig_clk", "en_dig_clk_to_dft")
    for mem in FAMILY:
        r = _dft(dft, r, mem["out"] + _ws(mem["w"]), mem["out"] + "_to_dft" + _ws(mem["w"]))

    # ========================= regmap 页 =========================
    rm = wb.create_sheet("regmap")
    _set(rm, 2, {"A": "BLOCK", "D": "Reg_Name", "F": "Reg Type", "G": "Signal_Name", "H": "Address",
                 "I": "default", "J": "15", "Y": "0", "Z": "suffix", "AA": "suffix2",
                 "AB": "top_output", "AE": "Owner"})
    r = 3
    # 直连寄存器 / refbuf cone
    r = _regmap(rm, r, "TOP_EN", "RW", "en_pll", 0, 0, ADDR_TOP_EN, "to_logic,to_dft", "to_pll_ctrl")
    r = _regmap(rm, r, "TOP_EN", "RW", "en_dig_clk", 8, 8, ADDR_TOP_EN, "to_logic", "to_logic")
    r = _regmap(rm, r, "TOP_EN", "RW", "clk_force_on", 9, 9, ADDR_TOP_EN, "to_dft", "to_pll_crg")
    r = _regmap(rm, r, "TESTMODE1", "RW", "testmode", 15, 15, ADDR_TESTMODE, "to_logic", "to_logic")
    r = _regmap(rm, r, "TESTMODE1", "RW", "d_en_refbuf", 14, 14, ADDR_TESTMODE, "to_logic", "to_logic")
    r = _regmap(rm, r, "REG_0x17", "RW", "d_pll_line_ctrl_mode", 14, 14, ADDR_PLL_LINE, "to_logic", "to_logic")
    # readro_reg_41(虚拟/线控 RO)
    r = _regmap(rm, r, "readro_reg_41", "RO", "d_bt_lp_bt_mode_sel", 0, 0, ADDR_RO41, "to_logic", "")
    r = _regmap(rm, r, "readro_reg_41", "RO", "d_bt_lp_linectrl_rx_en", 1, 1, ADDR_RO41, "to_logic", "")
    r = _regmap(rm, r, "readro_reg_41", "RO", "d_bt_lp_pll_dig_dft_iddq_mode", 2, 2, ADDR_RO41, "to_logic", "")
    r = _regmap(rm, r, "readro_reg_41", "RO", "d_bt_lp_pll_en", 3, 3, ADDR_RO41, "to_logic", "")
    # readro_reg_42/43/44(线控数据 RO，bus)
    r = _regmap(rm, r, "readro_reg_42", "RO", "d_bt_lp_linectrl_lna_agc[2:0]", 2, 0, ADDR_RO42, "to_logic", "")
    r = _regmap(rm, r, "readro_reg_42", "RO", "d_bt_lp_linectrl_lpf_agc[2:0]", 6, 4, ADDR_RO42, "to_logic", "")
    r = _regmap(rm, r, "readro_reg_43", "RO", "d_bt_lp_linectrl_rx_dcoc_i[6:0]", 6, 0, ADDR_RO43, "to_logic", "")
    r = _regmap(rm, r, "readro_reg_43", "RO", "d_bt_lp_linectrl_rx_dcoc_q[6:0]", 14, 8, ADDR_RO43, "to_logic", "")
    r = _regmap(rm, r, "readro_reg_44", "RO", "d_bt_lp_linectrl_tsensor[3:0]", 3, 0, ADDR_RO44, "to_logic", "")
    # RW 本地控制 / 选路位(@2D)
    r = _regmap(rm, r, "RX_LOCAL", "RW", "d_bt_lp_linelocal_mode_ctrl", 0, 0, ADDR_RX_LOCAL, "to_logic", "", "Wan Xu")
    r = _regmap(rm, r, "RX_LOCAL", "RW", "d_bt_lp_bt_mode_sel_local", 1, 1, ADDR_RX_LOCAL, "to_logic", "", "Wan Xu")
    r = _regmap(rm, r, "RX_LOCAL", "RW", "d_bt_lp_linelocal_tsensor_ctrl", 2, 2, ADDR_RX_LOCAL, "to_logic", "", "Wan Xu")
    r = _regmap(rm, r, "RX_LOCAL", "RW", "d_bt_lp_rx_en_local", 4, 4, ADDR_RX_LOCAL, "to_logic", "", "Wan Xu")
    # RW 本地数据(bus，@2E/2F/32)
    r = _regmap(rm, r, "DCOC_LOCAL", "RW", "d_bt_lp_local_rx_dcoc_i[6:0]", 6, 0, ADDR_DCOC_LOCAL, "to_logic", "", "Wan Xu")
    r = _regmap(rm, r, "DCOC_LOCAL", "RW", "d_bt_lp_local_rx_dcoc_q[6:0]", 14, 8, ADDR_DCOC_LOCAL, "to_logic", "", "Wan Xu")
    r = _regmap(rm, r, "AGC_LOCAL", "RW", "d_bt_lp_local_lna_agc[2:0]", 2, 0, ADDR_AGC_LOCAL, "to_logic", "", "Wan Xu")
    r = _regmap(rm, r, "AGC_LOCAL", "RW", "d_bt_lp_local_lpf_agc[2:0]", 6, 4, ADDR_AGC_LOCAL, "to_logic", "", "Wan Xu")
    r = _regmap(rm, r, "TSENSOR_LOCAL", "RW", "d_bt_lp_tsensor_local[3:0]", 3, 0, ADDR_TSENS_LOCAL, "to_logic", "", "Jiao Yexiang")
    # lna_itrim 8 trim 数据源(RW, _to_mux)
    for k in range(1, 5):
        r = _regmap(rm, r, "ITRIM_A", "RW", "d_bt_lp_lna_itrim_t%d[3:0]" % k,
                    4 * (k - 1) + 3, 4 * (k - 1), ADDR_ITRIM_A, "to_mux", "", "Jiao Yexiang")
    for k in range(5, 9):
        r = _regmap(rm, r, "ITRIM_B", "RW", "d_bt_lp_lna_itrim_t%d[3:0]" % k,
                    4 * (k - 5) + 3, 4 * (k - 5), ADDR_ITRIM_B, "to_mux", "", "Jiao Yexiang")
    # READro_reg_38：pll_lock_indicator(RO 回读,⑤)
    r = _regmap(rm, r, "READro_reg_38", "RO", "pll_lock_indicator", 15, 15, ADDR_RO38, "to_dft", "")

    # ========================= total_memory_map 页(精简) =========================
    tmm = wb.create_sheet("total_memory_map")
    r = 1
    r = _tmm_reg(tmm, r, "TOP_EN", ADDR_TOP_EN, "h0200")
    r = _tmm_field(tmm, r, "en_dig_clk", "8")
    r = _tmm_field(tmm, r, "clk_force_on", "9")
    r = _tmm_field(tmm, r, "en_pll", "0")
    r = _tmm_reg(tmm, r, "RX_LOCAL", ADDR_RX_LOCAL)
    r = _tmm_field(tmm, r, "d_bt_lp_linelocal_mode_ctrl", "0")
    r = _tmm_field(tmm, r, "d_bt_lp_bt_mode_sel_local", "1")
    r = _tmm_field(tmm, r, "d_bt_lp_linelocal_tsensor_ctrl", "2")
    r = _tmm_field(tmm, r, "d_bt_lp_rx_en_local", "4")
    r = _tmm_reg(tmm, r, "AGC_LOCAL", ADDR_AGC_LOCAL)
    r = _tmm_field(tmm, r, "d_bt_lp_local_lna_agc[2:0]", "2:0")
    r = _tmm_field(tmm, r, "d_bt_lp_local_lpf_agc[2:0]", "6:4")
    r = _tmm_reg(tmm, r, "DCOC_LOCAL", ADDR_DCOC_LOCAL)
    r = _tmm_field(tmm, r, "d_bt_lp_local_rx_dcoc_i[6:0]", "6:0")
    r = _tmm_field(tmm, r, "d_bt_lp_local_rx_dcoc_q[6:0]", "14:8")
    r = _tmm_reg(tmm, r, "TSENSOR_LOCAL", ADDR_TSENS_LOCAL)
    r = _tmm_field(tmm, r, "d_bt_lp_tsensor_local[3:0]", "3:0")
    r = _tmm_reg(tmm, r, "READro_reg_38", ADDR_RO38)
    r = _tmm_field(tmm, r, "pll_lock_indicator", "15", "N", "RO 回读(FSM/模拟状态,无 cone)")

    # ========================= level_shift 页(④)=========================
    lvl = wb.create_sheet("level_shift")
    _set(lvl, 2, {"A": "level_shift_input", "B": "level", "C": "level_shift_output_name",
                  "D": "suffix", "E": "top_out", "F": "Owner", "G": "Exported Signals",
                  "H": "Owner of exported Signals"})
    r = 3
    r = _ls(lvl, r, "d_en_refbuf_to_ls", "d_en_refbuf_ls")
    r = _ls(lvl, r, "d_en_pfd_to_ls", "d_en_pfd_ls")

    # ========================= for_test 页(① 7 条真族金标准)=========================
    ft = wb.create_sheet("for_test")
    _fill_fortest_header(ft)
    r = 4
    for n, mem in enumerate(FAMILY, start=2):
        r = _ft_block(ft, r, n, mem)
    # for_test_gen_tc：同结构(生成版占位，放同样 7 条)
    ftg = wb.create_sheet("for_test_gen_tc")
    _fill_fortest_header(ftg)
    r = 4
    for n, mem in enumerate(FAMILY, start=2):
        r = _ft_block(ftg, r, n, mem)

    # ========================= Topout 页(要验清单 B 列 + 独立寄存器图 C–G)=========================
    top = wb.create_sheet("Topout")
    _set(top, 2, {"A": "Owner", "B": "TOP Out Signal", "C": "Reg Name", "D": "Offset",
                  "E": "Reg Description", "F": "Address", "G": "ResetValue"})
    sigs = (["clk_force_on", "en_dig_clk", "d_bt_lp_lna_itrim[3:0]", "d_en_refbuf_ls"]
            + [m["out"] + _ws(m["w"]) for m in FAMILY]
            + ["pll_lock_indicator"])
    regs = [("TOPrw_reg_1", "10'h1", "[8] en_dig_clk\n[9] clk_force_on\n[0] en_pll", "16'h0200"),
            ("RXrw_reg_45", "10'h2D", "[0] linelocal_mode_ctrl\n[4] rx_en_local ...", "16'h0000"),
            ("READro_reg_38", "ld_dig_done", "[15] pll_lock_indicator ...", "16'd1")]
    r = 3
    for i, s in enumerate(sorted(sigs)):
        m = {"A": "Yao Wang", "B": s}
        if i < len(regs):
            rn, off, desc, rst = regs[i]
            m.update({"C": rn, "D": off, "E": desc, "G": rst})
        _set(top, r, m)
        r += 1

    # ========================= SignalPath 页(rx_en 全解析块；开发期对照)=========================
    sp = wb.create_sheet("SignalPath")
    _set(sp, 4, {"B": "Reg:d41", "C": "d_bt_lp_pll_dig_dft_iddq_mode", "D": "A",
                 "E": "d_logic_bt_lp_rx_en_to_logic", "F": "logic:1", "G": "(A?C:B)&(~D)"})
    _set(sp, 5, {"F": "A", "G": "d_bt_lp_linelocal_mode_ctrl_to_logic", "H": "Reg:d45",
                 "I": "d_bt_lp_linelocal_mode_ctrl", "J": "10'h2D", "K": "[0:0]", "L": "d0"})
    _set(sp, 6, {"F": "B", "G": "d_bt_lp_linectrl_rx_en_to_logic", "H": "Reg:d41",
                 "I": "d_bt_lp_linectrl_rx_en", "J": "d_bt_lp_linectrl_rx_en", "K": "[1:1]", "L": "d0"})
    _set(sp, 7, {"F": "C", "G": "d_bt_lp_rx_en_local_to_logic", "H": "Reg:d45",
                 "I": "d_bt_lp_rx_en_local", "J": "10'h2D", "K": "[4:4]", "L": "d0"})
    _set(sp, 8, {"F": "D", "G": "d_bt_lp_pll_dig_dft_iddq_mode_to_logic", "H": "Reg:d41",
                 "I": "d_bt_lp_pll_dig_dft_iddq_mode", "K": "[2:2]", "L": "d0"})

    # ========================= main 页(top_module_name + 版本史金句)=========================
    mn = wb.create_sheet("main")
    _set(mn, 7, {"A": "项目名称", "D": "Hi1108 Pilot", "F": "项目版本", "G": "1.0"})
    _set(mn, 10, {"A": "top_module_name", "D": "BT_LP_DREG"})
    _set(mn, 23, {"A": "更新描述", "G": "更新日期", "H": "修改人员"})
    _set(mn, 34, {"A": "1）Top out 输出不带后缀，suffix 支持两个，topout 默认的把后缀去掉。",
                  "G": "2022-10-20", "H": "张立国"})
    _set(mn, 37, {"A": "Topout 更新，增加了地址列的两个标志模式 TopVar / VaVar",
                  "G": "2022-11-26", "H": "张立国"})

    # ========================= NamingRule 页(命名约定，精简)=========================
    nr = wb.create_sheet("NamingRule")
    _set(nr, 5, {"A": "A_Prefix", "B": "说明"})
    _set(nr, 6, {"A": "d", "B": "dig2ana/不区分方向"})
    _set(nr, 7, {"A": "rb", "B": "ana2dig/readback"})
    _set(nr, 19, {"A": "Signal_Type"})
    _set(nr, 21, {"A": "rw", "B": "一般读写寄存器(默认)"})
    _set(nr, 22, {"A": "ro", "B": "只读"})
    _set(nr, 23, {"A": "linectrl", "B": "线控"})
    _set(nr, 24, {"A": "local", "B": "本地控制"})
    _set(nr, 25, {"A": "mode", "B": "线控/寄存器控选择"})
    _set(nr, 35, {"A": "Suffix", "B": "说明"})
    _set(nr, 36, {"A": "to_logic"})
    _set(nr, 37, {"A": "to_mux"})
    _set(nr, 38, {"A": "to_mux,to_logic"})

    # ========================= iddq / ISO 页(本表空，只 schema)=========================
    iddq = wb.create_sheet("iddq")
    _set(iddq, 2, {"A": "A", "B": "B", "C": "C", "D": "out put signal", "E": "logic expression",
                   "F": "abserve", "G": "suffix", "H": "owner", "I": "Exported Signals",
                   "J": "Owner of exported Signals"})
    iso = wb.create_sheet("ISO")
    _set(iso, 2, {"A": "iso_input", "B": "iso_en", "C": "output_name", "D": "path of input/ iso /output",
                  "E": "power", "F": "Cell_name", "G": "Owner", "H": "Exported Signals",
                  "I": "Owner of exported Signals", "L": "ISO_top_in", "O": "RO"})

    wb.save(path)
    return path


def _fill_fortest_header(ft):
    _set(ft, 2, {"A": "case", "B": "输入寄存器地址", "C": "value", "D": "待验证输出信号", "E": "预计输出",
                 "F": "验证信号的输入信号", "G": "Signal Value(bin)", "H": "Reg(dec)", "I": "Addr",
                 "J": "RO", "K": "Bits", "L": "Bit Addr", "M": "Bit ADDR", "N": "1",
                 "O": "T0", "P": "T1", "Q": "T2", "R": "T3", "S": "T4", "T": "T5"})


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "mirror_btlp_dreg.xlsx"
    build(out)
    print("已生成 BT_LP / Topout-rooted 镜像 Excel:", out)
