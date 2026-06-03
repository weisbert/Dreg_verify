# -*- coding: utf-8 -*-
"""
inspect_mux.py — 探查 Dreg Excel『mux 页』的排版结构，导出文本给 Claude 分析。

背景: 除 logic 页外，mux 页也要纳入验证 (2026-06-03 新需求)。在定生成规则前，
先把 mux 页的真实排版摸清楚: 列结构 / 表达式 / 序号 / 与 logic·regmap 页的关系。

用法 (在真表所在机器上跑):
    python inspect_mux.py "Hi1108_Pilot_BT_LP_DREG_95P_28May.xlsx"
        [--out 输出.txt]      输出文件 (默认 <excel名>_mux_inspect.txt)
        [--sheet NAME]        指定 mux 页名 (默认自动找名字含 mux 的页)
        [--rows N]            竖排样本行数 (默认 15)
        [--refs a,b,c]        参考信号关键词 (默认用 designer MUX test .sv 里的信号)
        [--mask-owners]       把 owner 类列的人名替换成 OwnerN
        [--no-formulas]       不读公式 (快; 默认读, 公式格后附 {=原文})
        [--no-cross]          跳过跨页交叉引用 (只看 mux 页本身, 输出最短)
        [--find-max N]        每页关键词命中行上限 (默认 60)

跑完把生成的 txt 全文贴回来即可。

依赖: openpyxl  ->  pip install openpyxl
限制: 仅支持 .xlsx。
"""

import argparse
import os
import re
import sys

# Windows 控制台常是 GBK，打印中文会崩；统一改 UTF-8（文件输出本就 UTF-8）
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass


def safe_print(s):
    """reconfigure 失败的环境(重定向/老终端)下, GBK 控制台打 ⚠/中文不崩。
    关键: 报告 txt 写盘在 print 之前, 收尾消息绝不能因编码反过来吓到用户『脚本失败』。"""
    try:
        print(s)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, "encoding", None) or "ascii"
        print(s.encode(enc, "replace").decode(enc, "replace"))

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少 openpyxl，请先运行:  pip install openpyxl")

# expr 求值器可选: 有就对 mux 页疑似表达式列做"能否解析"自检
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from dreg_verify import expr as _expr
except Exception:  # noqa: BLE001
    _expr = None

# ── designer MUX test .sv 中出现的参考信号 (2026-06-03 用户提供, 用于跨页锚定) ──
#   assert_5  : d_bt_lp_dcoc_swp_I        = 0x2F bit10,    T0/T1
#   assert_8  : d_bt_lp_dcoc_swp_Q        = 0x2F bit11,    T0/T1
#   assert_11 : d_bt_lp_dcoc_itrim[2:0]   = 0x2F bit14:12, T0..T7 穷举
#   force     : d_bt_lp_linectrl_lpf_agc[2:0]=16'h5 + RF_WRITE(0x2D,0)+(0x2F,0)  (line 路径)
DEFAULT_REFS = "dcoc_swp_I,dcoc_swp_Q,dcoc_itrim,linectrl_lpf_agc,lpf_agc"

SCAN_CAP = 20000        # 单页最多扫描行数 (for_test 近万行, 防 max_row 撑爆)
PROFILE_DISTINCT = 25   # 列画像最多列出多少个去重取值
_owner_map = {}

# 疑似"表达式"的特征: 三元 ?: / verilog 常量 'b 'h / 位运算 ~ & | ^ / 拼接 {}
_EXPR_HINT_RE = re.compile(r"\?|~|&|\||\^|\{|\d+'[sS]?[bBoOdDhH]")


# ───────────────────────────── 基础工具 ─────────────────────────────
def cell_str(value, maxlen=0):
    """单元格转文本; maxlen=0 不截断。浮点整数去掉 .0; 换行转 \\n。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        s = str(int(value))
    else:
        s = str(value)
    s = s.replace("\n", "\\n").replace("\r", "")
    if maxlen and len(s) > maxlen:
        s = s[:maxlen] + "…(+%d)" % (len(s) - maxlen)
    return s


def is_blank_row(row):
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in row)


def strip_sig(name):
    """去位宽 [..] 与 _to_logic/_to_mux 后缀, 返回小写基名, 用于跨页匹配。"""
    s = "" if name is None else str(name).strip()
    s = re.sub(r"\[[^\]]*\]", "", s)
    for suf in ("_to_logic", "_to_mux"):
        if s.lower().endswith(suf):
            s = s[: -len(suf)]
    return s.strip().lower()


def mask_owner(text):
    key = str(text).strip().lower()
    if not key:
        return text
    if key not in _owner_map:
        _owner_map[key] = "Owner%d" % (len(_owner_map) + 1)
    return _owner_map[key]


def read_rows(ws, cap=SCAN_CAP):
    """整页读成 [[v,...],...]，截断到 cap 行。read_only 模式下逐行流式。"""
    rows = []
    upto = min(ws.max_row or 0, cap)
    if upto <= 0:
        return rows
    for r in ws.iter_rows(min_row=1, max_row=upto, values_only=True):
        rows.append(list(r))
    return rows


# 已知 owner 列 (来自 Dreg 真表结构): logic.P / regmap.AE。表头检测不到时兜底。
KNOWN_OWNER_COLS = {"logic": [15], "regmap": [30]}
_ROW_CACHE = {}


def get_rows(wb, sn, args):
    """带缓存读取某页全部行; --mask-owners 时在加载层统一脱敏 owner 列,
    保证之后所有 dump (原文/竖排/跨页) 都不会漏人名。"""
    if sn in _ROW_CACHE:
        return _ROW_CACHE[sn]
    rows = read_rows(wb[sn])
    if args.mask_owners and rows:
        header_row = detect_header_row(rows)
        mask_cols = find_owner_cols(rows, header_row)
        if not mask_cols:
            mask_cols = set(KNOWN_OWNER_COLS.get(sn.lower(), []))
        for ri0 in range(header_row, len(rows)):
            row = rows[ri0]
            for c in mask_cols:
                if c < len(row) and row[c] is not None and str(row[c]).strip():
                    row[c] = mask_owner(row[c])
    _ROW_CACHE[sn] = rows
    return rows


def read_formulas(wbf, sheet_name, cap=SCAN_CAP):
    """从 data_only=False 工作簿读公式; 返回 {行号(1基): row}; wbf 为 None → None。"""
    if wbf is None:
        return None
    try:
        wsf = wbf[sheet_name]
    except KeyError:
        return None
    d = {}
    upto = min(wsf.max_row or 0, cap)
    if upto <= 0:
        return d
    for ri, row in enumerate(wsf.iter_rows(min_row=1, max_row=upto, values_only=True),
                             start=1):
        d[ri] = row
    return d


def formula_suffix(ri, c, fdict):
    """该格若是公式 → ' {=原文}', 否则 ''。"""
    if fdict is None:
        return ""
    frow = fdict.get(ri)
    fv = frow[c] if (frow is not None and c < len(frow)) else None
    if isinstance(fv, str) and fv.startswith("="):
        return " {%s}" % cell_str(fv)
    return ""


def detect_header_row(rows):
    """真表头通常在第 2 行 (第 1 行是合并大标题)。
    判据: 第 2 行非空格子数 >= 第 1 行 → 表头=2, 否则=1。"""
    if len(rows) < 2:
        return 1
    n1 = sum(1 for v in rows[0] if v is not None and str(v).strip() != "")
    n2 = sum(1 for v in rows[1] if v is not None and str(v).strip() != "")
    return 2 if n2 >= n1 else 1


def row_contains(row, keywords):
    """整行任意单元格(小写)是否含任一关键词(子串)。"""
    text = " ".join("" if v is None else str(v) for v in row).lower()
    return any(k in text for k in keywords)


def render_row(rows, ri0, max_col, fdict=None, maxlen=80):
    """横排渲染一行 (ri0 为 0 基下标): '[row N] A=.. | C=..'，跳过空格子。
    注: owner 脱敏已在 get_rows 加载层做掉, 这里不用再管。"""
    cells = []
    row = rows[ri0]
    for c in range(min(max_col, len(row))):
        letter = get_column_letter(c + 1)
        v = row[c]
        suf = formula_suffix(ri0 + 1, c, fdict)
        empty = v is None or (isinstance(v, str) and v.strip() == "")
        if empty and not suf:
            continue
        cells.append("%s=%s%s" % (letter, cell_str(v, maxlen), suf))
    return "[row %d] %s" % (ri0 + 1, " | ".join(cells))


# ───────────────────────────── SECTION 0: 全部 sheet 总览 ─────────────────────────────
def dump_overview(out, wb, mux_sheets):
    out.append("=" * 70)
    out.append("==== SECTION 0: 全部 sheet 总览 ====")
    out.append("%-28s %-14s %s" % ("sheet 名", "维度(行x列)", "备注"))
    for sn in wb.sheetnames:
        ws = wb[sn]
        dim = "%sx%s" % (ws.max_row, ws.max_column)
        note = ""
        low = sn.lower()
        if sn in mux_sheets:
            note = "← 本次探查目标 (名字含 mux)"
        elif low == "logic":
            note = "(已支持: logic 页)"
        elif "for_test" in low or "fortest" in low or low.endswith("tc"):
            note = "(块状测试页: 将查参考信号整块)"
        elif low in ("regmap", "total_memory_map"):
            note = "(地址/字段来源页)"
        out.append("%-28s %-14s %s" % (sn, dim, note))
    if not mux_sheets:
        out.append("")
        out.append("⚠ 没有任何 sheet 名含 'mux' —— 请用 --sheet 指定 mux 页的真实名字后重跑。")


# ───────────────────────────── SECTION 1: mux 页结构 ─────────────────────────────
def dump_raw_head(out, rows, max_col, fdict, n=5):
    out.append("--- 1a. 前 %d 行原文 (横排, 判断表头在第几行) ---" % n)
    for ri0 in range(min(n, len(rows))):
        out.append(render_row(rows, ri0, max_col, fdict, maxlen=60))


def dump_column_profile(out, rows, max_col, header_row):
    """每列: 字母 / 行1标题 / 行2标题 / 非空数 / 去重数 / 取值或样本。
    这是『排版』的核心 —— 哪列是什么一目了然。"""
    out.append("")
    out.append("--- 1b. 列画像 (数据从第 %d 行起统计) ---" % (header_row + 1))
    h1 = rows[0] if rows else []
    h2 = rows[1] if len(rows) > 1 else []
    for c in range(max_col):
        letter = get_column_letter(c + 1)
        vals = []
        for ri0 in range(header_row, len(rows)):
            row = rows[ri0]
            v = row[c] if c < len(row) else None
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            vals.append(cell_str(v))
        t1 = cell_str(h1[c] if c < len(h1) else None, 40)
        t2 = cell_str(h2[c] if c < len(h2) else None, 40)
        if not vals and not t1 and not t2:
            continue
        distinct = list(dict.fromkeys(vals))
        head = "col %s [行1:%s | 行2:%s]" % (letter, t1 or "-", t2 or "-")
        if not vals:
            out.append("  %s: (无数据)" % head)
            continue
        if len(distinct) <= PROFILE_DISTINCT:
            detail = "取值=%s" % distinct
        else:
            detail = "样本=%s" % distinct[:8]
        out.append("  %s: 非空=%d, 去重=%d, %s" % (head, len(vals), len(distinct), detail))


def dump_vertical_samples(out, rows, max_col, header_row, fdict, n):
    """前 n 个非空数据行竖排, 全列不截断 —— 看一行长什么样。"""
    out.append("")
    out.append("--- 1c. 样本数据行竖排 (前 %d 个非空行, 不截断) ---" % n)
    h2 = rows[header_row - 1] if len(rows) >= header_row else []
    shown = 0
    for ri0 in range(header_row, len(rows)):
        if shown >= n:
            break
        row = rows[ri0]
        if is_blank_row(row):
            continue
        out.append("[row %d]" % (ri0 + 1))
        for c in range(min(max_col, len(row))):
            letter = get_column_letter(c + 1)
            v = row[c]
            suf = formula_suffix(ri0 + 1, c, fdict)
            empty = v is None or (isinstance(v, str) and v.strip() == "")
            if empty and not suf:
                continue
            hname = cell_str(h2[c] if c < len(h2) else None, 30)
            out.append("    %-3s [%s]: %s%s" % (letter, hname, cell_str(v), suf))
        shown += 1
    if shown == 0:
        out.append("  (无非空数据行)")


def struct_form(expr_text):
    """表达式 → 结构形态: 变量→V 常量→K, 算子保留 (A?C:B 与 X?Y:Z 同形态 V?V:V)。"""
    if _expr is not None:
        try:
            toks = _expr.tokenize(expr_text)
            parts = []
            for t in toks:
                if t.kind == "EOF":
                    break
                if t.kind == "ID":
                    parts.append("V")
                elif t.kind in ("BASED", "DEC"):
                    parts.append("K")
                else:
                    parts.append(t.text)
            return "".join(parts)
        except Exception:  # noqa: BLE001
            pass
    s = re.sub(r"\d+'[sS]?[bBoOdDhH][0-9a-fA-FxXzZ_\?]+", "K", expr_text)
    s = re.sub(r"[A-Za-z_]\w*", "V", s)
    return re.sub(r"\s+", "", s)


def dump_expr_columns(out, rows, max_col, header_row):
    """1d. 表达式列侦测: 哪些列像表达式 → 全部不同表达式 + 形态归并 + expr.py 解析自检。
    直接决定: mux 页能否复用 logic 页的表达式求值器, 还是要扩展语法。"""
    out.append("")
    out.append("--- 1d. 表达式列侦测 (求值器可用: %s) ---" % ("是" if _expr else "否, 仅形态归并"))
    found_any = False
    for c in range(max_col):
        letter = get_column_letter(c + 1)
        exprs = {}      # 原文 -> 出现次数
        n_str = 0
        for ri0 in range(header_row, len(rows)):
            row = rows[ri0]
            v = row[c] if c < len(row) else None
            if not isinstance(v, str) or v.strip() == "":
                continue
            n_str += 1
            if _EXPR_HINT_RE.search(v):
                exprs[v.strip()] = exprs.get(v.strip(), 0) + 1
        if not exprs or sum(exprs.values()) < max(3, n_str * 0.2):
            continue   # 命中太少, 不算表达式列
        found_any = True
        out.append("")
        out.append("  ★ col %s 疑似表达式列: %d 格命中, %d 种不同表达式"
                   % (letter, sum(exprs.values()), len(exprs)))
        # 形态归并 + 解析自检
        by_form = {}
        parse_fail = []
        for e, cnt in exprs.items():
            form = struct_form(e)
            f = by_form.setdefault(form, {"count": 0, "exprs": []})
            f["count"] += cnt
            f["exprs"].append(e)
            if _expr is not None:
                try:
                    node = _expr.parse(e)
                    widths = {vv: 1 for vv in _expr.collect_vars(node)}
                    _expr.classify_vars(node, _expr.Env(widths))
                except Exception as ex:  # noqa: BLE001
                    parse_fail.append((e, str(ex)))
        out.append("    按结构形态归并:")
        for form, f in sorted(by_form.items(), key=lambda kv: -kv[1]["count"]):
            uniq = list(dict.fromkeys(f["exprs"]))
            out.append("      [%3d次] %s" % (f["count"], form))
            for e in uniq[:4]:
                out.append("               例: %s" % e)
            if len(uniq) > 4:
                out.append("               ...(+%d 个同形态)" % (len(uniq) - 4))
        if _expr is not None:
            if parse_fail:
                out.append("    ⚠ 现有求值器解析失败 %d 种 (mux 功能要扩展 expr.py):" % len(parse_fail))
                for e, err in parse_fail:
                    out.append("      表达式: %s" % e)
                    out.append("        错误: %s" % err)
            else:
                out.append("    ✅ 全部表达式现有求值器可解析")
    if not found_any:
        out.append("  (没有发现表达式列 —— mux 页可能用『结构化列』而非表达式描述选择关系)")


def dump_serial_columns(out, rows, max_col, header_row):
    """1e. 整数序号列侦测 (assert_5/8/11 的编号来源候选): 几乎全是整数且基本递增的列。"""
    out.append("")
    out.append("--- 1e. 整数序号列候选 (designer .sv 的 assert_5/8/11 编号来源) ---")
    found = False
    for c in range(max_col):
        letter = get_column_letter(c + 1)
        nums, n_nonempty = [], 0
        for ri0 in range(header_row, len(rows)):
            row = rows[ri0]
            v = row[c] if c < len(row) else None
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            n_nonempty += 1
            if isinstance(v, (int, float)) and float(v).is_integer():
                nums.append(int(v))
            elif isinstance(v, str) and v.strip().isdigit():
                nums.append(int(v.strip()))
        if len(nums) < 5 or n_nonempty == 0 or len(nums) / n_nonempty < 0.9:
            continue
        inc = sum(1 for i in range(1, len(nums)) if nums[i] > nums[i - 1])
        if inc / max(1, len(nums) - 1) < 0.8:
            continue
        found = True
        has5811 = all(x in nums for x in (5, 8, 11))
        out.append("  col %s: %d 个整数, 范围 %d..%d, 递增率 %.0f%%%s"
                   % (letter, len(nums), min(nums), max(nums),
                      100.0 * inc / max(1, len(nums) - 1),
                      "  ← 含 5,8,11 (与 .sv 编号吻合!)" if has5811 else ""))
    if not found:
        out.append("  (没有发现明显的递增整数列)")


def find_owner_cols(rows, header_row):
    """表头含 owner 的列下标集合 (0 基), 供 --mask-owners。"""
    cols = set()
    for hr in range(min(header_row, len(rows))):
        row = rows[hr]
        for c, v in enumerate(row):
            if isinstance(v, str) and "owner" in v.lower():
                cols.add(c)
    return cols


def dump_mux_sheet(out, wb, wbf, sn, args):
    ws = wb[sn]
    rows = get_rows(wb, sn, args)
    max_col = ws.max_column or max((len(r) for r in rows), default=0)
    fdict = read_formulas(wbf, sn)
    out.append("")
    out.append("=" * 70)
    out.append("==== SECTION 1: mux 页结构 [%s] ====" % sn)
    out.append("dimensions: %s 行 x %s 列 (实际扫描 %d 行)" % (ws.max_row, max_col, len(rows)))
    if not rows:
        out.append("(空表)")
        return rows, 1, max_col
    header_row = detect_header_row(rows)
    out.append("表头行判定: 第 %d 行 (数据从第 %d 行起)" % (header_row, header_row + 1))

    dump_raw_head(out, rows, max_col, fdict)
    dump_column_profile(out, rows, max_col, header_row)
    dump_vertical_samples(out, rows, max_col, header_row, fdict, args.rows)
    dump_expr_columns(out, rows, max_col, header_row)
    dump_serial_columns(out, rows, max_col, header_row)
    return rows, header_row, max_col


# ───────────────────────────── SECTION 2: 参考信号跨页定位 ─────────────────────────────
def is_block_sheet(sheet_name):
    low = sheet_name.lower()
    return ("for_test" in low) or ("fortest" in low) or low.endswith("_tc") or low.endswith("tc")


def dump_ref_rows_in_sheet(out, wb, wbf, sn, keywords, args, vertical=False, header_row=2):
    """普通页: 含参考关键词的行。vertical=True 时竖排全 dump (用于 mux 页命中行)。"""
    ws = wb[sn]
    rows = get_rows(wb, sn, args)
    max_col = ws.max_column or max((len(r) for r in rows), default=0)
    fdict = read_formulas(wbf, sn)
    out.append("")
    out.append("---- [%s] 含参考信号的行 ----" % sn)
    h = rows[header_row - 1] if len(rows) >= header_row else []
    matched = 0
    for ri0 in range(len(rows)):
        if row_contains(rows[ri0], keywords):
            if vertical:
                out.append("[row %d]" % (ri0 + 1))
                row = rows[ri0]
                for c in range(min(max_col, len(row))):
                    v = row[c]
                    suf = formula_suffix(ri0 + 1, c, fdict)
                    empty = v is None or (isinstance(v, str) and v.strip() == "")
                    if empty and not suf:
                        continue
                    hname = cell_str(h[c] if c < len(h) else None, 30)
                    out.append("    %-3s [%s]: %s%s"
                               % (get_column_letter(c + 1), hname, cell_str(v), suf))
            else:
                out.append(render_row(rows, ri0, max_col, fdict, maxlen=60))
            matched += 1
            if matched >= args.find_max:
                out.append("...(达每页上限 %d, 截断)" % args.find_max)
                break
    out.append("(本页命中 %d 行)" % matched)
    return matched


def dump_ref_blocks_in_fortest(out, wb, wbf, sn, keywords, args):
    """块状页 (for_test/*_tc): A 列=输出名界定块; 命中参考信号的块整块 dump (横向 T 向量)。
    这是 mux 生成规则的黄金对照 —— designer 的 .sv 就是从这种页生成的。"""
    rows = get_rows(wb, sn, args)
    fdict = read_formulas(wbf, sn)
    out.append("")
    out.append("---- [%s] 参考信号整块 (含横向 T 向量) ----" % sn)
    n = len(rows)
    # 块起点: A 列基名非空且与上一个非空 A 不同
    starts, last = [], None
    for ri0 in range(n):
        a = strip_sig(rows[ri0][0] if rows[ri0] else None)
        if a and a != last:
            starts.append(ri0)
        if a:
            last = a
    dumped = 0
    for si, start in enumerate(starts):
        a_base = strip_sig(rows[start][0])
        if not any(k in a_base for k in keywords):
            continue
        end = starts[si + 1] if si + 1 < len(starts) else n
        end = min(end, start + args.block_tail)
        out.append("")
        out.append("-- 块 A=%s (行 %d..%d) --"
                   % (cell_str(rows[start][0]), start + 1, end))
        for ri0 in range(start, end):
            row = rows[ri0]
            if is_blank_row(row):
                continue
            out.append(render_row(rows, ri0, len(row), fdict, maxlen=50))
        dumped += 1
        if dumped >= 8:
            out.append("...(块数达上限 8, 截断)")
            break
    if dumped == 0:
        out.append("  ⚠ A 列没有命中参考信号的块; 退化为含关键词的行:")
        dump_ref_rows_in_sheet(out, wb, wbf, sn, keywords, args)
    else:
        out.append("(共 dump %d 个块)" % dumped)


def dump_refs(out, wb, wbf, mux_sheets, keywords, args):
    out.append("")
    out.append("=" * 70)
    out.append("==== SECTION 2: 参考信号跨页定位 ====")
    out.append("关键词 (来自 designer MUX test .sv): %s" % sorted(keywords))
    # 2a. mux 页自身: 命中行竖排 (核心)
    for sn in mux_sheets:
        dump_ref_rows_in_sheet(out, wb, wbf, sn, keywords, args, vertical=True)
    # 2b/2c/2d. 其它页
    for sn in wb.sheetnames:
        if sn in mux_sheets:
            continue
        low = sn.lower()
        if is_block_sheet(sn):
            dump_ref_blocks_in_fortest(out, wb, wbf, sn, keywords, args)
        elif low in ("logic", "regmap", "total_memory_map") or "memory_map" in low:
            dump_ref_rows_in_sheet(out, wb, wbf, sn, keywords, args)
        else:
            # 其余页 (NamingRule/main/SignalPath/level_shift...) 也扫, 但只报命中数+前几行
            rows = get_rows(wb, sn, args)
            hits = [ri0 for ri0 in range(len(rows)) if row_contains(rows[ri0], keywords)]
            if hits:
                out.append("")
                out.append("---- [%s] 命中 %d 行 (前 5 行) ----" % (sn, len(hits)))
                for ri0 in hits[:5]:
                    out.append(render_row(rows, ri0, len(rows[ri0]), None, maxlen=60))


# ───────────────────────────── SECTION 3: mux ↔ 其它页 关系 ─────────────────────────────
def collect_names(wb, sheet_name, col_letters, args, header_row=2):
    """收集某页某些列的全部基名 (strip_sig 后) → set。页不存在返回空集。"""
    name = next((s for s in wb.sheetnames if s.lower() == sheet_name.lower()), None)
    if name is None:
        return set()
    rows = get_rows(wb, name, args)
    names = set()
    from openpyxl.utils import column_index_from_string
    for letter in col_letters:
        idx = column_index_from_string(letter) - 1
        for ri0 in range(header_row, len(rows)):
            row = rows[ri0]
            v = row[idx] if idx < len(row) else None
            b = strip_sig(v)
            if b:
                names.add(b)
    return names


def dump_cross_identity(out, wb, mux_sheets, mux_data, args):
    """SECTION 3: mux 页每列的内容"身份识别" —— 这列的值像谁?
    对照: logic K 列(输出名) / logic A..J(输入名) / regmap G(Signal_Name) / tmm A(Field Name)。
    一眼看出 mux 页哪列是『来自 logic 的输入』、哪列是『寄存器字段』、哪列是『最终输出』。"""
    out.append("")
    out.append("=" * 70)
    out.append("==== SECTION 3: mux 页列内容身份识别 (与 logic/regmap/tmm 交叉) ====")

    logic_k = collect_names(wb, "logic", ["K"], args)
    logic_in = collect_names(wb, "logic", list("ABCDEFGHIJ"), args)
    regmap_g = collect_names(wb, "regmap", ["G"], args)
    tmm_a = collect_names(wb, "total_memory_map", ["A"], args, header_row=0)
    out.append("对照集合规模: logic.K=%d, logic.A-J=%d, regmap.G=%d, tmm.A=%d"
               % (len(logic_k), len(logic_in), len(regmap_g), len(tmm_a)))

    # logic 页 M=to_mux 的行: 它们的 K 名应该会出现在 mux 页 (衔接点)
    to_mux_names = set()
    logic_name = next((s for s in wb.sheetnames if s.lower() == "logic"), None)
    if logic_name:
        rows = get_rows(wb, logic_name, args)
        for ri0 in range(2, len(rows)):
            row = rows[ri0]
            m = str(row[12]).lower() if len(row) > 12 and row[12] is not None else ""   # M 列
            if "mux" in m:
                k = strip_sig(row[10] if len(row) > 10 else None)   # K 列
                if k:
                    to_mux_names.add(k)
        out.append("logic 页 M 列含 'mux' (to_mux) 的行: %d 个, K 名样本: %s"
                   % (len(to_mux_names), sorted(to_mux_names)[:10]))

    for sn in mux_sheets:
        rows, header_row, max_col = mux_data[sn]
        out.append("")
        out.append("---- [%s] 每列与各对照集合的匹配率 ----" % sn)
        out.append("    (匹配=该列某格 strip 位宽/后缀后与对照集合相等; 子串命中不算)")
        for c in range(max_col):
            letter = get_column_letter(c + 1)
            vals = []
            for ri0 in range(header_row, len(rows)):
                row = rows[ri0]
                v = row[c] if c < len(row) else None
                b = strip_sig(v)
                if b:
                    vals.append(b)
            if len(vals) < 3:
                continue
            uniq = set(vals)
            n = len(uniq)
            hits = {
                "logic.K": len(uniq & logic_k),
                "logic.A-J": len(uniq & logic_in),
                "to_mux行K": len(uniq & to_mux_names),
                "regmap.G": len(uniq & regmap_g),
                "tmm.A": len(uniq & tmm_a),
            }
            if all(v == 0 for v in hits.values()):
                continue
            parts = ["%s:%d/%d" % (k, v, n) for k, v in hits.items() if v > 0]
            out.append("  col %s: %s" % (letter, ", ".join(parts)))
        out.append("  (没列出的列 = 与任何对照集合零匹配, 多半是数值/类型/备注列)")


# ───────────────────────────── main ─────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="探查 Dreg Excel mux 页排版")
    ap.add_argument("excel", help="Excel 文件路径 (.xlsx)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--sheet", default=None, help="指定 mux 页名 (默认自动找名字含 mux 的页)")
    ap.add_argument("--rows", type=int, default=15, help="竖排样本行数 (默认 15)")
    ap.add_argument("--refs", default=DEFAULT_REFS,
                    help="参考信号关键词, 逗号分隔 (默认: designer MUX test .sv 里的信号)")
    ap.add_argument("--mask-owners", action="store_true")
    ap.add_argument("--no-formulas", action="store_true")
    ap.add_argument("--no-cross", action="store_true", help="跳过 SECTION 2/3 跨页分析")
    ap.add_argument("--find-max", type=int, default=60)
    ap.add_argument("--block-tail", type=int, default=120,
                    help="for_test 整块最多带多少行 (默认 120)")
    args = ap.parse_args()

    if not os.path.isfile(args.excel):
        sys.exit("找不到文件: %s" % args.excel)
    if not args.excel.lower().endswith(".xlsx"):
        sys.exit("只支持 .xlsx，请先在 Excel 里另存为 .xlsx")

    wb = openpyxl.load_workbook(args.excel, data_only=True, read_only=True)
    wbf = None
    if not args.no_formulas:
        try:
            wbf = openpyxl.load_workbook(args.excel, data_only=False, read_only=True)
        except Exception as ex:  # noqa: BLE001
            safe_print("⚠ 读取公式失败 (退化为只读缓存值): %s" % ex)

    # 找 mux 页
    if args.sheet:
        mux_sheets = [s for s in wb.sheetnames if s == args.sheet]
        if not mux_sheets:
            wb.close()
            sys.exit("找不到指定 sheet '%s'；现有页: %s" % (args.sheet, wb.sheetnames))
    else:
        mux_sheets = [s for s in wb.sheetnames if "mux" in s.lower()]

    keywords = {k.strip().lower() for k in args.refs.split(",") if k.strip()}

    out = []
    out.append("Dreg mux 页探查报告 (inspect_mux.py)")
    out.append("文件: %s" % os.path.basename(args.excel))
    out.append("目标 mux 页: %s" % (mux_sheets or "(未找到)"))
    out.append("公式同时导出: %s" % ("是 (公式格后附 {=...})" if wbf is not None else "否"))
    out.append("")

    # SECTION 0
    dump_overview(out, wb, mux_sheets)

    # SECTION 1
    mux_data = {}
    for sn in mux_sheets:
        mux_data[sn] = dump_mux_sheet(out, wb, wbf, sn, args)

    # SECTION 2 / 3
    if not args.no_cross and mux_sheets:
        dump_refs(out, wb, wbf, mux_sheets, keywords, args)
        dump_cross_identity(out, wb, mux_sheets, mux_data, args)
    elif not mux_sheets:
        # 没找到 mux 页时, 参考信号定位照样跑 —— 帮我们找到 mux 内容到底放在哪页
        out.append("")
        out.append("(未找到 mux 页, 改为全表搜索参考信号, 帮助定位 mux 内容所在页)")
        dump_refs(out, wb, wbf, [], keywords, args)

    wb.close()
    if wbf is not None:
        wbf.close()

    out_path = args.out if args.out else os.path.splitext(args.excel)[0] + "_mux_inspect.txt"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(out) + "\n")
    safe_print("已导出 mux 页探查报告: %s" % out_path)
    safe_print("共 %d 行。请把该 txt 全文贴回来。" % len(out))
    if not mux_sheets:
        safe_print("⚠ 未找到名字含 'mux' 的页 —— 报告里 SECTION 0 列出了全部页名, "
                   "确认后用 --sheet <页名> 重跑。")


if __name__ == "__main__":
    main()
